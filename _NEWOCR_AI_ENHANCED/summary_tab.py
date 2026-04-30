"""
summary_tab.py — DocExtract Pro
=================================
"Look-Up Summary" tab: persistent database-backed view of ALL applicants
ever processed by the Look-Up tab, across sessions.

DEDUPLICATION POLICY (client_id):
  • A non-empty client_id is the primary unique key.
  • On insert/merge: if a row with the same client_id already exists,
    only NULL / empty fields in the existing row are patched with the
    incoming values — existing data is NEVER overwritten.
  • On startup: if the DB already contains duplicate client_ids (legacy
    data), they are collapsed into one row (most-complete record kept).

DATA POPULATION SOURCES:
  • From Look-Up tab (automatic on extraction):
      applicant_name, residence_address, office_address,
      income_items, income_total, business_items, business_total,
      household_items, household_total, net_income,
      petrol_risk, transport_risk, results_json, page_map,
      amort_history_total, source_file, status, session_id
  • From "Other Data" import (manual):
      client_id, pn, industry_name, loan_balance
  • From "Amort." import (manual):
      amort_current_total
  • From "P.Loan" import (manual):
      principal_loan, maturity, interest_rate, plus expanded fields
"""

import re
import io
import csv
import json
import threading
import time
import queue
import tkinter as tk
import tkinter.ttk as ttk
import customtkinter as ctk
from pathlib import Path
from datetime import datetime
from tkinter import filedialog, messagebox

from app_constants import (
    NAVY_DEEP, NAVY_LIGHT, NAVY_MID, NAVY_PALE, NAVY_MIST, NAVY_GHOST,
    WHITE, OFF_WHITE, CARD_WHITE,
    LIME_BRIGHT, LIME_DARK, LIME_MID, LIME_PALE, LIME_MIST,
    TXT_NAVY, TXT_SOFT, TXT_MUTED, TXT_ON_LIME,
    ACCENT_RED, ACCENT_GOLD, ACCENT_SUCCESS,
    BORDER_LIGHT, BORDER_MID,
    SIDEBAR_BG, SIDEBAR_ITEM, SIDEBAR_HVR,
    F, FF, FMONO,
)

import os
from dotenv import load_dotenv
load_dotenv()

# ═══════════════════════════════════════════════════════════════════════
#  AUDIT LOG HELPER
# ═══════════════════════════════════════════════════════════════════════

def _log_action(self, action: str, description: str):
    """Write an audit log entry to the logs table."""
    try:
        user_id = getattr(self, "_current_user_id", None)
        email   = getattr(self, "_current_username", None) or ""
        with _db_connect() as conn:
            cur = conn.cursor()
            cur.execute(
                "INSERT INTO logs (user_id, email, action, description, time) "
                "VALUES (%s, %s, %s, %s, NOW())",
                (user_id, email, action, description)
            )
            conn.commit()
            cur.close()
    except Exception as e:
        print(f"[log_action] failed: {e}")

PAD          = 20
PAGE_SIZE    = 50
HDR_BG       = "#93C47D"
HDR_FG       = "#FFFFFF"
ROW_BG_EVEN  = "#F3F9F0"
ROW_BG_ODD   = WHITE
NET_GREEN    = "#1F6B28"
TOT_BG       = "#D9EAD3"
SEC_BG       = "#E8F0FA"
SEC_FG       = NAVY_MID

STATUS_COLORS = {
    "done":    ("#F0FDF4", "#166534", "✓ Done"),
    "error":   ("#FEF2F2", "#991B1B", "✗ Error"),
    "running": ("#FFFBEB", "#92400E", "⟳ Running"),
    "waiting": ("#F3F4F6", "#6B7280", "… Waiting"),
}

# ── Virtual columns: derived from results_json at render time, not real DB cols.
# ── These must never be used in ORDER BY or WHERE clauses.
_VIRTUAL_COLS = {
    "spouse_info",
    "personal_assets",
    "business_assets",
    "business_inventory",
}

# ── TABLE_COLS definition ─────────────────────────────────────────────
# (db_col, label, width_px, is_monetary, is_text_block)
# Columns populated by Look-Up are marked with ★ in comments.
TABLE_COLS = [
    # ── Populated by Other Data import ───────────────────────────────
    ("client_id",           "Client ID",                           130, False, False),
    ("pn",                  "PN",                                  100, False, False),
    # ── ★ Populated by Look-Up ────────────────────────────────────────
    ("applicant_name",      "Applicant",                           200, False, False),
    ("residence_address",   "Residence Address",                   220, False, True),
    ("office_address",      "Office Address",                      180, False, True),
    # ── Populated by Other Data import ───────────────────────────────
    ("industry_name",       "Industry Name",                       160, False, False),
    # ── ★ Virtual — derived from results_json (Look-Up) ──────────────
    ("spouse_info",         "Spouse Info",                         220, False, True),
    ("personal_assets",     "Personal Assets",                     220, False, True),
    ("business_assets",     "Business Assets",                     220, False, True),
    ("business_inventory",  "Business Inventory",                  200, False, True),
    # ── ★ Populated by Look-Up ────────────────────────────────────────
    ("income_items",        "Source of Income",                    200, False, True),
    ("income_total",        "Total Income",                        130, True,  False),
    ("business_items",      "Business Expenses",                   200, False, True),
    ("business_total",      "Total Business",                      130, True,  False),
    ("household_items",     "Household Expenses",                  200, False, True),
    ("household_total",     "Total Household",                     130, True,  False),
    ("net_income",          "Total Net Income",                    130, True,  False),
    # ── ★ Populated by Look-Up (real DB column) ───────────────────────
    ("amort_history_total", "Total Amort. History",                150, True,  False),
    # ── Populated by Amort. import ────────────────────────────────────
    ("amort_current_total", "Total Current Amort.",                150, True,  False),
    # ── Populated by Other Data import ───────────────────────────────
    ("loan_balance",        "Loan Balance",                        150, True,  False),
    # ── Populated by P.Loan import ────────────────────────────────────
    ("principal_loan",      "Principal Loan",                      150, True,  False),
    ("maturity",            "Maturity",                            140, False, False),
    ("interest_rate",       "Interest Rate",                       120, False, False),
    # ── Populated by P.Loan import (expanded fields) ──────────────────
    ("branch",              "Branch",                              130, False, False),
    ("loan_class_name",     "Loan Class",                          140, False, False),
    ("product_name",        "Product Name",                        160, False, False),
    ("loan_date",           "Loan Date",                           120, False, False),
    ("term_unit",           "Term Unit",                           90,  False, False),
    ("term",                "Term",                                80,  False, False),
    ("security",            "Security",                            160, False, True ),
    ("release_tag",         "Release Tag",                         120, False, False),
    ("loan_status",         "Loan Status",                         120, False, False),
    ("ao_name",             "AO Name",                             160, False, False),
]

_EDIT_ACTION_COL = "_edit_action"
TREE_COLS = [c[0] for c in TABLE_COLS] + [_EDIT_ACTION_COL]

LOOKUP_ROWS = [
    ("cibi_place_of_work",      "CI/BI Report",      "Office Address"),
    ("cibi_temp_residence",     "CI/BI Report",      "Residence Address"),
    ("cibi_spouse",             "CI/BI Report",      "Spouse / Employment"),
    ("cibi_spouse_office",      "CI/BI Report",      "Spouse Office Address"),
    ("cibi_personal_assets",    "CI/BI Report",      "Personal Assets"),
    ("cibi_business_assets",    "CI/BI Report",      "Business Assets"),
    ("cibi_business_inventory", "CI/BI Report",      "Business Inventory"),
    ("cibi_petrol_products",    "CI/BI Report",      "Petrol / Plastics / PVC Risk"),
    ("cibi_transport_services", "CI/BI Report",      "Transport Services Risk"),
    ("credit_history_amort",    "CI/BI Report",      "Credit History Amort."),
    ("income_remittance",       "Cashflow Analysis", "Source of Income"),
    ("cfa_business_expenses",   "Cashflow Analysis", "Business Expenses"),
    ("cfa_household_expenses",  "Cashflow Analysis", "Household / Personal Expenses"),
    ("ws_food_grocery",         "Worksheet",         "Food / Grocery"),
    ("ws_fuel_transport",       "Worksheet",         "Fuel and Transportation"),
    ("ws_electricity",          "Worksheet",         "Electricity Expense"),
    ("ws_fertilizer",           "Worksheet",         "Fertilizer"),
    ("ws_forwarding",           "Worksheet",         "Forwarding / Trucking / Hauling"),
    ("ws_fuel_diesel",          "Worksheet",         "Fuel / Gas / Diesel"),
    ("ws_equipment",            "Worksheet",         "Cost of Rent of Equipment"),
]
NON_MONETARY = {"cibi_place_of_work", "cibi_temp_residence",
                "cibi_spouse", "cibi_spouse_office",
                "cibi_personal_assets", "cibi_business_assets",
                "cibi_business_inventory"}


# ═══════════════════════════════════════════════════════════════════════
#  HELPERS — derive display fields from results_json
# ═══════════════════════════════════════════════════════════════════════

def _extract_spouse_info(results: dict) -> str:
    parts = []

    raw_spouse = results.get("cibi_spouse", {})
    spouse_items = raw_spouse.get("items", []) if isinstance(raw_spouse, dict) else []

    raw_office = results.get("cibi_spouse_office", {})
    office_items = raw_office.get("items", []) if isinstance(raw_office, dict) else []

    def _clean(text: str) -> str:
        return re.sub(r"\s*\[N/A\]\s*$", "", text).strip()

    for item in spouse_items:
        parts.append(_clean(item))
    for item in office_items:
        cleaned = _clean(item)
        if cleaned:
            parts.append(f"Office: {cleaned}")

    return "  ·  ".join(parts) if parts else ""


def _extract_asset_items(results: dict, key: str) -> str:
    raw = results.get(key, {})
    if not isinstance(raw, dict):
        return ""
    items = raw.get("items", [])
    return "  ·  ".join(items) if items else ""


def _parse_amort_history_total(results_json_str: str) -> float | None:
    """
    Derive amort_history_total from results_json string.
    Returns float or None.
    """
    try:
        blob = json.loads(results_json_str or "{}")
        data = blob.get("credit_history_amort", {})
        raw  = data.get("total", "") if isinstance(data, dict) else ""
        if raw:
            cleaned = re.sub(r"[^\d.]", "", str(raw).replace(",", ""))
            return float(cleaned) if cleaned else None
    except Exception:
        pass
    return None


# ═══════════════════════════════════════════════════════════════════════
#  EXPORT CHECKLIST DIALOG
# ═══════════════════════════════════════════════════════════════════════

_CHECKLIST_PREVIEW_COLS = [
    ("Applicant",                                180, False),
    ("Client ID",                                 90, False),
    ("PN",                                        80, False),
    ("Industry Name",                            120, False),
    ("Spouse Info",                              160, False),
    ("Personal Assets",                          160, False),
    ("Business Assets",                          160, False),
    ("Business Inventory",                       140, False),
    ("Total Source Of Income",                   120, True),
    ("Total Business Expenses",                  120, True),
    ("Total Household / Personal Expenses",      130, True),
    ("Total Net Income",                         120, True),
    ("Total Amortization History",               130, True),
    ("Total Current Amortization",               130, True),
    ("Loan Balance",                             110, True),
    ("Principal Loan",                           120, True),
    ("Maturity",                                 120, False),
    ("Interest Rate",                            110, False),
    # ── P.Loan expanded ───────────────────────────────────────────────
    ("Branch",                                   110, False),
    ("Loan Class",                               120, False),
    ("Product Name",                             140, False),
    ("Loan Date",                                100, False),
    ("Term Unit",                                80,  False),
    ("Term",                                     70,  False),
    ("Loan Amount",                              120, True ),
    ("Loan Status",                              100, False),
    ("AO Name",                                  140, False),
]

_ALL_EXPORT_COLS = [
    ("Client ID",                                 90, False),
    ("PN",                                        80, False),
    ("Applicant",                                180, False),
    ("Residence Address",                        160, False),
    ("Office Address",                           140, False),
    ("Industry Name",                            120, False),
    ("Spouse Info",                              200, False),
    ("Personal Assets",                          200, False),
    ("Business Assets",                          200, False),
    ("Business Inventory",                       180, False),
    ("Source of Income",                         160, False),
    ("Total Source Of Income",                   120, True),
    ("Business Expenses",                        160, False),
    ("Total Business Expenses",                  120, True),
    ("Household / Personal Expenses",            160, False),
    ("Total Household / Personal Expenses",      130, True),
    ("Total Net Income",                         120, True),
    ("Total Amortization History",               130, True),
    ("Total Current Amortization",               130, True),
    ("Loan Balance",                             110, True),
    ("Principal Loan",                           120, True),
    ("Maturity",                                 120, False),
    ("Interest Rate",                            110, False),
    # ── P.Loan expanded ───────────────────────────────────────────────
    ("Branch",                                   120, False),
    ("Loan Class",                               130, False),
    ("Product Name",                             150, False),
    ("Loan Date",                                110, False),
    ("Term Unit",                                90,  False),
    ("Term",                                     80,  False),
    ("Security",                                 160, False),
    ("Release Tag",                              110, False),
    ("Loan Amount",                              130, True ),
    ("Loan Status",                              110, False),
    ("AO Name",                                  150, False),
]


def _checklist_fmt_currency(val) -> str:
    if val in (None, ""):
        return "—"
    try:
        return f"P{float(val):,.2f}"
    except Exception:
        return str(val) or "—"


def _apply_checklist_tree_style():
    style = ttk.Style()
    style.configure(
        "Checklist.Treeview",
        background=WHITE, foreground=TXT_NAVY, fieldbackground=WHITE,
        rowheight=30, font=("Segoe UI", 9), borderwidth=0, relief="flat",
    )
    style.configure(
        "Checklist.Treeview.Heading",
        background=HDR_BG, foreground=HDR_FG,
        font=("Segoe UI", 9, "bold"), relief="flat", borderwidth=0, padding=(6, 6),
    )
    style.map("Checklist.Treeview.Heading",
        background=[("active", "#7AB567")], relief=[("active", "flat")])
    style.map("Checklist.Treeview",
        background=[("selected", "#C8E6C9")], foreground=[("selected", NAVY_DEEP)])


def _show_export_checklist(parent: tk.Widget, flat_rows: list) -> tuple | None:
    if not flat_rows:
        return ([], [c[0] for c in _ALL_EXPORT_COLS])

    _apply_checklist_tree_style()

    win = tk.Toplevel(parent)
    win.title("Select Rows & Columns to Export")
    win.configure(bg=CARD_WHITE)
    win.resizable(True, True)
    win.grab_set()

    win.minsize(760, 500)
    win.state("zoomed")

    result = [None]
    checked:     dict[str, bool] = {str(i): True for i in range(len(flat_rows))}
    col_checked: dict[str, bool] = {c[0]: True for c in _ALL_EXPORT_COLS}

    hdr = tk.Frame(win, bg=NAVY_DEEP)
    hdr.pack(fill="x")
    title_f = tk.Frame(hdr, bg=NAVY_DEEP)
    title_f.pack(side="left", padx=16, pady=10)
    tk.Label(title_f, text="Export Preview — Select Rows & Columns to Include",
             font=("Segoe UI", 13, "bold"), fg=WHITE, bg=NAVY_DEEP).pack(anchor="w")
    tk.Label(title_f,
             text="Tick the rows and columns you want. Only checked items will appear in the Excel file.",
             font=("Segoe UI", 8), fg="#8DA8C8", bg=NAVY_DEEP).pack(anchor="w", pady=(1, 0))

    count_var = tk.StringVar()
    tk.Label(hdr, textvariable=count_var, font=("Segoe UI", 11, "bold"),
             fg="#B9F5A0", bg=NAVY_DEEP, padx=16).pack(side="right", pady=10)

    def _refresh_count():
        nr = sum(1 for v in checked.values() if v)
        nc = sum(1 for v in col_checked.values() if v)
        count_var.set(f"{nr} / {len(flat_rows)} rows  ·  {nc} / {len(_ALL_EXPORT_COLS)} cols")

    _refresh_count()

    col_outer = tk.Frame(win, bg="#E8F0FB",
                         highlightbackground=BORDER_MID, highlightthickness=1)
    col_outer.pack(fill="x", padx=16, pady=(8, 0))
    col_hdr = tk.Frame(col_outer, bg="#E8F0FB")
    col_hdr.pack(fill="x", padx=8, pady=(6, 2))
    tk.Label(col_hdr, text="📋  Columns to export:",
             font=("Segoe UI", 8, "bold"), fg=NAVY_MID, bg="#E8F0FB").pack(side="left")

    def _col_set_all(val: bool):
        for k, v in _col_vars.items():
            v.set(val); col_checked[k] = val
        _refresh_count()

    tk.Button(col_hdr, text="All", font=("Segoe UI", 7, "bold"),
              fg=TXT_ON_LIME, bg=LIME_MID, activebackground=LIME_BRIGHT,
              activeforeground=TXT_ON_LIME, relief="flat", bd=0,
              padx=8, pady=2, cursor="hand2",
              command=lambda: _col_set_all(True)).pack(side="right", padx=(4, 0))
    tk.Button(col_hdr, text="None", font=("Segoe UI", 7, "bold"),
              fg=TXT_NAVY, bg="#D8E4F4", activebackground="#C0D0EC",
              relief="flat", bd=0, padx=8, pady=2, cursor="hand2",
              command=lambda: _col_set_all(False)).pack(side="right", padx=(0, 4))

    col_grid = tk.Frame(col_outer, bg="#E8F0FB")
    col_grid.pack(fill="x", padx=8, pady=(0, 6))
    for _ci in range(9):
        col_grid.columnconfigure(_ci, weight=1)
    _col_vars: dict[str, tk.BooleanVar] = {}
    COLS_PER_ROW = 9

    for idx, (col_key, col_w, is_mon) in enumerate(_ALL_EXPORT_COLS):
        var = tk.BooleanVar(value=True)
        _col_vars[col_key] = var

        def _on_toggle(k=col_key, v=var):
            col_checked[k] = v.get(); _refresh_count()

        row_idx = idx // COLS_PER_ROW; col_idx = idx % COLS_PER_ROW
        pill = tk.Frame(col_grid, bg="#D8E8FB",
                        highlightbackground="#B8CFF0", highlightthickness=1)
        pill.grid(row=row_idx, column=col_idx, padx=3, pady=3, sticky="w")
        tk.Checkbutton(pill, text=col_key, variable=var,
                       font=("Segoe UI", 8), fg=TXT_NAVY, bg="#D8E8FB",
                       activebackground="#C0D4F0", selectcolor=WHITE,
                       relief="flat", bd=0, padx=6, pady=3,
                       command=_on_toggle).pack()

    toolbar = tk.Frame(win, bg="#F0F4FA",
                       highlightbackground=BORDER_MID, highlightthickness=1)
    toolbar.pack(fill="x", padx=16, pady=(8, 0))

    def _mk_toolbar_btn(text, bg, fg, hov, cmd):
        return tk.Button(toolbar, text=text, font=("Segoe UI", 8, "bold"),
                         fg=fg, bg=bg, activebackground=hov, activeforeground=fg,
                         relief="flat", bd=0, padx=10, pady=4,
                         cursor="hand2", command=cmd)

    _mk_toolbar_btn("☑  Select All Rows", LIME_MID, TXT_ON_LIME, LIME_BRIGHT,
                    lambda: _set_all(True)).pack(side="left", padx=(8, 4), pady=6)
    _mk_toolbar_btn("☐  Deselect All Rows", "#E0E8F0", TXT_NAVY, "#C8D8EC",
                    lambda: _set_all(False)).pack(side="left", padx=(0, 12), pady=6)

    tk.Frame(toolbar, bg=BORDER_MID, width=1).pack(side="left", fill="y", pady=4)
    tk.Label(toolbar, text="Quick filter:", font=("Segoe UI", 8),
             fg=TXT_MUTED, bg="#F0F4FA").pack(side="left", padx=(12, 4), pady=6)
    filter_var = tk.StringVar()
    filter_var.trace_add("write", lambda *_: _apply_filter())
    tk.Entry(toolbar, textvariable=filter_var, font=("Segoe UI", 9),
             fg=TXT_NAVY, bg=WHITE, relief="solid", bd=1,
             insertbackground=NAVY_MID, width=24).pack(side="left", pady=6, ipady=3)

    CHECK_COL = "#check"
    TCOLS     = [CHECK_COL] + [c[0] for c in _CHECKLIST_PREVIEW_COLS]

    tbl_outer = tk.Frame(win, bg=BORDER_LIGHT)
    tbl_outer.pack(fill="both", expand=True, padx=16, pady=(6, 0))
    tbl_wrap = tk.Frame(tbl_outer, bg=CARD_WHITE)
    tbl_wrap.pack(fill="both", expand=True, padx=1, pady=1)
    tbl_wrap.rowconfigure(0, weight=1); tbl_wrap.columnconfigure(0, weight=1)

    vscroll = tk.Scrollbar(tbl_wrap, orient="vertical", relief="flat",
                           troughcolor=OFF_WHITE, bg=BORDER_LIGHT, width=8, bd=0)
    vscroll.grid(row=0, column=1, sticky="ns")
    hscroll = tk.Scrollbar(tbl_wrap, orient="horizontal", relief="flat",
                           troughcolor=OFF_WHITE, bg=BORDER_LIGHT, bd=0)
    hscroll.grid(row=1, column=0, columnspan=2, sticky="ew")

    tree = ttk.Treeview(tbl_wrap, columns=TCOLS, show="headings",
                        style="Checklist.Treeview",
                        yscrollcommand=vscroll.set, xscrollcommand=hscroll.set,
                        selectmode="browse")
    tree.grid(row=0, column=0, sticky="nsew")
    vscroll.config(command=tree.yview); hscroll.config(command=tree.xview)
    tree.bind("<Enter>", lambda e: tree.bind_all(
        "<MouseWheel>",
        lambda ev: tree.yview_scroll(int(-1 * (ev.delta / 120)), "units")))
    tree.bind("<Leave>", lambda e: tree.unbind_all("<MouseWheel>"))

    _all_hdr = [True]

    def _header_toggle():
        new_val = not _all_hdr[0]; _all_hdr[0] = new_val; _set_all(new_val)

    tree.heading(CHECK_COL, text="☑ All", command=_header_toggle)
    tree.column(CHECK_COL, width=62, minwidth=50, anchor="center", stretch=False)
    for col_name, col_w, is_mon in _CHECKLIST_PREVIEW_COLS:
        tree.heading(col_name, text=col_name)
        tree.column(col_name, width=col_w, minwidth=50,
                    anchor="e" if is_mon else "w", stretch=False)

    tree.tag_configure("even_on",  background=ROW_BG_EVEN, foreground=TXT_NAVY)
    tree.tag_configure("odd_on",   background=ROW_BG_ODD,  foreground=TXT_NAVY)
    tree.tag_configure("even_off", background="#F0F0F0",    foreground="#BBBBBB")
    tree.tag_configure("odd_off",  background="#EBEBEB",    foreground="#BBBBBB")

    _visible_iids: list[str] = []

    def _row_values(idx: int, row_dict: dict) -> list:
        vals = ["☑" if checked[str(idx)] else "☐"]
        for col_name, _, is_mon in _CHECKLIST_PREVIEW_COLS:
            raw = row_dict.get(col_name, "")
            if is_mon:
                vals.append(_checklist_fmt_currency(raw))
            else:
                vals.append(str(raw) if raw is not None else "")
        return vals

    def _row_tag(idx: int) -> str:
        even = (idx % 2 == 0)
        return ("even_on" if even else "odd_on") if checked[str(idx)] \
               else ("even_off" if even else "odd_off")

    def _apply_filter():
        term = filter_var.get().strip().lower()
        tree.delete(*tree.get_children()); _visible_iids.clear()
        for i, row_dict in enumerate(flat_rows):
            iid = str(i)
            if term:
                haystack = " ".join(
                    str(row_dict.get(c, "") or "").lower()
                    for c, _, __ in _CHECKLIST_PREVIEW_COLS)
                if term not in haystack:
                    continue
            _visible_iids.append(iid)
            tree.insert("", "end", iid=iid,
                        values=_row_values(i, row_dict), tags=(_row_tag(i),))
        _refresh_count()

    _apply_filter()

    def _update_row_display(iid: str):
        idx = int(iid)
        tree.item(iid, values=_row_values(idx, flat_rows[idx]),
                  tags=(_row_tag(idx),))

    def _toggle_row(iid: str):
        checked[iid] = not checked[iid]; _update_row_display(iid); _refresh_count()
        n = sum(1 for v in checked.values() if v)
        if n == len(flat_rows):
            _all_hdr[0] = True; tree.heading(CHECK_COL, text="☑ All")
        elif n == 0:
            _all_hdr[0] = False; tree.heading(CHECK_COL, text="☐ All")
        else:
            _all_hdr[0] = False; tree.heading(CHECK_COL, text="— All")

    def _set_all(val: bool):
        _all_hdr[0] = val
        for iid in checked:
            checked[iid] = val
        for iid in _visible_iids:
            _update_row_display(iid)
        tree.heading(CHECK_COL, text="☑ All" if val else "☐ All")
        _refresh_count()

    tree.bind("<ButtonRelease-1>", lambda e: _toggle_row(tree.identify_row(e.y))
              if tree.identify_row(e.y) else None)
    tree.bind("<space>", lambda e: _toggle_row(tree.focus()) if tree.focus() else None)

    btn_bar = tk.Frame(win, bg=CARD_WHITE,
                       highlightbackground=BORDER_MID, highlightthickness=1)
    btn_bar.pack(fill="x", padx=16, pady=(4, 12))

    def _mk_btn(parent, text, bg, fg, hov, cmd):
        return tk.Button(parent, text=text, font=("Segoe UI", 9, "bold"),
                         fg=fg, bg=bg, activebackground=hov, activeforeground=fg,
                         relief="flat", bd=0, padx=18, pady=7,
                         cursor="hand2", command=cmd)

    def _on_export():
        selected_rows = [flat_rows[int(i)]
                         for i in sorted(checked, key=int) if checked[i]]
        selected_cols = [c[0] for c in _ALL_EXPORT_COLS if col_checked.get(c[0], True)]
        result[0] = (selected_rows, selected_cols); win.destroy()

    def _on_cancel():
        result[0] = None; win.destroy()

    _mk_btn(btn_bar, "✕  Cancel", "#F5F5F5", TXT_SOFT, "#E0E0E0",
            _on_cancel).pack(side="right", padx=(4, 8), pady=6)
    _mk_btn(btn_bar, "📊  Export Selected to Excel", LIME_MID, TXT_ON_LIME, LIME_BRIGHT,
            _on_export).pack(side="right", padx=(0, 4), pady=6)
    tk.Label(btn_bar, textvariable=count_var, font=("Segoe UI", 9),
             fg=TXT_MUTED, bg=CARD_WHITE).pack(side="left", padx=10, pady=6)

    win.protocol("WM_DELETE_WINDOW", _on_cancel)
    parent.wait_window(win)
    return result[0]


# ═══════════════════════════════════════════════════════════════════════
#  DATABASE LAYER
# ═══════════════════════════════════════════════════════════════════════

def _db_connect():
    import psycopg2
    conn = psycopg2.connect(
        host=os.getenv("DB_HOST"),
        port=int(os.getenv("DB_PORT", 5432)),
        dbname=os.getenv("DB_NAME"),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"),
    )
    return conn


# ── All columns that can be patched (never overwrite existing non-null data)
_PATCHABLE_COLS = [
    # ── Set by Look-Up ───────────────────────────────────────────────
    "applicant_name", "residence_address", "office_address",
    "income_items", "income_total",
    "business_items", "business_total",
    "household_items", "household_total",
    "net_income", "petrol_risk", "transport_risk",
    "results_json", "page_map",
    "amort_history_total",          # ← real DB column, set by Look-Up
    "source_file", "status",
    # ── Set by Other Data import ─────────────────────────────────────
    "client_id", "pn", "industry_name", "loan_balance", "amortized_cost",
    # ── Set by Amort. import ─────────────────────────────────────────
    "amort_current_total",
    # ── Set by P.Loan import ─────────────────────────────────────────
    "principal_loan", "maturity", "interest_rate",
    "branch", "loan_class_name", "product_name",
    "loan_date", "term_unit", "term", "security", "release_tag",
    "loan_amount", "loan_status", "ao_name",
]


def _patch_existing(conn, existing_id: int, incoming: dict) -> None:
    """Fill in only columns that are currently NULL/empty. Never overwrites."""
    cur = conn.cursor()
    cur.execute("SELECT * FROM applicants WHERE id=%s", (existing_id,))
    cols_desc = [desc[0] for desc in cur.description]
    row = cur.fetchone()
    cur.close()
    if not row:
        return
    existing = dict(zip(cols_desc, row))

    parts, vals = [], []
    for col in _PATCHABLE_COLS:
        if col not in incoming:
            continue
        existing_val = existing.get(col)
        is_empty = (existing_val is None or
                    (isinstance(existing_val, str) and existing_val.strip() == ""))
        if is_empty and incoming[col] not in (None, ""):
            parts.append(f"{col}=%s")
            vals.append(incoming[col])

    if parts:
        cur2 = conn.cursor()
        cur2.execute(
            f"UPDATE applicants SET {', '.join(parts)} WHERE id=%s",
            vals + [existing_id]
        )
        cur2.close()


def _db_init():
    pass


def _db_deduplicate_client_ids() -> int:
    removed = 0
    with _db_connect() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT client_id, COUNT(*) as cnt
            FROM applicants
            WHERE client_id IS NOT NULL AND TRIM(client_id) != ''
            GROUP BY client_id HAVING COUNT(*) > 1
        """)
        dupes = cur.fetchall()

        for (cid, _cnt) in dupes:
            cur.execute(
                "SELECT * FROM applicants WHERE client_id=%s ORDER BY id ASC",
                (cid,))
            cols_desc = [desc[0] for desc in cur.description]
            rows = [dict(zip(cols_desc, r)) for r in cur.fetchall()]

            def _score(r):
                return sum(1 for v in r.values()
                           if v is not None and str(v).strip() != "")

            rows.sort(key=_score, reverse=True)
            keeper_id = rows[0]["id"]
            keeper    = rows[0]

            for dup_row in rows[1:]:
                for col in _PATCHABLE_COLS:
                    if col not in keeper:
                        continue
                    keeper_val = keeper.get(col)
                    is_empty   = (keeper_val is None or
                                  (isinstance(keeper_val, str) and
                                   keeper_val.strip() == ""))
                    incoming   = dup_row.get(col)
                    has_value  = (incoming is not None and
                                  str(incoming).strip() != "")
                    if is_empty and has_value:
                        cur.execute(
                            f"UPDATE applicants SET {col}=%s WHERE id=%s",
                            (incoming, keeper_id))
                        keeper[col] = incoming
                cur.execute("DELETE FROM applicants WHERE id=%s", (dup_row["id"],))
                removed += 1

        conn.commit()
        cur.close()
    return removed


def _db_upsert(session_id: str, row_data: dict) -> int:
    with _db_connect() as conn:
        cur = conn.cursor()
        existing_id = None
        client_id   = (row_data.get("client_id") or "").strip()

        if client_id:
            cur.execute(
                "SELECT id FROM applicants WHERE TRIM(client_id)=%s",
                (client_id,))
            row = cur.fetchone()
            if row:
                existing_id = row[0]

        if existing_id is None:
            cur.execute(
                "SELECT id FROM applicants WHERE session_id=%s AND source_file=%s",
                (session_id, row_data.get("source_file", "")))
            row = cur.fetchone()
            if row:
                existing_id = row[0]

        if existing_id is not None:
            _patch_existing(conn, existing_id, row_data)
            conn.commit()
            cur.close()
            return existing_id

        cur.execute("""
            INSERT INTO applicants (
                session_id, processed_at, source_file, status,
                applicant_name, residence_address, office_address,
                income_items, income_total,
                business_items, business_total,
                household_items, household_total,
                net_income, petrol_risk, transport_risk,
                results_json, page_map,
                amort_history_total, amort_current_total,
                client_id, pn, industry_name,
                loan_balance, amortized_cost,
                principal_loan, maturity, interest_rate,
                branch, loan_class_name, product_name,
                loan_date, term_unit, term, security, release_tag,
                loan_status, ao_name
            ) VALUES (
                %(session_id)s, %(processed_at)s, %(source_file)s, %(status)s,
                %(applicant_name)s, %(residence_address)s, %(office_address)s,
                %(income_items)s, %(income_total)s,
                %(business_items)s, %(business_total)s,
                %(household_items)s, %(household_total)s,
                %(net_income)s, %(petrol_risk)s, %(transport_risk)s,
                %(results_json)s, %(page_map)s,
                %(amort_history_total)s, %(amort_current_total)s,
                %(client_id)s, %(pn)s, %(industry_name)s,
                %(loan_balance)s, %(amortized_cost)s,
                %(principal_loan)s, %(maturity)s, %(interest_rate)s,
                %(branch)s, %(loan_class_name)s, %(product_name)s,
                %(loan_date)s, %(term_unit)s, %(term)s, %(security)s, %(release_tag)s,
                %(loan_status)s, %(ao_name)s
            ) RETURNING id
        """, row_data)
        new_id = cur.fetchone()[0]
        conn.commit()
        cur.close()
        return new_id


# ── Safe sortable columns (real DB columns only, no virtuals) ──────────
_SORTABLE_COLS = (
    {c[0] for c in TABLE_COLS}
    - _VIRTUAL_COLS
    | {"processed_at", "session_id", "source_file", "id"}
)


def _db_query(search: str = "", session_id: str = "",
              sort_col: str = "processed_at", sort_asc: bool = False,
              offset: int = 0, limit: int = PAGE_SIZE,
              adv_filters: dict = None) -> tuple:
    order_col = sort_col if sort_col in _SORTABLE_COLS else "processed_at"
    direction = "ASC" if sort_asc else "DESC"
    where_parts, params = [], []

    terms = [t.strip() for t in search.split(",") if t.strip()] if search else []
    for term in terms:
        like = f"%{term}%"
        where_parts.append(
            "(applicant_name ILIKE %s OR residence_address ILIKE %s "
            "OR office_address ILIKE %s OR income_items ILIKE %s "
            "OR business_items ILIKE %s OR household_items ILIKE %s "
            "OR source_file ILIKE %s OR client_id ILIKE %s "
            "OR pn ILIKE %s OR industry_name ILIKE %s)")
        params.extend([like] * 10)

    if session_id:
        where_parts.append("session_id = %s")
        params.append(session_id)

    _EXACT_MATCH_COLS = {"client_id"}

    if adv_filters:
        for col, values in adv_filters.items():
            if col not in _SORTABLE_COLS or not values:
                continue
            if col in _EXACT_MATCH_COLS:
                placeholders = " OR ".join(
                    [f"TRIM(UPPER({col})) = TRIM(UPPER(%s))" for _ in values])
                where_parts.append(f"({placeholders})")
                params.extend(values)
            else:
                placeholders = " OR ".join([f"{col} ILIKE %s" for _ in values])
                where_parts.append(f"({placeholders})")
                params.extend([f"%{v}%" for v in values])

    where = ("WHERE " + " AND ".join(where_parts)) if where_parts else ""
    with _db_connect() as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT COUNT(*) FROM applicants {where}", params)
        total = cur.fetchone()[0]
        cur.execute(
            f"SELECT * FROM applicants {where} "
            f"ORDER BY {order_col} {direction} LIMIT %s OFFSET %s",
            params + [limit, offset])
        cols_desc = [desc[0] for desc in cur.description]
        rows = [dict(zip(cols_desc, r)) for r in cur.fetchall()]
        cur.close()
    return rows, total


def _db_totals(session_id: str = "", search: str = "",
               adv_filters: dict = None) -> dict:
    where_parts, params = [], []

    terms = [t.strip() for t in search.split(",") if t.strip()] if search else []
    for term in terms:
        like = f"%{term}%"
        where_parts.append(
            "(applicant_name ILIKE %s OR residence_address ILIKE %s "
            "OR office_address ILIKE %s OR income_items ILIKE %s "
            "OR business_items ILIKE %s OR household_items ILIKE %s "
            "OR source_file ILIKE %s OR client_id ILIKE %s "
            "OR pn ILIKE %s OR industry_name ILIKE %s)")
        params.extend([like] * 10)

    if session_id:
        where_parts.append("session_id = %s")
        params.append(session_id)

    _EXACT_MATCH_COLS = {"client_id"}

    if adv_filters:
        for col, values in adv_filters.items():
            if col not in _SORTABLE_COLS or not values:
                continue
            if col in _EXACT_MATCH_COLS:
                placeholders = " OR ".join(
                    [f"TRIM(UPPER({col})) = TRIM(UPPER(%s))" for _ in values])
                where_parts.append(f"({placeholders})")
                params.extend(values)
            else:
                placeholders = " OR ".join([f"{col} ILIKE %s" for _ in values])
                where_parts.append(f"({placeholders})")
                params.extend([f"%{v}%" for v in values])

    where = ("WHERE " + " AND ".join(where_parts)) if where_parts else ""
    with _db_connect() as conn:
        cur = conn.cursor()
        cur.execute(f"""
            SELECT
                COUNT(*)                                          AS total,
                SUM(CASE WHEN status='done'  THEN 1 ELSE 0 END)  AS done,
                SUM(CASE WHEN status='error' THEN 1 ELSE 0 END)  AS errors,
                SUM(income_total)                                 AS income,
                SUM(net_income)                                   AS net,
                SUM(amort_current_total)                          AS amort_current
            FROM applicants {where}
        """, params)
        cols_desc = [desc[0] for desc in cur.description]
        row = cur.fetchone()
        cur.close()
    return dict(zip(cols_desc, row)) if row else {}


def _db_delete_row(row_id: int):
    with _db_connect() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM applicants WHERE id=%s", (row_id,))
        conn.commit()
        cur.close()


def _db_clear_all():
    with _db_connect() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM applicants")
        conn.commit()
        cur.close()


def _db_update_amort_current(row_id: int, value: float) -> bool:
    with _db_connect() as conn:
        cur = conn.cursor()
        cur.execute("UPDATE applicants SET amort_current_total=%s WHERE id=%s",
                    (value, row_id))
        conn.commit()
        cur.close()
    return True


def _db_update_amort_all(matches: list, value: float) -> int:
    count = 0
    with _db_connect() as conn:
        cur = conn.cursor()
        for row_id, _ in matches:
            cur.execute(
                "UPDATE applicants SET amort_current_total=%s WHERE id=%s",
                (value, row_id))
            count += 1
        conn.commit()
        cur.close()
    return count


def _db_update_other_data_all(matches: list, client_id: str, pn_joined: str,
                               industry_name: str, loan_balance,
                               amortized_cost) -> int:
    count = 0
    with _db_connect() as conn:
        cur = conn.cursor()
        for row_id, _ in matches:
            parts, vals = [], []
            if client_id:
                parts.append("client_id=%s");      vals.append(client_id)
            if pn_joined:
                parts.append("pn=%s");             vals.append(pn_joined)
            if industry_name:
                parts.append("industry_name=%s");  vals.append(industry_name)
            if loan_balance is not None:
                parts.append("loan_balance=%s");   vals.append(loan_balance)
            if amortized_cost is not None:
                parts.append("amortized_cost=%s"); vals.append(amortized_cost)
            if parts:
                cur.execute(
                    f"UPDATE applicants SET {', '.join(parts)} WHERE id=%s",
                    vals + [row_id])
                count += 1
        conn.commit()
        cur.close()
    return count


# ───────────────────────────────────────────────────────────────────────
#  CHANGE 2 (added constants + _db_update_cell)
# ───────────────────────────────────────────────────────────────────────

# Columns the user may edit directly in the table.
# Virtual columns (derived from results_json) and read-only audit
# fields (session_id, processed_at, source_file, status, page_map,
# results_json, petrol_risk, transport_risk) are intentionally excluded.
_EDITABLE_COLS = {
    "client_id", "pn", "applicant_name",
    # virtual columns — written back into results_json
    "spouse_info", "personal_assets", "business_assets", "business_inventory",
    # item-text columns — real DB columns
    "income_items", "business_items", "household_items",
    "residence_address", "office_address", "industry_name",
    "income_total", "business_total", "household_total", "net_income",
    "amort_history_total", "amort_current_total",
    "loan_balance",
    "principal_loan", "maturity", "interest_rate",
    "branch", "loan_class_name", "product_name",
    "loan_date", "term_unit", "term", "security", "release_tag",
    "loan_amount", "loan_status", "ao_name",
}

# Subset of _EDITABLE_COLS that must be stored as REAL in SQLite.
_MONETARY_COLS = {
    "income_total", "business_total", "household_total", "net_income",
    "amort_history_total", "amort_current_total",
    "loan_balance", "principal_loan",
    "loan_amount",
}


def _db_update_cell(row_id: int, col_name: str, raw_value: str) -> str:
    """
    Write a single edited cell back to the DB.

    Returns the normalised display string (formatted if monetary),
    or raises ValueError if the value cannot be parsed for monetary cols.
    """
    if col_name not in _EDITABLE_COLS:
        raise ValueError(f"Column '{col_name}' is not editable.")

    if col_name in _MONETARY_COLS:
        cleaned = re.sub(r"[^\d.]", "", raw_value.replace(",", "").strip())
        if cleaned == "":
            db_val      = None
            display_val = "—"
        else:
            try:
                db_val      = float(cleaned)
                display_val = f"P{db_val:,.2f}"
            except ValueError:
                raise ValueError(
                    f"'{raw_value}' is not a valid number for '{col_name}'.")
    else:
        db_val      = raw_value.strip() or None
        display_val = raw_value.strip() if raw_value.strip() else "—"

    with _db_connect() as conn:
        cur = conn.cursor()
        cur.execute(
            f"UPDATE applicants SET {col_name}=%s WHERE id=%s",
            (db_val, row_id))
        conn.commit()
        cur.close()

    return display_val


# Map from virtual column name → (results_json top-level key, sub-key)
# For spouse_info two keys are involved; handled specially in the function.
_VIRTUAL_TO_JSON: dict[str, tuple] = {
    "spouse_info":        ("cibi_spouse",             "items"),   # special
    "personal_assets":    ("cibi_personal_assets",    "items"),
    "business_assets":    ("cibi_business_assets",    "items"),
    "business_inventory": ("cibi_business_inventory", "items"),
}


def _db_update_virtual_cell(row_id: int, col_name: str, raw_value: str) -> str:
    """
    Write an edited virtual-column value back into results_json.

    The display format uses '  ·  ' as an item separator (set by _render_tree).
    We split on that, strip, and store back as a JSON items list.

    For spouse_info the plain items go into cibi_spouse["items"]; the
    'Office: …' prefixed items go into cibi_spouse_office["items"].

    Returns the normalised display string (items joined with '  ·  ').
    """
    if col_name not in _VIRTUAL_TO_JSON:
        raise ValueError(f"'{col_name}' is not a known virtual column.")

    raw_stripped = raw_value.strip()
    if raw_stripped in ("", "—"):
        items = []
    else:
        items = [p.strip() for p in raw_stripped.split("  ·  ") if p.strip()]

    with _db_connect() as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT results_json FROM applicants WHERE id=%s", (row_id,))
        row = cur.fetchone()
        if row is None:
            raise ValueError(f"Row id={row_id} not found.")

        try:
            blob = json.loads(row[0] or "{}")
        except Exception:
            blob = {}

        if col_name == "spouse_info":
            plain_items  = [it for it in items if not it.startswith("Office: ")]
            office_items = [it[len("Office: "):] for it in items
                            if it.startswith("Office: ")]

            if "cibi_spouse" not in blob or not isinstance(blob["cibi_spouse"], dict):
                blob["cibi_spouse"] = {}
            blob["cibi_spouse"]["items"] = plain_items

            if "cibi_spouse_office" not in blob or \
               not isinstance(blob["cibi_spouse_office"], dict):
                blob["cibi_spouse_office"] = {}
            blob["cibi_spouse_office"]["items"] = office_items
        else:
            json_key, sub_key = _VIRTUAL_TO_JSON[col_name]
            if json_key not in blob or not isinstance(blob[json_key], dict):
                blob[json_key] = {}
            blob[json_key][sub_key] = items

        cur.execute(
            "UPDATE applicants SET results_json=%s WHERE id=%s",
            (json.dumps(blob, ensure_ascii=False), row_id))
        conn.commit()
        cur.close()

    return "  ·  ".join(items) if items else "—"


# ═══════════════════════════════════════════════════════════════════════
#  NAME-MATCHING FUNCTIONS  (similarity-based)
# ═══════════════════════════════════════════════════════════════════════

SUFFIXES = {"JR", "SR", "II", "III", "IV", "V", "ESQ", "PHD", "MD", "CPA"}

# PATCH: lowered similarity threshold from 0.72 to 0.68
SIMILARITY_THRESHOLD = 0.68

try:
    from rapidfuzz import fuzz as _rfuzz
    def _similarity(a: str, b: str) -> float:
        return _rfuzz.token_sort_ratio(a, b) / 100.0
except ImportError:
    from difflib import SequenceMatcher
    def _similarity(a: str, b: str) -> float:
        a_sorted = " ".join(sorted(a.split()))
        b_sorted = " ".join(sorted(b.split()))
        return SequenceMatcher(None, a_sorted, b_sorted).ratio()


def _normalise_for_sim(name: str) -> str:
    """
    Normalise a name for similarity comparison.
    - Flips 'LAST, FIRST' format to 'FIRST LAST'
    - Uppercases, removes punctuation
    - Strips suffixes (JR, SR, etc.)
    - Strips single-letter initials (e.g. 'I.' -> removed)
    """
    if "," in name:
        parts = name.split(",", 1)
        name  = parts[1].strip() + " " + parts[0].strip()
    tokens = re.split(r"[\s.]+", name.strip().upper())
    tokens = [re.sub(r"[^A-Z]", "", t) for t in tokens]
    tokens = [t for t in tokens if t and t not in SUFFIXES]
    tokens = [t for t in tokens if len(t) > 1]  # strip bare initials
    return " ".join(tokens)


def _extract_first_last(norm_name: str) -> tuple:
    """
    From a normalised name string, extract:
      - first token  (first name)
      - last_key     (last name, compound-aware)
      - middle_tokens (everything in between)

    Handles compound Filipino/Spanish surnames that begin with a
    particle: DELA, DE, DEL, LOS, LAS, SAN, SANTA, VAN, VON, etc.
    When the second-to-last token is a known particle the last TWO
    tokens are joined as the surname key, e.g.:
        ['JUAN', 'DELA', 'CRUZ']  ->  first='JUAN', last='DELA CRUZ'
        ['MARIA', 'DE', 'LOS', 'SANTOS'] ->  first='MARIA', last='DE LOS SANTOS'  (*)

    (*) 3-token particles: the loop walks backwards collecting
        consecutive particle tokens so DE LOS SANTOS is fully captured.

    Returns: (first, last_key, middle_tokens)
    """
    PARTICLES = {
        "DE", "DEL", "DELA", "LOS", "LAS", "SAN", "SANTA",
        "VAN", "VON", "DER", "DEN", "TEN", "TER", "LE", "LA",
    }

    tokens = norm_name.split()
    if not tokens:
        return ("", "", [])
    if len(tokens) == 1:
        return (tokens[0], tokens[0], [])

    first = tokens[0]

    # Walk backwards from the end collecting consecutive particle tokens
    # so that multi-word surnames like DE LOS SANTOS are fully captured.
    surname_start = len(tokens) - 1          # index of last token
    while surname_start > 1 and tokens[surname_start - 1] in PARTICLES:
        surname_start -= 1

    last_key = " ".join(tokens[surname_start:])  # e.g. "DELA CRUZ"
    middle   = tokens[1:surname_start]           # tokens between first and surname

    return (first, last_key, middle)


def _first_last_match(norm_a: str, norm_b: str) -> bool:
    """
    Returns True if two normalised names refer to the same person,
    defined as: same first name AND same compound-aware last name,
    regardless of middle name / initial presence.

    Examples that return True:
      JUAN DELA CRUZ          <-> JUAN ISIDRO DELA CRUZ
      JUAN DELA CRUZ          <-> JUAN I DELA CRUZ          (initial stripped by normalise)
      DELA CRUZ JUAN ISIDRO   <-> JUAN DELA CRUZ            (post comma-flip)
      MARIA DE LOS SANTOS     <-> MARIA CLARA DE LOS SANTOS

    Examples that return False:
      JUAN DELA CRUZ          <-> PEDRO DELA CRUZ           (different first)
      JUAN DELA CRUZ          <-> JUAN DELA SANTOS          (different last)
    """
    first_a, last_a, _ = _extract_first_last(norm_a)
    first_b, last_b, _ = _extract_first_last(norm_b)

    if not first_a or not last_a or not first_b or not last_b:
        return False

    return first_a == first_b and last_a == last_b


def _db_candidates_all():
    with _db_connect() as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT id, applicant_name, processed_at FROM applicants "
            "WHERE applicant_name IS NOT NULL ORDER BY processed_at DESC")
        cols_desc = [desc[0] for desc in cur.description]
        rows = [dict(zip(cols_desc, r)) for r in cur.fetchall()]
        cur.close()
    return rows


# ── UPDATED FUNCTION ──────────────────────────────────────────────────
# The change: first+last match now takes priority over fuzzy scoring.
def _resolve_name_similarity(client_name: str,
                              threshold: float = SIMILARITY_THRESHOLD) -> tuple:
    candidates = _db_candidates_all()
    needle     = _normalise_for_sim(client_name)
    scored     = []

    for c in candidates:
        hay   = _normalise_for_sim(c["applicant_name"])
        score = _similarity(needle, hay)

        # ── Primary: first + last name match regardless of middle ──────
        # Handles: "JUAN DELA CRUZ"        <-> "JUAN ISIDRO DELA CRUZ"
        #          "JUAN IBIS DELA CRUZ"   <-> "JUAN DELA CRUZ"
        #          "DELA CRUZ, JUAN"       <-> "JUAN DELA CRUZ"    (post-flip)
        #          "DELA CRUZ, JUAN ISIDRO" <-> "JUAN DELA CRUZ"
        #          "DELA CRUZ, JUAN I."    <-> "JUAN DELA CRUZ"
        if _first_last_match(needle, hay):
            floored = max(score, threshold)
            scored.append((floored, c["id"], c["applicant_name"]))
            continue  # already matched — skip fuzzy check to avoid double-add

        # ── Secondary: fuzzy score meets threshold ────────────────────
        if score >= threshold:
            scored.append((score, c["id"], c["applicant_name"]))

    if not scored:
        return [], ""

    top  = max(s for s, _, __ in scored)
    hits = [(rid, rname) for score, rid, rname in scored if score >= top - 0.02]

    label = "exact" if top == 1.0 else ("high" if top >= 0.90 else "similar")
    return hits, label


# ═══════════════════════════════════════════════════════════════════════
#  PUBLIC WRITER  (called from lookup_tab.py)
# ═══════════════════════════════════════════════════════════════════════

def db_save_applicant(session_id: str, results: dict):
    """
    Persist one applicant's Look-Up results to SQLite.

    Only data that comes directly from the Look-Up extraction is written
    here. client_id, pn, industry_name, loan_balance, amortized_cost,
    principal_loan, maturity, interest_rate are left as NULL/empty
    intentionally — they are populated later via the import buttons.
    """
    _db_init()
    gate = results.get("_gate_data", {})

    def _items(key):
        return "\n".join(results.get(key, {}).get("items", []))

    def _total(key):
        t = results.get(key, {}).get("total", "")
        if not t:
            return None
        try:
            return float(re.sub(r"[^\d.]", "", t.replace(",", "")))
        except Exception:
            return None

    # ── Net income: prefer explicit Gemini value, fall back to calculated ──
    raw_net = str(results.get("_cfa_net_income", "")).strip()
    net_val = None
    if raw_net:
        try:
            net_val = float(re.sub(r"[^\d.]", "", raw_net.replace(",", "")))
        except Exception:
            pass
    if net_val is None:
        inc = _total("income_remittance") or 0
        biz = _total("cfa_business_expenses") or 0
        hh  = _total("cfa_household_expenses") or 0
        if inc or biz or hh:
            net_val = inc - biz - hh

    petrol    = bool(results.get("cibi_petrol_products",    {}).get("items"))
    transport = bool(results.get("cibi_transport_services", {}).get("items"))

    # ── Slim results_json (exclude internal _ keys) ────────────────────
    slim = {k: v for k, v in results.items()
            if not k.startswith("_") and isinstance(v, dict)}

    results_json_str = json.dumps(slim, ensure_ascii=False)

    # ── amort_history_total derived from the already-parsed results ────
    amort_hist = _parse_amort_history_total(results_json_str)

    row_data = {
        "session_id":        session_id,
        "processed_at":      datetime.now().isoformat(timespec="seconds"),
        "source_file":       results.get("_source_file", ""),
        "status":            "done",
        # ── Look-Up extracted fields ───────────────────────────────────
        "applicant_name":    results.get("_applicant_name", ""),
        "residence_address": gate.get("residence_address", ""),
        "office_address":    gate.get("office_address", ""),
        "income_items":      _items("income_remittance"),
        "income_total":      _total("income_remittance"),
        "business_items":    _items("cfa_business_expenses"),
        "business_total":    _total("cfa_business_expenses"),
        "household_items":   _items("cfa_household_expenses"),
        "household_total":   _total("cfa_household_expenses"),
        "net_income":        net_val,
        "petrol_risk":       1 if petrol    else 0,
        "transport_risk":    1 if transport else 0,
        "results_json":      results_json_str,
        "page_map":          results.get("_page_map", ""),
        "amort_history_total": amort_hist,
        # ── Intentionally NULL — set later by import buttons ──────────
        "amort_current_total": None,
        "client_id":           "",
        "pn":                  "",
        "industry_name":       "",
        "loan_balance":        None,
        "amortized_cost":      None,
        "principal_loan":      None,
        "maturity":            "",
        "interest_rate":       "",
        # Expanded P.Loan fields (initially empty)
        "branch":              "",
        "loan_class_name":     "",
        "product_name":        "",
        "loan_date":           "",
        "term_unit":           "",
        "term":                "",
        "security":            "",
        "release_tag":         "",
        "loan_status":         "",
        "ao_name":             "",
    }
    _db_upsert(session_id, row_data)


# ═══════════════════════════════════════════════════════════════════════
#  TREEVIEW STYLE
# ═══════════════════════════════════════════════════════════════════════

def _apply_tree_style():
    style = ttk.Style()
    style.theme_use("clam")
    style.configure("Summary.Treeview",
        background=WHITE, foreground=TXT_NAVY, fieldbackground=WHITE,
        rowheight=36, font=("Segoe UI", 9), borderwidth=0, relief="flat")
    style.configure("Summary.Treeview.Heading",
        background=NAVY_DEEP, foreground=WHITE,
        font=("Segoe UI", 9, "bold"), relief="flat", borderwidth=0, padding=(10, 8))
    style.map("Summary.Treeview.Heading",
        background=[("active", NAVY_LIGHT)], relief=[("active", "flat")])
    style.map("Summary.Treeview",
        background=[("selected", "#C8E6C9")], foreground=[("selected", NAVY_DEEP)])


# ═══════════════════════════════════════════════════════════════════════
#  PANEL BUILDER
# ═══════════════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════════════
#  ADVANCED FILTER DIALOG
# ═══════════════════════════════════════════════════════════════════════

# Columns available in the advanced filter (real DB cols only, user-facing)
_ADV_FILTER_COLS = [
    ("applicant_name",   "Applicant Name"),
    ("client_id",        "Client ID"),
    ("pn",               "PN"),
    ("industry_name",    "Industry Name"),
    ("residence_address","Residence Address"),
    ("office_address",   "Office Address"),
    ("loan_status",      "Loan Status"),
    ("ao_name",          "AO Name"),
    ("branch",           "Branch"),
    ("loan_class_name",  "Loan Class"),
    ("product_name",     "Product Name"),
    ("maturity",         "Maturity"),
    ("interest_rate",    "Interest Rate"),
    ("term_unit",        "Term Unit"),
    ("release_tag",      "Release Tag"),
]


def _open_advanced_filter(self):
    win = tk.Toplevel(self)
    win.title("Advanced Column Filter")
    win.configure(bg=CARD_WHITE)
    win.resizable(True, True)
    win.grab_set()

    p_x = self.winfo_rootx(); p_y = self.winfo_rooty()
    p_w = self.winfo_width(); p_h = self.winfo_height()
    w_w, w_h = 680, 560
    win.geometry(f"{w_w}x{w_h}+{p_x + (p_w-w_w)//2}+{p_y + (p_h-w_h)//2}")
    win.minsize(500, 400)

    # ── Header ────────────────────────────────────────────────────────
    hdr = tk.Frame(win, bg=NAVY_DEEP)
    hdr.pack(fill="x")
    tk.Label(hdr, text="⧉  Advanced Column Filter",
             font=("Segoe UI", 12, "bold"), fg=WHITE, bg=NAVY_DEEP,
             padx=16, pady=10).pack(side="left")
    tk.Label(hdr,
             text="Each row = one column filter. Values in a column are OR-matched.\nDifferent columns are AND-matched.",
             font=("Segoe UI", 8), fg="#8DA8C8", bg=NAVY_DEEP,
             padx=16, justify="left").pack(side="left", pady=8)

    # ── Scrollable body ───────────────────────────────────────────────
    body_outer = tk.Frame(win, bg=CARD_WHITE)
    body_outer.pack(fill="both", expand=True, padx=16, pady=(10, 0))

    canvas  = tk.Canvas(body_outer, bg=CARD_WHITE, highlightthickness=0)
    vscroll = tk.Scrollbar(body_outer, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=vscroll.set)
    vscroll.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)

    rows_frame = tk.Frame(canvas, bg=CARD_WHITE)
    cwin = canvas.create_window((0, 0), window=rows_frame, anchor="nw")

    def _on_cfg(e):
        canvas.configure(scrollregion=canvas.bbox("all"))
        canvas.itemconfig(cwin, width=canvas.winfo_width())
    rows_frame.bind("<Configure>", _on_cfg)
    canvas.bind("<Configure>", _on_cfg)
    canvas.bind_all("<MouseWheel>",
        lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

    # ── Column header labels ───────────────────────────────────────────
    col_hdr = tk.Frame(rows_frame, bg="#E8EEF8")
    col_hdr.pack(fill="x", pady=(0, 4))
    tk.Label(col_hdr, text="Column", font=("Segoe UI", 8, "bold"),
             fg=NAVY_MID, bg="#E8EEF8", width=22, anchor="w",
             padx=8, pady=4).grid(row=0, column=0, sticky="w")
    tk.Label(col_hdr, text="Filter Values  (comma-separated, OR logic)",
             font=("Segoe UI", 8, "bold"),
             fg=NAVY_MID, bg="#E8EEF8", padx=8, pady=4).grid(row=0, column=1, sticky="w")
    tk.Label(col_hdr, text="", bg="#E8EEF8", width=4).grid(row=0, column=2)

    # ── Filter rows (one per column) ──────────────────────────────────
    filter_vars: dict[str, tk.StringVar] = {}

    current = getattr(self, "_sum_adv_filters", {})

    for i, (db_col, label) in enumerate(_ADV_FILTER_COLS):
        row_bg = ROW_BG_EVEN if i % 2 == 0 else ROW_BG_ODD
        row_f  = tk.Frame(rows_frame, bg=row_bg,
                          highlightbackground=BORDER_LIGHT, highlightthickness=1)
        row_f.pack(fill="x", pady=1)
        row_f.columnconfigure(1, weight=1)

        tk.Label(row_f, text=label, font=("Segoe UI", 8, "bold"),
                 fg=NAVY_DEEP, bg=row_bg, width=22, anchor="w",
                 padx=8, pady=6).grid(row=0, column=0, sticky="w")

        # Pre-populate from existing active filters
        existing = ", ".join(current.get(db_col, []))
        var = tk.StringVar(value=existing)
        filter_vars[db_col] = var

        entry = tk.Entry(row_f, textvariable=var, font=("Segoe UI", 9),
                         fg=TXT_NAVY, bg=WHITE if not existing else "#FFFDE7",
                         relief="solid", bd=1,
                         insertbackground=NAVY_MID)
        entry.grid(row=0, column=1, sticky="ew", padx=(0, 4), pady=5)

        # Yellow bg when typed, white when empty
        def _on_var_change(v=var, e=entry):
            e.config(bg="#FFFDE7" if v.get().strip() else WHITE)
        var.trace_add("write", lambda *_, v=var, e=entry: _on_var_change(v, e))

        # Clear button per row
        def _clear_row(v=var):
            v.set("")
        tk.Button(row_f, text="✕", font=("Segoe UI", 7),
                  fg=TXT_MUTED, bg=row_bg, activebackground=ACCENT_RED,
                  activeforeground=WHITE, relief="flat", bd=0,
                  padx=6, pady=4, cursor="hand2",
                  command=_clear_row).grid(row=0, column=2, padx=(0, 4))

    # ── Hint label ────────────────────────────────────────────────────
    hint = tk.Frame(win, bg="#F0F4FA",
                    highlightbackground=BORDER_MID, highlightthickness=1)
    hint.pack(fill="x", padx=16, pady=(6, 0))
    tk.Label(hint,
             text='💡  Example: Industry Name → "Agriculture, Wholesale"  '
                  'shows rows where industry contains Agriculture OR Wholesale.',
             font=("Segoe UI", 8), fg=TXT_MUTED, bg="#F0F4FA",
             pady=5, padx=10, anchor="w").pack(fill="x")

    # ── Button bar ────────────────────────────────────────────────────
    btn_bar = tk.Frame(win, bg=CARD_WHITE,
                       highlightbackground=BORDER_MID, highlightthickness=1)
    btn_bar.pack(fill="x", padx=16, pady=(6, 12))

    active_count_var = tk.StringVar()

    def _count_active():
        n = sum(1 for v in filter_vars.values() if v.get().strip())
        active_count_var.set(
            f"{n} column filter(s) active" if n else "No filters active")

    for v in filter_vars.values():
        v.trace_add("write", lambda *_: _count_active())
    _count_active()

    tk.Label(btn_bar, textvariable=active_count_var,
             font=("Segoe UI", 8), fg=LIME_DARK, bg=CARD_WHITE,
             padx=10).pack(side="left", pady=8)

    def _clear_all_filters():
        for v in filter_vars.values():
            v.set("")

    def _apply():
        new_filters = {}
        for db_col, var in filter_vars.items():
            raw = var.get().strip()
            if raw:
                # Split on comma, strip each, drop empties
                values = [t.strip() for t in raw.split(",") if t.strip()]
                if values:
                    new_filters[db_col] = values
        self._sum_adv_filters = new_filters
        self._sum_page        = 0

        # Update the indicator label on the main panel
        n = len(new_filters)
        if n:
            cols_used = ", ".join(
                label for db_col, label in _ADV_FILTER_COLS
                if db_col in new_filters)
            self._sum_adv_filter_lbl.config(
                text=f"⧉ {n} filter(s): {cols_used}")
            self._sum_adv_filter_btn.configure(
                fg_color=LIME_MID, text_color=TXT_ON_LIME)
        else:
            self._sum_adv_filter_lbl.config(text="")
            self._sum_adv_filter_btn.configure(
                fg_color="#2A3A6C", text_color=WHITE)

        _load_and_render(self)
        canvas.unbind_all("<MouseWheel>")
        win.destroy()

    def _on_cancel():
        canvas.unbind_all("<MouseWheel>")
        win.destroy()

    tk.Button(btn_bar, text="Clear All", font=("Segoe UI", 8),
              fg=TXT_SOFT, bg="#F0F0F0", activebackground="#E0E0E0",
              relief="flat", bd=0, padx=12, pady=6, cursor="hand2",
              command=_clear_all_filters).pack(side="right", padx=(4, 8), pady=6)
    tk.Button(btn_bar, text="✕  Cancel", font=("Segoe UI", 8),
              fg=TXT_SOFT, bg="#F0F0F0", activebackground="#E0E0E0",
              relief="flat", bd=0, padx=12, pady=6, cursor="hand2",
              command=_on_cancel).pack(side="right", padx=(0, 4), pady=6)
    ctk.CTkButton(btn_bar, text="✔  Apply Filters", command=_apply,
                  width=120, height=30, corner_radius=6,
                  fg_color=LIME_MID, hover_color=LIME_BRIGHT,
                  text_color=TXT_ON_LIME,
                  font=FF(8, "bold")).pack(side="right", padx=(0, 4), pady=6)

    win.protocol("WM_DELETE_WINDOW", _on_cancel)

def _build_lookup_summary_panel(self, parent):
    _db_init()
    _db_init_custom_docs()
    _apply_tree_style()

    # Thread-safe UI callback channel for background DB writes
    self._sum_bg_queue = queue.Queue()
    self._sum_bg_poller_started = False
    self._sum_save_inflight = False
    _start_summary_ui_poller(self)

    outer = tk.Frame(parent, bg=CARD_WHITE)
    self._lookup_summary_frame = outer
    main = tk.Frame(outer, bg=CARD_WHITE)
    main.pack(fill="both", expand=True)

    # ── Header band ───────────────────────────────────────────────────
    header_band = tk.Frame(main, bg=NAVY_DEEP)
    header_band.pack(fill="x", padx=PAD, pady=(16, 0))
    title_block = tk.Frame(header_band, bg=NAVY_DEEP)
    title_block.pack(side="left", padx=16, pady=12)
    tk.Label(title_block, text="Look-Up Summary",
             font=F(15, "bold"), fg=WHITE, bg=NAVY_DEEP).pack(anchor="w")
    tk.Label(title_block,
             text="Persistent across sessions  ·  SQLite back-end  ·  live view",
             font=F(8), fg="#8DA8C8", bg=NAVY_DEEP).pack(anchor="w", pady=(2, 0))

    btn_block = tk.Frame(header_band, bg=NAVY_DEEP)
    btn_block.pack(side="right", padx=12, pady=10)

    self._sum_export_csv_btn = ctk.CTkButton(
        btn_block, text="⬇  CSV", command=lambda: _export_csv(self),
        width=68, height=30, corner_radius=6, fg_color="transparent",
        hover_color="#1E3A5F", text_color="#8DA8C8", font=FF(8, "bold"),
        border_width=1, border_color="#2E4E72")
    self._sum_export_csv_btn.pack(side="left", padx=(0, 4))

    self._sum_export_xl_btn = ctk.CTkButton(
        btn_block, text="📊  Excel", command=lambda: _export_excel(self),
        width=74, height=30, corner_radius=6, fg_color=LIME_MID,
        hover_color=LIME_BRIGHT, text_color=TXT_ON_LIME,
        font=FF(8, "bold"), border_width=0)
    self._sum_export_xl_btn.pack(side="left", padx=(0, 4))

    self._sum_clear_all_btn = ctk.CTkButton(
        btn_block, text="🗑  Clear All", command=lambda: _clear_all(self),
        width=80, height=30, corner_radius=6, fg_color="#3D1010",
        hover_color="#5C1A1A", text_color="#FF8A80",
        font=FF(8, "bold"), border_width=0)
    self._sum_clear_all_btn.pack(side="left")

    self._sum_import_ploan_btn = ctk.CTkButton(
        btn_block, text="⬆  P.Loan", command=lambda: _import_ploan_file(self),
        width=74, height=30, corner_radius=6, fg_color="#1A4A3C",
        hover_color="#256050", text_color="#A0FFD8",
        font=FF(8, "bold"), border_width=0)
    self._sum_import_ploan_btn.pack(side="left", padx=(4, 0))

    self._sum_merge_db_btn = ctk.CTkButton(
        btn_block, text="⛁  Merge DB", command=lambda: _merge_db_files(self),
        width=80, height=30, corner_radius=6, fg_color="#2D4A1E",
        hover_color="#3D6128", text_color="#B9F5A0",
        font=FF(8, "bold"), border_width=0)
    self._sum_merge_db_btn.pack(side="left", padx=(4, 0))

    self._sum_dedup_btn = ctk.CTkButton(
        btn_block, text="🔗  Dedup", command=lambda: _run_dedup(self),
        width=72, height=30, corner_radius=6, fg_color="#4A2D1E",
        hover_color="#623D28", text_color="#FFD0A0",
        font=FF(8, "bold"), border_width=0)
    self._sum_dedup_btn.pack(side="left", padx=(4, 0))

    self._sum_validate_btn = ctk.CTkButton(
        btn_block, text="✔  Validate", command=lambda: _validate_clients(self),
        width=80, height=30, corner_radius=6, fg_color="#1A4A2A",
        hover_color="#256035", text_color="#A0FFB8",
        font=FF(8, "bold"), border_width=0)
    self._sum_validate_btn.pack(side="left", padx=(4, 0))

    self._sum_adv_delete_btn = ctk.CTkButton(
    btn_block, text="🗑  Adv. Delete", command=lambda: _advanced_delete(self),
    width=90, height=30, corner_radius=6, fg_color="#5C1A1A",
    hover_color="#7A2020", text_color="#FF8A80",
    font=FF(8, "bold"), border_width=0)
    self._sum_adv_delete_btn.pack(side="left", padx=(4, 0))
    
    self._sum_del_col_btn = ctk.CTkButton(
    btn_block, text="🗑  Delete Col", command=lambda: _delete_custom_column(self),
    width=90, height=30, corner_radius=6, fg_color="#4A1A3A",
    hover_color="#6A2050", text_color="#FFB0D8",
    font=FF(8, "bold"), border_width=0)
    self._sum_del_col_btn.pack(side="left", padx=(4, 0))

    self._sum_add_doc_btn = ctk.CTkButton(
    btn_block, text="📄  Add Doc",
    command=lambda: _open_add_doc_dialog(self),
    width=82, height=30, corner_radius=6,
    fg_color="#1A2E5C", hover_color="#243E7A",
    text_color="#A0C8FF",
    font=FF(8, "bold"), border_width=0)
    self._sum_add_doc_btn.pack(side="left", padx=(4, 0))

    # ── Controls row ──────────────────────────────────────────────────
    controls_row = tk.Frame(main, bg="#F0F4FA",
                            highlightbackground=BORDER_MID, highlightthickness=1)
    controls_row.pack(fill="x", padx=PAD)

    stats_group = tk.Frame(controls_row, bg="#F0F4FA")
    stats_group.pack(side="left", padx=10, pady=8)
    self._sum_stat_labels = {}
    for key, label, color, pill_bg in [
        ("total",         "Total",          NAVY_DEEP,      "#E8EEF8"),
        ("done",          "Done",           ACCENT_SUCCESS, "#F0FDF4"),
        ("errors",        "Errors",         ACCENT_RED,     "#FFF0F0"),
        ("income",        "Income",         NAVY_MID,       "#EEF3FA"),
        ("net",           "Net",            NET_GREEN,      "#F0FDF4"),
        ("amort_current", "Curr. Amort.",   ACCENT_GOLD,    "#FFFBF0"),
    ]:
        pill = tk.Frame(stats_group, bg=pill_bg,
                        highlightbackground=BORDER_MID, highlightthickness=1)
        pill.pack(side="left", padx=(0, 6))
        tk.Label(pill, text=label, font=F(7, "bold"), fg=TXT_MUTED,
                 bg=pill_bg, padx=10, pady=3).pack()
        lbl = tk.Label(pill, text="—", font=F(11, "bold"),
                       fg=color, bg=pill_bg, padx=10, pady=4)
        lbl.pack()
        self._sum_stat_labels[key] = lbl

    tk.Frame(controls_row, bg=BORDER_MID, width=1).pack(
        side="left", fill="y", pady=6, padx=4)

    search_wrap = tk.Frame(controls_row, bg=WHITE,
                           highlightbackground=BORDER_MID, highlightthickness=1)
    search_wrap.pack(side="left", fill="x", expand=True, padx=8, pady=8)
    tk.Label(search_wrap, text="🔍", font=("Segoe UI Emoji", 9),
             bg=WHITE, fg=NAVY_PALE).pack(side="left", padx=(8, 2))
    self._sum_search_var = tk.StringVar()
    self._sum_search_var.trace_add("write", lambda *a: _on_search_change(self))
    _HINT = "Search all fields… separate terms with commas for AND filtering"
    _se = tk.Entry(search_wrap, textvariable=self._sum_search_var,
                   font=F(9), fg=TXT_MUTED, bg=WHITE, relief="flat", bd=0,
                   insertbackground=NAVY_MID)
    _se.pack(side="left", fill="x", expand=True, pady=6)
    _se.insert(0, _HINT)

    def _se_focus_in(e):
        if self._sum_search_var.get() == _HINT:
            _se.delete(0, "end"); _se.config(fg=TXT_NAVY)

    def _se_focus_out(e):
        if not self._sum_search_var.get().strip():
            _se.config(fg=TXT_MUTED); _se.insert(0, _HINT)

    _se.bind("<FocusIn>",  _se_focus_in)
    _se.bind("<FocusOut>", _se_focus_out)

    self._sum_adv_filter_btn = ctk.CTkButton(
        controls_row, text="⧉  Filter",
        command=lambda: _open_advanced_filter(self),
        width=74, height=30, corner_radius=6,
        fg_color="#2A3A6C", hover_color="#3A4A8C",
        text_color=WHITE, font=FF(8, "bold"), border_width=0)
    self._sum_adv_filter_btn.pack(side="left", padx=(4, 0), pady=8)

    self._sum_adv_filter_lbl = tk.Label(
        controls_row, text="", font=F(7), fg=LIME_MID, bg="#F0F4FA")
    self._sum_adv_filter_lbl.pack(side="left", padx=(2, 0), pady=8)

    right_ctrl = tk.Frame(controls_row, bg="#F0F4FA")
    right_ctrl.pack(side="right", padx=10, pady=8)
    self._sum_save_status_after = None
    self._sum_save_status_lbl = tk.Label(
        right_ctrl, text="", font=F(8, "bold"),
        fg=ACCENT_SUCCESS, bg="#F0F4FA")
    self._sum_save_status_lbl.pack(side="top", anchor="e")
    self._sum_count_lbl = tk.Label(right_ctrl, text="",
                                    font=F(8), fg=TXT_SOFT, bg="#F0F4FA")
    self._sum_count_lbl.pack(side="top", anchor="e")
    pg_sub = tk.Frame(right_ctrl, bg="#F0F4FA")
    pg_sub.pack(side="top", pady=(4, 0))
    self._sum_prev_btn = ctk.CTkButton(
        pg_sub, text="◀", command=lambda: _page_prev(self),
        width=30, height=24, corner_radius=5, fg_color=CARD_WHITE,
        hover_color=NAVY_MIST, text_color=NAVY_MID, font=FF(8, "bold"),
        border_width=1, border_color=BORDER_MID, state="disabled")
    self._sum_prev_btn.pack(side="left", padx=(0, 4))
    self._sum_page_lbl = tk.Label(pg_sub, text="Page 1",
                                  font=F(8), fg=TXT_SOFT, bg="#F0F4FA")
    self._sum_page_lbl.pack(side="left", padx=4)
    self._sum_next_btn = ctk.CTkButton(
        pg_sub, text="▶", command=lambda: _page_next(self),
        width=30, height=24, corner_radius=5, fg_color=CARD_WHITE,
        hover_color=NAVY_MIST, text_color=NAVY_MID, font=FF(8, "bold"),
        border_width=1, border_color=BORDER_MID, state="disabled")
    self._sum_next_btn.pack(side="left", padx=(4, 0))

    # ── Treeview ──────────────────────────────────────────────────────
    tbl_outer = tk.Frame(main, bg=BORDER_LIGHT)
    tbl_outer.pack(fill="both", expand=True, padx=PAD, pady=(8, PAD))
    tbl_wrap = tk.Frame(tbl_outer, bg=CARD_WHITE)
    tbl_wrap.pack(fill="both", expand=True, padx=1, pady=1)
    tbl_wrap.rowconfigure(0, weight=1); tbl_wrap.columnconfigure(0, weight=1)

    vscroll = tk.Scrollbar(tbl_wrap, orient="vertical", relief="flat",
                           troughcolor=OFF_WHITE, bg=BORDER_LIGHT, width=8, bd=0)
    vscroll.grid(row=0, column=1, sticky="ns")
    hscroll = tk.Scrollbar(tbl_wrap, orient="horizontal", relief="flat",
                           troughcolor=OFF_WHITE, bg=BORDER_LIGHT, bd=0)
    hscroll.grid(row=1, column=0, columnspan=2, sticky="ew")

    self._sum_tree = ttk.Treeview(
        tbl_wrap, columns=TREE_COLS, show="headings",
        style="Summary.Treeview",
        yscrollcommand=vscroll.set, xscrollcommand=hscroll.set,
        selectmode="browse")
    self._sum_tree.grid(row=0, column=0, sticky="nsew")
    vscroll.config(command=self._sum_tree.yview)
    hscroll.config(command=self._sum_tree.xview)

    for db_col, label, width_px, is_mon, is_txt in TABLE_COLS:
        anchor = "e" if is_mon else "w"
        self._sum_tree.heading(db_col, text=label,
                               command=lambda c=db_col: _sort_by(self, c))
        self._sum_tree.column(db_col, width=width_px, minwidth=60,
                              anchor=anchor, stretch=False)
    self._sum_tree.heading(_EDIT_ACTION_COL, text="Edit")
    self._sum_tree.column(_EDIT_ACTION_COL, width=64, minwidth=54,
                          anchor="center", stretch=False)

    self._sum_tree.tag_configure("even", background=ROW_BG_EVEN)
    self._sum_tree.tag_configure("odd",  background=ROW_BG_ODD)

    # ── CHANGE 4 (replace the single <Double-1> bind with two lines) ──
    self._sum_tree.bind("<Double-1>", lambda e: _on_tree_double_click(self, e))
    self._sum_tree.bind("<Return>",   lambda e: _on_tree_return_key(self, e))

    self._sum_tree.bind("<Button-3>", lambda e: _on_tree_right_click(self, e))
    self._sum_tree.bind("<Enter>",
        lambda e: self._sum_tree.bind_all("<MouseWheel>",
            lambda ev: self._sum_tree.yview_scroll(int(-1*(ev.delta/120)), "units")))
    self._sum_tree.bind("<Leave>", lambda e: self._sum_tree.unbind_all("<MouseWheel>"))

    self._sum_sort_col       = "processed_at"
    self._sum_sort_asc       = False
    self._sum_page           = 0
    self._sum_total_rows     = 0
    self._sum_session_filter = ""
    self._sum_search_after   = None
    self._sum_row_data       = {}
    self._sum_adv_filters    = {}
    self._sum_active_cell_entry = None
    _refresh_summary(self)


# ═══════════════════════════════════════════════════════════════════════
#  DATA LOADING + RENDERING
# ═══════════════════════════════════════════════════════════════════════

def _on_search_change(self):
    if "separate terms with commas" in self._sum_search_var.get():
        return
    if self._sum_search_after:
        self.after_cancel(self._sum_search_after)
    self._sum_search_after = self.after(300, lambda: _load_and_render(self))


def _refresh_summary(self):
    _load_and_render(self)


def _load_and_render(self):
    # If the grid is about to re-render (sort/search/paginate), close any
    # inline editor so it can't "float" over the refreshed Treeview.
    _teardown_summary_cell_entry(self, getattr(self, "_sum_active_cell_entry", None))
    # ── Sync custom columns into tree definition ───────────────────────
    try:
        custom_cols = _db_get_custom_columns()
    except Exception:
        custom_cols = []

    prev_custom = getattr(self, "_sum_custom_cols", [])
    self._sum_custom_cols = custom_cols

    # If custom columns changed, rebuild the Treeview column definitions
    if [c[0] for c in custom_cols] != [c[0] for c in prev_custom]:
        new_tree_cols = (
            [c[0] for c in TABLE_COLS]
            + [c[0] for c in custom_cols]
            + [_EDIT_ACTION_COL]
        )
        self._sum_tree.configure(columns=new_tree_cols)
        # Re-apply standard headings
        for db_col, label, width_px, is_mon, _ in TABLE_COLS:
            self._sum_tree.heading(db_col, text=label,
                command=lambda c=db_col: _sort_by(self, c))
            self._sum_tree.column(db_col, width=width_px, minwidth=60,
                anchor="e" if is_mon else "w", stretch=False)
        # Apply custom headings
        for db_col, display_label, _cid in custom_cols:
            self._sum_tree.heading(db_col, text=display_label)
            self._sum_tree.column(db_col, width=180, minwidth=80,
                anchor="w", stretch=False)
        # Re-apply edit column
        self._sum_tree.heading(_EDIT_ACTION_COL, text="Edit")
        self._sum_tree.column(_EDIT_ACTION_COL, width=64, minwidth=54,
            anchor="center", stretch=False)
    raw    = self._sum_search_var.get().strip()
    search = "" if "separate terms with commas" in raw else raw
    offset = self._sum_page * PAGE_SIZE
    rows, total = _db_query(
        search=search, session_id=self._sum_session_filter,
        sort_col=self._sum_sort_col, sort_asc=self._sum_sort_asc,
        offset=offset, limit=PAGE_SIZE,
        adv_filters=getattr(self, "_sum_adv_filters", {}))
    self._sum_total_rows = total
    _update_stats(self)
    _update_pagination(self, total)
    _render_tree(self, rows)
    shown_start = offset + 1 if total > 0 else 0
    shown_end   = min(offset + PAGE_SIZE, total)
    self._sum_count_lbl.config(
        text=f"{shown_start}–{shown_end} of {total} applicant(s)")


def _update_stats(self):
    raw    = self._sum_search_var.get().strip()
    search = "" if "separate terms with commas" in raw else raw
    tots   = _db_totals(session_id=self._sum_session_filter, search=search,
                        adv_filters=getattr(self, "_sum_adv_filters", {}))

    def _fmt(val):
        if val is None:
            return "—"
        try:
            return f"P{float(val):,.0f}"
        except Exception:
            return "—"

    self._sum_stat_labels["total"].config(text=str(tots.get("total", 0)))
    self._sum_stat_labels["done"].config(text=str(tots.get("done", 0)))
    self._sum_stat_labels["errors"].config(text=str(tots.get("errors", 0)))
    self._sum_stat_labels["income"].config(text=_fmt(tots.get("income")))
    self._sum_stat_labels["net"].config(text=_fmt(tots.get("net")))
    self._sum_stat_labels["amort_current"].config(text=_fmt(tots.get("amort_current")))


def _update_pagination(self, total: int):
    total_pages = max(1, -(-total // PAGE_SIZE))
    cur_page    = self._sum_page + 1
    self._sum_page_lbl.config(text=f"Page {cur_page} / {total_pages}")
    self._sum_prev_btn.configure(
        state="normal" if self._sum_page > 0 else "disabled")
    self._sum_next_btn.configure(
        state="normal" if cur_page < total_pages else "disabled")


def _page_prev(self):
    _teardown_summary_cell_entry(self, getattr(self, "_sum_active_cell_entry", None))
    if self._sum_page > 0:
        self._sum_page -= 1; _load_and_render(self)


def _page_next(self):
    _teardown_summary_cell_entry(self, getattr(self, "_sum_active_cell_entry", None))
    total_pages = max(1, -(-self._sum_total_rows // PAGE_SIZE))
    if self._sum_page + 1 < total_pages:
        self._sum_page += 1; _load_and_render(self)


def _fmt_money(val) -> str:
    if val in (None, ""):
        return "—"
    try:
        return f"P{float(val):,.2f}"
    except Exception:
        return str(val) or "—"


def _render_tree(self, rows):
    _teardown_summary_cell_entry(self, getattr(self, "_sum_active_cell_entry", None))
    self._sum_tree.delete(*self._sum_tree.get_children())
    self._sum_row_data = {}

    for i, row in enumerate(rows):
        row    = dict(row)
        row_id = row.get("id")

        # ── Parse results_json once for all virtual fields ─────────────
        try:
            results_blob = json.loads(row.get("results_json", "") or "{}")
        except Exception:
            results_blob = {}

        # ── Virtual columns — derived at render time ───────────────────
        row["spouse_info"]        = _extract_spouse_info(results_blob)
        row["personal_assets"]    = _extract_asset_items(results_blob, "cibi_personal_assets")
        row["business_assets"]    = _extract_asset_items(results_blob, "cibi_business_assets")
        row["business_inventory"] = _extract_asset_items(results_blob, "cibi_business_inventory")

        # ── amort_history_total: use real DB value; backfill if missing ─
        if row.get("amort_history_total") is None and results_blob:
            row["amort_history_total"] = _parse_amort_history_total(
                row.get("results_json", ""))

        tag    = "even" if i % 2 == 0 else "odd"
        values = []
        for db_col, label, width_px, is_monetary, is_text_block in TABLE_COLS:
            raw = row.get(db_col, "") or ""
            if is_monetary:
                values.append(_fmt_money(raw))
            elif is_text_block:
                values.append(str(raw).replace("\n", "  ·  "))
            else:
                values.append(str(raw))

        # ── NEW: append custom extraction columns ──────────────────────
        custom_cols = getattr(self, "_sum_custom_cols", [])
        for db_col, _label, _cid in custom_cols:
            values.append(str(row.get(db_col, "") or ""))

        values.append("✏")
        self._sum_tree.insert("", "end", iid=str(row_id),
                              values=values, tags=(tag,))
        self._sum_row_data[str(row_id)] = row


def _sort_by(self, col_key: str):
    _teardown_summary_cell_entry(self, getattr(self, "_sum_active_cell_entry", None))
    # ── Virtual columns cannot be sorted by DB; fall back to name ─────
    if col_key in _VIRTUAL_COLS:
        col_key = "applicant_name"

    if self._sum_sort_col == col_key:
        self._sum_sort_asc = not self._sum_sort_asc
    else:
        self._sum_sort_col = col_key
        self._sum_sort_asc = True

    self._sum_page = 0
    for db_col, label, _, __, ___ in TABLE_COLS:
        active = (self._sum_sort_col == db_col)
        ind = (" ▲" if active and self._sum_sort_asc else
               " ▼" if active else "")
        self._sum_tree.heading(db_col, text=label + ind)
    _load_and_render(self)


# ═══════════════════════════════════════════════════════════════════════
#  ROW INTERACTION (including CHANGE 3 + CHANGE 5)
# ═══════════════════════════════════════════════════════════════════════

def _teardown_summary_cell_entry(app, entry: tk.Entry | None) -> None:
    """Close an inline summary cell editor without running its FocusOut commit."""
    if entry is None:
        return
    try:
        if not entry.winfo_exists():
            return
        for seq in ("<FocusOut>", "<Return>", "<KP_Enter>", "<Escape>", "<Tab>"):
            try:
                entry.unbind(seq)
            except tk.TclError:
                pass
        entry.destroy()
    except tk.TclError:
        pass
    if getattr(app, "_sum_active_cell_entry", None) is entry:
        app._sum_active_cell_entry = None


def _set_summary_edit_lock(self, locked: bool, reason: str = "") -> None:
    """
    While locked, users can still sort/scroll/search, but cannot open a new
    inline cell editor until the pending DB save finishes.
    """
    self._sum_edit_locked = bool(locked)
    try:
        if locked:
            self._sum_tree.configure(cursor="watch")
            if hasattr(self, "_sum_save_status_lbl"):
                self._sum_save_status_lbl.config(
                    text="Saving to database…", fg=ACCENT_GOLD)
        else:
            self._sum_tree.configure(cursor="")
    except Exception:
        pass


def _summary_ui_dispatch(self, fn) -> None:
    """Enqueue a UI callback to run on the Tk main thread."""
    try:
        q = getattr(self, "_sum_bg_queue", None)
        if q is not None:
            q.put(fn)
            return
    except Exception:
        pass
    # Fallback path: if queue dispatch is unavailable, schedule directly.
    try:
        self.after(0, fn)
    except Exception:
        pass


def _start_summary_ui_poller(self) -> None:
    if getattr(self, "_sum_bg_poller_started", False):
        return
    self._sum_bg_poller_started = True

    def _poll():
        q = getattr(self, "_sum_bg_queue", None)
        if q is not None:
            try:
                while True:
                    fn = q.get_nowait()
                    try:
                        if callable(fn):
                            fn()
                    except Exception:
                        pass
            except queue.Empty:
                pass
        try:
            self.after(80, _poll)
        except Exception:
            self._sum_bg_poller_started = False

    try:
        self.after(80, _poll)
    except Exception:
        self._sum_bg_poller_started = False


def _summary_display_from_raw(col_id: str, raw_value: str) -> tuple[str, float | None]:
    """
    Convert user-edited raw text into:
      - display string (as shown in Treeview)
      - parsed numeric value for monetary cols (else None)
    Mirrors _db_update_cell formatting so UI can update immediately.
    """
    if col_id in _MONETARY_COLS:
        cleaned = re.sub(r"[^\d.]", "", (raw_value or "").replace(",", "").strip())
        if cleaned == "":
            return "—", None
        val = float(cleaned)  # may raise ValueError (handled by caller)
        return f"P{val:,.2f}", val
    txt = (raw_value or "").strip()
    return (txt if txt else "—"), None


def _summary_virtual_display_and_store(raw_value: str) -> tuple[str, str]:
    """
    For virtual cols, normalise user entry to the app's separator format.
    Returns:
      - display_val to show in the tree ("  ·  " separated, or "—")
      - store_raw to send to _db_update_virtual_cell (same format)
    """
    s = (raw_value or "").strip()
    if not s or s in ("—",):
        return "—", ""
    # Allow users to type multi-line; convert to separator format.
    s = s.replace("\r\n", "\n")
    parts = [p.strip() for p in re.split(r"\n+| {2}· {2}", s) if p.strip()]
    joined = "  ·  ".join(parts)
    return (joined if joined else "—"), (joined if joined else "")


def _summary_tree_cell_bbox(tree: ttk.Treeview, iid: str, col_name: str):
    """Bounding box for a data cell; retries after scroll. Returns (x,y,w,h) or None."""
    tree.focus(iid)
    tree.selection_set(iid)
    for _ in range(3):
        tree.update_idletasks()
        b = tree.bbox(iid, col_name)
        if b:
            return b
        tree.see(iid)
    try:
        n = TREE_COLS.index(col_name) + 1
        tree.update_idletasks()
        return tree.bbox(iid, f"#{n}")
    except (ValueError, tk.TclError):
        return None


def _summary_edit_values_equivalent(col_id: str, before: str, after: str) -> bool:
    """True if the inline editor value is unchanged from when editing started."""
    if col_id in _MONETARY_COLS:
        def _num_part(s: str) -> str:
            s = (s or "").strip()
            if s in ("—", "–", "-", ""):
                return ""
            return re.sub(r"[^\d.]", "", s.replace(",", ""))

        b, a = _num_part(before), _num_part(after)
        if not b and not a:
            return True
        try:
            return float(b) == float(a)
        except ValueError:
            return before.strip() == after.strip()
    b = (before or "").strip().replace("\r\n", "\n")
    a = (after or "").strip().replace("\r\n", "\n")
    return b == a


def _confirm_cell_modify(self) -> bool:
    """
    Frameless modal with custom light header (no OS title bar / no X).
    Shown after the user edits a cell and commits, before persisting to the DB.
    Returns True if the user clicks Confirm, False for Cancel.
    """
    result = [False]
    win = tk.Toplevel(self)
    win.configure(bg=CARD_WHITE)
    win.resizable(False, False)
    win.overrideredirect(True)
    win.transient(self)
    win.grab_set()
    win.lift(self)
    win.focus_force()

    p_x = self.winfo_rootx()
    p_y = self.winfo_rooty()
    p_w = self.winfo_width()
    p_h = self.winfo_height()
    w_w, w_h = 420, 168
    win.geometry(
        f"{w_w}x{w_h}+{p_x + (p_w - w_w) // 2}+{p_y + (p_h - w_h) // 2}")

    shell = tk.Frame(win, bg=BORDER_MID)
    shell.pack(fill="both", expand=True, padx=1, pady=1)
    root = tk.Frame(shell, bg=CARD_WHITE)
    root.pack(fill="both", expand=True)

    def _finish(ok: bool):
        result[0] = ok
        try:
            win.grab_release()
        except Exception:
            pass
        try:
            win.destroy()
        except Exception:
            pass

    hdr = tk.Frame(root, bg="#E8EEF8")
    hdr.pack(fill="x")
    tk.Label(
        hdr,
        text="Confirm edit",
        font=("Segoe UI", 10, "bold"),
        fg=NAVY_MID,
        bg="#E8EEF8",
        padx=14,
        pady=8,
        anchor="w",
    ).pack(fill="x")

    body = tk.Frame(root, bg=CARD_WHITE)
    body.pack(fill="both", expand=True, padx=18, pady=(12, 14))
    tk.Label(
        body,
        text="Are you sure you want to modify this cell?",
        font=("Segoe UI", 10),
        fg=TXT_NAVY,
        bg=CARD_WHITE,
        wraplength=380,
        justify="left",
    ).pack(anchor="w")

    btn_f = tk.Frame(body, bg=CARD_WHITE)
    btn_f.pack(fill="x", pady=(16, 0))

    ctk.CTkButton(
        btn_f, text="Cancel", command=lambda: _finish(False),
        width=92, height=30, corner_radius=6,
        fg_color="#E8ECF2", hover_color="#DDE2EA",
        text_color=TXT_NAVY, font=FF(8, "bold"),
    ).pack(side="right", padx=(8, 0))
    ctk.CTkButton(
        btn_f, text="Confirm", command=lambda: _finish(True),
        width=92, height=30, corner_radius=6,
        fg_color=LIME_MID, hover_color=LIME_BRIGHT,
        text_color=TXT_ON_LIME, font=FF(8, "bold"),
    ).pack(side="right")

    win.bind("<Escape>", lambda e: _finish(False))

    try:
        # Even if something goes wrong, never leave a global grab behind.
        win.wait_window(win)
    finally:
        try:
            win.grab_release()
        except Exception:
            pass
    return result[0]

def _ask_edit_reason(self, col_label: str) -> str | None:
    """
    Show a pop-up asking the user for a reason before submitting an edit for approval.
    Returns the reason string, or None if the user cancelled.
    """
    result = [None]
    win = tk.Toplevel(self)
    win.configure(bg=CARD_WHITE)
    win.resizable(False, False)
    win.overrideredirect(True)
    win.transient(self)
    win.grab_set()
    win.lift(self)
    win.focus_force()

    p_x = self.winfo_rootx()
    p_y = self.winfo_rooty()
    p_w = self.winfo_width()
    p_h = self.winfo_height()
    w_w, w_h = 460, 230
    win.geometry(f"{w_w}x{w_h}+{p_x + (p_w - w_w) // 2}+{p_y + (p_h - w_h) // 2}")

    shell = tk.Frame(win, bg=BORDER_MID)
    shell.pack(fill="both", expand=True, padx=1, pady=1)
    root = tk.Frame(shell, bg=CARD_WHITE)
    root.pack(fill="both", expand=True)

    hdr = tk.Frame(root, bg="#E8EEF8")
    hdr.pack(fill="x")
    tk.Label(hdr, text=f"Reason for changing: {col_label}",
             font=("Segoe UI", 10, "bold"), fg=NAVY_MID,
             bg="#E8EEF8", padx=14, pady=8, anchor="w").pack(fill="x")

    body = tk.Frame(root, bg=CARD_WHITE)
    body.pack(fill="both", expand=True, padx=18, pady=(10, 0))

    tk.Label(body, text="Please provide a reason. Your edit will be sent for approval.",
             font=("Segoe UI", 9), fg=TXT_MUTED, bg=CARD_WHITE,
             wraplength=420, justify="left").pack(anchor="w")

    reason_var = tk.StringVar()
    reason_entry = tk.Entry(body, textvariable=reason_var,
                            font=("Segoe UI", 10), fg=TXT_NAVY, bg=WHITE,
                            relief="solid", bd=1, insertbackground=NAVY_MID)
    reason_entry.pack(fill="x", pady=(8, 0), ipady=5)
    reason_entry.focus_set()

    err_lbl = tk.Label(body, text="", font=("Segoe UI", 8),
                       fg=ACCENT_RED, bg=CARD_WHITE)
    err_lbl.pack(anchor="w")

    btn_f = tk.Frame(body, bg=CARD_WHITE)
    btn_f.pack(fill="x", pady=(8, 0))

    def _submit():
        if not reason_var.get().strip():
            err_lbl.config(text="Reason is required.")
            reason_entry.focus_set()
            return
        result[0] = reason_var.get().strip()
        try: win.grab_release()
        except Exception: pass
        win.destroy()

    def _cancel():
        result[0] = None
        try: win.grab_release()
        except Exception: pass
        win.destroy()

    ctk.CTkButton(btn_f, text="Cancel", command=_cancel,
                  width=92, height=30, corner_radius=6,
                  fg_color="#E8ECF2", hover_color="#DDE2EA",
                  text_color=TXT_NAVY, font=FF(8, "bold")).pack(side="right", padx=(8, 0))
    ctk.CTkButton(btn_f, text="📨  Send for Approval", command=_submit,
                  width=150, height=30, corner_radius=6,
                  fg_color=LIME_MID, hover_color=LIME_BRIGHT,
                  text_color=TXT_ON_LIME, font=FF(8, "bold")).pack(side="right")

    win.bind("<Return>", lambda e: _submit())
    win.bind("<Escape>", lambda e: _cancel())

    try:
        win.wait_window(win)
    finally:
        try: win.grab_release()
        except Exception: pass

    return result[0]

def _db_submit_edit_request(row_id: int, applicant_name: str,
                             col_name: str, old_value: str,
                             new_value: str, reason: str,
                             requested_by: str):
    with _db_connect() as conn:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO edit_requests "
            "(applicant_id, applicant_name, col_name, old_value, new_value, "
            " reason, requested_by) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s)",
            (row_id, applicant_name, col_name,
             old_value, new_value, reason, requested_by)
        )
        conn.commit()
        cur.close()
def _db_submit_delete_column_request(col_name: str, reason: str,
                                      requested_by: str):
    """Submit a column deletion request for admin approval."""
    with _db_connect() as conn:
        cur = conn.cursor()
        # Use 0 as a sentinel applicant_id for schema-level operations
        # (avoids NOT NULL constraint; 0 never matches a real applicant row)
        cur.execute(
            "INSERT INTO edit_requests "
            "(applicant_id, applicant_name, col_name, old_value, new_value, "
            " reason, requested_by) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s)",
            (0, "[COLUMN DELETION]", col_name,
             col_name, "__DROP_COLUMN__", reason, requested_by)
        )
        conn.commit()
        cur.close()


# ── CHANGE 3 (add _start_cell_edit) ───────────────────────────────────
def _start_cell_edit(self, iid: str, col_id: str):
    """
    Pop a lightweight Entry widget directly over the clicked Treeview
    cell so the user can type a new value.

    Behaviour:
      <Return> or <KP_Enter>  — confirm dialog (if changed), then save and close
      <Tab>                   — same, then move to next editable column
      <Escape>                — cancel, restore original value
      FocusOut                — same as Return (if focus leaves the cell)
    """
    if col_id not in _EDITABLE_COLS:
        return
    if getattr(self, "_sum_edit_locked", False):
        # Recover from stale lock state if no save is actually running.
        if getattr(self, "_sum_save_inflight", False):
            return
        _set_summary_edit_lock(self, False, "stale_lock_recover")

    tree   = self._sum_tree
    row_id = int(iid)

    prev_entry = getattr(self, "_sum_active_cell_entry", None)
    if prev_entry is not None:
        _teardown_summary_cell_entry(self, prev_entry)

    # ── Get the cell's pixel bounding box ─────────────────────────────
    bbox = _summary_tree_cell_bbox(tree, iid, col_id)
    if not bbox:
        return
    x, y, width, height = bbox

    # ── Current raw value ─────────────────────────────────────────────
    col_index   = TREE_COLS.index(col_id)
    cur_display = tree.item(iid, "values")[col_index]

    if col_id in _MONETARY_COLS:
        # Strip "P" prefix, commas, and "—" placeholder
        cur_edit = re.sub(r"[P,]", "", cur_display).strip()
        if cur_edit == "—":
            cur_edit = ""
    elif col_id in _VIRTUAL_TO_JSON or col_id in {
            "income_items", "business_items", "household_items"}:
        # Tree renders newlines as '  ·  '; restore for editing
        cur_edit = cur_display.replace("  ·  ", "\n") \
            if cur_display not in ("—", "") else ""
    else:
        cur_edit = "" if cur_display in ("—", "") else cur_display

    # ── Build the overlay Entry ────────────────────────────────────────
    var   = tk.StringVar(value=cur_edit)
    entry = tk.Entry(
        tree,
        textvariable=var,
        font=("Segoe UI", 9),
        fg=TXT_NAVY,
        bg="#FFFDE7",                   # soft yellow — "edit mode" signal
        insertbackground=NAVY_MID,
        relief="solid",
        bd=1,
        highlightthickness=1,
        highlightbackground=LIME_MID,   # green border, matches app theme
        highlightcolor=LIME_BRIGHT,
    )
    entry.place(x=x, y=y, width=max(width, 120), height=height)
    self._sum_active_cell_entry = entry
    entry.focus_set()
    entry.select_range(0, "end")

    # Pause mousewheel on the treeview while the Entry is open so the
    # row cannot scroll out from under the widget.
    tree.unbind_all("<MouseWheel>")

    committed = [False]   # guard against double-commit on FocusOut + Return

    def _commit(event=None):
        if committed[0]:
            return
        new_raw = var.get()

        if _summary_edit_values_equivalent(col_id, cur_edit, new_raw):
            committed[0] = True
            entry.destroy()
            return

        if not _confirm_cell_modify(self):
            _cancel()
            return

        # ── Ask for reason; submit for approval instead of direct write ──
        col_label = next(
            (label for db_col, label, *_ in TABLE_COLS if db_col == col_id),
            col_id)
        reason = _ask_edit_reason(self, col_label)
        if reason is None:
            _cancel()
            return

        # Queue the edit request and show pending notice — do NOT write to DB yet
        committed[0] = True
        applicant = (self._sum_row_data.get(iid, {})
                     .get("applicant_name", "") or f"id={row_id}")
        requested_by = getattr(self, "_current_username", "") or "unknown"

        try:
            if col_id in _VIRTUAL_TO_JSON:
                display_val, store_raw = _summary_virtual_display_and_store(new_raw)
            else:
                display_val, _ = _summary_display_from_raw(col_id, new_raw)
                store_raw = new_raw
        except ValueError as exc:
            entry.config(bg="#FFE0E0", highlightbackground=ACCENT_RED)
            messagebox.showerror("Invalid Value", str(exc))
            entry.focus_set()
            return

        entry.destroy()

        def _submit_worker():
            try:
                _db_submit_edit_request(
                    row_id, applicant, col_id,
                    cur_display, display_val, reason, requested_by)
                def _done():
                    if hasattr(self, "_sum_save_status_lbl"):
                        self._sum_save_status_lbl.config(
                            text="Edit submitted for approval", fg=ACCENT_GOLD)
                        try:
                            if getattr(self, "_sum_save_status_after", None):
                                self.after_cancel(self._sum_save_status_after)
                        except Exception:
                            pass
                        self._sum_save_status_after = self.after(
                            3000, lambda: self._sum_save_status_lbl.config(text=""))
                    _log_action(self, "edit_requested",
                                f"[{col_id}] '{cur_display}' → '{display_val}' "
                                f"reason='{reason}' ({applicant})")
                _summary_ui_dispatch(self, _done)
            except Exception as exc:
                err = str(exc)
                _summary_ui_dispatch(self,
                    lambda: messagebox.showerror("Submit Failed", err))

        threading.Thread(target=_submit_worker, daemon=True).start()
        return   # stop here — skip the rest of _commit (no live DB write)

        # ── Optimistic UI update (fast) + background DB write (no UI freeze) ──

        # ── Optimistic UI update (fast) + background DB write (no UI freeze) ──
        try:
            if col_id in _VIRTUAL_TO_JSON:
                display_val, store_raw = _summary_virtual_display_and_store(new_raw)
                parsed_num = None
            else:
                display_val, parsed_num = _summary_display_from_raw(col_id, new_raw)
                store_raw = new_raw
        except ValueError as exc:
            entry.config(bg="#FFE0E0", highlightbackground=ACCENT_RED)
            messagebox.showerror("Invalid Value", str(exc))
            entry.focus_set()
            return

        committed[0] = True

        # Update just this one cell in the treeview immediately — no full reload
        vals = list(tree.item(iid, "values"))
        vals[col_index] = display_val
        tree.item(iid, values=vals)

        # Patch the in-memory row cache optimistically so detail window stays current
        if iid in self._sum_row_data:
            if col_id in _MONETARY_COLS:
                self._sum_row_data[iid][col_id] = parsed_num
            elif col_id in _VIRTUAL_TO_JSON:
                self._sum_row_data[iid][col_id] = (store_raw.strip() or None)
                try:
                    blob = json.loads(self._sum_row_data[iid].get("results_json") or "{}")
                    items = [p.strip() for p in (store_raw or "").split("  ·  ") if p.strip()]
                    if col_id == "spouse_info":
                        plain  = [it for it in items if not it.startswith("Office: ")]
                        office = [it[len("Office: "):] for it in items if it.startswith("Office: ")]
                        blob.setdefault("cibi_spouse", {})["items"] = plain
                        blob.setdefault("cibi_spouse_office", {})["items"] = office
                    else:
                        json_key, sub_key = _VIRTUAL_TO_JSON[col_id]
                        blob.setdefault(json_key, {})[sub_key] = items
                    self._sum_row_data[iid]["results_json"] = json.dumps(blob, ensure_ascii=False)
                except Exception:
                    pass
            else:
                self._sum_row_data[iid][col_id] = (store_raw.strip() or None)

        entry.destroy()
        _update_stats(self)

        self._sum_save_inflight = True
        _set_summary_edit_lock(self, True, "saving")
        t0 = time.perf_counter()
        prior_display = cur_display
        prior_raw_cache = None
        if iid in self._sum_row_data:
            prior_raw_cache = self._sum_row_data[iid].get(col_id)

        def _worker():
            err = None
            try:
                if col_id in _VIRTUAL_TO_JSON:
                    _db_update_virtual_cell(row_id, col_id, store_raw)
                else:
                    _db_update_cell(row_id, col_id, store_raw)
            except Exception as e:
                err = e

            def _done():
                self._sum_save_inflight = False
                _set_summary_edit_lock(self, False)
                if err:
                    # Revert UI + cache on failure
                    try:
                        vals2 = list(tree.item(iid, "values"))
                        vals2[col_index] = prior_display
                        tree.item(iid, values=vals2)
                    except Exception:
                        pass
                    if iid in self._sum_row_data:
                        self._sum_row_data[iid][col_id] = prior_raw_cache
                    messagebox.showerror("Save Failed", str(err))
                else:
                    # ── Audit log on success ───────────────────────────
                    try:
                        applicant = (self._sum_row_data.get(iid, {})
                                     .get("applicant_name", "") or f"id={row_id}")
                        _log_action(self, "edit_cell",
                                    f"[{col_id}] '{prior_display}' → '{display_val}'  ({applicant})")
                    except Exception:
                        pass
                    if hasattr(self, "_sum_save_status_lbl"):
                        ms = int((time.perf_counter() - t0) * 1000)
                        self._sum_save_status_lbl.config(
                            text=f"Saved to database ({ms} ms)", fg=ACCENT_SUCCESS)
                        try:
                            if getattr(self, "_sum_save_status_after", None):
                                self.after_cancel(self._sum_save_status_after)
                        except Exception:
                            pass
                        self._sum_save_status_after = self.after(
                            1600, lambda: self._sum_save_status_lbl.config(text=""))

            _summary_ui_dispatch(self, _done)

        threading.Thread(target=_worker, daemon=True).start()

    def _cancel(event=None):
        committed[0] = True
        # Defensive cleanup in case a prior failed callback left a stale lock.
        if not getattr(self, "_sum_save_inflight", False):
            _set_summary_edit_lock(self, False, "cancel_cleanup")
        entry.destroy()

    def _on_focus_out(event=None):
        # Auto-commit on blur so edits are saved/logged even without Enter.
        # Defer by one tick so focus settles first (avoids duplicate commits).
        if committed[0]:
            return
        self.after_idle(lambda: (None if committed[0] else _commit()))

    def _tab_next(event=None):
        """Commit and jump to the next editable column on the same row."""
        _commit()
        if not committed[0]:
            return "break"   # validation error or confirm cancelled
        # Find the next editable column to the right
        start = col_index + 1
        for next_idx in range(start, len(TREE_COLS)):
            next_col = TREE_COLS[next_idx]
            if next_col in _EDITABLE_COLS:
                self.after(50, lambda c=next_col: _start_cell_edit(self, iid, c))
                break
        return "break"   # suppress default Tab focus behaviour

    entry.bind("<Return>",   _commit)
    entry.bind("<KP_Enter>", _commit)
    entry.bind("<Escape>",   _cancel)
    entry.bind("<Tab>",      _tab_next)
    entry.bind("<FocusOut>", _on_focus_out)

    # Restore mousewheel when the entry is destroyed
    def _on_entry_destroy(e):
        try:
            tree.bind_all(
                "<MouseWheel>",
                lambda ev: tree.yview_scroll(
                    int(-1 * (ev.delta / 120)), "units"))
        except tk.TclError:
            pass
        if getattr(self, "_sum_active_cell_entry", None) is entry:
            self._sum_active_cell_entry = None

    entry.bind("<Destroy>", _on_entry_destroy)


def _open_row_edit_dialog(self, row_id: int):
    """Open a form dialog to edit all summary table columns for one row."""
    iid = str(row_id)
    row = dict(getattr(self, "_sum_row_data", {}).get(iid, {}) or {})
    if not row:
        messagebox.showerror("Edit Row", "Could not load row data.")
        return

    dlg = tk.Toplevel(self)
    dlg.title(f"Edit Row — {row.get('applicant_name', '') or f'id={row_id}'}")
    dlg.configure(bg=CARD_WHITE)
    dlg.resizable(True, True)
    dlg.transient(self)
    dlg.grab_set()

    p_x = self.winfo_rootx()
    p_y = self.winfo_rooty()
    p_w = self.winfo_width()
    p_h = self.winfo_height()
    w_w, w_h = 820, 700
    dlg.geometry(
        f"{w_w}x{w_h}+{p_x + (p_w - w_w) // 2}+{p_y + (p_h - w_h) // 2}")
    dlg.minsize(700, 520)

    hdr = tk.Frame(dlg, bg=NAVY_DEEP)
    hdr.pack(fill="x")
    tk.Label(
        hdr,
        text=f"✏  Edit Summary Row  ·  ID {row_id}",
        font=("Segoe UI", 12, "bold"),
        fg=WHITE,
        bg=NAVY_DEEP,
        padx=16,
        pady=10,
        anchor="w",
    ).pack(fill="x")

    outer = tk.Frame(dlg, bg=CARD_WHITE)
    outer.pack(fill="both", expand=True, padx=16, pady=12)
    canvas = tk.Canvas(outer, bg=CARD_WHITE, highlightthickness=0)
    vscroll = tk.Scrollbar(outer, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=vscroll.set)
    vscroll.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)

    form = tk.Frame(canvas, bg=CARD_WHITE)
    cwin = canvas.create_window((0, 0), window=form, anchor="nw")

    def _on_cfg(_e=None):
        canvas.configure(scrollregion=canvas.bbox("all"))
        canvas.itemconfig(cwin, width=canvas.winfo_width())

    form.bind("<Configure>", _on_cfg)
    canvas.bind("<Configure>", _on_cfg)

    widgets: dict[str, object] = {}
    label_by_col = {db_col: label for db_col, label, *_ in TABLE_COLS}
    list_like_cols = {"income_items", "business_items", "household_items"}

    for idx, (db_col, label, _w, _is_mon, is_text_block) in enumerate(TABLE_COLS):
        row_bg = ROW_BG_EVEN if idx % 2 == 0 else ROW_BG_ODD
        line = tk.Frame(form, bg=row_bg, highlightbackground="#E5EAF3", highlightthickness=1)
        line.pack(fill="x", pady=1)
        tk.Label(
            line, text=f"{label}:", font=F(8, "bold"), fg=NAVY_DEEP,
            bg=row_bg, width=22, anchor="w", padx=8, pady=6
        ).pack(side="left")

        raw = row.get(db_col, "")
        if db_col in _MONETARY_COLS:
            init_val = "" if raw in (None, "") else str(raw)
        else:
            init_val = "" if raw in (None, "") else str(raw)

        if db_col in _VIRTUAL_TO_JSON or db_col in list_like_cols:
            init_val = init_val.replace("  ·  ", "\n")

        if is_text_block or db_col in _VIRTUAL_TO_JSON or db_col in list_like_cols:
            txt = tk.Text(
                line, height=3, font=("Segoe UI", 9), fg=TXT_NAVY, bg=WHITE,
                insertbackground=NAVY_MID, relief="solid", bd=1, wrap="word"
            )
            txt.pack(side="left", fill="x", expand=True, padx=(0, 8), pady=6)
            txt.insert("1.0", init_val)
            widgets[db_col] = txt
        else:
            var = tk.StringVar(value=init_val)
            ent = tk.Entry(
                line, textvariable=var, font=("Segoe UI", 9), fg=TXT_NAVY, bg=WHITE,
                insertbackground=NAVY_MID, relief="solid", bd=1
            )
            ent.pack(side="left", fill="x", expand=True, padx=(0, 8), pady=6)
            widgets[db_col] = var

    btn_row = tk.Frame(dlg, bg=CARD_WHITE)
    btn_row.pack(fill="x", padx=16, pady=(0, 14))

    def _read_raw(col: str) -> str:
        w = widgets[col]
        if isinstance(w, tk.StringVar):
            val = w.get().strip()
        else:
            val = w.get("1.0", "end-1c").strip()
        if col in _VIRTUAL_TO_JSON or col in list_like_cols:
            parts = [p.strip() for p in val.splitlines() if p.strip()]
            return "  ·  ".join(parts)
        return val

    def _save():
        if getattr(self, "_sum_edit_locked", False) and getattr(self, "_sum_save_inflight", False):
            messagebox.showinfo("Please wait", "A previous save is still in progress.")
            return

        changes: list[tuple[str, str]] = []
        for db_col, _label, _w, _is_mon, _is_txt in TABLE_COLS:
            before_raw = "" if row.get(db_col) is None else str(row.get(db_col))
            after_raw = _read_raw(db_col)
            if db_col in _VIRTUAL_TO_JSON or db_col in list_like_cols:
                before_raw = before_raw.replace("\n", "  ·  ")
            if not _summary_edit_values_equivalent(db_col, before_raw, after_raw):
                # Format old value the same way as display so the approval panel shows "P12,345.67" not "12345.67"
                if db_col in _MONETARY_COLS:
                    try:
                        old_display = f"P{float(before_raw):,.2f}" if before_raw.strip() else "—"
                    except Exception:
                        old_display = before_raw
                elif db_col in _VIRTUAL_TO_JSON or db_col in list_like_cols:
                    old_display = before_raw.replace("\n", "  ·  ")
                else:
                    old_display = before_raw or "—"
                changes.append((db_col, after_raw, old_display))

        if not changes:
            dlg.destroy()
            return

        # ── Ask reason once for all changes ──────────────────────────────
        col_labels_changed = ", ".join(
            label_by_col.get(db_col, db_col) for db_col, _, __ in changes)
        reason = _ask_edit_reason(self, col_labels_changed)
        if reason is None:
            return   # user cancelled — keep dialog open

        requested_by = getattr(self, "_current_username", "") or "unknown"
        applicant    = row.get("applicant_name", "") or f"id={row_id}"

        def _submit_worker():
            errors = []
            for db_col, new_raw, old_raw in changes:
                try:
                    if db_col in _VIRTUAL_TO_JSON:
                        display_val, _ = _summary_virtual_display_and_store(new_raw)
                        old_display    = old_raw
                    else:
                        display_val, _ = _summary_display_from_raw(db_col, new_raw)
                        old_display    = old_raw

                    _db_submit_edit_request(
                        row_id, applicant, db_col,
                        old_display, display_val, reason, requested_by)
                except Exception as exc:
                    errors.append(f"{db_col}: {exc}")

            def _done():
                if errors:
                    messagebox.showerror("Submit Failed",
                                        "Some fields failed:\n" + "\n".join(errors))
                else:
                    if hasattr(self, "_sum_save_status_lbl"):
                        self._sum_save_status_lbl.config(
                            text="Edit(s) submitted for approval", fg=ACCENT_GOLD)
                        try:
                            if getattr(self, "_sum_save_status_after", None):
                                self.after_cancel(self._sum_save_status_after)
                        except Exception:
                            pass
                        self._sum_save_status_after = self.after(
                            3000, lambda: self._sum_save_status_lbl.config(text=""))
                    _log_action(self, "edit_requested",
                                f"({applicant}) — {len(changes)} field(s) submitted via row edit dialog "
                                f"reason='{reason}'")
                dlg.destroy()

            _summary_ui_dispatch(self, _done)

        threading.Thread(target=_submit_worker, daemon=True).start()

    ctk.CTkButton(
        btn_row, text="💾  Save Changes", command=_save,
        width=150, height=32, corner_radius=7,
        fg_color=LIME_MID, hover_color=LIME_BRIGHT,
        text_color=TXT_ON_LIME, font=FF(9, "bold")
    ).pack(side="right")
    ctk.CTkButton(
        btn_row, text="Cancel", command=dlg.destroy,
        width=110, height=32, corner_radius=7,
        fg_color="#E8ECF2", hover_color="#DDE2EA",
        text_color=TXT_NAVY, font=FF(9, "bold")
    ).pack(side="right", padx=(0, 8))


# ── CHANGE 5 (replace _on_tree_double_click and add _on_tree_return_key) ──
def _on_tree_double_click(self, event):
    """
    Double-click routing:
      • Editable column   → open inline cell editor
      • Non-editable col  → open the detail window (original behaviour)
    """
    tree   = self._sum_tree
    iid    = tree.identify_row(event.y)
    col_id = tree.identify_column(event.x)   # e.g. "#3"

    if not iid:
        return

    # Convert Treeview "#N" index to db_col name
    try:
        col_index = int(col_id.lstrip("#")) - 1
        db_col    = TREE_COLS[col_index]
    except (ValueError, IndexError):
        db_col = ""

    if db_col == _EDIT_ACTION_COL:
        _open_row_edit_dialog(self, int(iid))
        return
    row = self._sum_row_data.get(iid)
    if row:
        _open_detail_window(self, row)


def _on_tree_return_key(self, event):
    """
    Pressing <Return> on a focused row opens the inline editor on the
    first editable column (applicant_name), letting power users navigate
    and edit without the mouse.
    """
    iid = self._sum_tree.focus()
    if iid:
        row = self._sum_row_data.get(iid)
        if row:
            _open_detail_window(self, row)


def _on_tree_right_click(self, event):
    iid = self._sum_tree.identify_row(event.y)
    if not iid:
        return
    self._sum_tree.selection_set(iid)
    self._sum_tree.focus(iid)
    menu = tk.Menu(self._sum_tree, tearoff=0, bg=CARD_WHITE, fg=TXT_NAVY,
                   activebackground=NAVY_MIST, font=("Segoe UI", 9))
    menu.add_command(label="✏  Edit Row", command=lambda: _open_row_edit_dialog(self, int(iid)))
    menu.add_separator()
    menu.add_command(label="👁  View Details",
                     command=lambda: _open_detail_window(
                         self, self._sum_row_data.get(iid, {})))
    menu.add_separator()
    menu.add_command(label="✕  Delete",
                     command=lambda: _delete_row(self, int(iid)))
    try:
        menu.tk_popup(event.x_root, event.y_root)
    finally:
        menu.grab_release()


def _open_detail_window(self, row: dict):
    win = tk.Toplevel(self)
    win.title(f"Detail — {row.get('applicant_name', '')}")
    win.geometry("780x640")
    win.configure(bg=CARD_WHITE)
    win.grab_set()
    canvas  = tk.Canvas(win, bg=CARD_WHITE, highlightthickness=0)
    vscroll = tk.Scrollbar(win, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=vscroll.set)
    vscroll.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)
    body = tk.Frame(canvas, bg=CARD_WHITE)
    cwin = canvas.create_window((0, 0), window=body, anchor="nw")

    def _on_cfg(e):
        canvas.configure(scrollregion=canvas.bbox("all"))
        canvas.itemconfig(cwin, width=canvas.winfo_width())

    body.bind("<Configure>", _on_cfg)
    canvas.bind("<Configure>", _on_cfg)
    canvas.bind_all("<MouseWheel>",
                    lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))
    _build_detail_panel(self, row, body)

    def _close():
        # Prevent "stuck grab" UI lockups after closing the detail window.
        try:
            win.grab_release()
        except Exception:
            pass
        try:
            win.destroy()
        except Exception:
            pass

    win.protocol("WM_DELETE_WINDOW", _close)
    win.bind("<Escape>", lambda e: _close())

    ctk.CTkButton(body, text="Close", command=_close,
                  width=100, height=32, corner_radius=7,
                  fg_color=NAVY_LIGHT, hover_color=NAVY_PALE,
                  text_color=WHITE, font=FF(9, "bold")).pack(pady=(12, 16))


def _build_detail_panel(self, row: dict, parent: tk.Frame):
    info_fields = [
        ("Client ID",            row.get("client_id",         "—") or "—"),
        ("PN",                   row.get("pn",                "—") or "—"),
        ("Applicant",            row.get("applicant_name",    "—") or "—"),
        ("Industry",             row.get("industry_name",     "—") or "—"),
        ("Residence Address",    row.get("residence_address", "—") or "—"),
        ("Office Address",       row.get("office_address",    "—") or "—"),
        ("Loan Balance",         _fmt_money(row.get("loan_balance"))   or "—"),
        ("Principal Loan",       _fmt_money(row.get("principal_loan")) or "—"),
        ("Maturity",             row.get("maturity",          "—") or "—"),
        ("Interest Rate",        row.get("interest_rate",     "—") or "—"),
        ("Branch",               row.get("branch",            "—") or "—"),
        ("Loan Class",           row.get("loan_class_name",   "—") or "—"),
        ("Product Name",         row.get("product_name",      "—") or "—"),
        ("Loan Date",            row.get("loan_date",         "—") or "—"),
        ("Term Unit",            row.get("term_unit",         "—") or "—"),
        ("Term",                 row.get("term",              "—") or "—"),
        ("Security",             row.get("security",          "—") or "—"),
        ("Release Tag",          row.get("release_tag",       "—") or "—"),
        ("Loan Amount",          _fmt_money(row.get("loan_amount"))         or "—"),
        ("Loan Status",          row.get("loan_status",       "—") or "—"),
        ("AO Name",              row.get("ao_name",           "—") or "—"),
        ("Amort. History",       _fmt_money(row.get("amort_history_total")) or "—"),
        ("Curr. Amort.",         _fmt_money(row.get("amort_current_total")) or "—"),
        ("Source File",          row.get("source_file",       "—") or "—"),
        ("Processed At",         (row.get("processed_at", "") or "")[:16].replace("T", "  ")),
        ("Session",              (row.get("session_id",   "") or "")[:19].replace("T", "  ")),
    ]
    # Keep the same table-like UI style across all sections, including metadata.
    PAD_X        = 16
    row_idx      = 0
    sec_bar = tk.Frame(parent, bg=SEC_BG)
    sec_bar.pack(fill="x", padx=PAD_X, pady=(8, 0))
    tk.Label(sec_bar, text="  CLIENT PROFILE",
             font=F(8, "bold"), fg=SEC_FG, bg=SEC_BG, pady=4).pack(side="left")

    def _render_profile_row(field_label: str, value: str):
        nonlocal row_idx
        row_bg = ROW_BG_EVEN if row_idx % 2 == 0 else ROW_BG_ODD
        row_idx += 1
        row_f = tk.Frame(parent, bg=row_bg,
                         highlightbackground="#E5EAF3", highlightthickness=1)
        row_f.pack(fill="x", padx=PAD_X)
        tk.Label(row_f, text=field_label, font=F(8, "bold"), fg=NAVY_DEEP,
                 bg=row_bg, padx=8, pady=6, anchor="w", width=26).pack(side="left")
        tk.Label(row_f, text="—", font=F(8),
                 fg=TXT_MUTED, bg=row_bg, padx=8, width=14, anchor="e").pack(side="left")
        tk.Label(row_f, text=str(value or "—"), font=F(8),
                 fg=TXT_NAVY if str(value or "").strip() and str(value) != "—" else TXT_MUTED,
                 bg=row_bg, padx=8, anchor="w",
                 wraplength=440, justify="left").pack(side="left", fill="x", expand=True)

    for field_label, value in info_fields:
        _render_profile_row(field_label, value)

    if row.get("petrol_risk"):
        _render_profile_row("Petrol Risk", "⚠ Petrol Risk")
    if row.get("transport_risk"):
        _render_profile_row("Transport Risk", "⚠ Transport Risk")

    try:
        results = json.loads(row.get("results_json", "") or "{}")
    except Exception:
        results = {}

    last_section = None
    # Continue row striping from metadata block for a single visual rhythm.

    def _normalize_items(raw_items):
        out = []
        for it in (raw_items or []):
            s = str(it or "").strip()
            if s:
                out.append(s)
        return out

    def _pretty_label_from_key(key: str) -> str:
        s = str(key or "").strip().replace("_", " ")
        return " ".join(part.capitalize() for part in s.split()) if s else "Unnamed Parameter"

    def _render_param_row(section: str, field_label: str, items, total, non_monetary: bool):
        nonlocal last_section, row_idx
        if section != last_section:
            last_section = section
            sec_bar = tk.Frame(parent, bg=SEC_BG)
            sec_bar.pack(fill="x", padx=PAD_X, pady=(8, 0))
            tk.Label(sec_bar, text=f"  {section.upper()}",
                     font=F(8, "bold"), fg=SEC_FG, bg=SEC_BG, pady=4).pack(side="left")

        row_bg = ROW_BG_EVEN if row_idx % 2 == 0 else ROW_BG_ODD
        row_idx += 1
        row_f = tk.Frame(parent, bg=row_bg,
                         highlightbackground="#E5EAF3", highlightthickness=1)
        row_f.pack(fill="x", padx=PAD_X)
        tk.Label(row_f, text=field_label, font=F(8, "bold"), fg=NAVY_DEEP,
                 bg=row_bg, padx=8, pady=6, anchor="w", width=26).pack(side="left")
        amt_txt = total if (total and not non_monetary) else "—"
        tk.Label(row_f, text=amt_txt,
                 font=F(9, "bold") if amt_txt != "—" else F(8),
                 fg=NAVY_MID if amt_txt != "—" else TXT_MUTED,
                 bg=row_bg, padx=8, width=14, anchor="e").pack(side="left")
        items = _normalize_items(items)
        det_txt = ("\n".join(f"• {it}" for it in items) if items else "No data found")
        tk.Label(row_f, text=det_txt, font=F(8),
                 fg=TXT_NAVY if items else TXT_MUTED,
                 bg=row_bg, padx=8, anchor="w",
                 wraplength=440, justify="left").pack(side="left", fill="x", expand=True)

    known_lookup_keys = {k for (k, _s, _lbl) in LOOKUP_ROWS}

    # Render canonical rows first, preserving existing section order.
    for key, section, field_label in LOOKUP_ROWS:
        field_data = results.get(key, {})
        items = field_data.get("items", []) if isinstance(field_data, dict) else []
        total = field_data.get("total", "") if isinstance(field_data, dict) else ""
        _render_param_row(section, field_label, items, total, key in NON_MONETARY)

    # Render additional parameters (if any) using the same summary row format.
    extra_rows = []
    for key, val in (results or {}).items():
        if key in known_lookup_keys or str(key).startswith("_"):
            continue
        if isinstance(val, dict):
            items = val.get("items", [])
            total = val.get("total", "")
            non_monetary = not bool(str(total or "").strip())
        elif isinstance(val, list):
            items = val
            total = ""
            non_monetary = True
        else:
            raw = str(val or "").strip()
            items = [raw] if raw else []
            total = ""
            non_monetary = True
        extra_rows.append(("Other Parameters", _pretty_label_from_key(key), items, total, non_monetary))

    for section, field_label, items, total, non_monetary in sorted(extra_rows, key=lambda x: x[1].lower()):
        _render_param_row(section, field_label, items, total, non_monetary)

    page_map = row.get("page_map", "") or ""
    if page_map:
        pm = tk.Frame(parent, bg=CARD_WHITE)
        pm.pack(fill="x", padx=PAD_X, pady=(8, 12))
        tk.Label(pm, text="Page Map:", font=F(7, "bold"),
                 fg=TXT_MUTED, bg=CARD_WHITE).pack(anchor="w")
        tk.Label(pm, text=page_map, font=FMONO(7), fg=TXT_SOFT,
                 bg=CARD_WHITE, justify="left", anchor="w").pack(anchor="w", padx=8)
    else:
        tk.Frame(parent, bg=CARD_WHITE, height=10).pack()


# ═══════════════════════════════════════════════════════════════════════
#  DELETE / CLEAR / DEDUP
# ═══════════════════════════════════════════════════════════════════════

def _delete_row(self, row_id: int):
    if not messagebox.askyesno("Delete Record",
            "Remove this applicant from the database?\n\nThis cannot be undone."):
        return
    _db_delete_row(row_id)
    _refresh_summary(self)

def _delete_custom_column(self):
    try:
        custom_cols = _db_get_custom_columns()
    except Exception as exc:
        messagebox.showerror("Delete Column", f"Could not load custom columns:\n{exc}")
        return

    if not custom_cols:
        messagebox.showinfo("Delete Column", "No custom columns found to delete.")
        return

    win = tk.Toplevel(self)
    win.title("Delete Custom Column")
    win.configure(bg=CARD_WHITE)
    win.resizable(False, False)
    win.grab_set()

    p_x = self.winfo_rootx(); p_y = self.winfo_rooty()
    p_w = self.winfo_width(); p_h = self.winfo_height()
    w_w, w_h = 520, 380
    win.geometry(f"{w_w}x{w_h}+{p_x + (p_w - w_w)//2}+{p_y + (p_h - w_h)//2}")

    hdr = tk.Frame(win, bg=NAVY_DEEP)
    hdr.pack(fill="x")
    tk.Label(hdr, text="🗑  Delete Custom Column",
             font=("Segoe UI", 12, "bold"), fg=WHITE, bg=NAVY_DEEP,
             padx=16, pady=10).pack(side="left")

    body = tk.Frame(win, bg=CARD_WHITE)
    body.pack(fill="both", expand=True, padx=20, pady=16)

    tk.Label(body,
             text="Select a custom column to permanently remove from the database.\n"
                  "This will DROP the column and all its data from every row.",
             font=("Segoe UI", 9), fg=TXT_MUTED, bg=CARD_WHITE,
             justify="left", wraplength=460).pack(anchor="w", pady=(0, 12))

    col_display_map = {
        f"{display_label}  ({db_col})": db_col
        for db_col, display_label, _ in custom_cols
    }
    choice_var = tk.StringVar(value=list(col_display_map.keys())[0])

    col_menu = ttk.Combobox(
        body,
        textvariable=choice_var,
        values=list(col_display_map.keys()),
        state="readonly",
        font=("Segoe UI", 9),
        width=55)
    col_menu.current(0)
    col_menu.pack(fill="x", pady=(0, 16))

    preview_var = tk.StringVar(value="")
    tk.Label(body, textvariable=preview_var,
             font=("Segoe UI", 9, "bold"),
             fg=ACCENT_RED, bg=CARD_WHITE, anchor="w").pack(fill="x")

    def _update_preview(*_):
        db_col = col_display_map.get(choice_var.get(), "")
        if not db_col:
            return
        try:
            with _db_connect() as conn:
                cur = conn.cursor()
                cur.execute(
                    f"SELECT COUNT(*) FROM applicants "
                    f"WHERE {db_col} IS NOT NULL AND TRIM({db_col}::text) != ''")
                count = cur.fetchone()[0]
                cur.close()
            preview_var.set(
                f"⚠  {count} row(s) have data in this column — all will be lost.")
        except Exception as exc:
            preview_var.set(f"Error: {exc}")

    col_menu.bind("<<ComboboxSelected>>", _update_preview)
    _update_preview()

    btn_bar = tk.Frame(win, bg=CARD_WHITE,
                       highlightbackground=BORDER_MID, highlightthickness=1)
    btn_bar.pack(fill="x", padx=20, pady=(12, 14))

    def _on_delete():
        db_col = col_display_map.get(choice_var.get(), "")
        if not db_col:
            return

        try:
            win.grab_release()
        except Exception:
            pass
        win.destroy()

        if not messagebox.askyesno(
                "Confirm Delete Column",
                f"Permanently DROP column:\n\n  {db_col}\n\n"
                f"This cannot be undone. Continue?",
                icon="warning"):
            return

        reason = _ask_edit_reason(self, f"Delete column: {db_col}")
        if reason is None:
            return

        requested_by = getattr(self, "_current_username", "") or "unknown"

        def _submit_worker():
            try:
                _db_submit_delete_column_request(db_col, reason, requested_by)
                def _done():
                    win.destroy()
                    if hasattr(self, "_sum_save_status_lbl"):
                        self._sum_save_status_lbl.config(
                            text="Column deletion submitted for approval",
                            fg=ACCENT_GOLD)
                        try:
                            if getattr(self, "_sum_save_status_after", None):
                                self.after_cancel(self._sum_save_status_after)
                        except Exception:
                            pass
                        self._sum_save_status_after = self.after(
                            3000, lambda: self._sum_save_status_lbl.config(text=""))
                    _log_action(self, "delete_col_requested",
                                f"Column '{db_col}' deletion submitted for approval "
                                f"reason='{reason}'")
                _summary_ui_dispatch(self, _done)
            except Exception as exc:
                err = str(exc)
                _summary_ui_dispatch(self,
                    lambda: messagebox.showerror("Submit Failed", err))

        threading.Thread(target=_submit_worker, daemon=True).start()

    tk.Button(btn_bar, text="✕  Cancel",
              font=("Segoe UI", 9, "bold"), fg=TXT_SOFT, bg="#F0F0F0",
              activebackground="#E0E0E0", relief="flat", bd=0,
              padx=14, pady=7, cursor="hand2",
              command=win.destroy).pack(side="right", padx=(4, 8), pady=8)

    ctk.CTkButton(btn_bar, text="🗑  Delete Column",
                  command=_on_delete,
                  width=140, height=32, corner_radius=6,
                  fg_color="#7A2020", hover_color="#9B2226",
                  text_color=WHITE,
                  font=FF(9, "bold")).pack(side="right", padx=(0, 4), pady=8)

    win.protocol("WM_DELETE_WINDOW", win.destroy)

def _advanced_delete(self):
    DELETE_COLS = [
        ("applicant_name",   "Applicant Name"),
        ("client_id",        "Client ID"),
        ("pn",               "PN"),
        ("industry_name",    "Industry Name"),
        ("residence_address","Residence Address"),
        ("office_address",   "Office Address"),
        ("loan_status",      "Loan Status"),
        ("ao_name",          "AO Name"),
        ("branch",           "Branch"),
        ("loan_class_name",  "Loan Class"),
        ("product_name",     "Product Name"),
        ("maturity",         "Maturity"),
        ("interest_rate",    "Interest Rate"),
        ("term_unit",        "Term Unit"),
        ("release_tag",      "Release Tag"),
        ("source_file",      "Source File"),
        ("session_id",       "Session ID"),
        ("status",           "Status"),
    ]

    win = tk.Toplevel(self)
    win.title("Advanced Delete (Clear Column Data)")
    win.configure(bg=CARD_WHITE)
    win.resizable(False, False)
    win.grab_set()

    p_x = self.winfo_rootx(); p_y = self.winfo_rooty()
    p_w = self.winfo_width(); p_h = self.winfo_height()
    w_w, w_h = 580, 380
    win.geometry(f"{w_w}x{w_h}+{p_x + (p_w - w_w)//2}+{p_y + (p_h - w_h)//2}")

    hdr = tk.Frame(win, bg=NAVY_DEEP)
    hdr.pack(fill="x")
    tk.Label(hdr, text="🗑  Advanced Delete — Clear Column Data",
             font=("Segoe UI", 12, "bold"), fg=WHITE, bg=NAVY_DEEP,
             padx=16, pady=10).pack(side="left")
    tk.Label(hdr,
             text="Clears (sets to blank) the chosen column\n"
                  "only for rows whose value matches the supplied filter.",
             font=("Segoe UI", 8), fg="#8DA8C8", bg=NAVY_DEEP,
             padx=16, justify="left").pack(side="left", pady=8)

    body = tk.Frame(win, bg=CARD_WHITE)
    body.pack(fill="both", expand=True, padx=20, pady=16)

    # ── Column to CLEAR ───────────────────────────────────────────────
    tk.Label(body, text="Column to clear:",
             font=("Segoe UI", 9, "bold"), fg=TXT_NAVY, bg=CARD_WHITE,
             anchor="w").pack(fill="x", pady=(0, 4))

    col_var = tk.StringVar(value=DELETE_COLS[0][0])
    col_labels = {db_col: label for db_col, label in DELETE_COLS}
    display_to_db = {
        f"{label}  ({db_col})": db_col
        for db_col, label in DELETE_COLS
    }

    col_menu = ttk.Combobox(
        body,
        textvariable=col_var,
        values=[f"{label}  ({db_col})" for db_col, label in DELETE_COLS],
        state="readonly",
        font=("Segoe UI", 9),
        width=48,
    )
    col_menu.current(0)
    col_menu.pack(fill="x", pady=(0, 12))

    # ── Filter value(s) ───────────────────────────────────────────────
    tk.Label(body,
             text="Clear only rows where that column matches  "
                  "(comma-separated, OR logic).\n"
                  "Leave blank to clear ALL rows for that column.",
             font=("Segoe UI", 8), fg=TXT_MUTED, bg=CARD_WHITE,
             anchor="w", justify="left").pack(fill="x", pady=(0, 4))

    val_var = tk.StringVar()
    val_entry = tk.Entry(body, textvariable=val_var,
                         font=("Segoe UI", 10), fg=TXT_NAVY, bg=WHITE,
                         relief="solid", bd=1, insertbackground=NAVY_MID)
    val_entry.pack(fill="x", ipady=5, pady=(0, 6))
    val_entry.focus_set()

    # ── Match mode + All button ───────────────────────────────────────
    match_mode = tk.StringVar(value="partial")
    mode_row = tk.Frame(body, bg=CARD_WHITE)
    mode_row.pack(fill="x", pady=(0, 10))
    tk.Label(mode_row, text="Match mode:",
             font=("Segoe UI", 8), fg=TXT_MUTED, bg=CARD_WHITE).pack(side="left")
    tk.Radiobutton(mode_row, text="Partial (LIKE)", variable=match_mode,
                   value="partial", font=("Segoe UI", 8),
                   fg=TXT_NAVY, bg=CARD_WHITE, activebackground=CARD_WHITE,
                   selectcolor=WHITE).pack(side="left", padx=(8, 0))
    tk.Radiobutton(mode_row, text="Exact", variable=match_mode,
                   value="exact", font=("Segoe UI", 8),
                   fg=TXT_NAVY, bg=CARD_WHITE, activebackground=CARD_WHITE,
                   selectcolor=WHITE).pack(side="left", padx=(8, 0))

    def _fill_all():
        db_col = display_to_db.get(col_menu.get(), col_var.get())
        try:
            with _db_connect() as conn:
                cur = conn.cursor()
                cur.execute(
                    f"SELECT DISTINCT {db_col} FROM applicants "
                    f"WHERE {db_col} IS NOT NULL AND TRIM({db_col}::text) != '' "
                    f"ORDER BY {db_col}")
                rows = cur.fetchall()
                cur.close()
            all_vals = [str(r[0]).strip() for r in rows if r[0]]
            val_var.set(", ".join(all_vals))
        except Exception as exc:
            preview_var.set(f"Error: {exc}")

    tk.Button(mode_row, text="All", font=("Segoe UI", 7, "bold"),
              fg=TXT_ON_LIME, bg=LIME_MID, activebackground=LIME_BRIGHT,
              activeforeground=TXT_ON_LIME, relief="flat", bd=0,
              padx=8, pady=2, cursor="hand2",
              command=_fill_all).pack(side="left", padx=(12, 0))

    # ── Live preview ──────────────────────────────────────────────────
    preview_var = tk.StringVar(value="")
    tk.Label(body, textvariable=preview_var,
             font=("Segoe UI", 9, "bold"),
             fg=ACCENT_RED, bg=CARD_WHITE, anchor="w").pack(fill="x")

    def _update_preview(*_):
        db_col = display_to_db.get(col_menu.get(), col_var.get())
        raw    = val_var.get().strip()
        try:
            count = _count_col_matches(db_col, raw, match_mode.get())
            preview_var.set(
                f"⚠  {count} row(s) will have [{col_labels.get(db_col, db_col)}] cleared."
                if count > 0 else "✓  No matching rows found.")
        except Exception as exc:
            preview_var.set(f"Error: {exc}")

    val_var.trace_add("write",   _update_preview)
    match_mode.trace_add("write", _update_preview)
    col_menu.bind("<<ComboboxSelected>>",
                  lambda e: (_update_preview(), _update_preview()))  # refresh twice so col updates

    # ── Buttons ───────────────────────────────────────────────────────
    btn_bar = tk.Frame(win, bg=CARD_WHITE,
                       highlightbackground=BORDER_MID, highlightthickness=1)
    btn_bar.pack(fill="x", padx=20, pady=(4, 14))

    def _on_clear():
        db_col = display_to_db.get(col_menu.get(), col_var.get())
        raw    = val_var.get().strip()
        count  = _count_col_matches(db_col, raw, match_mode.get())
        if count == 0:
            messagebox.showinfo("Advanced Delete", "No rows match — nothing changed.")
            return
        col_display = col_labels.get(db_col, db_col)
        filter_desc = f'where value matches:  "{raw}"' if raw else "in ALL rows"
        if not messagebox.askyesno(
                "Confirm Clear",
                f"This will blank out the [{col_display}] column\n"
                f"{filter_desc}\n\n"
                f"Affects {count} row(s).  This cannot be undone.  Continue?",
                icon="warning"):
            return
        cleared = _db_clear_column(db_col, raw, match_mode.get())
        win.destroy()
        _refresh_summary(self)
        messagebox.showinfo("Advanced Delete",
                            f"✓  [{col_display}] cleared on {cleared} row(s).")

    tk.Button(btn_bar, text="✕  Cancel",
              font=("Segoe UI", 9, "bold"), fg=TXT_SOFT, bg="#F0F0F0",
              activebackground="#E0E0E0", relief="flat", bd=0,
              padx=14, pady=7, cursor="hand2",
              command=win.destroy).pack(side="right", padx=(4, 8), pady=8)

    ctk.CTkButton(btn_bar, text="🗑  Clear Column Data",
                  command=_on_clear,
                  width=160, height=32, corner_radius=6,
                  fg_color="#7A2020", hover_color="#9B2226",
                  text_color=WHITE,
                  font=FF(9, "bold")).pack(side="right", padx=(0, 4), pady=8)

    win.protocol("WM_DELETE_WINDOW", win.destroy)


def _count_col_matches(db_col: str, raw_filter: str, mode: str) -> int:
    """Count rows that would be affected by the clear operation."""
    with _db_connect() as conn:
        cur = conn.cursor()
        if not raw_filter:
            cur.execute(
                f"SELECT COUNT(*) FROM applicants "
                f"WHERE {db_col} IS NOT NULL AND TRIM(CAST({db_col} AS TEXT)) != ''")
        else:
            values = [v.strip() for v in raw_filter.split(",") if v.strip()]
            if mode == "exact":
                placeholders = " OR ".join(
                    [f"TRIM(UPPER(CAST({db_col} AS TEXT))) = TRIM(UPPER(%s))"
                     for _ in values])
            else:
                placeholders = " OR ".join(
                    [f"CAST({db_col} AS TEXT) ILIKE %s" for _ in values])
                values = [f"%{v}%" for v in values]
            cur.execute(
                f"SELECT COUNT(*) FROM applicants WHERE {placeholders}", values)
        result = cur.fetchone()
        cur.close()
    return result[0] if result else 0


def _db_clear_column(db_col: str, raw_filter: str, mode: str) -> int:
    """Set the column to NULL for all matching rows. Returns rows affected."""
    with _db_connect() as conn:
        cur = conn.cursor()
        if not raw_filter:
            cur.execute(f"UPDATE applicants SET {db_col} = NULL")
        else:
            values = [v.strip() for v in raw_filter.split(",") if v.strip()]
            if mode == "exact":
                placeholders = " OR ".join(
                    [f"TRIM(UPPER(CAST({db_col} AS TEXT))) = TRIM(UPPER(%s))"
                     for _ in values])
            else:
                placeholders = " OR ".join(
                    [f"CAST({db_col} AS TEXT) ILIKE %s" for _ in values])
                values = [f"%{v}%" for v in values]
            cur.execute(
                f"UPDATE applicants SET {db_col} = NULL WHERE {placeholders}",
                values)
        affected = cur.rowcount
        conn.commit()
        cur.close()
    return affected


def _clear_all(self):
    totals = _db_totals()
    total  = totals.get("total", 0) or 0
    if total == 0:
        messagebox.showinfo("Clear All", "There are no records to delete.")
        return
    if not messagebox.askyesno(
            "Clear All Records",
            f"This will permanently delete ALL {total} applicant record(s) "
            f"from the database.\n\nThis cannot be undone.\n\nContinue?",
            icon="warning"):
        return
    _db_clear_all()
    _refresh_summary(self)


def _run_dedup(self):
    if not messagebox.askyesno(
            "Deduplicate by Client ID",
            "This will scan the database and merge any rows that share the "
            "same Client ID.\n\n"
            "• The most complete record is kept.\n"
            "• Missing fields are filled in from duplicates.\n"
            "• Duplicate rows are deleted.\n\nContinue?"):
        return

    _flash_btn(self, self._sum_dedup_btn, "⟳  Working…", 60_000)

    def _worker():
        try:
            removed = _db_deduplicate_client_ids()
            self.after(0, lambda: _refresh_summary(self))
            self.after(0, lambda: (
                _flash_btn(self, self._sum_dedup_btn, "✓  Done!", 2500),
                messagebox.showinfo(
                    "Deduplication Complete",
                    (f"Deduplication finished.\n\n"
                     f"✓  Duplicate rows removed : {removed}\n\n"
                     f"All unique Client IDs now have a single consolidated record.")
                    if removed > 0 else
                    "No duplicate Client IDs found — the database is already clean."
                )
            ))
        except Exception as exc:
            err = str(exc)
            self.after(0, lambda: (
                _flash_btn(self, self._sum_dedup_btn, "✗  Error", 3000),
                messagebox.showerror("Dedup Error", err)
            ))

    threading.Thread(target=_worker, daemon=True).start()


# ═══════════════════════════════════════════════════════════════════════
#  AMORTIZATION IMPORT  (CHANGE 1)
# ═══════════════════════════════════════════════════════════════════════

def _import_amort_file(self):
    path = filedialog.askopenfilename(
        title="Import Amortization Values",
        filetypes=[("Excel & CSV files", "*.xlsx *.csv"),
                   ("Excel files", "*.xlsx"),
                   ("CSV files", "*.csv"),
                   ("All files", "*.*")])
    if not path:
        return

    _flash_btn(self, self._sum_import_amort_btn, "⟳  Reading…", 60_000)

    def _worker():
        try:
            if path.lower().endswith(".csv"):
                import csv as _csv
                with open(path, newline="", encoding="utf-8-sig") as f:
                    reader  = _csv.DictReader(f)
                    records = [dict(row) for row in reader]
                all_cols = list(records[0].keys()) if records else []
            else:
                import openpyxl
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
                if header_row is None:
                    raise ValueError("The Excel file appears to be empty.")
                all_cols = [
                    str(cell.value).strip() if cell.value is not None else ""
                    for cell in header_row
                ]
                records = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if all(v is None for v in row):
                        continue
                    records.append({
                        all_cols[i]: (str(v).strip() if v is not None else "")
                        for i, v in enumerate(row) if i < len(all_cols)
                    })
                wb.close()

            if not records:
                raise ValueError("No data rows found in the file.")

            def _find_col(cols, *keywords):
                for kw in keywords:
                    kw_norm = re.sub(r"[\s_]", "", kw.lower())
                    for c in cols:
                        c_norm = re.sub(r"[\s_]", "", c.lower())
                        if kw_norm in c_norm:
                            return c
                return None

            # ── Detect required columns ────────────────────────────────
            col_client   = _find_col(all_cols, "applicant", "client", "name")
            col_amort    = _find_col(all_cols,
                                     "monthlypaymentamount", "monthly payment amount",
                                     "monthlypayment", "paymentamount",
                                     "currentamort", "totalcurrentamort", "amort")
            # ── client_id column (optional but preferred) ──────────────
            col_clientid = _find_col(all_cols, "clientid", "client id",
                                     "client_id", "cid")

            missing = []
            if not col_client: missing.append("Applicant / Client")
            if not col_amort:  missing.append("MonthlyPaymentAmount / Amortization")
            if missing:
                raise ValueError(
                    f"Could not detect column(s): {', '.join(missing)}\n\n"
                    f"File has: {', '.join(all_cols)}")

            # ── Aggregate amort values per client ──────────────────────
            aggregated: dict[str, dict] = {}
            bad_rows:   list[tuple]     = []

            for file_row in records:
                client_name = str(file_row.get(col_client)   or "").strip()
                raw_val     = str(file_row.get(col_amort)    or "").strip()
                raw_cid     = str(file_row.get(col_clientid) or "").strip() \
                              if col_clientid else ""

                if not client_name:
                    continue

                try:
                    cleaned   = re.sub(r"[^\d.]", "", raw_val.replace(",", ""))
                    amort_val = float(cleaned) if cleaned else None
                except Exception:
                    amort_val = None

                name_key = client_name.upper()

                if amort_val is None:
                    bad_rows.append((name_key, f"bad value: '{raw_val}'"))
                    continue

                if name_key not in aggregated:
                    aggregated[name_key] = {
                        "name":      client_name,
                        "client_id": raw_cid.upper(),
                        "total":     0.0,
                    }
                if not aggregated[name_key]["client_id"] and raw_cid:
                    aggregated[name_key]["client_id"] = raw_cid.upper()

                aggregated[name_key]["total"] += amort_val

            # ── Build client_id → db row id lookup ─────────────────────
            with _db_connect() as _conn:
                _cur = _conn.cursor()
                _cur.execute(
                    "SELECT client_id, id FROM applicants "
                    "WHERE client_id IS NOT NULL AND TRIM(client_id) != ''")
                cid_to_dbid: dict[str, int] = {
                    str(r[0]).strip().upper(): r[1]
                    for r in _cur.fetchall()
                }
                _cur.close()

            # ── Match & update ─────────────────────────────────────────
            updated_by_id   = []
            updated_by_name = []
            updated_relaxed = []
            skipped_names   = list(bad_rows)

            for name_key, entry in aggregated.items():
                amort_val   = entry["total"]
                client_name = entry["name"]
                file_cid    = entry["client_id"]

                # PRIMARY: match by client_id
                if file_cid and file_cid in cid_to_dbid:
                    db_id = cid_to_dbid[file_cid]
                    _db_update_amort_current(db_id, amort_val)
                    updated_by_id.append((name_key, db_id))
                    continue

                # FALLBACK: name similarity
                hits, sim_label = _resolve_name_similarity(client_name)
                if hits:
                    _db_update_amort_all(hits, amort_val)
                    if sim_label in ("exact", "high"):
                        updated_by_name.append((name_key, hits[0][1]))
                    else:
                        updated_relaxed.append((name_key, hits[0][1]))
                else:
                    reason = ("client_id not in DB, no name match"
                              if file_cid else "no client_id, no name match")
                    skipped_names.append((name_key, reason))

            self.after(0, lambda: _refresh_summary(self))

            cid_detected = f"'{col_clientid}'" if col_clientid else "not detected"
            msg  = "Amort. import complete.\n\n"
            msg += f"Client ID column      : {cid_detected}\n"
            msg += f"✓  Matched by ID      : {len(updated_by_id)} record(s)\n"
            msg += f"✓  Matched by name    : {len(updated_by_name)} record(s)\n"
            msg += f"–  Skipped            : {len(skipped_names)} row(s)\n"

            if not col_clientid:
                msg += "\nℹ  No Client ID column found in file — used name matching only.\n"

            if updated_relaxed:
                msg += f"\n⚠  {len(updated_relaxed)} matched via relaxed similarity — please verify:\n"
                for file_n, db_id in updated_relaxed[:10]:
                    msg += f"  • File: {file_n}  →  DB id: {db_id}\n"
                if len(updated_relaxed) > 10:
                    msg += f"  … and {len(updated_relaxed) - 10} more\n"

            if skipped_names:
                msg += "\nSkipped rows:\n"
                for name, reason in skipped_names[:10]:
                    msg += f"  • {name}  ({reason})\n"
                if len(skipped_names) > 10:
                    msg += f"  … and {len(skipped_names) - 10} more"

            self.after(0, lambda: (
                _flash_btn(self, self._sum_import_amort_btn, "✓  Done!", 2500),
                messagebox.showinfo("Amort. Import Result", msg)
            ))
        except Exception as exc:
            err = str(exc)
            self.after(0, lambda: (
                _flash_btn(self, self._sum_import_amort_btn, "✗  Error", 3000),
                messagebox.showerror("Import Error", err)
            ))

    threading.Thread(target=_worker, daemon=True).start()


# ═══════════════════════════════════════════════════════════════════════
#  OTHER DATA IMPORT  (PATCHED)
# ═══════════════════════════════════════════════════════════════════════

def _import_other_data_file(self):
    path = filedialog.askopenfilename(
        title="Import Other Data (Client ID / PN / Industry / Loan Balance)",
        filetypes=[("Excel & CSV files", "*.xlsx *.csv"),
                   ("Excel files", "*.xlsx"),
                   ("CSV files", "*.csv"),
                   ("All files", "*.*")])
    if not path:
        return

    _flash_btn(self, self._sum_import_other_btn, "⟳  Reading…", 60_000)

    def _worker():
        try:
            if path.lower().endswith(".csv"):
                import csv as _csv
                with open(path, newline="", encoding="utf-8-sig") as f:
                    reader  = _csv.DictReader(f)
                    records = [dict(row) for row in reader]
                all_cols = list(records[0].keys()) if records else []
            else:
                import openpyxl
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
                if header_row is None:
                    raise ValueError("The Excel file appears to be empty.")
                all_cols = [
                    str(cell.value).strip() if cell.value is not None else ""
                    for cell in header_row
                ]
                records = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if all(v is None for v in row):
                        continue
                    records.append({
                        all_cols[i]: (str(v).strip() if v is not None else "")
                        for i, v in enumerate(row) if i < len(all_cols)
                    })
                wb.close()

            if not records:
                raise ValueError("No data rows found in the file.")

            def _find_col(cols, *keywords):
                for kw in keywords:
                    kw_norm = re.sub(r"[\s_\-]", "", kw.lower())
                    for c in cols:
                        c_norm = re.sub(r"[\s_\-]", "", c.lower())
                        if kw_norm == c_norm or kw_norm in c_norm:
                            return c
                return None

            col_name      = _find_col(all_cols, "clientname", "client name",
                                      "applicant", "applicantname", "name")
            col_clientid  = _find_col(all_cols, "clientid", "client id",
                                      "client_id", "cid")
            col_pn        = _find_col(all_cols, "pnid", "pn id", "pn_id",
                                      "pn", "promissorynote")
            col_industry  = _find_col(all_cols, "industryname", "industry name",
                                      "industry_name", "industry")
            col_loanbal   = _find_col(all_cols, "loanbalance", "loan balance",
                                      "loan_balance", "loanbal", "balance")
            col_amortcost = _find_col(all_cols, "amortizedcost", "amortized cost",
                                      "amortized_cost", "amortcost",
                                      "amortised cost", "amortisedcost")

            if not col_name:
                raise ValueError(
                    f"Could not detect a client name column.\n\n"
                    f"File has: {', '.join(all_cols)}")

            if not any([col_clientid, col_pn, col_industry,
                        col_loanbal, col_amortcost]):
                raise ValueError(
                    f"No data columns found (clientid / pnid / industryname / "
                    f"loanbalance / amortizedcost).\n\n"
                    f"File has: {', '.join(all_cols)}")

            # ── B-1: pn_collect keyed on normalised name ───────────────
            # Using _normalise_for_sim ensures "DELA CRUZ, JUAN" and
            # "JUAN DELA CRUZ" accumulate into the same PN bucket.
            pn_collect: dict[str, list] = {}
            if col_pn:
                for file_row in records:
                    client_name = str(file_row.get(col_name) or "").strip()
                    if not client_name:
                        continue
                    pn_val   = str(file_row.get(col_pn) or "").strip()
                    norm_key = _normalise_for_sim(client_name)   # ← normalised
                    if norm_key not in pn_collect:
                        pn_collect[norm_key] = []
                    if pn_val and pn_val not in pn_collect[norm_key]:
                        pn_collect[norm_key].append(pn_val)

            def _agg_numeric(col_key, records, col_name_field, bad_list, label):
                agg = {}
                for file_row in records:
                    client_name = str(
                        file_row.get(col_name_field) or "").strip()
                    if not client_name:
                        continue
                    raw_val  = str(file_row.get(col_key) or "").strip()
                    # ← also normalise numeric-agg keys
                    norm_key = _normalise_for_sim(client_name)
                    try:
                        cleaned = re.sub(r"[^\d.]", "",
                                         raw_val.replace(",", ""))
                        val     = float(cleaned) if cleaned else None
                    except Exception:
                        val = None
                    if val is None:
                        if raw_val:
                            bad_list.append(
                                (norm_key, f"bad {label}: '{raw_val}'"))
                        if norm_key not in agg:
                            agg[norm_key] = None
                    else:
                        agg[norm_key] = (agg.get(norm_key) or 0.0) + val
                return agg

            loan_bal_bad  = []
            amortcost_bad = []
            loan_bal_agg  = (
                _agg_numeric(col_loanbal,   records, col_name,
                             loan_bal_bad,  "loanbalance")
                if col_loanbal   else {})
            amortcost_agg = (
                _agg_numeric(col_amortcost, records, col_name,
                             amortcost_bad, "amortizedcost")
                if col_amortcost else {})

            # ── B-2: dedup keyed on normalised name ────────────────────
            # "DELA CRUZ, JUAN" normalises to "JUAN DELA CRUZ" — same key
            # as a row that already spells it in natural order.
            seen_norm_keys: set[str] = set()
            deduped: list[dict]      = []
            dup_count = 0

            for file_row in records:
                client_name = str(file_row.get(col_name) or "").strip()
                if not client_name:
                    continue
                norm_key = _normalise_for_sim(client_name)
                if norm_key in seen_norm_keys:
                    dup_count += 1
                    continue
                seen_norm_keys.add(norm_key)
                pn_joined = "\n".join(
                    pn_collect.get(norm_key, [])) if col_pn else ""
                deduped.append({
                    "name":           client_name,   # original display name
                    "norm_key":       norm_key,       # normalised form
                    "client_id":      str(
                        file_row.get(col_clientid) or "").strip()
                        if col_clientid else "",
                    "pn_joined":      pn_joined,
                    "industry":       str(
                        file_row.get(col_industry) or "").strip()
                        if col_industry else "",
                    "loan_balance":   loan_bal_agg.get(norm_key)
                                      if col_loanbal   else None,
                    "amortized_cost": amortcost_agg.get(norm_key)
                                      if col_amortcost else None,
                })

            # ── B-3: pre-match by client_id (mirrors Amort import) ─────
            # Build a lookup of client_id → db row id from the current DB.
            updated_by_cid:  list = []
            name_match_queue: list[dict] = []

            if col_clientid:
                with _db_connect() as _conn:
                    _cur = _conn.cursor()
                    _cur.execute(
                        "SELECT client_id, id FROM applicants "
                        "WHERE client_id IS NOT NULL "
                        "AND TRIM(client_id) != ''")
                    cid_to_dbid: dict[str, int] = {
                        str(r[0]).strip().upper(): r[1]
                        for r in _cur.fetchall()
                    }
                    _cur.close()

                for entry in deduped:
                    file_cid = entry["client_id"].upper()
                    if file_cid and file_cid in cid_to_dbid:
                        db_id = cid_to_dbid[file_cid]
                        rows_written = _db_update_other_data_all(
                            [(db_id, None)],
                            entry["client_id"],
                            entry["pn_joined"],
                            entry["industry"],
                            entry["loan_balance"],
                            entry["amortized_cost"],
                        )
                        if rows_written:
                            updated_by_cid.append(
                                (entry["norm_key"], db_id))
                    else:
                        # No client_id match — queue for name matching
                        name_match_queue.append(entry)
            else:
                # No client_id column in file — everything goes to name match
                name_match_queue = deduped

            # ── Name-similarity matching for remaining entries ─────────
            updated_strict   = []
            updated_relaxed  = []
            skipped_no_match = []

            for entry in name_match_queue:
                hits, sim_label = _resolve_name_similarity(entry["name"])
                if not hits:
                    skipped_no_match.append(
                        entry["norm_key"]); continue
                rows_written = _db_update_other_data_all(
                    hits,
                    entry["client_id"],
                    entry["pn_joined"],
                    entry["industry"],
                    entry["loan_balance"],
                    entry["amortized_cost"],
                )
                if rows_written:
                    if sim_label == "exact":
                        updated_strict.append(
                            (entry["norm_key"], hits[0][1]))
                    else:
                        updated_relaxed.append(
                            (entry["norm_key"], hits[0][1]))

            _db_deduplicate_client_ids()
            self.after(0, lambda: _refresh_summary(self))

            # ── B-4: updated summary message ──────────────────────────
            total_updated = (len(updated_by_cid)
                             + len(updated_strict)
                             + len(updated_relaxed))
            cols_imported = ", ".join(filter(None, [
                "Client ID"            if col_clientid  else "",
                "PN (all)"             if col_pn        else "",
                "Industry Name"        if col_industry  else "",
                "Loan Balance"         if col_loanbal   else "",
                "Total Amortized Cost" if col_amortcost else "",
            ]))
            msg  = "Other Data import complete.\n\n"
            msg += f"Columns imported       : {cols_imported}\n"
            msg += f"✓  Matched by ID       : {len(updated_by_cid)} record(s)\n"
            msg += f"✓  Matched by name     : {len(updated_strict) + len(updated_relaxed)} record(s)\n"
            msg += f"↩  Duplicates skipped  : {dup_count}\n"
            msg += f"–  No DB match         : {len(skipped_no_match)} name(s)\n"
            if loan_bal_bad:
                msg += (f"⚠  Unparseable loan balance values: "
                        f"{len(loan_bal_bad)}\n")
            if amortcost_bad:
                msg += (f"⚠  Unparseable amortized cost values: "
                        f"{len(amortcost_bad)}\n")
            if not col_clientid:
                msg += ("\nℹ  No Client ID column found — "
                        "used name matching only.\n")
            if updated_relaxed:
                msg += (f"\n⚠  {len(updated_relaxed)} matched via "
                        f"similarity — please verify:\n")
                for file_n, db_id in updated_relaxed[:10]:
                    msg += f"  • File: {file_n}  →  DB id: {db_id}\n"
                if len(updated_relaxed) > 10:
                    msg += f"  … and {len(updated_relaxed) - 10} more\n"
            if skipped_no_match:
                msg += "\nNames with no DB match:\n"
                for name in skipped_no_match[:10]:
                    msg += f"  • {name}\n"
                if len(skipped_no_match) > 10:
                    msg += f"  … and {len(skipped_no_match) - 10} more"

            self.after(0, lambda: (
                _flash_btn(self, self._sum_import_other_btn, "✓  Done!", 2500),
                messagebox.showinfo("Other Data Import Result", msg)
            ))
        except Exception as exc:
            err = str(exc)
            self.after(0, lambda: (
                _flash_btn(self, self._sum_import_other_btn, "✗  Error", 3000),
                messagebox.showerror("Other Data Import Error", err)
            ))

    threading.Thread(target=_worker, daemon=True).start()


# ═══════════════════════════════════════════════════════════════════════
#  PRINCIPAL LOAN IMPORT
# ═══════════════════════════════════════════════════════════════════════

def _normalize_name_for_matching(name: str) -> str:
    """Normalize name for matching by removing suffixes and extra spaces."""
    if not name:
        return ""
    normalized = name.upper().strip()
    suffixes = [" JR", " SR", " II", " III", " IV", " JR.", " SR.", " JR", " SR"]
    for suffix in suffixes:
        if normalized.endswith(suffix):
            normalized = normalized[:-len(suffix)]
    normalized = " ".join(normalized.split())
    return normalized


def _import_ploan_file(self):
    path = filedialog.askopenfilename(
        title="Import Principal Loan Data",
        filetypes=[("Excel & CSV files", "*.xlsx *.csv"),
                   ("Excel files", "*.xlsx"),
                   ("CSV files", "*.csv"),
                   ("All files", "*.*")])
    if not path:
        return

    _flash_btn(self, self._sum_import_ploan_btn, "⟳  Reading…", 60_000)

    def _worker():
        try:
            # ── 1. Read file ─────────────────────────────────────────────
            if path.lower().endswith(".csv"):
                import csv as _csv
                with open(path, newline="", encoding="utf-8-sig") as f:
                    reader  = _csv.DictReader(f)
                    records = [dict(row) for row in reader]
                all_cols = list(records[0].keys()) if records else []
            else:
                import openpyxl
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
                if header_row is None:
                    raise ValueError("The Excel file appears to be empty.")
                all_cols = [
                    str(cell.value).strip() if cell.value is not None else ""
                    for cell in header_row
                ]
                records = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if all(v is None for v in row):
                        continue
                    records.append({
                        all_cols[i]: (str(v).strip() if v is not None else "")
                        for i, v in enumerate(row) if i < len(all_cols)
                    })
                wb.close()

            if not records:
                raise ValueError("No data rows found in the file.")

            # ── 2. Column detection ──────────────────────────────────────
            def _find_col(cols, *keywords):
                for kw in keywords:
                    kw_norm = re.sub(r"[\s_\-]", "", kw.lower())
                    for c in cols:
                        c_norm = re.sub(r"[\s_\-]", "", c.lower())
                        if kw_norm == c_norm:
                            return c
                for kw in keywords:
                    kw_norm = re.sub(r"[\s_\-]", "", kw.lower())
                    for c in cols:
                        c_norm = re.sub(r"[\s_\-]", "", c.lower())
                        if kw_norm in c_norm:
                            return c
                return None

            col_clientname = _find_col(all_cols,
                                       "clientname", "client name", "client_name",
                                       "applicant", "applicantname", "name")
            if not col_clientname:
                raise ValueError(
                    f"Could not detect a client name column.\n\n"
                    f"File has: {', '.join(all_cols)}")

            col_clientid_file = _find_col(all_cols,
                                          "clientid", "client id", "client_id", "cid")

            COL_MAP_DEFS = [
                ("client_id",           False, "clientid", "client id", "client_id", "cid"),
                ("pn",                  False, "pnid", "pn id", "pn_id", "pn", "promissorynote"),
                ("branch",              False, "branch", "branchname", "branch name"),
                ("loan_class_name",     False, "loanclassname", "loan class name", "loan class", "loanclass"),
                ("product_name",        False, "productname", "product name", "product_name", "product"),
                ("industry_name",       False, "industryname", "industry name", "industry_name", "industry"),
                ("loan_date",           False, "loandate", "loan date", "loan_date", "dateofrelease", "releasedate"),
                ("maturity",            False, "maturity", "maturitydate", "maturity date", "duedate", "due date"),
                ("interest_rate",       False, "interest", "interestrate", "interest rate", "interest_rate", "rate", "intrate"),
                ("term_unit",           False, "termunit", "term unit", "term_unit", "paymentfrequency", "frequency"),
                ("term",                False, "term"),
                ("security",            False, "security", "collateral", "securitydescription"),
                ("release_tag",         False, "releasetag", "release tag", "release_tag", "tag"),
                ("loan_amount",         True,  "loanamount", "loan amount", "principalloan", "principal loan", "amount"),
                ("loan_balance",        True,  "loanbalance", "loan balance"),
                ("amort_current_total", True,  "ammortization", "amortization"),
                ("loan_status",         False, "loanstatus", "loan status", "status", "accountstatus"),
                ("ao_name",             False, "aoname", "ao name", "ao_name", "accountofficer", "account officer", "ao"),
            ]

            detected: dict[str, tuple] = {}
            for entry in COL_MAP_DEFS:
                db_col      = entry[0]
                is_monetary = entry[1]
                keywords    = entry[2:]
                file_col    = _find_col(all_cols, *keywords)
                detected[db_col] = (file_col, is_monetary)

            # ── 3. Aggregate all rows per client ─────────────────────────
            aggregated: dict[str, dict] = {}

            for file_row in records:
                client_name = str(file_row.get(col_clientname) or "").strip()
                if not client_name:
                    continue

                original_key = client_name.upper()

                if original_key not in aggregated:
                    bucket: dict = {"_display_name": client_name, "_original_key": original_key}
                    for db_col, (file_col, is_monetary) in detected.items():
                        bucket[db_col] = 0.0 if is_monetary else []
                    aggregated[original_key] = bucket

                bucket = aggregated[original_key]

                for db_col, (file_col, is_monetary) in detected.items():
                    if file_col is None:
                        continue
                    raw = str(file_row.get(file_col) or "").strip()

                    if is_monetary:
                        cleaned = re.sub(r"[^\d.]", "", raw.replace(",", ""))
                        try:
                            val = float(cleaned) if cleaned else None
                        except ValueError:
                            val = None
                        if val is not None:
                            bucket[db_col] = (bucket[db_col] or 0.0) + val
                    else:
                        if raw and raw not in bucket[db_col]:
                            bucket[db_col].append(raw)

            # ── 4. Flatten aggregated buckets ────────────────────────────
            write_ready: dict[str, dict] = {}

            for key, bucket in aggregated.items():
                record: dict = {
                    "_display_name": bucket["_display_name"],
                    "_original_key": bucket["_original_key"]
                }
                for db_col, (file_col, is_monetary) in detected.items():
                    if file_col is None:
                        record[db_col] = None
                        continue
                    raw_val = bucket[db_col]
                    if is_monetary:
                        record[db_col] = raw_val if raw_val != 0.0 else None
                    else:
                        joined = ", ".join(raw_val) if raw_val else None
                        record[db_col] = joined
                write_ready[key] = record

            # ══════════════════════════════════════════════════════════════
            #  LOCAL NAME-MATCHING HELPERS
            #
            #  _abbreviate_middle(norm_name)
            #  ──────────────────────────────
            #  Collapses every middle token to its first letter while
            #  keeping compound surnames (DELA CRUZ, DE LOS SANTOS, SAN
            #  PEDRO, etc.) fully intact.
            #
            #  Examples:
            #    "JUAN CARLOS DELA CRUZ"   → "JUAN C DELA CRUZ"
            #    "MARIA THERESA SAN PEDRO" → "MARIA T SAN PEDRO"
            #    "JUAN DELA CRUZ"          → "JUAN DELA CRUZ"  (no middle)
            #    "JOSE P RIZAL"            → "JOSE P RIZAL"    (already initial)
            #
            #  _middle_aware_score(a, b)
            #  ──────────────────────────
            #  Returns the BEST similarity across four representations:
            #    (a_full vs b_full), (a_abbr vs b_abbr),
            #    (a_full vs b_abbr), (a_abbr vs b_full)
            #
            #  Taking the MAX ensures we never penalise a pair just because
            #  one source stored "CARLOS" and the other stored "C".
            # ══════════════════════════════════════════════════════════════

            _COMPOUND_PREFIXES = {
                "DELA", "DE", "DEL", "LOS", "LAS", "SAN", "SANTA",
                "SANTO", "NG", "NGA",
            }

            def _abbreviate_middle(norm_name: str) -> str:
                tokens = norm_name.split()
                if len(tokens) <= 2:
                    return norm_name

                # Walk from the right to find where the surname starts,
                # absorbing compound-prefix tokens (DELA, DE, SAN, etc.)
                surname_tokens: list[str] = []
                i = len(tokens) - 1
                surname_tokens.insert(0, tokens[i])
                i -= 1
                while i > 0 and tokens[i] in _COMPOUND_PREFIXES:
                    surname_tokens.insert(0, tokens[i])
                    i -= 1

                # tokens[0]       = first name
                # tokens[1 .. i]  = middle name(s) / initials  (may be empty)
                # surname_tokens  = last name (compound-safe)
                first      = tokens[0]
                middle     = tokens[1: i + 1]
                abbreviated = [t[0] for t in middle if t]

                return " ".join([first] + abbreviated + surname_tokens)

            def _middle_aware_score(name_a: str, name_b: str) -> float:
                a_full = name_a
                b_full = name_b
                a_abbr = _abbreviate_middle(name_a)
                b_abbr = _abbreviate_middle(name_b)
                return max(
                    _similarity(a_full, b_full),
                    _similarity(a_abbr, b_abbr),
                    _similarity(a_full, b_abbr),
                    _similarity(a_abbr, b_full),
                )

            # ══════════════════════════════════════════════════════════════
            #  PHASE A — CLIENT ID ASSIGNMENT
            #
            #  Strategy (in priority order):
            #    1. File has a Client ID AND it already exists in DB
            #       → direct match, confirm name similarity (loose check)
            #    2. File has a Client ID but it's NOT yet in DB
            #       → find best DB row by name, assign if confident
            #    3. File has NO Client ID
            #       → name-only matching via _resolve_name_similarity
            #
            #  _best_db_match pipeline (best score among all three wins):
            #    1. _first_last_match  → compound-surname aware, middle-
            #                            agnostic → score floored at 0.90
            #    2. _middle_aware_score ≥ 0.82 → full / abbreviated middle
            #                            compared in all four combinations
            #    3. raw _similarity ≥ 0.80 → plain token-sort fuzzy fallback
            # ══════════════════════════════════════════════════════════════

            # Fetch ALL current DB rows once
            with _db_connect() as _conn:
                _cur = _conn.cursor()
                _cur.execute(
                    "SELECT id, applicant_name, client_id FROM applicants "
                    "WHERE applicant_name IS NOT NULL")
                cols_desc = [desc[0] for desc in _cur.description]
                db_all_rows = [dict(zip(cols_desc, r)) for r in _cur.fetchall()]
                _cur.close()

            # Pre-build normalised-name → db row list for fast iteration
            db_norm_index: list[tuple[str, dict]] = [
                (_normalise_for_sim(r["applicant_name"]), r)
                for r in db_all_rows
                if r.get("applicant_name")
            ]

            # client_id → db_id map
            existing_cid_map: dict[str, int] = {
                str(r["client_id"]).strip().upper(): r["id"]
                for r in db_all_rows
                if r.get("client_id") and str(r["client_id"]).strip()
            }

            def _best_db_match(file_display_name: str) -> tuple[dict | None, float, str]:
                """
                Find the single best-matching DB row for *file_display_name*.
                Returns (db_row | None, score, reason_str).

                Pipeline:
                  1. _first_last_match       → floored at 0.90
                  2. _middle_aware_score ≥ 0.82
                  3. raw _similarity    ≥ 0.80
                """
                file_norm = _normalise_for_sim(file_display_name)
                if not file_norm:
                    return None, 0.0, "empty_name"

                best_row    = None
                best_score  = 0.0
                best_reason = "no_match"

                for db_norm, db_row in db_norm_index:
                    if not db_norm:
                        continue

                    if _first_last_match(file_norm, db_norm):
                        raw_fuzzy = _similarity(file_norm, db_norm)
                        score     = max(raw_fuzzy, 0.90)
                        reason    = "first_last_match"
                    else:
                        ma_score = _middle_aware_score(file_norm, db_norm)
                        if ma_score >= 0.82:
                            score  = ma_score
                            reason = "middle_aware_match"
                        else:
                            raw_fuzzy = _similarity(file_norm, db_norm)
                            if raw_fuzzy >= 0.80:
                                score  = raw_fuzzy
                                reason = "fuzzy_match"
                            else:
                                continue

                    if score > best_score or (
                            score == best_score
                            and reason == "first_last_match"
                            and best_reason != "first_last_match"):
                        best_score  = score
                        best_row    = db_row
                        best_reason = reason

                return best_row, best_score, best_reason

            # Tracking counters / logs
            cid_assigned_count = 0
            cid_updated_count  = 0
            no_match_count     = 0
            match_details      = []
            conflict_details   = []

            for key, record in write_ready.items():
                display_name   = record["_display_name"]
                file_cid       = str(record.get("client_id") or "").strip()
                file_cid_upper = file_cid.upper() if file_cid else ""

                # ── Case 1: file has a CID that already exists in DB ──────
                if file_cid_upper and file_cid_upper in existing_cid_map:
                    existing_db_id = existing_cid_map[file_cid_upper]
                    existing_row   = next(
                        (r for r in db_all_rows if r["id"] == existing_db_id), None)
                    if existing_row:
                        existing_name = existing_row.get("applicant_name") or ""
                        file_norm = _normalise_for_sim(display_name)
                        db_norm   = _normalise_for_sim(existing_name)
                        if _first_last_match(file_norm, db_norm):
                            score  = max(_similarity(file_norm, db_norm), 0.90)
                            reason = "first_last_match"
                        else:
                            ma = _middle_aware_score(file_norm, db_norm)
                            score  = ma
                            reason = "middle_aware_match" if ma >= 0.82 else "fuzzy"

                        if score >= 0.75:
                            with _db_connect() as _conn:
                                cur = _conn.cursor()
                                cur.execute(
                                    "UPDATE applicants SET client_id=%s WHERE id=%s",
                                    (file_cid, existing_db_id))
                                _conn.commit()
                                cur.close()
                            cid_updated_count += 1
                            match_details.append(
                                f"✓ Confirmed: '{display_name}' → CID {file_cid} "
                                f"(DB: '{existing_name}', {reason}, {score:.0%})")
                        else:
                            conflict_details.append(
                                f"⚠ Conflict: file '{display_name}' CID {file_cid} "
                                f"already owned by DB '{existing_name}' ({score:.0%})")
                    continue

                # ── Case 2: file has a CID, but it's not yet in the DB ────
                if file_cid_upper:
                    best_row, best_score, best_reason = _best_db_match(display_name)
                    if best_row and best_score >= 0.80:
                        db_id = best_row["id"]
                        with _db_connect() as _conn:
                            cur = _conn.cursor()
                            cur.execute(
                                "UPDATE applicants SET client_id=%s WHERE id=%s",
                                (file_cid, db_id))
                            _conn.commit()
                            cur.close()
                        existing_cid_map[file_cid_upper] = db_id
                        for r in db_all_rows:
                            if r["id"] == db_id:
                                r["client_id"] = file_cid
                                break
                        cid_assigned_count += 1
                        match_details.append(
                            f"✓ Assigned: '{display_name}' → CID {file_cid} "
                            f"(DB: '{best_row['applicant_name']}', "
                            f"{best_reason}, {best_score:.0%})")
                    else:
                        no_match_count += 1
                        best_name = best_row["applicant_name"] if best_row else "—"
                        match_details.append(
                            f"✗ No match: '{display_name}' CID {file_cid} "
                            f"(best DB: '{best_name}', {best_score:.0%})")
                    continue

                # ── Case 3: file has NO CID — use module-level resolver ───
                hits, sim_label = _resolve_name_similarity(display_name)
                if hits:
                    match_details.append(
                        f"ℹ  No CID in file for '{display_name}' "
                        f"({sim_label} name match found, no CID assigned)")
                else:
                    no_match_count += 1
                    match_details.append(
                        f"✗ No match: '{display_name}' (no CID in file, "
                        f"no name match found)")

            # ══════════════════════════════════════════════════════════════
            #  PHASE A.5 — BACKFILL CIDs FOR DB ROWS STILL WITHOUT ONE
            #
            #  For every DB row still missing a client_id after Phase A,
            #  we search all file rows that carry a CID and use the same
            #  three-tier matching pipeline (first_last → middle_aware →
            #  fuzzy) with the same 0.80 threshold.
            #
            #  _middle_aware_score is especially valuable here: a DB record
            #  with "JUAN CARLOS DELA CRUZ" will now correctly match a file
            #  row for "JUAN C DELA CRUZ" (and vice-versa).
            #
            #  CID collision guard: if the best CID is already owned by
            #  another DB row, we skip and log it as a conflict.
            # ══════════════════════════════════════════════════════════════

            file_cid_name_pairs: list[tuple[str, str]] = []
            for key, record in write_ready.items():
                fcid = str(record.get("client_id") or "").strip()
                if fcid:
                    file_cid_name_pairs.append(
                        (_normalise_for_sim(record["_display_name"]), fcid)
                    )

            backfill_assigned = 0
            backfill_log      = []

            if file_cid_name_pairs:
                with _db_connect() as _conn:
                    _cur = _conn.cursor()
                    _cur.execute(
                        "SELECT id, applicant_name FROM applicants "
                        "WHERE (client_id IS NULL OR TRIM(client_id) = '') "
                        "AND applicant_name IS NOT NULL")
                    cols_desc = [desc[0] for desc in _cur.description]
                    no_cid_rows = [dict(zip(cols_desc, r)) for r in _cur.fetchall()]
                    _cur.close()

                for db_row in no_cid_rows:
                    db_norm = _normalise_for_sim(db_row["applicant_name"])
                    if not db_norm:
                        continue

                    best_cid       = None
                    best_score     = 0.0
                    best_reason    = ""
                    best_file_name = ""

                    for file_norm, fcid in file_cid_name_pairs:
                        # Three-tier match (same as _best_db_match)
                        if _first_last_match(file_norm, db_norm):
                            raw_fuzzy = _similarity(file_norm, db_norm)
                            score     = max(raw_fuzzy, 0.90)
                            reason    = "first_last_match"
                        else:
                            ma_score = _middle_aware_score(file_norm, db_norm)
                            if ma_score >= 0.82:
                                score  = ma_score
                                reason = "middle_aware_match"
                            else:
                                raw_fuzzy = _similarity(file_norm, db_norm)
                                if raw_fuzzy >= 0.80:
                                    score  = raw_fuzzy
                                    reason = "fuzzy_match"
                                else:
                                    continue

                        if score > best_score:
                            best_score     = score
                            best_cid       = fcid
                            best_reason    = reason
                            best_file_name = file_norm

                    if best_cid and best_score >= 0.80:
                        best_cid_upper = best_cid.upper()
                        if best_cid_upper not in existing_cid_map:
                            with _db_connect() as _conn:
                                cur = _conn.cursor()
                                cur.execute(
                                    "UPDATE applicants SET client_id=%s WHERE id=%s",
                                    (best_cid, db_row["id"]))
                                _conn.commit()
                                cur.close()
                            existing_cid_map[best_cid_upper] = db_row["id"]
                            for r in db_all_rows:
                                if r["id"] == db_row["id"]:
                                    r["client_id"] = best_cid
                                    break
                            backfill_assigned += 1
                            backfill_log.append(
                                f"✓ Backfilled: DB '{db_row['applicant_name']}' "
                                f"← CID {best_cid} "
                                f"(file: '{best_file_name}', "
                                f"{best_reason}, {best_score:.0%})")
                        else:
                            owner_id   = existing_cid_map[best_cid_upper]
                            owner_name = next(
                                (r["applicant_name"] for r in db_all_rows
                                 if r["id"] == owner_id), "unknown")
                            backfill_log.append(
                                f"⚠ Skip backfill: CID {best_cid} already owned by "
                                f"DB '{owner_name}' — skipped for "
                                f"'{db_row['applicant_name']}' ({best_score:.0%})")

            # ══════════════════════════════════════════════════════════════
            #  PHASE B — WRITE LOAN DATA (all other columns)
            #  Keyed on Client ID wherever possible, name-match fallback.
            # ══════════════════════════════════════════════════════════════

            with _db_connect() as _conn:
                _cur = _conn.cursor()
                _cur.execute(
                    "SELECT client_id, id FROM applicants "
                    "WHERE client_id IS NOT NULL AND TRIM(client_id) != ''")
                cid_to_dbid: dict[str, int] = {
                    str(r[0]).strip().upper(): r[1]
                    for r in _cur.fetchall()
                }
                _cur.close()

            loan_balance_updated = 0
            other_data_updated   = 0
            write_log            = []

            for key, record in write_ready.items():
                display_name = record["_display_name"]
                file_cid     = str(record.get("client_id") or "").strip().upper()

                db_id = None
                if file_cid and file_cid in cid_to_dbid:
                    db_id = cid_to_dbid[file_cid]
                else:
                    best_row, best_score, best_reason = _best_db_match(display_name)
                    if best_row and best_score >= 0.80:
                        db_id = best_row["id"]
                        write_log.append(
                            f"ℹ  Data write via name match: '{display_name}' "
                            f"→ DB id {db_id} ({best_reason}, {best_score:.0%})")

                if db_id is None:
                    write_log.append(
                        f"✗ Cannot write data — no DB row found for '{display_name}'")
                    continue

                parts, vals = [], []
                for db_col, (file_col, is_monetary) in detected.items():
                    if db_col == "client_id":
                        continue
                    val = record.get(db_col)
                    if val is None:
                        continue
                    parts.append(f"{db_col}=%s")
                    vals.append(val)

                if record.get("loan_amount") is not None:
                    parts.append("principal_loan=%s")
                    vals.append(record["loan_amount"])

                if parts:
                    with _db_connect() as _conn:
                        cur = _conn.cursor()
                        cur.execute(
                            f"UPDATE applicants SET {', '.join(parts)} WHERE id=%s",
                            vals + [db_id])
                        _conn.commit()
                        cur.close()
                    if record.get("loan_balance") is not None:
                        loan_balance_updated += 1
                    other_data_updated += 1

            self.after(0, lambda: _refresh_summary(self))

            # ── Result message ────────────────────────────────────────────
            msg  = "=" * 60 + "\n"
            msg += "PRINCIPAL LOAN IMPORT — RESULTS\n"
            msg += "=" * 60 + "\n\n"
            msg += f"✓  Client IDs assigned   : {cid_assigned_count}\n"
            msg += f"✓  Client IDs confirmed  : {cid_updated_count}\n"
            msg += f"✓  Client IDs backfilled : {backfill_assigned}\n"
            msg += f"✓  Loan data rows written: {other_data_updated}\n"
            msg += f"✓  Loan balances updated : {loan_balance_updated}\n"
            msg += f"✗  No match found        : {no_match_count}\n"

            if conflict_details:
                msg += f"\n⚠  CID conflicts ({len(conflict_details)}):\n"
                for d in conflict_details[:10]:
                    msg += f"  {d}\n"
                if len(conflict_details) > 10:
                    msg += f"  … and {len(conflict_details) - 10} more\n"

            if backfill_log:
                msg += f"\n--- BACKFILL LOG ({len(backfill_log)} entries) ---\n"
                for entry in backfill_log[:20]:
                    msg += f"  {entry}\n"
                if len(backfill_log) > 20:
                    msg += f"  … and {len(backfill_log) - 20} more\n"

            if write_log:
                msg += f"\n--- DATA WRITE LOG ({len(write_log)} entries) ---\n"
                for entry in write_log[:15]:
                    msg += f"  {entry}\n"
                if len(write_log) > 15:
                    msg += f"  … and {len(write_log) - 15} more\n"

            if match_details:
                msg += f"\n--- CLIENT ID MATCH DETAILS ---\n"
                for detail in match_details[:25]:
                    msg += f"  {detail}\n"
                if len(match_details) > 25:
                    msg += f"  … and {len(match_details) - 25} more\n"

            self.after(0, lambda: (
                _flash_btn(self, self._sum_import_ploan_btn, "✓  Done!", 2500),
                messagebox.showinfo("P.Loan Import Result", msg)
            ))

        except Exception as exc:
            err = str(exc)
            self.after(0, lambda: (
                _flash_btn(self, self._sum_import_ploan_btn, "✗  Error", 3000),
                messagebox.showerror("P.Loan Import Error", err)
            ))

    threading.Thread(target=_worker, daemon=True).start()


# ═══════════════════════════════════════════════════════════════════════
#  MERGE DB
# ═══════════════════════════════════════════════════════════════════════

def _merge_db_files(self):
    """Merge one or more SQLite .db files into the PostgreSQL database."""
    import sqlite3 as _sqlite3
    paths = filedialog.askopenfilenames(
        title="Select SQLite DB files to merge into PostgreSQL",
        filetypes=[("SQLite DB files", "*.db"), ("All files", "*.*")])
    if not paths:
        return

    _flash_btn(self, self._sum_merge_db_btn, "⟳  Merging…", 60_000)

    _COLS = [
        "session_id", "processed_at", "source_file", "status",
        "applicant_name", "residence_address", "office_address",
        "income_items", "income_total",
        "business_items", "business_total",
        "household_items", "household_total",
        "net_income", "petrol_risk", "transport_risk",
        "results_json", "page_map",
        "amort_history_total", "amort_current_total",
        "client_id", "pn", "industry_name",
        "loan_balance", "amortized_cost",
        "principal_loan", "maturity", "interest_rate",
        "branch", "loan_class_name", "product_name",
        "loan_date", "term_unit", "term", "security", "release_tag",
        "loan_amount", "loan_status", "ao_name",
    ]
    _PG_INSERT = (
        f"INSERT INTO applicants ({', '.join(_COLS)}) "
        f"VALUES ({', '.join(['%s'] * len(_COLS))})"
    )

    def _worker():
        total_inserted = 0; total_skipped = 0; total_patched = 0
        file_results   = []

        try:
            for src in paths:
                try:
                    # Read source SQLite file
                    s_conn = _sqlite3.connect(str(src), timeout=10)
                    s_conn.row_factory = _sqlite3.Row
                    src_rows = [dict(r) for r in s_conn.execute(
                        "SELECT * FROM applicants").fetchall()]
                    s_conn.close()
                except Exception as e:
                    file_results.append((Path(src).name, 0, 0, 0, str(e)))
                    continue

                ins = skp = pat = 0

                with _db_connect() as out_conn:
                    pg_cur = out_conn.cursor()

                    # Build lookup sets from PostgreSQL
                    pg_cur.execute(
                        "SELECT client_id, id FROM applicants "
                        "WHERE client_id IS NOT NULL AND TRIM(client_id) != ''")
                    existing_by_clientid = {
                        str(r[0]).strip().upper(): r[1]
                        for r in pg_cur.fetchall()}

                    pg_cur.execute(
                        "SELECT session_id, source_file FROM applicants")
                    existing_primary = {
                        (r[0], r[1]) for r in pg_cur.fetchall()}

                    pg_cur.execute(
                        "SELECT applicant_name, source_file FROM applicants")
                    existing_fallback = {
                        (str(r[0]).strip().upper(), str(r[1]).strip().upper())
                        for r in pg_cur.fetchall()}

                    for rd in src_rows:
                        incoming_cid = str(rd.get("client_id") or "").strip().upper()
                        pk = (rd.get("session_id", ""), rd.get("source_file", ""))
                        fk = (
                            str(rd.get("applicant_name") or "").strip().upper(),
                            str(rd.get("source_file")    or "").strip().upper(),
                        )

                        if incoming_cid and incoming_cid in existing_by_clientid:
                            _patch_existing(out_conn,
                                            existing_by_clientid[incoming_cid], rd)
                            pat += 1; continue

                        if pk in existing_primary or fk in existing_fallback:
                            skp += 1; continue

                        pg_cur.execute(_PG_INSERT,
                                       [rd.get(c) for c in _COLS])
                        if incoming_cid:
                            pg_cur.execute("SELECT lastval()")
                            new_id = pg_cur.fetchone()[0]
                            existing_by_clientid[incoming_cid] = new_id
                        existing_primary.add(pk)
                        existing_fallback.add(fk)
                        ins += 1

                    out_conn.commit()
                    pg_cur.close()

                file_results.append((Path(src).name, ins, pat, skp, None))
                total_inserted += ins
                total_patched  += pat
                total_skipped  += skp

            _db_deduplicate_client_ids()
            self.after(0, lambda: _refresh_summary(self))

            msg  = "Merge complete.\n\n"
            msg += f"✓  Inserted : {total_inserted:,} new record(s)\n"
            msg += f"🔧  Patched  : {total_patched:,} existing record(s)\n"
            msg += f"–  Skipped  : {total_skipped:,} (exact duplicates)\n\n"
            msg += "Per file:\n"
            for fname, ins, pat, skp, err in file_results:
                if err:
                    msg += f"  ✗  {fname}  →  Error: {err}\n"
                else:
                    msg += (f"  ✓  {fname}  →  {ins:,} inserted, "
                            f"{pat:,} patched, {skp:,} skipped\n")

            self.after(0, lambda: (
                _flash_btn(self, self._sum_merge_db_btn, "✓  Done!", 2500),
                messagebox.showinfo("Merge DB Result", msg)
            ))
        except Exception as exc:
            err = str(exc)
            self.after(0, lambda: (
                _flash_btn(self, self._sum_merge_db_btn, "✗  Error", 3000),
                messagebox.showerror("Merge DB Error", err)
            ))

    threading.Thread(target=_worker, daemon=True).start()


# ═══════════════════════════════════════════════════════════════════════
#  EXPORT HELPERS
# ═══════════════════════════════════════════════════════════════════════

def _get_all_filtered_rows(self) -> list:
    raw    = self._sum_search_var.get().strip()
    search = "" if "separate terms with commas" in raw else raw
    rows, _ = _db_query(
        search=search, session_id=self._sum_session_filter,
        sort_col=self._sum_sort_col, sort_asc=self._sum_sort_asc,
        offset=0, limit=100_000,
        adv_filters=getattr(self, "_sum_adv_filters", {}))
    return [dict(r) for r in rows]


def _row_to_export_dict(row: dict) -> dict:
    def _fmt(val):
        try:
            return float(val) if val not in (None, "") else None
        except Exception:
            return None

    # Derive virtual fields from results_json
    try:
        results = json.loads(row.get("results_json", "") or "{}")
    except Exception:
        results = {}

    spouse_info        = _extract_spouse_info(results)
    personal_assets    = _extract_asset_items(results, "cibi_personal_assets")
    business_assets    = _extract_asset_items(results, "cibi_business_assets")
    business_inventory = _extract_asset_items(results, "cibi_business_inventory")

    # amort_history_total: prefer real DB column, fall back to JSON parse
    amort_hist = row.get("amort_history_total")
    if amort_hist is None:
        amort_hist = _parse_amort_history_total(row.get("results_json", ""))

    return {
        "Client ID":                           row.get("client_id",          "") or "",
        "PN":                                  row.get("pn",                 "") or "",
        "Applicant":                           row.get("applicant_name",     "") or "",
        "Residence Address":                   row.get("residence_address",  "") or "",
        "Office Address":                      row.get("office_address",     "") or "",
        "Industry Name":                       row.get("industry_name",      "") or "",
        "Spouse Info":                         spouse_info,
        "Personal Assets":                     personal_assets,
        "Business Assets":                     business_assets,
        "Business Inventory":                  business_inventory,
        "Source of Income":                    row.get("income_items",       "") or "",
        "Total Source Of Income":              _fmt(row.get("income_total")),
        "Business Expenses":                   row.get("business_items",     "") or "",
        "Total Business Expenses":             _fmt(row.get("business_total")),
        "Household / Personal Expenses":       row.get("household_items",    "") or "",
        "Total Household / Personal Expenses": _fmt(row.get("household_total")),
        "Total Net Income":                    _fmt(row.get("net_income")),
        "Total Amortization History":          _fmt(amort_hist),
        "Total Current Amortization":          _fmt(row.get("amort_current_total")),
        "Loan Balance":                        _fmt(row.get("loan_balance")),
        "Principal Loan":                      _fmt(row.get("principal_loan")),
        "Maturity":                            row.get("maturity",           "") or "",
        "Interest Rate":                       row.get("interest_rate",      "") or "",
        # ── P.Loan expanded fields ─────────────────────────────────────
        "Branch":                              row.get("branch",             "") or "",
        "Loan Class":                          row.get("loan_class_name",    "") or "",
        "Product Name":                        row.get("product_name",       "") or "",
        "Loan Date":                           row.get("loan_date",          "") or "",
        "Term Unit":                           row.get("term_unit",          "") or "",
        "Term":                                row.get("term",               "") or "",
        "Security":                            row.get("security",           "") or "",
        "Release Tag":                         row.get("release_tag",        "") or "",
        "Loan Amount":                         _fmt(row.get("loan_amount")),
        "Loan Status":                         row.get("loan_status",        "") or "",
        "AO Name":                             row.get("ao_name",            "") or "",
    }


def _export_csv(self):
    rows = _get_all_filtered_rows(self)
    if not rows:
        return
    path = filedialog.asksaveasfilename(
        title="Export to CSV", defaultextension=".csv",
        filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        initialfile=f"LookUp_Summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
    if not path:
        return
    flat    = [_row_to_export_dict(r) for r in rows]
    headers = list(flat[0].keys()) if flat else []
    try:
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(flat)
        _flash_btn(self, self._sum_export_csv_btn, "✓  Saved!", 2000)
    except Exception as e:
        _flash_btn(self, self._sum_export_csv_btn, f"Error: {e}", 3000)


def _export_excel(self):
    rows = _get_all_filtered_rows(self)
    if not rows:
        messagebox.showinfo("Export", "No records to export.")
        return

    flat = [_row_to_export_dict(r) for r in rows]
    if not flat:
        return

    checklist_result = _show_export_checklist(self, flat)
    if checklist_result is None:
        return
    flat, selected_col_keys = checklist_result
    if not flat:
        messagebox.showinfo("Export", "No rows were selected — nothing exported.")
        return
    if not selected_col_keys:
        messagebox.showinfo("Export", "No columns were selected — nothing exported.")
        return

    path = filedialog.asksaveasfilename(
        title="Export to Excel", defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialfile=f"LookUp_Summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    if not path:
        return

    def _worker():
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            all_possible_headers = [c[0] for c in _ALL_EXPORT_COLS]
            selected_set         = set(selected_col_keys)
            headers              = [h for h in all_possible_headers
                                    if h in selected_set]
            if not headers:
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Look-Up Summary"

            hdr_fill  = PatternFill("solid", fgColor="93C47D")
            tot_fill  = PatternFill("solid", fgColor="D9EAD3")
            avg_fill  = PatternFill("solid", fgColor="B8D0E5")
            even_fill = PatternFill("solid", fgColor="FFFFFF")
            odd_fill  = PatternFill("solid", fgColor="F3F9F0")
            hdr_font  = Font(name="Roboto", bold=True, color="FFFFFF", size=10)
            body_font = Font(name="Roboto", size=9)
            bold_font = Font(name="Roboto", bold=True, size=9)
            net_font  = Font(name="Roboto", bold=True, size=9, color="1F6B28")
            tot_font  = Font(name="Roboto", bold=True, size=10)
            tot_font_j= Font(name="Roboto", bold=True, size=10, color="1F6B28")
            avg_font  = Font(name="Roboto", bold=True, size=10, color="1A3A5C")
            thin      = Side(style="thin",   color="CCCCCC")
            med       = Side(style="medium", color="555555")
            cell_bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
            tot_bdr   = Border(left=med,  right=med,  top=med,  bottom=med)
            wrap_al   = Alignment(horizontal="left",  vertical="top", wrap_text=True)
            right_al  = Alignment(horizontal="right", vertical="top")
            right_c   = Alignment(horizontal="right", vertical="center")
            left_c    = Alignment(horizontal="left",  vertical="center")
            CURRENCY  = '#,##0.00;(#,##0.00);"-"'

            TOTAL_COLS = {
                "Total Source Of Income", "Total Business Expenses",
                "Total Household / Personal Expenses",
                "Total Amortization History", "Total Current Amortization",
                "Loan Balance", "Principal Loan",
                "Loan Amount",
            }
            NET_COL  = "Total Net Income"
            SUM_COLS = TOTAL_COLS | {NET_COL}

            col_widths = {
                "Client ID": 14, "PN": 12,
                "Applicant": 22, "Residence Address": 30, "Office Address": 26,
                "Industry Name": 20,
                "Spouse Info": 28, "Personal Assets": 28,
                "Business Assets": 28, "Business Inventory": 24,
                "Source of Income": 32, "Total Source Of Income": 22,
                "Business Expenses": 32, "Total Business Expenses": 22,
                "Household / Personal Expenses": 36,
                "Total Household / Personal Expenses": 24,
                "Total Net Income": 20,
                "Total Amortization History": 26,
                "Total Current Amortization": 26,
                "Loan Balance": 22,
                "Principal Loan": 22, "Maturity": 20, "Interest Rate": 18,
                "Branch": 18, "Loan Class": 20, "Product Name": 22,
                "Loan Date": 16, "Term Unit": 14,
                "Term": 10, "Security": 24, "Release Tag": 16,
                "Loan Amount": 20,
                "Loan Status": 16, "AO Name": 22,
            }

            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.font = hdr_font; cell.fill = hdr_fill
                cell.alignment = wrap_al; cell.border = cell_bdr
                ws.column_dimensions[get_column_letter(ci)].width = \
                    col_widths.get(h, 20)
            ws.row_dimensions[1].height = 28

            for ri, row_dict in enumerate(flat, 2):
                bg_fill    = even_fill if (ri - 2) % 2 == 0 else odd_fill
                text_lines = []
                for ci, h in enumerate(headers, 1):
                    val  = row_dict.get(h)
                    cell = ws.cell(row=ri, column=ci,
                                   value=val if val is not None else None)
                    cell.fill = bg_fill; cell.border = cell_bdr
                    if h == NET_COL:
                        cell.number_format = CURRENCY
                        cell.font = net_font; cell.alignment = right_al
                    elif h in TOTAL_COLS:
                        cell.number_format = CURRENCY
                        cell.font = bold_font; cell.alignment = right_al
                    elif h in ("Applicant", "Client ID", "PN"):
                        cell.font = bold_font; cell.alignment = wrap_al
                    else:
                        cell.font = body_font; cell.alignment = wrap_al
                        if val:
                            text_lines.append(len(str(val).split("\n")))
                max_lines = max(text_lines) if text_lines else 1
                ws.row_dimensions[ri].height = max(18, min(max_lines * 15, 150))

            first_data = 2; last_data = len(flat) + 1
            total_row  = last_data + 1; avg_row = last_data + 2

            for ci, h in enumerate(headers, 1):
                col_l = get_column_letter(ci)
                cell  = ws.cell(row=total_row, column=ci)
                cell.fill = tot_fill; cell.border = tot_bdr
                if h == "Applicant":
                    cell.value = "TOTAL"; cell.font = tot_font; cell.alignment = left_c
                elif h == NET_COL:
                    cell.value = f"=SUM({col_l}{first_data}:{col_l}{last_data})"
                    cell.number_format = CURRENCY
                    cell.font = tot_font_j; cell.alignment = right_c
                elif h in SUM_COLS:
                    cell.value = f"=SUM({col_l}{first_data}:{col_l}{last_data})"
                    cell.number_format = CURRENCY
                    cell.font = tot_font; cell.alignment = right_c
                else:
                    cell.font = tot_font; cell.alignment = left_c

                cell_avg = ws.cell(row=avg_row, column=ci)
                cell_avg.fill = avg_fill; cell_avg.border = tot_bdr
                if h == "Applicant":
                    cell_avg.value = "AVERAGE"
                    cell_avg.font = avg_font; cell_avg.alignment = left_c
                elif h == NET_COL:
                    cell_avg.value = f"=AVERAGE({col_l}{first_data}:{col_l}{last_data})"
                    cell_avg.number_format = CURRENCY
                    cell_avg.font = avg_font; cell_avg.alignment = right_c
                elif h in SUM_COLS:
                    cell_avg.value = f"=AVERAGE({col_l}{first_data}:{col_l}{last_data})"
                    cell_avg.number_format = CURRENCY
                    cell_avg.font = avg_font; cell_avg.alignment = right_c
                else:
                    cell_avg.font = avg_font; cell_avg.alignment = left_c

            ws.row_dimensions[total_row].height = 22
            ws.row_dimensions[avg_row].height   = 22
            ws.freeze_panes = "A2"
            wb.save(path)

            self.after(0, lambda: (
                _flash_btn(self, self._sum_export_xl_btn, "✓  Saved!", 2000),
                messagebox.showinfo("Export Excel",
                    f"Saved {len(flat):,} record(s) · {len(headers)} column(s)\n"
                    f"with TOTAL and AVERAGE rows to:\n{path}")
            ))
        except ImportError:
            self.after(0, lambda: (
                _flash_btn(self, self._sum_export_xl_btn, "openpyxl missing", 3000),
                messagebox.showerror("Export Excel Error",
                    "openpyxl is not installed.\nRun: pip install openpyxl")
            ))
        except Exception as e:
            self.after(0, lambda err=str(e): (
                _flash_btn(self, self._sum_export_xl_btn, "✗  Error", 3000),
                messagebox.showerror("Export Excel Error", err)
            ))

    threading.Thread(target=_worker, daemon=True).start()


# ═══════════════════════════════════════════════════════════════════════
#  VALIDATE CLIENT ID ↔ APPLICANT NAME
# ═══════════════════════════════════════════════════════════════════════

def _validate_clients(self):
    path = filedialog.askopenfilename(
        title="Select Reference File (Client ID + Client Name)",
        filetypes=[("Excel & CSV files", "*.xlsx *.csv"),
                   ("Excel files", "*.xlsx"),
                   ("CSV files", "*.csv"),
                   ("All files", "*.*")])
    if not path:
        return

    out_path = filedialog.asksaveasfilename(
        title="Save Validation Report as…",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialfile=f"Validation_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    if not out_path:
        return

    _flash_btn(self, self._sum_validate_btn, "⟳  Validating…", 60_000)

    def _worker():
        try:
            def _find_col(cols, *keywords):
                for kw in keywords:
                    kw_norm = re.sub(r"[\s_\-]", "", kw.lower())
                    for c in cols:
                        c_norm = re.sub(r"[\s_\-]", "", c.lower())
                        if kw_norm == c_norm or kw_norm in c_norm:
                            return c
                return None

            if path.lower().endswith(".csv"):
                import csv as _csv
                with open(path, newline="", encoding="utf-8-sig") as f:
                    reader      = _csv.DictReader(f)
                    ref_records = [dict(row) for row in reader]
                all_cols = list(ref_records[0].keys()) if ref_records else []
            else:
                import openpyxl as _oxl
                wb = _oxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
                if header_row is None:
                    raise ValueError("Reference file appears to be empty.")
                all_cols = [
                    str(cell.value).strip() if cell.value is not None else ""
                    for cell in header_row
                ]
                ref_records = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if all(v is None for v in row):
                        continue
                    ref_records.append({
                        all_cols[i]: (str(v).strip() if v is not None else "")
                        for i, v in enumerate(row) if i < len(all_cols)
                    })
                wb.close()

            if not ref_records:
                raise ValueError("No data rows found in the reference file.")

            col_cid  = _find_col(all_cols, "clientid", "client id", "client_id", "cid")
            col_name = _find_col(all_cols, "clientname", "client name", "applicant",
                                 "applicantname", "name")
            if not col_cid:
                raise ValueError(
                    f"Could not detect a Client ID column.\n\nFile has: {', '.join(all_cols)}")
            if not col_name:
                raise ValueError(
                    f"Could not detect a Client Name column.\n\nFile has: {', '.join(all_cols)}")

            ref_map:       dict[str, str] = {}
            ref_order:     list[tuple]    = []
            seen_ref_cids: set[str]       = set()

            for rec in ref_records:
                cid  = str(rec.get(col_cid,  "") or "").strip()
                name = str(rec.get(col_name, "") or "").strip()
                cid_up = cid.upper()
                if cid and name and cid_up not in seen_ref_cids:
                    ref_map[cid_up] = name
                    ref_order.append((cid, name))
                    seen_ref_cids.add(cid_up)

            with _db_connect() as conn:
                cur = conn.cursor()
                cur.execute(
                    "SELECT id, client_id, applicant_name, pn, industry_name, "
                    "residence_address, office_address, "
                    "income_total, business_total, household_total, net_income, "
                    "amort_history_total, amort_current_total, "
                    "loan_balance, amortized_cost, "
                    "principal_loan, maturity, interest_rate, "
                    "branch, loan_class_name, product_name, loan_date, security, loan_status, ao_name, "
                    "loan_amount "
                    "FROM applicants")
                cols_desc = [desc[0] for desc in cur.description]
                db_rows = [dict(zip(cols_desc, r)) for r in cur.fetchall()]
                cur.close()

            matched:    list[tuple] = []
            unmatched:  list[tuple] = []
            no_cid:     list[tuple] = []
            # ── CRITICAL: db_cid_set tracks every non-empty client_id
            #    that EXISTS in the DB, regardless of name match result.
            #    This is what we compare the reference list against.
            db_cid_set: set[str]   = set()

            for r in db_rows:
                db_cid  = str(r.get("client_id",      "") or "").strip()
                db_name = str(r.get("applicant_name", "") or "").strip()

                if not db_cid:
                    no_cid.append((db_cid, db_name or "(no name)"))
                    continue

                # Add to set BEFORE any match logic so reference comparison is complete
                db_cid_up = db_cid.upper()
                db_cid_set.add(db_cid_up)

                if db_cid_up not in ref_map:
                    unmatched.append((db_cid, db_name, "Client ID not in reference"))
                    continue

                ref_name = ref_map[db_cid_up]
                score    = _similarity(
                    _normalise_for_sim(db_name.upper()),
                    _normalise_for_sim(ref_name.upper()))
                if score >= SIMILARITY_THRESHOLD:
                    ref_cid_original = next(
                        (cid for cid, nm in ref_order if cid.upper() == db_cid_up),
                        db_cid
                    )
                    matched.append((db_cid, db_name, ref_cid_original, ref_name))
                else:
                    unmatched.append((db_cid, db_name,
                                      f"Name mismatch — reference: {ref_name}"))

            # ── Every reference entry whose client_id does NOT appear
            #    anywhere in the DB (matched, unmatched, or otherwise)
            ref_not_in_db = [
                (cid, name) for cid, name in ref_order
                if cid.upper() not in db_cid_set
            ]

            CHECKED_COLS = [
                ("client_id",           "Client ID"),
                ("pn",                  "PN"),
                ("applicant_name",      "Applicant Name"),
                ("residence_address",   "Residence Address"),
                ("office_address",      "Office Address"),
                ("industry_name",       "Industry Name"),
                ("income_total",        "Total Income"),
                ("business_total",      "Total Business Expenses"),
                ("household_total",     "Total Household Expenses"),
                ("net_income",          "Total Net Income"),
                ("amort_history_total", "Total Amortization History"),
                ("amort_current_total", "Total Current Amortization"),
                ("loan_balance",        "Loan Balance"),
                ("principal_loan",      "Principal Loan"),
                ("maturity",            "Maturity"),
                ("interest_rate",       "Interest Rate"),
                ("branch",              "Branch"),
                ("loan_class_name",     "Loan Class"),
                ("product_name",        "Product Name"),
                ("loan_date",           "Loan Date"),
                ("security",            "Security"),
                ("loan_status",         "Loan Status"),
                ("ao_name",             "AO Name"),
                ("loan_amount",         "Loan Amount"),
            ]

            missing_info_rows: list[dict] = []
            for r in db_rows:
                missing_cols = []
                for col_key, col_label in CHECKED_COLS:
                    val = r.get(col_key)
                    if val is None or (isinstance(val, str) and val.strip() == ""):
                        missing_cols.append(col_label)
                if missing_cols:
                    missing_info_rows.append({
                        "client_id":      r.get("client_id",      "") or "",
                        "applicant_name": r.get("applicant_name", "") or "(no name)",
                        "missing":        missing_cols,
                    })

            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb_out   = openpyxl.Workbook()
            thin     = Side(style="thin",   color="CCCCCC")
            med      = Side(style="medium", color="555555")
            cell_bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
            hdr_bdr  = Border(left=med,  right=med,  top=med,  bottom=med)
            wrap_al  = Alignment(horizontal="left",   vertical="top",  wrap_text=True)
            ctr_al   = Alignment(horizontal="center", vertical="center")
            left_c   = Alignment(horizontal="left",   vertical="center")
            hdr_font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=10)
            body_f   = Font(name="Segoe UI", size=9)
            bold_f   = Font(name="Segoe UI", bold=True, size=9)

            fill_hdr_green = PatternFill("solid", fgColor="2D6A4F")
            fill_hdr_red   = PatternFill("solid", fgColor="9B2226")
            fill_hdr_gray  = PatternFill("solid", fgColor="4A4E69")
            fill_hdr_blue  = PatternFill("solid", fgColor="1A3A5C")
            fill_match     = PatternFill("solid", fgColor="D8F3DC")
            fill_unmatch   = PatternFill("solid", fgColor="FFE8E8")
            fill_nocid     = PatternFill("solid", fgColor="FFF3CD")
            fill_notindb   = PatternFill("solid", fgColor="E8F0FF")
            fill_section   = PatternFill("solid", fgColor="F0F4FA")
            fill_missing   = PatternFill("solid", fgColor="FFF8E8")

            def _hc(ws, row, col, value, fill):
                c = ws.cell(row=row, column=col, value=value)
                c.font = hdr_font; c.fill = fill
                c.alignment = ctr_al; c.border = hdr_bdr
                return c

            def _bc(ws, row, col, value, fill, fnt=None, al=None):
                c = ws.cell(row=row, column=col, value=value)
                c.font = fnt or body_f; c.fill = fill
                c.alignment = al or wrap_al; c.border = cell_bdr
                return c

            def _section_label(ws, row, text, fill_bg, txt_color, ncols):
                c = ws.cell(row=row, column=1, value=text)
                c.font = Font(name="Segoe UI", bold=True, size=9, color=txt_color)
                c.fill = fill_bg; c.alignment = wrap_al; c.border = cell_bdr
                if ncols > 1:
                    ws.merge_cells(start_row=row, start_column=1,
                                   end_row=row,   end_column=ncols)
                ws.row_dimensions[row].height = 18

            # ── Sheet 1: Summary ───────────────────────────────────────
            ws1 = wb_out.active
            ws1.title = "Summary"
            ws1.column_dimensions["A"].width = 38
            ws1.column_dimensions["B"].width = 18

            ws1.cell(row=1, column=1,
                     value="Validation Report — Summary").font = \
                Font(name="Segoe UI", bold=True, size=13, color="1A3A5C")
            ws1.cell(row=1, column=1).alignment = ctr_al
            ws1.merge_cells("A1:B1")
            ws1.row_dimensions[1].height = 28

            ws1.cell(row=2, column=1,
                     value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = \
                Font(name="Segoe UI", size=8, color="888888")
            ws1.merge_cells("A2:B2")
            ws1.row_dimensions[2].height = 16

            total_db    = len(db_rows)
            n_matched   = len(matched)
            n_unmatched = len(unmatched)
            n_no_cid    = len(no_cid)
            n_notindb   = len(ref_not_in_db)
            n_miss_rows = len(missing_info_rows)
            rate        = f"{n_matched / total_db * 100:.1f}%" if total_db else "N/A"

            _hc(ws1, 3, 1, "Category", fill_hdr_gray)
            _hc(ws1, 3, 2, "Count",    fill_hdr_gray)
            ws1.row_dimensions[3].height = 22

            for ri, (label, count, fill) in enumerate([
                ("Total Records in DB",                  total_db,     fill_section),
                ("✓  Matched (ID + Name correct)",        n_matched,    fill_match),
                ("✗  Unmatched (ID or Name mismatch)",    n_unmatched,  fill_unmatch),
                ("⚠  No Client ID in DB (unvalidatable)", n_no_cid,     fill_nocid),
                ("—  In Reference but not in DB",         n_notindb,    fill_notindb),
                ("⚠  Records with missing column data",   n_miss_rows,  fill_missing),
                ("Total Unique Entries in Reference",     len(ref_map), fill_section),
            ], 4):
                _bc(ws1, ri, 1, label, fill, bold_f, left_c)
                _bc(ws1, ri, 2, count, fill, bold_f, ctr_al)
                ws1.row_dimensions[ri].height = 20

            rate_row = 4 + 7
            c1 = ws1.cell(row=rate_row, column=1, value="Match Rate")
            c1.font = Font(name="Segoe UI", bold=True, size=10, color="1F6B28")
            c1.alignment = left_c
            c2 = ws1.cell(row=rate_row, column=2, value=rate)
            c2.font = Font(name="Segoe UI", bold=True, size=10, color="1F6B28")
            c2.alignment = ctr_al
            ws1.row_dimensions[rate_row].height = 22
            ws1.freeze_panes = "A4"

            # ── Sheet 2: Unmatched & Not in DB ─────────────────────────
            ws2 = wb_out.create_sheet("Unmatched & Not In DB")
            ws2.column_dimensions["A"].width = 18
            ws2.column_dimensions["B"].width = 30
            ws2.column_dimensions["C"].width = 38
            ws2.column_dimensions["D"].width = 16
            _hc(ws2, 1, 1, "Client ID",     fill_hdr_red)
            _hc(ws2, 1, 2, "Applicant",     fill_hdr_red)
            _hc(ws2, 1, 3, "Reason / Note", fill_hdr_red)
            _hc(ws2, 1, 4, "Issue Type",    fill_hdr_red)
            ws2.row_dimensions[1].height = 24

            ri = 2

            # ── SECTION A: DB records unmatched ────────────────────────
            _section_label(ws2, ri,
                "SECTION A — DB Records: Unmatched (Client ID or Name mismatch)",
                PatternFill("solid", fgColor="FFE8E8"), "9B2226", 4)
            ri += 1
            if unmatched:
                for db_cid, db_name, reason in unmatched:
                    issue = "Name Mismatch" if "mismatch" in reason.lower() else "ID Not in Ref"
                    _bc(ws2, ri, 1, db_cid,  fill_unmatch, bold_f)
                    _bc(ws2, ri, 2, db_name, fill_unmatch)
                    _bc(ws2, ri, 3, reason,  fill_unmatch)
                    _bc(ws2, ri, 4, issue,   fill_unmatch, bold_f, ctr_al)
                    ws2.row_dimensions[ri].height = 18
                    ri += 1
            else:
                c = ws2.cell(row=ri, column=1, value="✓  No unmatched DB records.")
                c.font = Font(name="Segoe UI", bold=True, size=9, color="1F6B28")
                c.fill = fill_unmatch; c.alignment = wrap_al
                ws2.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=4)
                ws2.row_dimensions[ri].height = 18
                ri += 1

            # ── SECTION A2: DB records with no client ID ───────────────
            ri += 1  # blank spacer row
            _section_label(ws2, ri,
                "SECTION A2 — DB Records with No Client ID (cannot be validated)",
                PatternFill("solid", fgColor="FFF3CD"), "7D5A00", 4)
            ri += 1
            if no_cid:
                for raw_cid, name in no_cid:
                    _bc(ws2, ri, 1, raw_cid or "(empty)",      fill_nocid, bold_f)
                    _bc(ws2, ri, 2, name,                       fill_nocid)
                    _bc(ws2, ri, 3, "Missing Client ID in DB",  fill_nocid)
                    _bc(ws2, ri, 4, "No ID",                    fill_nocid, bold_f, ctr_al)
                    ws2.row_dimensions[ri].height = 18
                    ri += 1
            else:
                c = ws2.cell(row=ri, column=1, value="✓  All DB records have a Client ID.")
                c.font = Font(name="Segoe UI", bold=True, size=9, color="1F6B28")
                c.fill = fill_nocid; c.alignment = wrap_al
                ws2.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=4)
                ws2.row_dimensions[ri].height = 18
                ri += 1

            # ── SECTION B: Reference entries NOT found in DB ───────────
            # This is the key section — shows EVERY reference client whose
            # client_id does not exist anywhere in the DB table.
            ri += 1  # blank spacer row
            _section_label(ws2, ri,
                f"SECTION B — Reference Entries with NO Matching Client ID in DB  ({len(ref_not_in_db):,} records)",
                PatternFill("solid", fgColor="E8F0FF"), "1A3A5C", 4)
            ri += 1
            if ref_not_in_db:
                for ref_cid, ref_name in ref_not_in_db:
                    _bc(ws2, ri, 1, ref_cid,                         fill_notindb, bold_f)
                    _bc(ws2, ri, 2, ref_name,                         fill_notindb)
                    _bc(ws2, ri, 3, "Client ID not found in DB",      fill_notindb)
                    _bc(ws2, ri, 4, "Not in DB",                      fill_notindb, bold_f, ctr_al)
                    ws2.row_dimensions[ri].height = 18
                    ri += 1
            else:
                c = ws2.cell(row=ri, column=1,
                             value="✓  All reference entries exist in the DB.")
                c.font = Font(name="Segoe UI", bold=True, size=9, color="1F6B28")
                c.fill = fill_notindb; c.alignment = wrap_al
                ws2.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=4)
                ws2.row_dimensions[ri].height = 18

            ws2.freeze_panes = "A2"

            # ── Sheet 3: Missing Info ───────────────────────────────────
            ws3 = wb_out.create_sheet("Missing Info")
            ws3.column_dimensions["A"].width = 18
            ws3.column_dimensions["B"].width = 30
            ws3.column_dimensions["C"].width = 16
            ws3.column_dimensions["D"].width = 55
            _hc(ws3, 1, 1, "Client ID",      fill_hdr_red)
            _hc(ws3, 1, 2, "Applicant",       fill_hdr_red)
            _hc(ws3, 1, 3, "Missing Count",   fill_hdr_red)
            _hc(ws3, 1, 4, "Missing Columns", fill_hdr_red)
            ws3.row_dimensions[1].height = 24

            if missing_info_rows:
                missing_info_rows.sort(key=lambda x: len(x["missing"]), reverse=True)
                for ri3, entry in enumerate(missing_info_rows, 2):
                    n_miss = len(entry["missing"])
                    fill_row = (PatternFill("solid", fgColor="FFD0D0") if n_miss >= 8
                                else PatternFill("solid", fgColor="FFE8CC") if n_miss >= 4
                                else fill_missing)
                    _bc(ws3, ri3, 1, entry["client_id"],          fill_row, bold_f)
                    _bc(ws3, ri3, 2, entry["applicant_name"],      fill_row)
                    _bc(ws3, ri3, 3, n_miss,                       fill_row, bold_f, ctr_al)
                    _bc(ws3, ri3, 4, ", ".join(entry["missing"]),  fill_row)
                    ws3.row_dimensions[ri3].height = 20
            else:
                c = ws3.cell(row=2, column=1,
                             value="✓  All records have complete information.")
                c.font = Font(name="Segoe UI", bold=True, size=10, color="1F6B28")
                c.alignment = wrap_al
                ws3.merge_cells("A2:D2")
                ws3.row_dimensions[2].height = 22
            ws3.freeze_panes = "A2"

            # ── Sheet 4: Matched Comparison ─────────────────────────────
            ws4 = wb_out.create_sheet("Matched Comparison")
            ws4.column_dimensions["A"].width = 20
            ws4.column_dimensions["B"].width = 32
            ws4.column_dimensions["C"].width = 20
            ws4.column_dimensions["D"].width = 32
            ws4.column_dimensions["E"].width = 18

            _hc(ws4, 1, 1, "Client ID (DB)",          fill_hdr_green)
            _hc(ws4, 1, 2, "Applicant (DB)",           fill_hdr_green)
            _hc(ws4, 1, 3, "Client ID (Reference)",    fill_hdr_green)
            _hc(ws4, 1, 4, "Client Name (Reference)",  fill_hdr_green)
            _hc(ws4, 1, 5, "Name Similarity %",        fill_hdr_green)
            ws4.row_dimensions[1].height = 24

            if matched:
                for ri4, (db_cid, db_name, ref_cid, ref_name) in enumerate(matched, 2):
                    score_pct = round(
                        _similarity(
                            _normalise_for_sim(db_name.upper()),
                            _normalise_for_sim(ref_name.upper())
                        ) * 100, 1
                    )
                    fill_row = (fill_match if score_pct == 100
                                else PatternFill("solid", fgColor="C7F9CC"))
                    _bc(ws4, ri4, 1, db_cid,         fill_row, bold_f)
                    _bc(ws4, ri4, 2, db_name,         fill_row)
                    _bc(ws4, ri4, 3, ref_cid,         fill_row, bold_f)
                    _bc(ws4, ri4, 4, ref_name,        fill_row)
                    _bc(ws4, ri4, 5, f"{score_pct}%", fill_row, bold_f, ctr_al)
                    ws4.row_dimensions[ri4].height = 18
            else:
                c = ws4.cell(row=2, column=1, value="No matched records found.")
                c.font = Font(name="Segoe UI", bold=True, size=10, color="888888")
                ws4.merge_cells("A2:E2")
                ws4.row_dimensions[2].height = 22
            ws4.freeze_panes = "A2"

            # ── Sheet 5: Reference Not In DB (dedicated full sheet) ─────
            # Standalone sheet so it's never truncated or mixed with other data
            ws5 = wb_out.create_sheet("Ref Not In DB")
            ws5.column_dimensions["A"].width = 22
            ws5.column_dimensions["B"].width = 38
            ws5.column_dimensions["C"].width = 20

            _hc(ws5, 1, 1, "Client ID (Reference)",  fill_hdr_blue)
            _hc(ws5, 1, 2, "Client Name (Reference)", fill_hdr_blue)
            _hc(ws5, 1, 3, "Status",                  fill_hdr_blue)
            ws5.row_dimensions[1].height = 24

            ws5.cell(row=2, column=1,
                     value=f"Total: {len(ref_not_in_db):,} reference entries have no matching Client ID in the DB.").font = \
                Font(name="Segoe UI", bold=True, size=9, color="1A3A5C")
            ws5.cell(row=2, column=1).alignment = wrap_al
            ws5.merge_cells("A2:C2")
            ws5.row_dimensions[2].height = 18

            if ref_not_in_db:
                for ri5, (ref_cid, ref_name) in enumerate(ref_not_in_db, 3):
                    _bc(ws5, ri5, 1, ref_cid,              fill_notindb, bold_f)
                    _bc(ws5, ri5, 2, ref_name,              fill_notindb)
                    _bc(ws5, ri5, 3, "Not in DB",           fill_notindb, bold_f, ctr_al)
                    ws5.row_dimensions[ri5].height = 18
            else:
                c = ws5.cell(row=3, column=1,
                             value="✓  All reference entries exist in the DB.")
                c.font = Font(name="Segoe UI", bold=True, size=10, color="1F6B28")
                c.alignment = wrap_al
                ws5.merge_cells("A3:C3")
                ws5.row_dimensions[3].height = 22
            ws5.freeze_panes = "A3"

            wb_out.save(out_path)

            msg  = "Validation complete.\n\n"
            msg += f"Total DB records           : {total_db:,}\n"
            msg += f"✓  Matched                 : {n_matched:,}\n"
            msg += f"✗  Unmatched               : {n_unmatched:,}\n"
            msg += f"⚠  No Client ID in DB      : {n_no_cid:,}\n"
            msg += f"—  In Ref, not in DB       : {n_notindb:,}\n"
            msg += f"⚠  Missing col info        : {n_miss_rows:,} record(s)\n"
            msg += f"Match Rate                 : {rate}\n\n"
            msg += f"Report saved to:\n{out_path}"

            self.after(0, lambda: (
                _flash_btn(self, self._sum_validate_btn, "✓  Done!", 2500),
                messagebox.showinfo("Validation Result", msg)
            ))
        except ImportError:
            self.after(0, lambda: (
                _flash_btn(self, self._sum_validate_btn, "openpyxl missing", 3000),
                messagebox.showerror("Validate Error",
                    "openpyxl is not installed.\nRun: pip install openpyxl")
            ))
        except Exception as exc:
            err = str(exc)
            self.after(0, lambda: (
                _flash_btn(self, self._sum_validate_btn, "✗  Error", 3000),
                messagebox.showerror("Validate Error", err)
            ))

    threading.Thread(target=_worker, daemon=True).start()


# ═══════════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════════

def _flash_btn(self, btn, msg: str, ms: int):
    try:
        orig = btn.cget("text")
        btn.configure(text=msg)
        self.after(ms, lambda: btn.configure(text=orig)
                   if btn.winfo_exists() else None)
    except Exception:
        pass


def lookup_summary_notify(self):
    # Do not refresh while a cell edit is in flight — a full re-render
    # would wipe _sum_row_data and destroy the in-progress treeview item,
    # making the row un-editable until the next manual refresh.
    if getattr(self, "_sum_edit_locked", False):
        return
    if getattr(self, "_current_tab", "") == "lookup_summary":
        _refresh_summary(self)
    elif hasattr(self, "_sum_stat_labels"):
        _update_stats(self)

# ═══════════════════════════════════════════════════════════════════════
#  CUSTOM DOC — DB HELPERS
# ═══════════════════════════════════════════════════════════════════════

def _db_init_custom_docs():
    """Tables already exist in DB — just verify connectivity."""
    try:
        with _db_connect() as conn:
            cur = conn.cursor()
            cur.execute("SELECT 1 FROM custom_doc_configs LIMIT 1")
            cur.execute("SELECT 1 FROM custom_extractions LIMIT 1")
            cur.close()
    except Exception as e:
        print(f"[_db_init_custom_docs] Warning: {e}")


def _db_save_custom_doc_config(title: str, file_type: str,
                                labels: list, col_titles: list) -> int:
    with _db_connect() as conn:
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO custom_doc_configs (title, file_type, labels, col_titles) "
            "VALUES (%s, %s, %s, %s) RETURNING id",
            (title, file_type,
             json.dumps(labels), json.dumps(col_titles)))
        new_id = cur.fetchone()[0]
        conn.commit()
        cur.close()
    return new_id


def _db_load_custom_doc_configs() -> list:
    with _db_connect() as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT id, title, file_type, labels, col_titles "
            "FROM custom_doc_configs ORDER BY created_at DESC")
        rows = cur.fetchall()
        cur.close()
    return [
        {"id": r[0], "title": r[1], "file_type": r[2],
         "labels": r[3], "col_titles": r[4]}
        for r in rows
    ]


def _db_delete_custom_doc_config(config_id: int):
    with _db_connect() as conn:
        cur = conn.cursor()
        cur.execute(
            "DELETE FROM custom_doc_configs WHERE id=%s", (config_id,))
        conn.commit()
        cur.close()


def _db_upsert_custom_extraction(client_id: str, config_id: int,
                                  config_title: str, data: dict):
    with _db_connect() as conn:
        cur = conn.cursor()

        # ── Upsert into custom_extractions (existing logic) ────────────
        cur.execute(
            "SELECT id FROM custom_extractions "
            "WHERE client_id=%s AND config_id=%s",
            (client_id, config_id))
        existing = cur.fetchone()
        if existing:
            cur.execute(
                "UPDATE custom_extractions "
                "SET data=%s, extracted_at=NOW() WHERE id=%s",
                (json.dumps(data), existing[0]))
        else:
            cur.execute(
                "INSERT INTO custom_extractions "
                "(client_id, config_id, config_title, data) "
                "VALUES (%s, %s, %s, %s)",
                (client_id, config_id, config_title,
                 json.dumps(data)))

        # ── NEW: Ensure a column exists in applicants for each key ─────
        # Column name format: custom_<config_id>_<snake_cased_key>
        for col_title, value in data.items():
            safe_col = "custom_" + re.sub(r"[^a-z0-9]", "_",
                       f"{config_id}_{col_title}".lower()).strip("_")
            # Add column if it doesn't exist yet
            # Check if column exists first, then ALTER TABLE if needed
            cur.execute("""
                SELECT 1 FROM information_schema.columns
                WHERE table_name='applicants' AND column_name=%s
            """, (safe_col,))
            if not cur.fetchone():
                # safe_col is already sanitised by the regex above —
                # only a-z, 0-9 and _ can appear in it, so f-string is safe here
                cur.execute(f"ALTER TABLE applicants ADD COLUMN {safe_col} TEXT")

            # Write the value into the matching applicant row
            cur.execute(
                f"UPDATE applicants SET {safe_col}=%s "
                "WHERE TRIM(UPPER(client_id))=TRIM(UPPER(%s))",
                (value, client_id))

        conn.commit()
        cur.close()

def _db_get_custom_columns() -> list[tuple[str, str, int]]:
    """
    Returns list of (db_col, display_label, config_id) for all
    custom columns that exist in the applicants table.
    Format: custom_<config_id>_<key>
    """
    result = []
    with _db_connect() as conn:
        cur = conn.cursor()
        # Get all custom_* columns from applicants
        cur.execute("""
            SELECT column_name FROM information_schema.columns
            WHERE table_name='applicants'
            AND column_name LIKE 'custom_%'
            ORDER BY column_name
        """)
        col_rows = [r[0] for r in cur.fetchall()]

        # Get config titles for display labels
        cur.execute("SELECT id, title FROM custom_doc_configs")
        config_map = {r[0]: r[1] for r in cur.fetchall()}
        cur.close()

    for col_name in col_rows:
        # Parse: custom_<config_id>_<rest>
        parts = col_name.split("_", 2)  # ['custom', '<id>', '<key>']
        if len(parts) < 3:
            continue
        try:
            config_id = int(parts[1])
        except ValueError:
            continue
        key_part      = parts[2].replace("_", " ").title()
        config_title  = config_map.get(config_id, f"Config {config_id}")
        display_label = config_title
        result.append((col_name, display_label, config_id))

    return result


def _db_verify_client_id_exists(client_id: str) -> bool:
    with _db_connect() as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT 1 FROM applicants "
            "WHERE TRIM(UPPER(client_id))=TRIM(UPPER(%s)) LIMIT 1",
            (client_id,))
        result = cur.fetchone()
        cur.close()
    return result is not None


# ═══════════════════════════════════════════════════════════════════════
#  CUSTOM DOC — GEMINI EXTRACTION
# ═══════════════════════════════════════════════════════════════════════

def _gemini_extract_custom_labels(pdf_bytes: bytes,
                                   labels: list,
                                   file_type_hint: str,
                                   api_key: str) -> dict:
    try:
        from google import genai
        from google.genai import types as gt
    except ImportError:
        raise RuntimeError(
            "google-genai not installed. Run: pip install google-genai")

    client = genai.Client(api_key=api_key)
    label_lines = "\n".join(
        f'  {i+1}. "{lbl.strip()}"' for i, lbl in enumerate(labels))
    file_hint = (
        f"The document type is: {file_type_hint}.\n"
        if file_type_hint else "")
    json_template = "{\n" + ",\n".join(
        f'  "{lbl.strip()}": ""' for lbl in labels) + "\n}"

    prompt = f"""You are a data extraction assistant for a Philippine rural bank.
This is a scanned loan-application document.
{file_hint}
Extract the following labelled fields. Each label is a printed field
name on the form. Return the handwritten or typed value next to it.
If a label is absent or its value is blank, return "".

Fields to extract:
{label_lines}

Rules:
- Copy values exactly as written on the document.
- For multi-line values (e.g. addresses), join with a single space.
- For amounts, include the peso sign and commas exactly as written.
- NEVER invent or guess. Missing or blank = empty string "".
- Return ONLY valid JSON — no explanation, no markdown fences.

Required JSON format:
{json_template}"""

    resp = client.models.generate_content(
        model="gemini-2.5-flash",
        contents=[
            gt.Part.from_bytes(
                data=pdf_bytes, mime_type="application/pdf"),
            prompt,
        ],
        config=gt.GenerateContentConfig(temperature=0.0),
    )
    raw     = resp.text or ""
    cleaned = re.sub(r"```(?:json)?", "", raw).strip().strip("`").strip()
    m       = re.search(r"\{[\s\S]*\}", cleaned)
    if not m:
        return {}
    try:
        return json.loads(m.group(0))
    except Exception:
        return {}


# ═══════════════════════════════════════════════════════════════════════
#  CUSTOM DOC — ADD DOC DIALOG
# ═══════════════════════════════════════════════════════════════════════

# FILE TYPE CATEGORIES
_DOC_FILE_TYPES = [
    "PDF Document",
    "Scanned Image (JPG/PNG)",
    "Word Document (.docx)",
    "Text File (.txt)",
    "Other Document",
]

_DATA_FILE_TYPES = [
    "Excel Spreadsheet (.xlsx)",
    "CSV File (.csv)",
]

_ALL_FILE_TYPES = _DOC_FILE_TYPES + _DATA_FILE_TYPES


def _open_add_doc_dialog(self):
    _db_init_custom_docs()

    api_key = None
    for attr in ("_gemini_api_key", "gemini_api_key", "_api_key"):
        if hasattr(self, attr):
            api_key = getattr(self, attr)
            break
    if not api_key:
        try:
            from app_constants import GEMINI_API_KEY
            api_key = GEMINI_API_KEY
        except Exception:
            pass
    if not api_key:
        import os
        api_key = (os.environ.get("GEMINI_API_KEY") or
                   os.environ.get("GOOGLE_API_KEY"))
    if not api_key:
        messagebox.showerror(
            "Gemini Key Missing",
            "No Gemini API key found.\n"
            "Add GEMINI_API_KEY to app_constants.py or set the "
            "GEMINI_API_KEY environment variable.")
        return

    win = tk.Toplevel(self)
    win.overrideredirect(True)
    win.configure(bg=LIME_MID)  # lime border outline
    win.grab_set()

    p_x = self.winfo_rootx(); p_y = self.winfo_rooty()
    p_w = self.winfo_width();  p_h = self.winfo_height()
    w_w, w_h = 820, 720
    win.geometry(
        f"{w_w}x{w_h}+{p_x + (p_w - w_w) // 2}"
        f"+{p_y + (p_h - w_h) // 2}")
    win.minsize(700, 600)

    # ── Drag support ──────────────────────────────────────────────────
    _drag = {"x": 0, "y": 0}
    def _drag_start(e): _drag["x"] = e.x_root; _drag["y"] = e.y_root
    def _drag_move(e):
        dx = e.x_root - _drag["x"]; dy = e.y_root - _drag["y"]
        x  = win.winfo_x() + dx;    y  = win.winfo_y() + dy
        win.geometry(f"+{x}+{y}")
        _drag["x"] = e.x_root;      _drag["y"] = e.y_root

    # ── 2px lime border shell ─────────────────────────────────────────
    border_shell = tk.Frame(win, bg=LIME_MID, padx=2, pady=2)
    border_shell.pack(fill="both", expand=True)

    # ── Sidebar + main layout ─────────────────────────────────────────
    body_outer = tk.Frame(border_shell, bg=NAVY_DEEP)
    body_outer.pack(fill="both", expand=True)

    # Left sidebar (mirrors login screen sidebar)
    sidebar = tk.Frame(body_outer, bg=NAVY_DEEP, width=220)
    sidebar.pack(side="left", fill="y")
    sidebar.pack_propagate(False)

    # Sidebar logo block
    # ── Sidebar logo block (draggable) ────────────────────────────────
    logo_block = tk.Frame(sidebar, bg=NAVY_DEEP)
    logo_block.pack(fill="x", padx=20, pady=(28, 0))
    logo_block.bind("<ButtonPress-1>",   _drag_start)
    logo_block.bind("<B1-Motion>",       _drag_move)

    logo_circle = tk.Frame(logo_block, bg=LIME_MID, width=38, height=38)
    logo_circle.pack(side="left")
    logo_circle.pack_propagate(False)
    lc_lbl = tk.Label(logo_circle, text="📄", font=("Segoe UI Emoji", 14),
                      bg=LIME_MID, fg=NAVY_DEEP)
    lc_lbl.place(relx=0.5, rely=0.5, anchor="center")
    lc_lbl.bind("<ButtonPress-1>", _drag_start)
    lc_lbl.bind("<B1-Motion>",     _drag_move)

    title_block = tk.Frame(logo_block, bg=NAVY_DEEP)
    title_block.pack(side="left", padx=(10, 0))
    title_block.bind("<ButtonPress-1>", _drag_start)
    title_block.bind("<B1-Motion>",     _drag_move)
    for txt, fnt, clr in [
        ("ADD DOC",          ("Segoe UI", 11, "bold"), WHITE),
        ("Custom Extraction",("Segoe UI", 8),          "#8DA8C8"),
    ]:
        l = tk.Label(title_block, text=txt, font=fnt, fg=clr,
                     bg=NAVY_DEEP, anchor="w")
        l.pack(anchor="w")
        l.bind("<ButtonPress-1>", _drag_start)
        l.bind("<B1-Motion>",     _drag_move)

    tk.Frame(sidebar, bg="#1E2D45", height=1).pack(fill="x", padx=20, pady=(18, 0))

    # Sidebar description
    desc_frame = tk.Frame(sidebar, bg=NAVY_DEEP)
    desc_frame.pack(fill="x", padx=20, pady=(16, 0))
    desc_frame.bind("<ButtonPress-1>", _drag_start)
    desc_frame.bind("<B1-Motion>",     _drag_move)
    dl = tk.Label(desc_frame,
                  text="Configure labels,\nthen run against\nany file and link\nto a Client ID.",
                  font=("Segoe UI", 8), fg="#6A84A0",
                  bg=NAVY_DEEP, justify="left", anchor="w")
    dl.pack(anchor="w")
    dl.bind("<ButtonPress-1>", _drag_start)
    dl.bind("<B1-Motion>",     _drag_move)

    tk.Frame(sidebar, bg="#1E2D45", height=1).pack(fill="x", padx=20, pady=(20, 0))

    # ── Clickable sidebar nav items ───────────────────────────────────
    _sidebar_nav_frames: dict[str, tk.Frame] = {}
    _sidebar_nav_labels: dict[str, tk.Label] = {}

    def _update_sidebar_nav(active_name: str):
        for n, f in _sidebar_nav_frames.items():
            is_act = (n == active_name)
            f.config(bg="#1E2D45" if is_act else NAVY_DEEP)
            _sidebar_nav_labels[n].config(
                bg="#1E2D45" if is_act else NAVY_DEEP,
                fg=LIME_MID  if is_act else "#6A84A0",
                font=("Segoe UI", 9, "bold" if is_act else "normal"))

    for nav_label, nav_icon in [
        ("Add Extraction",  "＋"),
        ("Added Extracted", "≡"),
    ]:
        nav_f = tk.Frame(sidebar, bg=NAVY_DEEP, cursor="hand2")
        nav_f.pack(fill="x", padx=10, pady=1)
        nav_lbl = tk.Label(nav_f,
                           text=f"  {nav_icon}  {nav_label}",
                           font=("Segoe UI", 9), fg="#6A84A0",
                           bg=NAVY_DEEP, padx=10, pady=8,
                           anchor="w", cursor="hand2")
        nav_lbl.pack(fill="x")
        _sidebar_nav_frames[nav_label] = nav_f
        _sidebar_nav_labels[nav_label] = nav_lbl
        nav_f.bind("<Button-1>",   lambda e, n=nav_label: _switch_tab(n))
        nav_lbl.bind("<Button-1>", lambda e, n=nav_label: _switch_tab(n))

    # Spacer at bottom of sidebar + version tag
    tk.Frame(sidebar, bg=NAVY_DEEP).pack(fill="both", expand=True)
    tk.Frame(sidebar, bg="#1E2D45", height=1).pack(fill="x", padx=20, pady=(0, 8))
    tk.Label(sidebar, text="DocExtract Pro  ·  v2.0",
             font=("Segoe UI", 7), fg="#3A5070",
             bg=NAVY_DEEP, pady=8).pack()

    # Vertical divider
    tk.Frame(body_outer, bg="#1E2D45", width=1).pack(side="left", fill="y")

    # Right content panel
    right_panel = tk.Frame(body_outer, bg=WHITE)
    right_panel.pack(side="left", fill="both", expand=True)

    # Content header
    hdr = tk.Frame(right_panel, bg=WHITE)
    hdr.pack(fill="x", padx=32, pady=(28, 0))
    _right_title_lbl = tk.Label(hdr, text="Add Doc",
             font=("Segoe UI", 20, "bold"), fg=NAVY_DEEP,
             bg=WHITE, anchor="w")
    _right_title_lbl.pack(anchor="w")
    _right_sub_lbl = tk.Label(hdr,
             text="Configure extraction labels and run against any document file.",
             font=("Segoe UI", 9), fg="#6A84A0",
             bg=WHITE, anchor="w")
    _right_sub_lbl.pack(anchor="w", pady=(2, 0))

    # Close button top-right of right panel
    close_btn = tk.Label(right_panel, text="✕",
                         font=("Segoe UI", 13), fg="#9AAFC5",
                         bg=WHITE, cursor="hand2", padx=14, pady=10)
    close_btn.place(relx=1.0, rely=0.0, anchor="ne")
    close_btn.bind("<Button-1>",  lambda e: win.destroy())
    close_btn.bind("<Enter>",     lambda e: close_btn.config(fg=ACCENT_RED))
    close_btn.bind("<Leave>",     lambda e: close_btn.config(fg="#9AAFC5"))

    # Make header draggable
    for w in (hdr, _right_title_lbl, _right_sub_lbl):
        w.bind("<ButtonPress-1>", _drag_start)
        w.bind("<B1-Motion>",     _drag_move)
    tk.Frame(right_panel, bg="#E8EEF8", height=1).pack(fill="x", padx=32, pady=(14, 0))

    content_area = tk.Frame(right_panel, bg=WHITE)
    content_area.pack(fill="both", expand=True)

    tab_frames: dict[str, tk.Frame] = {}

    def _switch_tab(name: str):
        for f in tab_frames.values():
            f.pack_forget()
        tab_frames[name].pack(fill="both", expand=True)
        _update_sidebar_nav(name)
        # Update right-panel header title
        _right_title_lbl.config(
            text="Add Doc" if name == "Add Extraction" else "Added Extracted")
        _right_sub_lbl.config(
            text="Configure extraction labels and run against any document file."
                 if name == "Add Extraction"
                 else "Run a saved config against a file and link it to a Client ID.")

    for tab_name in ("Add Extraction", "Added Extracted"):
        f = tk.Frame(content_area, bg=WHITE)
        tab_frames[tab_name] = f

    # ── TAB 1: Add Extraction ─────────────────────────────────────────
    # ── TAB 1: Add Extraction ─────────────────────────────────────────
    t1 = tab_frames["Add Extraction"]
    t1_body = tk.Frame(t1, bg=WHITE)
    t1_body.pack(fill="both", expand=True, padx=32, pady=16)

    def _lbl(parent, text):
        tk.Label(parent, text=text,
                 font=("Segoe UI", 8, "bold"), fg="#374151",
                 bg=WHITE, anchor="w").pack(fill="x", pady=(0, 4))

    def _hint(parent, text):
        tk.Label(parent, text=text,
                 font=("Segoe UI", 7), fg="#9AAFC5",
                 bg=WHITE, anchor="w").pack(fill="x", pady=(0, 10))

    # ── Document Title ────────────────────────────────────────────────
    _lbl(t1_body, "Document Title  (shown as the button label in 'Added Extracted')")
    title_var = tk.StringVar()
    tk.Entry(t1_body, textvariable=title_var,
             font=("Segoe UI", 10), fg=TXT_NAVY, bg=WHITE,
             relief="solid", bd=1,
             insertbackground=NAVY_MID).pack(fill="x", ipady=4, pady=(0, 10))

    # ── File Type Dropdown ────────────────────────────────────────────
    _lbl(t1_body, "File Type  (select the type of document to process)")

    filetype_var = tk.StringVar(value=_ALL_FILE_TYPES[0])

    filetype_combo = ttk.Combobox(
        t1_body,
        textvariable=filetype_var,
        values=_ALL_FILE_TYPES,
        state="readonly",
        font=("Segoe UI", 9),
        width=60,
    )
    filetype_combo.pack(fill="x", pady=(0, 4))

    # ── Mode indicator label ──────────────────────────────────────────
    mode_lbl = tk.Label(
        t1_body, text="",
        font=("Segoe UI", 8, "bold"), fg=NAVY_MID,
        bg=CARD_WHITE, anchor="w")
    mode_lbl.pack(fill="x", pady=(0, 8))

    # ── Labels section ────────────────────────────────────────────────
    labels_frame = tk.Frame(t1_body, bg=WHITE)
    labels_frame.pack(fill="x")
    _lbl(labels_frame, "Labels to Fetch  (comma-separated — printed field names on the form)")
    labels_txt = tk.Text(
        labels_frame, height=3,
        font=("Segoe UI", 9), fg=TXT_NAVY, bg=WHITE,
        relief="solid", bd=1,
        insertbackground=NAVY_MID, wrap="word")
    labels_txt.pack(fill="x", pady=(0, 2))
    _hint(labels_frame, "e.g.  Credit Score, Risk Rating, Loan Purpose, Remarks")

    # ── Column Titles section ─────────────────────────────────────────
    coltitles_frame = tk.Frame(t1_body, bg=WHITE)
    coltitles_frame.pack(fill="x")
    _lbl(coltitles_frame, "Column Titles  (comma-separated, same order as Labels — stored as keys in DB)")
    coltitles_txt = tk.Text(
        coltitles_frame, height=3,
        font=("Segoe UI", 9), fg=TXT_NAVY, bg=WHITE,
        relief="solid", bd=1,
        insertbackground=NAVY_MID, wrap="word")
    coltitles_txt.pack(fill="x", pady=(0, 2))
    _hint(coltitles_frame, "Must be the same count as Labels.  e.g.  Credit Score, Risk Rating, Loan Purpose, Remarks")

    # ── Dynamic enable/disable based on file type ─────────────────────
    def _is_data_type(ft: str) -> bool:
        return ft in _DATA_FILE_TYPES

    def _set_widget_state(widget, enabled: bool):
        """Enable or disable a Text widget visually."""
        if enabled:
            widget.config(
                state="normal",
                bg=WHITE,
                fg=TXT_NAVY,
                relief="solid")
        else:
            widget.config(
                state="disabled",
                bg="#F0F0F0",
                fg="#AAAAAA",
                relief="flat")

    def _on_filetype_change(*_):
        ft = filetype_var.get()
        is_data = _is_data_type(ft)

        if is_data:
            # CSV/Excel: Column Titles required, Labels disabled
            _set_widget_state(labels_txt,     enabled=False)
            _set_widget_state(coltitles_txt,  enabled=True)
            mode_lbl.config(
                text="📊  Data file mode — Column Titles are required. Labels field is not used.",
                fg="#1A4A2A")
        else:
            # Document/PDF: Labels required, Column Titles auto-derived (disabled)
            _set_widget_state(labels_txt,    enabled=True)
            _set_widget_state(coltitles_txt, enabled=False)
            mode_lbl.config(
                text="📄  Document mode — Labels are required. Column Titles will mirror the Labels.",
                fg="#1A2E5C")

    filetype_combo.bind("<<ComboboxSelected>>", _on_filetype_change)
    _on_filetype_change()  # apply initial state

    # ── Error label ───────────────────────────────────────────────────
    err_lbl_t1 = tk.Label(t1_body, text="",
                           font=("Segoe UI", 8), fg=ACCENT_RED,
                           bg=WHITE, anchor="w")
    err_lbl_t1.pack(fill="x")

    # ── Save Config ───────────────────────────────────────────────────
    def _save_config():
        title     = title_var.get().strip()
        file_type = filetype_var.get().strip()
        ft        = file_type
        is_data   = _is_data_type(ft)

        if not title:
            err_lbl_t1.config(text="Document title is required.", fg=ACCENT_RED)
            return

        if is_data:
            # CSV/Excel mode — only column titles matter
            cols_raw   = coltitles_txt.get("1.0", "end-1c").strip()
            col_titles = [x.strip() for x in cols_raw.split(",") if x.strip()]
            if not col_titles:
                err_lbl_t1.config(
                    text="Column Titles are required for data file types.",
                    fg=ACCENT_RED)
                return
            # Labels mirror column titles for storage consistency
            labels = col_titles[:]
        else:
            # Document mode — labels required; column titles = labels
            labels_raw = labels_txt.get("1.0", "end-1c").strip()
            labels     = [x.strip() for x in labels_raw.split(",") if x.strip()]
            if not labels:
                err_lbl_t1.config(
                    text="Labels are required for document/PDF types.",
                    fg=ACCENT_RED)
                return
            # Column titles automatically mirror labels
            col_titles = labels[:]

        err_lbl_t1.config(text="")
        try:
            _db_save_custom_doc_config(title, file_type, labels, col_titles)
            err_lbl_t1.config(
                text=f'✓  "{title}" saved — switch to "Added Extracted" to run it.',
                fg=ACCENT_SUCCESS)
            _reload_run_tab()
        except Exception as exc:
            err_lbl_t1.config(text=f"Save failed: {exc}", fg=ACCENT_RED)

    tk.Frame(t1_body, bg=WHITE, height=10).pack()
    ctk.CTkButton(t1_body, text="Save Config →",
                  command=_save_config,
                  width=180, height=38, corner_radius=8,
                  fg_color=NAVY_DEEP, hover_color=NAVY_LIGHT,
                  text_color=WHITE,
                  font=FF(10, "bold")).pack(anchor="w")

    # ── TAB 2: Added Extracted ────────────────────────────────────────
    t2 = tab_frames["Added Extracted"]

    t2_info = tk.Frame(t2, bg="#F5F7FA",
                       highlightbackground=BORDER_MID,
                       highlightthickness=1)
    t2_info.pack(fill="x", padx=16, pady=(12, 4))
    tk.Label(t2_info,
             text="Click a config button to upload a file and extract data.\n"
                  "You will be asked for the applicant's Client ID first.",
             font=("Segoe UI", 8), fg=TXT_MUTED, bg="#F0F4FA",
             padx=10, pady=8, justify="left").pack(anchor="w")

    status_var = tk.StringVar(value="")
    status_lbl = tk.Label(t2, textvariable=status_var,
                           font=("Segoe UI", 9, "bold"),
                           fg=ACCENT_SUCCESS, bg=WHITE, pady=3)
    status_lbl.pack(fill="x", padx=16)

    t2_scroll_outer = tk.Frame(t2, bg=WHITE)
    t2_scroll_outer.pack(fill="both", expand=True, padx=16, pady=4)

    t2_canvas  = tk.Canvas(t2_scroll_outer, bg=WHITE,
                            highlightthickness=0)
    t2_vscroll = tk.Scrollbar(t2_scroll_outer, orient="vertical",
                               command=t2_canvas.yview)
    t2_canvas.configure(yscrollcommand=t2_vscroll.set)
    t2_vscroll.pack(side="right", fill="y")
    t2_canvas.pack(side="left", fill="both", expand=True)

    t2_inner = tk.Frame(t2_canvas, bg=WHITE)
    t2_cwin  = t2_canvas.create_window((0, 0), window=t2_inner, anchor="nw")

    def _on_t2_cfg(e=None):
        t2_canvas.configure(scrollregion=t2_canvas.bbox("all"))
        t2_canvas.itemconfig(t2_cwin, width=t2_canvas.winfo_width())
    t2_inner.bind("<Configure>", _on_t2_cfg)
    t2_canvas.bind("<Configure>", _on_t2_cfg)

    def _ask_client_id(config_title: str):
        result = [None]
        dlg = tk.Toplevel(win)
        dlg.configure(bg=CARD_WHITE)
        dlg.resizable(False, False)
        dlg.overrideredirect(True)
        dlg.transient(win)
        dlg.grab_set()
        dlg.lift(win)
        dlg.focus_force()

        d_w, d_h = 440, 200
        dlg.geometry(
            f"{d_w}x{d_h}"
            f"+{win.winfo_rootx() + (w_w - d_w) // 2}"
            f"+{win.winfo_rooty() + (w_h - d_h) // 2}")

        shell = tk.Frame(dlg, bg=BORDER_MID)
        shell.pack(fill="both", expand=True, padx=1, pady=1)
        root = tk.Frame(shell, bg=CARD_WHITE)
        root.pack(fill="both", expand=True)

        hdr2 = tk.Frame(root, bg="#E8EEF8")
        hdr2.pack(fill="x")
        tk.Label(hdr2, text=f"Link to Client ID — {config_title}",
                 font=("Segoe UI", 10, "bold"), fg=NAVY_MID,
                 bg="#E8EEF8", padx=14, pady=8, anchor="w").pack(fill="x")

        body2 = tk.Frame(root, bg=CARD_WHITE)
        body2.pack(fill="both", expand=True, padx=16, pady=(10, 0))
        tk.Label(body2,
                 text="Enter the applicant's Client ID.\n"
                      "The extraction will be saved and linked to this ID.",
                 font=("Segoe UI", 9), fg=TXT_MUTED,
                 bg=CARD_WHITE, justify="left").pack(anchor="w")

        # Load all client IDs from DB for the dropdown
        def _load_client_ids():
            try:
                with _db_connect() as conn:
                    cur = conn.cursor()
                    cur.execute(
                        "SELECT client_id, applicant_name FROM applicants "
                        "WHERE client_id IS NOT NULL AND TRIM(client_id) != '' "
                        "ORDER BY applicant_name ASC")
                    rows = cur.fetchall()
                    cur.close()
                return [f"{r[0]}  —  {r[1]}" for r in rows if r[0]]
            except Exception:
                return []

        cid_choices = _load_client_ids()
        cid_var = tk.StringVar()
        cid_entry = ttk.Combobox(
            body2, textvariable=cid_var,
            values=cid_choices,
            font=("Segoe UI", 10),
            width=40)
        cid_entry.pack(fill="x", ipady=4, pady=(8, 0))
        cid_entry.focus_set()

        err2 = tk.Label(body2, text="",
                        font=("Segoe UI", 8), fg=ACCENT_RED, bg=CARD_WHITE)
        err2.pack(anchor="w")

        def _confirm_cid():
            raw = cid_var.get().strip()
            # Strip the " — Name" suffix if selected from dropdown
            cid = raw.split("  —  ")[0].strip() if "  —  " in raw else raw
            if not cid:
                err2.config(text="Client ID is required."); return
            if not _db_verify_client_id_exists(cid):
                err2.config(
                    text=f'Client ID "{cid}" not found. '
                         f'Check spelling or import this applicant first.')
                return
            result[0] = cid
            try: dlg.grab_release()
            except Exception: pass
            dlg.destroy()

        def _cancel_cid():
            result[0] = None
            try: dlg.grab_release()
            except Exception: pass
            dlg.destroy()

        btn_row2 = tk.Frame(body2, bg=CARD_WHITE)
        btn_row2.pack(fill="x", pady=(8, 0))
        ctk.CTkButton(btn_row2, text="Cancel", command=_cancel_cid,
                      width=90, height=28, corner_radius=6,
                      fg_color="#E8ECF2", hover_color="#DDE2EA",
                      text_color=TXT_NAVY, font=FF(8, "bold")).pack(side="right", padx=(8, 0))
        ctk.CTkButton(btn_row2, text="✔  Confirm", command=_confirm_cid,
                      width=100, height=28, corner_radius=6,
                      fg_color=LIME_MID, hover_color=LIME_BRIGHT,
                      text_color=TXT_ON_LIME, font=FF(8, "bold")).pack(side="right")

        cid_entry.bind("<Return>", lambda e: _confirm_cid())
        dlg.bind("<Escape>", lambda e: _cancel_cid())
        try:
            dlg.wait_window(dlg)
        finally:
            try: dlg.grab_release()
            except Exception: pass
        return result[0]

    def _run_extraction(config: dict):
        ft      = config.get("file_type", "")
        is_data = _is_data_type(ft)

        # Choose appropriate file dialog filter
        if is_data:
            filetypes = [("Data files", "*.xlsx *.csv"),
                        ("Excel files", "*.xlsx"),
                        ("CSV files", "*.csv"),
                        ("All files", "*.*")]
        else:
            filetypes = [("Document files", "*.pdf *.docx *.txt *.jpg *.jpeg *.png"),
                        ("PDF files", "*.pdf"),
                        ("All files", "*.*")]

        path = filedialog.askopenfilename(
            title=f"Upload file for: {config['title']}",
            filetypes=filetypes)
        if not path:
            return

        labels     = config["labels"]
        col_titles = config["col_titles"]

        def _ui_status(msg, color=ACCENT_GOLD):
            win.after(0, lambda: (
                status_var.set(msg),
                status_lbl.config(fg=color)))

        def _worker():
            try:
                if is_data:
                    # ── CSV / Excel extraction ─────────────────────────────
                    _ui_status(f"⟳  Reading data file '{Path(path).name}'…")
                    import openpyxl as _oxl
                    import csv as _csv_mod

                    if path.lower().endswith(".csv"):
                        with open(path, newline="", encoding="utf-8-sig") as f:
                            reader  = _csv_mod.DictReader(f)
                            records = [dict(r) for r in reader]
                        file_cols = list(records[0].keys()) if records else []
                    else:
                        wb  = _oxl.load_workbook(path, read_only=True, data_only=True)
                        ws  = wb.active
                        hdr = next(ws.iter_rows(min_row=1, max_row=1), None)
                        if hdr is None:
                            raise ValueError("File appears to be empty.")
                        file_cols = [str(c.value).strip() if c.value else "" for c in hdr]
                        records = []
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if all(v is None for v in row):
                                continue
                            records.append({
                                file_cols[i]: (str(v).strip() if v is not None else "")
                                for i, v in enumerate(row) if i < len(file_cols)
                            })
                        wb.close()

                    if not records:
                        raise ValueError("No data rows found in the file.")

                    # ── Detect client_id column in file ────────────────────
                    def _find_col(cols, *keywords):
                        for kw in keywords:
                            kw_norm = re.sub(r"[\s_\-]", "", kw.lower())
                            for c in cols:
                                if re.sub(r"[\s_\-]", "", c.lower()) == kw_norm:
                                    return c
                        for kw in keywords:
                            kw_norm = re.sub(r"[\s_\-]", "", kw.lower())
                            for c in cols:
                                if kw_norm in re.sub(r"[\s_\-]", "", c.lower()):
                                    return c
                        return None

                    col_cid = _find_col(file_cols,
                                        "clientid", "client id", "client_id", "cid")
                    if not col_cid:
                        raise ValueError(
                            f"No Client ID column found in file.\n\n"
                            f"File has: {', '.join(file_cols)}\n\n"
                            f"Add a 'Client ID' column to the file so each row "
                            f"can be matched to the correct applicant.")

                    # ── Process each row by its client_id ─────────────────
                    saved_count   = 0
                    skipped_count = 0
                    not_found     = []

                    for file_row in records:
                        client_id = str(file_row.get(col_cid, "") or "").strip()
                        if not client_id:
                            skipped_count += 1
                            continue

                        if not _db_verify_client_id_exists(client_id):
                            not_found.append(client_id)
                            continue

                        # Map col_titles → file columns for this row
                        data_to_store: dict[str, str] = {}
                        for ct in col_titles:
                            ct_norm = re.sub(r"[\s_\-]", "", ct.lower())
                            matched_col = next(
                                (fc for fc in file_cols
                                if re.sub(r"[\s_\-]", "", fc.lower()) == ct_norm),
                                None)
                            if matched_col is None:
                                matched_col = next(
                                    (fc for fc in file_cols
                                    if ct_norm in re.sub(r"[\s_\-]", "", fc.lower())),
                                    None)
                            val = str(file_row.get(matched_col, "") or "").strip() \
                                if matched_col else ""
                            data_to_store[ct] = val

                        if any(v for v in data_to_store.values()):
                            _db_upsert_custom_extraction(
                                client_id, config["id"],
                                config["title"], data_to_store)
                            saved_count += 1
                            _log_action(
                                self, "custom_extract",
                                f"[{config['title']}] "
                                f"{len(data_to_store)} field(s) → "
                                f"client_id={client_id}")

                    _refresh_summary(self)

                    msg = f"✓  Saved {saved_count} record(s)"
                    if skipped_count:
                        msg += f"  ·  {skipped_count} skipped (no client ID)"
                    if not_found:
                        msg += f"  ·  {len(not_found)} client ID(s) not in DB"
                    _ui_status(msg, ACCENT_SUCCESS)

                else:
                    # ── PDF / Document — still ask for client ID manually ──
                    # (PDFs don't have structured rows with client IDs)
                    client_id = _ask_client_id(config["title"])
                    if not client_id:
                        return

                    pdf_bytes = Path(path).read_bytes()
                    _ui_status(
                        f"⟳  Extracting {len(labels)} field(s) "
                        f"from '{Path(path).name}'…")

                    extracted = _gemini_extract_custom_labels(
                        pdf_bytes, labels, ft, api_key)

                    if not extracted:
                        _ui_status(
                            "✗  Gemini returned no data — check the document.",
                            ACCENT_RED)
                        return

                    data_to_store = {
                        col_title: str(extracted.get(label, "") or "").strip()
                        for label, col_title in zip(labels, col_titles)
                    }

                    if not data_to_store:
                        _ui_status("✗  No data could be extracted.", ACCENT_RED)
                        return

                    _db_upsert_custom_extraction(
                        client_id, config["id"],
                        config["title"], data_to_store)

                    _refresh_summary(self)
                    _ui_status(
                        f"✓  Saved {len(data_to_store)} field(s) "
                        f"for Client ID: {client_id}",
                        ACCENT_SUCCESS)
                    _log_action(
                        self, "custom_extract",
                        f"[{config['title']}] "
                        f"{len(data_to_store)} field(s) → "
                        f"client_id={client_id}")

            except Exception as exc:
                _ui_status(f"✗  Error: {exc}", ACCENT_RED)

        threading.Thread(target=_worker, daemon=True).start()

    def _reload_run_tab():
        for w in t2_inner.winfo_children():
            w.destroy()

        configs = _db_load_custom_doc_configs()

        if not configs:
            tk.Label(t2_inner,
                     text='No extraction configs yet.\n'
                          'Go to "Add Extraction" to create one.',
                     font=("Segoe UI", 9), fg=TXT_MUTED,
                     bg=CARD_WHITE, pady=20).pack()
            return

        for cfg in configs:
            ft       = cfg.get("file_type", "")
            is_data  = _is_data_type(ft)
            type_tag = "📊" if is_data else "📄"

            row_f = tk.Frame(t2_inner, bg=ROW_BG_EVEN,
                             highlightbackground=BORDER_MID,
                             highlightthickness=1)
            row_f.pack(fill="x", pady=3)

            info_f = tk.Frame(row_f, bg=ROW_BG_EVEN)
            info_f.pack(side="left", fill="x", expand=True, padx=10, pady=8)

            tk.Label(info_f, text=f"{type_tag}  {cfg['title']}",
                     font=("Segoe UI", 10, "bold"),
                     fg=NAVY_DEEP, bg=ROW_BG_EVEN).pack(anchor="w")

            if ft:
                tk.Label(info_f, text=f"Type: {ft}",
                         font=("Segoe UI", 8), fg=TXT_MUTED,
                         bg=ROW_BG_EVEN).pack(anchor="w")

            if is_data:
                preview = ", ".join(cfg["col_titles"][:5])
                extra   = len(cfg["col_titles"]) - 5
                lbl_txt = f"Columns: {preview}" + (f" … +{extra} more" if extra > 0 else "")
            else:
                preview = ", ".join(cfg["labels"][:5])
                extra   = len(cfg["labels"]) - 5
                lbl_txt = f"Labels: {preview}" + (f" … +{extra} more" if extra > 0 else "")

            tk.Label(info_f, text=lbl_txt,
                     font=("Segoe UI", 7), fg=TXT_SOFT,
                     bg=ROW_BG_EVEN, wraplength=420,
                     justify="left").pack(anchor="w")

            btn_f = tk.Frame(row_f, bg=ROW_BG_EVEN)
            btn_f.pack(side="right", padx=8, pady=8)

            ctk.CTkButton(
                btn_f,
                text=f"📂  Run: {cfg['title']}",
                command=lambda c=cfg: _run_extraction(c),
                width=180, height=32, corner_radius=6,
                fg_color=NAVY_LIGHT, hover_color=NAVY_PALE,
                text_color=WHITE,
                font=FF(9, "bold"),
            ).pack(side="left", padx=(0, 6))

            ctk.CTkButton(
                btn_f, text="🗑",
                command=lambda cid=cfg["id"], ct=cfg["title"]: (
                    _db_delete_custom_doc_config(cid)
                    or _reload_run_tab()
                ) if messagebox.askyesno(
                    "Delete Config",
                    f'Delete "{ct}"?\n\n'
                    f'All saved extractions for this config '
                    f'will also be deleted.') else None,
                width=36, height=32, corner_radius=6,
                fg_color="#3D1010", hover_color="#5C1A1A",
                text_color="#FF8A80",
                font=FF(9, "bold"),
            ).pack(side="left")

    _reload_run_tab()

    btn_bar = tk.Frame(right_panel, bg=WHITE,
                       highlightbackground="#E8EEF8",
                       highlightthickness=1)
    btn_bar.pack(fill="x", padx=32, pady=(4, 16))
    ctk.CTkButton(btn_bar, text="Close",
                  command=win.destroy,
                  width=100, height=32, corner_radius=6,
                  fg_color="#F0F4FA", hover_color="#E0E8F0",
                  text_color=TXT_NAVY,
                  font=FF(9, "bold")).pack(side="right", pady=8)

    win.protocol("WM_DELETE_WINDOW", win.destroy)
    _update_sidebar_nav("Add Extraction")
    _switch_tab("Add Extraction")


# ═══════════════════════════════════════════════════════════════════════
#  ATTACH
# ═══════════════════════════════════════════════════════════════════════

def attach(cls):
    cls._build_lookup_summary_panel = _build_lookup_summary_panel
    cls.lookup_summary_notify        = lookup_summary_notify