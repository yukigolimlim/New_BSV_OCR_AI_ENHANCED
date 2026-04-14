"""
lu_ui.py — LU Analysis: all tkinter / matplotlib / reportlab UI code
=====================================================================
Updated for Look-Up Summary sheet format.

Tabs:
  1. Analysis   — individual client search + per-client risk scorecard
  2. Charts     — sector breakdown charts (5 canonical sectors)
  3. Risk Simulator — expense inflation simulator
  4. Sector vs Loan Balance — how much of total loan balance each sector represents
  5. Report     — print-ready text report + PDF/Excel export

PATCH: Search bar now filters by sector name across ALL tabs.
       A sector-filter pill appears when a sector is matched.
       Export functions respect the active filter.
"""

import re
import tkinter as tk
import customtkinter as ctk
from pathlib import Path
from tkinter import filedialog, messagebox
from datetime import datetime

from lu_core import (
    GENERAL_CLIENT, _RISK_ORDER, _SCORE_BANDS,
    SECTOR_EXPENSE_RISK, SECTOR_WHOLESALE, SECTOR_AGRICULTURE,
    SECTOR_TRANSPORT, SECTOR_REMITTANCE, SECTOR_CONSUMER, SECTOR_OTHER,
    _compute_risk_score, _fmt_value, run_lu_analysis,
)

# ── Optional heavy deps ────────────────────────────────────────────────
try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

try:
    import matplotlib
    matplotlib.use("TkAgg")
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    import matplotlib.ticker
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    _HAS_MPL = True
except ImportError:
    _HAS_MPL = False

try:
    from reportlab.lib.pagesizes import A4, landscape as rl_landscape
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, PageBreak
    )
    _HAS_RL = True
except ImportError:
    _HAS_RL = False


# ══════════════════════════════════════════════════════════════════════
#  COLOUR / FONT CONSTANTS
# ══════════════════════════════════════════════════════════════════════

_NAVY_DEEP    = "#0A1628"
_NAVY_MID     = "#1A3A6B"
_NAVY_LIGHT   = "#1E4080"
_NAVY_MIST    = "#EEF3FB"
_NAVY_GHOST   = "#D6E4F7"
_NAVY_PALE    = "#4A6FA5"
_WHITE        = "#FFFFFF"
_CARD_WHITE   = "#FAFBFD"
_OFF_WHITE    = "#F5F7FA"
_BORDER_LIGHT = "#E2E8F5"
_BORDER_MID   = "#C5D0E8"
_TXT_NAVY     = "#1A2B4A"
_TXT_SOFT     = "#6B7FA3"
_TXT_MUTED    = "#9AAACE"
_TXT_ON_LIME  = "#0A2010"
_LIME_BRIGHT  = "#B8FF57"
_LIME_MID     = "#8FD14F"
_LIME_DARK    = "#5A9E28"
_LIME_PALE    = "#D4F5A0"
_ACCENT_RED   = "#E53E3E"
_ACCENT_GOLD  = "#D4A017"
_ACCENT_SUCCESS = "#2E7D32"

_RISK_COLOR    = {"HIGH": _ACCENT_RED, "MODERATE": _ACCENT_GOLD, "LOW": _ACCENT_SUCCESS}
_RISK_BG       = {"HIGH": "#FFF5F5",   "MODERATE": "#FFFBF0",    "LOW": "#F0FBE8"}
_RISK_BADGE_BG = {"HIGH": "#FFE8E8",   "MODERATE": "#FFF3CD",    "LOW": "#DCEDC8"}

_CLIENT_HERO_BG = {
    "CRITICAL": "#2D0A0A", "HIGH": "#1E0A0A",
    "MODERATE": "#1A1400", "LOW":  "#0A1A0A", "N/A": "#0A1628",
}
_CLIENT_HERO_ACCENT = {
    "CRITICAL": "#FF4444", "HIGH": "#E53E3E",
    "MODERATE": "#D4A017", "LOW":  "#2E7D32", "N/A": "#4A6FA5",
}

_MPL_HIGH = "#E53E3E"
_MPL_MOD  = "#D4A017"
_MPL_LOW  = "#2E7D32"
_MPL_NAVY = "#1A3A6B"
_MPL_BG   = "#FAFBFD"

_SIM_BAR_BASE = "#4A6FA5"
_SIM_BAR_SIM  = "#E53E3E"

# Sector colours for charts / loan balance tab
_SECTOR_COLORS = {
    SECTOR_WHOLESALE:   "#1A3A6B",
    SECTOR_AGRICULTURE: "#2E7D32",
    SECTOR_TRANSPORT:   "#D4A017",
    SECTOR_REMITTANCE:  "#8B5CF6",
    SECTOR_CONSUMER:    "#E53E3E",
    SECTOR_OTHER:       "#9AAACE",
}

_SECTOR_ICON = {
    SECTOR_WHOLESALE:   "🏪",
    SECTOR_AGRICULTURE: "🐟",
    SECTOR_TRANSPORT:   "🚛",
    SECTOR_REMITTANCE:  "💸",
    SECTOR_CONSUMER:    "🏦",
    SECTOR_OTHER:       "🏠",
}

# The 5 charted sectors (for Charts tab)
_CHART_SECTORS = [
    SECTOR_WHOLESALE, SECTOR_AGRICULTURE,
    SECTOR_TRANSPORT, SECTOR_REMITTANCE, SECTOR_CONSUMER,
]

# All known sectors including OTHER
_ALL_SECTORS = _CHART_SECTORS + [SECTOR_OTHER]


def F(size, weight="normal"):
    return ("Segoe UI", size, weight)

def FF(size, weight="normal"):
    return ctk.CTkFont(family="Segoe UI", size=size, weight=weight)


# ══════════════════════════════════════════════════════════════════════
#  SMART MOUSEWHEEL BINDING
# ══════════════════════════════════════════════════════════════════════

def _bind_mousewheel(canvas: tk.Canvas):
    def _on_enter(e):
        canvas.bind_all("<MouseWheel>",
                        lambda ev: canvas.yview_scroll(int(-1*(ev.delta/120)), "units"))
    def _on_leave(e):
        canvas.unbind_all("<MouseWheel>")
    canvas.bind("<Enter>", _on_enter)
    canvas.bind("<Leave>", _on_leave)


# ══════════════════════════════════════════════════════════════════════
#  SCROLLABLE FRAME HELPER
# ══════════════════════════════════════════════════════════════════════

def _make_scrollable(parent, bg=None):
    """Return (outer_frame, inner_frame, canvas) with scroll bar."""
    bg = bg or _CARD_WHITE
    outer  = tk.Frame(parent, bg=bg)
    outer.pack(fill="both", expand=True)
    sb     = tk.Scrollbar(outer, relief="flat", troughcolor=_OFF_WHITE,
                          bg=_BORDER_LIGHT, width=8, bd=0)
    sb.pack(side="right", fill="y")
    canvas = tk.Canvas(outer, bg=bg, highlightthickness=0, yscrollcommand=sb.set)
    canvas.pack(side="left", fill="both", expand=True)
    sb.config(command=canvas.yview)
    inner  = tk.Frame(canvas, bg=bg)
    win    = canvas.create_window((0,0), window=inner, anchor="nw")
    inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.bind("<Configure>", lambda e: canvas.itemconfig(win, width=e.width))
    _bind_mousewheel(canvas)
    return outer, inner, canvas


# ══════════════════════════════════════════════════════════════════════
#  FILTER STATE HELPERS
# ══════════════════════════════════════════════════════════════════════

def _lu_get_active_sectors(self) -> list:
    """
    Return the list of sector names currently filtered, or None if no sector filter.
    When a sector filter is active, _lu_filtered_sectors is set to a non-empty list.
    """
    return getattr(self, "_lu_filtered_sectors", None)


def _lu_get_filtered_all_data(self) -> dict:
    """
    Return a view of all_data filtered to the active sector(s).
    If no sector filter is active, returns the original all_data unchanged.
    Used by Charts, Sector vs Loan Balance, Report, and Simulator tabs.
    """
    all_data = self._lu_all_data
    active_sectors = _lu_get_active_sectors(self)
    if not active_sectors:
        return all_data

    # Filter general list and sector_map
    filtered_general = [r for r in all_data.get("general", [])
                        if r.get("sector") in active_sectors]
    filtered_sector_map = {s: v for s, v in all_data.get("sector_map", {}).items()
                           if s in active_sectors}

    # Recompute totals for the filtered set
    filtered_totals = {
        "loan_balance": sum(r.get("loan_balance") or 0 for r in filtered_general),
        "total_net":    sum(r.get("net_income")   or 0 for r in filtered_general),
    }

    # Rebuild clients dict
    filtered_clients = {r["client"]: r for r in filtered_general}

    return {
        "general":    filtered_general,
        "sector_map": filtered_sector_map,
        "totals":     filtered_totals,
        "clients":    filtered_clients,
    }


def _lu_update_filter_pill(self):
    """
    Show or hide the sector-filter pill label next to the search bar.
    The pill shows which sector(s) are active and has an ✕ to clear.
    """
    # Remove existing pill if any
    existing = getattr(self, "_lu_filter_pill", None)
    if existing:
        try:
            existing.destroy()
        except Exception:
            pass
        self._lu_filter_pill = None

    active = _lu_get_active_sectors(self)
    if not active:
        return

    sector_text = " · ".join(active)
    pill_frame = tk.Frame(self._lu_client_bar, bg="#1E4080",
                          highlightbackground=_LIME_MID, highlightthickness=1)
    pill_frame.pack(side="left", padx=(6, 0), pady=10)

    icon = _SECTOR_ICON.get(active[0], "🏭") if len(active) == 1 else "🏭"
    tk.Label(pill_frame,
             text=f"  {icon}  SECTOR: {sector_text}  ",
             font=F(8, "bold"), fg=_LIME_MID, bg="#1E4080",
             padx=6, pady=3).pack(side="left")

    clear_btn = tk.Label(pill_frame, text=" ✕ ", font=F(8, "bold"),
                         fg=_ACCENT_RED, bg="#1E4080", cursor="hand2", padx=2)
    clear_btn.pack(side="left")
    clear_btn.bind("<Button-1>", lambda e: _lu_clear_sector_filter(self))

    self._lu_filter_pill = pill_frame


def _lu_clear_sector_filter(self):
    """Clear any active sector filter and return to general view."""
    self._lu_filtered_sectors = None
    self._lu_search_var.set("")
    _lu_update_filter_pill(self)
    _lu_on_client_change(self, GENERAL_CLIENT)


# ══════════════════════════════════════════════════════════════════════
#  MAIN PANEL BUILDER
# ══════════════════════════════════════════════════════════════════════

def _build_lu_analysis_panel(self, parent):
    self._lu_analysis_frame = tk.Frame(parent, bg=_CARD_WHITE)

    # ── Top header bar ─────────────────────────────────────────────────
    hdr_bar = tk.Frame(self._lu_analysis_frame, bg=_NAVY_DEEP, height=56)
    hdr_bar.pack(fill="x")
    hdr_bar.pack_propagate(False)

    hdr_inner = tk.Frame(hdr_bar, bg=_NAVY_DEEP)
    hdr_inner.pack(side="left", fill="y", padx=(28,0))
    tk.Label(hdr_inner, text="📈  LU Analysis",
             font=F(14,"bold"), fg=_WHITE, bg=_NAVY_DEEP).pack(side="left", anchor="center")
    tk.Label(hdr_inner, text=" — Sector & Expense Risk Scanner",
             font=F(9), fg=_TXT_MUTED, bg=_NAVY_DEEP).pack(side="left", anchor="center")

    self._lu_active_view = tk.StringVar(value="analysis")
    tab_frame = tk.Frame(hdr_bar, bg=_NAVY_DEEP)
    tab_frame.pack(side="left", padx=16, fill="y")
    for label, view in [("Analysis","analysis"), ("Charts","charts"),
                        ("Risk Simulator","simulator"),
                        ("Sector vs Loan Balance","loanbal"),
                        ("Report","report")]:
        tk.Button(tab_frame, text=label, font=F(8,"bold"),
                  bg=_NAVY_MID, fg=_WHITE,
                  activebackground=_LIME_MID, activeforeground=_TXT_ON_LIME,
                  relief="flat", padx=10, pady=0, cursor="hand2",
                  command=lambda v=view: _lu_switch_view(self, v)
                  ).pack(side="left", padx=2, pady=12, ipady=4)

    self._lu_export_btn = ctk.CTkButton(
        hdr_bar, text="💾  Export", command=lambda: _lu_show_export_menu(self),
        width=100, height=34, corner_radius=6,
        fg_color=_LIME_DARK, hover_color=_LIME_MID,
        text_color=_TXT_ON_LIME, font=FF(9,"bold"), state="disabled")
    self._lu_export_btn.pack(side="right", padx=(0,8), pady=11)

    self._lu_load_btn = ctk.CTkButton(
        hdr_bar, text="📂  Load Excel File", command=lambda: _lu_browse_file(self),
        width=160, height=34, corner_radius=6,
        fg_color=_LIME_MID, hover_color=_LIME_BRIGHT,
        text_color=_TXT_ON_LIME, font=FF(9,"bold"))
    self._lu_load_btn.pack(side="right", padx=(0,4), pady=11)

    self._lu_rescan_btn = ctk.CTkButton(
        hdr_bar, text="🔄  Re-Scan", command=lambda: _lu_run_analysis(self),
        width=90, height=34, corner_radius=6,
        fg_color=_NAVY_LIGHT, hover_color=_NAVY_MID,
        text_color=_WHITE, font=FF(9,"bold"), state="disabled")
    self._lu_rescan_btn.pack(side="right", padx=(0,4), pady=11)

    # ── File info strip ────────────────────────────────────────────────
    self._lu_file_strip = tk.Frame(self._lu_analysis_frame, bg=_OFF_WHITE, height=30)
    self._lu_file_strip.pack(fill="x")
    self._lu_file_strip.pack_propagate(False)
    self._lu_file_lbl = tk.Label(
        self._lu_file_strip,
        text="No file loaded  —  click  📂 Load Excel File  to begin",
        font=F(8), fg=_TXT_SOFT, bg=_OFF_WHITE)
    self._lu_file_lbl.pack(side="left", padx=28, pady=6)
    self._lu_status_lbl = tk.Label(self._lu_file_strip, text="",
                                   font=F(8,"bold"), fg=_LIME_DARK, bg=_OFF_WHITE)
    self._lu_status_lbl.pack(side="right", padx=28)
    tk.Frame(self._lu_analysis_frame, bg=_BORDER_LIGHT, height=1).pack(fill="x")

    # ── Client search + selector bar ───────────────────────────────────
    self._lu_client_bar = tk.Frame(self._lu_analysis_frame, bg=_NAVY_MIST, height=50)
    self._lu_client_bar.pack(fill="x")
    self._lu_client_bar.pack_propagate(False)

    self._lu_mode_badge = tk.Label(self._lu_client_bar, text="  GENERAL VIEW  ",
                                   font=F(8,"bold"), fg=_WHITE, bg=_NAVY_MID, padx=10, pady=4)
    self._lu_mode_badge.pack(side="left", padx=(14,8), pady=12)

    # Search entry
    tk.Label(self._lu_client_bar, text="🔍", font=F(11),
             fg=_NAVY_PALE, bg=_NAVY_MIST).pack(side="left", padx=(4,2), pady=12)
    self._lu_search_var = tk.StringVar()
    self._lu_search_entry = ctk.CTkEntry(
        self._lu_client_bar, textvariable=self._lu_search_var,
        placeholder_text="Search client, ID, PN, or sector name…  (blank = all)",
        width=340, height=28, corner_radius=4,
        fg_color=_WHITE, text_color=_TXT_NAVY,
        border_color=_BORDER_MID, font=FF(9))
    self._lu_search_entry.pack(side="left", pady=10)
    self._lu_search_var.trace_add("write", lambda *_: _lu_filter_by_search(self))

    # Clear button
    ctk.CTkButton(self._lu_client_bar, text="✕", width=28, height=28,
                  corner_radius=4, fg_color=_BORDER_MID, hover_color=_ACCENT_RED,
                  text_color=_TXT_NAVY, font=FF(9,"bold"),
                  command=lambda: _lu_clear_sector_filter(self)
                  ).pack(side="left", padx=(2,0), pady=10)

    # Filter pill placeholder (populated dynamically)
    self._lu_filter_pill = None

    # Hidden client var still needed for internal state
    self._lu_client_var = tk.StringVar(value=GENERAL_CLIENT)
    self._lu_client_dropdown = None

    self._lu_client_count_lbl = tk.Label(self._lu_client_bar, text="",
                                         font=F(8), fg=_TXT_SOFT, bg=_NAVY_MIST)
    self._lu_client_count_lbl.pack(side="right", padx=20)
    tk.Frame(self._lu_analysis_frame, bg=_BORDER_MID, height=1).pack(fill="x")

    # ── View container ─────────────────────────────────────────────────
    self._lu_view_container = tk.Frame(self._lu_analysis_frame, bg=_CARD_WHITE)
    self._lu_view_container.pack(fill="both", expand=True)

    # Analysis view (scrollable)
    self._lu_analysis_view = tk.Frame(self._lu_view_container, bg=_CARD_WHITE)
    self._lu_analysis_view.place(relx=0, rely=0, relwidth=1, relheight=1)
    self._lu_scroll_outer, self._lu_results_inner, self._lu_canvas = _make_scrollable(
        self._lu_analysis_view, _CARD_WHITE)
    self._lu_results_frame = self._lu_results_inner

    # Sub-panels
    self._lu_charts_view = tk.Frame(self._lu_view_container, bg=_CARD_WHITE)
    _build_charts_panel(self, self._lu_charts_view)

    self._lu_simulator_view = tk.Frame(self._lu_view_container, bg=_CARD_WHITE)
    _build_simulator_panel(self, self._lu_simulator_view)

    self._lu_loanbal_view = tk.Frame(self._lu_view_container, bg=_CARD_WHITE)
    _build_loanbal_panel(self, self._lu_loanbal_view)

    self._lu_report_view = tk.Frame(self._lu_view_container, bg=_CARD_WHITE)
    _build_report_panel(self, self._lu_report_view)

    self._lu_filepath      = None
    self._lu_results       = []
    self._lu_all_data      = {}
    self._lu_active_client = GENERAL_CLIENT
    self._lu_filtered_sectors = None   # NEW: active sector filter

    _lu_show_placeholder(self)
    _lu_switch_view(self, "analysis")


# ══════════════════════════════════════════════════════════════════════
#  CLIENT SELECTOR + SEARCH  (patched: sector search + cross-tab sync)
# ══════════════════════════════════════════════════════════════════════

def _lu_on_client_change(self, value: str):
    self._lu_active_client = value
    is_general = (value == GENERAL_CLIENT)

    # Respect active sector filter when in general mode
    active_sectors = _lu_get_active_sectors(self)
    all_clients = self._lu_all_data.get("clients", {})

    if is_general:
        if active_sectors:
            # Sector filter active — show only clients in those sectors
            results = [r for r in all_clients.values()
                       if r.get("sector") in active_sectors]
        else:
            results = list(all_clients.values())
    else:
        results = [all_clients[value]] if value in all_clients else []

    self._lu_results = results

    if is_general:
        if active_sectors:
            sector_text = " · ".join(active_sectors)
            icon = _SECTOR_ICON.get(active_sectors[0], "🏭") if len(active_sectors) == 1 else "🏭"
            self._lu_mode_badge.config(
                text=f"  {icon}  SECTOR FILTER  ",
                bg="#1E4080", fg=_LIME_MID)
        else:
            self._lu_mode_badge.config(text="  GENERAL VIEW  ", bg=_NAVY_MID, fg=_WHITE)
    else:
        rec = all_clients.get(value, {})
        label = rec.get("score_label", "N/A")
        accent = _CLIENT_HERO_ACCENT.get(label, _NAVY_PALE)
        self._lu_mode_badge.config(text=f"  PER-CLIENT  ·  {value[:35]}  ",
                                   bg=accent, fg=_WHITE)

    _lu_render_results(self, results)

    # Sync all other tabs with the current filter state
    view = self._lu_active_view.get()
    if view == "charts":      _charts_render(self)
    elif view == "simulator": _sim_populate(self)
    elif view == "report":    _report_render(self)
    elif view == "loanbal":   _loanbal_render(self)


def _lu_filter_by_search(self):
    """
    Filter by client name, ID, PN, or SECTOR NAME.
    Sector matches set _lu_filtered_sectors and propagate to all tabs.
    """
    query   = self._lu_search_var.get().strip().lower()
    clients = list(self._lu_all_data.get("clients", {}).keys())
    all_clients = self._lu_all_data.get("clients", {})

    if not query:
        self._lu_filtered_sectors = None
        _lu_update_filter_pill(self)
        _lu_on_client_change(self, GENERAL_CLIENT)
        return

    # ── Check for sector match first ──────────────────────────────────
    matched_sectors = [
        s for s in _ALL_SECTORS
        if query in s.lower()
    ]

    if matched_sectors and not any(
        query in c.lower()
        or query in str(all_clients[c].get("client_id","")).lower()
        or query in str(all_clients[c].get("pn","")).lower()
        for c in clients
    ):
        # Pure sector match — no client names match
        self._lu_filtered_sectors = matched_sectors
        _lu_update_filter_pill(self)
        _lu_on_client_change(self, GENERAL_CLIENT)
        return

    # ── Check for client match (name / ID / PN) ───────────────────────
    matched_clients = [
        c for c in clients
        if query in c.lower()
        or query in str(all_clients[c].get("client_id","")).lower()
        or query in str(all_clients[c].get("pn","")).lower()
    ]

    # If query matches both a sector AND client names, prefer sector filter
    # but also check if query is clearly a sector keyword
    if matched_sectors and matched_clients:
        # If sector match is strong (full word), prefer sector
        sector_exact = any(s.lower() == query for s in matched_sectors)
        if sector_exact:
            self._lu_filtered_sectors = matched_sectors
            _lu_update_filter_pill(self)
            _lu_on_client_change(self, GENERAL_CLIENT)
            return
        # Otherwise fall through to client match below

    if matched_clients:
        # Clear any sector filter when showing client results
        self._lu_filtered_sectors = None
        _lu_update_filter_pill(self)

        if len(matched_clients) == 1:
            _lu_on_client_change(self, matched_clients[0])
        else:
            self._lu_active_client = GENERAL_CLIENT
            self._lu_mode_badge.config(
                text=f"  {len(matched_clients)} CLIENTS MATCHED  ", bg=_NAVY_PALE, fg=_WHITE)
            results = [all_clients[c] for c in matched_clients if c in all_clients]
            self._lu_results = results
            _lu_render_results(self, results)
    elif matched_sectors:
        # Sector-only match (mixed query path)
        self._lu_filtered_sectors = matched_sectors
        _lu_update_filter_pill(self)
        _lu_on_client_change(self, GENERAL_CLIENT)
    else:
        self._lu_filtered_sectors = None
        _lu_update_filter_pill(self)
        self._lu_mode_badge.config(text="  NO MATCH  ", bg=_ACCENT_RED, fg=_WHITE)
        _lu_render_results(self, [])


def _lu_populate_client_dropdown(self):
    clients      = list(self._lu_all_data.get("clients", {}).keys())
    client_count = len(clients)
    self._lu_client_var.set(GENERAL_CLIENT)
    self._lu_active_client = GENERAL_CLIENT
    self._lu_filtered_sectors = None
    _lu_update_filter_pill(self)
    self._lu_mode_badge.config(text="  GENERAL VIEW  ", bg=_NAVY_MID, fg=_WHITE)
    suffix = "client" if client_count == 1 else "clients"
    self._lu_client_count_lbl.config(text=f"{client_count} {suffix} loaded")


# ══════════════════════════════════════════════════════════════════════
#  VIEW SWITCHER  (patched: re-render current tab with filter state)
# ══════════════════════════════════════════════════════════════════════

def _lu_switch_view(self, view: str):
    self._lu_active_view.set(view)
    views = {
        "analysis":  self._lu_analysis_view,
        "charts":    self._lu_charts_view,
        "simulator": self._lu_simulator_view,
        "loanbal":   self._lu_loanbal_view,
        "report":    self._lu_report_view,
    }
    for name, frame in views.items():
        if name == view:
            frame.place(relx=0, rely=0, relwidth=1, relheight=1)
        else:
            frame.place_forget()

    # Each tab receives the filtered data when it becomes active
    if view == "simulator" and self._lu_all_data:
        print(f"[DEBUG] calling _sim_populate, lu_results count={len(self._lu_results)}")
        self._sim_populate()
    if view == "charts"    and self._lu_all_data: self._charts_render()
    if view == "report"    and self._lu_all_data: self._report_render()
    if view == "loanbal"   and self._lu_all_data: self._loanbal_render()


# ══════════════════════════════════════════════════════════════════════
#  ANALYSIS VIEW — RENDERERS
# ══════════════════════════════════════════════════════════════════════

def _lu_render_results(self, results: list[dict]):
    for w in self._lu_results_frame.winfo_children():
        w.destroy()

    for w in self._lu_analysis_view.winfo_children():
        if w is not self._lu_scroll_outer:
            w.destroy()

    if not results:
        self._lu_scroll_outer.pack(fill="both", expand=True)
        ph = tk.Frame(self._lu_results_frame, bg=_CARD_WHITE)
        ph.pack(expand=True, fill="both", pady=60)
        tk.Label(ph, text="⚠️  No clients found.",
                 font=F(11), fg=_ACCENT_GOLD, bg=_CARD_WHITE).pack(pady=20)
        tk.Label(ph, text="Load an Excel file with a Look-Up Summary sheet.",
                 font=F(9), fg=_TXT_SOFT, bg=_CARD_WHITE, justify="center").pack()
        return

    is_general = (self._lu_active_client == GENERAL_CLIENT)

    if is_general:
        self._lu_scroll_outer.pack_forget()
        direct = tk.Frame(self._lu_analysis_view, bg=_CARD_WHITE)
        direct.pack(fill="both", expand=True)
        _lu_render_general_view(self, results, direct)
    else:
        self._lu_scroll_outer.pack(fill="both", expand=True)
        _lu_render_client_view(self, results)
        self._lu_canvas.yview_moveto(0)


def _lu_render_general_view(self, results: list[dict], parent: tk.Frame):
    import tkinter.ttk as ttk

    # ── Summary stats strip ───────────────────────────────────────────
    total_lb  = sum(r.get("loan_balance") or 0 for r in results)
    total_net = sum(r.get("net_income")   or 0 for r in results)
    sector_counts: dict[str, int] = {}
    for r in results:
        sector_counts[r["sector"]] = sector_counts.get(r["sector"], 0) + 1

    # Show filter context in stats bar
    active_sectors = _lu_get_active_sectors(self)
    stats_bg = "#0E2040" if active_sectors else _NAVY_MIST
    stats_border = _LIME_MID if active_sectors else _BORDER_MID

    stats_bar = tk.Frame(parent, bg=stats_bg,
                         highlightbackground=stats_border, highlightthickness=1)
    stats_bar.pack(fill="x", padx=20, pady=(16, 0))

    if active_sectors:
        # Show sector filter indicator in stats bar
        filter_lbl = tk.Frame(stats_bar, bg=stats_bg)
        filter_lbl.pack(side="left", padx=(12, 0), pady=10)
        icon = _SECTOR_ICON.get(active_sectors[0], "🏭") if len(active_sectors) == 1 else "🏭"
        tk.Label(filter_lbl,
                 text=f"{icon}  Filtered: {' · '.join(active_sectors)}",
                 font=F(8, "bold"), fg=_LIME_MID, bg=stats_bg).pack(anchor="w")
        tk.Label(filter_lbl,
                 text="Showing sector subset only",
                 font=F(7), fg=_TXT_MUTED, bg=stats_bg).pack(anchor="w")

    for lbl, val in [
        ("👥 Clients",          str(len(results))),
        ("💰 Total Loan Bal",   f"₱{total_lb:,.2f}"),
        ("📈 Total Net Income", f"₱{total_net:,.2f}"),
        ("🏭 Sectors",          str(len(sector_counts))),
    ]:
        c = tk.Frame(stats_bar, bg=stats_bg)
        c.pack(side="left", padx=20, pady=10)
        tk.Label(c, text=lbl, font=F(7),        fg=_TXT_SOFT,  bg=stats_bg).pack(anchor="w")
        tk.Label(c, text=val, font=F(12,"bold"), fg=_WHITE if active_sectors else _TXT_NAVY,
                 bg=stats_bg).pack(anchor="w")

    tk.Label(parent, text="Click any row to view full client scorecard",
             font=F(7), fg=_TXT_MUTED, bg=_CARD_WHITE).pack(anchor="e", padx=20, pady=(4, 2))

    # ── Treeview style ────────────────────────────────────────────────
    style = ttk.Style()
    style.theme_use("default")
    style.configure("LU.Treeview",
                    background=_WHITE, foreground=_TXT_NAVY,
                    rowheight=26, fieldbackground=_WHITE,
                    bordercolor=_BORDER_MID, font=("Segoe UI", 9))
    style.configure("LU.Treeview.Heading",
                    background=_NAVY_MID, foreground=_WHITE,
                    font=("Segoe UI", 9, "bold"), relief="flat", borderwidth=0)
    style.map("LU.Treeview.Heading",
              background=[("active", _NAVY_LIGHT)])
    style.map("LU.Treeview",
              background=[("selected", _NAVY_GHOST)],
              foreground=[("selected", _TXT_NAVY)])
    style.layout("LU.Treeview", [
        ("LU.Treeview.treearea", {"sticky": "nswe"})
    ])

    tree_frame = tk.Frame(parent, bg=_CARD_WHITE)
    tree_frame.pack(fill="both", expand=True, padx=20, pady=(0, 16))

    vsb = ttk.Scrollbar(tree_frame, orient="vertical")
    vsb.pack(side="right", fill="y")
    hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
    hsb.pack(side="bottom", fill="x")

    COLS = ("id", "pn", "client", "sector", "industry",
            "total_source", "net_income", "principal_loan", "loan_balance", "risk")
    COL_LABELS = {
        "id":             "ID",
        "pn":             "PN",
        "client":         "Borrower's Name",
        "sector":         "Sector",
        "industry":       "Industry",
        "total_source":   "Total Source of Income",
        "net_income":     "Net Income",
        "principal_loan": "Principal Loan",
        "loan_balance":   "Loan Balance",
        "risk":           "Risk",
    }
    COL_WIDTHS = {
        "id":             75,
        "pn":             120,
        "client":         70,
        "sector":         200,
        "industry":       230,
        "total_source":   140,
        "net_income":     115,
        "principal_loan": 125,
        "loan_balance":   125,
        "risk":           110,
    }
    COL_ANCHOR = {
        "id":             "center",
        "pn":             "center",
        "client":         "w",
        "sector":         "center",
        "industry":       "center",
        "total_source":   "center",
        "net_income":     "center",
        "principal_loan": "center",
        "loan_balance":   "center",
        "risk":           "center",
    }

    tree = ttk.Treeview(tree_frame, columns=COLS, show="headings",
                        style="LU.Treeview",
                        yscrollcommand=vsb.set,
                        xscrollcommand=hsb.set)
    tree.pack(side="left", fill="both", expand=True)
    vsb.config(command=tree.yview)
    hsb.config(command=tree.xview)

    for col in COLS:
        tree.heading(col, text=COL_LABELS[col],
                     command=lambda c=col: _tv_sort(tree, c, False))
        tree.column(col,
                    width=COL_WIDTHS[col],
                    minwidth=40,
                    anchor=COL_ANCHOR[col],
                    stretch=(col == "client"))

    RISK_ICONS = {
        "CRITICAL": "🔴", "HIGH": "🟠",
        "MODERATE": "🟡", "LOW":  "🟢", "N/A": "⚪",
    }
    tree.tag_configure("HIGH",     background="#FFF5F5", foreground=_ACCENT_RED)
    tree.tag_configure("MODERATE", background="#FFFBF0", foreground=_ACCENT_GOLD)
    tree.tag_configure("LOW",      background="#F0FBE8", foreground=_ACCENT_SUCCESS)
    tree.tag_configure("CRITICAL", background="#FFF5F5", foreground="#B71C1C")
    tree.tag_configure("NA",       background=_WHITE,    foreground=_TXT_MUTED)
    tree.tag_configure("alt",      background=_OFF_WHITE)

    _iid_to_client: dict[str, str] = {}

    for idx, rec in enumerate(results):
        rl       = rec.get("score_label", "N/A")
        icon     = RISK_ICONS.get(rl, "⚪")
        sec      = rec.get("sector", "—")
        sec_icon = _SECTOR_ICON.get(sec, "")

        def _fmt(v):
            return f"₱{v:,.2f}" if v else "—"

        values = (
            rec.get("client_id", "—"),
            rec.get("pn", "—"),
            rec.get("client", "—"),
            f"{sec_icon} {sec}",
            (rec.get("industry") or "—")[:40],
            _fmt(rec.get("total_source")),
            _fmt(rec.get("net_income")),
            _fmt(rec.get("principal_loan")),
            _fmt(rec.get("loan_balance")),
            f"{icon} {rl}",
        )

        tag = rl if rl in ("HIGH", "MODERATE", "LOW", "CRITICAL") else "NA"
        if tag == "NA" and idx % 2 == 1:
            tag = "alt"

        iid = tree.insert("", "end", values=values, tags=(tag,))
        _iid_to_client[iid] = rec.get("client", "")

    def _tree_scroll(event):
        tree.yview_scroll(int(-1 * (event.delta / 120)), "units")
        return "break"

    tree.bind("<MouseWheel>", _tree_scroll)

    def _on_row_click(event):
        iid = tree.identify_row(event.y)
        if not iid:
            return
        client_name = _iid_to_client.get(iid, "")
        if client_name:
            self._lu_client_var.set(client_name)
            _lu_on_client_change(self, client_name)

    tree.bind("<ButtonRelease-1>", _on_row_click)
    tree.bind("<Return>", _on_row_click)
    self._lu_general_tree = tree


def _lu_render_client_card(self, parent, rec: dict):
    """Kept for compatibility — no longer called in general view."""
    pass


def _tv_sort(tree: "tk.Treeview", col: str, reverse: bool):
    """Sort treeview by column when header is clicked."""
    data = [(tree.set(iid, col), iid) for iid in tree.get_children("")]

    def _key(val):
        v = val[0].replace("₱", "").replace(",", "").strip()
        try:
            return (0, float(v))
        except ValueError:
            return (1, v.lower())

    data.sort(key=_key, reverse=reverse)
    for idx, (_, iid) in enumerate(data):
        tree.move(iid, "", idx)

    tree.heading(col, command=lambda c=col: _tv_sort(tree, c, not reverse))


def _lu_render_client_view(self, results: list[dict]):
    """Full per-client scorecard for a single client."""
    if not results:
        return
    rec = results[0]
    client_name = rec["client"]
    label       = rec.get("score_label","N/A")
    score       = rec.get("score", 0.0)
    hero_bg     = _CLIENT_HERO_BG.get(label, _NAVY_DEEP)
    hero_accent = _CLIENT_HERO_ACCENT.get(label, _NAVY_PALE)

    pad = tk.Frame(self._lu_results_frame, bg=_CARD_WHITE)
    pad.pack(fill="both", expand=True)

    # ── Hero ─────────────────────────────────────────────────────────
    hero = tk.Frame(pad, bg=hero_bg)
    hero.pack(fill="x")
    hi = tk.Frame(hero, bg=hero_bg)
    hi.pack(fill="x", padx=28, pady=20)

    left = tk.Frame(hi, bg=hero_bg)
    left.pack(side="left", fill="y")
    tk.Label(left, text="PER-CLIENT ANALYSIS",
             font=F(7,"bold"), fg=hero_accent, bg=hero_bg).pack(anchor="w")
    tk.Label(left, text=f"👤  {client_name}",
             font=F(18,"bold"), fg=_WHITE, bg=hero_bg).pack(anchor="w", pady=(4,2))
    icon = _SECTOR_ICON.get(rec.get("sector",""), "📋")
    tk.Label(left, text=f"{icon}  {rec.get('sector','—')}  ·  {rec.get('industry','—')[:50]}",
             font=F(9), fg=hero_accent, bg=hero_bg).pack(anchor="w")

    right = tk.Frame(hi, bg=hero_bg)
    right.pack(side="right", fill="y")
    score_icons = {"CRITICAL":"🔴","HIGH":"🟠","MODERATE":"🟡","LOW":"🟢","N/A":"⚪"}
    tk.Label(right, text=score_icons.get(label,"⚪"),
             font=("Segoe UI Emoji",32), bg=hero_bg).pack()
    tk.Label(right, text=label,
             font=F(16,"bold"), fg=hero_accent, bg=hero_bg).pack()
    tk.Label(right, text=f"Risk Score  {score:.2f}",
             font=F(9), fg=_WHITE, bg=hero_bg).pack()

    # ── Financial summary row ─────────────────────────────────────────
    fin_bar = tk.Frame(pad, bg=_NAVY_MIST)
    fin_bar.pack(fill="x")
    for lbl, val in [
        ("Client ID",          rec.get("client_id","—")),
        ("PN",                 rec.get("pn","—")),
        ("Total Source",       f"₱{rec['total_source']:,.2f}" if rec.get("total_source") else "—"),
        ("Total Business Exp", f"₱{rec['total_biz']:,.2f}"    if rec.get("total_biz")    else "—"),
        ("Total Hhld Exp",     f"₱{rec['total_hhld']:,.2f}"   if rec.get("total_hhld")   else "—"),
        ("Net Income",         f"₱{rec['net_income']:,.2f}"   if rec.get("net_income")   else "—"),
        ("Amort History",      f"₱{rec['amort_history']:,.2f}" if rec.get("amort_history") else "—"),
        ("Current Amort",      f"₱{rec['current_amort']:,.2f}" if rec.get("current_amort") else "—"),
        ("Loan Balance",       f"₱{rec['loan_balance']:,.2f}" if rec.get("loan_balance") else "—"),
    ]:
        c = tk.Frame(fin_bar, bg=_NAVY_MIST)
        c.pack(side="left", padx=12, pady=10)
        tk.Label(c, text=lbl,  font=F(7),     fg=_TXT_SOFT, bg=_NAVY_MIST).pack(anchor="w")
        tk.Label(c, text=val,  font=F(9,"bold"), fg=_TXT_NAVY, bg=_NAVY_MIST).pack(anchor="w")

    tk.Frame(pad, bg=_BORDER_MID, height=1).pack(fill="x")

    # ── Risk chips ────────────────────────────────────────────────────
    expenses = rec.get("expenses", [])
    h_count = sum(1 for e in expenses if e["risk"]=="HIGH")
    m_count = sum(1 for e in expenses if e["risk"]=="MODERATE")
    l_count = sum(1 for e in expenses if e["risk"]=="LOW")
    total_e = len(expenses)

    chips_bar = tk.Frame(pad, bg=hero_bg)
    chips_bar.pack(fill="x")
    chips_inner = tk.Frame(chips_bar, bg=hero_bg)
    chips_inner.pack(side="left", padx=28, pady=(0,14))
    for text, color, bg_chip, count in [
        ("🔴  HIGH RISK",   _ACCENT_RED,     "#FFE8E8", h_count),
        ("🟡  MODERATE",    _ACCENT_GOLD,    "#FFF3CD", m_count),
        ("🟢  LOW RISK",    _ACCENT_SUCCESS, "#DCEDC8", l_count),
        ("📋  TOTAL ITEMS", _TXT_NAVY,       "#F0F4FF", total_e),
    ]:
        chip = tk.Frame(chips_inner, bg=bg_chip,
                        highlightbackground=color, highlightthickness=1)
        chip.pack(side="left", padx=(0,10))
        tk.Label(chip, text=text,       font=F(7,"bold"), fg=color, bg=bg_chip, padx=8, pady=3).pack()
        tk.Label(chip, text=str(count), font=F(14,"bold"), fg=color, bg=bg_chip, padx=8, pady=2).pack()

    tk.Frame(pad, bg=_BORDER_MID, height=1).pack(fill="x")

    if rec.get("source_income"):
        src_frame = tk.Frame(pad, bg=_NAVY_MIST,
                             highlightbackground=_BORDER_MID, highlightthickness=1)
        src_frame.pack(fill="x", padx=28, pady=(14,0))
        tk.Label(src_frame, text="💰  Source of Income",
                 font=F(9,"bold"), fg=_NAVY_MID, bg=_NAVY_MIST,
                 padx=12, pady=6).pack(anchor="w")
        tk.Label(src_frame, text=rec["source_income"],
                 font=F(8), fg=_TXT_NAVY, bg=_NAVY_MIST,
                 padx=12, pady=8, anchor="w", justify="left",
                 wraplength=800).pack(fill="x")

    inner_pad = tk.Frame(pad, bg=_CARD_WHITE)
    inner_pad.pack(fill="both", expand=True, padx=28, pady=16)
    tk.Label(inner_pad, text="Expense Risk Breakdown",
             font=F(11,"bold"), fg=_TXT_NAVY, bg=_CARD_WHITE).pack(anchor="w", pady=(0,10))
    _lu_render_sector_card(self, inner_pad, rec, show_client=False)


def _lu_render_sector_card(self, parent, result: dict, show_client: bool = True):
    sector   = result.get("sector","—")
    client   = result.get("client","—")
    expenses = result.get("expenses",[])

    card = tk.Frame(parent, bg=_WHITE,
                    highlightbackground=_BORDER_MID, highlightthickness=1)
    card.pack(fill="x", pady=(0,16))

    card_hdr = tk.Frame(card, bg=_NAVY_MID)
    card_hdr.pack(fill="x")
    icon = _SECTOR_ICON.get(sector, "📋")
    tk.Label(card_hdr, text=f"{icon}  {sector}",
             font=F(12,"bold"), fg=_WHITE, bg=_NAVY_MID).pack(side="left", padx=16, pady=10)
    if show_client:
        tk.Label(card_hdr, text=f"👤 {client}",
                 font=F(8,"bold"), fg=_LIME_PALE, bg=_NAVY_MID).pack(side="left", padx=(0,4))

    score, score_label, score_fg, score_bg = _compute_risk_score(expenses)
    icons = {"CRITICAL":"🔴","HIGH":"🟠","MODERATE":"🟡","LOW":"🟢"}
    tk.Label(card_hdr, text=f"{icons.get(score_label,'⚪')} {score_label}  {score:.2f}",
             font=F(8,"bold"), fg=score_fg, bg=score_bg,
             padx=10, pady=4).pack(side="left", padx=8, pady=8)

    high_n = sum(1 for e in expenses if e["risk"]=="HIGH")
    mod_n  = sum(1 for e in expenses if e["risk"]=="MODERATE")
    low_n  = sum(1 for e in expenses if e["risk"]=="LOW")
    chips  = tk.Frame(card_hdr, bg=_NAVY_MID)
    chips.pack(side="right", padx=16, pady=8)
    for text, color, bg in [
        (f"HIGH ×{high_n}",   _ACCENT_RED,     "#FFE8E8"),
        (f"MOD ×{mod_n}",     _ACCENT_GOLD,    "#FFF3CD"),
        (f"LOW ×{low_n}",     _ACCENT_SUCCESS, "#DCEDC8"),
    ]:
        tk.Label(chips, text=text, font=F(7,"bold"), fg=color, bg=bg,
                 padx=8, pady=3).pack(side="left", padx=(0,4))

    col_hdr = tk.Frame(card, bg=_OFF_WHITE)
    col_hdr.pack(fill="x")
    for col_idx, lbl in enumerate(["Expense Item","Risk Level","Impact Reason","Value / Amount"]):
        col_hdr.columnconfigure(col_idx, weight=3 if col_idx != 1 else 1)
        tk.Label(col_hdr, text=lbl, font=F(8,"bold"), fg=_NAVY_PALE,
                 bg=_OFF_WHITE, padx=12, pady=6, anchor="w"
                 ).grid(row=0, column=col_idx, sticky="ew")
    tk.Frame(card, bg=_BORDER_LIGHT, height=1).pack(fill="x")

    for idx, exp in enumerate(expenses):
        risk   = exp["risk"]
        row_bg = _RISK_BG.get(risk, _WHITE) if idx % 2 == 0 else _WHITE
        row    = tk.Frame(card, bg=row_bg)
        row.pack(fill="x")
        for c in range(4):
            row.columnconfigure(c, weight=3 if c != 1 else 1)
        tk.Label(row, text=exp["name"], font=F(9,"bold"),
                 fg=_TXT_NAVY, bg=row_bg, padx=12, pady=8, anchor="w"
                 ).grid(row=0, column=0, sticky="ew")
        badge_frame = tk.Frame(row, bg=row_bg)
        badge_frame.grid(row=0, column=1, sticky="w", padx=4)
        tk.Label(badge_frame, text=risk, font=F(8,"bold"),
                 fg=_RISK_COLOR.get(risk, _TXT_SOFT),
                 bg=_RISK_BADGE_BG.get(risk, _OFF_WHITE),
                 padx=10, pady=4).pack(pady=6)
        tk.Label(row, text=exp["reason"], font=F(8), fg=_TXT_SOFT,
                 bg=row_bg, padx=12, pady=8, anchor="w",
                 wraplength=340, justify="left").grid(row=0, column=2, sticky="ew")
        tk.Label(row, text=exp["value_str"], font=F(8),
                 fg=_TXT_NAVY if exp["has_values"] else _TXT_MUTED,
                 bg=row_bg, padx=12, pady=8, anchor="w",
                 wraplength=220, justify="left").grid(row=0, column=3, sticky="ew")
        tk.Frame(card, bg=_BORDER_LIGHT, height=1).pack(fill="x")


# ══════════════════════════════════════════════════════════════════════
#  CHARTS PANEL  (patched: respects active sector filter)
# ══════════════════════════════════════════════════════════════════════

def _build_charts_panel(self, parent):
    hdr = tk.Frame(parent, bg=_NAVY_MID, height=38)
    hdr.pack(fill="x")
    hdr.pack_propagate(False)
    self._charts_hdr_lbl = tk.Label(
        hdr, text="📊  Sector Charts — Wholesale/Retail · Agriculture · Transport · Remittance · Consumer Loan",
        font=F(10,"bold"), fg=_WHITE, bg=_NAVY_MID)
    self._charts_hdr_lbl.pack(side="left", padx=20, pady=8)
    self._charts_body = tk.Frame(parent, bg=_CARD_WHITE)
    self._charts_body.pack(fill="both", expand=True)
    _charts_show_placeholder(self)


def _charts_show_placeholder(self):
    for w in self._charts_body.winfo_children(): w.destroy()
    tk.Label(self._charts_body, text="Run an analysis first to view sector charts.",
             font=F(10), fg=_TXT_MUTED, bg=_CARD_WHITE).pack(pady=60)


def _charts_render(self):
    plt.close("all")
    for w in self._charts_body.winfo_children(): w.destroy()

    if not _HAS_MPL:
        tk.Label(self._charts_body,
                 text="matplotlib is not installed.\nRun:  pip install matplotlib",
                 font=F(10), fg=_ACCENT_RED, bg=_CARD_WHITE).pack(pady=40)
        return

    # Use filtered data if a sector filter is active
    all_data   = _lu_get_filtered_all_data(self)
    sector_map = all_data.get("sector_map", {})
    general    = all_data.get("general", [])

    if not general:
        _charts_show_placeholder(self)
        return

    # Update header to reflect filter state
    active_sectors = _lu_get_active_sectors(self)
    if active_sectors:
        sector_text = " · ".join(active_sectors)
        self._charts_hdr_lbl.config(
            text=f"📊  Charts — Filtered: {sector_text}",
            fg=_LIME_MID)
    else:
        self._charts_hdr_lbl.config(
            text="📊  Sector Charts — Wholesale/Retail · Agriculture · Transport · Remittance · Consumer Loan",
            fg=_WHITE)

    _, inner, _ = _make_scrollable(self._charts_body, _CARD_WHITE)
    pad = tk.Frame(inner, bg=_CARD_WHITE)
    pad.pack(fill="both", expand=True, padx=24, pady=16)

    # Filter which sectors to chart
    chart_sectors = [s for s in _CHART_SECTORS if s in sector_map]

    def _embed(fig, frame):
        FigureCanvasTkAgg(fig, master=frame).get_tk_widget().pack(
            fill="both", expand=True, padx=4, pady=4)

    # ── Chart 1: Client count per sector ─────────────────────────────
    tk.Label(pad, text="Client Distribution by Sector",
             font=F(11,"bold"), fg=_TXT_NAVY, bg=_CARD_WHITE).pack(anchor="w", pady=(0,8))
    fig1 = None
    try:
        fig1, ax1 = plt.subplots(figsize=(9, 3.5))
        fig1.patch.set_facecolor(_MPL_BG); ax1.set_facecolor(_MPL_BG)
        sec_names  = [s for s in chart_sectors if s in sector_map]
        sec_counts = [len(sector_map[s]) for s in sec_names]
        colors     = [_SECTOR_COLORS.get(s, _MPL_NAVY) for s in sec_names]
        if sec_names:
            bars = ax1.bar(sec_names, sec_counts, color=colors, edgecolor=_MPL_BG,
                           linewidth=1.5, width=0.55)
            ax1.set_ylabel("Number of Clients", fontsize=9, color=_MPL_NAVY)
            ax1.tick_params(axis="x", labelsize=8, rotation=10)
            ax1.tick_params(axis="y", labelsize=8)
            ax1.spines[["top","right"]].set_visible(False)
            for bar, val in zip(bars, sec_counts):
                ax1.text(bar.get_x()+bar.get_width()/2,
                         bar.get_height()+0.05,
                         str(val), ha="center", va="bottom",
                         fontsize=9, fontweight="bold", color=_MPL_NAVY)
        else:
            ax1.text(0.5,0.5,"No sector data", ha="center", va="center",
                     transform=ax1.transAxes, color=_TXT_MUTED)
        fig1.tight_layout(pad=1.2)
        c1 = tk.Frame(pad, bg=_WHITE, highlightbackground=_BORDER_MID, highlightthickness=1)
        c1.pack(fill="x", pady=(0,16))
        _embed(fig1, c1)
    except Exception: pass
    finally:
        if fig1: plt.close(fig1)

    # ── Chart 2: Risk distribution doughnut per sector ────────────────
    tk.Label(pad, text="Risk Distribution by Sector",
             font=F(11,"bold"), fg=_TXT_NAVY, bg=_CARD_WHITE).pack(anchor="w", pady=(0,8))
    row_charts = tk.Frame(pad, bg=_CARD_WHITE)
    row_charts.pack(fill="x", pady=(0,16))

    for sector in chart_sectors:
        recs = sector_map.get(sector, [])
        fig_s = None
        try:
            all_exp = [e for r in recs for e in r.get("expenses",[])]
            h = sum(1 for e in all_exp if e["risk"]=="HIGH")
            m = sum(1 for e in all_exp if e["risk"]=="MODERATE")
            l = sum(1 for e in all_exp if e["risk"]=="LOW")
            total = h + m + l
            fig_s, ax_s = plt.subplots(figsize=(3.0, 2.8))
            fig_s.patch.set_facecolor(_MPL_BG); ax_s.set_facecolor(_MPL_BG)
            if total > 0:
                non_zero = [(v,c,lab) for v,c,lab in zip(
                    [h,m,l],[_MPL_HIGH,_MPL_MOD,_MPL_LOW],["HIGH","MOD","LOW"]) if v > 0]
                vals,cols,labs = zip(*non_zero) if non_zero else ([1],["#ccc"],["N/A"])
                wedges, _ = ax_s.pie(vals, colors=list(cols), startangle=90,
                                     wedgeprops=dict(width=0.5, edgecolor=_MPL_BG, linewidth=1.5))
                ax_s.text(0,0,str(len(recs)), ha="center", va="center",
                          fontsize=13, fontweight="bold", color=_MPL_NAVY)
                ax_s.text(0,-0.25,"clients", ha="center", va="center",
                          fontsize=7, color="#6B7FA3")
            else:
                ax_s.text(0.5,0.5,"No data", ha="center", va="center",
                          transform=ax_s.transAxes, color=_TXT_MUTED, fontsize=8)
            icon = _SECTOR_ICON.get(sector,"")
            ax_s.set_title(f"{icon} {sector}", fontsize=8,
                           color=_SECTOR_COLORS.get(sector,_MPL_NAVY), fontweight="bold", pad=4,
                           wrap=True)
            fig_s.tight_layout(pad=0.8)
            cf = tk.Frame(row_charts, bg=_WHITE,
                          highlightbackground=_BORDER_MID, highlightthickness=1)
            cf.pack(side="left", fill="both", expand=True, padx=(0,8))
            _embed(fig_s, cf)
        except Exception: pass
        finally:
            if fig_s: plt.close(fig_s)

    # ── Chart 3: Total Loan Balance by sector ─────────────────────────
    tk.Label(pad, text="Total Loan Balance by Sector",
             font=F(11,"bold"), fg=_TXT_NAVY, bg=_CARD_WHITE).pack(anchor="w", pady=(0,8))
    fig3 = None
    try:
        fig3, ax3 = plt.subplots(figsize=(9, 3.5))
        fig3.patch.set_facecolor(_MPL_BG); ax3.set_facecolor(_MPL_BG)
        sec_names = [s for s in chart_sectors if s in sector_map]
        sec_vals  = [sum(r.get("loan_balance") or 0 for r in sector_map[s]) for s in sec_names]
        colors    = [_SECTOR_COLORS.get(s,_MPL_NAVY) for s in sec_names]
        if sec_names and any(v > 0 for v in sec_vals):
            bars = ax3.bar(sec_names, sec_vals, color=colors, edgecolor=_MPL_BG,
                           linewidth=1.5, width=0.55)
            ax3.yaxis.set_major_formatter(
                matplotlib.ticker.FuncFormatter(lambda x,_: f"₱{x/1e6:.1f}M" if x>=1e6 else f"₱{x:,.0f}"))
            ax3.tick_params(axis="x", labelsize=8, rotation=10)
            ax3.tick_params(axis="y", labelsize=8)
            ax3.spines[["top","right"]].set_visible(False)
            max_v = max(sec_vals) if sec_vals else 1
            for bar, val in zip(bars, sec_vals):
                lbl = f"₱{val/1e6:.2f}M" if val>=1e6 else f"₱{val:,.0f}"
                ax3.text(bar.get_x()+bar.get_width()/2,
                         val + max_v*0.01, lbl,
                         ha="center", va="bottom", fontsize=7, color=_MPL_NAVY)
        else:
            ax3.text(0.5,0.5,"No loan balance data", ha="center", va="center",
                     transform=ax3.transAxes, color=_TXT_MUTED)
        fig3.tight_layout(pad=1.2)
        c3 = tk.Frame(pad, bg=_WHITE, highlightbackground=_BORDER_MID, highlightthickness=1)
        c3.pack(fill="x", pady=(0,16))
        _embed(fig3, c3)
    except Exception: pass
    finally:
        if fig3: plt.close(fig3)

    plt.close("all")


# ══════════════════════════════════════════════════════════════════════
#  SECTOR vs LOAN BALANCE TAB
# ══════════════════════════════════════════════════════════════════════

_SECTOR_COLS = [
    ("Sector",              3, 200, "w"),
    ("# Clients",           1,  80, "center"),
    ("Total Loan Balance",  2, 170, "center"),
    ("% of Total",          2, 130, "center"),
    ("Avg Loan per Client", 2, 160, "center"),
    ("Avg Net Income",      2, 150, "center"),
    ("Risk Profile",        1, 100, "center"),
]

_CLIENT_COLS = [
    ("Client",          3, 200, "w"),
    ("ID",              1,  60, "center"),
    ("Sector",          2, 160, "center"),
    ("Principal Loan",  2, 130, "center"),
    ("Loan Balance",    2, 130, "center"),
    ("% of Total",      1, 100, "center"),
    ("Net Income",      2, 120, "center"),
    ("Current Amort",   2, 130, "center"),
    ("Risk",            1,  90, "center"),
]


def _make_table_frame(parent, col_specs):
    tf = tk.Frame(parent, bg=_WHITE)
    tf.pack(fill="x", padx=0, pady=0)
    for ci, (_, weight, min_px, _anchor) in enumerate(col_specs):
        tf.columnconfigure(ci, weight=weight, minsize=min_px)
    return tf


def _table_header(table_frame, col_specs):
    hdr_bg = _NAVY_MID
    bg_strip = tk.Frame(table_frame, bg=hdr_bg)
    bg_strip.grid(row=0, column=0, columnspan=len(col_specs), sticky="nsew")
    bg_strip.lower()
    for ci, (label, _w, _min, anchor) in enumerate(col_specs):
        cell = tk.Label(
            table_frame, text=label,
            font=F(8, "bold"), fg=_WHITE, bg=hdr_bg,
            anchor=anchor, padx=10, pady=9
        )
        cell.grid(row=0, column=ci, sticky="nsew", ipadx=0)


def _table_divider(table_frame, row_idx, n_cols, color=_BORDER_LIGHT):
    div = tk.Frame(table_frame, bg=color, height=1)
    div.grid(row=row_idx, column=0, columnspan=n_cols, sticky="ew")


# ══════════════════════════════════════════════════════════════════════
#  BUILD LOANBAL PANEL
# ══════════════════════════════════════════════════════════════════════

def _build_loanbal_panel(self, parent):
    hdr = tk.Frame(parent, bg=_NAVY_MID, height=46)
    hdr.pack(fill="x")
    hdr.pack_propagate(False)

    self._loanbal_hdr_lbl = tk.Label(
        hdr, text="📊  Sector vs Total Loan Balance  —  Exposure Analysis",
        font=F(10, "bold"), fg=_WHITE, bg=_NAVY_MID)
    self._loanbal_hdr_lbl.pack(side="left", padx=20, pady=12)

    self._loanbal_export_btn = ctk.CTkButton(
        hdr,
        text="💾  Export",
        command=lambda: _loanbal_show_export_menu(self),
        width=110, height=30, corner_radius=6,
        fg_color=_LIME_DARK, hover_color=_LIME_MID,
        text_color=_TXT_ON_LIME, font=FF(9, "bold"),
        state="disabled"
    )
    self._loanbal_export_btn.pack(side="right", padx=16, pady=8)

    self._loanbal_body = tk.Frame(parent, bg=_CARD_WHITE)
    self._loanbal_body.pack(fill="both", expand=True)
    tk.Label(self._loanbal_body,
             text="Run an analysis first to view loan balance exposure.",
             font=F(10), fg=_TXT_MUTED, bg=_CARD_WHITE).pack(pady=60)


# ══════════════════════════════════════════════════════════════════════
#  LOANBAL EXPORT MENU  (unchanged — exports full data always)
# ══════════════════════════════════════════════════════════════════════

def _loanbal_show_export_menu(self):
    """Show popup export menu for the Sector vs Loan Balance tab.
    NOTE: Export always uses the FULL unfiltered dataset by design.
    """
    # Determine label context
    active_sectors = _lu_get_active_sectors(self)
    if active_sectors:
        sector_text = " · ".join(active_sectors)
        pdf_label = f"📄  Export PDF — Full Report (filter: {sector_text[:30]})"
        xl_label  = f"📊  Export Excel — Full Report (filter: {sector_text[:30]})"
    else:
        pdf_label = "📄  Export PDF — Sector & Client Loan Balance"
        xl_label  = "📊  Export Excel — Sector & Client Loan Balance"

    menu = tk.Menu(
        self._lu_loanbal_view, tearoff=0,
        font=F(9), bg=_WHITE, fg=_TXT_NAVY,
        activebackground=_NAVY_GHOST, activeforeground=_NAVY_DEEP,
        relief="flat", bd=1
    )
    menu.add_command(label=pdf_label, command=lambda: _loanbal_export_pdf(self))
    menu.add_command(label=xl_label,  command=lambda: _loanbal_export_excel(self))
    try:
        btn = self._loanbal_export_btn
        btn.update_idletasks()
        x = btn.winfo_rootx()
        y = btn.winfo_rooty() + btn.winfo_height()
        menu.tk_popup(x, y)
    finally:
        menu.grab_release()


# ══════════════════════════════════════════════════════════════════════
#  LOANBAL PDF EXPORT  (unchanged — always exports full data)
# ══════════════════════════════════════════════════════════════════════

def _loanbal_export_pdf(self):
    if not self._lu_all_data:
        messagebox.showwarning("No Data", "Run an analysis first.")
        return
    if not _HAS_RL:
        messagebox.showerror("Missing Library",
                             "reportlab is not installed.\nRun:  pip install reportlab")
        return

    default_name = f"LoanBalance_Exposure_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    path = filedialog.asksaveasfilename(
        title="Save Loan Balance PDF",
        defaultextension=".pdf",
        filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
        initialfile=default_name
    )
    if not path:
        return

    try:
        # Always use full unfiltered data for export
        _generate_loanbal_pdf(self._lu_all_data, path,
                              filepath=self._lu_filepath or "")
        messagebox.showinfo("Export Complete", f"PDF saved to:\n{path}")
    except Exception as ex:
        messagebox.showerror("PDF Export Error", str(ex))


def _generate_loanbal_pdf(all_data, out_path, filepath=""):
    from reportlab.lib.pagesizes import A4, landscape as rl_landscape

    PAGE        = rl_landscape(A4)
    doc         = SimpleDocTemplate(
        out_path, pagesize=PAGE,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.8*cm, bottomMargin=1.8*cm
    )

    styles  = getSampleStyleSheet()
    navy    = rl_colors.HexColor("#1A3A6B")
    lime    = rl_colors.HexColor("#5A9E28")
    white   = rl_colors.white
    off     = rl_colors.HexColor("#F5F7FA")
    mist    = rl_colors.HexColor("#EEF3FB")
    border  = rl_colors.HexColor("#C5D0E8")
    red     = rl_colors.HexColor("#E53E3E")
    gold    = rl_colors.HexColor("#D4A017")
    green   = rl_colors.HexColor("#2E7D32")
    dark_bg = rl_colors.HexColor("#0A1628")
    lime_mid = rl_colors.HexColor("#8FD14F")

    title_s  = ParagraphStyle("LBTitle",  parent=styles["Title"],
                               fontSize=16, textColor=navy, spaceAfter=2)
    sub_s    = ParagraphStyle("LBSub",    parent=styles["Normal"],
                               fontSize=8, textColor=rl_colors.HexColor("#9AAACE"), leading=10)
    h1_s     = ParagraphStyle("LBH1",     parent=styles["Normal"],
                               fontSize=12, textColor=white, leading=16, spaceAfter=0)
    h2_s     = ParagraphStyle("LBH2",     parent=styles["Normal"],
                               fontSize=10, textColor=navy, leading=14, spaceBefore=10)
    body_s   = ParagraphStyle("LBBody",   parent=styles["Normal"],
                               fontSize=8, textColor=rl_colors.HexColor("#1A2B4A"), leading=11)
    muted_s  = ParagraphStyle("LBMuted",  parent=styles["Normal"],
                               fontSize=7, textColor=rl_colors.HexColor("#9AAACE"), leading=10)

    RISK_RL  = {"HIGH": red, "MODERATE": gold, "LOW": green, "N/A": rl_colors.grey}
    SEC_RL   = {
        "Wholesale/Retail": rl_colors.HexColor("#1A3A6B"),
        "Agriculture":      rl_colors.HexColor("#2E7D32"),
        "Transport":        rl_colors.HexColor("#D4A017"),
        "Remittance":       rl_colors.HexColor("#8B5CF6"),
        "Consumer Loan":    rl_colors.HexColor("#E53E3E"),
        "Other":            rl_colors.HexColor("#9AAACE"),
    }

    general    = all_data.get("general", [])
    sector_map = all_data.get("sector_map", {})
    totals     = all_data.get("totals", {})
    grand_lb   = totals.get("loan_balance", 0) or 0
    now        = datetime.now().strftime("%B %d, %Y  %H:%M")
    fname      = Path(filepath).name if filepath else "—"

    story = []

    story.append(Paragraph("Sector vs Loan Balance — Exposure Analysis", title_s))
    story.append(Paragraph(f"File: {fname}    Generated: {now}    "
                            f"Total Clients: {len(general)}    "
                            f"Grand Total Loan Balance: ₱{grand_lb:,.2f}", sub_s))
    story.append(Spacer(1, 0.3*cm))
    story.append(HRFlowable(width="100%", thickness=1.5, color=navy))
    story.append(Spacer(1, 0.4*cm))

    story.append(Paragraph("Sector Loan Balance Breakdown", h2_s))
    story.append(Spacer(1, 0.2*cm))

    all_sectors = [s for s in _CHART_SECTORS if s in sector_map]
    from lu_core import SECTOR_OTHER as _SO
    if _SO in sector_map and _SO not in all_sectors:
        all_sectors.append(_SO)

    sector_rows_data = []
    for sector in all_sectors:
        recs    = sector_map.get(sector, [])
        n       = len(recs)
        s_lb    = sum(r.get("loan_balance") or 0 for r in recs)
        s_net   = sum(r.get("net_income")   or 0 for r in recs)
        pct     = (s_lb / grand_lb * 100) if grand_lb > 0 else 0.0
        s_pl    = sum(r.get("principal_loan") or 0 for r in recs)
        avg_lb  = s_pl / n  if n > 0 else 0.0
        avg_net = s_net / n if n > 0 else 0.0
        all_exp = [e for r in recs for e in r.get("expenses", [])]
        _, risk_label, _, _ = _compute_risk_score(all_exp)
        sector_rows_data.append((sector, n, s_lb, pct, avg_lb, avg_net, risk_label))

    sector_rows_data.sort(key=lambda x: -x[2])

    pg_w   = PAGE[0] - 3*cm
    s_cols = [pg_w*0.22, pg_w*0.09, pg_w*0.17, pg_w*0.12,
              pg_w*0.17, pg_w*0.14, pg_w*0.09]

    sec_hdr_s = ParagraphStyle("SecHdr", parent=styles["Normal"],
                               fontSize=8, textColor=white, leading=11)

    sec_tbl_data = [[
        Paragraph("<b>Sector</b>",              sec_hdr_s),
        Paragraph("<b># Clients</b>",           sec_hdr_s),
        Paragraph("<b>Total Loan Balance</b>",  sec_hdr_s),
        Paragraph("<b>% of Total</b>",          sec_hdr_s),
        Paragraph("<b>Avg Loan / Client</b>",   sec_hdr_s),
        Paragraph("<b>Avg Net Income</b>",      sec_hdr_s),
        Paragraph("<b>Risk Profile</b>",        sec_hdr_s),
    ]]

    for sector, n, s_lb, pct, avg_lb, avg_net, risk_label in sector_rows_data:
        icon     = _SECTOR_ICON.get(sector, "")
        col_hex  = _SECTOR_COLORS.get(sector, "#1A3A6B").lstrip("#")
        risk_col = RISK_RL.get(risk_label, rl_colors.grey)
        sec_tbl_data.append([
            Paragraph(f"<font color='#{col_hex}'><b>{icon}  {sector}</b></font>", body_s),
            Paragraph(str(n), body_s),
            Paragraph(f"<b>₱{s_lb:,.2f}</b>", body_s),
            Paragraph(f"{pct:.1f}%", body_s),
            Paragraph(f"₱{avg_lb:,.2f}", body_s),
            Paragraph(f"₱{avg_net:,.2f}" if avg_net else "—", body_s),
            Paragraph(
                f"<font color='#{_rgb_hex(risk_col)}'><b>{risk_label}</b></font>",
                body_s
            ),
        ])

    sec_tbl_data.append([
        Paragraph("<b>GRAND TOTAL</b>", body_s),
        Paragraph(f"<b>{len(general)}</b>", body_s),
        Paragraph(f"<b>₱{grand_lb:,.2f}</b>", body_s),
        Paragraph("<b>100.0%</b>", body_s),
        Paragraph(f"<b>₱{grand_lb/len(general):,.2f}</b>" if general else "—", body_s),
        Paragraph("—", body_s),
        Paragraph("—", body_s),
    ])

    sec_style = TableStyle([
        ("BACKGROUND",   (0, 0), (-1, 0),  navy),
        ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("TEXTCOLOR",    (0, 0), (-1, 0),  white),
        ("FONTSIZE",     (0, 0), (-1, -1), 8),
        ("LEADING",      (0, 0), (-1, -1), 11),
        ("TOPPADDING",   (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
        ("LEFTPADDING",  (0, 0), (-1, -1), 6),
        ("BOX",          (0, 0), (-1, -1), 0.5, border),
        ("INNERGRID",    (0, 0), (-1, -1), 0.3, border),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        *[("BACKGROUND", (0, i), (-1, i), off if i % 2 == 0 else white)
          for i in range(1, len(sec_tbl_data) - 1)],
        ("BACKGROUND",   (0, -1), (-1, -1), mist),
        ("FONTNAME",     (0, -1), (-1, -1), "Helvetica-Bold"),
        ("LINEABOVE",    (0, -1), (-1, -1), 1.2, navy),
    ])

    sec_tbl = Table(sec_tbl_data, colWidths=s_cols, repeatRows=1)
    sec_tbl.setStyle(sec_style)
    story.append(sec_tbl)

    story.append(PageBreak())
    story.append(Paragraph("Individual Client Loan Balance", title_s))
    story.append(Paragraph(
        f"File: {fname}    Generated: {now}    "
        f"Sorted by: Loan Balance (Descending)    "
        f"Grand Total: ₱{grand_lb:,.2f}", sub_s))
    story.append(Spacer(1, 0.3*cm))
    story.append(HRFlowable(width="100%", thickness=1.5, color=navy))
    story.append(Spacer(1, 0.4*cm))

    c_cols = [pg_w*0.07, pg_w*0.07, pg_w*0.16, pg_w*0.12,
              pg_w*0.09, pg_w*0.09, pg_w*0.07, pg_w*0.09,
              pg_w*0.08, pg_w*0.08, pg_w*0.09, pg_w*0.09]

    hdr_s = ParagraphStyle("CLIHdr", parent=styles["Normal"],
                           fontSize=7, textColor=white, leading=9)

    cli_tbl_data = [[
        Paragraph("<b>Client ID</b>",       hdr_s),
        Paragraph("<b>PN</b>",              hdr_s),
        Paragraph("<b>Client</b>",          hdr_s),
        Paragraph("<b>Sector</b>",          hdr_s),
        Paragraph("<b>Principal Loan</b>",  hdr_s),
        Paragraph("<b>Loan Balance</b>",    hdr_s),
        Paragraph("<b>% of Total</b>",      hdr_s),
        Paragraph("<b>Net Income</b>",      hdr_s),
        Paragraph("<b>Current Amort</b>",   hdr_s),
        Paragraph("<b>Amort History</b>",   hdr_s),
        Paragraph("<b>Total Source</b>",    hdr_s),
        Paragraph("<b>Risk</b>",            hdr_s),
    ]]

    clients_sorted = sorted(general, key=lambda r: -(r.get("loan_balance") or 0))
    for rec in clients_sorted:
        lb      = rec.get("loan_balance") or 0
        pl      = rec.get("principal_loan") or 0
        net     = rec.get("net_income") or 0
        amrt_c  = rec.get("current_amort") or 0
        amrt_h  = rec.get("amort_history") or 0
        total_s = rec.get("total_source") or 0
        pct     = (lb / grand_lb * 100) if grand_lb > 0 else 0.0
        rl      = rec.get("score_label", "N/A")
        sec     = rec.get("sector", "—")
        col_hex = _SECTOR_COLORS.get(sec, "#4A6FA5").lstrip("#")
        risk_col = RISK_RL.get(rl, rl_colors.grey)
        icon    = _SECTOR_ICON.get(sec, "")

        cli_tbl_data.append([
            Paragraph(rec.get("client_id", "—"), muted_s),
            Paragraph(rec.get("pn", "—"), muted_s),
            Paragraph(f"<b>{rec['client'][:30]}</b>", body_s),
            Paragraph(f"<font color='#{col_hex}'> {sec[:18]}</font>", body_s),
            Paragraph(f"{pl:,.2f}" if pl else "—", body_s),
            Paragraph(f"<b>{lb:,.2f}</b>", body_s),
            Paragraph(f"{pct:.2f}%", body_s),
            Paragraph(f"{net:,.2f}" if net else "—", body_s),
            Paragraph(f"{amrt_c:,.2f}" if amrt_c else "—", body_s),
            Paragraph(f"{amrt_h:,.2f}" if amrt_h else "—", body_s),
            Paragraph(f"{total_s:,.2f}" if total_s else "—", body_s),
            Paragraph(
                f"<font color='#{_rgb_hex(risk_col)}'><b>{rl}</b></font>",
                body_s
            ),
        ])

    total_net_all  = sum(r.get("net_income") or 0 for r in general)
    total_pl_all   = sum(r.get("principal_loan") or 0 for r in general)
    total_amrt_c   = sum(r.get("current_amort") or 0 for r in general)
    total_amrt_h   = sum(r.get("amort_history") or 0 for r in general)
    total_src_all  = sum(r.get("total_source") or 0 for r in general)
    cli_tbl_data.append([
        Paragraph("<b>TOTAL</b>", body_s),
        Paragraph(f"<b>{len(general)}</b>", body_s),
        Paragraph("—", body_s),
        Paragraph("—", body_s),
        Paragraph(f"<b>{total_pl_all:,.2f}</b>", body_s),
        Paragraph(f"<b>₱{grand_lb:,.2f}</b>", body_s),
        Paragraph("<b>100.00%</b>", body_s),
        Paragraph(f"<b>{total_net_all:,.2f}</b>", body_s),
        Paragraph(f"<b>{total_amrt_c:,.2f}</b>", body_s),
        Paragraph(f"<b>{total_amrt_h:,.2f}</b>", body_s),
        Paragraph(f"<b>{total_src_all:,.2f}</b>", body_s),
        Paragraph("—", body_s),
    ])

    cli_style = TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  navy),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  white),
        ("FONTSIZE",      (0, 0), (-1, -1), 7),
        ("LEADING",       (0, 0), (-1, -1), 10),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING",   (0, 0), (-1, -1), 5),
        ("BOX",           (0, 0), (-1, -1), 0.5, border),
        ("INNERGRID",     (0, 0), (-1, -1), 0.3, border),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        *[("BACKGROUND", (0, i), (-1, i), off if i % 2 == 0 else white)
          for i in range(1, len(cli_tbl_data) - 1)],
        ("BACKGROUND",    (0, -1), (-1, -1), mist),
        ("FONTNAME",      (0, -1), (-1, -1), "Helvetica-Bold"),
        ("LINEABOVE",     (0, -1), (-1, -1), 1.2, navy),
    ])

    cli_tbl = Table(cli_tbl_data, colWidths=c_cols, repeatRows=1)
    cli_tbl.setStyle(cli_style)
    story.append(cli_tbl)

    doc.build(story)


# ══════════════════════════════════════════════════════════════════════
#  LOANBAL EXCEL EXPORT  (unchanged — always exports full data)
# ══════════════════════════════════════════════════════════════════════

def _loanbal_export_excel(self):
    if not self._lu_all_data:
        messagebox.showwarning("No Data", "Run an analysis first.")
        return
    if not _HAS_OPENPYXL:
        messagebox.showerror("Missing Library",
                             "openpyxl is not installed.\nRun:  pip install openpyxl")
        return

    default_name = f"LoanBalance_Exposure_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    path = filedialog.asksaveasfilename(
        title="Save Loan Balance Excel",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialfile=default_name
    )
    if not path:
        return

    try:
        # Always use full unfiltered data for export
        _generate_loanbal_excel(self._lu_all_data, path,
                                filepath=self._lu_filepath or "")
        messagebox.showinfo("Export Complete", f"Excel saved to:\n{path}")
    except Exception as ex:
        messagebox.showerror("Excel Export Error", str(ex))


def _generate_loanbal_excel(all_data, out_path, filepath=""):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def fill(hex_col):
        return PatternFill("solid", fgColor=hex_col.lstrip("#"))

    def thin_border():
        s = Side(style="thin", color="C5D0E8")
        return Border(left=s, right=s, top=s, bottom=s)

    def thick_bottom():
        thin = Side(style="thin",   color="C5D0E8")
        thck = Side(style="medium", color="1A3A6B")
        return Border(left=thin, right=thin, top=thin, bottom=thck)

    FILLS = {
        "hdr":   fill("#93C47D"),
        "alt":   fill("#F5F7FA"),
        "white": fill("#FFFFFF"),
        "total": fill("#D6E4F7"),
    }
    RISK_FC = {"HIGH": "E53E3E", "MODERATE": "D4A017", "LOW": "2E7D32", "N/A": "9AAACE"}
    NUM_FMT = '#,##0.00;\\(#,##0.00\\);"-"'

    general    = all_data.get("general", [])
    sector_map = all_data.get("sector_map", {})
    totals     = all_data.get("totals", {})
    grand_lb   = totals.get("loan_balance", 0) or 0
    now        = datetime.now().strftime("%Y-%m-%d %H:%M")
    fname      = Path(filepath).name if filepath else "—"

    ws1 = wb.create_sheet("Sector Breakdown")

    ws1.merge_cells("A1:G1")
    ws1["A1"] = "Sector vs Loan Balance — Exposure Analysis"
    ws1["A1"].font      = Font(bold=True, size=14, color="0A1628")
    ws1["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[1].height = 24

    ws1.merge_cells("A2:G2")
    ws1["A2"] = f"File: {fname}    Generated: {now}    Grand Total: ₱{grand_lb:,.2f}    Clients: {len(general)}"
    ws1["A2"].font      = Font(size=8, color="9AAACE")
    ws1["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[2].height = 16

    col_widths_s = [30, 12, 22, 14, 22, 20, 14]
    hdrs_s = ["Sector", "# Clients", "Total Loan Balance",
              "% of Total", "Avg Loan per Client", "Avg Net Income", "Risk Profile"]
    for ci, (w, h) in enumerate(zip(col_widths_s, hdrs_s), 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    for ci, h in enumerate(hdrs_s, 1):
        c = ws1.cell(4, ci, h)
        c.fill      = FILLS["hdr"]
        c.font      = Font(bold=True, color="FFFFFF", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin_border()
    ws1.row_dimensions[4].height = 22

    all_sectors = [s for s in _CHART_SECTORS if s in sector_map]
    from lu_core import SECTOR_OTHER as _SO
    if _SO in sector_map and _SO not in all_sectors:
        all_sectors.append(_SO)

    sector_rows_data = []
    for sector in all_sectors:
        recs    = sector_map.get(sector, [])
        n       = len(recs)
        s_lb    = sum(r.get("loan_balance") or 0 for r in recs)
        s_net   = sum(r.get("net_income")   or 0 for r in recs)
        pct     = (s_lb / grand_lb * 100) if grand_lb > 0 else 0.0
        avg_lb  = s_lb / n  if n > 0 else 0.0
        avg_net = s_net / n if n > 0 else 0.0
        all_exp = [e for r in recs for e in r.get("expenses", [])]
        _, risk_label, _, _ = _compute_risk_score(all_exp)
        sector_rows_data.append((sector, n, s_lb, pct, avg_lb, avg_net, risk_label))

    sector_rows_data.sort(key=lambda x: -x[2])

    data_start = 5
    for idx, (sector, n, s_lb, pct, avg_lb, avg_net, risk_label) in enumerate(sector_rows_data):
        row_num   = data_start + idx
        row_fill  = FILLS["alt"] if idx % 2 == 0 else FILLS["white"]
        icon      = _SECTOR_ICON.get(sector, "")
        sec_color = _SECTOR_COLORS.get(sector, "#4A6FA5").lstrip("#")
        risk_fc   = RISK_FC.get(risk_label, "9AAACE")

        values = [f"{icon}  {sector}", n, s_lb, pct / 100, avg_lb,
                  avg_net if avg_net else None, risk_label]
        fmts   = [None, "0", '#,##0.00', '0.00%', '#,##0.00', '#,##0.00', None]
        bolds  = [True, False, True, False, False, False, True]
        colors = [sec_color, "1A2B4A", "1A2B4A", "1A2B4A", "1A2B4A", "1A2B4A", risk_fc]

        for ci, (val, fmt, bold, color) in enumerate(zip(values, fmts, bolds, colors), 1):
            c = ws1.cell(row_num, ci, val)
            c.fill      = row_fill
            c.font      = Font(bold=bold, size=9, color=color)
            c.border    = thin_border()
            c.alignment = Alignment(
                horizontal="left" if ci == 1 else "center",
                vertical="center", indent=1 if ci == 1 else 0)
            if fmt:
                c.number_format = fmt
        ws1.row_dimensions[row_num].height = 18

    total_row = data_start + len(sector_rows_data)
    avg_grand = grand_lb / len(general) if general else 0
    grand_vals = ["GRAND TOTAL", len(general), grand_lb, 1.0, avg_grand, None, "—"]
    grand_fmts = [None, "0", '#,##0.00', '0.00%', '#,##0.00', None, None]
    for ci, (val, fmt) in enumerate(zip(grand_vals, grand_fmts), 1):
        c = ws1.cell(total_row, ci, val)
        c.fill      = FILLS["total"]
        c.font      = Font(bold=True, size=9, color="0A1628")
        c.border    = thick_bottom()
        c.alignment = Alignment(
            horizontal="left" if ci == 1 else "center",
            vertical="center", indent=1 if ci == 1 else 0)
        if fmt:
            c.number_format = fmt
    ws1.row_dimensions[total_row].height = 20
    ws1.freeze_panes = "A5"

    ws2 = wb.create_sheet("Client Breakdown")

    ws2.merge_cells("A1:L1")
    ws2["A1"] = "Client Loan Balance — Individual Breakdown"
    ws2["A1"].font      = Font(bold=True, size=14, color="0A1628")
    ws2["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws2.row_dimensions[1].height = 24

    ws2.merge_cells("A2:L2")
    ws2["A2"] = (
        f"File: {fname}    Generated: {now}    "
        f"Total Clients: {len(general)}    "
        f"Grand Total Loan Balance: ₱{grand_lb:,.2f}    "
        f"Sorted by: Loan Balance (Descending)"
    )
    ws2["A2"].font      = Font(size=8, color="9AAACE")
    ws2["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws2.row_dimensions[2].height = 16

    ws2.row_dimensions[3].height = 6

    CLI_HDRS = [
        "Client ID", "PN", "Client", "Sector",
        "Principal Loan", "Loan Balance", "% of Total",
        "Net Income", "Current Amort", "Amort History", "Total Source", "Risk Label",
    ]
    CLI_WIDTHS = [12, 14, 32, 24, 18, 18, 12, 18, 16, 16, 18, 14]
    CLI_NUM_COLS = {4, 5, 7, 8, 9, 10}
    CLI_PCT_COLS = {6}
    PCT_FMT = '0.00%'

    for ci, (h, w) in enumerate(zip(CLI_HDRS, CLI_WIDTHS), 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
        c = ws2.cell(4, ci, h)
        c.fill      = FILLS["hdr"]
        c.font      = Font(bold=True, color="FFFFFF", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin_border()
    ws2.row_dimensions[4].height = 22

    clients_sorted = sorted(general, key=lambda r: -(r.get("loan_balance") or 0))
    sum_lb = sum_net = sum_amrt_c = sum_amrt_h = sum_src = 0.0

    for idx, rec in enumerate(clients_sorted):
        row_num  = 5 + idx
        lb       = rec.get("loan_balance") or 0
        net      = rec.get("net_income") or 0
        amrt_c   = rec.get("current_amort") or 0
        amrt_h   = rec.get("amort_history") or 0
        total_s  = rec.get("total_source") or 0
        pct      = (lb / grand_lb) if grand_lb > 0 else 0.0
        rl       = rec.get("score_label", "N/A")
        pl       = rec.get("principal_loan") or 0
        sec      = rec.get("sector", "—")
        icon     = _SECTOR_ICON.get(sec, "")
        sec_fc   = _SECTOR_COLORS.get(sec, "#4A6FA5").lstrip("#")
        risk_fc  = RISK_FC.get(rl, "9AAACE")
        row_fill = FILLS["alt"] if idx % 2 == 0 else FILLS["white"]

        sum_lb     += lb
        sum_net    += net
        sum_amrt_c += amrt_c
        sum_amrt_h += amrt_h
        sum_src    += total_s

        row_vals = [
            rec.get("client_id", ""),
            rec.get("pn", ""),
            rec.get("client", ""),
            f"{icon}  {sec}",
            pl     if pl     else None,
            lb     if lb     else None,
            pct,
            net    if net    else None,
            amrt_c if amrt_c else None,
            amrt_h if amrt_h else None,
            total_s if total_s else None,
            rl,
        ]

        for ci, val in enumerate(row_vals, 1):
            col_idx = ci - 1
            c = ws2.cell(row_num, ci, val)
            c.fill   = row_fill
            c.border = thin_border()

            if col_idx in CLI_NUM_COLS:
                c.font          = Font(bold=True, size=9, color="1A2B4A")
                c.alignment     = Alignment(horizontal="right", vertical="center")
                c.number_format = NUM_FMT
            elif col_idx in CLI_PCT_COLS:
                c.font          = Font(bold=False, size=9, color="1A2B4A")
                c.alignment     = Alignment(horizontal="center", vertical="center")
                c.number_format = PCT_FMT
            elif col_idx == 11:
                c.font      = Font(bold=True, size=9, color=risk_fc)
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 3:
                c.font      = Font(bold=False, size=9, color=sec_fc)
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 0:
                c.font      = Font(bold=False, size=9, color="6B7FA3")
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 2:
                c.font      = Font(bold=True, size=9, color="1A2B4A")
                c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            else:
                c.font      = Font(bold=False, size=9, color="6B7FA3")
                c.alignment = Alignment(horizontal="center", vertical="center")

        ws2.row_dimensions[row_num].height = 18

    total_row = 5 + len(clients_sorted)
    ws2.row_dimensions[total_row].height = 20

    sum_pl = sum(r.get("principal_loan") or 0 for r in clients_sorted)
    total_vals = ["GRAND TOTAL", "", "", "", sum_pl, sum_lb, 1.0,
                  sum_net, sum_amrt_c, sum_amrt_h, sum_src, ""]
    total_fmts = [None, None, None, None, NUM_FMT, NUM_FMT, PCT_FMT,
                  NUM_FMT, NUM_FMT, NUM_FMT, NUM_FMT, None]

    for ci, (val, fmt) in enumerate(zip(total_vals, total_fmts), 1):
        col_idx = ci - 1
        c = ws2.cell(total_row, ci, val)
        c.fill   = FILLS["total"]
        c.font   = Font(bold=True, size=9, color="0A1628")
        c.border = thick_bottom()
        if col_idx in CLI_NUM_COLS or col_idx in CLI_PCT_COLS:
            c.alignment = Alignment(horizontal="right", vertical="center")
        elif col_idx == 0:
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        else:
            c.alignment = Alignment(horizontal="center", vertical="center")
        if fmt:
            c.number_format = fmt

    ws2.freeze_panes = "A5"
    wb.save(out_path)


# ══════════════════════════════════════════════════════════════════════
#  LOANBAL RENDER  (patched: respects active sector filter for display)
# ══════════════════════════════════════════════════════════════════════

def _loanbal_render(self):
    for w in self._loanbal_body.winfo_children():
        w.destroy()
    plt.close("all")

    # Use filtered data for display only; export always uses full data
    all_data   = _lu_get_filtered_all_data(self)
    general    = all_data.get("general", [])
    sector_map = all_data.get("sector_map", {})
    totals     = all_data.get("totals", {})
    grand_lb   = totals.get("loan_balance", 0) or 0

    # Update header
    active_sectors = _lu_get_active_sectors(self)
    if active_sectors:
        sector_text = " · ".join(active_sectors)
        self._loanbal_hdr_lbl.config(
            text=f"📊  Loan Balance — Filtered: {sector_text}",
            fg=_LIME_MID)
    else:
        self._loanbal_hdr_lbl.config(
            text="📊  Sector vs Total Loan Balance  —  Exposure Analysis",
            fg=_WHITE)

    self._loanbal_export_btn.configure(state="normal")

    if not general:
        tk.Label(self._loanbal_body, text="No data available for this filter.",
                 font=F(10), fg=_TXT_MUTED, bg=_CARD_WHITE).pack(pady=60)
        return

    _, inner, canvas_scroll = _make_scrollable(self._loanbal_body, _CARD_WHITE)
    pad = tk.Frame(inner, bg=_CARD_WHITE)
    pad.pack(fill="both", expand=True, padx=20, pady=14)

    # ── Grand-total card ──────────────────────────────────────────────
    grand_card = tk.Frame(pad, bg=_NAVY_DEEP,
                          highlightbackground=_NAVY_MID, highlightthickness=1)
    grand_card.pack(fill="x", pady=(0, 14))
    gc_inner = tk.Frame(grand_card, bg=_NAVY_DEEP)
    gc_inner.pack(fill="both", padx=22, pady=16)

    # Left side — main balance figure
    left_gc = tk.Frame(gc_inner, bg=_NAVY_DEEP)
    left_gc.pack(side="left", fill="y")

    lbl_icon_text = (
        f"💰  FILTERED LOAN BALANCE  ·  {' · '.join(active_sectors)}"
        if active_sectors else "💰  GRAND TOTAL LOAN BALANCE"
    )
    lbl_icon_fg = _LIME_MID if active_sectors else _TXT_MUTED
    tk.Label(left_gc, text=lbl_icon_text,
             font=F(9, "bold"), fg=lbl_icon_fg, bg=_NAVY_DEEP).pack(anchor="w")

    tk.Label(left_gc, text=f"₱{grand_lb:,.2f}",
             font=F(22, "bold"), fg=_LIME_MID, bg=_NAVY_DEEP).pack(anchor="w")

    sector_count = len(sector_map)
    tk.Label(left_gc,
             text=f"{len(general)} clients  ·  {sector_count} sector(s)",
             font=F(9), fg=_TXT_SOFT, bg=_NAVY_DEEP).pack(anchor="w")

    if active_sectors:
        tk.Label(left_gc,
                 text="⚠  Export button exports the full unfiltered dataset",
                 font=F(7), fg=_ACCENT_GOLD, bg=_NAVY_DEEP).pack(anchor="w", pady=(4, 0))

    # Right side — secondary stats (Total Net Income + Avg Loan per Client)
    right_gc = tk.Frame(gc_inner, bg=_NAVY_DEEP)
    right_gc.pack(side="right", anchor="e", pady=(4, 0))
    total_net = sum(r.get("net_income") or 0 for r in general)
    avg_loan  = (grand_lb / len(general)) if general else 0.0
    for lbl_text, val_text in [
        ("Total Net Income",    f"₱{total_net:,.2f}"),
        ("Avg Loan per Client", f"₱{avg_loan:,.2f}"),
    ]:
        row_f = tk.Frame(right_gc, bg=_NAVY_DEEP)
        row_f.pack(anchor="e", pady=3)
        tk.Label(row_f, text=lbl_text, font=F(8),
                 fg=_TXT_MUTED, bg=_NAVY_DEEP).pack(side="left", padx=(0, 12))
        tk.Label(row_f, text=val_text, font=F(11, "bold"),
                 fg=_WHITE, bg=_NAVY_DEEP).pack(side="left")

    # ── Section heading ───────────────────────────────────────────────
    tk.Label(pad, text="Sector Loan Balance Breakdown",
             font=F(11, "bold"), fg=_TXT_NAVY, bg=_CARD_WHITE
             ).pack(anchor="w", pady=(0, 6))

    # ── Sector table ──────────────────────────────────────────────────
    sector_tf = _make_table_frame(pad, _SECTOR_COLS)
    _table_header(sector_tf, _SECTOR_COLS)
    _table_divider(sector_tf, 1, len(_SECTOR_COLS), _NAVY_MID)

    all_sectors = [s for s in _CHART_SECTORS if s in sector_map]
    if SECTOR_OTHER in sector_map and SECTOR_OTHER not in all_sectors:
        all_sectors.append(SECTOR_OTHER)

    sector_rows_data = []
    for sector in all_sectors:
        recs    = sector_map.get(sector, [])
        n       = len(recs)
        s_lb    = sum(r.get("loan_balance") or 0 for r in recs)
        s_net   = sum(r.get("net_income")   or 0 for r in recs)
        pct     = (s_lb / grand_lb * 100) if grand_lb > 0 else 0.0
        avg_lb  = s_lb / n  if n > 0 else 0.0
        avg_net = s_net / n if n > 0 else 0.0
        all_exp = [e for r in recs for e in r.get("expenses", [])]
        _, risk_label, _, _ = _compute_risk_score(all_exp)
        sector_rows_data.append((sector, n, s_lb, pct, avg_lb, avg_net, risk_label))

    sector_rows_data.sort(key=lambda x: -x[2])

    grid_row = 2
    for idx, (sector, n, s_lb, pct, avg_lb, avg_net, risk_label) in enumerate(sector_rows_data):
        row_bg    = _CARD_WHITE if idx % 2 == 0 else _OFF_WHITE
        col_color = _SECTOR_COLORS.get(sector, _NAVY_MID)
        icon      = _SECTOR_ICON.get(sector, "📋")
        risk_fg   = _RISK_COLOR.get(risk_label, _TXT_SOFT)
        risk_bg   = _RISK_BADGE_BG.get(risk_label, _OFF_WHITE)

        # Highlight active sector rows
        if active_sectors and sector in active_sectors:
            row_bg = "#0E2040"

        stripe = tk.Frame(sector_tf, bg=row_bg)
        stripe.grid(row=grid_row, column=0, columnspan=len(_SECTOR_COLS), sticky="nsew")
        stripe.lower()

        name_fg = _LIME_MID if (active_sectors and sector in active_sectors) else col_color
        val_fg  = _WHITE    if (active_sectors and sector in active_sectors) else _TXT_NAVY

        # Col 0: Sector name — left aligned, sector colour, bold
        tk.Label(sector_tf, text=f"  {icon}  {sector}", font=F(9, "bold"),
                 fg=name_fg, bg=row_bg, anchor="w", padx=8, pady=11
                 ).grid(row=grid_row, column=0, sticky="nsew")
        # Col 1: # Clients — centered
        tk.Label(sector_tf, text=str(n), font=F(9),
                 fg=val_fg, bg=row_bg, anchor="center", padx=6, pady=11
                 ).grid(row=grid_row, column=1, sticky="nsew")
        # Col 2: Total Loan Balance — bold, right-ish
        tk.Label(sector_tf, text=f"₱{s_lb:,.2f}", font=F(9, "bold"),
                 fg=val_fg, bg=row_bg, anchor="e", padx=14, pady=11
                 ).grid(row=grid_row, column=2, sticky="nsew")
        # Col 3: % of Total with mini bar
        pct_cell = tk.Frame(sector_tf, bg=row_bg)
        pct_cell.grid(row=grid_row, column=3, sticky="nsew", padx=8, pady=6)
        tk.Label(pct_cell, text=f"{pct:.1f}%", font=F(9, "bold"),
                 fg=name_fg, bg=row_bg, anchor="w").pack(anchor="w", pady=(3, 1))
        bar_outer = tk.Frame(pct_cell, bg=_BORDER_LIGHT, height=5)
        bar_outer.pack(fill="x", pady=(0, 2))
        bar_outer.pack_propagate(False)
        fill_w = max(3, int(100 * pct / 100))
        tk.Frame(bar_outer, bg=col_color, height=5, width=fill_w).place(x=0, y=0, relheight=1)
        # Col 4: Avg Loan per Client
        tk.Label(sector_tf, text=f"₱{avg_lb:,.2f}", font=F(9),
                 fg=val_fg, bg=row_bg, anchor="e", padx=14, pady=11
                 ).grid(row=grid_row, column=4, sticky="nsew")
        # Col 5: Avg Net Income
        tk.Label(sector_tf, text=f"₱{avg_net:,.2f}" if avg_net else "—",
                 font=F(9), fg=val_fg, bg=row_bg, anchor="e", padx=14, pady=11
                 ).grid(row=grid_row, column=5, sticky="nsew")
        # Col 6: Risk badge — outlined pill centred in cell
        badge_cell = tk.Frame(sector_tf, bg=row_bg)
        badge_cell.grid(row=grid_row, column=6, sticky="nsew", pady=9, padx=10)
        tk.Label(badge_cell, text=risk_label, font=F(8, "bold"),
                 fg=risk_fg, bg=risk_bg,
                 padx=12, pady=5,
                 relief="flat",
                 highlightbackground=risk_fg, highlightthickness=1
                 ).pack(anchor="center")

        grid_row += 1
        div = tk.Frame(sector_tf, bg=_BORDER_LIGHT, height=1)
        div.grid(row=grid_row, column=0, columnspan=len(_SECTOR_COLS), sticky="ew")
        grid_row += 1

    # ── Pie / bar chart ───────────────────────────────────────────────
    if _HAS_MPL and sector_rows_data:
        tk.Label(pad, text="Loan Balance Share by Sector",
                 font=F(11, "bold"), fg=_TXT_NAVY, bg=_CARD_WHITE
                 ).pack(anchor="w", pady=(20, 8))
        fig_pie = None
        try:
            fig_pie, (ax_pie, ax_bar) = plt.subplots(1, 2, figsize=(11, 4.8))
            fig_pie.patch.set_facecolor(_MPL_BG)
            for ax in (ax_pie, ax_bar):
                ax.set_facecolor(_MPL_BG)

            valid = [(s, lb, pct) for s, n, lb, pct, *_ in sector_rows_data if lb > 0]
            if valid:
                snames  = [x[0] for x in valid]
                svals   = [x[1] for x in valid]
                spcts   = [x[2] for x in valid]
                scolors = [_SECTOR_COLORS.get(s, _MPL_NAVY) for s in snames]
                labels  = [f"{_SECTOR_ICON.get(s,'')}\n{s}\n{p:.1f}%"
                           for s, p in zip(snames, spcts)]

                wedges, _ = ax_pie.pie(
                    svals, colors=scolors, startangle=90,
                    wedgeprops=dict(width=0.55, edgecolor=_MPL_BG, linewidth=2))
                ax_pie.legend(wedges, labels, loc="lower center", fontsize=7,
                              frameon=False, ncol=3, bbox_to_anchor=(0.5, -0.22))
                total_str = (f"₱{grand_lb/1e6:.2f}M" if grand_lb >= 1e6
                             else f"₱{grand_lb:,.0f}")
                ax_pie.text(0, 0.1, total_str, ha="center", va="center",
                            fontsize=10, fontweight="bold", color=_MPL_NAVY)
                ax_pie.text(0, -0.15, "total", ha="center", va="center",
                            fontsize=8, color="#6B7FA3")
                ax_pie.set_title("Loan Balance Share",
                                 fontsize=10, color=_MPL_NAVY,
                                 fontweight="bold", pad=8)

                short_names = [f"{_SECTOR_ICON.get(s,'')} {s[:22]}" for s in snames]
                bars = ax_bar.barh(short_names, svals, color=scolors,
                                   edgecolor=_MPL_BG, linewidth=1.2, height=0.55)
                ax_bar.invert_yaxis()
                ax_bar.xaxis.set_major_formatter(
                    matplotlib.ticker.FuncFormatter(
                        lambda x, _: f"₱{x/1e6:.1f}M" if x >= 1e6 else f"₱{x:,.0f}"))
                ax_bar.tick_params(axis="x", labelsize=7)
                ax_bar.tick_params(axis="y", labelsize=8)
                ax_bar.spines[["top", "right"]].set_visible(False)
                max_v = max(svals) if svals else 1
                for bar, val, pct in zip(bars, svals, spcts):
                    ax_bar.text(val + max_v * 0.01,
                                bar.get_y() + bar.get_height() / 2,
                                f"{pct:.1f}%", va="center",
                                fontsize=8, fontweight="bold", color=_MPL_NAVY)
                ax_bar.set_title("Loan Balance by Sector",
                                 fontsize=10, color=_MPL_NAVY,
                                 fontweight="bold", pad=8)

            fig_pie.tight_layout(pad=1.6)
            c_pie = tk.Frame(pad, bg=_WHITE,
                             highlightbackground=_BORDER_MID, highlightthickness=1)
            c_pie.pack(fill="x", pady=(0, 14))
            FigureCanvasTkAgg(fig_pie, master=c_pie).get_tk_widget().pack(
                fill="both", expand=True, padx=4, pady=4)
        except Exception:
            pass
        finally:
            if fig_pie:
                plt.close(fig_pie)

    # ── Per-client table ──────────────────────────────────────────────
    tk.Label(pad, text="Individual Client Loan Balance",
             font=F(11, "bold"), fg=_TXT_NAVY, bg=_CARD_WHITE
             ).pack(anchor="w", pady=(14, 6))

    client_tf = _make_table_frame(pad, _CLIENT_COLS)
    _table_header(client_tf, _CLIENT_COLS)
    _table_divider(client_tf, 1, len(_CLIENT_COLS), _NAVY_MID)

    clients_sorted = sorted(general, key=lambda r: -(r.get("loan_balance") or 0))
    grid_row = 2
    for idx, rec in enumerate(clients_sorted):
        lb      = rec.get("loan_balance") or 0
        pl      = rec.get("principal_loan") or 0
        net     = rec.get("net_income") or 0
        amrt    = rec.get("current_amort") or 0
        pct     = (lb / grand_lb * 100) if grand_lb > 0 else 0.0
        rl      = rec.get("score_label", "N/A")
        sec     = rec.get("sector", "—")
        row_bg  = _CARD_WHITE if idx % 2 == 0 else _OFF_WHITE
        col_clr = _SECTOR_COLORS.get(sec, _NAVY_MID)
        risk_fg = _RISK_COLOR.get(rl, _TXT_SOFT)
        risk_bg = _RISK_BADGE_BG.get(rl, _OFF_WHITE)

        stripe = tk.Frame(client_tf, bg=row_bg)
        stripe.grid(row=grid_row, column=0,
                    columnspan=len(_CLIENT_COLS), sticky="nsew")
        stripe.lower()

        tk.Label(client_tf, text=f"  {rec['client'][:30]}", font=F(9, "bold"),
                 fg=_TXT_NAVY, bg=row_bg, anchor="w", padx=8, pady=10
                 ).grid(row=grid_row, column=0, sticky="nsew")
        tk.Label(client_tf, text=rec.get("client_id", "—"), font=F(8),
                 fg=_TXT_SOFT, bg=row_bg, anchor="center", padx=4, pady=10
                 ).grid(row=grid_row, column=1, sticky="nsew")
        tk.Label(client_tf,
                 text=f"{_SECTOR_ICON.get(sec,'')} {sec[:22]}",
                 font=F(8), fg=col_clr, bg=row_bg, anchor="center", padx=6, pady=10
                 ).grid(row=grid_row, column=2, sticky="nsew")
        tk.Label(client_tf, text=f"₱{pl:,.2f}" if pl else "—", font=F(9),
                 fg=_TXT_NAVY, bg=row_bg, anchor="e", padx=14, pady=10
                 ).grid(row=grid_row, column=3, sticky="nsew")
        tk.Label(client_tf, text=f"₱{lb:,.2f}", font=F(9, "bold"),
                 fg=_TXT_NAVY, bg=row_bg, anchor="e", padx=14, pady=10
                 ).grid(row=grid_row, column=4, sticky="nsew")

        pct_cell = tk.Frame(client_tf, bg=row_bg)
        pct_cell.grid(row=grid_row, column=5, sticky="nsew", padx=8, pady=5)
        tk.Label(pct_cell, text=f"{pct:.2f}%", font=F(8, "bold"),
                 fg=col_clr, bg=row_bg, anchor="w").pack(anchor="w", pady=(3, 1))
        bar_outer = tk.Frame(pct_cell, bg=_BORDER_LIGHT, height=4)
        bar_outer.pack(fill="x", pady=(0, 2))
        bar_outer.pack_propagate(False)
        bw = max(2, int(80 * pct / 100))
        tk.Frame(bar_outer, bg=col_clr, height=4, width=bw).place(x=0, y=0, relheight=1)

        tk.Label(client_tf, text=f"₱{net:,.2f}" if net else "—",
                 font=F(9), fg=_TXT_NAVY, bg=row_bg, anchor="e", padx=14, pady=10
                 ).grid(row=grid_row, column=6, sticky="nsew")
        tk.Label(client_tf, text=f"₱{amrt:,.2f}" if amrt else "—",
                 font=F(9), fg=_TXT_NAVY, bg=row_bg, anchor="e", padx=14, pady=10
                 ).grid(row=grid_row, column=7, sticky="nsew")

        badge_cell = tk.Frame(client_tf, bg=row_bg)
        badge_cell.grid(row=grid_row, column=8, sticky="nsew", pady=9, padx=10)
        tk.Label(badge_cell, text=rl, font=F(8, "bold"),
                 fg=risk_fg, bg=risk_bg, padx=10, pady=4,
                 highlightbackground=risk_fg, highlightthickness=1
                 ).pack(anchor="center")

        grid_row += 1
        div = tk.Frame(client_tf, bg=_BORDER_LIGHT, height=1)
        div.grid(row=grid_row, column=0, columnspan=len(_CLIENT_COLS), sticky="ew")
        grid_row += 1

    plt.close("all")


# ══════════════════════════════════════════════════════════════════════
#  REPORT PANEL  (patched: respects active sector filter)
# ══════════════════════════════════════════════════════════════════════

def _build_report_panel(self, parent):
    hdr = tk.Frame(parent, bg=_NAVY_MID, height=38)
    hdr.pack(fill="x")
    hdr.pack_propagate(False)
    tk.Label(hdr, text="🖨  Print-Ready Report",
             font=F(10,"bold"), fg=_WHITE, bg=_NAVY_MID).pack(side="left", padx=20, pady=8)
    ctk.CTkButton(hdr, text="🖨  Print This Report",
                  command=lambda: _report_print(self),
                  width=140, height=26, corner_radius=4,
                  fg_color=_LIME_DARK, hover_color=_LIME_MID,
                  text_color=_TXT_ON_LIME, font=FF(8,"bold")
                  ).pack(side="right", padx=12, pady=6)

    body = tk.Frame(parent, bg=_CARD_WHITE)
    body.pack(fill="both", expand=True)
    rsb = tk.Scrollbar(body, relief="flat", troughcolor=_OFF_WHITE,
                       bg=_BORDER_LIGHT, width=8, bd=0)
    rsb.pack(side="right", fill="y")
    self._report_text = tk.Text(
        body, font=("Consolas",9), fg=_TXT_NAVY, bg=_WHITE,
        relief="flat", bd=0, padx=28, pady=20, wrap="word",
        yscrollcommand=rsb.set, state="disabled")
    self._report_text.pack(side="left", fill="both", expand=True)
    rsb.config(command=self._report_text.yview)

    t = self._report_text
    t.tag_configure("title",         font=("Consolas",13,"bold"), foreground=_NAVY_DEEP)
    t.tag_configure("h1",            font=("Consolas",11,"bold"), foreground=_NAVY_MID)
    t.tag_configure("h2",            font=("Consolas",10,"bold"), foreground=_NAVY_LIGHT)
    t.tag_configure("high",          font=("Consolas", 9,"bold"), foreground=_ACCENT_RED)
    t.tag_configure("mod",           font=("Consolas", 9,"bold"), foreground=_ACCENT_GOLD)
    t.tag_configure("low",           font=("Consolas", 9,"bold"), foreground=_ACCENT_SUCCESS)
    t.tag_configure("muted",         font=("Consolas", 8),        foreground=_TXT_MUTED)
    t.tag_configure("normal",        font=("Consolas", 9),        foreground=_TXT_NAVY)
    t.tag_configure("rule",          font=("Consolas", 8),        foreground=_BORDER_MID)
    t.tag_configure("client_name",   font=("Consolas",11,"bold"), foreground=_NAVY_MID)
    t.tag_configure("general_tag",   font=("Consolas", 9,"bold"), foreground=_LIME_DARK)
    t.tag_configure("perclient_tag", font=("Consolas", 9,"bold"), foreground="#8B5CF6")
    t.tag_configure("sector_tag",    font=("Consolas", 9,"bold"), foreground=_LIME_MID)
    for key, fg in [("score_critical","#B71C1C"),("score_high",_ACCENT_RED),
                    ("score_moderate",_ACCENT_GOLD),("score_low",_ACCENT_SUCCESS)]:
        t.tag_configure(key, font=("Consolas",9,"bold"), foreground=fg)
    _report_show_placeholder(self)


def _report_show_placeholder(self):
    self._report_text.config(state="normal")
    self._report_text.delete("1.0","end")
    self._report_text.insert("end","Run an analysis first to generate the report.","muted")
    self._report_text.config(state="disabled")


def _report_render(self):
    t          = self._report_text
    client     = self._lu_active_client
    is_general = (client == GENERAL_CLIENT)
    all_data   = self._lu_all_data

    # Use filtered data for report
    filtered_data = _lu_get_filtered_all_data(self)
    active_sectors = _lu_get_active_sectors(self)

    if is_general:
        results = filtered_data.get("general", [])
    else:
        results = [all_data["clients"][client]] if client in all_data.get("clients",{}) else []

    now   = datetime.now().strftime("%B %d, %Y  %H:%M")
    fname = Path(self._lu_filepath).name if self._lu_filepath else "—"
    rule  = "─"*78+"\n"
    dbl   = "═"*78+"\n"
    RISK_TAG = {"HIGH":"high","MODERATE":"mod","LOW":"low"}

    t.config(state="normal")
    t.delete("1.0","end")
    t.insert("end", dbl,"rule")
    t.insert("end", "  LU RISK ANALYSIS REPORT\n","title")
    t.insert("end", dbl,"rule")
    t.insert("end", f"  File     : {fname}\n","normal")
    t.insert("end", f"  Generated: {now}\n","normal")

    if is_general:
        totals = filtered_data.get("totals",{})
        if active_sectors:
            t.insert("end","  Mode     : ","normal")
            t.insert("end",f"SECTOR FILTER — {' · '.join(active_sectors)}\n","sector_tag")
            t.insert("end",f"  Sectors  : {', '.join(active_sectors)}\n","normal")
        else:
            t.insert("end","  Mode     : ","normal")
            t.insert("end","GENERAL VIEW — All Clients\n","general_tag")
        t.insert("end",f"  Clients  : {len(results)}\n","normal")
        t.insert("end",f"  Total Loan Balance : ₱{totals.get('loan_balance',0):,.2f}\n","normal")
        t.insert("end",f"  Total Net Income   : ₱{totals.get('total_net',0):,.2f}\n","normal")
    else:
        rec = results[0] if results else {}
        t.insert("end","  Mode     : ","normal")
        t.insert("end","PER-CLIENT VIEW\n","perclient_tag")
        t.insert("end","  Client   : ","normal")
        t.insert("end",f"{rec.get('client','—')}\n","client_name")
        t.insert("end",f"  Client ID: {rec.get('client_id','—')}\n","normal")
        t.insert("end",f"  PN       : {rec.get('pn','—')}\n","normal")
        t.insert("end",f"  Industry : {rec.get('industry','—')}\n","normal")
        t.insert("end",f"  Sector   : {rec.get('sector','—')}\n","normal")
        t.insert("end",f"  Loan Bal : ₱{rec.get('loan_balance') or 0:,.2f}\n","normal")
        t.insert("end",f"  Net Inc  : ₱{rec.get('net_income') or 0:,.2f}\n","normal")
        t.insert("end","  Risk Score: ","normal")
        lbl = rec.get("score_label","N/A")
        t.insert("end",f"{lbl}  ({rec.get('score',0):.2f})\n",f"score_{lbl.lower()}")

    t.insert("end", dbl,"rule")
    t.insert("end","\n")

    for rec in results:
        icon = _SECTOR_ICON.get(rec.get("sector",""),"📋")
        t.insert("end",rule,"rule")
        t.insert("end",f"  {icon}  CLIENT: ","h1")
        t.insert("end",f"{rec['client']}\n","client_name")
        t.insert("end",f"  Industry: {rec.get('industry','—')}  |  Sector: {rec.get('sector','—')}\n","normal")
        t.insert("end",f"  ID: {rec.get('client_id','—')}  PN: {rec.get('pn','—')}\n","normal")
        t.insert("end",f"  Total Source Income: ₱{rec.get('total_source') or 0:,.2f}  "
                       f"Net Income: ₱{rec.get('net_income') or 0:,.2f}  "
                       f"Loan Balance: ₱{rec.get('loan_balance') or 0:,.2f}\n","normal")
        t.insert("end",rule,"rule")

        exps = rec.get("expenses",[])
        h = sum(1 for e in exps if e["risk"]=="HIGH")
        m = sum(1 for e in exps if e["risk"]=="MODERATE")
        l = sum(1 for e in exps if e["risk"]=="LOW")
        t.insert("end","  Risk summary: ","normal")
        t.insert("end",f"HIGH×{h} ","high")
        t.insert("end",f"MODERATE×{m} ","mod")
        t.insert("end",f"LOW×{l}\n\n","low")

        col_w = [26,10,44]
        t.insert("end",f"  {'EXPENSE ITEM':<{col_w[0]}} {'RISK':<{col_w[1]}} {'IMPACT REASON'}\n","h2")
        t.insert("end","  "+"─"*76+"\n","rule")
        for exp in exps:
            tag  = RISK_TAG.get(exp["risk"],"normal")
            name = exp["name"] if len(exp["name"])<=col_w[0] else exp["name"][:col_w[0]-1]+"…"
            rsn  = exp["reason"][:col_w[2]] if len(exp["reason"])>col_w[2] else exp["reason"]
            t.insert("end",f"  {name:<{col_w[0]}} {exp['risk']:<{col_w[1]}} {rsn}\n",tag)
            t.insert("end","  "+" "*(col_w[0]+col_w[1]+1)+f"↳ {exp['value_str']}\n","muted")
        t.insert("end","\n")

    t.insert("end",dbl,"rule")
    t.insert("end","  END OF REPORT\n","muted")
    t.insert("end",dbl,"rule")
    t.config(state="disabled")
    t.yview_moveto(0)


def _report_print(self):
    try:
        import subprocess, platform, tempfile, os
        content = self._report_text.get("1.0","end")
        with tempfile.NamedTemporaryFile(mode="w", suffix=".txt",
                                         delete=False, encoding="utf-8") as f:
            f.write(content); tmp = f.name
        if platform.system() == "Windows":
            os.startfile(tmp, "print")
        else:
            subprocess.run(["lpr", tmp])
    except Exception as ex:
        messagebox.showerror("Print Error", str(ex))


# ══════════════════════════════════════════════════════════════════════
#  MAIN EXPORT MENU  (top-bar Export button)
# ══════════════════════════════════════════════════════════════════════

def _lu_show_export_menu(self):
    client     = self._lu_active_client
    is_general = (client == GENERAL_CLIENT)
    active_sectors = _lu_get_active_sectors(self)

    if active_sectors:
        sector_text = " · ".join(active_sectors)
        label_pdf = f"📄  Export PDF — Sector: {sector_text[:35]}"
        label_xl  = f"📊  Export Excel — Sector: {sector_text[:35]}"
    elif is_general:
        label_pdf = "📄  Export General PDF report"
        label_xl  = "📊  Export General Excel workbook"
    else:
        label_pdf = f"📄  Export PDF — {client}"
        label_xl  = f"📊  Export Excel — {client}"

    menu = tk.Menu(self._lu_analysis_frame, tearoff=0,
                   font=F(9), bg=_WHITE, fg=_TXT_NAVY,
                   activebackground=_NAVY_GHOST, activeforeground=_NAVY_DEEP,
                   relief="flat", bd=1)
    menu.add_command(label=label_pdf, command=lambda: _export_pdf(self))
    menu.add_command(label=label_xl,  command=lambda: _export_excel(self))
    menu.add_separator()
    menu.add_command(label="🖨  Print report", command=lambda: _report_print(self))
    try:
        menu.tk_popup(
            self._lu_export_btn.winfo_rootx(),
            self._lu_export_btn.winfo_rooty() + self._lu_export_btn.winfo_height())
    finally:
        menu.grab_release()


def _rgb_hex(color) -> str:
    try:
        return f"{int(color.red*255):02X}{int(color.green*255):02X}{int(color.blue*255):02X}"
    except Exception:
        return "000000"


def _export_pdf(self):
    if not self._lu_all_data:
        messagebox.showwarning("No Data","Run an analysis first."); return
    if not _HAS_RL:
        messagebox.showerror("Missing Library",
                             "reportlab is not installed.\nRun:  pip install reportlab"); return
    client     = self._lu_active_client
    is_general = (client == GENERAL_CLIENT)
    active_sectors = _lu_get_active_sectors(self)

    if active_sectors:
        sector_slug = "_".join(s.replace("/","_").replace(" ","_") for s in active_sectors)
        default_name = f"LU_Risk_Sector_{sector_slug}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    elif is_general:
        default_name = f"LU_Risk_General_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    else:
        default_name = f"LU_Risk_{client.replace(' ','_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"

    path = filedialog.asksaveasfilename(
        title="Save PDF Report", defaultextension=".pdf",
        filetypes=[("PDF files","*.pdf"),("All files","*.*")],
        initialfile=default_name)
    if not path: return
    try:
        # Use filtered results for PDF
        filtered_data = _lu_get_filtered_all_data(self)
        if is_general or active_sectors:
            results = filtered_data.get("general", [])
            cn = None
        else:
            all_data = self._lu_all_data
            results = [all_data["clients"][client]] if client in all_data.get("clients",{}) else []
            cn = client
        _generate_pdf(results, path,
                      filepath=self._lu_filepath or "",
                      client_name=cn,
                      sector_filter=active_sectors)
        messagebox.showinfo("Export Complete", f"PDF saved to:\n{path}")
    except Exception as ex:
        messagebox.showerror("PDF Export Error", str(ex))


def _generate_pdf(results, out_path, filepath="", client_name=None, sector_filter=None):
    styles  = getSampleStyleSheet()
    doc     = SimpleDocTemplate(out_path, pagesize=A4,
                                leftMargin=2*cm, rightMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
    navy    = rl_colors.HexColor("#1A3A6B")
    red     = rl_colors.HexColor("#E53E3E")
    gold    = rl_colors.HexColor("#D4A017")
    green   = rl_colors.HexColor("#2E7D32")
    white   = rl_colors.white
    off     = rl_colors.HexColor("#F5F7FA")
    border  = rl_colors.HexColor("#C5D0E8")
    crimson = rl_colors.HexColor("#B71C1C")

    title_style  = ParagraphStyle("LUTitle",  parent=styles["Title"],    fontSize=18, textColor=navy,  spaceAfter=4)
    h1_style     = ParagraphStyle("LUH1",     parent=styles["Heading1"], fontSize=13, textColor=white, spaceAfter=4)
    body_style   = ParagraphStyle("LUBody",   parent=styles["Normal"],   fontSize=8,  leading=11, textColor=rl_colors.HexColor("#1A2B4A"))
    muted_style  = ParagraphStyle("LUMuted",  parent=styles["Normal"],   fontSize=7,  leading=10, textColor=rl_colors.HexColor("#9AAACE"))
    client_style = ParagraphStyle("LUClient", parent=styles["Normal"],   fontSize=13, leading=16, textColor=navy, spaceAfter=2)
    mode_style   = ParagraphStyle("LUMode",   parent=styles["Normal"],   fontSize=9,  leading=12, textColor=rl_colors.HexColor("#5A9E28"))

    RISK_COLOR_RL  = {"HIGH":red,"MODERATE":gold,"LOW":green}
    SCORE_COLOR_RL = {"CRITICAL":crimson,"HIGH":red,"MODERATE":gold,"LOW":green}

    story = [Paragraph("LU Risk Analysis Report", title_style)]
    now   = datetime.now().strftime("%B %d, %Y  %H:%M")
    fname = Path(filepath).name if filepath else "—"

    if sector_filter:
        sector_text = " · ".join(sector_filter)
        story.append(Paragraph(f"SECTOR FILTER — {sector_text}", mode_style))
        story.append(Paragraph(f"Clients in filter: {len(results)}", body_style))
    elif client_name and results:
        rec = results[0]
        story.append(Paragraph("PER-CLIENT VIEW", mode_style))
        story.append(Paragraph(f"<b>Client: {client_name}</b>", client_style))
        story.append(Paragraph(f"Industry: {rec.get('industry','—')}  |  Sector: {rec.get('sector','—')}", body_style))
        story.append(Paragraph(
            f"Loan Balance: ₱{rec.get('loan_balance') or 0:,.2f}   Net Income: ₱{rec.get('net_income') or 0:,.2f}",
            body_style))
        score, label, fg_hex, _ = _compute_risk_score(rec.get("expenses",[]))
        sc = SCORE_COLOR_RL.get(label, green)
        story.append(Paragraph(
            f"Risk Score: <font color='#{_rgb_hex(sc)}'><b>{label} ({score:.2f})</b></font>",
            body_style))
    else:
        story.append(Paragraph("GENERAL VIEW — All Clients", mode_style))

    story += [
        Paragraph(f"File: {fname}    Generated: {now}", muted_style),
        Spacer(1, 0.3*cm),
        HRFlowable(width="100%", thickness=1, color=border),
        Spacer(1, 0.4*cm),
    ]

    for rec in results:
        story.append(PageBreak())
        icon   = _SECTOR_ICON.get(rec.get("sector",""),"")
        client_lbl = rec.get("client","—")
        hdr_tbl = Table([[Paragraph(f"{icon}  {client_lbl}", h1_style)]],
                        colWidths=[17*cm])
        hdr_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1),navy),
            ("TOPPADDING",(0,0),(-1,-1),8), ("BOTTOMPADDING",(0,0),(-1,-1),8),
            ("LEFTPADDING",(0,0),(-1,-1),10),
        ]))
        story += [hdr_tbl, Spacer(1,0.1*cm)]

        story.append(Paragraph(
            f"Industry: {rec.get('industry','—')}  |  Sector: {rec.get('sector','—')}  |  "
            f"ID: {rec.get('client_id','—')}  PN: {rec.get('pn','—')}",
            muted_style))
        story.append(Paragraph(
            f"Total Source: ₱{rec.get('total_source') or 0:,.2f}   "
            f"Net Income: ₱{rec.get('net_income') or 0:,.2f}   "
            f"Loan Balance: ₱{rec.get('loan_balance') or 0:,.2f}   "
            f"Current Amort: ₱{rec.get('current_amort') or 0:,.2f}",
            body_style))
        story.append(Spacer(1,0.2*cm))

        col_w    = [4.5*cm,2.2*cm,6.3*cm,4*cm]
        tbl_data = [[Paragraph(f"<b>{h}</b>", body_style)
                     for h in ["Expense Item","Risk","Impact Reason","Value / Amount"]]]
        row_bgs  = []
        for exp in rec.get("expenses",[]):
            rc = RISK_COLOR_RL.get(exp["risk"],green)
            tbl_data.append([
                Paragraph(f"<b>{exp['name']}</b>", body_style),
                Paragraph(f"<font color='#{_rgb_hex(rc)}'><b>{exp['risk']}</b></font>", body_style),
                Paragraph(exp["reason"], body_style),
                Paragraph(exp["value_str"], muted_style if not exp["has_values"] else body_style),
            ])
            row_bgs.append({"HIGH":rl_colors.HexColor("#FFF5F5"),
                            "MODERATE":rl_colors.HexColor("#FFFBF0"),
                            "LOW":rl_colors.white}.get(exp["risk"],rl_colors.white))

        tbl_style = [
            ("BACKGROUND",(0,0),(-1,0),off), ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),8), ("LEADING",(0,0),(-1,-1),11),
            ("BOX",(0,0),(-1,-1),0.5,border), ("INNERGRID",(0,0),(-1,-1),0.3,border),
            ("TOPPADDING",(0,0),(-1,-1),5), ("BOTTOMPADDING",(0,0),(-1,-1),5),
            ("LEFTPADDING",(0,0),(-1,-1),6), ("VALIGN",(0,0),(-1,-1),"TOP"),
        ]
        for i, bg in enumerate(row_bgs):
            tbl_style.append(("BACKGROUND",(0,i+1),(-1,i+1),bg))

        exp_tbl = Table(tbl_data, colWidths=col_w, repeatRows=1)
        exp_tbl.setStyle(TableStyle(tbl_style))
        story += [exp_tbl, Spacer(1,0.4*cm)]

    doc.build(story)


def _export_excel(self):
    if not self._lu_all_data:
        messagebox.showwarning("No Data","Run an analysis first."); return
    if not _HAS_OPENPYXL:
        messagebox.showerror("Missing Library","openpyxl is not installed."); return
    client     = self._lu_active_client
    is_general = (client == GENERAL_CLIENT)
    active_sectors = _lu_get_active_sectors(self)

    if active_sectors:
        sector_slug = "_".join(s.replace("/","_").replace(" ","_") for s in active_sectors)
        default_name = f"LU_Risk_Sector_{sector_slug}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    elif is_general:
        default_name = f"LU_Risk_General_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    else:
        default_name = f"LU_Risk_{client.replace(' ','_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    path = filedialog.asksaveasfilename(
        title="Save Excel Workbook", defaultextension=".xlsx",
        filetypes=[("Excel files","*.xlsx"),("All files","*.*")],
        initialfile=default_name)
    if not path: return
    try:
        filtered_data = _lu_get_filtered_all_data(self)
        if is_general or active_sectors:
            results = filtered_data.get("general", [])
            cn = None
        else:
            all_data = self._lu_all_data
            results = [all_data["clients"][client]] if client in all_data.get("clients",{}) else []
            cn = client
        _generate_excel(results, path,
                        client_name=cn,
                        all_data=filtered_data,
                        sector_filter=active_sectors)
        messagebox.showinfo("Export Complete", f"Excel saved to:\n{path}")
    except Exception as ex:
        messagebox.showerror("Excel Export Error", str(ex))


def _generate_excel(results, out_path, client_name=None, all_data=None, sector_filter=None):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def fill(hex_col):
        return PatternFill("solid", fgColor=hex_col.lstrip("#"))
    def border_thin():
        s = Side(style="thin", color="C5D0E8")
        return Border(left=s, right=s, top=s, bottom=s)

    FILLS = {
        "header": fill("#1A3A6B"), "sector": fill("#1E4080"),
        "HIGH":   fill("#FFF5F5"), "MODERATE": fill("#FFFBF0"),
        "LOW":    fill("#F0FBE8"), "col_hdr": fill("#F5F7FA"),
    }
    RISK_FONT_COLOR = {"HIGH":"E53E3E","MODERATE":"D4A017","LOW":"2E7D32"}

    ws_sum = wb.create_sheet("Summary")
    for col, w in [(1,30),(2,40)]:
        ws_sum.column_dimensions[get_column_letter(col)].width = w

    ws_sum["A1"] = "LU Risk Analysis Report"
    ws_sum["A1"].font = Font(bold=True, size=14, color="0A1628")

    row = 2
    if sector_filter:
        mode_txt = f"Sector Filter — {' · '.join(sector_filter)}"
    elif client_name:
        mode_txt = "Per-Client"
    else:
        mode_txt = "General View — All Clients"

    ws_sum.cell(row,1,"Mode").font = Font(bold=True,size=9,color="5A9E28")
    ws_sum.cell(row,2,mode_txt).font = Font(size=9,color="5A9E28")
    row += 1
    ws_sum.cell(row,1,f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(size=9,color="9AAACE")
    row += 2

    for col_idx, txt in enumerate(["Metric","Value"],1):
        c = ws_sum.cell(row,col_idx,txt)
        c.fill = FILLS["header"]; c.font = Font(bold=True,color="FFFFFF",size=10)
        c.alignment = Alignment(horizontal="center",vertical="center")
        c.border = border_thin()
    row += 1

    totals = (all_data or {}).get("totals",{})
    for lbl, val in [
        ("Clients", len(results)),
        ("Total Loan Balance", f"₱{totals.get('loan_balance',0):,.2f}"),
        ("Total Net Income",   f"₱{totals.get('total_net',0):,.2f}"),
    ]:
        ws_sum.cell(row,1,lbl).border = border_thin()
        ws_sum.cell(row,2,val).border = border_thin()
        row += 1

    ws_data = wb.create_sheet("Client Data")
    hdrs = ["Client ID","PN","Applicant","Industry","Sector","Residence",
            "Source Income","Total Source","Biz Expenses (Total)","Hhld Expenses (Total)",
            "Net Income","Amort History","Current Amort","Loan Balance","Risk Score","Risk Label"]
    for ci, h in enumerate(hdrs,1):
        c = ws_data.cell(1,ci,h)
        c.fill = FILLS["header"]; c.font = Font(bold=True,color="FFFFFF",size=9)
        c.alignment = Alignment(horizontal="center",vertical="center")
        c.border = border_thin()
        ws_data.column_dimensions[get_column_letter(ci)].width = max(12,len(h)+4)

    for ri, rec in enumerate(results,2):
        rl = rec.get("score_label","N/A")
        fc = RISK_FONT_COLOR.get(rl,"2E7D32")
        row_fill = FILLS.get(rl, FILLS["LOW"])
        for ci, val in enumerate([
            rec.get("client_id",""), rec.get("pn",""),
            rec.get("client",""),   rec.get("industry",""),
            rec.get("sector",""),   rec.get("residence",""),
            rec.get("source_income","")[:200] if rec.get("source_income") else "",
            rec.get("total_source") or 0,
            rec.get("total_biz") or 0,
            rec.get("total_hhld") or 0,
            rec.get("net_income") or 0,
            rec.get("amort_history") or 0,
            rec.get("current_amort") or 0,
            rec.get("loan_balance") or 0,
            rec.get("score",0),
            rl,
        ],1):
            c = ws_data.cell(ri,ci,val)
            c.fill   = row_fill
            c.border = border_thin()
            c.font   = Font(size=9, color=fc if ci==16 else "1A2B4A",
                            bold=(ci==16))
            c.alignment = Alignment(vertical="top",
                                    horizontal="right" if isinstance(val,float) else "left",
                                    indent=1)

    for rec in results:
        safe_name = rec["client"][:28].replace("/","_").replace("\\","_")
        ws = wb.create_sheet(safe_name)
        for ci, w in enumerate([28,12,52,30],1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        ws.merge_cells("A1:D1")
        ws["A1"].value     = f"{rec['client']}  |  {rec.get('sector','—')}"
        ws["A1"].fill      = FILLS["sector"]
        ws["A1"].font      = Font(bold=True,color="FFFFFF",size=12)
        ws["A1"].alignment = Alignment(vertical="center",horizontal="left",indent=1)

        ws.merge_cells("A2:D2")
        ws["A2"].value = (f"Loan Balance: ₱{rec.get('loan_balance') or 0:,.2f}   "
                          f"Net Income: ₱{rec.get('net_income') or 0:,.2f}   "
                          f"Risk: {rec.get('score_label','N/A')}")
        ws["A2"].fill  = fill("#EEF3FB")
        ws["A2"].font  = Font(size=9,color="1A3A6B")
        ws["A2"].alignment = Alignment(vertical="center",horizontal="left",indent=1)

        for ci, hdr_text in enumerate(["Expense Item","Risk Level","Impact Reason","Value / Amount"],1):
            c = ws.cell(3,ci,hdr_text)
            c.fill = FILLS["col_hdr"]; c.font = Font(bold=True,size=9,color="4A6FA5")
            c.alignment = Alignment(vertical="center",horizontal="left",indent=1)
            c.border = border_thin()

        for idx, exp in enumerate(rec.get("expenses",[]),start=4):
            row_fill = FILLS.get(exp["risk"],FILLS["LOW"])
            risk_col = RISK_FONT_COLOR.get(exp["risk"],"2E7D32")
            for ci, (val,bold,color,halign,wrap) in enumerate([
                (exp["name"],True,"1A2B4A","left",False),
                (exp["risk"],True,risk_col,"center",False),
                (exp["reason"],False,"6B7FA3","left",True),
                (exp["value_str"],False,
                 "9AAACE" if not exp["has_values"] else "1A2B4A","left",True),
            ],1):
                c = ws.cell(idx,ci,val)
                c.fill = row_fill
                c.font = Font(bold=bold,size=9 if ci<3 else 8,color=color,
                              italic=(ci==4 and not exp["has_values"]))
                c.alignment = Alignment(vertical="top",horizontal=halign,
                                        wrap_text=wrap,indent=1)
                c.border = border_thin()
            ws.row_dimensions[idx].height = 30

    wb.save(out_path)


# ══════════════════════════════════════════════════════════════════════
#  RISK SIMULATOR  (patched: uses filtered data)
# ══════════════════════════════════════════════════════════════════════

def _build_simulator_panel(self, parent):
    hdr = tk.Frame(parent, bg=_NAVY_MID, height=38)
    hdr.pack(fill="x")
    hdr.pack_propagate(False)
    self._sim_hdr_lbl = tk.Label(
        hdr, text="⚙️  Inflation / Cost-Shock Simulator",
        font=F(10,"bold"), fg=_WHITE, bg=_NAVY_MID)
    self._sim_hdr_lbl.pack(side="left", padx=20, pady=8)

    ctrl = tk.Frame(parent, bg=_OFF_WHITE, height=46)
    ctrl.pack(fill="x")
    ctrl.pack_propagate(False)
    tk.Label(ctrl, text="Global %:", font=F(9,"bold"),
             fg=_NAVY_MID, bg=_OFF_WHITE).pack(side="left", padx=(16,4), pady=12)
    self._sim_global_var = tk.StringVar(value="0")
    ctk.CTkEntry(ctrl, textvariable=self._sim_global_var, width=70, height=26,
                 corner_radius=4, fg_color=_WHITE, text_color=_TXT_NAVY,
                 border_color=_BORDER_MID, font=FF(9)
                 ).pack(side="left", pady=10)
    ctk.CTkButton(ctrl, text="Apply All", command=lambda: _sim_apply_global(self),
                  width=80, height=26, corner_radius=4,
                  fg_color=_NAVY_LIGHT, hover_color=_NAVY_MID,
                  text_color=_WHITE, font=FF(8,"bold")
                  ).pack(side="left", padx=6, pady=10)
    ctk.CTkButton(ctrl, text="Reset", command=lambda: _sim_reset(self),
                  width=70, height=26, corner_radius=4,
                  fg_color=_ACCENT_RED, hover_color="#C53030",
                  text_color=_WHITE, font=FF(8,"bold")
                  ).pack(side="left", padx=(0,6), pady=10)
    tk.Frame(parent, bg=_BORDER_LIGHT, height=1).pack(fill="x")

    cards_frame = tk.Frame(parent, bg=_NAVY_MIST)
    cards_frame.pack(fill="x")
    _build_sim_summary_cards(self, cards_frame)

    inc_bar = tk.Frame(parent, bg=_NAVY_DEEP, height=38)
    inc_bar.pack(fill="x")
    inc_bar.pack_propagate(False)
    self._sim_income_lbl = tk.Label(
        inc_bar, text="TOTAL SOURCE OF INCOME  —  Load a file to begin",
        font=F(9,"bold"), fg=_TXT_MUTED, bg=_NAVY_DEEP)
    self._sim_income_lbl.pack(side="left", padx=20, pady=10)
    self._sim_surplus_lbl = tk.Label(inc_bar, text="", font=F(9,"bold"),
                                     fg=_LIME_MID, bg=_NAVY_DEEP)
    self._sim_surplus_lbl.pack(side="right", padx=20, pady=10)
    tk.Frame(parent, bg=_BORDER_LIGHT, height=1).pack(fill="x")

    split = tk.Frame(parent, bg=_CARD_WHITE)
    split.pack(fill="both", expand=True)

    left_frame = tk.Frame(split, bg=_CARD_WHITE)
    left_frame.pack(side="left", fill="both", expand=True)

    right_frame = tk.Frame(split, bg=_CARD_WHITE,
                           highlightbackground=_BORDER_MID, highlightthickness=1,
                           width=280)
    right_frame.pack(side="right", fill="y")
    right_frame.pack_propagate(False)
    tk.Label(right_frame, text="Expense Chart", font=F(8,"bold"),
             fg=_TXT_SOFT, bg=_CARD_WHITE).pack(pady=(8,0))
    self._sim_chart_canvas = tk.Canvas(right_frame, bg=_CARD_WHITE, width=268,
                                       highlightthickness=0)
    self._sim_chart_canvas.pack(fill="both", expand=True, padx=4, pady=4)

    # PATCH: replace scrollable frame with plain Frame
    self._sim_scroll_frame = tk.Frame(left_frame, bg=_CARD_WHITE)
    self._sim_scroll_frame.pack(fill="both", expand=True)

    self._sim_expenses   = []
    self._sim_sliders    = {}
    self._sim_net_income = 0.0
    _sim_show_placeholder(self)


def _build_sim_summary_cards(self, parent):
    for title, attr, color in [
        ("Total Source of Income", "_sim_lbl_income",  _ACCENT_SUCCESS),
        ("Base Total Expenses",    "_sim_lbl_base",    _TXT_NAVY),
        ("Simulated Total",        "_sim_lbl_sim",     _TXT_NAVY),
        ("Total Increase (₱)",     "_sim_lbl_inc",     _ACCENT_RED),
        ("Surplus / Deficit",      "_sim_lbl_surplus", _ACCENT_SUCCESS),
    ]:
        card = tk.Frame(parent, bg=_NAVY_MIST,
                        highlightbackground=_NAVY_GHOST, highlightthickness=1)
        card.pack(side="left", fill="x", expand=True, padx=6, pady=8)
        tk.Label(card, text=title, font=F(7), fg=_TXT_SOFT, bg=_NAVY_MIST).pack(anchor="w", padx=10, pady=(6,0))
        lbl = tk.Label(card, text="—", font=F(13,"bold"), fg=color, bg=_NAVY_MIST)
        lbl.pack(anchor="w", padx=10, pady=(0,6))
        setattr(self, attr, lbl)


def _sim_show_placeholder(self):
    for w in self._sim_scroll_frame.winfo_children(): w.destroy()
    tk.Label(self._sim_scroll_frame,
             text="Run an analysis first to unlock the simulator.",
             font=F(10), fg=_TXT_MUTED, bg=_CARD_WHITE).pack(pady=60)
    _sim_draw_chart(self)


def _sim_populate(self):
    if not hasattr(self, '_sim_hdr_lbl') or not self._sim_hdr_lbl.winfo_exists():
        return
    # Use filtered data so simulator respects sector filter
    filtered_data    = _lu_get_filtered_all_data(self)
    active_sectors   = _lu_get_active_sectors(self)
    client           = self._lu_active_client
    is_general       = (client == GENERAL_CLIENT)
    all_clients_dict = filtered_data.get("clients", {})
    all_clients_list = list(all_clients_dict.values())

    if is_general or active_sectors:
        recs = all_clients_list
    else:
        recs = [all_clients_dict[client]] if client in all_clients_dict else []
    print(f"[SIM DEBUG] active_sectors={active_sectors}")
    print(f"[SIM DEBUG] recs count={len(recs)}")
    print(f"[SIM DEBUG] first rec expenses={recs[0].get('expenses', [])[:2] if recs else 'EMPTY'}")
    # Update simulator header
    if active_sectors:
        self._sim_hdr_lbl.config(
            text=f"⚙️  Simulator — Filtered: {' · '.join(active_sectors)}",
            fg=_LIME_MID)
    else:
        self._sim_hdr_lbl.config(
            text="⚙️  Inflation / Cost-Shock Simulator",
            fg=_WHITE)

    net_income = sum((r.get("total_source") or 0) for r in recs)
    self._sim_net_income = net_income

    accumulated: dict[str,dict] = {}
    for rec in recs:
        for exp in rec.get("expenses",[]):
            if exp["total"] <= 0: continue
            name = exp["name"]
            if name not in accumulated:
                accumulated[name] = dict(exp)
            else:
                accumulated[name]["total"] += exp["total"]
                if _RISK_ORDER.get(exp["risk"],9) < _RISK_ORDER.get(accumulated[name]["risk"],9):
                    accumulated[name]["risk"]   = exp["risk"]
                    accumulated[name]["reason"] = exp["reason"]
                accumulated[name]["value_str"] = _fmt_value([accumulated[name]["total"]])

    self._sim_expenses = list(accumulated.values())
    self._sim_sliders  = {}

    for w in list(self._sim_scroll_frame.winfo_children()):
        try: w.destroy()
        except: pass

    if not self._sim_expenses:
        tk.Label(self._sim_scroll_frame,
                 text="No numeric expense data found.",
                 font=F(9), fg=_TXT_MUTED, bg=_CARD_WHITE, justify="center").pack(pady=60)
        return

    hdr = tk.Frame(self._sim_scroll_frame, bg=_OFF_WHITE)
    hdr.pack(fill="x", pady=(8,0))
    for col, text, w in [
        (0,"Expense Item",220),(1,"Risk",60),(2,"Base Amount",120),
        (3,"Inflation Rate (%)",80),(4,"Extra Cost",120),(5,"Simulated",120)
    ]:
        tk.Label(hdr, text=text, font=F(8,"bold"), fg=_NAVY_PALE, bg=_OFF_WHITE,
                 width=w//8, anchor="w", padx=6, pady=5
                 ).grid(row=0, column=col, sticky="ew", padx=(0,2))

    tk.Frame(self._sim_scroll_frame, bg=_BORDER_MID, height=1).pack(fill="x")

    for idx, exp in enumerate(self._sim_expenses):
        try:
            var = tk.StringVar(value="0")
            self._sim_sliders[exp["name"]] = var
            _sim_build_expense_row(self, self._sim_scroll_frame, exp, var, idx)
        except Exception:
            continue

    _sim_refresh(self)


def _sim_build_expense_row(self, parent, exp, var, idx):
    risk   = exp["risk"]
    row_bg = _RISK_BG.get(risk, _WHITE) if idx % 2 == 0 else _WHITE
    row    = tk.Frame(parent, bg=row_bg)
    row.pack(fill="x")

    tk.Label(row, text=exp["name"], font=F(9,"bold"),
             fg=_TXT_NAVY, bg=row_bg, anchor="w", padx=8, pady=6, width=26
             ).grid(row=0, column=0, sticky="ew")
    tk.Label(row, text=risk[:3], font=F(7,"bold"),
             fg=_RISK_COLOR.get(risk,_TXT_SOFT), bg=_RISK_BADGE_BG.get(risk,_OFF_WHITE),
             padx=4, pady=3).grid(row=0, column=1, padx=4, pady=6)
    tk.Label(row, text=f"₱{exp['total']:,.2f}" if exp["total"]>0 else "—",
             font=F(9), fg=_TXT_NAVY, bg=row_bg, anchor="e", padx=6, width=14
             ).grid(row=0, column=2, sticky="ew")
    rate_entry = ctk.CTkEntry(row, textvariable=var, width=80, height=26, corner_radius=4,
                              font=FF(9), fg_color=_WHITE, text_color=_TXT_NAVY,
                              border_color=_RISK_COLOR.get(risk, _BORDER_MID),
                              placeholder_text="0")
    rate_entry.grid(row=0, column=3, padx=8, pady=6)
    rate_entry.bind("<Return>",   lambda e, ex=exp: _sim_on_slide(self, ex, var.get()))
    rate_entry.bind("<FocusOut>", lambda e, ex=exp: _sim_on_slide(self, ex, var.get()))
    extra_lbl = tk.Label(row, text="—", font=F(9), fg=_ACCENT_RED,
                         bg=row_bg, anchor="e", padx=6, width=14)
    extra_lbl.grid(row=0, column=4, sticky="ew")
    sim_lbl = tk.Label(row, text="—", font=F(9,"bold"), fg=_TXT_NAVY,
                       bg=row_bg, anchor="e", padx=6, width=14)
    sim_lbl.grid(row=0, column=5, sticky="ew")

    var._extra_lbl = extra_lbl
    var._sim_lbl   = sim_lbl
    var._base      = exp["total"]
    tk.Frame(parent, bg=_BORDER_LIGHT, height=1).pack(fill="x")


def _sim_on_slide(self, exp, value):
    try:    pct = float(value)
    except: pct = 0.0
    # PATCH: allow any non-negative value (no upper clamp)
    if pct < 0.0:
        pct = 0.0
    self._sim_sliders[exp["name"]].set(str(pct))
    _sim_refresh(self)


def _sim_apply_global(self):
    try:    pct = float(self._sim_global_var.get())
    except: pct = 0.0
    # PATCH: allow any non-negative value (no upper clamp)
    if pct < 0.0:
        pct = 0.0
    if not self._sim_sliders and self._lu_all_data: _sim_populate(self)
    for var in self._sim_sliders.values(): var.set(str(pct))
    _sim_refresh(self)


def _sim_reset(self):
    self._sim_global_var.set("0")
    if not self._sim_sliders and self._lu_all_data: _sim_populate(self)
    for var in self._sim_sliders.values(): var.set("0")
    _sim_refresh(self)


def _sim_refresh(self):
    base_total = sim_total = 0.0
    for exp in getattr(self, "_sim_expenses", []):
        pct = 0.0
        var = self._sim_sliders.get(exp["name"])
        if var:
            try: pct = float(var.get())
            except: pass
        base  = exp["total"]
        extra = base * pct / 100.0
        sim   = base + extra
        base_total += base
        sim_total  += sim
        if var and hasattr(var,"_extra_lbl"):
            try:
                var._extra_lbl.config(
                    text=f"+₱{extra:,.2f}" if extra>0 else "—",
                    fg=_ACCENT_RED if extra>0 else _TXT_MUTED)
                var._sim_lbl.config(
                    text=f"₱{sim:,.2f}" if base>0 else "—", fg=_TXT_NAVY)
            except: pass

    increase   = sim_total - base_total
    net_income = getattr(self, "_sim_net_income", 0.0) or 0.0
    surplus    = net_income - sim_total

    if hasattr(self, "_sim_lbl_base"):
        try:
            self._sim_lbl_income.config(
                text=f"₱{net_income:,.2f}" if net_income else "—",
                fg=_ACCENT_SUCCESS)
            self._sim_lbl_base.config(
                text=f"₱{base_total:,.2f}" if base_total else "—")
            self._sim_lbl_sim.config(
                text=f"₱{sim_total:,.2f}" if base_total else "—")
            self._sim_lbl_inc.config(
                text=f"+₱{increase:,.2f}" if increase>0 else "₱0.00",
                fg=_ACCENT_RED if increase>0 else _TXT_NAVY)
            if net_income:
                surplus_txt = f"{'▲' if surplus>=0 else '▼'} ₱{abs(surplus):,.2f}"
                self._sim_lbl_surplus.config(
                    text=surplus_txt,
                    fg=_ACCENT_SUCCESS if surplus>=0 else _ACCENT_RED)
            else:
                self._sim_lbl_surplus.config(text="—", fg=_TXT_MUTED)
        except: pass

    if hasattr(self, "_sim_income_lbl"):
        try:
            if net_income:
                self._sim_income_lbl.config(
                    text=f"TOTAL SOURCE OF INCOME  ₱{net_income:,.2f}",
                    fg=_LIME_MID)
                self._sim_surplus_lbl.config(
                    text=(f"SURPLUS  ₱{surplus:,.2f}" if surplus >= 0
                          else f"DEFICIT  ▲ ₱{abs(surplus):,.2f}"),
                    fg=_LIME_MID if surplus >= 0 else _ACCENT_RED)
            else:
                self._sim_income_lbl.config(
                    text="TOTAL SOURCE OF INCOME  —  Load a file to begin",
                    fg=_TXT_MUTED)
                self._sim_surplus_lbl.config(text="", fg=_LIME_MID)
        except: pass

    _sim_draw_chart(self)


def _sim_draw_chart(self):
    c = getattr(self,"_sim_chart_canvas",None)
    if c is None: return
    try:
        if not c.winfo_exists(): return
        c.delete("all")
    except: return

    sim_expenses = getattr(self, "_sim_expenses", None)
    if not sim_expenses:
        return
    expenses = [e for e in sim_expenses if e["total"] > 0]
    if not expenses:
        try:
            c.create_text(134,100, text="No numeric data\nto chart.",
                          fill=_TXT_MUTED, font=F(9), justify="center")
        except: pass
        return

    MAX_BARS = 20
    expenses = expenses[:MAX_BARS]
    try:
        c.update_idletasks(); W = c.winfo_width()
    except: W = 268
    if W < 10: W = 268

    bar_h, gap, label_h = 14, 8, 12
    n = len(expenses)
    margin_top = 10
    H = margin_top + n * (bar_h*2 + gap + label_h) + 10
    try: c.config(height=H)
    except: pass

    margin_left = margin_right = 10
    bar_area_w  = max(W - margin_left - margin_right, 50)
    row_h  = bar_h*2 + gap + label_h
    half_h = bar_h

    try:
        max_val = max(
            e["total"] + e["total"] * (float(self._sim_sliders.get(e["name"],tk.StringVar(value="0")).get() or 0)/100)
            for e in expenses)
    except: max_val = 1
    if not max_val or max_val <= 0: max_val = 1

    for i, exp in enumerate(expenses):
        pct = 0.0
        var = self._sim_sliders.get(exp["name"])
        if var:
            try: pct = float(var.get())
            except: pass
        base = exp["total"]
        sim  = base + base * pct / 100.0
        y_mid = margin_top + i * row_h + row_h // 2
        try:
            bw = int(bar_area_w * (base / max_val))
            c.create_rectangle(margin_left, y_mid-half_h,
                               margin_left+bw, y_mid,
                               fill=_SIM_BAR_BASE, outline="")
            sw = int(bar_area_w * (sim / max_val))
            c.create_rectangle(margin_left, y_mid,
                               margin_left+sw, y_mid+half_h,
                               fill=_SIM_BAR_SIM, outline="")
            short = exp["name"] if len(exp["name"])<=16 else exp["name"][:15]+"…"
            c.create_text(margin_left+4, y_mid-half_h-1,
                          text=short, anchor="sw",
                          font=("Segoe UI",7), fill=_TXT_SOFT)
        except: continue
    try: c.update_idletasks()
    except: pass


# ══════════════════════════════════════════════════════════════════════
#  PLACEHOLDER + FILE BROWSING + RUN
# ══════════════════════════════════════════════════════════════════════

def _lu_show_placeholder(self):
    for w in self._lu_results_frame.winfo_children(): w.destroy()
    ph = tk.Frame(self._lu_results_frame, bg=_CARD_WHITE)
    ph.pack(expand=True, fill="both", pady=60)
    tk.Label(ph, text="📊", font=("Segoe UI Emoji",40),
             bg=_CARD_WHITE, fg=_TXT_MUTED).pack()
    tk.Label(ph, text="No analysis yet",
             font=F(14,"bold"), fg=_TXT_SOFT, bg=_CARD_WHITE).pack(pady=(8,4))
    tk.Label(ph,
             text=("Load a Look-Up Summary Excel file to scan all clients.\n\n"
                   "Detects sectors:  Wholesale/Retail · Agriculture (Fishing & Forestry) · "
                   "Transport · Remittance · Consumer Loan\n\n"
                   "Reads columns:  Client ID · PN · Applicant · Industry Name · "
                   "Source of Income · Business Expenses · Household Expenses · "
                   "Net Income · Loan Balance\n\n"
                   "Search by client name, ID, PN, or sector name.\n"
                   "Type a sector name (e.g. 'Transport') to filter all tabs by that sector."),
             font=F(9), fg=_TXT_MUTED, bg=_CARD_WHITE, justify="center").pack()


def _lu_browse_file(self):
    path = filedialog.askopenfilename(
        title="Select Excel File for LU Analysis",
        filetypes=[("Excel files","*.xlsx *.xlsm *.xls"),("All files","*.*")])
    if not path: return
    self._lu_filepath = path
    self._lu_file_lbl.config(text=f"📊  {Path(path).name}", fg=_TXT_NAVY)
    self._lu_rescan_btn.configure(state="normal")
    _lu_run_analysis(self)


def _lu_run_analysis(self):
    if not self._lu_filepath: return
    self._lu_status_lbl.config(text="⏳  Scanning…", fg=_ACCENT_GOLD)
    self._lu_load_btn.configure(state="disabled")
    self._lu_rescan_btn.configure(state="disabled")
    self._lu_export_btn.configure(state="disabled")
    self.update_idletasks()
    try:
        all_data = run_lu_analysis(self._lu_filepath)
        self._lu_all_data = all_data
        # Reset filter state on new scan
        self._lu_filtered_sectors = None
        _lu_update_filter_pill(self)
        _lu_populate_client_dropdown(self)

        general = all_data.get("general",[])
        self._lu_results = general
        _lu_render_results(self, general)

        n_clients  = len(all_data.get("clients",{}))
        n_sectors  = len(all_data.get("sector_map",{}))
        totals     = all_data.get("totals",{})
        self._lu_status_lbl.config(
            text=(f"✅  {n_clients} client(s) · {n_sectors} sector(s) · "
                  f"₱{totals.get('loan_balance',0):,.0f} total loan balance"),
            fg=_LIME_DARK)
        self._lu_export_btn.configure(state="normal")
        if getattr(self, "_loanbal_export_btn", None) is not None:
            self._loanbal_export_btn.configure(state="normal")
        _sim_populate(self)
    except Exception as exc:
        _lu_show_error(self, str(exc))
        self._lu_status_lbl.config(text="❌  Error during scan", fg=_ACCENT_RED)
    finally:
        self._lu_load_btn.configure(state="normal")
        self._lu_rescan_btn.configure(state="normal")


def _lu_show_error(self, msg):
    for w in self._lu_results_frame.winfo_children(): w.destroy()
    err = tk.Frame(self._lu_results_frame, bg=_CARD_WHITE)
    err.pack(expand=True, fill="both", pady=60)
    tk.Label(err, text="❌", font=("Segoe UI Emoji",32), bg=_CARD_WHITE).pack()
    tk.Label(err, text="Analysis failed",
             font=F(13,"bold"), fg=_ACCENT_RED, bg=_CARD_WHITE).pack(pady=(8,4))
    tk.Label(err, text=msg, font=F(9), fg=_TXT_SOFT,
             bg=_CARD_WHITE, wraplength=500, justify="center").pack()


# ══════════════════════════════════════════════════════════════════════
#  ATTACH
# ══════════════════════════════════════════════════════════════════════

def attach(cls):
    """Attach all LU Analysis methods to DocExtractorApp."""
    cls._build_lu_analysis_panel    = _build_lu_analysis_panel
    cls._lu_show_placeholder         = _lu_show_placeholder
    cls._lu_browse_file              = _lu_browse_file
    cls._lu_run_analysis             = _lu_run_analysis
    cls._lu_show_error               = _lu_show_error
    cls._lu_render_results           = _lu_render_results
    cls._lu_render_general_view      = _lu_render_general_view
    cls._lu_render_client_view       = _lu_render_client_view
    cls._lu_render_client_card       = _lu_render_client_card
    cls._lu_render_sector_card       = _lu_render_sector_card
    cls._lu_on_client_change         = _lu_on_client_change
    cls._lu_filter_by_search         = _lu_filter_by_search
    cls._lu_populate_client_dropdown = _lu_populate_client_dropdown
    cls._lu_get_active_sectors       = _lu_get_active_sectors
    cls._lu_get_filtered_all_data    = _lu_get_filtered_all_data
    cls._lu_update_filter_pill       = _lu_update_filter_pill
    cls._lu_clear_sector_filter      = _lu_clear_sector_filter
    cls._build_charts_panel          = _build_charts_panel
    cls._charts_show_placeholder     = _charts_show_placeholder
    cls._charts_render               = _charts_render
    cls._build_loanbal_panel         = _build_loanbal_panel
    cls._loanbal_render              = _loanbal_render
    cls._loanbal_show_export_menu    = _loanbal_show_export_menu
    cls._loanbal_export_pdf          = _loanbal_export_pdf
    cls._loanbal_export_excel        = _loanbal_export_excel
    cls._generate_loanbal_pdf        = _generate_loanbal_pdf
    cls._generate_loanbal_excel      = _generate_loanbal_excel
    cls._build_report_panel          = _build_report_panel
    cls._report_show_placeholder     = _report_show_placeholder
    cls._report_render               = _report_render
    cls._report_print                = _report_print
    cls._lu_show_export_menu         = _lu_show_export_menu
    cls._lu_export_btn               = None
    cls._loanbal_export_btn          = None
    cls._build_simulator_panel       = _build_simulator_panel
    cls._build_sim_summary_cards     = _build_sim_summary_cards
    cls._lu_switch_view              = _lu_switch_view
    cls._sim_show_placeholder        = _sim_show_placeholder
    cls._sim_populate                = _sim_populate
    cls._sim_build_expense_row       = _sim_build_expense_row
    cls._sim_on_slide                = _sim_on_slide
    cls._sim_apply_global            = _sim_apply_global
    cls._sim_reset                   = _sim_reset
    cls._sim_refresh                 = _sim_refresh
    cls._sim_draw_chart              = _sim_draw_chart
    cls._export_pdf                  = _export_pdf
    cls._export_excel                = _export_excel
    cls._generate_pdf                = _generate_pdf
    cls._generate_excel              = _generate_excel
    cls._tv_sort                     = staticmethod(_tv_sort)