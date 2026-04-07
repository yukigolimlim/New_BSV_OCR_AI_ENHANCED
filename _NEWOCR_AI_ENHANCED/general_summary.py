"""
general_summary.py — DocExtract Pro
=====================================
"General Summary" tab: persistent database-backed view of ALL applicants
ever processed by the General Look-Up tab, across sessions.

Uses a SEPARATE database from summary_tab.py:
  Folder : general_lookup_summary_results/
  DB file: general_applicants.db

Compatibility: aligned with general_lookup.py (19-row LOOKUP_ROWS including
spouse, assets, and credit_history_amort fields).
"""

import re
import io
import csv
import json
import sqlite3
import threading
import tkinter as tk
import tkinter.ttk as ttk
import customtkinter as ctk
from pathlib import Path
from datetime import datetime
from tkinter import filedialog, messagebox

from app_constants import (
    NAVY_DEEP, NAVY_LIGHT, NAVY_MID, NAVY_PALE, NAVY_MIST, NAVY_GHOST,
    WHITE, OFF_WHITE, CARD_WHITE,
    LIME_BRIGHT, LIME_DARK, LIME_MID, LIME_PALE, LIME_MIST,
    TXT_NAVY, TXT_SOFT, TXT_MUTED, TXT_ON_LIME,
    ACCENT_RED, ACCENT_GOLD, ACCENT_SUCCESS,
    BORDER_LIGHT, BORDER_MID,
    SIDEBAR_BG, SIDEBAR_ITEM, SIDEBAR_HVR,
    F, FF, FMONO,
)

# ── Separate DB for General Look-Up (does NOT share with summary_tab) ─
DB_DIR  = Path(__file__).parent / "general_lookup_summary_results"
DB_PATH = DB_DIR / "general_applicants.db"

PAD          = 20
PAGE_SIZE    = 50
HDR_BG       = "#93C47D"
HDR_FG       = "#FFFFFF"
ROW_BG_EVEN  = "#F3F9F0"
ROW_BG_ODD   = WHITE
NET_GREEN    = "#1F6B28"
TOT_BG       = "#D9EAD3"
SEC_BG       = "#E8F0FA"
SEC_FG       = NAVY_MID

# Placeholder text shown in the search box when empty
_SEARCH_HINT = "Search all fields… separate terms with commas for AND filtering"

STATUS_COLORS = {
    "done":    ("#F0FDF4", "#166534", "✓ Done"),
    "error":   ("#FEF2F2", "#991B1B", "✗ Error"),
    "running": ("#FFFBEB", "#92400E", "⟳ Running"),
    "waiting": ("#F3F4F6", "#6B7280", "… Waiting"),
}

TABLE_COLS = [
    ("applicant_name",      "Applicant",               200, False, False),
    ("residence_address",   "Residence Address",        220, False, True),
    ("office_address",      "Office Address",           180, False, True),
    ("spouse",              "Spouse / Employment",      200, False, True),
    ("spouse_office",       "Spouse Office Address",    200, False, True),
    ("assets",              "Assets (Personal & Biz)",  240, False, True),
    ("income_items",        "Source of Income",         200, False, True),
    ("income_total",        "Total Income",             130, True,  False),
    ("business_items",      "Business Expenses",        200, False, True),
    ("business_total",      "Total Business",           130, True,  False),
    ("household_items",     "Household Expenses",       200, False, True),
    ("household_total",     "Total Household",          130, True,  False),
    ("net_income",          "Total Net Income",         130, True,  False),
    ("amort_history_total", "Total Amort. History",     150, True,  False),
    ("amort_current_total", "Total Current Amort.",     150, True,  False),
]

TREE_COLS = [c[0] for c in TABLE_COLS]

_VIRTUAL_COLS = {"spouse", "spouse_office", "assets", "amort_history_total"}

LOOKUP_ROWS = [
    ("cibi_place_of_work",      "CI/BI Report",      "Office Address"),
    ("cibi_temp_residence",     "CI/BI Report",      "Residence Address"),
    ("cibi_spouse",             "CI/BI Report",      "Spouse / Employment"),
    ("cibi_spouse_office",      "CI/BI Report",      "Spouse Office Address"),
    ("cibi_personal_assets",    "CI/BI Report",      "Personal Assets"),
    ("cibi_business_assets",    "CI/BI Report",      "Business Assets"),
    ("cibi_petrol_products",    "CI/BI Report",      "Petrol / Plastics / PVC Risk"),
    ("cibi_transport_services", "CI/BI Report",      "Transport Services Risk"),
    ("credit_history_amort",    "CI/BI Report",      "Credit History Amort."),
    ("income_remittance",       "Cashflow Analysis", "Source of Income"),
    ("cfa_business_expenses",   "Cashflow Analysis", "Business Expenses"),
    ("cfa_household_expenses",  "Cashflow Analysis", "Household / Personal Expenses"),
    ("ws_food_grocery",         "Worksheet",         "Food / Grocery"),
    ("ws_fuel_transport",       "Worksheet",         "Fuel and Transportation"),
    ("ws_electricity",          "Worksheet",         "Electricity Expense"),
    ("ws_fertilizer",           "Worksheet",         "Fertilizer"),
    ("ws_forwarding",           "Worksheet",         "Forwarding / Trucking / Hauling"),
    ("ws_fuel_diesel",          "Worksheet",         "Fuel / Gas / Diesel"),
    ("ws_equipment",            "Worksheet",         "Cost of Rent of Equipment"),
]

NON_MONETARY = {
    "cibi_place_of_work",
    "cibi_temp_residence",
    "cibi_spouse",
    "cibi_spouse_office",
    "cibi_personal_assets",
    "cibi_business_assets",
}


# ═══════════════════════════════════════════════════════════════════════
#  DATABASE LAYER
# ═══════════════════════════════════════════════════════════════════════

def _db_connect() -> sqlite3.Connection:
    DB_DIR.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(DB_PATH), timeout=10, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def _db_init():
    with _db_connect() as conn:
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS applicants (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id          TEXT    NOT NULL,
            processed_at        TEXT    NOT NULL,
            source_file         TEXT,
            status              TEXT    DEFAULT 'done',
            applicant_name      TEXT,
            residence_address   TEXT,
            office_address      TEXT,
            income_items        TEXT,
            income_total        REAL,
            business_items      TEXT,
            business_total      REAL,
            household_items     TEXT,
            household_total     REAL,
            net_income          REAL,
            petrol_risk         INTEGER DEFAULT 0,
            transport_risk      INTEGER DEFAULT 0,
            results_json        TEXT,
            page_map            TEXT,
            amort_current_total REAL
        );
        CREATE INDEX IF NOT EXISTS idx_session   ON applicants(session_id);
        CREATE INDEX IF NOT EXISTS idx_name      ON applicants(applicant_name COLLATE NOCASE);
        CREATE INDEX IF NOT EXISTS idx_status    ON applicants(status);
        CREATE INDEX IF NOT EXISTS idx_processed ON applicants(processed_at);
        """)
        cols = [r[1] for r in conn.execute("PRAGMA table_info(applicants)").fetchall()]
        if "amort_current_total" not in cols:
            conn.execute("ALTER TABLE applicants ADD COLUMN amort_current_total REAL")


def _db_upsert(session_id: str, row_data: dict) -> int:
    with _db_connect() as conn:
        existing = conn.execute(
            "SELECT id FROM applicants WHERE session_id=? AND source_file=?",
            (session_id, row_data.get("source_file", ""))
        ).fetchone()
        if existing:
            conn.execute("""
                UPDATE applicants SET
                    processed_at=:processed_at, status=:status,
                    applicant_name=:applicant_name, residence_address=:residence_address,
                    office_address=:office_address, income_items=:income_items,
                    income_total=:income_total, business_items=:business_items,
                    business_total=:business_total, household_items=:household_items,
                    household_total=:household_total, net_income=:net_income,
                    petrol_risk=:petrol_risk, transport_risk=:transport_risk,
                    results_json=:results_json, page_map=:page_map,
                    amort_current_total=:amort_current_total
                WHERE id=:id
            """, {**row_data, "id": existing["id"]})
            return existing["id"]
        else:
            cur = conn.execute("""
                INSERT INTO applicants (
                    session_id, processed_at, source_file, status,
                    applicant_name, residence_address, office_address,
                    income_items, income_total, business_items, business_total,
                    household_items, household_total, net_income,
                    petrol_risk, transport_risk, results_json, page_map,
                    amort_current_total
                ) VALUES (
                    :session_id, :processed_at, :source_file, :status,
                    :applicant_name, :residence_address, :office_address,
                    :income_items, :income_total, :business_items, :business_total,
                    :household_items, :household_total, :net_income,
                    :petrol_risk, :transport_risk, :results_json, :page_map,
                    :amort_current_total
                )
            """, row_data)
            return cur.lastrowid


def _db_query(search: str = "", session_id: str = "",
              sort_col: str = "processed_at", sort_asc: bool = False,
              offset: int = 0, limit: int = PAGE_SIZE) -> tuple:
    _DB_COLS = {c[0] for c in TABLE_COLS if c[0] not in _VIRTUAL_COLS}
    _DB_COLS |= {"processed_at", "session_id", "source_file", "id"}
    order_col = sort_col if sort_col in _DB_COLS else "processed_at"
    direction = "ASC" if sort_asc else "DESC"
    where_parts, params = [], []

    # ── Comma-separated AND search (adopted from summary_tab) ─────────
    # Each comma-separated term must independently match at least one field.
    terms = [t.strip() for t in search.split(",") if t.strip()] if search else []
    for term in terms:
        like = f"%{term}%"
        where_parts.append(
            "(applicant_name LIKE ? OR residence_address LIKE ? "
            "OR office_address LIKE ? OR income_items LIKE ? "
            "OR business_items LIKE ? OR household_items LIKE ? "
            "OR source_file LIKE ?)"
        )
        params.extend([like] * 7)

    if session_id:
        where_parts.append("session_id = ?")
        params.append(session_id)

    where = ("WHERE " + " AND ".join(where_parts)) if where_parts else ""
    with _db_connect() as conn:
        total = conn.execute(
            f"SELECT COUNT(*) FROM applicants {where}", params).fetchone()[0]
        rows  = conn.execute(
            f"SELECT * FROM applicants {where} "
            f"ORDER BY {order_col} {direction} LIMIT ? OFFSET ?",
            params + [limit, offset]).fetchall()
    return rows, total


def _db_totals(session_id: str = "", search: str = "") -> dict:
    where_parts, params = [], []

    # ── Comma-separated AND search (adopted from summary_tab) ─────────
    terms = [t.strip() for t in search.split(",") if t.strip()] if search else []
    for term in terms:
        like = f"%{term}%"
        where_parts.append(
            "(applicant_name LIKE ? OR residence_address LIKE ? "
            "OR office_address LIKE ? OR income_items LIKE ? "
            "OR business_items LIKE ? OR household_items LIKE ? "
            "OR source_file LIKE ?)"
        )
        params.extend([like] * 7)

    if session_id:
        where_parts.append("session_id = ?")
        params.append(session_id)

    where = ("WHERE " + " AND ".join(where_parts)) if where_parts else ""
    with _db_connect() as conn:
        row = conn.execute(f"""
            SELECT COUNT(*) as total,
                SUM(CASE WHEN status='done'  THEN 1 ELSE 0 END) as done,
                SUM(CASE WHEN status='error' THEN 1 ELSE 0 END) as errors,
                SUM(income_total)    as income,
                SUM(business_total)  as business,
                SUM(household_total) as household,
                SUM(net_income)      as net
            FROM applicants {where}
        """, params).fetchone()
    return dict(row) if row else {}


def _db_delete_row(row_id: int):
    with _db_connect() as conn:
        conn.execute("DELETE FROM applicants WHERE id=?", (row_id,))


def _db_clear_all():
    with _db_connect() as conn:
        conn.execute("DELETE FROM applicants")


# ═══════════════════════════════════════════════════════════════════════
#  NAME-MATCHING FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════

SUFFIXES = {"JR", "SR", "II", "III", "IV", "V", "ESQ", "PHD", "MD", "CPA"}


def _normalise_name(name: str, drop_initials: bool = False) -> str:
    tokens = re.split(r"[\s,]+", name.strip().upper())
    tokens = [re.sub(r"\.", "", t) for t in tokens]
    tokens = [t for t in tokens if t not in SUFFIXES]
    if drop_initials:
        tokens = [t for t in tokens if len(t) > 1]
    else:
        tokens = [t for t in tokens if t]
    return "".join(sorted(tokens))


def _reorder_lastname_first(name: str) -> str:
    if "," not in name:
        return name
    parts     = name.split(",", 1)
    lastname  = parts[0].strip()
    firstname = parts[1].strip()
    return f"{firstname} {lastname}"


def _firstlast_key(name: str) -> str:
    reordered = _reorder_lastname_first(name)
    tokens = re.split(r"[\s,]+", reordered.strip().upper())
    tokens = [re.sub(r"\.", "", t) for t in tokens if t]
    tokens = [t for t in tokens if t not in SUFFIXES]
    if len(tokens) <= 2:
        return "".join(sorted(tokens))
    return "".join(sorted([tokens[0], tokens[-1]]))


def _db_find_amort_match(name_key: str) -> list:
    with _db_connect() as conn:
        candidates = conn.execute(
            "SELECT id, applicant_name, processed_at FROM applicants "
            "WHERE applicant_name IS NOT NULL ORDER BY processed_at DESC"
        ).fetchall()
    return [
        (c["id"], c["applicant_name"])
        for c in candidates
        if _normalise_name(_reorder_lastname_first(c["applicant_name"])) == name_key
    ]


def _db_find_amort_match_relaxed(name_key_relaxed: str) -> list:
    with _db_connect() as conn:
        candidates = conn.execute(
            "SELECT id, applicant_name, processed_at FROM applicants "
            "WHERE applicant_name IS NOT NULL ORDER BY processed_at DESC"
        ).fetchall()
    return [
        (c["id"], c["applicant_name"])
        for c in candidates
        if _normalise_name(
            _reorder_lastname_first(c["applicant_name"]),
            drop_initials=True
        ) == name_key_relaxed
    ]


def _db_find_amort_match_firstlast(key_firstlast: str) -> list:
    with _db_connect() as conn:
        candidates = conn.execute(
            "SELECT id, applicant_name, processed_at FROM applicants "
            "WHERE applicant_name IS NOT NULL ORDER BY processed_at DESC"
        ).fetchall()
    return [
        (c["id"], c["applicant_name"])
        for c in candidates
        if _firstlast_key(c["applicant_name"]) == key_firstlast
    ]


def _db_update_amort_current(row_id: int, value: float) -> bool:
    with _db_connect() as conn:
        conn.execute(
            "UPDATE applicants SET amort_current_total=? WHERE id=?",
            (value, row_id))
    return True


def _db_update_amort_all(matches: list, value: float) -> int:
    count = 0
    for row_id, _ in matches:
        _db_update_amort_current(row_id, value)
        count += 1
    return count


# ═══════════════════════════════════════════════════════════════════════
#  PUBLIC WRITER  (called from general_lookup.py)
# ═══════════════════════════════════════════════════════════════════════

def db_save_applicant(session_id: str, results: dict):
    """
    Persist one applicant's extraction results to the General SQLite DB.
    Called by general_lookup._general_process_single_file after Gemini returns.
    """
    _db_init()
    gate = results.get("_gate_data", {})

    def _items(key):
        d = results.get(key, {})
        return "\n".join(d.get("items", []) if isinstance(d, dict) else [])

    def _total(key):
        d = results.get(key, {})
        t = d.get("total", "") if isinstance(d, dict) else ""
        if not t:
            return None
        try:
            return float(re.sub(r"[^\d.]", "", str(t).replace(",", "")))
        except Exception:
            return None

    raw_net = str(results.get("_cfa_net_income", "")).strip()
    net_val = None
    if raw_net:
        try:
            net_val = float(re.sub(r"[^\d.]", "", raw_net.replace(",", "")))
        except Exception:
            pass
    if net_val is None:
        inc = _total("income_remittance")     or 0
        biz = _total("cfa_business_expenses") or 0
        hh  = _total("cfa_household_expenses") or 0
        if inc or biz or hh:
            net_val = inc - biz - hh

    petrol    = bool(results.get("cibi_petrol_products",    {}).get("items"))
    transport = bool(results.get("cibi_transport_services", {}).get("items"))

    slim = {k: v for k, v in results.items()
            if not k.startswith("_") and isinstance(v, dict)}

    row_data = {
        "session_id":          session_id,
        "processed_at":        datetime.now().isoformat(timespec="seconds"),
        "source_file":         results.get("_source_file", ""),
        "status":              "done",
        "applicant_name":      results.get("_applicant_name", ""),
        "residence_address":   gate.get("residence_address", ""),
        "office_address":      gate.get("office_address", ""),
        "income_items":        _items("income_remittance"),
        "income_total":        _total("income_remittance"),
        "business_items":      _items("cfa_business_expenses"),
        "business_total":      _total("cfa_business_expenses"),
        "household_items":     _items("cfa_household_expenses"),
        "household_total":     _total("cfa_household_expenses"),
        "net_income":          net_val,
        "petrol_risk":         1 if petrol    else 0,
        "transport_risk":      1 if transport else 0,
        "results_json":        json.dumps(slim, ensure_ascii=False),
        "page_map":            results.get("_page_map", ""),
        "amort_current_total": None,
    }
    _db_upsert(session_id, row_data)


# ═══════════════════════════════════════════════════════════════════════
#  TREEVIEW STYLE
# ═══════════════════════════════════════════════════════════════════════

def _apply_tree_style():
    style = ttk.Style()
    style.theme_use("clam")
    style.configure("GeneralSummary.Treeview",
        background=WHITE, foreground=TXT_NAVY, fieldbackground=WHITE,
        rowheight=36, font=("Segoe UI", 9), borderwidth=0, relief="flat")
    style.configure("GeneralSummary.Treeview.Heading",
        background=HDR_BG, foreground=HDR_FG,
        font=("Segoe UI", 9, "bold"), relief="flat", borderwidth=0, padding=(8, 8))
    style.map("GeneralSummary.Treeview.Heading",
        background=[("active", "#7AB567")], relief=[("active", "flat")])
    style.map("GeneralSummary.Treeview",
        background=[("selected", "#C8E6C9")], foreground=[("selected", NAVY_DEEP)])


# ═══════════════════════════════════════════════════════════════════════
#  PANEL BUILDER
# ═══════════════════════════════════════════════════════════════════════

def _build_general_summary_panel(self, parent):
    _db_init()
    _apply_tree_style()

    outer = tk.Frame(parent, bg=CARD_WHITE)
    self._general_summary_frame = outer
    main = tk.Frame(outer, bg=CARD_WHITE)
    main.pack(fill="both", expand=True)

    # ── Header band ───────────────────────────────────────────────────
    header_band = tk.Frame(main, bg=NAVY_DEEP)
    header_band.pack(fill="x", padx=PAD, pady=(16, 0))
    title_block = tk.Frame(header_band, bg=NAVY_DEEP)
    title_block.pack(side="left", padx=16, pady=12)
    tk.Label(title_block, text="General Summary",
             font=F(15, "bold"), fg=WHITE, bg=NAVY_DEEP).pack(anchor="w")
    tk.Label(title_block,
             text="Persistent across sessions  ·  SQLite backed  ·  live view",
             font=F(8), fg="#8DA8C8", bg=NAVY_DEEP).pack(anchor="w", pady=(2, 0))

    btn_block = tk.Frame(header_band, bg=NAVY_DEEP)
    btn_block.pack(side="right", padx=12, pady=10)
    self._gen_sum_export_csv_btn = ctk.CTkButton(
        btn_block, text="⬇  CSV", command=lambda: _export_csv(self),
        width=80, height=30, corner_radius=6, fg_color="transparent",
        hover_color="#1E3A5F", text_color="#8DA8C8", font=FF(8, "bold"),
        border_width=1, border_color="#2E4E72")
    self._gen_sum_export_csv_btn.pack(side="left", padx=(0, 4))
    self._gen_sum_export_xl_btn = ctk.CTkButton(
        btn_block, text="📊  Excel", command=lambda: _export_excel(self),
        width=88, height=30, corner_radius=6, fg_color=LIME_MID,
        hover_color=LIME_BRIGHT, text_color=TXT_ON_LIME, font=FF(8, "bold"),
        border_width=0)
    self._gen_sum_export_xl_btn.pack(side="left", padx=(0, 4))
    self._gen_sum_refresh_btn = ctk.CTkButton(
        btn_block, text="↺  Refresh", command=lambda: _refresh_summary(self),
        width=88, height=30, corner_radius=6, fg_color="#1A3A5C",
        hover_color="#1E4A72", text_color=WHITE, font=FF(8, "bold"),
        border_width=0)
    self._gen_sum_refresh_btn.pack(side="left", padx=(0, 4))
    self._gen_sum_clear_all_btn = ctk.CTkButton(
        btn_block, text="🗑  Clear All", command=lambda: _clear_all(self),
        width=96, height=30, corner_radius=6, fg_color="#3D1010",
        hover_color="#5C1A1A", text_color="#FF8A80", font=FF(8, "bold"),
        border_width=0)
    self._gen_sum_clear_all_btn.pack(side="left")
    self._gen_sum_import_amort_btn = ctk.CTkButton(
        btn_block, text="⬆  Amort.", command=lambda: _import_amort_file(self),
        width=96, height=30, corner_radius=6, fg_color="#1A3A5C",
        hover_color="#1E4A72", text_color=WHITE, font=FF(8, "bold"),
        border_width=0)
    self._gen_sum_import_amort_btn.pack(side="left", padx=(4, 0))
    self._gen_sum_merge_db_btn = ctk.CTkButton(
        btn_block, text="⛁  Merge DB", command=lambda: _merge_db_files(self),
        width=100, height=30, corner_radius=6, fg_color="#2D4A1E",
        hover_color="#3D6128", text_color="#B9F5A0", font=FF(8, "bold"),
        border_width=0)
    self._gen_sum_merge_db_btn.pack(side="left", padx=(4, 0))
    self._gen_sum_merge_xl_btn = ctk.CTkButton(
        btn_block, text="⛁  Merge Excel", command=lambda: _merge_excel_files(self),
        width=110, height=30, corner_radius=6, fg_color="#1E3D4A",
        hover_color="#255262", text_color="#A0E4F5", font=FF(8, "bold"),
        border_width=0)
    self._gen_sum_merge_xl_btn.pack(side="left", padx=(4, 0))

    # ── Controls row ──────────────────────────────────────────────────
    controls_row = tk.Frame(main, bg="#F0F4FA",
                            highlightbackground=BORDER_MID, highlightthickness=1)
    controls_row.pack(fill="x", padx=PAD, pady=(0, 0))

    stats_group = tk.Frame(controls_row, bg="#F0F4FA")
    stats_group.pack(side="left", padx=10, pady=8)
    self._gen_sum_stat_labels = {}
    for key, label, color, pill_bg in [
        ("total",  "Total",  NAVY_DEEP,      "#E8EEF8"),
        ("done",   "Done",   ACCENT_SUCCESS, "#F0FDF4"),
        ("errors", "Errors", ACCENT_RED,     "#FFF0F0"),
        ("income", "Income", NAVY_MID,       "#EEF3FA"),
        ("net",    "Net",    NET_GREEN,      "#F0FDF4"),
    ]:
        pill = tk.Frame(stats_group, bg=pill_bg,
                        highlightbackground=BORDER_MID, highlightthickness=1)
        pill.pack(side="left", padx=(0, 6))
        tk.Label(pill, text=label, font=F(7, "bold"), fg=TXT_MUTED,
                 bg=pill_bg, padx=10, pady=3).pack()
        lbl = tk.Label(pill, text="—", font=F(11, "bold"),
                       fg=color, bg=pill_bg, padx=10, pady=4)
        lbl.pack()
        self._gen_sum_stat_labels[key] = lbl

    tk.Frame(controls_row, bg=BORDER_MID, width=1).pack(
        side="left", fill="y", pady=6, padx=4)

    # ── Search box with placeholder hint (adopted from summary_tab) ───
    search_wrap = tk.Frame(controls_row, bg=WHITE,
                           highlightbackground=BORDER_MID, highlightthickness=1)
    search_wrap.pack(side="left", fill="x", expand=True, padx=8, pady=8)
    tk.Label(search_wrap, text="🔍", font=("Segoe UI Emoji", 9),
             bg=WHITE, fg=NAVY_PALE).pack(side="left", padx=(8, 2))

    self._gen_sum_search_var = tk.StringVar()
    self._gen_sum_search_var.trace_add("write", lambda *a: _on_search_change(self))

    _se = tk.Entry(search_wrap, textvariable=self._gen_sum_search_var,
                   font=F(9), fg=TXT_MUTED, bg=WHITE, relief="flat", bd=0,
                   insertbackground=NAVY_MID)
    _se.pack(side="left", fill="x", expand=True, pady=6)
    _se.insert(0, _SEARCH_HINT)

    def _se_focus_in(e):
        if self._gen_sum_search_var.get() == _SEARCH_HINT:
            _se.delete(0, "end")
            _se.config(fg=TXT_NAVY)

    def _se_focus_out(e):
        if not self._gen_sum_search_var.get().strip():
            _se.config(fg=TXT_MUTED)
            _se.insert(0, _SEARCH_HINT)

    _se.bind("<FocusIn>",  _se_focus_in)
    _se.bind("<FocusOut>", _se_focus_out)
    # ──────────────────────────────────────────────────────────────────

    right_ctrl = tk.Frame(controls_row, bg="#F0F4FA")
    right_ctrl.pack(side="right", padx=10, pady=8)
    self._gen_sum_count_lbl = tk.Label(right_ctrl, text="",
                                       font=F(8), fg=TXT_SOFT, bg="#F0F4FA")
    self._gen_sum_count_lbl.pack(side="top", anchor="e")
    pg_sub = tk.Frame(right_ctrl, bg="#F0F4FA")
    pg_sub.pack(side="top", pady=(4, 0))
    self._gen_sum_prev_btn = ctk.CTkButton(
        pg_sub, text="◀", command=lambda: _page_prev(self),
        width=30, height=24, corner_radius=5, fg_color=CARD_WHITE,
        hover_color=NAVY_MIST, text_color=NAVY_MID, font=FF(8, "bold"),
        border_width=1, border_color=BORDER_MID, state="disabled")
    self._gen_sum_prev_btn.pack(side="left", padx=(0, 4))
    self._gen_sum_page_lbl = tk.Label(pg_sub, text="Page 1",
                                      font=F(8), fg=TXT_SOFT, bg="#F0F4FA")
    self._gen_sum_page_lbl.pack(side="left", padx=4)
    self._gen_sum_next_btn = ctk.CTkButton(
        pg_sub, text="▶", command=lambda: _page_next(self),
        width=30, height=24, corner_radius=5, fg_color=CARD_WHITE,
        hover_color=NAVY_MIST, text_color=NAVY_MID, font=FF(8, "bold"),
        border_width=1, border_color=BORDER_MID, state="disabled")
    self._gen_sum_next_btn.pack(side="left", padx=(4, 0))

    # ── Table ─────────────────────────────────────────────────────────
    tbl_outer = tk.Frame(main, bg=BORDER_LIGHT)
    tbl_outer.pack(fill="both", expand=True, padx=PAD, pady=(8, PAD))
    tbl_wrap = tk.Frame(tbl_outer, bg=CARD_WHITE)
    tbl_wrap.pack(fill="both", expand=True, padx=1, pady=1)
    tbl_wrap.rowconfigure(0, weight=1)
    tbl_wrap.columnconfigure(0, weight=1)

    vscroll = tk.Scrollbar(tbl_wrap, orient="vertical", relief="flat",
                           troughcolor=OFF_WHITE, bg=BORDER_LIGHT, width=8, bd=0)
    vscroll.grid(row=0, column=1, sticky="ns")
    hscroll = tk.Scrollbar(tbl_wrap, orient="horizontal", relief="flat",
                           troughcolor=OFF_WHITE, bg=BORDER_LIGHT, bd=0)
    hscroll.grid(row=1, column=0, columnspan=2, sticky="ew")

    self._gen_sum_tree = ttk.Treeview(
        tbl_wrap, columns=TREE_COLS, show="headings",
        style="GeneralSummary.Treeview",
        yscrollcommand=vscroll.set, xscrollcommand=hscroll.set,
        selectmode="browse")
    self._gen_sum_tree.grid(row=0, column=0, sticky="nsew")
    vscroll.config(command=self._gen_sum_tree.yview)
    hscroll.config(command=self._gen_sum_tree.xview)

    for db_col, label, width_px, is_mon, is_txt in TABLE_COLS:
        anchor = "e" if is_mon else "w"
        self._gen_sum_tree.heading(db_col, text=label,
                                   command=lambda c=db_col: _sort_by(self, c))
        self._gen_sum_tree.column(db_col, width=width_px, minwidth=80,
                                  anchor=anchor, stretch=False)

    self._gen_sum_tree.tag_configure("even", background=ROW_BG_EVEN)
    self._gen_sum_tree.tag_configure("odd",  background=ROW_BG_ODD)
    self._gen_sum_tree.tag_configure("net",  foreground=NET_GREEN)
    self._gen_sum_tree.bind("<Double-1>", lambda e: _on_tree_double_click(self, e))
    self._gen_sum_tree.bind("<Button-3>", lambda e: _on_tree_right_click(self, e))
    self._gen_sum_tree.bind("<Enter>",
        lambda e: self._gen_sum_tree.bind_all("<MouseWheel>",
            lambda ev: self._gen_sum_tree.yview_scroll(
                int(-1 * (ev.delta / 120)), "units")))
    self._gen_sum_tree.bind("<Leave>",
        lambda e: self._gen_sum_tree.unbind_all("<MouseWheel>"))

    # ── Internal state ────────────────────────────────────────────────
    self._gen_sum_sort_col       = "processed_at"
    self._gen_sum_sort_asc       = False
    self._gen_sum_page           = 0
    self._gen_sum_total_rows     = 0
    self._gen_sum_session_filter = ""
    self._gen_sum_search_after   = None
    self._gen_sum_row_data       = {}
    _refresh_summary(self)


# ═══════════════════════════════════════════════════════════════════════
#  DATA LOADING + RENDERING
# ═══════════════════════════════════════════════════════════════════════

def _is_placeholder(self) -> bool:
    """Returns True when the search box is showing the hint text."""
    return _SEARCH_HINT in self._gen_sum_search_var.get()


def _on_search_change(self):
    """
    Debounced handler — ignores changes while the placeholder is active
    so typing into a fresh search box doesn't trigger a spurious query.
    Adopted from summary_tab._on_search_change.
    """
    if _is_placeholder(self):
        return
    if self._gen_sum_search_after:
        self.after_cancel(self._gen_sum_search_after)
    self._gen_sum_search_after = self.after(300, lambda: _load_and_render(self))


def _refresh_summary(self):
    _load_and_render(self)


def _active_search(self) -> str:
    """Return the real search string, or '' when the placeholder is showing."""
    raw = self._gen_sum_search_var.get().strip()
    return "" if _is_placeholder(self) else raw


def _load_and_render(self):
    search = _active_search(self)
    offset = self._gen_sum_page * PAGE_SIZE
    rows, total = _db_query(
        search=search, session_id=self._gen_sum_session_filter,
        sort_col=self._gen_sum_sort_col, sort_asc=self._gen_sum_sort_asc,
        offset=offset, limit=PAGE_SIZE)
    self._gen_sum_total_rows = total
    _update_stats(self)
    _update_pagination(self, total)
    _render_tree(self, rows)
    shown_start = offset + 1 if total > 0 else 0
    shown_end   = min(offset + PAGE_SIZE, total)
    self._gen_sum_count_lbl.config(
        text=f"{shown_start}–{shown_end} of {total} applicant(s)")


def _update_stats(self):
    search = _active_search(self)
    tots   = _db_totals(session_id=self._gen_sum_session_filter, search=search)
    def _fmt(val):
        if val is None: return "—"
        try:    return f"P{float(val):,.0f}"
        except: return "—"
    self._gen_sum_stat_labels["total"].config(text=str(tots.get("total", 0)))
    self._gen_sum_stat_labels["done"].config(text=str(tots.get("done", 0)))
    self._gen_sum_stat_labels["errors"].config(text=str(tots.get("errors", 0)))
    self._gen_sum_stat_labels["income"].config(text=_fmt(tots.get("income")))
    self._gen_sum_stat_labels["net"].config(text=_fmt(tots.get("net")))


def _update_pagination(self, total: int):
    total_pages = max(1, -(-total // PAGE_SIZE))
    cur_page    = self._gen_sum_page + 1
    self._gen_sum_page_lbl.config(text=f"Page {cur_page} / {total_pages}")
    self._gen_sum_prev_btn.configure(
        state="normal" if self._gen_sum_page > 0 else "disabled")
    self._gen_sum_next_btn.configure(
        state="normal" if cur_page < total_pages else "disabled")


def _page_prev(self):
    if self._gen_sum_page > 0:
        self._gen_sum_page -= 1
        _load_and_render(self)


def _page_next(self):
    total_pages = max(1, -(-self._gen_sum_total_rows // PAGE_SIZE))
    if self._gen_sum_page + 1 < total_pages:
        self._gen_sum_page += 1
        _load_and_render(self)


def _fmt_money(val) -> str:
    if val in (None, ""): return "—"
    try:    return f"P{float(val):,.2f}"
    except: return str(val) or "—"


def _extract_amort_history_total(results_blob: dict):
    try:
        amort_data = results_blob.get("credit_history_amort", {})
        raw_total  = amort_data.get("total", "") if isinstance(amort_data, dict) else ""
        if raw_total:
            cleaned = re.sub(r"[^\d.]", "", str(raw_total).replace(",", ""))
            return float(cleaned) if cleaned else None
    except Exception:
        pass
    return None


def _render_tree(self, rows):
    self._gen_sum_tree.delete(*self._gen_sum_tree.get_children())
    self._gen_sum_row_data = {}

    for i, row in enumerate(rows):
        row    = dict(row)
        row_id = row.get("id")

        try:
            results_blob = json.loads(row.get("results_json", "") or "{}")
        except Exception:
            results_blob = {}
        row["amort_history_total"] = _extract_amort_history_total(results_blob)
        row["amort_current_total"] = row.get("amort_current_total")

        def _items_flat(key, _blob=results_blob):
            d = _blob.get(key, {})
            items = d.get("items", []) if isinstance(d, dict) else []
            return "  ·  ".join(items) if items else ""

        row["spouse"]        = _items_flat("cibi_spouse")
        row["spouse_office"] = _items_flat("cibi_spouse_office")
        personal = _items_flat("cibi_personal_assets")
        business = _items_flat("cibi_business_assets")
        if personal and business:
            row["assets"] = f"[Personal] {personal}  |  [Business] {business}"
        else:
            row["assets"] = personal or business

        tag    = "even" if i % 2 == 0 else "odd"
        values = []
        for db_col, label, width_px, is_monetary, is_text_block in TABLE_COLS:
            raw = row.get(db_col, "") or ""
            if is_monetary:
                values.append(_fmt_money(raw))
            elif is_text_block:
                values.append(str(raw).replace("\n", "  ·  "))
            else:
                values.append(str(raw))
        self._gen_sum_tree.insert("", "end", iid=str(row_id),
                                  values=values, tags=(tag,))
        self._gen_sum_row_data[str(row_id)] = row


def _sort_by(self, col_key: str):
    if col_key in _VIRTUAL_COLS:
        return
    if self._gen_sum_sort_col == col_key:
        self._gen_sum_sort_asc = not self._gen_sum_sort_asc
    else:
        self._gen_sum_sort_col = col_key
        self._gen_sum_sort_asc = True
    self._gen_sum_page = 0
    for db_col, label, _, __, ___ in TABLE_COLS:
        active = (self._gen_sum_sort_col == db_col)
        ind = (" ▲" if active and self._gen_sum_sort_asc
               else " ▼" if active else "")
        self._gen_sum_tree.heading(db_col, text=label + ind)
    _load_and_render(self)


# ═══════════════════════════════════════════════════════════════════════
#  ROW INTERACTION
# ═══════════════════════════════════════════════════════════════════════

def _on_tree_double_click(self, event):
    iid = self._gen_sum_tree.focus()
    if not iid: return
    row = self._gen_sum_row_data.get(iid)
    if row: _open_detail_window(self, row)


def _on_tree_right_click(self, event):
    iid = self._gen_sum_tree.identify_row(event.y)
    if not iid: return
    self._gen_sum_tree.selection_set(iid)
    self._gen_sum_tree.focus(iid)
    menu = tk.Menu(self._gen_sum_tree, tearoff=0, bg=CARD_WHITE, fg=TXT_NAVY,
                   activebackground=NAVY_MIST, font=("Segoe UI", 9))
    menu.add_command(label="👁  View Details",
        command=lambda: _open_detail_window(
            self, self._gen_sum_row_data.get(iid, {})))
    menu.add_separator()
    menu.add_command(label="✕  Delete",
        command=lambda: _delete_row(self, int(iid)))
    try:
        menu.tk_popup(event.x_root, event.y_root)
    finally:
        menu.grab_release()


def _open_detail_window(self, row: dict):
    win = tk.Toplevel(self)
    win.title(f"Detail — {row.get('applicant_name', '')}")
    win.geometry("780x640")
    win.configure(bg=CARD_WHITE)
    win.grab_set()
    canvas  = tk.Canvas(win, bg=CARD_WHITE, highlightthickness=0)
    vscroll = tk.Scrollbar(win, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=vscroll.set)
    vscroll.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)
    body = tk.Frame(canvas, bg=CARD_WHITE)
    cwin = canvas.create_window((0, 0), window=body, anchor="nw")
    def _on_cfg(e):
        canvas.configure(scrollregion=canvas.bbox("all"))
        canvas.itemconfig(cwin, width=canvas.winfo_width())
    body.bind("<Configure>", _on_cfg)
    canvas.bind("<Configure>", _on_cfg)
    canvas.bind_all("<MouseWheel>",
        lambda e: canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"))
    _build_detail_panel(self, row, body)
    ctk.CTkButton(body, text="Close", command=win.destroy,
                  width=100, height=32, corner_radius=7,
                  fg_color=NAVY_LIGHT, hover_color=NAVY_PALE,
                  text_color=WHITE, font=FF(9, "bold")).pack(pady=(12, 16))


def _build_detail_panel(self, row: dict, parent: tk.Frame):
    info_strip = tk.Frame(parent, bg=NAVY_MIST)
    info_strip.pack(fill="x")
    for label, value in [
        ("Applicant",         row.get("applicant_name",    "—") or "—"),
        ("Residence Address", row.get("residence_address", "—") or "—"),
        ("Office Address",    row.get("office_address",    "—") or "—"),
        ("Source File",       row.get("source_file",       "—") or "—"),
        ("Processed At",
         (row.get("processed_at", "") or "")[:16].replace("T", "  ")),
        ("Session",
         (row.get("session_id", "") or "")[:19].replace("T", "  ")),
    ]:
        col = tk.Frame(info_strip, bg=NAVY_MIST)
        col.pack(side="left", padx=12, pady=8, anchor="w")
        tk.Label(col, text=label, font=F(7, "bold"),
                 fg=TXT_MUTED, bg=NAVY_MIST).pack(anchor="w")
        tk.Label(col, text=value or "—", font=F(9, "bold"),
                 fg=NAVY_DEEP, bg=NAVY_MIST,
                 wraplength=180, justify="left").pack(anchor="w")

    flags = tk.Frame(info_strip, bg=NAVY_MIST)
    flags.pack(side="right", padx=12, pady=8)
    if row.get("petrol_risk"):
        tk.Label(flags, text="⚠ Petrol Risk", font=F(8, "bold"),
                 fg=ACCENT_RED, bg=NAVY_MIST).pack(anchor="e")
    if row.get("transport_risk"):
        tk.Label(flags, text="⚠ Transport Risk", font=F(8, "bold"),
                 fg=ACCENT_RED, bg=NAVY_MIST).pack(anchor="e")

    tk.Frame(parent, bg=BORDER_MID, height=1).pack(fill="x")

    try:
        results = json.loads(row.get("results_json", "") or "{}")
    except Exception:
        results = {}

    last_section = None
    PAD_X        = 16

    for idx, (key, section, field_label) in enumerate(LOOKUP_ROWS):
        field_data = results.get(key, {})
        items  = field_data.get("items", []) if isinstance(field_data, dict) else []
        total  = field_data.get("total", "") if isinstance(field_data, dict) else ""
        non_m  = key in NON_MONETARY

        if section != last_section:
            last_section = section
            sec_bar = tk.Frame(parent, bg=SEC_BG)
            sec_bar.pack(fill="x", padx=PAD_X, pady=(8, 0))
            tk.Label(sec_bar, text=f"  {section.upper()}",
                     font=F(8, "bold"), fg=SEC_FG, bg=SEC_BG,
                     pady=4).pack(side="left")

        row_bg = ROW_BG_EVEN if idx % 2 == 0 else ROW_BG_ODD
        row_f  = tk.Frame(parent, bg=row_bg,
                          highlightbackground="#E5EAF3", highlightthickness=1)
        row_f.pack(fill="x", padx=PAD_X)
        tk.Label(row_f, text=field_label, font=F(8, "bold"),
                 fg=NAVY_DEEP, bg=row_bg,
                 padx=8, pady=6, anchor="w", width=26).pack(side="left")

        amt_txt = total if (total and not non_m) else "—"
        tk.Label(row_f, text=amt_txt,
                 font=F(9, "bold") if amt_txt != "—" else F(8),
                 fg=NAVY_MID if amt_txt != "—" else TXT_MUTED,
                 bg=row_bg, padx=8, width=14, anchor="e").pack(side="left")

        det_txt = ("\n".join(f"• {it}" for it in items)
                   if items else "No data found")
        tk.Label(row_f, text=det_txt, font=F(8),
                 fg=TXT_NAVY if items else TXT_MUTED,
                 bg=row_bg, padx=8, anchor="w",
                 wraplength=440, justify="left").pack(
                     side="left", fill="x", expand=True)

    page_map = row.get("page_map", "") or ""
    if page_map:
        pm = tk.Frame(parent, bg=CARD_WHITE)
        pm.pack(fill="x", padx=PAD_X, pady=(8, 12))
        tk.Label(pm, text="Page Map:", font=F(7, "bold"),
                 fg=TXT_MUTED, bg=CARD_WHITE).pack(anchor="w")
        tk.Label(pm, text=page_map, font=FMONO(7), fg=TXT_SOFT,
                 bg=CARD_WHITE, justify="left", anchor="w").pack(
                     anchor="w", padx=8)
    else:
        tk.Frame(parent, bg=CARD_WHITE, height=10).pack()


# ═══════════════════════════════════════════════════════════════════════
#  DELETE / CLEAR
# ═══════════════════════════════════════════════════════════════════════

def _delete_row(self, row_id: int):
    if not messagebox.askyesno("Delete Record",
            "Remove this applicant from the General database?\n\nThis cannot be undone."):
        return
    _db_delete_row(row_id)
    _refresh_summary(self)


def _clear_all(self):
    totals = _db_totals()
    total  = totals.get("total", 0) or 0
    if total == 0:
        messagebox.showinfo("Clear All", "There are no records to delete.")
        return
    if not messagebox.askyesno("Clear All Records",
            f"This will permanently delete ALL {total} applicant record(s) "
            f"from the General database.\n\nThis cannot be undone.\n\nContinue?",
            icon="warning"):
        return
    _db_clear_all()
    _refresh_summary(self)


# ═══════════════════════════════════════════════════════════════════════
#  AMORTIZATION IMPORT
# ═══════════════════════════════════════════════════════════════════════

def _import_amort_file(self):
    path = filedialog.askopenfilename(
        title="Import Amortization Values",
        filetypes=[
            ("Excel & CSV files", "*.xlsx *.csv"),
            ("Excel files",       "*.xlsx"),
            ("CSV files",         "*.csv"),
            ("All files",         "*.*"),
        ])
    if not path:
        return

    _flash_btn(self, self._gen_sum_import_amort_btn, "⟳  Reading…", 60_000)

    def _worker():
        try:
            if path.lower().endswith(".csv"):
                import csv as _csv
                with open(path, newline="", encoding="utf-8-sig") as f:
                    reader  = _csv.DictReader(f)
                    records = [dict(row) for row in reader]
                all_cols = list(records[0].keys()) if records else []
            else:
                import openpyxl
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
                if header_row is None:
                    raise ValueError("The Excel file appears to be empty.")
                all_cols = [
                    str(cell.value).strip() if cell.value is not None else ""
                    for cell in header_row
                ]
                records = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if all(v is None for v in row):
                        continue
                    records.append({
                        all_cols[i]: (str(v).strip() if v is not None else "")
                        for i, v in enumerate(row)
                        if i < len(all_cols)
                    })
                wb.close()

            if not records:
                raise ValueError("No data rows found in the file.")

            def _find_col(cols, *keywords):
                for kw in keywords:
                    kw_norm = re.sub(r"[\s_]", "", kw.lower())
                    for c in cols:
                        c_norm = re.sub(r"[\s_]", "", c.lower())
                        if kw_norm in c_norm:
                            return c
                return None

            col_client = _find_col(all_cols, "applicant", "client", "name")
            col_amort  = _find_col(all_cols,
                                   "monthlypaymentamount", "monthly payment amount",
                                   "monthlypayment",       "paymentamount",
                                   "currentamort",         "totalcurrentamort",
                                   "amort")

            missing = []
            if not col_client: missing.append("Applicant / Client")
            if not col_amort:  missing.append("MonthlyPaymentAmount / Amortization")
            if missing:
                raise ValueError(
                    f"Could not detect column(s): {', '.join(missing)}\n\n"
                    f"File has: {', '.join(all_cols)}")

            aggregated = {}
            bad_rows   = []

            for file_row in records:
                client_name = str(file_row.get(col_client) or "").strip()
                raw_val     = str(file_row.get(col_amort)  or "").strip()
                display     = client_name.upper()
                if not client_name:
                    continue
                try:
                    cleaned   = re.sub(r"[^\d.]", "", raw_val.replace(",", ""))
                    amort_val = float(cleaned) if cleaned else None
                except Exception:
                    amort_val = None
                if amort_val is None:
                    bad_rows.append((display, f"bad value: '{raw_val}'"))
                    continue
                if client_name in aggregated:
                    aggregated[client_name] += amort_val
                else:
                    aggregated[client_name]  = amort_val

            updated_strict  = []
            updated_relaxed = []
            skipped_names   = list(bad_rows)

            for client_name, amort_val in aggregated.items():
                display   = client_name.upper()
                reordered = _reorder_lastname_first(client_name)

                key_strict = _normalise_name(reordered)
                hits = _db_find_amort_match(key_strict)
                if hits:
                    _db_update_amort_all(hits, amort_val)
                    updated_strict.append((display, hits[0][1]))
                    continue

                key_relaxed = _normalise_name(reordered, drop_initials=True)
                hits = _db_find_amort_match_relaxed(key_relaxed)
                if hits:
                    _db_update_amort_all(hits, amort_val)
                    updated_relaxed.append((display, hits[0][1]))
                    continue

                key_raw_relaxed = _normalise_name(client_name, drop_initials=True)
                hits = _db_find_amort_match_relaxed(key_raw_relaxed)
                if hits:
                    _db_update_amort_all(hits, amort_val)
                    updated_relaxed.append((display, hits[0][1]))
                    continue

                key_firstlast = _firstlast_key(client_name)
                hits = _db_find_amort_match_firstlast(key_firstlast)
                if hits:
                    _db_update_amort_all(hits, amort_val)
                    updated_relaxed.append((display, hits[0][1]))
                    continue

                skipped_names.append((display, "no DB match"))

            self.after(0, lambda: _refresh_summary(self))

            total_updated = len(updated_strict) + len(updated_relaxed)
            msg  = "Import complete.\n\n"
            msg += f"✓  Updated  : {total_updated} record(s)\n"
            msg += f"–  Skipped  : {len(skipped_names)} row(s)\n"
            if updated_relaxed:
                msg += (f"\n⚠  {len(updated_relaxed)} matched via relaxed/first-last "
                        "pass (middle name ignored) — please verify:\n")
                for file_n, db_n in updated_relaxed[:10]:
                    msg += f"  • File: {file_n}  →  DB: {db_n}\n"
                if len(updated_relaxed) > 10:
                    msg += f"  … and {len(updated_relaxed) - 10} more\n"
            if skipped_names:
                msg += "\nSkipped rows:\n"
                for name, reason in skipped_names[:10]:
                    msg += f"  • {name}  ({reason})\n"
                if len(skipped_names) > 10:
                    msg += f"  … and {len(skipped_names) - 10} more"

            self.after(0, lambda: (
                _flash_btn(self, self._gen_sum_import_amort_btn, "✓  Done!", 2500),
                messagebox.showinfo("Amort. Import Result", msg)
            ))

        except Exception as exc:
            err = str(exc)
            self.after(0, lambda: (
                _flash_btn(self, self._gen_sum_import_amort_btn, "✗  Error", 3000),
                messagebox.showerror("Import Error", err)
            ))

    threading.Thread(target=_worker, daemon=True).start()


# ═══════════════════════════════════════════════════════════════════════
#  MERGE DB
# ═══════════════════════════════════════════════════════════════════════

def _merge_db_files(self):
    paths = filedialog.askopenfilenames(
        title="Select DB files to merge into General database",
        filetypes=[("SQLite DB files", "*.db"), ("All files", "*.*")])
    if not paths:
        return

    src_paths = [p for p in paths if Path(p).resolve() != DB_PATH.resolve()]
    if not src_paths:
        messagebox.showwarning("Merge DB",
            "All selected files are the current General database — nothing to merge.")
        return

    _flash_btn(self, self._gen_sum_merge_db_btn, "⟳  Merging…", 60_000)

    def _worker():
        _COLS = [
            "session_id", "processed_at", "source_file", "status",
            "applicant_name", "residence_address", "office_address",
            "income_items", "income_total", "business_items", "business_total",
            "household_items", "household_total", "net_income",
            "petrol_risk", "transport_risk", "results_json", "page_map",
            "amort_current_total",
        ]
        _INSERT = (
            f"INSERT INTO applicants ({', '.join(_COLS)}) "
            f"VALUES ({', '.join(['?' for _ in _COLS])})"
        )

        def _ensure_col(conn, col, col_type):
            existing = [r[1] for r in
                        conn.execute("PRAGMA table_info(applicants)").fetchall()]
            if col not in existing:
                conn.execute(f"ALTER TABLE applicants ADD COLUMN {col} {col_type}")

        total_inserted = total_skipped = 0
        file_results   = []

        try:
            with _db_connect() as out_conn:
                _ensure_col(out_conn, "amort_current_total", "REAL")

                for src in src_paths:
                    try:
                        s_conn = sqlite3.connect(str(src), timeout=10)
                        s_conn.row_factory = sqlite3.Row
                        _ensure_col(s_conn, "amort_current_total", "REAL")
                        rows = s_conn.execute(
                            "SELECT * FROM applicants").fetchall()
                        s_conn.close()
                    except Exception as e:
                        file_results.append((Path(src).name, 0, 0, str(e)))
                        continue

                    existing_primary = {
                        (r[0], r[1])
                        for r in out_conn.execute(
                            "SELECT session_id, source_file FROM applicants"
                        ).fetchall()
                    }
                    existing_fallback = {
                        (str(r[0]).strip().upper(), str(r[1]).strip().upper())
                        for r in out_conn.execute(
                            "SELECT applicant_name, source_file FROM applicants"
                        ).fetchall()
                    }

                    ins = skp = 0
                    for row in rows:
                        rd = dict(row)
                        pk = (rd.get("session_id", ""), rd.get("source_file", ""))
                        fk = (
                            str(rd.get("applicant_name") or "").strip().upper(),
                            str(rd.get("source_file")    or "").strip().upper(),
                        )
                        if pk in existing_primary or fk in existing_fallback:
                            skp += 1
                            continue
                        out_conn.execute(_INSERT, [rd.get(c) for c in _COLS])
                        existing_primary.add(pk)
                        existing_fallback.add(fk)
                        ins += 1

                    out_conn.commit()
                    file_results.append((Path(src).name, ins, skp, None))
                    total_inserted += ins
                    total_skipped  += skp

            self.after(0, lambda: _refresh_summary(self))

            msg  = "Merge complete.\n\n"
            msg += f"✓  Inserted : {total_inserted:,} record(s)\n"
            msg += f"–  Skipped  : {total_skipped:,} (duplicates)\n\nPer file:\n"
            for fname, ins, skp, err in file_results:
                if err:
                    msg += f"  ✗  {fname}  →  Error: {err}\n"
                else:
                    msg += f"  ✓  {fname}  →  {ins:,} inserted, {skp:,} skipped\n"

            self.after(0, lambda: (
                _flash_btn(self, self._gen_sum_merge_db_btn, "✓  Done!", 2500),
                messagebox.showinfo("Merge DB Result", msg)
            ))

        except Exception as exc:
            err = str(exc)
            self.after(0, lambda: (
                _flash_btn(self, self._gen_sum_merge_db_btn, "✗  Error", 3000),
                messagebox.showerror("Merge DB Error", err)
            ))

    threading.Thread(target=_worker, daemon=True).start()


# ═══════════════════════════════════════════════════════════════════════
#  MERGE EXCEL
# ═══════════════════════════════════════════════════════════════════════

def _merge_excel_files(self):
    paths = filedialog.askopenfilenames(
        title="Select Excel summaries to merge",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
    if not paths or len(paths) < 1:
        return

    out_path = filedialog.asksaveasfilename(
        title="Save merged Excel as…",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialfile=(
            f"Merged_General_Summary_"
            f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"))
    if not out_path:
        return

    _flash_btn(self, self._gen_sum_merge_xl_btn, "⟳  Merging…", 60_000)

    def _worker():
        _HEADERS = [
            "Applicant", "Residence Address", "Office Address",
            "Spouse / Employment", "Spouse Office Address",
            "Assets (Personal & Biz)",
            "Source of Income", "Total Source Of Income",
            "Business Expenses", "Total Business Expenses",
            "Household / Personal Expenses",
            "Total Household / Personal Expenses",
            "Total Net Income",
            "Total Amortization History", "Total Current Amortization",
        ]
        _MONETARY = {
            "Total Source Of Income", "Total Business Expenses",
            "Total Household / Personal Expenses",
            "Total Amortization History", "Total Current Amortization",
        }
        _NET_COL = "Total Net Income"
        _COL_WIDTHS = {
            "Applicant": 22, "Residence Address": 30, "Office Address": 26,
            "Spouse / Employment": 28, "Spouse Office Address": 26,
            "Assets (Personal & Biz)": 34,
            "Source of Income": 32, "Total Source Of Income": 22,
            "Business Expenses": 32, "Total Business Expenses": 22,
            "Household / Personal Expenses": 36,
            "Total Household / Personal Expenses": 24,
            "Total Net Income": 20,
            "Total Amortization History": 26,
            "Total Current Amortization": 26,
        }

        def _to_float(val):
            if val is None or val == "": return None
            try:
                return float(str(val).replace(",", "")
                             .replace("(", "-").replace(")", ""))
            except Exception:
                return None

        def _read_one(path):
            import openpyxl as _xl
            wb = _xl.load_workbook(str(path), read_only=True, data_only=True)
            ws = wb.active
            rows_iter  = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if header_row is None:
                wb.close(); return [], []
            file_hdrs = [str(h).strip() if h is not None else ""
                         for h in header_row]
            col_map   = {h: i for i, h in enumerate(file_hdrs) if h in _HEADERS}
            records   = []
            for row in rows_iter:
                first = str(row[0]).strip().upper() if row[0] is not None else ""
                if first == "TOTAL" or all(v is None for v in row):
                    continue
                rec = {h: (row[col_map[h]] if h in col_map else None)
                       for h in _HEADERS}
                records.append(rec)
            wb.close()
            return records, [h for h in _HEADERS if h not in col_map]

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            all_rows     = []
            seen_names   = set()
            dup_rows     = []
            file_results = []

            for path in paths:
                rows, missing = _read_one(path)
                ins = 0
                for row in rows:
                    name = str(row.get("Applicant") or "").strip().upper()
                    if name in seen_names:
                        dup_rows.append((name, Path(path).name))
                        continue
                    seen_names.add(name)
                    all_rows.append(row)
                    ins += 1
                file_results.append((Path(path).name, ins, missing))

            if not all_rows:
                raise ValueError("No data rows found across selected files.")

            hdr_fill  = PatternFill("solid", fgColor="93C47D")
            tot_fill  = PatternFill("solid", fgColor="D9EAD3")
            even_fill = PatternFill("solid", fgColor="FFFFFF")
            odd_fill  = PatternFill("solid", fgColor="F3F9F0")
            hdr_font  = Font(name="Roboto", bold=True, color="FFFFFF", size=10)
            body_font = Font(name="Roboto", size=9)
            bold_font = Font(name="Roboto", bold=True, size=9)
            net_font  = Font(name="Roboto", bold=True, size=9, color="1F6B28")
            tot_font  = Font(name="Roboto", bold=True, size=10)
            tot_font_j = Font(name="Roboto", bold=True, size=10, color="1F6B28")
            thin      = Side(style="thin",   color="CCCCCC")
            med       = Side(style="medium", color="555555")
            cell_bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
            tot_bdr   = Border(left=med,  right=med,  top=med,  bottom=med)
            wrap_al   = Alignment(horizontal="left",  vertical="top", wrap_text=True)
            right_al  = Alignment(horizontal="right", vertical="top")
            right_c   = Alignment(horizontal="right", vertical="center")
            left_c    = Alignment(horizontal="left",  vertical="center")
            CURRENCY  = '#,##0.00;(#,##0.00);"-"'
            SUM_COLS  = _MONETARY | {_NET_COL}

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "General Summary (Merged)"

            for ci, h in enumerate(_HEADERS, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.font = hdr_font; cell.fill = hdr_fill
                cell.alignment = wrap_al; cell.border = cell_bdr
                ws.column_dimensions[get_column_letter(ci)].width = \
                    _COL_WIDTHS.get(h, 20)
            ws.row_dimensions[1].height = 28

            for ri, row_dict in enumerate(all_rows, 2):
                bg_fill    = even_fill if (ri - 2) % 2 == 0 else odd_fill
                text_lines = []
                for ci, h in enumerate(_HEADERS, 1):
                    val  = row_dict.get(h)
                    fval = _to_float(val) if h in SUM_COLS else None
                    cell = ws.cell(row=ri, column=ci,
                                   value=fval if fval is not None
                                   else (val if val is not None else None))
                    cell.fill = bg_fill; cell.border = cell_bdr
                    if h == _NET_COL:
                        cell.number_format = CURRENCY
                        cell.font = net_font; cell.alignment = right_al
                    elif h in _MONETARY:
                        cell.number_format = CURRENCY
                        cell.font = bold_font; cell.alignment = right_al
                    elif h == "Applicant":
                        cell.font = bold_font; cell.alignment = wrap_al
                    else:
                        cell.font = body_font; cell.alignment = wrap_al
                        if val:
                            text_lines.append(len(str(val).split("\n")))
                ws.row_dimensions[ri].height = max(
                    18, min((max(text_lines) if text_lines else 1) * 15, 150))

            first_data = 2
            last_data  = len(all_rows) + 1
            tot_row    = last_data + 1
            for ci, h in enumerate(_HEADERS, 1):
                cell  = ws.cell(row=tot_row, column=ci)
                col_l = get_column_letter(ci)
                cell.fill = tot_fill; cell.border = tot_bdr
                if h == "Applicant":
                    cell.value = "TOTAL"
                    cell.font = tot_font; cell.alignment = left_c
                elif h == _NET_COL:
                    cell.value = f"=SUM({col_l}{first_data}:{col_l}{last_data})"
                    cell.number_format = CURRENCY
                    cell.font = tot_font_j; cell.alignment = right_c
                elif h in SUM_COLS:
                    cell.value = f"=SUM({col_l}{first_data}:{col_l}{last_data})"
                    cell.number_format = CURRENCY
                    cell.font = tot_font; cell.alignment = right_c
                else:
                    cell.font = tot_font; cell.alignment = left_c
            ws.row_dimensions[tot_row].height = 22
            ws.freeze_panes = "A2"

            wb.save(out_path)

            msg  = "Merge complete.\n\n"
            msg += f"✓  Total rows : {len(all_rows):,}\n"
            msg += f"–  Duplicates : {len(dup_rows):,} (skipped)\n\nPer file:\n"
            for fname, ins, missing in file_results:
                msg += f"  ✓  {fname}  →  {ins:,} rows added\n"
                if missing:
                    msg += (f"     ⚠ Missing cols (blanked): "
                            f"{', '.join(missing)}\n")
            if dup_rows:
                msg += "\nDuplicate applicants skipped (first file wins):\n"
                for name, src in dup_rows[:15]:
                    msg += f"  • {name}  (from {src})\n"
                if len(dup_rows) > 15:
                    msg += f"  … and {len(dup_rows) - 15} more\n"

            self.after(0, lambda: (
                _flash_btn(self, self._gen_sum_merge_xl_btn, "✓  Saved!", 2500),
                messagebox.showinfo("Merge Excel Result", msg)
            ))

        except ImportError:
            self.after(0, lambda: (
                _flash_btn(self, self._gen_sum_merge_xl_btn,
                           "openpyxl not installed", 3000),
                messagebox.showerror("Merge Excel Error",
                    "openpyxl is required.\nRun: pip install openpyxl")
            ))
        except Exception as exc:
            err = str(exc)
            self.after(0, lambda: (
                _flash_btn(self, self._gen_sum_merge_xl_btn, "✗  Error", 3000),
                messagebox.showerror("Merge Excel Error", err)
            ))

    threading.Thread(target=_worker, daemon=True).start()


# ═══════════════════════════════════════════════════════════════════════
#  EXPORT HELPERS
# ═══════════════════════════════════════════════════════════════════════

def _get_all_filtered_rows(self) -> list:
    search = _active_search(self)
    rows, _ = _db_query(
        search=search, session_id=self._gen_sum_session_filter,
        sort_col=self._gen_sum_sort_col, sort_asc=self._gen_sum_sort_asc,
        offset=0, limit=100_000)
    return [dict(r) for r in rows]


def _row_to_export_dict(row: dict) -> dict:
    def _fmt(val):
        try:    return float(val) if val not in (None, "") else None
        except: return None

    amort_total   = None
    spouse        = ""
    spouse_office = ""
    assets        = ""
    try:
        results_blob = json.loads(row.get("results_json", "") or "{}")
        amort_total  = _extract_amort_history_total(results_blob)

        def _if(key):
            d = results_blob.get(key, {})
            items = d.get("items", []) if isinstance(d, dict) else []
            return "  ·  ".join(items) if items else ""

        spouse        = _if("cibi_spouse")
        spouse_office = _if("cibi_spouse_office")
        personal      = _if("cibi_personal_assets")
        business      = _if("cibi_business_assets")
        if personal and business:
            assets = f"[Personal] {personal}  |  [Business] {business}"
        else:
            assets = personal or business
    except Exception:
        pass

    return {
        "Applicant":                           row.get("applicant_name",    "") or "",
        "Residence Address":                   row.get("residence_address", "") or "",
        "Office Address":                      row.get("office_address",    "") or "",
        "Spouse / Employment":                 spouse,
        "Spouse Office Address":               spouse_office,
        "Assets (Personal & Biz)":             assets,
        "Source of Income":                    row.get("income_items",      "") or "",
        "Total Source Of Income":              _fmt(row.get("income_total")),
        "Business Expenses":                   row.get("business_items",    "") or "",
        "Total Business Expenses":             _fmt(row.get("business_total")),
        "Household / Personal Expenses":       row.get("household_items",   "") or "",
        "Total Household / Personal Expenses": _fmt(row.get("household_total")),
        "Total Net Income":                    _fmt(row.get("net_income")),
        "Total Amortization History":          amort_total,
        "Total Current Amortization":          _fmt(row.get("amort_current_total")),
    }


def _export_csv(self):
    rows = _get_all_filtered_rows(self)
    if not rows:
        return
    path = filedialog.asksaveasfilename(
        title="Export to CSV", defaultextension=".csv",
        filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        initialfile=(
            f"General_Summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"))
    if not path:
        return
    flat    = [_row_to_export_dict(r) for r in rows]
    headers = list(flat[0].keys()) if flat else []
    try:
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(flat)
        _flash_btn(self, self._gen_sum_export_csv_btn, "✓  Saved!", 2000)
    except Exception as e:
        _flash_btn(self, self._gen_sum_export_csv_btn, f"Error: {e}", 3000)


def _export_excel(self):
    rows = _get_all_filtered_rows(self)
    if not rows:
        return
    path = filedialog.asksaveasfilename(
        title="Export to Excel", defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialfile=(
            f"General_Summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"))
    if not path:
        return

    def _worker():
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            flat    = [_row_to_export_dict(r) for r in rows]
            headers = list(flat[0].keys()) if flat else []

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "General Summary"

            hdr_fill  = PatternFill("solid", fgColor="93C47D")
            tot_fill  = PatternFill("solid", fgColor="D9EAD3")
            even_fill = PatternFill("solid", fgColor="FFFFFF")
            odd_fill  = PatternFill("solid", fgColor="F3F9F0")
            hdr_font  = Font(name="Roboto", bold=True, color="FFFFFF", size=10)
            body_font = Font(name="Roboto", size=9)
            bold_font = Font(name="Roboto", bold=True, size=9)
            net_font  = Font(name="Roboto", bold=True, size=9, color="1F6B28")
            tot_font  = Font(name="Roboto", bold=True, size=10)
            tot_font_j = Font(name="Roboto", bold=True, size=10, color="1F6B28")
            thin      = Side(style="thin",   color="CCCCCC")
            med       = Side(style="medium", color="555555")
            cell_bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
            tot_bdr   = Border(left=med,  right=med,  top=med,  bottom=med)
            wrap_al   = Alignment(horizontal="left",  vertical="top", wrap_text=True)
            right_al  = Alignment(horizontal="right", vertical="top")
            right_c   = Alignment(horizontal="right", vertical="center")
            left_c    = Alignment(horizontal="left",  vertical="center")
            CURRENCY  = '#,##0.00;(#,##0.00);"-"'

            TOTAL_COLS = {
                "Total Source Of Income",
                "Total Business Expenses",
                "Total Household / Personal Expenses",
                "Total Amortization History",
                "Total Current Amortization",
            }
            NET_COL  = "Total Net Income"
            SUM_COLS = TOTAL_COLS | {NET_COL}

            col_widths = {
                "Applicant": 22, "Residence Address": 30, "Office Address": 26,
                "Spouse / Employment": 28, "Spouse Office Address": 26,
                "Assets (Personal & Biz)": 34,
                "Source of Income": 32, "Total Source Of Income": 22,
                "Business Expenses": 32, "Total Business Expenses": 22,
                "Household / Personal Expenses": 36,
                "Total Household / Personal Expenses": 24,
                "Total Net Income": 20,
                "Total Amortization History": 26,
                "Total Current Amortization": 26,
            }

            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.font = hdr_font; cell.fill = hdr_fill
                cell.alignment = wrap_al; cell.border = cell_bdr
                ws.column_dimensions[get_column_letter(ci)].width = \
                    col_widths.get(h, 20)
            ws.row_dimensions[1].height = 28

            for ri, row_dict in enumerate(flat, 2):
                bg_fill    = even_fill if (ri - 2) % 2 == 0 else odd_fill
                text_lines = []
                for ci, h in enumerate(headers, 1):
                    val  = row_dict.get(h)
                    cell = ws.cell(row=ri, column=ci,
                                   value=val if val is not None else None)
                    cell.fill = bg_fill; cell.border = cell_bdr
                    if h == NET_COL:
                        cell.number_format = CURRENCY
                        cell.font = net_font; cell.alignment = right_al
                    elif h in TOTAL_COLS:
                        cell.number_format = CURRENCY
                        cell.font = bold_font; cell.alignment = right_al
                    elif h == "Applicant":
                        cell.font = bold_font; cell.alignment = wrap_al
                    else:
                        cell.font = body_font; cell.alignment = wrap_al
                        if val:
                            text_lines.append(len(str(val).split("\n")))
                max_lines = max(text_lines) if text_lines else 1
                ws.row_dimensions[ri].height = max(18, min(max_lines * 15, 150))

            first_data = 2
            last_data  = len(flat) + 1
            tot_row    = last_data + 1

            for ci, h in enumerate(headers, 1):
                cell  = ws.cell(row=tot_row, column=ci)
                col_l = get_column_letter(ci)
                cell.fill = tot_fill; cell.border = tot_bdr
                if h == "Applicant":
                    cell.value = "TOTAL"
                    cell.font = tot_font; cell.alignment = left_c
                elif h == NET_COL:
                    cell.value = f"=SUM({col_l}{first_data}:{col_l}{last_data})"
                    cell.number_format = CURRENCY
                    cell.font = tot_font_j; cell.alignment = right_c
                elif h in SUM_COLS:
                    cell.value = f"=SUM({col_l}{first_data}:{col_l}{last_data})"
                    cell.number_format = CURRENCY
                    cell.font = tot_font; cell.alignment = right_c
                else:
                    cell.font = tot_font; cell.alignment = left_c
            ws.row_dimensions[tot_row].height = 22
            ws.freeze_panes = "A2"

            wb.save(path)
            self.after(0, lambda: _flash_btn(
                self, self._gen_sum_export_xl_btn, "✓  Saved!", 2000))
        except ImportError:
            self.after(0, lambda: _flash_btn(
                self, self._gen_sum_export_xl_btn, "openpyxl not installed", 3000))
        except Exception as e:
            self.after(0, lambda err=str(e): _flash_btn(
                self, self._gen_sum_export_xl_btn, f"Error: {err[:40]}", 3000))

    threading.Thread(target=_worker, daemon=True).start()


# ═══════════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════════

def _flash_btn(self, btn, msg: str, ms: int):
    try:
        orig = btn.cget("text")
        btn.configure(text=msg)
        self.after(ms, lambda: (
            btn.configure(text=orig) if btn.winfo_exists() else None))
    except Exception:
        pass


def lookup_summary_notify(self):
    """
    Called by general_lookup after each file is saved to the General DB.
    If the General Summary tab is active, do a full refresh.
    Otherwise just update the stats pills.
    """
    if getattr(self, "_current_tab", "") == "general_summary":
        _refresh_summary(self)
    elif hasattr(self, "_gen_sum_stat_labels"):
        _update_stats(self)


# ═══════════════════════════════════════════════════════════════════════
#  ATTACH
# ═══════════════════════════════════════════════════════════════════════

def attach(cls):
    cls._build_general_summary_panel = _build_general_summary_panel
    cls.general_lookup_summary_notify  = lookup_summary_notify