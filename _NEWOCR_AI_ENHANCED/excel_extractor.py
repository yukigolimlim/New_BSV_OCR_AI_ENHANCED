"""
excel_extractor.py — DocExtract Pro
=====================================
Gemini 2.5 Flash → structured JSON → openpyxl → populated .xlsx

Pipeline
--------
  1. PDF/image file  →  convert pages to base64 images
  2. Gemini 2.5 Flash (multimodal) →  structured JSON of all fields
  3. Gemini 2.5 Flash (text)       →  map JSON fields → Excel cell addresses
  4. openpyxl                      →  write values into template or new workbook
  5. Return saved .xlsx path

Public API
----------
  extract_to_excel(file_path, api_key, template_path=None,
                   progress_cb=None)  →  Path
      Full pipeline. Returns path to the saved .xlsx file.

  extract_fields_json(file_path, api_key,
                      progress_cb=None)  →  dict
      Step 1+2 only. Returns raw extracted JSON dict.

  json_to_excel(data, api_key, template_path=None,
                output_path=None)  →  Path
      Step 3+4 only. Takes an already-extracted dict and writes Excel.
"""

from __future__ import annotations

import base64
import json
import os
import re
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Optional

# ── Output folder (Desktop / DocExtract_Files) ───────────────────────────────
def _output_dir() -> Path:
    desktop = Path.home() / "Desktop"
    if not desktop.exists():
        desktop = Path.home()
    out = desktop / "DocExtract_Files"
    out.mkdir(parents=True, exist_ok=True)
    return out


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 1 — FILE → BASE64 IMAGES
# ══════════════════════════════════════════════════════════════════════════════

def _file_to_images(file_path: str) -> list[dict]:
    """
    Convert a PDF or image file into a list of base64-encoded image dicts.
    Each dict: {"mime_type": "image/png", "data": "<base64>"}

    For PDFs  → each page becomes one PNG image (via pdf2image / pymupdf).
    For images → returned as-is (converted to PNG).
    """
    p = Path(file_path)
    ext = p.suffix.lower()
    images = []

    IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".bmp", ".tiff", ".tif", ".webp", ".gif"}

    if ext == ".pdf":
        # Try pymupdf first (faster, no external binary needed)
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(str(p))
            for page in doc:
                pix = page.get_pixmap(dpi=150)
                png_bytes = pix.tobytes("png")
                images.append({
                    "mime_type": "image/png",
                    "data": base64.b64encode(png_bytes).decode("utf-8"),
                })
            doc.close()
            return images
        except ImportError:
            pass

        # Fallback: pdf2image (needs poppler)
        try:
            from pdf2image import convert_from_path
            import io
            pages = convert_from_path(str(p), dpi=150)
            for page in pages:
                buf = io.BytesIO()
                page.save(buf, format="PNG")
                images.append({
                    "mime_type": "image/png",
                    "data": base64.b64encode(buf.getvalue()).decode("utf-8"),
                })
            return images
        except Exception:
            pass

        raise RuntimeError(
            "Cannot convert PDF to images.\n"
            "Install either:  pip install pymupdf\n"
            "           or:   pip install pdf2image  (+ poppler)"
        )

    elif ext in IMAGE_EXTS:
        try:
            from PIL import Image
            import io
            img = Image.open(str(p)).convert("RGB")
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            images.append({
                "mime_type": "image/png",
                "data": base64.b64encode(buf.getvalue()).decode("utf-8"),
            })
            return images
        except ImportError:
            # No PIL — read raw bytes and guess mime
            mime = {
                ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
                ".png": "image/png",  ".webp": "image/webp",
                ".gif": "image/gif",  ".bmp":  "image/bmp",
            }.get(ext, "image/png")
            images.append({
                "mime_type": mime,
                "data": base64.b64encode(p.read_bytes()).decode("utf-8"),
            })
            return images

    raise ValueError(f"Unsupported file type for image conversion: {ext}")


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 2 — GEMINI FLASH → STRUCTURED JSON
# ══════════════════════════════════════════════════════════════════════════════

_EXTRACT_SYSTEM = """
You are a precise document data extractor for Banco San Vicente (BSV), a Philippine rural bank.

Your task: extract EVERY data field visible in the document image(s) and return them
as a single flat or nested JSON object.

RULES:
1. Return ONLY valid JSON — no markdown fences, no explanations, no preamble.
2. Use the EXACT field names/labels as they appear in the document (translated to English if in Filipino).
3. For tables, return an array of row objects.
4. For financial figures, preserve the raw value as a string (e.g., "3,000,000") AND
   include a separate "_numeric" key with the plain number (e.g., 3000000).
5. If a field is blank or unreadable, use null.
6. Preserve dates in their original format (e.g., "26/03/1975").
7. Group related fields logically:
   - "borrower"      → personal info
   - "employment"    → work/income info
   - "loans"         → array of loan records
   - "credit_cards"  → array of card records
   - "legal"         → adverse/legal info
   - "collateral"    → asset/guarantee info
   - "co_borrower"   → co-borrower info
   - "references"    → character references
   - "financials"    → income/expense figures
   Use "other" for anything that doesn't fit.

EXAMPLE OUTPUT SHAPE (do not copy literally — use actual document fields):
{
  "document_type": "CIC Credit Report",
  "report_date": "14/05/2015",
  "borrower": {
    "full_name": "EMERSON REJANO MADAYAG REYES III",
    "date_of_birth": "26/03/1975",
    "tin": "123789123"
  },
  "employment": {
    "employer": "SOLAR TRADE",
    "gross_income": "3,000,000",
    "gross_income_numeric": 3000000,
    "frequency": "Annual"
  },
  "loans": [
    {
      "contract_code": "M00000628",
      "type": "Mortgage/Real Estate",
      "financed_amount": "100,000",
      "financed_amount_numeric": 100000,
      "outstanding": "96,666",
      "outstanding_numeric": 96666,
      "overdue": "6,668",
      "overdue_numeric": 6668,
      "status": "Active"
    }
  ]
}
""".strip()

_EXTRACT_USER = (
    "Extract all data fields from the attached document page(s) and return "
    "structured JSON as instructed. Do not add any text outside the JSON."
)


def _extract_text_from_structured(file_path: str) -> str:
    """
    Read text/tabular content directly from .xlsx, .xls, .csv, .docx, .txt, .md
    files — no image conversion needed.
    Returns a plain-text representation suitable for the Gemini text prompt.
    """
    p   = Path(file_path)
    ext = p.suffix.lower()

    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(p), data_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                parts.append(f"=== Sheet: {sheet_name} ===")
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(c.strip() for c in cells):
                        parts.append("\t".join(cells))
            return "\n".join(parts)
        except ImportError:
            # Fallback: pandas
            import pandas as pd
            xl = pd.ExcelFile(str(p))
            parts = []
            for sheet in xl.sheet_names:
                df = xl.parse(sheet)
                parts.append(f"=== Sheet: {sheet} ===")
                parts.append(df.to_string(index=False))
            return "\n\n".join(parts)

    if ext == ".csv":
        return p.read_text(encoding="utf-8", errors="replace")

    if ext in (".txt", ".md"):
        return p.read_text(encoding="utf-8", errors="replace")

    if ext in (".docx", ".doc"):
        try:
            import docx
            doc = docx.Document(str(p))
            return "\n".join(para.text for para in doc.paragraphs if para.text.strip())
        except ImportError:
            raise RuntimeError("pip install python-docx  to handle .docx files")

    raise ValueError(f"_extract_text_from_structured: unsupported extension {ext}")


# File types that can be read directly as text (no image conversion)
_TEXT_EXTS = {".xlsx", ".xls", ".csv", ".txt", ".md", ".docx", ".doc"}


def extract_fields_json(
    file_path: str,
    api_key: str,
    progress_cb: Optional[Callable] = None,
) -> dict:
    """
    Send the document to Gemini Flash and return a structured dict of all fields.

    Routing logic:
      • .xlsx / .xls / .csv / .txt / .md / .docx  → read as text, send as text prompt
      • .pdf / images                              → convert to base64 images
    """
    def _cb(pct, msg=""):
        if progress_cb:
            progress_cb(pct, msg)

    p   = Path(file_path)
    ext = p.suffix.lower()

    try:
        from google import genai as _genai
        from google.genai import types as _gtypes
    except ImportError:
        raise RuntimeError(
            "google-genai package not installed.\n"
            "Run:  pip install google-genai"
        )

    client = _genai.Client(api_key=api_key)

    # ── Branch A: structured/text files — no image conversion needed ─────
    if ext in _TEXT_EXTS:
        _cb(10, f"Reading {ext.lstrip('.')} file directly…")
        raw_text = _extract_text_from_structured(file_path)
        _cb(30, "Sending to Gemini Flash…")

        user_msg = (
            f"Document content ({ext.lstrip('.')} file):\n\n"
            f"{raw_text[:24_000]}\n\n"  # Flash text limit safety trim
            + _EXTRACT_USER
        )
        resp = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=[_gtypes.Content(
                role="user",
                parts=[_gtypes.Part(text=user_msg)]
            )],
            config=_gtypes.GenerateContentConfig(
                system_instruction=_EXTRACT_SYSTEM,
                max_output_tokens=65536,
            ),
        )
        raw = _safe_text(resp)
        _cb(70, "Parsing JSON response…")
        return _parse_json(raw)

    # ── Branch B: PDF / images — convert to base64 first ─────────────────
    _cb(5, "Converting file to images…")
    images = _file_to_images(file_path)
    total_pages = len(images)

    try:
        from google import genai as _genai
        from google.genai import types as _gtypes
    except ImportError:
        raise RuntimeError(
            "google-genai package not installed.\n"
            "Run:  pip install google-genai"
        )

    client = _genai.Client(api_key=api_key)

    # ── If single page, send in one call ─────────────────────────────────
    if total_pages == 1:
        _cb(20, "Sending to Gemini Flash…")
        parts = [
            _gtypes.Part(
                inline_data=_gtypes.Blob(
                    mime_type=images[0]["mime_type"],
                    data=base64.b64decode(images[0]["data"]),
                )
            ),
            _gtypes.Part(text=_EXTRACT_USER),
        ]
        resp = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=[_gtypes.Content(role="user", parts=parts)],
            config=_gtypes.GenerateContentConfig(
                system_instruction=_EXTRACT_SYSTEM,
                max_output_tokens=65536,
            ),
        )
        raw = _safe_text(resp)
        _cb(70, "Parsing JSON response…")
        return _parse_json(raw)

    # ── Multi-page: batch all pages in one call (up to 10) ───────────────
    # Batching saves API quota vs. one call per page
    _cb(15, f"Batching {total_pages} pages into Gemini Flash…")
    parts = []
    for i, img in enumerate(images[:10]):  # Flash supports up to 10 inline images
        parts.append(_gtypes.Part(text=f"[Page {i+1} of {total_pages}]"))
        parts.append(
            _gtypes.Part(
                inline_data=_gtypes.Blob(
                    mime_type=img["mime_type"],
                    data=base64.b64decode(img["data"]),
                )
            )
        )
    parts.append(_gtypes.Part(
        text=_EXTRACT_USER + "\nMerge all pages into one unified JSON object."
    ))

    _cb(30, "Waiting for Gemini Flash response…")
    resp = client.models.generate_content(
        model="gemini-2.5-flash",
        contents=[_gtypes.Content(role="user", parts=parts)],
        config=_gtypes.GenerateContentConfig(
            system_instruction=_EXTRACT_SYSTEM,
            max_output_tokens=65536,
        ),
    )
    raw = _safe_text(resp)
    _cb(70, "Parsing JSON response…")
    return _parse_json(raw)


def _safe_text(resp) -> str:
    try:
        return resp.text or ""
    except Exception:
        try:
            return "".join(
                p.text for p in resp.candidates[0].content.parts
                if hasattr(p, "text") and p.text
            )
        except Exception:
            return ""


def _parse_json(raw: str) -> dict:
    """
    Strip markdown fences and parse JSON.
    If the response was truncated (common with large forms), attempt to
    salvage it by closing unclosed braces before giving up.
    """
    clean = re.sub(r"^```(?:json)?\s*", "", raw.strip(), flags=re.I)
    clean = re.sub(r"\s*```$", "", clean.strip())

    # ── Attempt 1: clean parse ────────────────────────────────────────────
    try:
        return json.loads(clean)
    except json.JSONDecodeError:
        pass

    # ── Attempt 2: find the outermost {...} block ─────────────────────────
    m = re.search(r"\{[\s\S]+\}", clean)
    if m:
        try:
            return json.loads(m.group(0))
        except json.JSONDecodeError:
            pass

    # ── Attempt 3: truncated JSON — close unclosed braces/brackets ────────
    # Count open vs closed braces to figure out how many closers are missing
    truncated = m.group(0) if m else clean
    # Strip trailing comma + whitespace before we close
    truncated = re.sub(r",\s*$", "", truncated.rstrip())
    open_braces   = truncated.count("{") - truncated.count("}")
    open_brackets = truncated.count("[") - truncated.count("]")
    # Close in reverse order of what's open (brackets first, then braces)
    suffix = "]" * max(0, open_brackets) + "}" * max(0, open_braces)
    try:
        recovered = json.loads(truncated + suffix)
        recovered["_truncated"] = True   # flag so caller knows it was partial
        return recovered
    except json.JSONDecodeError:
        pass

    # ── Attempt 4: extract key:value pairs with regex as last resort ───────
    pairs = re.findall(r'"([^"]+)"\s*:\s*"([^"]*)"', clean)
    if pairs:
        result = {k: v for k, v in pairs}
        result["_parse_error"] = True
        result["_recovered_pairs"] = len(pairs)
        return result

    return {"_raw": raw[:2000], "_parse_error": True}


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 3 — JSON → CELL MAPPING (via Gemini Flash)
# ══════════════════════════════════════════════════════════════════════════════

_MAPPING_SYSTEM = """
You are an Excel cell-mapping assistant for Banco San Vicente (BSV).

Given:
  A) A JSON object of extracted document fields
  B) A list of Excel column headers / cell labels from the target template
     (or "NEW_WORKBOOK" if no template is used)

Your task: return a JSON array where each element maps one extracted value
to a specific Excel cell address.

Output format — return ONLY a JSON array, no markdown, no explanation:
[
  {"cell": "B5",  "value": "EMERSON REJANO MADAYAG REYES III", "label": "Borrower Name"},
  {"cell": "B6",  "value": "26/03/1975",                       "label": "Date of Birth"},
  {"cell": "C10", "value": 3000000,                            "label": "Gross Income"},
  ...
]

RULES:
1. If a template is provided, map fields to the NEAREST matching template cell.
2. If no template (NEW_WORKBOOK), lay out the data logically:
   - Column A = field label, Column B = field value
   - Group related fields with blank rows between sections
   - Start at row 3 (rows 1-2 are reserved for the document title)
   - Arrays/tables start at the next available row after the scalar fields,
     with their own header row.
3. Never map "_numeric" shadow keys — use the human-readable string value instead,
   except for cells that need numeric values for formulas (mark those explicitly).
4. Skip null values.
5. For array data (loans, credit_cards, etc.), return one mapping entry per cell,
   using the correct row number for each array item.
""".strip()


def _build_mapping(
    data: dict,
    api_key: str,
    template_headers: Optional[list[str]] = None,
) -> list[dict]:
    """
    Ask Gemini Flash to map extracted JSON fields → Excel cell addresses.
    Returns list of {"cell": "B5", "value": ..., "label": ...} dicts.
    """
    try:
        from google import genai as _genai
        from google.genai import types as _gtypes
    except ImportError:
        raise RuntimeError("pip install google-genai")

    client = _genai.Client(api_key=api_key)

    template_info = (
        f"Template headers / labels:\n{json.dumps(template_headers, indent=2)}"
        if template_headers
        else "Template: NEW_WORKBOOK (no template provided)"
    )

    user_msg = (
        f"Extracted data JSON:\n{json.dumps(data, indent=2, ensure_ascii=False)}\n\n"
        f"{template_info}\n\n"
        "Return the cell mapping array as instructed."
    )

    resp = client.models.generate_content(
        model="gemini-2.5-flash",
        contents=[_gtypes.Content(
            role="user",
            parts=[_gtypes.Part(text=user_msg)]
        )],
        config=_gtypes.GenerateContentConfig(
            system_instruction=_MAPPING_SYSTEM,
            max_output_tokens=4096,
        ),
    )
    raw = _safe_text(resp)
    clean = re.sub(r"^```(?:json)?\s*", "", raw.strip(), flags=re.I)
    clean = re.sub(r"\s*```$", "", clean.strip())
    try:
        result = json.loads(clean)
        if isinstance(result, list):
            return result
    except json.JSONDecodeError:
        m = re.search(r"\[[\s\S]+\]", clean)
        if m:
            try:
                return json.loads(m.group(0))
            except json.JSONDecodeError:
                pass
    return []


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 4 — WRITE TO EXCEL (openpyxl)
# ══════════════════════════════════════════════════════════════════════════════

from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter, column_index_from_string


# BSV colour palette
_NAVY      = "1C2E7A"
_NAVY_MID  = "3858C8"
_LIME      = "A8D818"
_LIME_PALE = "F2FFCC"
_GHOST     = "E8EEFF"
_WHITE     = "FFFFFF"
_BORDER    = "B0BEEC"
_MUTED     = "9AAACE"


def _thin_border(color=_BORDER) -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _write_excel(
    mapping: list[dict],
    data: dict,
    template_path: Optional[str],
    output_path: Path,
    doc_type: str = "Extracted Document",
) -> Path:
    """
    Write the cell mapping into an Excel workbook.
    If template_path is given, load it first; otherwise create a new workbook.
    """
    if template_path and Path(template_path).exists():
        wb = load_workbook(template_path)
        ws = wb.active
        _apply_template_mapping(ws, mapping)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Data"
        _build_new_workbook(ws, mapping, data, doc_type)

    # Add a metadata sheet
    _add_metadata_sheet(wb, doc_type)

    wb.save(str(output_path))
    return output_path


def _apply_template_mapping(ws, mapping: list[dict]) -> None:
    """Write mapped values into a pre-existing template sheet."""
    for entry in mapping:
        cell_addr = entry.get("cell", "")
        value     = entry.get("value")
        if not cell_addr or value is None:
            continue
        try:
            cell = ws[cell_addr]
            # Don't overwrite formula cells in the template
            if isinstance(cell.value, str) and cell.value.startswith("="):
                continue
            cell.value = value
            # Blue text for hardcoded inputs (industry standard)
            cell.font = Font(
                name="Arial", size=10,
                color="0000FF",  # blue = user input
            )
        except Exception:
            pass


def _build_new_workbook(
    ws,
    mapping: list[dict],
    data: dict,
    doc_type: str,
) -> None:
    """Build a clean, formatted BSV-style workbook from scratch."""

    # ── Title rows ────────────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = f"BSV — {doc_type}"
    title_cell.font  = Font(name="Arial", size=14, bold=True, color=_WHITE)
    title_cell.fill  = PatternFill("solid", start_color=_NAVY)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:D2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Extracted: {datetime.now().strftime('%d %b %Y  %H:%M')}"
    sub_cell.font  = Font(name="Arial", size=9, italic=True, color=_MUTED)
    sub_cell.fill  = PatternFill("solid", start_color=_GHOST)
    sub_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # ── Column widths ─────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    # ── Write mappings ────────────────────────────────────────────────────
    # Group by section (detect from label prefix or cell row gaps)
    current_section = None

    for entry in mapping:
        cell_addr = entry.get("cell", "")
        value     = entry.get("value")
        label     = entry.get("label", "")
        if not cell_addr or value is None:
            continue

        try:
            cell_obj = ws[cell_addr]
            # Determine column
            col_letter = re.match(r"([A-Z]+)", cell_addr.upper())
            col = col_letter.group(1) if col_letter else "B"

            if col == "A":
                # This is a label cell
                cell_obj.value = value
                cell_obj.font  = Font(name="Arial", size=10, bold=True, color=_NAVY_MID)
                cell_obj.fill  = PatternFill("solid", start_color=_GHOST)
                cell_obj.alignment = Alignment(horizontal="left", wrap_text=True)
                cell_obj.border = _thin_border()
            else:
                # This is a value cell
                cell_obj.value = value
                cell_obj.font  = Font(name="Arial", size=10, color="0000FF")
                cell_obj.fill  = PatternFill("solid", start_color=_WHITE)
                cell_obj.alignment = Alignment(horizontal="left", wrap_text=True)
                cell_obj.border = _thin_border()

                # Also write the label in column A of the same row
                row_num = cell_obj.row
                label_cell = ws.cell(row=row_num, column=1)
                if not label_cell.value:
                    label_cell.value = label
                    label_cell.font  = Font(name="Arial", size=10, bold=True,
                                            color=_NAVY_MID)
                    label_cell.fill  = PatternFill("solid", start_color=_GHOST)
                    label_cell.alignment = Alignment(horizontal="left", wrap_text=True)
                    label_cell.border = _thin_border()

        except Exception:
            pass

    # ── Section headers (detect by scanning for section-level keys) ───────
    _add_section_headers(ws, data)


def _add_section_headers(ws, data: dict) -> None:
    """
    Scan the workbook for rows where a new document section starts
    and insert a coloured section header row.
    This is a best-effort pass — it doesn't reorder existing rows.
    """
    section_labels = {
        "borrower":     "BORROWER INFORMATION",
        "employment":   "EMPLOYMENT & INCOME",
        "loans":        "LOAN CONTRACTS",
        "credit_cards": "CREDIT CARDS",
        "legal":        "LEGAL & ADVERSE INFORMATION",
        "collateral":   "COLLATERAL & GUARANTEES",
        "financials":   "FINANCIAL SUMMARY",
        "co_borrower":  "CO-BORROWER",
        "references":   "CHARACTER REFERENCES",
    }
    # Find rows that contain section-level label text and add fill
    for row in ws.iter_rows():
        for cell in row:
            if not cell.value:
                continue
            val_lower = str(cell.value).lower().strip()
            for key, header in section_labels.items():
                if key in val_lower or header.lower() in val_lower:
                    # Style this row as a section header
                    ws.merge_cells(
                        start_row=cell.row, start_column=1,
                        end_row=cell.row,   end_column=4
                    )
                    header_cell = ws.cell(row=cell.row, column=1)
                    header_cell.value = header
                    header_cell.font  = Font(
                        name="Arial", size=11, bold=True, color=_WHITE
                    )
                    header_cell.fill = PatternFill("solid", start_color=_NAVY_MID)
                    header_cell.alignment = Alignment(
                        horizontal="left", vertical="center"
                    )
                    ws.row_dimensions[cell.row].height = 22
                    break


def _add_metadata_sheet(wb: Workbook, doc_type: str) -> None:
    """Add a hidden metadata sheet with extraction info."""
    if "Metadata" in wb.sheetnames:
        return
    ms = wb.create_sheet("Metadata")
    ms.sheet_state = "hidden"
    ms["A1"] = "Generated by"
    ms["B1"] = "DocExtract Pro — Banco San Vicente"
    ms["A2"] = "Document Type"
    ms["B2"] = doc_type
    ms["A3"] = "Extraction Date"
    ms["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ms["A4"] = "Pipeline"
    ms["B4"] = "Gemini 2.5 Flash → JSON → openpyxl"


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 5 — TEMPLATE SCANNER (reads headers from an existing .xlsx template)
# ══════════════════════════════════════════════════════════════════════════════

def _scan_template_headers(template_path: str) -> list[str]:
    """
    Read an existing .xlsx template and return a flat list of
    non-empty cell labels/headers found in the first 60 rows.
    """
    headers = []
    try:
        wb = load_workbook(template_path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(max_row=60, values_only=True):
            for val in row:
                if val and isinstance(val, str) and val.strip():
                    headers.append(val.strip())
        wb.close()
    except Exception:
        pass
    return headers


# ══════════════════════════════════════════════════════════════════════════════
#  PUBLIC API — FULL PIPELINE
# ══════════════════════════════════════════════════════════════════════════════

def extract_to_excel(
    file_path:     str,
    api_key:       str,
    template_path: Optional[str] = None,
    output_stem:   Optional[str] = None,
    progress_cb:   Optional[Callable] = None,
) -> Path:
    """
    Full pipeline: PDF/image → Gemini Flash JSON → cell mapping → .xlsx

    Parameters
    ----------
    file_path     : path to the source PDF or image
    api_key       : Gemini API key
    template_path : optional path to a BSV Excel template (.xlsx)
    output_stem   : filename stem for the output file (no extension)
    progress_cb   : optional callback(pct: int, stage: str)

    Returns
    -------
    Path to the saved .xlsx file.
    """
    def _cb(pct, msg=""):
        if progress_cb:
            progress_cb(pct, msg)

    # ── Step 1+2: extract JSON ─────────────────────────────────────────────
    _cb(0, "Extracting fields with Gemini Flash…")
    data = extract_fields_json(file_path, api_key, progress_cb=progress_cb)

    doc_type = data.get("document_type", Path(file_path).stem)

    # ── Step 3: build cell mapping ─────────────────────────────────────────
    _cb(75, "Mapping fields to Excel cells…")
    template_headers = None
    if template_path and Path(template_path).exists():
        template_headers = _scan_template_headers(template_path)

    mapping = _build_mapping(data, api_key, template_headers)

    if not mapping:
        # Fallback: build a simple flat mapping without the AI mapper
        _cb(80, "Building fallback mapping…")
        mapping = _flat_fallback_mapping(data)

    # ── Step 4: write Excel ────────────────────────────────────────────────
    _cb(88, "Writing Excel file…")
    stem = output_stem or Path(file_path).stem
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = _output_dir() / f"{stem}_extracted_{timestamp}.xlsx"

    _write_excel(mapping, data, template_path, out_path, doc_type)

    _cb(100, "Done!")
    return out_path


def json_to_excel(
    data:          dict,
    api_key:       str,
    template_path: Optional[str] = None,
    output_path:   Optional[Path] = None,
) -> Path:
    """
    Step 3+4 only. Takes an already-extracted dict and writes to Excel.
    Useful when you already have extracted JSON from a previous call.
    """
    doc_type = data.get("document_type", "Extracted Document")
    template_headers = None
    if template_path and Path(template_path).exists():
        template_headers = _scan_template_headers(template_path)

    mapping = _build_mapping(data, api_key, template_headers)
    if not mapping:
        mapping = _flat_fallback_mapping(data)

    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = _output_dir() / f"bsv_extracted_{timestamp}.xlsx"

    return _write_excel(mapping, data, template_path, output_path, doc_type)


# ══════════════════════════════════════════════════════════════════════════════
#  FALLBACK MAPPING (no AI mapper — pure Python flat layout)
# ══════════════════════════════════════════════════════════════════════════════

def _flat_fallback_mapping(data: dict, start_row: int = 3) -> list[dict]:
    """
    Generate a simple A=label / B=value mapping without calling the AI mapper.
    Used as a safety net when the AI mapping call fails.
    """
    entries = []
    row = start_row

    def _add(label: str, value: Any) -> None:
        nonlocal row
        if value is None:
            return
        entries.append({"cell": f"A{row}", "value": label,  "label": label})
        entries.append({"cell": f"B{row}", "value": value,  "label": label})
        row += 1

    def _recurse(obj: Any, prefix: str = "") -> None:
        nonlocal row
        if isinstance(obj, dict):
            for k, v in obj.items():
                if k.endswith("_numeric") or k == "_raw":
                    continue
                full_key = f"{prefix}{k}".replace("_", " ").title()
                if isinstance(v, (dict, list)):
                    # Section header
                    entries.append({
                        "cell": f"A{row}",
                        "value": full_key.upper(),
                        "label": "__section__",
                    })
                    row += 1
                    _recurse(v, prefix=f"{k}_")
                    row += 1  # blank row after section
                else:
                    _add(full_key, v)
        elif isinstance(obj, list):
            for i, item in enumerate(obj):
                entries.append({
                    "cell": f"A{row}",
                    "value": f"Item {i+1}",
                    "label": "__section__",
                })
                row += 1
                _recurse(item, prefix="")
                row += 1

    _recurse(data)
    return entries