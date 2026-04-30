"""
lu_simulator_patch.py — Risk Simulator Tab
============================================
Inflation / Cost-Shock Simulator.

This file is the CANONICAL simulator implementation.
It supersedes _sim_* functions in lu_ui.py and the simulator patches
in lu_analysis_patch_v8.py.  Apply it AFTER lu_tab_analysis.attach(cls).

Standalone: imports only lu_core and lu_shared.
Attached to app class via attach(cls).

Key improvements over original lu_ui
--------------------------------------
  • Expense mix shown as a matplotlib pie (% of total simulated spend).
  • Chunked row construction via after() to keep UI responsive.
  • SIM_MAX_ROWS cap (50) to prevent widget explosion.
  • Respects active sector filter from lu_shared.
"""

import tkinter as tk
from tkinter import messagebox, filedialog
import tkinter.ttk as ttk
import customtkinter as ctk
import re

from lu_core import GENERAL_CLIENT, _RISK_ORDER, _fmt_value, get_high_risk_industries
from lu_shared import (
    F, FF, _bind_mousewheel,
    _NAVY_DEEP, _NAVY_MID, _NAVY_LIGHT, _NAVY_MIST, _NAVY_GHOST, _NAVY_PALE,
    _WHITE, _CARD_WHITE, _OFF_WHITE, _BORDER_LIGHT, _BORDER_MID,
    _TXT_NAVY, _TXT_SOFT, _TXT_MUTED, _TXT_ON_LIME,
    _LIME_MID, _LIME_DARK, _LIME_PALE,
    _ACCENT_RED, _ACCENT_GOLD, _ACCENT_SUCCESS,
    _RISK_COLOR, _RISK_BG, _RISK_BADGE_BG,
    _lu_filter_data_by_query,
    _lu_get_active_sectors, _lu_get_filtered_all_data,
    LU_CLIENT_TREE_SPEC, lu_format_lu_cell,
)

# ── Tuneable constants ──────────────────────────────────────────────
SIM_MAX_ROWS       = 50   # max expense rows shown
SIM_CHUNK_SIZE     = 10   # rows built per after() tick
SIM_CHART_MAX_BARS = 20   # max expense rows considered for chart
PIE_MAX_SLICES     = 10   # pie slices before aggregating to "Other"
SIM_TABLE_COLUMNS = (
    # (title, min_width_px, weight)
    ("Expense Item", 220, 5),
    ("Risk", 72, 1),
    ("Base Amount", 120, 2),
    ("Inflation Rate (%)", 100, 2),
    ("Extra Cost", 120, 2),
    ("Simulated", 120, 2),
)

SIM_CLIENT_TABLE_COLUMNS = (
    # (col_id, heading, min_width_px, anchor)
    ("client",               "Client Name",           210, "w"),
    ("industry",             "Industry",              150, "w"),
    ("base_total_expenses",  "Total Expenses (Base)", 150, "e"),
    ("sim_total_expenses", "Total Expenses (Sim)", 150, "e"),
    ("net_income", "Total Net Income (Base)", 160, "e"),
    ("sim_net_income", "Total Net Income (Simulated)", 185, "e"),
    ("pct_increase", "% Increase", 90, "center"),
    ("sim_increase", "Simulated Increase", 150, "e"),
    ("current_amort", "Total Current Amort", 150, "e"),
    ("pct_net_to_amort", "% Net → Amort", 120, "center"),
    ("sim_risk_label", "Risk Label", 90, "center"),
    ("risk_reasoning", "Risk Reasoning", 280, "w"),
)

SIM_CLIENT_PAGE_SIZE = 10

# ── Default risk range boundaries (% Net → Amort) ───────────────────
# Each entry: (min_inclusive, max_inclusive)
# Stored on the app instance as self._sim_risk_ranges at runtime.
_SIM_DEFAULT_RISK_RANGES = {
    "LOW":    (1.0,  35.0),
    "MEDIUM": (36.0, 70.0),
    "HIGH":   (71.0, float("inf")),
}

def _apply_sim_client_tree_style():
    """
    Match the Summary tab Treeview look-and-feel.
    (Same row height, heading padding, clam theme, and selection colors.)
    """
    style = ttk.Style()
    try:
        style.theme_use("clam")
    except Exception:
        pass
    style.configure(
        "SimSummary.Treeview",
        background=_WHITE,
        foreground=_TXT_NAVY,
        fieldbackground=_WHITE,
        rowheight=36,
        font=("Segoe UI", 9),
        borderwidth=0,
        relief="flat",
    )
    style.configure(
        "SimSummary.Treeview.Heading",
        background=_NAVY_DEEP,
        foreground=_WHITE,
        font=("Segoe UI", 9, "bold"),
        relief="flat",
        borderwidth=0,
        padding=(10, 8),
    )
    style.map(
        "SimSummary.Treeview.Heading",
        background=[("active", _NAVY_LIGHT)],
        relief=[("active", "flat")],
    )
    style.map(
        "SimSummary.Treeview",
        background=[("selected", "#C8E6C9")],
        foreground=[("selected", _NAVY_DEEP)],
    )

try:
    import matplotlib
    matplotlib.use("TkAgg")
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    _HAS_MPL = True
except ImportError:
    _HAS_MPL = False


# ══════════════════════════════════════════════════════════════════════
#  PANEL BUILDER (sets fixed canvas height)
# ══════════════════════════════════════════════════════════════════════

def _build_simulator_panel(self, parent):
    hdr = tk.Frame(parent, bg=_NAVY_MID, height=38)
    hdr.pack(fill="x")
    hdr.pack_propagate(False)
    self._sim_hdr_lbl = tk.Label(
        hdr, text="⚙️  Inflation / Cost-Shock Simulator",
        font=F(10, "bold"), fg=_WHITE, bg=_NAVY_MID)
    self._sim_hdr_lbl.pack(side="left", padx=20, pady=8)

    ctrl = tk.Frame(parent, bg=_OFF_WHITE, height=58)
    ctrl.pack(fill="x")
    ctrl.pack_propagate(False)
    # kept as attrs for compatibility with _sim_apply_global / _sim_reset / _sim_populate
    self._sim_global_var = tk.StringVar(value="0")
    self._sim_search_var = tk.StringVar()
    self._sim_match_lbl = tk.Label(
        ctrl, text="", font=F(8, "bold"), fg=_WHITE, bg=_OFF_WHITE, padx=8, pady=3)
    ctk.CTkButton(
        ctrl,
        text="⧉  Expense Simulator Table",
        command=lambda: _sim_open_expense_table_window(self),
        width=200,
        height=32,
        corner_radius=5,
        fg_color=_NAVY_LIGHT,
        hover_color=_NAVY_MID,
        text_color=_WHITE,
        font=FF(10, "bold"),
    ).pack(side="left", padx=(16, 0), pady=13)
    self._sim_industry_filter_btn = ctk.CTkButton(
        ctrl,
        text="Industry Checklist",
        command=lambda: _sim_open_industry_checklist(self),
        width=150,
        height=32,
        corner_radius=5,
        fg_color=_NAVY_LIGHT,
        hover_color=_NAVY_MID,
        text_color=_WHITE,
        font=FF(10, "bold"),
    )
    self._sim_industry_filter_btn.pack(side="left", padx=(8, 0), pady=10)
    self._sim_industry_filter_lbl = tk.Label(
        ctrl, text="", font=F(8, "bold"), fg=_TXT_SOFT, bg=_OFF_WHITE
    )
    self._sim_industry_filter_lbl.pack(side="left", padx=(8, 0), pady=10)
    ctk.CTkButton(
        ctrl,
        text="⚖  Risk Ranges",
        command=lambda: _sim_open_risk_ranges_dialog(self),
        width=130,
        height=32,
        corner_radius=5,
        fg_color=_NAVY_LIGHT,
        hover_color=_NAVY_MID,
        text_color=_WHITE,
        font=FF(10, "bold"),
    ).pack(side="left", padx=(8, 0), pady=10)
    tk.Frame(parent, bg=_BORDER_LIGHT, height=1).pack(fill="x")

    cards_frame = tk.Frame(parent, bg=_NAVY_MIST)
    cards_frame.pack(fill="x")
    _build_sim_summary_cards(self, cards_frame)

    # inc_bar removed — surplus/deficit label no longer displayed here.
    # Store dummy label attrs so _sim_refresh does not crash on hasattr checks.
    self._sim_income_lbl  = tk.Label(parent, text="")
    self._sim_surplus_lbl = tk.Label(parent, text="")

    # ── Outer scrollable body — one canvas scrolls everything ──────────
    body = tk.Frame(parent, bg=_CARD_WHITE)
    body.pack(fill="both", expand=True)

    _outer_vsb = tk.Scrollbar(body, orient="vertical", relief="flat",
                               troughcolor=_OFF_WHITE, bg=_BORDER_LIGHT, width=10, bd=0)
    _outer_vsb.pack(side="right", fill="y")
    _outer_canvas = tk.Canvas(body, bg=_CARD_WHITE, highlightthickness=0,
                               yscrollcommand=_outer_vsb.set)
    _outer_canvas.pack(side="left", fill="both", expand=True)
    _outer_vsb.config(command=_outer_canvas.yview)

    _outer_frame = tk.Frame(_outer_canvas, bg=_CARD_WHITE)
    _outer_win = _outer_canvas.create_window((0, 0), window=_outer_frame, anchor="nw")

    def _outer_on_frame_configure(e):
        _outer_canvas.configure(scrollregion=_outer_canvas.bbox("all"))
    _outer_frame.bind("<Configure>", _outer_on_frame_configure)

    def _outer_on_canvas_configure(e):
        _outer_canvas.itemconfig(_outer_win, width=e.width)
    _outer_canvas.bind("<Configure>", _outer_on_canvas_configure)

    def _outer_mousewheel(event):
        _outer_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
    _outer_canvas.bind("<Enter>", lambda _e: _outer_canvas.bind_all("<MouseWheel>", _outer_mousewheel))
    _outer_canvas.bind("<Leave>", lambda _e: _outer_canvas.unbind_all("<MouseWheel>"))

    # ── CLIENT IMPACT SECTION ─────────────────────────────────────────
    # Two stacked sub-frames inside a shared container.
    # Only one is visible at a time:
    #   _sim_client_table_view  — the treeview + pagination (default)
    #   _sim_client_detail_view — the full single-client panel with Back btn
    # ──────────────────────────────────────────────────────────────────
    _client_section = tk.Frame(_outer_frame, bg=_CARD_WHITE)
    _client_section.pack(fill="x", expand=False)

    # Section header bar removed. Back button is now rendered inside the detail
    # view content by _sim_show_client_details. Keep dummy attrs so show/hide
    # helpers do not crash.
    self._sim_section_title_lbl = tk.Label(parent, text="")
    self._sim_back_btn = None

    # ═══════════════════════════════════════════════════════════════════
    #  TABLE VIEW  (treeview + search + pagination)
    # ═══════════════════════════════════════════════════════════════════
    self._sim_client_table_view = tk.Frame(_client_section, bg=_CARD_WHITE)
    self._sim_client_table_view.pack(fill="x", expand=False)
    table_wrap = self._sim_client_table_view  # alias so existing code below works

    # Header row: label on left, search bar on right
    hdr_row = tk.Frame(table_wrap, bg=_CARD_WHITE)
    hdr_row.pack(fill="x", padx=20, pady=(10, 4))
    tk.Label(
        hdr_row,
        text="Client Impact (updates when you ramp expenses)",
        font=F(9, "bold"),
        fg=_TXT_SOFT,
        bg=_CARD_WHITE,
    ).pack(side="left")
    self._sim_client_search_var = tk.StringVar()
    _search_frame = tk.Frame(hdr_row, bg=_CARD_WHITE)
    _search_frame.pack(side="right")
    tk.Label(
        _search_frame,
        text="🔍",
        font=F(10),
        fg=_TXT_MUTED,
        bg=_CARD_WHITE,
    ).pack(side="left", padx=(0, 4))
    _client_search_entry = ctk.CTkEntry(
        _search_frame,
        textvariable=self._sim_client_search_var,
        width=220,
        height=28,
        corner_radius=6,
        fg_color=_WHITE,
        text_color=_TXT_NAVY,
        border_color=_BORDER_MID,
        font=FF(10),
        placeholder_text="Search client name…",
    )
    _client_search_entry.pack(side="left")
    def _client_search_refresh():
        self._sim_client_page = 0
        _sim_populate(self)
    _client_search_entry.bind("<KeyRelease>", lambda _e: _client_search_refresh())
    ctk.CTkButton(
        _search_frame,
        text="✕",
        width=28,
        height=28,
        corner_radius=6,
        fg_color=_OFF_WHITE,
        hover_color=_BORDER_MID,
        text_color=_TXT_MUTED,
        font=FF(9, "bold"),
        border_width=1,
        border_color=_BORDER_MID,
        command=lambda: [
            self._sim_client_search_var.set(""),
            _client_search_refresh(),
        ],
    ).pack(side="left", padx=(4, 0))

    tk.Label(
        table_wrap,
        text="Risk Label is based on % Net Income to Amortization. Use ⚖ Risk Ranges to configure thresholds.",
        font=F(7),
        fg=_TXT_MUTED,
        bg=_CARD_WHITE,
    ).pack(anchor="w", padx=20, pady=(0, 6))

    # Page controls
    pg_row = tk.Frame(table_wrap, bg=_CARD_WHITE)
    pg_row.pack(fill="x", padx=20, pady=(0, 6))
    self._sim_client_prev_btn = ctk.CTkButton(
        pg_row,
        text="◀ Prev",
        width=70,
        height=26,
        corner_radius=6,
        fg_color=_WHITE,
        hover_color=_NAVY_MIST,
        text_color=_NAVY_MID,
        font=FF(9, "bold"),
        border_width=1,
        border_color=_BORDER_MID,
        command=lambda: _sim_client_page_prev(self),
        state="disabled",
    )
    self._sim_client_prev_btn.pack(side="left")
    self._sim_client_page_lbl = tk.Label(
        pg_row,
        text="Page 1",
        font=F(8),
        fg=_TXT_SOFT,
        bg=_CARD_WHITE,
    )
    self._sim_client_page_lbl.pack(side="left", padx=10)
    self._sim_client_next_btn = ctk.CTkButton(
        pg_row,
        text="Next ▶",
        width=70,
        height=26,
        corner_radius=6,
        fg_color=_WHITE,
        hover_color=_NAVY_MIST,
        text_color=_NAVY_MID,
        font=FF(9, "bold"),
        border_width=1,
        border_color=_BORDER_MID,
        command=lambda: _sim_client_page_next(self),
        state="disabled",
    )
    self._sim_client_next_btn.pack(side="left")

    self._sim_export_btn = ctk.CTkButton(
        pg_row,
        text="💾  Simulator Clients Excel",
        width=100,
        height=26,
        corner_radius=6,
        fg_color=_WHITE,
        hover_color=_NAVY_MIST,
        text_color=_NAVY_MID,
        font=FF(9, "bold"),
        border_width=1,
        border_color=_BORDER_MID,
        command=lambda: _sim_export_client_impact_excel(self),
    )
    self._sim_export_btn.pack(side="right", padx=(0, 10))
    self._sim_export_high_risk_btn = ctk.CTkButton(
        pg_row,
        text="🔴  HIGH Risk Clients Excel",
        width=100,
        height=26,
        corner_radius=6,
        fg_color="#FFF0F0",
        hover_color="#FFD6D6",
        text_color=_ACCENT_RED,
        font=FF(9, "bold"),
        border_width=1,
        border_color=_ACCENT_RED,
        command=lambda: _sim_export_high_risk_clients_excel(self),
    )
    self._sim_export_high_risk_btn.pack(side="right", padx=(0, 6))
    self._sim_merge_excel_btn = ctk.CTkButton(
        pg_row,
        text="🧩  Merge Excel",
        width=92,
        height=26,
        corner_radius=6,
        fg_color="#EEF3FA",
        hover_color="#DCE7F6",
        text_color=_NAVY_MID,
        font=FF(9, "bold"),
        border_width=1,
        border_color=_BORDER_MID,
        command=lambda: _sim_merge_excel_files(self),
    )
    self._sim_merge_excel_btn.pack(side="right", padx=(0, 6))
    self._sim_client_count_lbl = tk.Label(
        pg_row,
        text="",
        font=F(8),
        fg=_TXT_MUTED,
        bg=_CARD_WHITE,
    )
    self._sim_client_count_lbl.pack(side="right")

    # Treeview
    tbl_box = tk.Frame(table_wrap, bg=_CARD_WHITE)
    tbl_box.pack(fill="x", padx=20, pady=(0, 16))

    hsb = ttk.Scrollbar(tbl_box, orient="horizontal")
    hsb.pack(side="bottom", fill="x")

    _apply_sim_client_tree_style()
    self._sim_client_tree = ttk.Treeview(
        tbl_box,
        columns=tuple(c[0] for c in SIM_CLIENT_TABLE_COLUMNS),
        show="headings",
        height=10,
        style="SimSummary.Treeview",
        xscrollcommand=hsb.set,
        selectmode="browse",
    )
    self._sim_client_tree.pack(side="top", fill="x")
    hsb.config(command=self._sim_client_tree.xview)

    self._sim_client_tree.bind(
        "<ButtonRelease-1>",
        lambda e: _sim_on_client_impact_row_activated(self, e),
    )
    self._sim_client_tree.bind(
        "<Return>",
        lambda e: _sim_on_client_impact_row_activated(self, e),
    )

    for col_id, heading, min_px, anchor in SIM_CLIENT_TABLE_COLUMNS:
        self._sim_client_tree.heading(col_id, text=heading)
        stretch = (col_id == "client")
        self._sim_client_tree.column(col_id, width=min_px, minwidth=min_px, anchor=anchor, stretch=stretch)

    # Risk tags
    self._sim_client_tree.tag_configure("HIGH",   background="#FFF5F5", foreground=_ACCENT_RED)
    self._sim_client_tree.tag_configure("MEDIUM", background="#FFFBF0", foreground=_ACCENT_GOLD)
    self._sim_client_tree.tag_configure("LOW",    background="#F0FBE8", foreground=_ACCENT_SUCCESS)
    self._sim_client_tree.tag_configure("NA",     background=_WHITE,    foreground=_TXT_MUTED)

    # Hover tag — blue highlight matching Analysis tab
    self._sim_client_tree.tag_configure(
        "row_hover",
        background="#D6E8FF",
        foreground=_NAVY_DEEP,
    )

    # Hand cursor
    self._sim_client_tree.configure(cursor="hand2")

    # Hover motion bindings
    _sim_tree_hovered_iid = [None]
    self._sim_client_tree._sim_hovered_iid_ref = _sim_tree_hovered_iid

    def _sim_tree_on_motion(event, _tree=self._sim_client_tree,
                            _hovered=_sim_tree_hovered_iid):
        iid = _tree.identify_row(event.y)
        prev = _hovered[0]
        if iid == prev:
            return
        if prev and prev in _tree.get_children(""):
            tags = [t for t in _tree.item(prev, "tags") if t != "row_hover"]
            _tree.item(prev, tags=tags)
        if iid and iid in _tree.get_children(""):
            tags = [t for t in _tree.item(iid, "tags") if t != "row_hover"]
            tags.append("row_hover")
            _tree.item(iid, tags=tags)
        _hovered[0] = iid

    def _sim_tree_on_leave(event, _tree=self._sim_client_tree,
                           _hovered=_sim_tree_hovered_iid):
        prev = _hovered[0]
        if prev and prev in _tree.get_children(""):
            tags = [t for t in _tree.item(prev, "tags") if t != "row_hover"]
            _tree.item(prev, tags=tags)
        _hovered[0] = None

    self._sim_client_tree.bind("<Motion>", _sim_tree_on_motion)
    self._sim_client_tree.bind("<Leave>",  _sim_tree_on_leave)

    # Hint strip at bottom of table view
    _tbl_hint = tk.Frame(table_wrap, bg=_NAVY_MIST,
                         highlightbackground=_BORDER_MID, highlightthickness=1)
    _tbl_hint.pack(fill="x", padx=20, pady=(0, 16))
    tk.Label(
        _tbl_hint,
        text="👆  Click any row to view the full client details.",
        font=F(8),
        fg=_TXT_SOFT,
        bg=_NAVY_MIST,
        anchor="w",
        padx=14,
        pady=8,
    ).pack(anchor="w")

    # ═══════════════════════════════════════════════════════════════════
    #  DETAIL VIEW  (single client — hidden until a row is clicked)
    # ═══════════════════════════════════════════════════════════════════
    self._sim_client_detail_view = tk.Frame(_client_section, bg=_CARD_WHITE)
    # NOT packed yet — shown on click, hidden on Back
    self._sim_client_detail_frame = self._sim_client_detail_view   # alias for _sim_show_client_details

    # Bottom padding spacer
    tk.Frame(_outer_frame, bg=_CARD_WHITE, height=20).pack(fill="x")

    self._sim_sliders    = {}
    self._sim_expenses   = []
    self._sim_build_job  = None
    self._sim_expenses_capped = False
    self._sim_expense_win = None
    self._sim_expense_search_var = tk.StringVar()
    self._sim_canvas = None
    self._sim_scroll_frame = None
    self._sim_net_income = 0.0
    self._sim_recs = []
    self._sim_client_page = 0
    self._sim_selected_industries = set()
    self._sim_manual_industries = set()
    self._sim_chart_holder = None  # chart removed; kept for compatibility
    _sim_show_placeholder(self)
def _sim_filter_data_by_industry_checklist(all_data, selected_industries):
    selected = {str(x).strip().lower() for x in (selected_industries or set()) if str(x).strip()}
    if not selected:
        return all_data

    splitter = re.compile(r"\s*(?:,|/|;|&|\band\b)\s*", re.I)

    def _industry_tokens(rec: dict) -> set[str]:
        tags = rec.get("industry_tags") or []
        if tags:
            return {str(x).strip().lower() for x in tags if str(x).strip()}
        raw = str((rec or {}).get("industry") or "").strip()
        if not raw:
            return set()
        return {tok.strip().lower() for tok in splitter.split(raw) if tok.strip()}

    base_general = list((all_data or {}).get("general", []))
    kept_general = [
        rec for rec in base_general
        if _industry_tokens(rec) & selected
    ]
    kept_clients = {str((r or {}).get("client") or ""): r for r in kept_general if (r or {}).get("client")}

    patched = dict(all_data or {})
    patched["general"] = kept_general
    patched["clients"] = kept_clients
    return patched


def _sim_open_industry_checklist(self):
    all_data = getattr(self, "_lu_all_data", None) or {}
    base_industries = {str(x).strip() for x in all_data.get("unique_industries", []) if str(x).strip()}
    manual_industries = set(getattr(self, "_sim_manual_industries", set()) or set())
    industries = sorted(base_industries | manual_industries, key=str.lower)
    if not industries:
        messagebox.showwarning("No Data", "Load and run LU analysis first.")
        return

    selected = set(getattr(self, "_sim_selected_industries", set()) or set())
    high_defaults = {str(x).strip().lower() for x in get_high_risk_industries() if str(x).strip()}
    if not selected:
        selected = {name for name in industries if name.lower() in high_defaults}

    dialog = ctk.CTkToplevel(self)
    dialog.title("Risk Simulator Industry Checklist")
    dialog.geometry("620x620")
    dialog.minsize(520, 480)
    dialog.transient(self)
    dialog.grab_set()
    dialog.configure(fg_color=_CARD_WHITE)

    hdr = tk.Frame(dialog, bg=_NAVY_MID, height=52)
    hdr.pack(fill="x")
    hdr.pack_propagate(False)
    tk.Label(
        hdr,
        text="☑  Risk Simulator Industry Checklist",
        font=F(11, "bold"),
        fg=_WHITE,
        bg=_NAVY_MID,
    ).pack(side="left", padx=16, pady=12)

    info = tk.Frame(dialog, bg=_NAVY_MIST, highlightbackground=_BORDER_MID, highlightthickness=1)
    info.pack(fill="x", padx=16, pady=(10, 6))
    tk.Label(
        info,
        text=(
            "Use checkboxes to filter the Risk Simulator by industry. "
            "By default, this follows HIGH industries from Risk Settings."
        ),
        font=F(8),
        fg=_TXT_SOFT,
        bg=_NAVY_MIST,
        anchor="w",
        justify="left",
    ).pack(fill="x", padx=10, pady=8)

    search_row = tk.Frame(dialog, bg=_CARD_WHITE)
    search_row.pack(fill="x", padx=16, pady=(4, 6))
    tk.Label(search_row, text="🔍", font=F(10), fg=_TXT_SOFT, bg=_CARD_WHITE).pack(side="left")
    search_var = tk.StringVar(value="")
    search_entry = ctk.CTkEntry(
        search_row,
        textvariable=search_var,
        width=380,
        height=28,
        corner_radius=4,
        fg_color=_WHITE,
        text_color=_TXT_NAVY,
        border_color=_BORDER_MID,
        font=FF(9),
        placeholder_text="Search industry...",
    )
    search_entry.pack(side="left", fill="x", expand=True, padx=(6, 0))

    add_row = tk.Frame(dialog, bg=_CARD_WHITE)
    add_row.pack(fill="x", padx=16, pady=(0, 6))
    tk.Label(
        add_row, text="Add Industry:", font=F(8, "bold"),
        fg=_NAVY_MID, bg=_CARD_WHITE
    ).pack(side="left", padx=(0, 6))
    add_var = tk.StringVar(value="")
    add_entry = ctk.CTkEntry(
        add_row,
        textvariable=add_var,
        width=260,
        height=28,
        corner_radius=4,
        fg_color=_WHITE,
        text_color=_TXT_NAVY,
        border_color=_BORDER_MID,
        font=FF(9),
        placeholder_text="e.g. Logistics",
    )
    add_entry.pack(side="left", padx=(0, 6))
    add_hint_lbl = tk.Label(add_row, text="", font=F(7), fg=_TXT_MUTED, bg=_CARD_WHITE)
    add_hint_lbl.pack(side="left", padx=(4, 0))

    list_wrap = tk.Frame(dialog, bg=_CARD_WHITE)
    list_wrap.pack(fill="both", expand=True, padx=16, pady=(0, 8))
    sb = tk.Scrollbar(list_wrap, relief="flat", troughcolor=_OFF_WHITE, bg=_BORDER_LIGHT, width=8, bd=0)
    sb.pack(side="right", fill="y")
    canvas = tk.Canvas(list_wrap, bg=_CARD_WHITE, highlightthickness=0, yscrollcommand=sb.set)
    canvas.pack(side="left", fill="both", expand=True)
    sb.config(command=canvas.yview)
    rows_frame = tk.Frame(canvas, bg=_CARD_WHITE)
    win = canvas.create_window((0, 0), window=rows_frame, anchor="nw")
    rows_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.bind("<Configure>", lambda e: canvas.itemconfig(win, width=e.width))
    canvas.bind("<Enter>", lambda _e: canvas.bind_all("<MouseWheel>", lambda ev: canvas.yview_scroll(int(-1 * (ev.delta / 120)), "units")))
    canvas.bind("<Leave>", lambda _e: canvas.unbind_all("<MouseWheel>"))

    row_widgets = []
    var_map = {}
    row_by_industry = {}

    def _add_industry_row(industry: str, preselect: bool = False):
        if industry in var_map:
            if preselect:
                var_map[industry].set(True)
            return
        idx = len(row_widgets)
        row_bg = _WHITE if idx % 2 == 0 else _OFF_WHITE
        row = tk.Frame(rows_frame, bg=row_bg)
        row.pack(fill="x")
        tk.Frame(row, bg=_BORDER_LIGHT, height=1).pack(fill="x")
        inner = tk.Frame(row, bg=row_bg)
        inner.pack(fill="x", padx=8, pady=4)
        var = tk.BooleanVar(value=preselect)
        var_map[industry] = var
        chk = tk.Checkbutton(
            inner,
            text=industry,
            variable=var,
            onvalue=True,
            offvalue=False,
            font=F(9),
            fg=_TXT_NAVY,
            bg=row_bg,
            activebackground=row_bg,
            anchor="w",
            justify="left",
            relief="flat",
            highlightthickness=0,
        )
        chk.pack(side="left", fill="x", expand=True, padx=(4, 0))
        row_widgets.append((row, industry))
        row_by_industry[industry] = row

    for industry in industries:
        _add_industry_row(industry, preselect=(industry in selected))

    def _apply_search(*_args):
        q = search_var.get().strip().lower()
        for row, industry in row_widgets:
            show = (not q) or (q in industry.lower())
            if show and not row.winfo_ismapped():
                row.pack(fill="x")
            elif (not show) and row.winfo_ismapped():
                row.pack_forget()

    search_var.trace_add("write", _apply_search)

    def _add_manual_industry():
        name = add_var.get().strip()
        if not name:
            add_hint_lbl.config(text="Enter an industry name.", fg=_ACCENT_RED)
            return
        existing = {k.lower(): k for k in var_map.keys()}
        if name.lower() in existing:
            _add_industry_row(existing[name.lower()], preselect=True)
            add_hint_lbl.config(text="Already exists; selected.", fg=_ACCENT_SUCCESS)
        else:
            _add_industry_row(name, preselect=True)
            add_hint_lbl.config(text=f"Added '{name}'.", fg=_ACCENT_SUCCESS)
        add_var.set("")
        _apply_search()
        add_entry.focus_set()

    def _remove_manual_industry():
        name = add_var.get().strip()
        if not name:
            add_hint_lbl.config(text="Enter an industry to remove.", fg=_ACCENT_RED)
            return
        existing = {k.lower(): k for k in var_map.keys()}
        canonical = existing.get(name.lower())
        if not canonical:
            add_hint_lbl.config(text="Industry not found.", fg=_ACCENT_RED)
            return
        if canonical in base_industries:
            add_hint_lbl.config(text="Cannot remove base industry from data.", fg=_ACCENT_RED)
            return
        row = row_by_industry.pop(canonical, None)
        if row is not None:
            try:
                row.destroy()
            except Exception:
                pass
        row_widgets[:] = [(r, n) for (r, n) in row_widgets if n != canonical]
        var_map.pop(canonical, None)
        selected.discard(canonical)
        add_hint_lbl.config(text=f"Removed '{canonical}'.", fg=_ACCENT_SUCCESS)
        add_var.set("")
        _apply_search()
        add_entry.focus_set()

    tk.Button(
        add_row,
        text="Add",
        font=F(8, "bold"),
        fg=_WHITE,
        bg=_NAVY_MID,
        activebackground=_NAVY_LIGHT,
        activeforeground=_WHITE,
        relief="flat",
        bd=0,
        padx=10,
        pady=5,
        cursor="hand2",
        command=_add_manual_industry,
    ).pack(side="left")
    tk.Button(
        add_row,
        text="Remove",
        font=F(8, "bold"),
        fg=_WHITE,
        bg=_ACCENT_RED,
        activebackground="#C53030",
        activeforeground=_WHITE,
        relief="flat",
        bd=0,
        padx=10,
        pady=5,
        cursor="hand2",
        command=_remove_manual_industry,
    ).pack(side="left", padx=(4, 0))
    add_entry.bind("<Return>", lambda _e: _add_manual_industry())

    def _set_all(v: bool):
        for _name, vv in var_map.items():
            vv.set(v)

    def _use_high_defaults():
        for name, vv in var_map.items():
            vv.set(name.lower() in high_defaults)

    def _apply_and_close():
        chosen = {name for name, vv in var_map.items() if vv.get()}
        self._sim_selected_industries = chosen
        self._sim_manual_industries = set(var_map.keys()) - base_industries
        _sim_populate(self)
        dialog.destroy()

    footer = tk.Frame(dialog, bg=_OFF_WHITE, highlightbackground=_BORDER_MID, highlightthickness=1)
    footer.pack(fill="x", padx=16, pady=(2, 14))
    tk.Button(
        footer,
        text="Select All",
        font=F(8, "bold"),
        fg=_TXT_NAVY,
        bg=_WHITE,
        relief="flat",
        bd=0,
        padx=10,
        pady=6,
        cursor="hand2",
        command=lambda: _set_all(True),
    ).pack(side="left", padx=(12, 4), pady=8)
    tk.Button(
        footer,
        text="Clear All",
        font=F(8, "bold"),
        fg=_TXT_SOFT,
        bg=_WHITE,
        relief="flat",
        bd=0,
        padx=10,
        pady=6,
        cursor="hand2",
        command=lambda: _set_all(False),
    ).pack(side="left", padx=4, pady=8)
    tk.Button(
        footer,
        text="Use HIGH from Settings",
        font=F(8, "bold"),
        fg=_ACCENT_RED,
        bg="#FFE8E8",
        relief="flat",
        bd=0,
        padx=10,
        pady=6,
        cursor="hand2",
        command=_use_high_defaults,
    ).pack(side="left", padx=4, pady=8)
    tk.Button(
        footer,
        text="Cancel",
        font=F(9),
        fg=_TXT_SOFT,
        bg=_OFF_WHITE,
        relief="flat",
        bd=0,
        padx=10,
        pady=8,
        cursor="hand2",
        command=dialog.destroy,
    ).pack(side="right", padx=(0, 4), pady=8)
    tk.Button(
        footer,
        text="  ✔  Apply Filter  ",
        font=F(9, "bold"),
        fg=_WHITE,
        bg=_NAVY_MID,
        activebackground=_NAVY_LIGHT,
        activeforeground=_WHITE,
        relief="flat",
        bd=0,
        padx=14,
        pady=8,
        cursor="hand2",
        command=_apply_and_close,
    ).pack(side="right", padx=12, pady=8)






def _sim_open_risk_ranges_dialog(self):
    """
    Settings dialog for configuring custom risk range boundaries.
    Mirrors the style of the Industry Checklist dialog.

    Stores result on self._sim_risk_ranges as:
        {"LOW": (min, max), "MEDIUM": (min, max), "HIGH": (min, max)}
    where HIGH max is float('inf').

    Validates that no two ranges share any overlapping percentage value.
    """
    # Load current ranges (or defaults)
    current = dict(getattr(self, "_sim_risk_ranges", None) or _SIM_DEFAULT_RISK_RANGES)

    dialog = ctk.CTkToplevel(self)
    dialog.title("Risk Range Settings")
    dialog.geometry("500x440")
    dialog.minsize(460, 400)
    dialog.resizable(False, False)
    dialog.transient(self)
    dialog.grab_set()
    dialog.configure(fg_color=_CARD_WHITE)

    # ── Header ────────────────────────────────────────────────────────
    hdr = tk.Frame(dialog, bg=_NAVY_MID, height=52)
    hdr.pack(fill="x")
    hdr.pack_propagate(False)
    tk.Label(
        hdr,
        text="⚖  Risk Range Settings",
        font=F(11, "bold"),
        fg=_WHITE,
        bg=_NAVY_MID,
    ).pack(side="left", padx=16, pady=12)

    # ── Info banner ───────────────────────────────────────────────────
    info = tk.Frame(dialog, bg=_NAVY_MIST, highlightbackground=_BORDER_MID, highlightthickness=1)
    info.pack(fill="x", padx=16, pady=(12, 6))
    tk.Label(
        info,
        text=(
            "Set the % Net → Amort boundaries for each risk level.\n"
            "HIGH max is always open-ended (any value above HIGH min is HIGH).\n"
            "Ranges must not overlap — each percentage point belongs to exactly one level."
        ),
        font=F(8),
        fg=_TXT_SOFT,
        bg=_NAVY_MIST,
        anchor="w",
        justify="left",
        wraplength=440,
    ).pack(fill="x", padx=10, pady=8)

    # ── Warning label (shown when overlap detected) ────────────────────
    warn_var = tk.StringVar(value="")
    warn_lbl = tk.Label(
        dialog,
        textvariable=warn_var,
        font=F(8, "bold"),
        fg=_ACCENT_RED,
        bg="#FFF5F5",
        anchor="w",
        justify="left",
        wraplength=440,
        padx=10,
        pady=6,
        relief="flat",
    )
    # packed only when there is a warning — we pack/forget it dynamically

    # ── Range input rows ──────────────────────────────────────────────
    RISKS = [
        ("LOW",    _ACCENT_SUCCESS, "#F0FBE8"),
        ("MEDIUM", _ACCENT_GOLD,   "#FFFBF0"),
        ("HIGH",   _ACCENT_RED,    "#FFF5F5"),
    ]

    fields: dict[str, dict] = {}   # "LOW" / "MEDIUM" / "HIGH" → {"min": var, "max": var}

    rows_frame = tk.Frame(dialog, bg=_CARD_WHITE)
    rows_frame.pack(fill="x", padx=16, pady=(4, 0))

    for risk, badge_color, row_bg in RISKS:
        lo, hi = current.get(risk, _SIM_DEFAULT_RISK_RANGES[risk])
        hi_display = "" if hi == float("inf") else str(int(hi) if hi == int(hi) else hi)
        lo_display = str(int(lo) if lo == int(lo) else lo)

        row = tk.Frame(rows_frame, bg=row_bg,
                       highlightbackground=_BORDER_MID, highlightthickness=1)
        row.pack(fill="x", pady=4)

        # Badge
        badge = tk.Frame(row, bg=badge_color, width=70)
        badge.pack(side="left", fill="y")
        badge.pack_propagate(False)
        tk.Label(
            badge,
            text=risk,
            font=F(9, "bold"),
            fg=badge_color if risk == "LOW" else badge_color,
            bg=row_bg,
            anchor="center",
        ).pack(expand=True)
        # Re-apply correct bg
        badge.configure(bg=row_bg)

        # Risk label as colored text
        tk.Label(
            row,
            text=risk,
            font=F(10, "bold"),
            fg=badge_color,
            bg=row_bg,
            width=8,
            anchor="w",
            padx=10,
        ).pack(side="left")

        # Min field
        tk.Label(row, text="Min %", font=F(8), fg=_TXT_SOFT, bg=row_bg, padx=(6)).pack(side="left")
        min_var = tk.StringVar(value=lo_display)
        min_entry = ctk.CTkEntry(
            row,
            textvariable=min_var,
            width=72,
            height=30,
            corner_radius=4,
            fg_color=_WHITE,
            text_color=_TXT_NAVY,
            border_color=badge_color,
            font=FF(10),
        )
        min_entry.pack(side="left", padx=(4, 8))

        # Max field
        if risk == "HIGH":
            tk.Label(row, text="Max %  ∞ (open-ended)", font=F(8), fg=_TXT_MUTED, bg=row_bg, padx=6).pack(side="left")
            max_var = tk.StringVar(value="")
            max_entry = None
        else:
            tk.Label(row, text="Max %", font=F(8), fg=_TXT_SOFT, bg=row_bg, padx=6).pack(side="left")
            max_var = tk.StringVar(value=hi_display)
            max_entry = ctk.CTkEntry(
                row,
                textvariable=max_var,
                width=72,
                height=30,
                corner_radius=4,
                fg_color=_WHITE,
                text_color=_TXT_NAVY,
                border_color=badge_color,
                font=FF(10),
            )
            max_entry.pack(side="left", padx=(4, 8))

        fields[risk] = {"min": min_var, "max": max_var, "min_entry": min_entry,
                        "max_entry": max_entry, "bg": row_bg, "color": badge_color}

    # ── Live validation ───────────────────────────────────────────────
    def _parse_field(var: tk.StringVar, allow_empty: bool = False):
        """Return float or None on parse error. Empty string → None if allow_empty."""
        s = var.get().strip()
        if not s and allow_empty:
            return None
        try:
            return float(s)
        except ValueError:
            return None

    def _validate(*_args):
        """
        Parse all fields, detect overlaps, and update warn_lbl.
        Returns (ranges_dict, error_msg) — error_msg is "" if valid.
        """
        errors = []
        parsed: dict[str, tuple] = {}

        for risk, fd in fields.items():
            lo = _parse_field(fd["min"])
            if lo is None:
                errors.append(f"{risk}: Min % must be a number.")
                continue
            if risk == "HIGH":
                hi = float("inf")
            else:
                hi = _parse_field(fd["max"])
                if hi is None:
                    errors.append(f"{risk}: Max % must be a number.")
                    continue
                if hi < lo:
                    errors.append(f"{risk}: Max % must be ≥ Min %.")
                    continue
            parsed[risk] = (lo, hi)

        if not errors and len(parsed) == 3:
            # Skip overlap check if any field is currently empty (user is mid-edit).
            all_filled = all(
                fd["min"].get().strip() and (fd["max"] is None or fd["max"].get().strip())
                for fd in fields.values()
            )
            if not all_filled:
                return None, []
            # Check for overlaps using boundary-point comparison only (no looping).
            # For each pair of ranges, check if their intervals overlap.
            overlap_msgs = []
            range_items = list(parsed.items())  # [("LOW", (lo, hi)), ...]
            for i in range(len(range_items)):
                for j in range(i + 1, len(range_items)):
                    r1, (lo1, hi1) = range_items[i]
                    r2, (lo2, hi2) = range_items[j]
                    hi1_cmp = hi1 if hi1 != float("inf") else float("inf")
                    hi2_cmp = hi2 if hi2 != float("inf") else float("inf")
                    # Two ranges [lo1,hi1] and [lo2,hi2] overlap if lo1 <= hi2 and lo2 <= hi1
                    if lo1 <= hi2_cmp and lo2 <= hi1_cmp:
                        msg = f"Overlap: {r1} ({lo1}–{'∞' if hi1==float('inf') else hi1}%) and {r2} ({lo2}–{'∞' if hi2==float('inf') else hi2}%) overlap."
                        if msg not in overlap_msgs:
                            overlap_msgs.append(msg)
            if overlap_msgs:
                errors.extend(overlap_msgs)

        if errors:
            warn_var.set("⚠  " + "  |  ".join(errors))
            warn_lbl.pack(fill="x", padx=16, pady=(2, 4), before=rows_frame)
            return None, errors
        else:
            warn_var.set("")
            try:
                warn_lbl.pack_forget()
            except Exception:
                pass
            return parsed, []

    # Debounced live validation — waits 600 ms after last keystroke before running.
    # This prevents flooding the warning label while the user is mid-edit (e.g. backspacing).
    _debounce_job = [None]
    def _validate_debounced(*_args):
        if _debounce_job[0] is not None:
            try:
                dialog.after_cancel(_debounce_job[0])
            except Exception:
                pass
        _debounce_job[0] = dialog.after(600, _validate)
    for fd in fields.values():
        fd["min"].trace_add("write", _validate_debounced)
        if fd["max"] is not None:
            fd["max"].trace_add("write", _validate_debounced)

    # ── Reset to defaults ─────────────────────────────────────────────
    def _reset_defaults():
        for risk, fd in fields.items():
            lo, hi = _SIM_DEFAULT_RISK_RANGES[risk]
            fd["min"].set(str(int(lo) if lo == int(lo) else lo))
            if fd["max"] is not None:
                fd["max"].set(str(int(hi) if hi == int(hi) else hi))
        warn_var.set("")
        try:
            warn_lbl.pack_forget()
        except Exception:
            pass

    # ── Apply ─────────────────────────────────────────────────────────
    def _apply_and_close():
        parsed, errors = _validate()
        if errors:
            return   # keep dialog open, warning already shown
        self._sim_risk_ranges = parsed
        _sim_populate(self)
        dialog.destroy()

    # ── Footer ────────────────────────────────────────────────────────
    footer = tk.Frame(dialog, bg=_OFF_WHITE, highlightbackground=_BORDER_MID, highlightthickness=1)
    footer.pack(fill="x", padx=16, pady=(10, 14), side="bottom")

    tk.Button(
        footer,
        text="Reset to Defaults",
        font=F(8, "bold"),
        fg=_TXT_SOFT,
        bg=_WHITE,
        relief="flat",
        bd=0,
        padx=10,
        pady=6,
        cursor="hand2",
        command=_reset_defaults,
    ).pack(side="left", padx=(12, 4), pady=8)

    tk.Button(
        footer,
        text="Cancel",
        font=F(9),
        fg=_TXT_SOFT,
        bg=_OFF_WHITE,
        relief="flat",
        bd=0,
        padx=10,
        pady=8,
        cursor="hand2",
        command=dialog.destroy,
    ).pack(side="right", padx=(0, 4), pady=8)

    tk.Button(
        footer,
        text="  ✔  Apply Ranges  ",
        font=F(9, "bold"),
        fg=_WHITE,
        bg=_NAVY_MID,
        activebackground=_NAVY_LIGHT,
        activeforeground=_WHITE,
        relief="flat",
        bd=0,
        padx=14,
        pady=8,
        cursor="hand2",
        command=_apply_and_close,
    ).pack(side="right", padx=12, pady=8)

    # Run initial validation silently — only show warning if the pre-filled
    # values are genuinely invalid (not just because the user hasn't typed yet).
    _validate_debounced()


def _build_sim_summary_cards(self, parent):
    for title, attr, color in [
        ("Total Net Income",       "_sim_lbl_income",  _ACCENT_SUCCESS),
        ("Base Total Expenses",    "_sim_lbl_base",    _TXT_NAVY),
        ("Simulated Total",        "_sim_lbl_sim",     _TXT_NAVY),
        ("Total Increase (₱)",     "_sim_lbl_inc",     _ACCENT_RED),
        ("Surplus / Deficit",      "_sim_lbl_surplus", _ACCENT_SUCCESS),
    ]:
        card = tk.Frame(parent, bg=_NAVY_MIST,
                        highlightbackground="#D6E4F7", highlightthickness=1)
        card.pack(side="left", fill="x", expand=True, padx=6, pady=8)
        tk.Label(card, text=title, font=F(7), fg=_TXT_SOFT, bg=_NAVY_MIST
                 ).pack(anchor="w", padx=10, pady=(6, 0))
        lbl = tk.Label(card, text="—", font=F(13, "bold"), fg=color, bg=_NAVY_MIST)
        lbl.pack(anchor="w", padx=10, pady=(0, 6))
        setattr(self, attr, lbl)


# ══════════════════════════════════════════════════════════════════════
#  PLACEHOLDER
# ══════════════════════════════════════════════════════════════════════

def _sim_show_placeholder(self):
    frame = getattr(self, "_sim_scroll_frame", None)
    if frame is not None:
        for w in frame.winfo_children():
            w.destroy()
        tk.Label(
            frame,
            text="Run an analysis first to unlock the simulator.",
            font=F(10),
            fg=_TXT_MUTED,
            bg=_CARD_WHITE,
        ).pack(pady=60)
    _sim_draw_chart(self)
    _sim_refresh_client_table(self)

    # Always return to table view and clear any client detail content.
    _sim_show_table_view(self)
    detail = getattr(self, "_sim_client_detail_frame", None)
    if detail is not None:
        for w in detail.winfo_children():
            try:
                w.destroy()
            except Exception:
                pass


# ══════════════════════════════════════════════════════════════════════
#  EXPENSE POPUP + POPULATE
# ══════════════════════════════════════════════════════════════════════

def _sim_render_expense_table_rows(self):
    """Render expense rows into the popup table (if open)."""
    frame = getattr(self, "_sim_scroll_frame", None)
    if frame is None:
        return
    try:
        if not frame.winfo_exists():
            return
    except Exception:
        return

    for w in list(frame.winfo_children()):
        try:
            w.destroy()
        except Exception:
            pass

    all_expenses = list(getattr(self, "_sim_expenses", []) or [])
    if not all_expenses:
        tk.Label(
            frame,
            text="No numeric expense data found.",
            font=F(9),
            fg=_TXT_MUTED,
            bg=_CARD_WHITE,
            justify="center",
        ).pack(pady=60)
        return

    # Apply search filter if present
    _exp_q = str(getattr(self, "_sim_expense_search_var", None) and
                 self._sim_expense_search_var.get() or "").strip().lower()
    if _exp_q:
        all_expenses = [e for e in all_expenses if _exp_q in str(e.get("name") or "").lower()]

    if getattr(self, "_sim_expenses_capped", False):
        tk.Label(
            frame,
            text=f"ℹ  Showing top {SIM_MAX_ROWS} expense rows (file has more).",
            font=F(8),
            fg=_ACCENT_GOLD,
            bg=_OFF_WHITE,
            padx=10,
            pady=4,
        ).pack(fill="x")

    hdr = tk.Frame(frame, bg=_OFF_WHITE)
    hdr.pack(fill="x", pady=(8, 0))
    for col, (_title, min_px, _wt) in enumerate(SIM_TABLE_COLUMNS):
        hdr.grid_columnconfigure(col, weight=1, minsize=min_px, uniform="sim_col")
    for col, (text, _min_px, _wt) in enumerate(SIM_TABLE_COLUMNS):
        tk.Label(
            hdr,
            text=text,
            font=F(8, "bold"),
            fg=_NAVY_PALE,
            bg=_OFF_WHITE,
            anchor="w" if col == 0 else "center",
            justify="left" if col == 0 else "center",
            padx=6,
            pady=5,
        ).grid(row=0, column=col, sticky="ew", padx=(0, 2))
    tk.Frame(frame, bg=_BORDER_MID, height=1).pack(fill="x")

    for idx, exp in enumerate(all_expenses):
        var = self._sim_sliders.get(exp["name"])
        if var is None:
            var = tk.DoubleVar(value=0.0)
            self._sim_sliders[exp["name"]] = var
        _sim_build_expense_row(self, frame, exp, var, idx)

    _sim_refresh(self)


def _sim_open_expense_table_window(self):
    """Open the expense simulator table in a separate window."""
    existing = getattr(self, "_sim_expense_win", None)
    if existing is not None:
        try:
            if existing.winfo_exists():
                existing.deiconify()
                existing.lift()
                existing.focus_force()
                return
        except Exception:
            pass

    win = tk.Toplevel(self)
    win.title("Expense Simulator Table")
    win.configure(bg=_CARD_WHITE)
    win.resizable(True, True)
    win.grab_set()
    self._sim_expense_win = win

    p_x = self.winfo_rootx()
    p_y = self.winfo_rooty()
    p_w = self.winfo_width()
    p_h = self.winfo_height()
    w_w, w_h = 980, 620
    win.geometry(f"{w_w}x{w_h}+{p_x + (p_w - w_w)//2}+{p_y + (p_h - w_h)//2}")
    win.minsize(760, 500)

    hdr = tk.Frame(win, bg=_NAVY_DEEP)
    hdr.pack(fill="x")
    tk.Label(
        hdr,
        text="⧉  Expense Simulator Table",
        font=F(11, "bold"),
        fg=_WHITE,
        bg=_NAVY_DEEP,
        padx=16,
        pady=10,
    ).pack(side="left")
    tk.Label(
        hdr,
        text="Adjust per-expense inflation here. Client Impact in the main tab updates automatically.",
        font=F(8),
        fg="#8DA8C8",
        bg=_NAVY_DEEP,
        padx=8,
        pady=10,
    ).pack(side="left")

    # ── Search bar ─────────────────────────────────────────────────────
    search_bar = tk.Frame(win, bg=_OFF_WHITE, height=46)
    search_bar.pack(fill="x")
    search_bar.pack_propagate(False)
    tk.Label(search_bar, text="🔍", font=F(10), fg=_TXT_MUTED,
             bg=_OFF_WHITE).pack(side="left", padx=(14, 4), pady=10)
    self._sim_expense_search_var = tk.StringVar()
    _exp_search_entry = ctk.CTkEntry(
        search_bar,
        textvariable=self._sim_expense_search_var,
        width=300,
        height=28,
        corner_radius=5,
        fg_color=_WHITE,
        text_color=_TXT_NAVY,
        border_color=_BORDER_MID,
        font=FF(10),
        placeholder_text="Search expense item…",
    )
    _exp_search_entry.pack(side="left", pady=9)
    _exp_search_entry.bind(
        "<KeyRelease>",
        lambda _e: _sim_render_expense_table_rows(self),
    )
    ctk.CTkButton(
        search_bar,
        text="✕",
        width=28,
        height=28,
        corner_radius=5,
        fg_color=_OFF_WHITE,
        hover_color=_BORDER_MID,
        text_color=_TXT_MUTED,
        font=FF(9, "bold"),
        border_width=1,
        border_color=_BORDER_MID,
        command=lambda: [
            self._sim_expense_search_var.set(""),
            _sim_render_expense_table_rows(self),
        ],
    ).pack(side="left", padx=(6, 0), pady=9)

    body = tk.Frame(win, bg=_CARD_WHITE)
    body.pack(fill="both", expand=True, padx=12, pady=12)
    sim_sb = tk.Scrollbar(body, relief="flat",
                          troughcolor=_OFF_WHITE, bg=_BORDER_LIGHT, width=8, bd=0)
    sim_sb.pack(side="right", fill="y")
    canvas = tk.Canvas(body, bg=_CARD_WHITE, highlightthickness=0,
                       yscrollcommand=sim_sb.set)
    canvas.pack(side="left", fill="both", expand=True)
    sim_sb.config(command=canvas.yview)
    frame = tk.Frame(canvas, bg=_CARD_WHITE)
    win_id = canvas.create_window((0, 0), window=frame, anchor="nw")
    frame.bind("<Configure>", lambda _e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.bind("<Configure>", lambda e: canvas.itemconfig(win_id, width=e.width))
    canvas.bind(
        "<Enter>",
        lambda _e: canvas.bind_all(
            "<MouseWheel>",
            lambda ev: canvas.yview_scroll(int(-1 * (ev.delta / 120)), "units"),
        ),
    )
    canvas.bind("<Leave>", lambda _e: canvas.unbind_all("<MouseWheel>"))

    self._sim_canvas = canvas
    self._sim_scroll_frame = frame
    _sim_render_expense_table_rows(self)

    def _on_close():
        try:
            canvas.unbind_all("<MouseWheel>")
        except Exception:
            pass
        self._sim_canvas = None
        self._sim_scroll_frame = None
        self._sim_expense_win = None
        try:
            self._sim_expense_search_var.set("")
        except Exception:
            pass
        win.destroy()

    win.protocol("WM_DELETE_WINDOW", _on_close)


def _sim_populate(self):
    if not hasattr(self, "_sim_hdr_lbl") or not self._sim_hdr_lbl.winfo_exists():
        return

    # Cancel any in-progress build
    if getattr(self, "_sim_build_job", None):
        try:
            self._sim_hdr_lbl.after_cancel(self._sim_build_job)
        except Exception:
            pass
        self._sim_build_job = None

    filtered_data  = _lu_get_filtered_all_data(self)
    selected_inds  = getattr(self, "_sim_selected_industries", set()) or set()
    filtered_data  = _sim_filter_data_by_industry_checklist(filtered_data, selected_inds)
    q              = getattr(self, "_sim_search_var", tk.StringVar(value="")).get().strip()
    filtered_data  = _lu_filter_data_by_query(filtered_data, q)
    match_count    = len(filtered_data.get("general", []))
    match_lbl      = getattr(self, "_sim_match_lbl", None)
    filter_lbl     = getattr(self, "_sim_industry_filter_lbl", None)
    if filter_lbl is not None:
        if selected_inds:
            selected_sorted = sorted(selected_inds, key=str.lower)
            preview = ", ".join(selected_sorted[:2])
            extra_count = len(selected_sorted) - 2
            suffix = f" (+{extra_count} more)" if extra_count > 0 else ""
            filter_lbl.config(text=f"Industry filter: {preview}{suffix}")
        else:
            filter_lbl.config(text="")
    if match_lbl is not None:
        if q:
            client_names = sorted({
                (r.get("client") or "").strip()
                for r in filtered_data.get("general", [])
                if r.get("client")
            })
            if len(client_names) == 1:
                match_lbl.config(text=client_names[0][:28], bg="#4A6FA5")
            else:
                match_lbl.config(text=f"{match_count} CLIENTS MATCHED", bg="#4A6FA5")
        else:
            match_lbl.config(text="", bg=_OFF_WHITE)
    active_sectors = _lu_get_active_sectors(self)
    client         = self._lu_active_client
    is_general     = (client == GENERAL_CLIENT)
    all_clients    = filtered_data.get("clients", {})

    # If the search narrows to a single client (or exact client key match),
    # treat it as a per-client simulator view so expenses always show.
    chosen_client = None
    if q:
        ql = q.strip().lower()
        exact = next((name for name in all_clients.keys() if str(name).strip().lower() == ql), None)
        if exact:
            chosen_client = exact
        else:
            client_names = sorted({
                (r.get("client") or "").strip()
                for r in filtered_data.get("general", [])
                if r.get("client")
            })
            if len(client_names) == 1:
                chosen_client = client_names[0]

    if chosen_client and chosen_client in all_clients:
        try:
            self._lu_active_client = chosen_client
        except Exception:
            pass
        recs = [all_clients[chosen_client]]
    else:
        recs = (list(all_clients.values())
                if (is_general or active_sectors)
                else ([all_clients[client]] if client in all_clients else []))

    # ── Apply client impact search bar filter to recs ──────────────────
    # This ensures the Expense Settings popup only shows expenses that
    # belong to the clients currently visible in the Client Impact table.
    _client_search_term = str(
        getattr(self, "_sim_client_search_var", None) and
        self._sim_client_search_var.get() or ""
    ).strip().lower()
    if _client_search_term:
        recs = [
            r for r in recs
            if _client_search_term in str((r or {}).get("client") or "").lower()
        ]

    # Update header
    if q:
        self._sim_hdr_lbl.config(
            text=f"⚙️  Simulator — Search: {q[:30]}",
            fg=_LIME_MID)
    elif active_sectors:
        self._sim_hdr_lbl.config(
            text=f"⚙️  Simulator — Filtered: {' · '.join(active_sectors)}",
            fg=_LIME_MID)
    else:
        self._sim_hdr_lbl.config(text="⚙️  Inflation / Cost-Shock Simulator", fg=_WHITE)

    # Use per-client net income (not total source) for simulator totals.
    net_income           = sum((r.get("net_income") or 0) for r in recs)
    self._sim_net_income = net_income
    self._sim_recs = list(recs)
    self._sim_client_page = 0

    accumulated: dict = {}
    for rec in recs:
        for exp in rec.get("expenses", []):
            name = str((exp or {}).get("name") or "").strip()
            if not name:
                continue
            try:
                total = float((exp or {}).get("total") or 0.0)
            except (TypeError, ValueError):
                total = 0.0
            if total <= 0:
                continue
            risk = str((exp or {}).get("risk") or "LOW").upper()
            if risk not in _RISK_ORDER:
                risk = "LOW"
            reason = str((exp or {}).get("reason") or "").strip()
            if name not in accumulated:
                accumulated[name] = {
                    "name": name,
                    "total": total,
                    "risk": risk,
                    "reason": reason,
                    "value_str": _fmt_value([total]),
                }
            else:
                accumulated[name]["total"] += total
                if _RISK_ORDER.get(risk, 9) < _RISK_ORDER.get(accumulated[name]["risk"], 9):
                    accumulated[name]["risk"] = risk
                    accumulated[name]["reason"] = reason
                accumulated[name]["value_str"] = _fmt_value([accumulated[name]["total"]])

    all_expenses = sorted(accumulated.values(),
                          key=lambda e: _RISK_ORDER.get(e["risk"], 9))

    # Apply row cap
    capped = False
    if len(all_expenses) > SIM_MAX_ROWS:
        all_expenses = all_expenses[:SIM_MAX_ROWS]
        capped       = True

    self._sim_expenses = all_expenses
    self._sim_expenses_capped = capped

    # Preserve previously-typed inflation rates so filter changes don't wipe them.
    _old_sliders = getattr(self, "_sim_sliders", {}) or {}
    self._sim_sliders = {}

    if not all_expenses:
        _sim_render_expense_table_rows(self)
        _sim_refresh(self)
        return

    # Pre-create DoubleVar objects, carrying forward any rate the user already typed.
    for exp in all_expenses:
        old_var = _old_sliders.get(exp["name"])
        try:
            old_val = float(old_var.get()) if old_var is not None else 0.0
        except Exception:
            old_val = 0.0
        var = tk.DoubleVar(value=old_val)
        self._sim_sliders[exp["name"]] = var

    _sim_render_expense_table_rows(self)
    _sim_refresh(self)


# ══════════════════════════════════════════════════════════════════════
#  ROW BUILDER
# ══════════════════════════════════════════════════════════════════════

def _sim_build_expense_row(self, parent, exp, var, idx):
    risk   = str(exp.get("risk") or "LOW").upper()
    if risk not in _RISK_ORDER:
        risk = "LOW"
    name = str(exp.get("name") or "Unnamed Expense")
    try:
        base_total = float(exp.get("total") or 0.0)
    except (TypeError, ValueError):
        base_total = 0.0
    row_bg = _RISK_BG.get(risk, _WHITE) if idx % 2 == 0 else _WHITE
    row    = tk.Frame(parent, bg=row_bg)
    row.pack(fill="x")
    for ci, (_title, min_px, _wt) in enumerate(SIM_TABLE_COLUMNS):
        row.grid_columnconfigure(ci, weight=1, minsize=min_px, uniform="sim_col")

    tk.Label(row, text=name, font=F(9, "bold"),
             fg=_TXT_NAVY, bg=row_bg, anchor="w", padx=8, pady=6
             ).grid(row=0, column=0, sticky="ew")
    tk.Label(row, text=risk, font=F(7, "bold"),
             fg=_RISK_COLOR.get(risk, _TXT_SOFT),
             bg=_RISK_BADGE_BG.get(risk, _OFF_WHITE),
             anchor="center", justify="center",
             padx=10, pady=3).grid(row=0, column=1, padx=4, pady=6, sticky="")
    tk.Label(row, text=f"₱{base_total:,.2f}" if base_total > 0 else "—",
             font=F(9), fg=_TXT_NAVY, bg=row_bg, anchor="center", justify="center", padx=6
             ).grid(row=0, column=2, sticky="ew")

    rate_entry = ctk.CTkEntry(row, textvariable=var, width=80, height=26, corner_radius=4,
                              font=FF(9), fg_color=_WHITE, text_color=_TXT_NAVY,
                              border_color=_RISK_COLOR.get(risk, _BORDER_MID),
                              placeholder_text="0")
    rate_entry.grid(row=0, column=3, padx=8, pady=6)
    rate_entry.bind("<Return>",   lambda e, ex=exp: _sim_on_slide(self, ex, var.get()))
    rate_entry.bind("<FocusOut>", lambda e, ex=exp: _sim_on_slide(self, ex, var.get()))

    extra_lbl = tk.Label(row, text="—", font=F(9), fg=_ACCENT_RED,
                         bg=row_bg, anchor="center", justify="center", padx=6)
    extra_lbl.grid(row=0, column=4, sticky="ew")
    sim_lbl = tk.Label(row, text="—", font=F(9, "bold"), fg=_TXT_NAVY,
                       bg=row_bg, anchor="center", justify="center", padx=6)
    sim_lbl.grid(row=0, column=5, sticky="ew")

    var._extra_lbl = extra_lbl
    var._sim_lbl   = sim_lbl
    var._base      = base_total
    tk.Frame(parent, bg=_BORDER_LIGHT, height=1).pack(fill="x")


# ══════════════════════════════════════════════════════════════════════
#  INTERACTION CALLBACKS
# ══════════════════════════════════════════════════════════════════════

def _sim_on_slide(self, exp, value):
    try:
        pct = float(value)
    except (ValueError, TypeError):
        pct = 0.0
    if pct < 0.0:
        pct = 0.0
    self._sim_sliders[exp["name"]].set(str(pct))
    _sim_refresh(self)


def _sim_apply_global(self):
    try:
        pct = float(self._sim_global_var.get())
    except (ValueError, TypeError):
        pct = 0.0
    if pct < 0.0:
        pct = 0.0
    if not self._sim_sliders and self._lu_all_data:
        _sim_populate(self)
    for var in self._sim_sliders.values():
        var.set(str(pct))
    _sim_refresh(self)


def _sim_reset(self):
    self._sim_global_var.set("0")
    if not self._sim_sliders and self._lu_all_data:
        _sim_populate(self)
    for var in self._sim_sliders.values():
        var.set("0")
    _sim_refresh(self)


def _sim_refresh(self):
    base_total = sim_total = 0.0
    for exp in getattr(self, "_sim_expenses", []):
        pct = 0.0
        var = self._sim_sliders.get(exp["name"])
        if var:
            try:
                pct = float(var.get())
            except (ValueError, TypeError):
                pass
        base  = exp["total"]
        extra = base * pct / 100.0
        sim   = base + extra
        base_total += base
        sim_total  += sim
        if var and hasattr(var, "_extra_lbl"):
            try:
                var._extra_lbl.config(
                    text=f"+₱{extra:,.2f}" if extra > 0 else "—",
                    fg=_ACCENT_RED if extra > 0 else _TXT_MUTED)
                var._sim_lbl.config(
                    text=f"₱{sim:,.2f}" if base > 0 else "—", fg=_TXT_NAVY)
            except Exception:
                pass

    increase   = sim_total - base_total
    net_income = getattr(self, "_sim_net_income", 0.0) or 0.0
    surplus    = net_income - sim_total

    if hasattr(self, "_sim_lbl_base"):
        try:
            self._sim_lbl_income.config(
                text=f"₱{net_income:,.2f}" if net_income else "—",
                fg=_ACCENT_SUCCESS)
            self._sim_lbl_base.config(
                text=f"₱{base_total:,.2f}" if base_total else "—")
            self._sim_lbl_sim.config(
                text=f"₱{sim_total:,.2f}" if base_total else "—")
            self._sim_lbl_inc.config(
                text=f"+₱{increase:,.2f}" if increase > 0 else "₱0.00",
                fg=_ACCENT_RED if increase > 0 else _TXT_NAVY)
            if net_income:
                surplus_txt = f"{'▲' if surplus >= 0 else '▼'} ₱{abs(surplus):,.2f}"
                self._sim_lbl_surplus.config(
                    text=surplus_txt,
                    fg=_ACCENT_SUCCESS if surplus >= 0 else _ACCENT_RED)
            else:
                self._sim_lbl_surplus.config(text="—", fg=_TXT_MUTED)
        except Exception:
            pass

    if hasattr(self, "_sim_income_lbl"):
        try:
            if net_income:
                self._sim_income_lbl.config(
                    text=f"TOTAL NET INCOME  ₱{net_income:,.2f}",
                    fg=_LIME_MID)
                self._sim_surplus_lbl.config(
                    text=(f"SURPLUS  ₱{surplus:,.2f}" if surplus >= 0
                          else f"DEFICIT  ▲ ₱{abs(surplus):,.2f}"),
                    fg=_LIME_MID if surplus >= 0 else _ACCENT_RED)
            else:
                self._sim_income_lbl.config(
                    text="TOTAL NET INCOME  —  Load a file to begin",
                    fg=_TXT_MUTED)
                self._sim_surplus_lbl.config(text="", fg=_LIME_MID)
        except Exception:
            pass

    _sim_draw_chart(self)
    _sim_refresh_client_table(self)


def _sim_pct_net_to_amort_label(pct: float, ranges: dict | None = None) -> str:
    """
    Risk Label rules for the simulator client table.
    Ranges default to _SIM_DEFAULT_RISK_RANGES but can be overridden at runtime
    via self._sim_risk_ranges (passed in as the `ranges` argument).

    Each range entry: {"LOW": (min, max), "MEDIUM": (min, max), "HIGH": (min, max)}
    Boundaries are inclusive on both ends. HIGH max may be float('inf').
    Falls back to LOW if pct matches no range.
    """
    if ranges is None:
        ranges = _SIM_DEFAULT_RISK_RANGES
    try:
        p = float(pct)
    except Exception:
        return "LOW"
    # Check HIGH first, then MEDIUM, then LOW so the most severe wins on overlap edge.
    for label in ("HIGH", "MEDIUM", "LOW"):
        lo, hi = ranges.get(label, (0.0, -1.0))
        if lo <= p <= hi:
            return label
    return "LOW"


def _sim_build_risk_reasoning(risk_label: str, pct_net_to_am: float) -> str:
    """Build user-facing simulator risk explanation text."""
    label = str(risk_label or "LOW").upper()
    base = (
        f"The client is {label} risk because they have "
        f"{float(pct_net_to_am):.1f}% of Net Income to Amortization."
    )
    if label == "HIGH":
        return (
            base
            + " Please review carefully — this client might have a special loan case."
        )
    return base


def _sim_amount_for_expense(self, base_amount: float, expense_name: str) -> tuple[float, float]:
    """Return (extra_cost, simulated_amount) for an expense name using current slider %."""
    pct = 0.0
    var = getattr(self, "_sim_sliders", {}).get(expense_name)
    if var is not None:
        try:
            pct = float(var.get() or 0.0)
        except Exception:
            pct = 0.0
    if pct < 0.0:
        pct = 0.0
    base = float(base_amount or 0.0)
    extra = base * pct / 100.0
    return (extra, base + extra)


def _sim_refresh_client_table(self):
    tree = getattr(self, "_sim_client_tree", None)
    if tree is None:
        return
    try:
        if not tree.winfo_exists():
            return
    except Exception:
        return

    recs_all = list(getattr(self, "_sim_recs", []) or [])
    page = int(getattr(self, "_sim_client_page", 0) or 0)

    # Used by click handler to reliably map a clicked Treeview row
    # back to the full client name (even if the visible label is truncated).
    self._sim_iid_to_client = {}
    self._sim_client_metrics_by_name = {}

    # Apply client name search filter
    _search_term = str(getattr(self, "_sim_client_search_var", None) and
                       self._sim_client_search_var.get() or "").strip().lower()
    if _search_term:
        recs_all = [
            r for r in recs_all
            if _search_term in str((r or {}).get("client") or "").lower()
        ]
    for iid in tree.get_children(""):
        tree.delete(iid)

    # Reset hover tracking on every repopulate so stale iids don't persist.
    try:
        tree._sim_hovered_iid_ref[0] = None
    except AttributeError:
        pass
    if not recs_all:
        tree.insert("", "end", values=("—",) + ("—",) * (len(SIM_CLIENT_TABLE_COLUMNS) - 1), tags=("NA",))
        _sim_update_client_pagination_ui(self, total_rows=0)
        return

    def _money(v: float | None) -> str:
        try:
            if v is None:
                return "—"
            return f"₱{float(v):,.2f}"
        except Exception:
            return "—"

    def _pct(v: float | None) -> str:
        try:
            if v is None:
                return "—"
            return f"{float(v):.1f}%"
        except Exception:
            return "—"

    rows = []
    for rec in recs_all:
        name = str((rec or {}).get("client") or "").strip() or "—"
        industry = str((rec or {}).get("industry") or "").strip() or "—"
        base_net = float((rec or {}).get("net_income") or 0.0)
        current_am = float((rec or {}).get("current_amort") or 0.0)

        # Build per-client base expense map from parsed expenses (same names used in simulator sliders).
        base_by_name: dict[str, float] = {}
        base_total_exp = 0.0
        for exp in (rec or {}).get("expenses", []) or []:
            nm = str((exp or {}).get("name") or "").strip()
            if not nm:
                continue
            try:
                amt = float((exp or {}).get("total") or 0.0)
            except Exception:
                amt = 0.0
            if amt <= 0:
                continue
            base_by_name[nm] = base_by_name.get(nm, 0.0) + amt
            base_total_exp += amt

        # Apply current slider ramp-ups to THIS client's expenses.
        extra_total = 0.0
        for nm, base_amt in base_by_name.items():
            extra, _sim = _sim_amount_for_expense(self, base_amt, nm)
            extra_total += extra

        sim_total_exp = base_total_exp + extra_total

        pct_inc = (extra_total / base_total_exp * 100.0) if base_total_exp > 0 else (0.0 if extra_total <= 0 else 100.0)

        # Simulated net income = Base net income minus the simulated cost increase
        sim_net_income = base_net - extra_total

        # % Net → Amort uses simulated net income (Total Current Amort / Total Net Income Simulated)
        if sim_net_income <= 0:
            pct_net_to_am = 999.0 if current_am > 0 else 0.0
        else:
            pct_net_to_am = (current_am / sim_net_income) * 100.0 if current_am > 0 else 0.0

        sim_risk = _sim_pct_net_to_amort_label(pct_net_to_am, getattr(self, "_sim_risk_ranges", None))
        risk_reasoning = _sim_build_risk_reasoning(sim_risk, pct_net_to_am)
        r = {
            "client": name,
            "industry": industry,
            "base_total_expenses": base_total_exp,
            "sim_total_expenses": sim_total_exp,
            "net_income": base_net,
            "sim_net_income": sim_net_income,
            "pct_increase": pct_inc,
            "sim_increase": extra_total,
            "current_amort": current_am,
            "pct_net_to_amort": pct_net_to_am,
            "sim_risk_label": sim_risk,
            "risk_reasoning": risk_reasoning,
        }
        rows.append(r)
        self._sim_client_metrics_by_name[name] = r

    # Highest risk first, then largest amort ratio.
    risk_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
    rows.sort(key=lambda r: (risk_order.get(r["sim_risk_label"], 9), -(r["pct_net_to_amort"] or 0.0)))

    total = len(rows)
    max_page = max(0, (total - 1) // SIM_CLIENT_PAGE_SIZE) if total else 0
    if page < 0:
        page = 0
    if page > max_page:
        page = max_page
    self._sim_client_page = page
    start = page * SIM_CLIENT_PAGE_SIZE
    end = min(start + SIM_CLIENT_PAGE_SIZE, total)
    page_rows = rows[start:end]
    _sim_update_client_pagination_ui(self, total_rows=total)

    for idx_in_page, r in enumerate(page_rows):
        # Keep full client name for matching with Analysis tab.
        full_client_name = r.get("client") or "—"
        iid = f"sim_client_{start + idx_in_page}"
        vals = (
            r["client"][:60],
            r["industry"],
            _money(r["base_total_expenses"]),
            _money(r["sim_total_expenses"]),
            _money(r["net_income"]),
            _money(r["sim_net_income"]),
            _pct(r["pct_increase"]),
            _money(r["sim_increase"]),
            _money(r["current_amort"]),
            _pct(r["pct_net_to_amort"]),
            r["sim_risk_label"],
            r["risk_reasoning"],
        )
        tag = r["sim_risk_label"] if r["sim_risk_label"] in ("HIGH", "MEDIUM", "LOW") else "NA"
        self._sim_iid_to_client[iid] = str(full_client_name)
        tree.insert("", "end", iid=iid, values=vals, tags=(tag,))


def _sim_show_table_view(self):
    """
    Hide the single-client detail view and restore the full treeview table.
    Called by the Back button and whenever the data is refreshed.
    """
    detail = getattr(self, "_sim_client_detail_view", None)
    table  = getattr(self, "_sim_client_table_view",  None)

    if detail is not None:
        try:
            detail.pack_forget()
        except Exception:
            pass
    if table is not None:
        try:
            table.pack(fill="x", expand=False)
        except Exception:
            pass


def _sim_show_detail_view(self, client_name: str):
    """
    Hide the treeview table, show the detail panel for *client_name*.
    Called when a row is clicked.
    """
    table  = getattr(self, "_sim_client_table_view",  None)
    detail = getattr(self, "_sim_client_detail_view", None)

    if table is not None:
        try:
            table.pack_forget()
        except Exception:
            pass
    if detail is not None:
        try:
            detail.pack(fill="x", expand=False, padx=0, pady=0)
        except Exception:
            pass

    # Render the detail content (Back button is rendered inside by _sim_show_client_details)
    _sim_show_client_details(self, client_name)


def _sim_on_client_impact_row_activated(self, event=None):
    """
    Mirror Analysis treeview behavior:
    when a user clicks a Risk Simulator client row, open that client's
    full details in the Analysis tab.
    """
    tree = getattr(self, "_sim_client_tree", None)
    if tree is None:
        return

    iid = None
    try:
        if event is not None and hasattr(event, "y"):
            iid = tree.identify_row(event.y)
        else:
            sel = tree.selection()
            iid = sel[0] if sel else None
    except Exception:
        iid = None

    client_name = getattr(self, "_sim_iid_to_client", {}).get(iid)
    if not client_name or client_name == "—":
        return

    # Keep the Analysis dropdown synchronized (so clicking Analysis tab shows
    # the correct selected client), but do the "all columns" display inside
    # the Simulator tab as requested.
    try:
        if getattr(self, "_lu_client_var", None) is not None:
            self._lu_client_var.set(client_name)
    except Exception:
        pass

    try:
        if hasattr(self, "_lu_on_client_change"):
            self._lu_on_client_change(client_name)
    except Exception:
        pass

    _sim_show_detail_view(self, client_name)


def _sim_show_client_details(self, client_name: str):
    """
    Render the full single-client panel inside _sim_client_detail_view.
    Called by _sim_show_detail_view after the table is hidden.
    """
    detail = getattr(self, "_sim_client_detail_frame", None)
    if detail is None:
        return

    # Wipe previous content
    for w in detail.winfo_children():
        try:
            w.destroy()
        except Exception:
            pass

    if not client_name or client_name == "—":
        return

    metrics = getattr(self, "_sim_client_metrics_by_name", {}) or {}
    m = metrics.get(client_name) or {}

    # Prefer full record from LU core output
    rec = (getattr(self, "_lu_all_data", None) or {}).get("clients", {}).get(client_name)
    if not rec:
        for rr in getattr(self, "_sim_recs", []) or []:
            if str((rr or {}).get("client") or "").strip() == client_name:
                rec = rr
                break

    sim_label = str(m.get("sim_risk_label") or "LOW").upper()
    badge_bg  = _RISK_BADGE_BG.get(sim_label, _OFF_WHITE)
    badge_fg  = _RISK_COLOR.get(sim_label, _TXT_MUTED)

    # ── Top-accent color per risk (matches the card screenshot) ────────
    _ACCENT_BAR = {
        "HIGH":   _ACCENT_RED,      # red
        "MEDIUM": _ACCENT_GOLD,     # gold/amber
        "LOW":    _ACCENT_SUCCESS,  # green
    }
    accent_color = _ACCENT_BAR.get(sim_label, _LIME_MID)

    # ── Single unified card (one accent bar, hero + metrics together) ──
    client_card = tk.Frame(detail, bg=_NAVY_DEEP)
    client_card.pack(fill="x")

    # Single colored top accent bar only
    tk.Frame(client_card, bg=accent_color, height=4).pack(fill="x")

    hero_inner = tk.Frame(client_card, bg=_NAVY_DEEP)
    hero_inner.pack(fill="x", padx=24, pady=16)

    # Back to Table button — lives in the same container as the client name
    ctk.CTkButton(
        hero_inner,
        text="◄  Back to Table",
        width=130,
        height=28,
        corner_radius=6,
        fg_color=_NAVY_LIGHT,
        hover_color=_NAVY_MID,
        text_color=_WHITE,
        font=FF(9, "bold"),
        command=lambda: _sim_show_table_view(self),
    ).pack(side="right", padx=(0, 4), pady=(0, 8))

    tk.Label(
        hero_inner,
        text=client_name,
        font=F(13, "bold"),
        fg=_WHITE,
        bg=_NAVY_DEEP,
        anchor="w",
        justify="left",
        wraplength=900,
    ).pack(anchor="w")

    badge_row = tk.Frame(hero_inner, bg=_NAVY_DEEP)
    badge_row.pack(anchor="w", pady=(8, 0))
    badge = tk.Frame(badge_row, bg=badge_bg,
                     highlightbackground=badge_fg, highlightthickness=1)
    badge.pack(side="left")
    tk.Label(
        badge,
        text=f"  {sim_label} RISK  ",
        font=F(9, "bold"),
        fg=badge_fg,
        bg=badge_bg,
        padx=10,
        pady=4,
    ).pack()

    # Reasoning beside badge
    tk.Label(
        badge_row,
        text=str(m.get("risk_reasoning") or ""),
        font=F(8),
        fg="#8DAACC",
        bg=_NAVY_DEEP,
        anchor="w",
        justify="left",
        wraplength=800,
        padx=14,
    ).pack(side="left", fill="x", expand=True)

    # Subtle inner divider — no second accent bar
    tk.Frame(client_card, bg="#1E3A5F", height=1).pack(fill="x", padx=0)

    summary = tk.Frame(client_card, bg=_NAVY_DEEP)
    summary.pack(fill="x")


    def _money(v):
        try:
            return f"₱{float(v):,.2f}"
        except Exception:
            return "—"

    def _pct(v):
        try:
            return f"{float(v):.1f}%"
        except Exception:
            return "—"

    metrics_pairs = [
        ("Total Expenses (Base)", _money(m.get("base_total_expenses"))),
        ("Total Expenses (Sim)",  _money(m.get("sim_total_expenses"))),
        ("Net Income (Base)",     _money(m.get("net_income"))),
        ("Net Income (Sim)",      _money(m.get("sim_net_income"))),
        ("% Expense Increase",    _pct(m.get("pct_increase"))),
        ("Simulated Increase",    _money(m.get("sim_increase"))),
        ("Current Amort",         _money(m.get("current_amort"))),
        ("% Net → Amort (Sim)",   _pct(m.get("pct_net_to_amort"))),
    ]
    for lbl, val in metrics_pairs:
        c = tk.Frame(summary, bg=_NAVY_DEEP)
        c.pack(side="left", padx=14, pady=12)
        tk.Label(c, text=lbl,  font=F(7),          fg="#8DAACC",  bg=_NAVY_DEEP).pack(anchor="w")
        tk.Label(c, text=val,  font=F(10, "bold"),  fg=_WHITE,     bg=_NAVY_DEEP).pack(anchor="w")

    tk.Frame(detail, bg=_BORDER_MID, height=1).pack(fill="x")

    # ── Full LU record — section title ───────────────────────────────
    rec_hdr = tk.Frame(detail, bg=_OFF_WHITE)
    rec_hdr.pack(fill="x")
    tk.Label(
        rec_hdr,
        text="Full Client Record  —  All Columns",
        font=F(9, "bold"),
        fg=_TXT_SOFT,
        bg=_OFF_WHITE,
        anchor="w",
        padx=20,
        pady=8,
    ).pack(anchor="w")
    tk.Frame(detail, bg=_BORDER_LIGHT, height=1).pack(fill="x")

    if not rec:
        tk.Label(
            detail,
            text="No full LU record found for this client.",
            font=F(9),
            fg=_ACCENT_RED,
            bg=_CARD_WHITE,
            wraplength=1100,
            justify="left",
            padx=20,
            pady=20,
        ).pack(anchor="w")
        return

    row_idx = 0
    for _cid, heading, field, _w, _a, kind in LU_CLIENT_TREE_SPEC:
        if field not in rec:
            continue
        raw_val = rec.get(field)
        if raw_val is None or str(raw_val).strip() == "":
            continue
        value = lu_format_lu_cell(rec, field, kind, text_limit=500)

        bg  = _WHITE if row_idx % 2 == 0 else _OFF_WHITE
        row = tk.Frame(detail, bg=bg)
        row.pack(fill="x", padx=0, pady=0)

        tk.Label(
            row,
            text=heading,
            font=F(8, "bold"),
            fg=_TXT_SOFT,
            bg=bg,
            width=28,
            anchor="nw",
            justify="left",
            padx=20,
            pady=7,
        ).pack(side="left", fill="y")

        tk.Label(
            row,
            text=value,
            font=F(9),
            fg=_TXT_NAVY,
            bg=bg,
            anchor="w",
            justify="left",
            wraplength=1000,
            padx=8,
            pady=7,
        ).pack(side="left", fill="x", expand=True)

        tk.Frame(detail, bg=_BORDER_LIGHT, height=1).pack(fill="x", pady=0)
        row_idx += 1

    # Bottom spacer
    tk.Frame(detail, bg=_CARD_WHITE, height=24).pack(fill="x")


def _sim_export_client_impact_excel(self):
    """
    Export the simulator "Client Impact" rows to Excel.

    - Uses current simulator slider values (what-if).
    - Uses the current LU industry filter already applied to `_sim_recs`.
    - Uses the current client search filter from `_sim_client_search_var`.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        from lu_loanbal_export_patch import _write_configuration_settings_sheet
    except ImportError:
        messagebox.showerror(
            "Missing Library",
            "openpyxl is not installed.\nRun:  pip install openpyxl",
            parent=self,
        )
        return

    recs_all = list(getattr(self, "_sim_recs", []) or [])
    if not recs_all:
        messagebox.showinfo(
            "No simulator data",
            "Run LU analysis and the simulator first, then try Export again.",
            parent=self,
        )
        return

    # Apply client name search filter (same logic as the on-screen table).
    _search_term = str(
        getattr(self, "_sim_client_search_var", None) and self._sim_client_search_var.get()
        or ""
    ).strip().lower()
    if _search_term:
        recs_all = [
            r for r in recs_all
            if _search_term in str((r or {}).get("client") or "").lower()
        ]
    if not recs_all:
        messagebox.showinfo(
            "No matching rows",
            "No simulator rows match the current client search filter.",
            parent=self,
        )
        return

    # Build per-client simulated totals and risk labels (same rules as the table).
    rows = []
    for rec in recs_all:
        name = str((rec or {}).get("client") or "").strip() or "—"
        industry = str((rec or {}).get("industry") or "").strip() or "—"
        base_net = float((rec or {}).get("net_income") or 0.0)
        current_am = float((rec or {}).get("current_amort") or 0.0)

        base_by_name: dict[str, float] = {}
        base_total_exp = 0.0
        for exp in (rec or {}).get("expenses", []) or []:
            nm = str((exp or {}).get("name") or "").strip()
            if not nm:
                continue
            try:
                amt = float((exp or {}).get("total") or 0.0)
            except Exception:
                amt = 0.0
            if amt <= 0:
                continue
            base_by_name[nm] = base_by_name.get(nm, 0.0) + amt
            base_total_exp += amt

        extra_total = 0.0
        for nm, base_amt in base_by_name.items():
            extra, _sim = _sim_amount_for_expense(self, base_amt, nm)
            extra_total += extra

        sim_total_exp = base_total_exp + extra_total
        pct_inc = (
            extra_total / base_total_exp * 100.0
            if base_total_exp > 0
            else (0.0 if extra_total <= 0 else 100.0)
        )

        # Simulated net income = Base net income minus simulated cost increase.
        sim_net_income = base_net - extra_total
        if sim_net_income <= 0:
            pct_net_to_am = 999.0 if current_am > 0 else 0.0
        else:
            pct_net_to_am = (
                (current_am / sim_net_income) * 100.0 if current_am > 0 else 0.0
            )

        sim_risk = _sim_pct_net_to_amort_label(pct_net_to_am, getattr(self, "_sim_risk_ranges", None))
        risk_reasoning = _sim_build_risk_reasoning(sim_risk, pct_net_to_am)
        rows.append({
            "client_id": str((rec or {}).get("client_id") or "").strip(),
            "pn": str((rec or {}).get("pn") or "").strip(),
            "client": name,
            "residence_address": str((rec or {}).get("residence_address") or "").strip(),
            "office_address": str((rec or {}).get("office_address") or "").strip(),
            "industry": industry,
            "loan_status": str((rec or {}).get("loan_status") or "").strip(),
            "ao_name": str((rec or {}).get("ao_name") or "").strip(),
            "product_name": str((rec or {}).get("product_name") or "").strip(),
            "loan_balance": float((rec or {}).get("loan_balance") or 0.0),
            "principal_loan": float((rec or {}).get("principal_loan") or 0.0),
            "base_total_expenses": base_total_exp,
            "sim_total_expenses": sim_total_exp,
            "net_income": base_net,
            "sim_net_income": sim_net_income,
            "pct_increase": pct_inc,
            "sim_increase": extra_total,
            "current_amort": current_am,
            "pct_net_to_amort": pct_net_to_am,
            "sim_risk_label": sim_risk,
            "risk_reasoning": risk_reasoning,
        })

    # Highest risk first, then larger amort ratio.
    risk_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
    rows.sort(key=lambda r: (risk_order.get(r["sim_risk_label"], 9), -(r["pct_net_to_amort"] or 0.0)))

    from tkinter import filedialog
    import getpass
    from datetime import datetime
    from pathlib import Path

    default_name = f"RiskSimulator_ClientImpact_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    path = filedialog.asksaveasfilename(
        parent=self,
        title="Save Simulator Client Impact Excel",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialfile=default_name,
    )
    if not path:
        return

    wb = openpyxl.Workbook()
    ws_cfg = wb.active
    ws_cfg.title = "Export settings"

    exported_by = str(getattr(self, "_current_username", "") or "").strip()
    if not exported_by:
        try:
            exported_by = (getpass.getuser() or "").strip()
        except Exception:
            exported_by = ""
    if not exported_by:
        exported_by = "Unknown user"

    selected_inds = sorted(getattr(self, "_sim_selected_industries", set()) or set(), key=str.lower)
    industry_note = "None (no industry filter)" if not selected_inds else " · ".join(selected_inds)
    search_note = _search_term if _search_term else "None"

    _write_configuration_settings_sheet(
        ws_cfg,
        fname=Path(str(getattr(self, "_lu_filepath", "") or "—")).name,
        generated_at=datetime.now().strftime("%B %d, %Y  %H:%M"),
        exported_by=exported_by,
        export_scope_note=(
            "Risk Simulator export — includes current simulator what-if values from the Client Impact table. "
            f"Industry filter: {industry_note}. Client search: {search_note}."
        ),
    )

    # Client Impact sheet.
    ws = wb.create_sheet("Client Impact")
    headers = [
        "Client ID",
        "PN",
        "Client",
        "Residence Address",
        "Office Address",
        "Industry",
        "Loan Status",
        "AO Name",
        "Product Name",
        "Loan Balance",
        "Principal Loan",
        "Total Expenses (Base)",
        "Total Expenses (Sim)",
        "Total Net Income (Base)",
        "Total Net Income (Sim)",
        "% Increase",
        "Simulated Increase",
        "Total Current Amort",
        "% Net → Amort",
        "Risk Label",
        "Risk Reasoning",
    ]

    hdr_fill = PatternFill("solid", fgColor=_NAVY_DEEP.lstrip("#"))
    hdr_font = Font(bold=True, size=10, color="FFFFFF")
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Column widths (Excel-ish, not perfect).
    col_widths = [14, 14, 24, 24, 24, 18, 16, 18, 20, 16, 16, 18, 18, 20, 20, 12, 18, 18, 12, 52]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    NUM_FMT = '"₱"#,##0.00'
    PCT_FMT = '0.0"%"'

    for idx, r in enumerate(rows, 0):
        row_num = 2 + idx
        ws.cell(row_num, 1, r["client_id"] or "—")
        ws.cell(row_num, 2, r["pn"] or "—")
        ws.cell(row_num, 3, r["client"])
        ws.cell(row_num, 4, r["residence_address"] or "—")
        ws.cell(row_num, 5, r["office_address"] or "—")
        ws.cell(row_num, 6, r["industry"])
        ws.cell(row_num, 7, r["loan_status"] or "—")
        ws.cell(row_num, 8, r["ao_name"] or "—")
        ws.cell(row_num, 9, r["product_name"] or "—")
        ws.cell(row_num, 10, r["loan_balance"]).number_format = NUM_FMT
        ws.cell(row_num, 11, r["principal_loan"]).number_format = NUM_FMT
        ws.cell(row_num, 12, r["base_total_expenses"]).number_format = NUM_FMT
        ws.cell(row_num, 13, r["sim_total_expenses"]).number_format = NUM_FMT
        ws.cell(row_num, 14, r["net_income"]).number_format = NUM_FMT
        ws.cell(row_num, 15, r["sim_net_income"]).number_format = NUM_FMT
        ws.cell(row_num, 16, r["pct_increase"]).number_format = PCT_FMT
        ws.cell(row_num, 17, r["sim_increase"]).number_format = NUM_FMT
        ws.cell(row_num, 18, r["current_amort"]).number_format = NUM_FMT
        ws.cell(row_num, 19, r["pct_net_to_amort"]).number_format = PCT_FMT
        ws.cell(row_num, 20, r["sim_risk_label"])
        ws.cell(row_num, 21, r["risk_reasoning"])

    ws.freeze_panes = "A2"
    wb.save(path)
    messagebox.showinfo("Export Complete", f"Excel saved to:\n{path}", parent=self)


def _sim_merge_excel_files(self):
    """
    Merge Simulator Client Impact exports with Summary-tab exports.

    Merge key is NAME-BASED (Client/Applicant), as requested.
    For duplicate names, preserve existing non-empty values and only patch
    missing fields from later files.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        messagebox.showerror(
            "Missing Library",
            "openpyxl is not installed.\nRun:  pip install openpyxl",
            parent=self,
        )
        return

    paths = filedialog.askopenfilenames(
        parent=self,
        title="Select Summary Excel/CSV files to merge",
        filetypes=[
            ("Excel & CSV files", "*.xlsx *.csv"),
            ("Excel files", "*.xlsx"),
            ("CSV files", "*.csv"),
            ("All files", "*.*"),
        ],
    )
    if not paths:
        return

    def _norm_text(v) -> str:
        return str(v or "").strip()

    def _row_name(row: dict) -> str:
        # Support both Summary and Simulator naming conventions.
        name = _norm_text(row.get("Applicant"))
        if not name:
            name = _norm_text(row.get("Client"))
        if not name:
            name = _norm_text(row.get("Client Name"))
        return name

    def _norm_name(name: str) -> str:
        s = _norm_text(name).upper()
        s = re.sub(r"[^A-Z0-9\s,.\-]", " ", s)
        s = re.sub(r"\s+", " ", s).strip()
        # Handle "LAST, FIRST" form to improve joins.
        if "," in s:
            parts = [p.strip() for p in s.split(",", 1)]
            if len(parts) == 2 and parts[0] and parts[1]:
                s = f"{parts[1]} {parts[0]}".strip()
        return s

    def _key_for_row(row: dict) -> str:
        name = _norm_name(_row_name(row))
        return f"NAME::{name}" if name else ""

    def _is_empty(v) -> bool:
        if v is None:
            return True
        if isinstance(v, str):
            return v.strip() in ("", "—", "-", "N/A", "n/a")
        return False

    def _read_rows_from_file(path: str) -> tuple[list[str], list[dict]]:
        if str(path).lower().endswith(".csv"):
            import csv as _csv
            with open(path, newline="", encoding="utf-8-sig") as f:
                reader = _csv.DictReader(f)
                headers = [str(h or "").strip() for h in (reader.fieldnames or [])]
                rows = []
                for r in reader:
                    rr = {str(k or "").strip(): ("" if v is None else str(v).strip())
                          for k, v in (r or {}).items()}
                    if any(str(v).strip() for v in rr.values()):
                        rows.append(rr)
                return headers, rows

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                return [], []
            headers = [str(c or "").strip() for c in all_rows[0]]
            rows = []
            for vals in all_rows[1:]:
                if vals is None:
                    continue
                row_dict = {}
                for idx, h in enumerate(headers):
                    if not h:
                        continue
                    vv = vals[idx] if idx < len(vals) else ""
                    row_dict[h] = "" if vv is None else str(vv).strip()
                if any(str(v).strip() for v in row_dict.values()):
                    rows.append(row_dict)
            return headers, rows
        finally:
            wb.close()

    all_headers: list[str] = []
    merged_by_key: dict[str, dict] = {}
    loaded_rows = 0

    for p in paths:
        try:
            headers, rows = _read_rows_from_file(p)
        except Exception as exc:
            messagebox.showerror("Merge Error", f"Failed reading:\n{p}\n\n{exc}", parent=self)
            return

        for h in headers:
            if h and h not in all_headers:
                all_headers.append(h)

        for row in rows:
            loaded_rows += 1
            k = _key_for_row(row)
            if not k:
                k = f"ROW::{loaded_rows}"
            existing = merged_by_key.get(k)
            if existing is None:
                merged_by_key[k] = dict(row)
                continue
            for col, incoming in row.items():
                if col not in existing:
                    existing[col] = incoming
                    continue
                if _is_empty(existing.get(col)) and not _is_empty(incoming):
                    existing[col] = incoming

    if not merged_by_key:
        messagebox.showinfo("Merge Excel", "No mergeable rows found.", parent=self)
        return

    preferred_order = [
        # Name identity first (both summary + simulator naming variants)
        "Applicant", "Client", "Client Name",
        # Summary-tab columns
        "Client ID", "PN", "Residence Address", "Office Address",
        "Industry Name", "Spouse Info", "Personal Assets", "Business Assets",
        "Business Inventory", "Source of Income", "Total Source Of Income",
        "Business Expenses", "Total Business Expenses",
        "Household / Personal Expenses", "Total Household / Personal Expenses",
        "Total Net Income", "Total Amortization History",
        "Total Current Amortization", "Loan Balance", "Principal Loan",
        "Maturity", "Interest Rate", "Branch", "Loan Class", "Product Name",
        "Loan Date", "Term Unit", "Term", "Security", "Release Tag",
        "Loan Amount", "Loan Status", "AO Name",
        # Simulator Client Impact columns
        "Industry",
        "Total Expenses (Base)", "Total Expenses (Sim)",
        "Total Net Income (Base)", "Total Net Income (Sim)",
        "% Increase", "Simulated Increase", "Total Current Amort",
        "% Net → Amort", "Risk Label", "Risk Reasoning",
    ]
    ordered_headers = [h for h in preferred_order if h in all_headers] + [
        h for h in all_headers if h not in preferred_order
    ]

    out_path = filedialog.asksaveasfilename(
        parent=self,
        title="Save Merged Excel",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialfile="Merged_Summary.xlsx",
    )
    if not out_path:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Merged Summary"

    hdr_fill = PatternFill("solid", fgColor="93C47D")
    hdr_font = Font(name="Roboto", bold=True, color="FFFFFF", size=10)
    body_font = Font(name="Roboto", size=9)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for ci, h in enumerate(ordered_headers, 1):
        c = ws.cell(1, ci, h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = left
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 22

    out_rows = list(merged_by_key.values())
    out_rows.sort(key=lambda r: (
        _norm_text(r.get("Client ID")).upper(),
        _norm_text(r.get("Applicant")).upper(),
    ))

    for ri, row in enumerate(out_rows, 2):
        for ci, h in enumerate(ordered_headers, 1):
            cell = ws.cell(ri, ci, row.get(h, ""))
            cell.font = body_font
            cell.alignment = left

    ws.freeze_panes = "A2"
    wb.save(out_path)
    messagebox.showinfo(
        "Merge Complete",
        f"Merged {len(paths)} file(s)\n"
        f"Input rows: {loaded_rows}\n"
        f"Output rows: {len(out_rows)}\n\n"
        f"Saved to:\n{out_path}",
        parent=self,
    )


def _sim_export_high_risk_clients_excel(self):
    """
    Export only HIGH-risk clients from the simulator "Client Impact" table to Excel.

    - HIGH risk is defined as % Net → Amort >= 71%.
    - Uses current simulator slider values (what-if).
    - Uses the current LU industry filter already applied to `_sim_recs`.
    - Uses the current client search filter from `_sim_client_search_var`.
    - Exports all columns identical to the Client Impact table, including Industry.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from lu_loanbal_export_patch import _write_configuration_settings_sheet
    except ImportError:
        messagebox.showerror(
            "Missing Library",
            "openpyxl is not installed.\nRun:  pip install openpyxl",
            parent=self,
        )
        return

    recs_all = list(getattr(self, "_sim_recs", []) or [])
    if not recs_all:
        messagebox.showinfo(
            "No simulator data",
            "Run LU analysis and the simulator first, then try Export again.",
            parent=self,
        )
        return

    # Apply client name search filter (same logic as the on-screen table).
    _search_term = str(
        getattr(self, "_sim_client_search_var", None) and self._sim_client_search_var.get()
        or ""
    ).strip().lower()
    if _search_term:
        recs_all = [
            r for r in recs_all
            if _search_term in str((r or {}).get("client") or "").lower()
        ]

    # Build per-client simulated totals and risk labels (same rules as the table).
    rows = []
    for rec in recs_all:
        name = str((rec or {}).get("client") or "").strip() or "—"
        industry = str((rec or {}).get("industry") or "").strip() or "—"
        base_net = float((rec or {}).get("net_income") or 0.0)
        current_am = float((rec or {}).get("current_amort") or 0.0)

        base_by_name: dict[str, float] = {}
        base_total_exp = 0.0
        for exp in (rec or {}).get("expenses", []) or []:
            nm = str((exp or {}).get("name") or "").strip()
            if not nm:
                continue
            try:
                amt = float((exp or {}).get("total") or 0.0)
            except Exception:
                amt = 0.0
            if amt <= 0:
                continue
            base_by_name[nm] = base_by_name.get(nm, 0.0) + amt
            base_total_exp += amt

        extra_total = 0.0
        for nm, base_amt in base_by_name.items():
            extra, _sim = _sim_amount_for_expense(self, base_amt, nm)
            extra_total += extra

        sim_total_exp = base_total_exp + extra_total
        pct_inc = (
            extra_total / base_total_exp * 100.0
            if base_total_exp > 0
            else (0.0 if extra_total <= 0 else 100.0)
        )

        sim_net_income = base_net - extra_total
        if sim_net_income <= 0:
            pct_net_to_am = 999.0 if current_am > 0 else 0.0
        else:
            pct_net_to_am = (
                (current_am / sim_net_income) * 100.0 if current_am > 0 else 0.0
            )

        sim_risk = _sim_pct_net_to_amort_label(pct_net_to_am, getattr(self, "_sim_risk_ranges", None))
        risk_reasoning = _sim_build_risk_reasoning(sim_risk, pct_net_to_am)
        rows.append({
            "client_id": str((rec or {}).get("client_id") or "").strip(),
            "pn": str((rec or {}).get("pn") or "").strip(),
            "client": name,
            "residence_address": str((rec or {}).get("residence_address") or "").strip(),
            "office_address": str((rec or {}).get("office_address") or "").strip(),
            "industry": industry,
            "loan_status": str((rec or {}).get("loan_status") or "").strip(),
            "ao_name": str((rec or {}).get("ao_name") or "").strip(),
            "product_name": str((rec or {}).get("product_name") or "").strip(),
            "loan_balance": float((rec or {}).get("loan_balance") or 0.0),
            "principal_loan": float((rec or {}).get("principal_loan") or 0.0),
            "base_total_expenses": base_total_exp,
            "sim_total_expenses": sim_total_exp,
            "net_income": base_net,
            "sim_net_income": sim_net_income,
            "pct_increase": pct_inc,
            "sim_increase": extra_total,
            "current_amort": current_am,
            "pct_net_to_amort": pct_net_to_am,
            "sim_risk_label": sim_risk,
            "risk_reasoning": risk_reasoning,
        })

    # Filter to HIGH risk only (>= 71% Net → Amort).
    high_risk_rows = [r for r in rows if r["sim_risk_label"] == "HIGH"]

    if not high_risk_rows:
        messagebox.showinfo(
            "No HIGH Risk Clients",
            (
                "No clients are currently classified as HIGH risk "
                "(% Net → Amort ≥ 71%) under the current simulator settings.\n\n"
                "Try increasing the inflation rates to see high-risk clients appear."
            ),
            parent=self,
        )
        return

    # Sort: largest amort ratio first (all HIGH, so sort by ratio descending).
    high_risk_rows.sort(key=lambda r: -(r["pct_net_to_amort"] or 0.0))

    from tkinter import filedialog
    import getpass
    from datetime import datetime
    from pathlib import Path

    default_name = f"RiskSimulator_HIGH_Risk_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    path = filedialog.asksaveasfilename(
        parent=self,
        title="Save HIGH Risk Clients Excel",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialfile=default_name,
    )
    if not path:
        return

    wb = openpyxl.Workbook()
    ws_cfg = wb.active
    ws_cfg.title = "Export settings"

    exported_by = str(getattr(self, "_current_username", "") or "").strip()
    if not exported_by:
        try:
            exported_by = (getpass.getuser() or "").strip()
        except Exception:
            exported_by = ""
    if not exported_by:
        exported_by = "Unknown user"

    selected_inds = sorted(getattr(self, "_sim_selected_industries", set()) or set(), key=str.lower)
    industry_note = "None (no industry filter)" if not selected_inds else " · ".join(selected_inds)
    search_note = _search_term if _search_term else "None"

    _write_configuration_settings_sheet(
        ws_cfg,
        fname=Path(str(getattr(self, "_lu_filepath", "") or "—")).name,
        generated_at=datetime.now().strftime("%B %d, %Y  %H:%M"),
        exported_by=exported_by,
        export_scope_note=(
            "HIGH Risk Clients export — includes only clients with % Net → Amort ≥ 71% "
            "under current simulator what-if values. "
            f"Industry filter: {industry_note}. Client search: {search_note}."
        ),
    )

    # HIGH Risk Clients sheet.
    ws = wb.create_sheet("HIGH Risk Clients")
    NUM_COLS = 21

    # ── Banner row (row 1) ────────────────────────────────────────────
    BANNER_COLOR = "C0392B"
    BANNER_TEXT  = "HIGH RISK CLIENTS — % Net Income to Amortization ≥ 71%"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
    banner_cell = ws.cell(1, 1, BANNER_TEXT)
    banner_cell.fill      = PatternFill("solid", fgColor=BANNER_COLOR)
    banner_cell.font      = Font(bold=True, size=11, color="FFFFFF")
    banner_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Summary row (row 2) ───────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NUM_COLS)
    summary_text = (
        f"Total HIGH risk clients: {len(high_risk_rows)}     "
        f"Exported: {datetime.now().strftime('%B %d, %Y  %H:%M')}     "
        f"Exported by: {exported_by}"
    )
    summary_cell = ws.cell(2, 1, summary_text)
    summary_cell.fill      = PatternFill("solid", fgColor="FADBD8")
    summary_cell.font      = Font(italic=True, size=9, color="7B241C")
    summary_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 16

    # ── Header row (row 3) ────────────────────────────────────────────
    headers = [
        "Client ID",
        "PN",
        "Client",
        "Residence Address",
        "Office Address",
        "Industry",
        "Loan Status",
        "AO Name",
        "Product Name",
        "Loan Balance",
        "Principal Loan",
        "Total Expenses (Base)",
        "Total Expenses (Sim)",
        "Total Net Income (Base)",
        "Total Net Income (Sim)",
        "% Increase",
        "Simulated Increase",
        "Total Current Amort",
        "% Net → Amort",
        "Risk Label",
        "Risk Reasoning",
    ]

    HDR_FILL  = PatternFill("solid", fgColor="922B21")
    HDR_FONT  = Font(bold=True, size=10, color="FFFFFF")
    HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side  = Side(style="thin", color="D5D8DC")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(3, ci, h)
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = HDR_ALIGN
        cell.border    = cell_border
    ws.row_dimensions[3].height = 30

    # ── Column widths ─────────────────────────────────────────────────
    col_widths = [14, 14, 24, 24, 24, 18, 16, 18, 20, 16, 16, 18, 18, 20, 20, 12, 18, 18, 12, 12, 52]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    NUM_FMT = '"₱"#,##0.00'
    PCT_FMT = '0.0"%"'

    FILL_ODD       = PatternFill("solid", fgColor="FDEDEC")
    FILL_EVEN      = PatternFill("solid", fgColor="FFFFFF")
    DATA_FONT      = Font(size=9, color="1A252F")
    DATA_FONT_BOLD = Font(bold=True, size=9, color="C0392B")
    DATA_ALIGN_L   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    DATA_ALIGN_C   = Alignment(horizontal="center", vertical="center")
    DATA_ALIGN_R   = Alignment(horizontal="right",  vertical="center")

    for idx, r in enumerate(high_risk_rows):
        row_num  = 4 + idx
        row_fill = FILL_ODD if idx % 2 == 0 else FILL_EVEN

        def _dc(col, value, fmt=None, align=DATA_ALIGN_R, font=DATA_FONT):
            cell = ws.cell(row_num, col, value)
            cell.fill      = row_fill
            cell.font      = font
            cell.border    = cell_border
            cell.alignment = align
            if fmt:
                cell.number_format = fmt
            return cell

        _dc(1,  r["client_id"] or "—",   align=DATA_ALIGN_L)
        _dc(2,  r["pn"] or "—",          align=DATA_ALIGN_L)
        _dc(3,  r["client"],             align=DATA_ALIGN_L)
        _dc(4,  r["residence_address"] or "—", align=DATA_ALIGN_L)
        _dc(5,  r["office_address"] or "—",    align=DATA_ALIGN_L)
        _dc(6,  r["industry"],           align=DATA_ALIGN_L)
        _dc(7,  r["loan_status"] or "—", align=DATA_ALIGN_L)
        _dc(8,  r["ao_name"] or "—",     align=DATA_ALIGN_L)
        _dc(9,  r["product_name"] or "—",align=DATA_ALIGN_L)
        _dc(10, r["loan_balance"],       fmt=NUM_FMT)
        _dc(11, r["principal_loan"],     fmt=NUM_FMT)
        _dc(12, r["base_total_expenses"],fmt=NUM_FMT)
        _dc(13, r["sim_total_expenses"], fmt=NUM_FMT)
        _dc(14, r["net_income"],         fmt=NUM_FMT)
        _dc(15, r["sim_net_income"],     fmt=NUM_FMT)
        _dc(16, r["pct_increase"],       fmt=PCT_FMT, align=DATA_ALIGN_C)
        _dc(17, r["sim_increase"],       fmt=NUM_FMT)
        _dc(18, r["current_amort"],      fmt=NUM_FMT)
        _dc(19, r["pct_net_to_amort"],   fmt=PCT_FMT, align=DATA_ALIGN_C)
        _dc(20, r["sim_risk_label"],     align=DATA_ALIGN_C, font=DATA_FONT_BOLD)
        _dc(21, r["risk_reasoning"],     align=DATA_ALIGN_L)
        ws.row_dimensions[row_num].height = 18

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(NUM_COLS)}{3 + len(high_risk_rows)}"

    wb.save(path)
    messagebox.showinfo(
        "Export Complete",
        f"HIGH Risk Clients Excel saved to:\n{path}\n\n"
        f"{len(high_risk_rows)} HIGH risk client(s) exported.",
        parent=self,
    )


def _sim_update_client_pagination_ui(self, *, total_rows: int):
    """Update Prev/Next buttons and page label for the simulator client table."""
    total = int(total_rows or 0)
    page = int(getattr(self, "_sim_client_page", 0) or 0)
    max_page = max(0, (total - 1) // SIM_CLIENT_PAGE_SIZE) if total else 0
    if page < 0:
        page = 0
    if page > max_page:
        page = max_page
    try:
        lbl = getattr(self, "_sim_client_page_lbl", None)
        if lbl is not None:
            lbl.config(text=f"Page {page + 1} of {max_page + 1 if total else 1}")
    except Exception:
        pass
    try:
        cnt = getattr(self, "_sim_client_count_lbl", None)
        if cnt is not None:
            start = page * SIM_CLIENT_PAGE_SIZE + 1 if total else 0
            end = min((page + 1) * SIM_CLIENT_PAGE_SIZE, total) if total else 0
            cnt.config(text=f"{start}–{end} of {total} client(s)")
    except Exception:
        pass
    try:
        prev_btn = getattr(self, "_sim_client_prev_btn", None)
        next_btn = getattr(self, "_sim_client_next_btn", None)
        if prev_btn is not None:
            prev_btn.configure(state=("normal" if total and page > 0 else "disabled"))
        if next_btn is not None:
            next_btn.configure(state=("normal" if total and page < max_page else "disabled"))
    except Exception:
        pass


def _sim_client_page_prev(self):
    self._sim_client_page = max(0, int(getattr(self, "_sim_client_page", 0) or 0) - 1)
    _sim_refresh_client_table(self)


def _sim_client_page_next(self):
    total = len(list(getattr(self, "_sim_recs", []) or []))
    max_page = max(0, (total - 1) // SIM_CLIENT_PAGE_SIZE) if total else 0
    self._sim_client_page = min(max_page, int(getattr(self, "_sim_client_page", 0) or 0) + 1)
    _sim_refresh_client_table(self)


# ══════════════════════════════════════════════════════════════════════
#  PIE CHART  (simulated expense mix — % of total)
# ══════════════════════════════════════════════════════════════════════

def _sim_draw_chart(self):
    holder = getattr(self, "_sim_chart_holder", None)
    if holder is None:
        return
    try:
        if not holder.winfo_exists():
            return
    except Exception:
        return

    for w in holder.winfo_children():
        w.destroy()

    expenses = [e for e in getattr(self, "_sim_expenses", []) if e["total"] > 0]
    if not expenses:
        tk.Label(
            holder,
            text="No numeric data\nto chart.",
            font=F(9),
            fg=_TXT_MUTED,
            bg=_CARD_WHITE,
            justify="center",
        ).pack(pady=40)
        return

    expenses = expenses[:SIM_CHART_MAX_BARS]

    def _sim_amount(exp):
        pct = 0.0
        var = self._sim_sliders.get(exp["name"])
        if var:
            try:
                pct = float(var.get() or 0)
            except Exception:
                pass
        base = float(exp["total"] or 0)
        return max(0.0, base + base * (pct / 100.0))

    pairs = [(e["name"], _sim_amount(e)) for e in expenses]
    pairs.sort(key=lambda x: -x[1])
    names = [p[0] for p in pairs]
    vals = [p[1] for p in pairs]

    if sum(vals) <= 0:
        tk.Label(
            holder,
            text="No simulated amounts\nto chart.",
            font=F(9),
            fg=_TXT_MUTED,
            bg=_CARD_WHITE,
            justify="center",
        ).pack(pady=40)
        return

    if len(pairs) > PIE_MAX_SLICES:
        top = pairs[: PIE_MAX_SLICES - 1]
        other_sum = sum(p[1] for p in pairs[PIE_MAX_SLICES - 1 :])
        names = [p[0] for p in top] + (["Other"] if other_sum > 0 else [])
        vals = [p[1] for p in top] + ([other_sum] if other_sum > 0 else [])

    if not _HAS_MPL:
        lines = [f"{n[:22]}{'…' if len(n) > 22 else ''}: {v/sum(vals)*100:.1f}%"
                 for n, v in zip(names, vals)]
        tk.Label(
            holder,
            text="matplotlib unavailable.\n\n" + "\n".join(lines[:12]),
            font=F(7),
            fg=_TXT_SOFT,
            bg=_CARD_WHITE,
            justify="left",
        ).pack(padx=6, pady=8)
        return

    def _short(n: str, w: int = 18) -> str:
        n = str(n or "").strip()
        return n if len(n) <= w else n[: w - 1] + "…"

    try:
        fig, ax = plt.subplots(figsize=(4.7, 4.9))
        fig.patch.set_facecolor(_CARD_WHITE)
        ax.set_facecolor(_CARD_WHITE)

        colors = [plt.cm.Pastel2(i % 8) for i in range(len(vals))]
        wedges, _texts, autotexts = ax.pie(
            vals,
            labels=None,
            colors=colors,
            startangle=90,
            counterclock=False,
            autopct=lambda p: f"{p:.1f}%" if p >= 4.5 else "",
            pctdistance=0.80,
            textprops={"fontsize": 8, "color": "#243B64", "fontweight": "bold"},
            wedgeprops={"width": 0.44, "linewidth": 1.0, "edgecolor": _CARD_WHITE},
        )
        for t in autotexts:
            t.set_fontsize(8)
        total_sim = sum(vals)
        ax.text(
            0, 0,
            f"Total\nP{total_sim:,.0f}",
            ha="center",
            va="center",
            fontsize=9,
            color="#365B8C",
            fontweight="bold",
        )
        ax.set_title("Share of Total (Simulated)", fontsize=9, color="#4A6FA5", pad=6)

        leg_labels = [_short(n, 22) for n in names]
        ax.legend(
            wedges,
            leg_labels,
            loc="upper center",
            bbox_to_anchor=(0.5, -0.07),
            ncol=2,
            fontsize=6.5,
            frameon=False,
        )
        fig.subplots_adjust(left=0.06, right=0.94, top=0.88, bottom=0.30)

        canvas = FigureCanvasTkAgg(fig, master=holder)
        widget = canvas.get_tk_widget()
        widget.config(width=370, height=370)
        widget.pack_propagate(False)
        widget.pack(fill="none", expand=False)
        plt.close(fig)
    except Exception:
        tk.Label(
            holder,
            text="Could not draw chart.",
            font=F(9),
            fg=_TXT_MUTED,
            bg=_CARD_WHITE,
        ).pack(pady=24)


# ══════════════════════════════════════════════════════════════════════
#  ATTACH
# ══════════════════════════════════════════════════════════════════════

def attach(cls):
    """
    Attach Risk Simulator methods to the app class.
    Call AFTER lu_tab_analysis.attach(cls).
    """
    cls._build_simulator_panel    = _build_simulator_panel
    cls._build_sim_summary_cards  = _build_sim_summary_cards
    cls._sim_show_placeholder     = _sim_show_placeholder
    cls._sim_open_expense_table_window = _sim_open_expense_table_window
    cls._sim_render_expense_table_rows = _sim_render_expense_table_rows
    cls._sim_populate             = _sim_populate
    cls._sim_build_expense_row    = _sim_build_expense_row
    cls._sim_on_slide             = _sim_on_slide
    cls._sim_apply_global         = _sim_apply_global
    cls._sim_reset                = _sim_reset
    cls._sim_refresh              = _sim_refresh
    cls._sim_draw_chart           = _sim_draw_chart
    cls._sim_export_client_impact_excel = _sim_export_client_impact_excel
    cls._sim_merge_excel_files = _sim_merge_excel_files
    cls._sim_export_high_risk_clients_excel = _sim_export_high_risk_clients_excel
    cls._sim_open_risk_ranges_dialog = _sim_open_risk_ranges_dialog
    cls._sim_on_client_impact_row_activated = _sim_on_client_impact_row_activated
    cls._sim_show_client_details  = _sim_show_client_details
    cls._sim_show_table_view      = _sim_show_table_view
    cls._sim_show_detail_view     = _sim_show_detail_view