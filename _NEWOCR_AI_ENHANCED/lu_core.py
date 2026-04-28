"""
lu_core.py — LU Analysis: dynamic industry‑based risk
=======================================================
Parses an Excel file (LATEST_DUMMY format) and extracts client data.
Risk uses industry tags, optional Product Name overrides, and user expense overrides.

Public functions:
  run_lu_analysis(filepath) -> dict
  set_high_risk_industries(industries)
  add_high_risk_industry(industry)
  remove_high_risk_industry(industry)
  get_high_risk_industries() -> list
  set_product_risk_overrides(mapping)   # atomic product name (lower) -> "HIGH"|"LOW"
  get_product_risk_overrides() -> dict
  split_product_name_tokens(s) -> list[str]   # "A, B" -> ["A","B"]
  lookup_product_risk_override(s) -> ("HIGH"|None, matched_part)
"""

import re
import unicodedata
from decimal import Decimal
from pathlib import Path
from datetime import datetime
import json

try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# ─────────────────────────────────────────────────────────────────────
#  GLOBAL HIGH‑RISK INDUSTRY SET (can be modified by UI)
# ─────────────────────────────────────────────────────────────────────
_HIGH_RISK_INDUSTRIES = set()
_MEDIUM_RISK_INDUSTRIES = set()
_SETTINGS_PATH = Path(__file__).with_name("lu_risk_settings.json")


def _load_risk_settings() -> None:
    """Best-effort load of persisted risk settings from JSON."""
    global _HIGH_RISK_INDUSTRIES, _MEDIUM_RISK_INDUSTRIES, _PRODUCT_RISK_OVERRIDES, _EXPENSE_RISK_OVERRIDES
    try:
        if not _SETTINGS_PATH.exists():
            return
        data = json.loads(_SETTINGS_PATH.read_text(encoding="utf-8") or "{}")
        inds = data.get("high_risk_industries") or []
        if isinstance(inds, list):
            _HIGH_RISK_INDUSTRIES = set(str(x) for x in inds if str(x).strip())
        med_inds = data.get("medium_risk_industries") or []
        if isinstance(med_inds, list):
            _MEDIUM_RISK_INDUSTRIES = set(str(x) for x in med_inds if str(x).strip())
        prod = data.get("product_risk_overrides") or {}
        if isinstance(prod, dict):
            set_product_risk_overrides(prod)
        exp = data.get("expense_risk_overrides") or {}
        if isinstance(exp, dict):
            set_expense_risk_overrides(exp)
    except Exception:
        # Never fail analysis due to settings file issues.
        return


def _save_risk_settings() -> None:
    """Best-effort persist of current risk settings to JSON."""
    try:
        payload = {
            "high_risk_industries": sorted({str(x) for x in _HIGH_RISK_INDUSTRIES if str(x).strip()}),
            "medium_risk_industries": sorted({str(x) for x in _MEDIUM_RISK_INDUSTRIES if str(x).strip()}),
            "product_risk_overrides": get_product_risk_overrides(),
            "expense_risk_overrides": get_expense_risk_overrides(),
        }
        _SETTINGS_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        return

def set_high_risk_industries(industries):
    """Replace the high‑risk industry set with a new list."""
    global _HIGH_RISK_INDUSTRIES, _MEDIUM_RISK_INDUSTRIES
    _HIGH_RISK_INDUSTRIES = set(industries)
    _MEDIUM_RISK_INDUSTRIES = {x for x in _MEDIUM_RISK_INDUSTRIES if x not in _HIGH_RISK_INDUSTRIES}
    _save_risk_settings()


def set_medium_risk_industries(industries):
    """Replace the medium-risk industry set with a new list."""
    global _MEDIUM_RISK_INDUSTRIES
    _MEDIUM_RISK_INDUSTRIES = set(industries)
    _MEDIUM_RISK_INDUSTRIES.difference_update(_HIGH_RISK_INDUSTRIES)
    _save_risk_settings()

def add_high_risk_industry(industry):
    """Mark a single industry as HIGH risk."""
    _HIGH_RISK_INDUSTRIES.add(industry)
    _save_risk_settings()

def remove_high_risk_industry(industry):
    """Remove an industry from the high‑risk set (becomes LOW risk)."""
    _HIGH_RISK_INDUSTRIES.discard(industry)
    _save_risk_settings()

def get_high_risk_industries():
    """Return the current list of high‑risk industries."""
    return list(_HIGH_RISK_INDUSTRIES)


def get_medium_risk_industries():
    """Return the current list of medium-risk industries."""
    return list(_MEDIUM_RISK_INDUSTRIES)


def format_lu_active_risk_settings_summary(max_len: int = 200) -> str:
    """
    Short description of active risk rules for export titles (not exhaustive).
    """
    n_ind = len(_HIGH_RISK_INDUSTRIES)
    n_mid_ind = len(_MEDIUM_RISK_INDUSTRIES)
    n_prod = len(_PRODUCT_RISK_OVERRIDES)
    n_exp = len(_EXPENSE_RISK_OVERRIDES)
    parts = [
        f"{n_ind} HI industry tag(s)" if n_ind else "0 HI industry tags",
        f"{n_mid_ind} MED industry tag(s)" if n_mid_ind else "0 MED industry tags",
        f"{n_prod} product override(s)" if n_prod else "0 product overrides",
        f"{n_exp} expense override(s)" if n_exp else "0 expense overrides",
        "auto fuel/LPG signal",
    ]
    s = "; ".join(parts)
    if len(s) > max_len:
        return s[: max_len - 1] + "…"
    return s


# ─────────────────────────────────────────────────────────────────────
#  PRODUCT NAME RISK OVERRIDES (HIGH / LOW per Product Name cell)
# ─────────────────────────────────────────────────────────────────────
_PRODUCT_RISK_OVERRIDES: dict[str, str] = {}
_EXPENSE_RISK_OVERRIDES: dict[str, str] = {}

_EXPENSE_WORD_FIXES = {
    "transpotation": "transportation",
    "trasportation": "transportation",
    "transporation": "transportation",
    "groceris": "groceries",
    "grocerry": "grocery",
    "eletricity": "electricity",
    "maintenace": "maintenance",
    "maintainance": "maintenance",
    "cigarts": "cigarettes",
    "cigartes": "cigarettes",
}

# Expense settings reference (mirrors expense_categories.pdf).
# Structure:
#   EXPENSE_CATEGORIES["household"|"business"][slug] = {"Title": [item variants...]}
EXPENSE_CATEGORIES: dict = {
    "household": {
        "food_groceries": {"Food & Groceries": [
            "Food", "Food & Groceries", "Food & Personal Expense", "Groceries",
            "Grocery", "Groceries Share", "Food & Groceries Share", "Food & Groceries (Share)",
            "Share for food & Groceries", "Groceries & Child Expense", "Palay Buying",
            "Rice Milling", "Rice Allowance", "Food & Groceries & Rice", "Food/Groceries",
            "Food and Groceries", "Allocated Meals (lunch)", "Food / Snack (work)",
            "Grocey (Client)",
        ]},
        "utilities_electricity": {"Utilities – Electricity": [
            "Light", "Electricity", "Electric", "Electric Bill", "Electric bill", "Electricity Bill",
        ]},
        "utilities_water": {"Utilities – Water": [
            "Water", "Water Bill", "Drinking Water", "Water *Mineral", "Mineral Water", "Tap Water",
            "BADEVCO amortization", "BADEVCO Payment",
        ]},
        "utilities_lpg": {"Utilities – Gas / LPG": [
            "LPG", "LPG/GASUL", "GASUL", "LPG (Solane)", "LPG gas", "LPG/Charcoal",
            "LPG Max Gas", "Charcoal", "Gasul", "Town Gas",
        ]},
        "utilities_internet_cable": {"Utilities – Internet & Cable": [
            "Internet", "Internet Fee", "Cable", "Cable & Internet", "Internet Connection",
            "Cable w/ Internet", "CATV", "CATV/Internet Fees", "Internet (PLDT)",
            "Internet (Converge)", "Internet (Globe)", "Wifi", "Wifi rental fee", "Pocket wifi",
            "Netflix", "Internet Allocation", "SAT-Lite Load",
        ]},
        "phone_load": {"Communication – Phone Load": [
            "CP Load", "CP load", "Load", "Cellphone Load", "Cell Load", "CP LOAD",
            "Load Allocation", "Load (Unli)", "CP Load Plan", "Telephone",
        ]},
        "fuel_transport": {"Fuel & Transportation": [
            "Gasoline", "Gas Expense", "Gas", "Gas Hillux", "Gas Tricycle", "Gasoline Hauler",
            "Motorcycle Gasoline", "Car Gasoline", "Gasoline (Mio)", "Gasoline (TMX)",
            "Gasoline (L300)", "Diesel", "Fuel", "Transportation/Gasoline", "Fare", "Transportation",
            "Transportation Expense", "Fare expense", "Fare/Allowance", "Transpo", "Car Pool",
            "Gas expense", "Gasoline Expense",
            # Public transport / for-hire (common free-text labels; must map here for risk overrides)
            "Tricycle Driving", "Tricycle Driver", "Tricycle Operation", "Tricycle Boundary",
            "Motorcycle Taxi", "Habal-habal", "Pedicab", "Motorela",
        ]},
        "vehicle_maintenance": {"Vehicle Maintenance": [
            "Change Oil", "Maintenance/Change Oil", "Motor Maintenance", "Tricycle Maintenance",
            "Motorcycle Maintenance", "Maintenance Allocation", "Vehicle Maintenance", "Car Maintenance",
            "Maintenance/Change Tire", "Change Tire", "Tires", "Tire", "LTO Renewal",
            "Hauler Maintenance", "Hauler Renewal", "Vehicle Renewal", "Motorcycle Renewal",
            "LTO Renewal & maintenance", "Comprehensive Insurance", "Car Maintenance & Renewal",
            "Change-oil", "Change-Tire", "Chattel Expenses", "Van Maintainance", "Franchise Renewal",
        ]},
        "education": {"Education": [
            "School Allowance", "Children's School Allowance", "Student Allowance", "Kids Allowance",
            "Allowance of Children", "Tuition Fee", "Tuition", "School Project", "Project Allocation",
            "School Expenses", "School Misc", "Miscellaneous fee", "Board & Lodging", "Boarding House",
            "Dorm", "Dormitory", "Dorm Fee", "Ed. Allowance", "Educational allowance", "E.S. Allowance",
            "H.S. Allowance", "Tutor", "Allocated School Supplies", "Monthly School Due",
        ]},
        "loans_amort": {"Loans & Amortizations": [
            "BSV Amortization", "BSV Loan Amortization", "BSV Loan Payment", "Pag-Ibig Housing",
            "Pag-ibig amortization", "Housing Loan amortization", "Car Loan Amortization",
            "Car Amortization", "Toyota Amortization", "Motorcycle Loan", "Motorcycle Amortization",
            "RBP Amortization", "RB Paracale Amortization", "Home Credit", "Home Credit Amortization",
            "Loan Payment", "Loan Installment", "Loan", "NCCB Loan Amortization",
            "Card Bank Amortization", "East West Bank", "PS Bank Amortization", "PSBank",
            "BDO Amortization", "REM Loan", "LBP Amortization", "Legazpi Savings Bank",
            "Landbank Amortization", "Motortrade Amortization", "Cardbank", "Cardbank loan",
            "ARDCI Amortization", "JMH Amortization", "JMH Microfinance", "Bangko ng Kabuhayan Amort",
            "SBCorp", "RCBC credit card amort", "Gcash Loan", "Fuse Financing", "SeaMoney",
            "Pawnshop Amortization", "Malayan Bank", "NCCB-Daet amortization", "SpayLater Payment",
            "GSIS Amortization", "Meralco Employee Loan Ass.", "Lucky Five Installment",
            "TFS Ammortization", "Union Bank", "BPI Banco", "CIMB Bank", "Net bank (Ggives)",
            "Card amortization", "Wigo Amortization", "TOYOTA FINANCING", "Lot Installment",
        ]},
        "credit_cards": {"Credit Card Payments": [
            "Credit Card Payment", "Credit Card", "Credit Card Allocation", "Credit cards",
            "credit card ave. payment", "UB Credit Card",
        ]},
        "insurance_gov": {"Insurance & Gov't Contributions": [
            "Insurance", "Life Insurance", "St. Peter", "St.Peter", "SSS/St. Peter",
            "Insular/FWD Insurance", "FWD Insurance", "Prulife UK", "BDO Life",
            "Insurance (Prulife)", "Comprehensive Insurance", "Fortuner Insurance",
            "SSS", "SSS Contribution", "Philhealth", "Pag-ibig Contribution",
            "Pag Ibig", "PhilHealth", "SSS & Philhealth", "Tithes", "St. Peter life plan",
        ]},
        "housing_rent": {"Housing & Rent": [
            "Apartment Rental", "House Rental", "House rental", "Rental", "Rent",
            "Housing Loan amortization", "House Rent", "Boarding House", "Dorm",
        ]},
        "household_help": {"Household Help": [
            "Helper", "House helper", "Housemaid", "Housekeeper", "Kasambahay", "Nanny",
            "Baby sitter", "Laundry", "Laundress", "Janitor Fee", "2 Kasambahay",
            "House keeper salary",
        ]},
        "health_medicine": {"Health & Medicine": [
            "Medicine", "Medicines", "Vitamins", "Vitamins & Fruits", "Vitamins and Fruits",
            "Vitamins and Medicines", "Milk", "Milk & Diapers", "Milk/Diaper/Vitamins", "Diapers",
            "Diaper", "Milk, Vitamins, Diapers & Water", "Medicine Maintenance", "Medical Allocation",
            "Medicine Allocation", "Check-up Allocation", "Medical Maintenance", "Therapy",
            "Maintenance medicines",
        ]},
        "personal_support": {"Personal & Family Support": [
            "Personal Expenses", "Personal Expense Allocation", "Other Personal Allocation",
            "Financial support", "Financial Share", "Remittance to parent", "Mother Allowance",
            "Mothers Allotment", "Allowance", "Kids Allowance", "Allowance sister", "Family Share",
            "Share for Utilities", "Utilities Share", "Monthly share", "Financial support (parents)",
            "Dog Expense", "Pet supplies", "Recreation Budget", "Other Expenses", "Miscellaneous",
            "Pawn Interest",
        ]},
        "savings_invest": {"Savings & Investments": [
            "Savings", "Personal Investment", "House construction",
        ]},
    },
    "business": {
        "cogs_purchases": {"Cost of Goods / Purchases": [
            "Purchases", "Cost of Sales", "Cost of Goods Per Mark-up", "Purchase Cost",
            "Purchased Cost", "Raw Materials/Purchases", "Buy & Sell", "Buy and Sell",
            "Merchandise Inventory Expense", "Bulk surplus", "BIGASAN", "FEEDS",
            "SARI-SARI STORE", "Sari-sari Store", "Sari Sari Store", "Softdrinks",
            "Liquor", "Beverages", "Cigarettes", "Rice Retailing", "Rice", "Egg",
            "Fish Vending", "Chicken", "Pork", "Meat Shop", "Vegetables", "Frozen Products",
            "Ice Cream", "Ice", "Gasoline Retailing", "General Merchandise", "Grocery Store Expense",
            "Pharmacy Expense", "Egg Retail", "Hardware", "RMDC Hardware/CHB/Aggregates",
            "Gravel and Sand", "Hollow Blocks Making", "Cost of Coconut", "Copra", "Corn Expenses",
            "Pineapple Farming Expenses", "Palay Expenses", "Hog Raising",
            "Average purchase price of fattener", "Eatery", "Restaurant/Carenderia", "Canteen",
            "Ukay Bundle", "Japan Surplus", "Copies", "LPG Outlet", "COS Eatery",
            "COS Laundry Shop", "COS Elegancia", "COFFEE VENDO", "LUTONG ULAM", "Piggery",
            "Sow Feed Expense", "Feed Expense", "Cost of Herbal Coffee", "Weekly Purchase",
            "Fruit Selling Cost", "Halloblocks", "Ingredients", "Packaging", "Plastic",
            "SMJ Copra", "Sari-sari Store Purchases", "FISH VENDING", "Trucking Service",
            "Operational Expenses for Coconut", "Operational Expenses for Copras",
            "Cost of Sales - Restaurant", "Atasha Frozen Goods Cost of Sales",
            "Atasha Resto Cost of Sales", "Buy and Sell of Rice", "Meat retailing",
        ]},
        "salaries_labor": {"Salaries & Labor": [
            "Salary", "Salaries", "Salaries and Wages", "Salary/Wages", "Salary/Employee",
            "On Call Cleaner", "Tindero Sahod", "Cook", "Pasahod", "Delivery Assistants",
            "EMPLOYEE and rice subsidy", "Food Allocation (employee)", "Food for the Employee",
            "Helper", "Boy/Helper", "Drivers", "Butcher", "Slaughter", "SSS (Driver)",
            "Labor for catering", "Rider", "Salary (Vice President)", "Salary (gasoline boy)",
            "Salary (Cashier)", "Salesman", "On-call technician", "2 Employee", "3 employees",
            "4 employees", "OverHead *3 employee", "School share", "HAULING driver and pahinante",
            "BOULDER ROCK & DRIVER PAHINANTE",
        ]},
        "rent_space": {"Rent & Space": [
            "Rent", "Rent Expense", "Space Rental", "Stall Rental Fee", "Rental",
            "Office Space Rental", "Land Expense (rent)", "House Rental", "Rental expense",
            "Rental @ Lavadas", "Rental @ CN", "Rent (Business place)", "Monthly Rentals",
        ]},
        "utilities_elec_water": {"Utilities – Electricity & Water": [
            "Electricity Expense", "Electric Bill", "Business Electric Bill", "Water Expense",
            "Water Bill", "Water", "Electricity", "SOAP AND ELECTRICITY",
            "Utilities (Water & Electricity)", "Electric and Water Bill", "LIGHT AT THE HBC",
            "Ebill Share",
        ]},
        "fuel_transport": {"Fuel & Transportation": [
            "Gasoline", "Gas Expense", "Gasoline Expense", "Gasoline (ULD & ADO)",
            "Gasoline/transportation", "Diesel", "Diesel Expense", "Fuel", "Transportation",
            "Transportation Fare", "Transportation of Goods", "Transportation and Accommodation",
            "Travel and Transportation", "Trucking Fee", "Trucking Service", "Trucking expense",
            "Baggage Fee", "Boundary", "Boundery", "Fare Expense", "Fare",
            "Tricycle Driving", "Tricycle Driver", "Tricycle Operation", "Tricycle Boundary",
            "Motorcycle Taxi", "Habal-habal", "Pedicab", "Motorela",
        ]},
        "vehicle_equipment_maint": {"Vehicle & Equipment Maintenance": [
            "Vehicle Maintenance", "Tricycle Maintenance", "Tricycle Expenses", "Tricycle Expense",
            "TRICYCLE EXPENSE", "Tricycle Registration", "Motorcycle Maintenance", "Car Maintenance",
            "Hauler Maintenance", "Motor Hauler Maintenance", "Machine Maintenance", "Truck Maintenance",
            "Repairs in Truck", "Chattel Expenses", "Chariot Maintenance", "Jeep Maintenance",
            "LTO Renewals", "Renewal", "Vehicle Renewal", "Yearly Renewal of Franchise",
            "Renewal of Registration", "Change Oil", "Change Tire", "Vehicle Maintenance Allocation",
            "Maintenance for carwash", "Maintenance Allocation (Truck)", "Maintenance Allocation (Molding Machine)",
            "Maintenance Allocation (Hauler)", "Maintenance motorcycle", "COS Motorshop",
            "Vehicle Maintenance (Hauler)",
        ]},
        "permits_taxes": {"Permits, Taxes & Licenses": [
            "Business Permit", "Business Permits", "Mayor's Permit", "Mayors Permit",
            "Brgy. Permits", "BRGY. PERMITS", "BIR", "Tax Payment", "Tax", "Business Tax",
            "Annual BIR Tax", "Quarterly Tax", "Real Property Tax", "Taxes & Licenses",
            "Taxes and Licenses", "Permit & Taxes", "Permits & Licenses", "Permits/Plasada",
            "Sanitary Permit", "Franchise Registration", "Environmental fee", "LTO/Tricycle Registration",
            "Business (Tax & Licenses)",
        ]},
        "operating_general": {"Operating Expenses (General)": [
            "Operating Expense", "Operating Expenses", "Operation Expense", "OPEX",
            "Other Business Expenses", "Other Business Expense", "Other Operating Expense",
            "Other Expenses", "Business Expense", "Business Expenses", "Buss expense",
            "Miscellaneous Expense", "Misc. Expense", "Regular Expenses per Branch",
            "Total Average Other Expenses", "RRMP Operation Expense", "Restobar other Expenses",
            "Total Operation Expenses", "Operating expesnes",
        ]},
        "loans_amort": {"Loans & Amortizations": [
            "Loan Amortization", "Loan payment: ORIX", "LBP/UCPB", "MPC SME loan payment",
            "Producer's Bank Amortization", "Home Credit Amortization (Gadget)", "Salmon Amortization",
            "BSV Amortization", "RBP Amortization", "BDO Amortization", "BPI amortization (credit card)",
            "Loan Payments: BSV", "Loan Payments (Card)", "Loan Payments (RB Paracale)", "BPI Banko",
            "RBJP amortization", "Term loans (CIC)", "Chattel loan", "ARDCI", "Card",
        ]},
        "supplies_materials": {"Supplies & Materials": [
            "Supplies", "Cleaning Materials", "Plastic", "Plastic gloves", "Plastic cup",
            "Dishwashing liquid", "Soap, Fabcon, Bleach, Alcohol Hand soap", "Tissue",
            "Amenity Kit", "Packaging", "Store Supplies", "Office Supplies", "Materials",
            "Laundry and Cleaning Materials", "Condiments and Supplies", "GASUL",
            "LPG (business)", "Internet (business)", "Distilled Water",
        ]},
        "bookkeeping_prof": {"Bookkeeping & Professional Fees": [
            "Bookkeeper", "Bookkeeper & Tax Expense", "Operating Exp. (Bookkeeping)",
            "Salary (Vice President)", "Accounting", "Professional fees",
        ]},
        "agri_farming": {"Agriculture & Farming": [
            "Piggery expense", "Palay Expenses", "Copra Expenses", "Corn Expenses",
            "Pineapple Farming Expenses", "Hog Raising", "Sow Expenses", "Sow Feed Expense",
            "Stud (Sow)", "Feed Expense (Fattener)", "Vitamins & Injection (Fattener)",
            "Planting Expense", "Farming Expense", "Rice Land", "Dogs Maintenance",
            "Copra Farming", "Gravel and Sand", "Boulder Rock", "Crudo", "NE of Hauler",
            "Operational Expenses for Coconut", "Operational Expenses for Copras",
            "Cost of Coconut", "Client's Copras Capital", "Hatchery",
        ]},
        "insurance_business": {"Insurance (Business)": [
            "Marine Insurance", "Comprehensive Insurance", "SSS/Pag ibig/Philhealth",
            "SSS (Driver)", "Employer contributions",
        ]},
        "other_misc": {"Other / Miscellaneous": [
            "Other expenses", "Water Refilling", "WATER REFFILLING", "Ice Making",
            "Internet Fee (business)", "Plasada", "Tracking Fee of Goods", "RBP Voluntary Savings",
            "Share to in-laws", "Snack Allowance", "Refreshments Cost", "Delivery Expense",
            "Right of way", "Sand", "Cake making", "Apartment Repair & Maintenance",
            "Maintenance materials", "Dental Chair", "Food (business)",
        ]},
    },
}

# Canonical expense titles that always contribute a fuel / LPG risk signal
# when present in parsed detail lines (aligned with EXPENSE_CATEGORIES slugs).
def _fuel_lpg_sensitive_canonical_titles() -> frozenset[str]:
    titles: set[str] = set()
    for section in ("household", "business"):
        sec = EXPENSE_CATEGORIES.get(section, {}) or {}
        for slug in ("fuel_transport", "utilities_lpg"):
            block = sec.get(slug) or {}
            if not isinstance(block, dict):
                continue
            for title in block.keys():
                if isinstance(title, str) and title.strip():
                    titles.add(title.strip())
    return frozenset(titles)


_FUEL_LPG_CANONICAL_TITLES = _fuel_lpg_sensitive_canonical_titles()


def _canon_title_is_fuel_or_lpg(canonical: str) -> bool:
    """True if canonicalized label is Fuel & Transportation or Utilities – Gas / LPG."""
    if not canonical or not str(canonical).strip():
        return False
    t = canonical.replace("–", "-").replace("—", "-").strip().lower()
    for ref in _FUEL_LPG_CANONICAL_TITLES:
        if ref.replace("–", "-").replace("—", "-").strip().lower() == t:
            return True
    return False


def _build_expense_keyword_groups() -> tuple[tuple[str, tuple[str, ...]], ...]:
    """
    Build keyword groups from EXPENSE_CATEGORIES so matching stays in sync
    with the PDF reference. Category title becomes the canonical label.
    """
    groups: list[tuple[str, list[str]]] = []
    for _section in ("household", "business"):
        sec = EXPENSE_CATEGORIES.get(_section, {}) or {}
        for _slug, block in sec.items():
            if not isinstance(block, dict):
                continue
            for title, items in block.items():
                kws: list[str] = []
                for it in items or []:
                    s = str(it or "").strip()
                    if s:
                        kws.append(s)
                # Also allow matching by the category title itself.
                kws.append(str(title))
                groups.append((str(title), kws))

    # Make longer/phrase keywords match first by sorting groups by max keyword length.
    groups.sort(key=lambda g: max((len(k) for k in g[1]), default=0), reverse=True)
    # Normalize keywords for matching, and expand phrases into token/bigram keys
    # so shorter user-entered phrases can still map into the correct category.
    # IMPORTANT: avoid generic tokens that cause false-positive category matches.
    _GENERIC_TOKENS = {
        "expense", "expenses", "cost", "costs", "allocation", "alloc", "payment", "payments",
        "loan", "loans", "amortization", "amortizations", "amort", "other", "general",
        "business", "personal", "household", "share", "total", "monthly", "fee", "fees",
        "misc", "miscellaneous", "and", "for", "the", "per", "of", "to",
    }
    out: list[tuple[str, tuple[str, ...]]] = []
    for title, kws in groups:
        norm = []
        seen = set()
        for k in kws:
            kk = unicodedata.normalize("NFKC", str(k)).strip().lower()
            kk = kk.replace("&", " and ")
            # Match normalization used by _normalize_expense_name()
            kk = re.sub(r"[^a-z0-9\s/]", " ", kk)
            kk = re.sub(r"\s+", " ", kk)
            if not kk:
                continue
            def _add(x: str):
                x = (x or "").strip()
                if not x:
                    return
                if x in seen:
                    return
                seen.add(x)
                norm.append(x)

            _add(kk)
            toks = [t for t in kk.split(" ") if t]
            # include bigrams only.
            # Single-token expansion caused false positives (e.g. "gasoline"
            # matching unrelated categories that contain "gasoline retailing").
            for i in range(len(toks) - 1):
                a, b = toks[i], toks[i + 1]
                if a in _GENERIC_TOKENS or b in _GENERIC_TOKENS:
                    continue
                bg = f"{a} {b}"
                _add(bg)
        out.append((title, tuple(norm)))
    return tuple(out)


_EXPENSE_KEYWORD_GROUPS: tuple[tuple[str, tuple[str, ...]], ...] = _build_expense_keyword_groups()


# Split Excel "Product Name" cells into atomic products (same separators as industry tags).
_PRODUCT_NAME_SPLIT_RE = re.compile(r"\s*(?:,|/|;|&|\band\b)\s*", re.I)


def split_product_name_tokens(product_name: str) -> list[str]:
    """
    Split one Product Name cell into atomic labels, e.g.
    'CHATTEL LOAN, MOTORCYCLE LOAN' -> ['CHATTEL LOAN', 'MOTORCYCLE LOAN'].
    """
    raw = str(product_name or "").strip()
    if not raw:
        return []
    parts = [p.strip() for p in _PRODUCT_NAME_SPLIT_RE.split(raw) if p.strip()]
    return parts if parts else [raw]


def lookup_product_risk_override(product_name: str) -> tuple[str | None, str]:
    """
    Match risk overrides against each atomic product in the cell.
    Returns ('HIGH'|'MEDIUM'|None, matched_atomic_name for messaging).
    Also accepts legacy keys where the entire cell was stored as one string.
    """
    raw = str(product_name or "").strip()
    if not raw:
        return (None, "")
    keys = _PRODUCT_RISK_OVERRIDES
    order = {"HIGH": 0, "MEDIUM": 1, "MODERATE": 1, "LOW": 2}
    best_lvl = None
    best_tok = ""
    for tok in split_product_name_tokens(raw):
        k = tok.strip().lower()
        if not k:
            continue
        lvl = str(keys.get(k) or "").upper()
        if lvl not in ("HIGH", "MEDIUM", "MODERATE", "LOW"):
            continue
        lvl = "MEDIUM" if lvl == "MODERATE" else lvl
        if best_lvl is None or order[lvl] < order[best_lvl]:
            best_lvl = lvl
            best_tok = tok.strip()
            if best_lvl == "HIGH":
                break
    if best_lvl is None:
        lvl = str(keys.get(raw.lower()) or "").upper()
        if lvl in ("HIGH", "MEDIUM", "MODERATE", "LOW"):
            best_lvl = "MEDIUM" if lvl == "MODERATE" else lvl
            best_tok = raw
    return (best_lvl, best_tok)


def set_product_risk_overrides(mapping: dict) -> None:
    """
    Replace product overrides.

    Keys should be atomic product names (comma-separated cells are split for matching).
    Product settings accept HIGH/MEDIUM; LOW is treated as "not set".
    """
    global _PRODUCT_RISK_OVERRIDES
    out: dict[str, str] = {}
    for k, v in (mapping or {}).items():
        val = str(v or "").strip().upper()
        if val == "MODERATE":
            val = "MEDIUM"
        if val not in ("HIGH", "MEDIUM"):
            continue
        raw = str(k or "").strip()
        if not raw:
            continue
        for part in split_product_name_tokens(raw):
            ks = part.strip().lower()
            if ks:
                out[ks] = val
    _PRODUCT_RISK_OVERRIDES = out
    _save_risk_settings()


def get_product_risk_overrides() -> dict[str, str]:
    """Lowercased product name -> HIGH or LOW."""
    return dict(_PRODUCT_RISK_OVERRIDES)


def set_expense_risk_overrides(mapping: dict) -> None:
    """
    Replace expense risk settings.

    Expense settings accept HIGH/MEDIUM.
    LOW entries are treated as "not set" and are not stored.
    This avoids a default-LOW for every category accidentally overriding
    other risk sources (like Industry HIGH).
    """
    global _EXPENSE_RISK_OVERRIDES
    out: dict[str, str] = {}
    for k, v in (mapping or {}).items():
        ks = _expense_override_key(k)
        if not ks:
            continue
        val = str(v or "").strip().upper()
        if val == "MODERATE":
            val = "MEDIUM"
        if val in ("HIGH", "MEDIUM"):
            out[ks] = val
    _EXPENSE_RISK_OVERRIDES = out
    _save_risk_settings()


def get_expense_risk_overrides() -> dict[str, str]:
    """Lowercased expense name -> HIGH or LOW."""
    return dict(_EXPENSE_RISK_OVERRIDES)


# Load persisted settings on import (best-effort).
_load_risk_settings()


def _unstick_common_expense_concatenations(s: str) -> str:
    """Split OCR-glued tokens so e.g. 'TireGasoline' matches fuel keywords."""
    s = str(s or "")
    s = re.sub(r"(?i)(tire)(gasoline|diesel)", r"\1 \2", s)
    s = re.sub(r"(?i)(change)\s*(tire)(gasoline|diesel)", r"\1 \2 \3", s)
    return s


# PH for-hire transport (see _normalize_expense_name and _expense_override_lookup_keys).
_PH_FOR_HIRE_LABEL_MARKERS = (
    "tricycle driving",
    "tricycle driver",
    "tricycle operation",
    "tricycle boundary",
    "habal-habal",
    "habal habal",
    "pedicab",
    "motorela",
    "motorcycle taxi",
)


def _normalize_expense_name(name: str) -> str:
    """
    Canonicalize expense labels so near-duplicates (typos/punctuation variants)
    are treated as the same logical expense.
    """
    s = unicodedata.normalize("NFKC", str(name or ""))
    s = _unstick_common_expense_concatenations(s)
    s = s.replace("&", " and ")
    s = re.sub(r"[^A-Za-z0-9\s/]", " ", s)
    s = re.sub(r"\s+", " ", s).strip().lower()
    if not s:
        return ""
    tokens = []
    for t in s.split():
        tokens.append(_EXPENSE_WORD_FIXES.get(t, t))
    s = " ".join(tokens)

    # For-hire / boundary-style labels (common in PH exports) must map to Fuel & Transportation
    # so they match PDF category risk overrides. Keyword-group sort order otherwise prefers
    # Vehicle & Equipment Maintenance (business) and breaks Fuel-only HIGH flags.
    if any(p in s for p in _PH_FOR_HIRE_LABEL_MARKERS):
        return "Fuel & Transportation"

    # Exact canonical-title match should win before fuzzy keyword matching.
    for canonical, _keywords in _EXPENSE_KEYWORD_GROUPS:
        ck = unicodedata.normalize("NFKC", str(canonical)).strip().lower()
        ck = ck.replace("&", " and ")
        ck = re.sub(r"[^a-z0-9\s/]", " ", ck)
        ck = re.sub(r"\s+", " ", ck).strip()
        if s == ck:
            return canonical

    def _matches_kw(text: str, kw: str) -> bool:
        kw = (kw or "").strip().lower()
        if not kw:
            return False
        # For multi-word / phrase keywords, \\b...\\b matching is unreliable.
        if any(ch.isspace() for ch in kw) or "/" in kw:
            return kw in text
        return bool(re.search(rf"\b{re.escape(kw)}\b", text))

    chosen = None
    for canonical, keywords in _EXPENSE_KEYWORD_GROUPS:
        if any(_matches_kw(s, kw) for kw in keywords):
            chosen = canonical
            break
    # Prefer fuel / LPG categories when the label clearly names motor fuel or LPG,
    # even if a broader bucket (e.g. vehicle maintenance) matched first.
    if chosen and re.search(r"\b(lpg|gasul|gasoline|diesel)\b", s):
        if not _canon_title_is_fuel_or_lpg(chosen):
            for canonical, keywords in _EXPENSE_KEYWORD_GROUPS:
                if not _canon_title_is_fuel_or_lpg(canonical):
                    continue
                if any(_matches_kw(s, kw) for kw in keywords):
                    chosen = canonical
                    break
    if chosen:
        return chosen

    return " ".join(w.capitalize() for w in s.split())


def _expense_override_key(name) -> str:
    """Stable lookup key for expense overrides."""
    return _normalize_expense_name(str(name or "")).lower()


def _expense_override_lookup_keys(raw_expense_name: str) -> list[str]:
    """
    Keys to check in _EXPENSE_RISK_OVERRIDES for one parsed expense line.
    For-hire tricycle labels normalize to Fuel & Transportation but should still
    match Vehicle HIGH flags (household vs business PDF titles use different keys).
    """
    name = str(raw_expense_name or "").strip()
    if not name:
        return []
    keys: list[str] = []
    primary = _expense_override_key(name)
    if primary:
        keys.append(primary)
    low = unicodedata.normalize("NFKC", name).strip().lower()
    if any(m in low for m in _PH_FOR_HIRE_LABEL_MARKERS):
        for ek in ("vehicle & equipment maintenance", "vehicle maintenance"):
            if ek not in keys:
                keys.append(ek)
    return keys


def _detail_text_indicates_fuel_or_lpg(detail: str) -> bool:
    """
    True if free-text household / business expense cells mention LPG, GASUL,
    gasoline, or diesel (handles lines with no bracketed amount).
    """
    if not detail or not str(detail).strip():
        return False
    s = unicodedata.normalize("NFKC", str(detail))
    s = _unstick_common_expense_concatenations(s).lower()
    if re.search(r"\b(lpg|gasul|gasoline|diesel)\b", s):
        return True
    if re.search(r"\blpg\s*/\s*gasul\b", s) or "lpg/gasul" in s:
        return True
    return False


def _apply_fuel_lpg_risk_to_expense_rows(
    expenses: list[dict],
    biz_detail: str,
    hhld_detail: str,
) -> str:
    """
    Tag matching parsed expense rows as HIGH risk for charts / _compute_risk_score,
    and return a short string for risk_reasoning (empty if no fuel/LPG signal).
    """
    detail_hit = _detail_text_indicates_fuel_or_lpg(biz_detail) or _detail_text_indicates_fuel_or_lpg(
        hhld_detail
    )
    matched_name = ""
    for item in expenses or []:
        nm = str((item or {}).get("name") or "").strip()
        if not nm or not _canon_title_is_fuel_or_lpg(nm):
            continue
        tot = float((item or {}).get("total") or 0.0)
        if tot > 0.0 or detail_hit:
            item["risk"] = "HIGH"
            item["reason"] = (
                f"Expense '{nm}' is classified as fuel / LPG–related (sensitive cost category)."
            )
            if not matched_name:
                matched_name = nm
    if detail_hit and not matched_name:
        return "Fuel / LPG-related wording in Business or Household expense details"
    return matched_name


# ─────────────────────────────────────────────────────────────────────
#  CONSTANTS
# ─────────────────────────────────────────────────────────────────────
GENERAL_CLIENT = "📊  General (All Clients)"
_MAX_HEADER_SCAN_ROWS = 10
_MAX_DATA_ROWS = 2000
SECTOR_WHOLESALE  = "Wholesale"
SECTOR_AGRICULTURE = "Agriculture"
SECTOR_TRANSPORT  = "Transport"
SECTOR_REMITTANCE = "Remittance"
SECTOR_CONSUMER   = "Consumer"
SECTOR_OTHER      = "Other"
# ─────────────────────────────────────────────────────────────────────
#  RISK SCORE CONSTANTS (for backward compatibility)
# ─────────────────────────────────────────────────────────────────────
_RISK_ORDER = {"HIGH": 0, "MODERATE": 1, "MEDIUM": 1, "LOW": 2}
_SCORE_BANDS = [
    (2.5, "CRITICAL", "#B71C1C", "#FFEBEE"),
    (1.8, "HIGH",     "#E53E3E", "#FFF5F5"),
    (1.2, "MODERATE", "#D4A017", "#FFFBF0"),
    (0.0, "LOW",      "#2E7D32", "#F0FBE8"),
]
_DEFAULT_RISK = ("LOW", "No specific sector-expense sensitivity rule defined; treat as low risk.")
_INDUSTRY_SPLIT_RE = re.compile(r"\s*(?:,|/|;|&|\band\b)\s*", re.I)
_CANONICAL_INDUSTRIES = (
    "Agriculture, Forestry and Fishing",
    "Accommodation and Food Services",
    "Construction",
    "Education",
    "Electricity, Gas, Steam and Air Conditioning Supply",
    "Household Use",
    "Human Health and Social Activities",
    "Manufacturing",
    "Mining and Quarrying",
    "Others Service Activities",
    "Real Estate Activities",
    "Trading (Wholesale and Retail), Repair of Motor Vehicle and Motorcycle",
    "Transportation and Storage",
)
_CANONICAL_INDUSTRY_PATTERNS = tuple(
    (
        name,
        re.compile(re.escape(name).replace(r"\ ", r"\s+"), re.I),
    )
    for name in sorted(_CANONICAL_INDUSTRIES, key=len, reverse=True)
)

def _compute_risk_score(expenses_or_dummy, industry: str = "", product_name: str = ""):
    """
    Compatibility scorer used by report/loan-balance/charts modules.
    Derives a simple score from provided expense risk tags.
    """
    # Optional `industry` / `product_name` for patched callers (ignored here).
    _ = industry
    _ = product_name
    expenses = expenses_or_dummy or []
    has_high = any((e or {}).get("risk") == "HIGH" for e in expenses)
    has_mod = any((e or {}).get("risk") == "MODERATE" for e in expenses)
    if has_high:
        return (1.8, "HIGH", "#E53E3E", "#FFF5F5")
    if has_mod:
        return (1.2, "MODERATE", "#D4A017", "#FFFBF0")
    return (0.0, "LOW", "#2E7D32", "#F0FBE8")
# ─────────────────────────────────────────────────────────────────────
#  CELL / NUMERIC HELPERS
# ─────────────────────────────────────────────────────────────────────
def _cell_str(cell) -> str:
    if cell is None:
        return ""
    val = getattr(cell, "value", None) if hasattr(cell, "value") else cell
    if val is None:
        return ""
    return str(val).strip()

def _asset_text_field(raw) -> str:
    """
    Personal / Business / Inventory columns are usually narrative text with
    embedded amounts and line breaks, not a single number. Preserve as string.
    """
    if raw is None:
        return ""
    if isinstance(raw, Decimal):
        v = float(raw)
        if v != v:
            return ""
        return f"{v:,.2f}" if v % 1 else str(int(v))
    if isinstance(raw, (int, float)):
        v = float(raw)
        if v != v:
            return ""
        return f"{v:,.2f}" if v % 1 else str(int(v))
    s = str(raw)
    if not s.strip():
        return ""
    s = unicodedata.normalize("NFKC", s)
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    return s.strip()


def _parse_numeric(raw) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, Decimal):
        v = float(raw)
        return v if v == v else None
    if isinstance(raw, (int, float)):
        v = float(raw)
        return v if v == v else None
    txt = str(raw).strip()
    if not txt or txt.startswith("="):
        return None
    negative = False
    # Support both (123.45) and [123.45] wrappers from OCR/table exports.
    if ((txt.startswith("(") and txt.endswith(")"))
            or (txt.startswith("[") and txt.endswith("]"))):
        txt = txt[1:-1].strip()
        # Parentheses usually mean negative in accounting formats.
        negative = raw is not None and str(raw).strip().startswith("(")
    txt = (txt.replace("₱", "").replace("$", "").replace("€", "")
              .replace("£", "").replace(",", "").strip().rstrip("%").strip())
    try:
        v = float(txt)
        return -v if negative else v
    except (ValueError, TypeError):
        return None
def _fmt_value(vals: list) -> str:
    """Stub for compatibility; returns a placeholder."""
    return "—"

def _numeric_total(vals: list) -> float:
    """Compatibility helper used by legacy/report code paths."""
    total = 0.0
    for v in vals or []:
        n = _parse_numeric(v)
        if n is not None:
            total += n
    return total


def _industry_to_sector(industry: str) -> str:
    """
    Map free-text industry labels into canonical LU sectors used by tabs.
    """
    ind = (industry or "").strip()
    low = ind.lower()
    if not low:
        return SECTOR_OTHER

    if any(k in low for k in (
        "wholesale", "retail", "trading", "store", "shop", "sari", "merch",
    )):
        return SECTOR_WHOLESALE
    if any(k in low for k in (
        "agri", "agriculture", "farming", "farm", "fish", "fisher", "livestock",
        "hog", "poultry", "crop",
    )):
        return SECTOR_AGRICULTURE
    if any(k in low for k in (
        "transport", "truck", "tricycle", "jeep", "taxi", "hauling", "delivery",
        "logistics", "driver",
    )):
        return SECTOR_TRANSPORT
    if any(k in low for k in (
        "remittance", "padala", "money transfer", "ofw", "pawn", "exchange",
    )):
        return SECTOR_REMITTANCE
    if any(k in low for k in (
        "consumer", "salary", "employee", "professional", "service", "personal",
        "loan",
    )):
        return SECTOR_CONSUMER
    return SECTOR_OTHER


def _extract_industry_tags(raw_industry: str) -> list[str]:
    """
    Split combined industry text into distinct tags.
    Example: "Construction, Transportation" -> ["Construction", "Transportation"].
    """
    raw = (raw_industry or "").strip()
    if not raw:
        return []
    # Protect canonical labels first so embedded commas (e.g. Trading...)
    # are not split into incorrect pseudo-industries.
    protected = raw
    token_to_name: dict[str, str] = {}
    for idx, (name, pat) in enumerate(_CANONICAL_INDUSTRY_PATTERNS):
        if not pat.search(protected):
            continue
        token = f"__IND_{idx}__"
        protected = pat.sub(f" {token} ", protected)
        token_to_name[token] = name

    parts = [p.strip() for p in _INDUSTRY_SPLIT_RE.split(protected) if p and p.strip()]
    tags = []
    seen = set()
    for p in parts:
        normalized = token_to_name.get(p, p)
        key = normalized.lower()
        if key in seen:
            continue
        seen.add(key)
        tags.append(normalized)

    # If the row only contained canonical matches and delimiters, ensure they
    # are still returned even when token boundaries are unusual.
    if token_to_name and not tags:
        for token, canonical in token_to_name.items():
            if token in protected:
                key = canonical.lower()
                if key not in seen:
                    seen.add(key)
                    tags.append(canonical)
    return tags


def _make_compat_expenses(rec: dict) -> list[dict]:
    """
    Build minimal expense rows expected by charts/simulator/report tabs.
    New LU model is industry-based, so we synthesize a single row that
    carries risk metadata and amount context.
    """
    risk = rec.get("score_label", "LOW")
    industry = rec.get("industry") or "Unspecified Industry"
    base_amount = rec.get("total_source")
    has_values = base_amount is not None and base_amount > 0
    reason = (
        f"Industry '{industry}' is marked as {risk} risk."
        if risk in ("HIGH", "MEDIUM", "MODERATE")
        else f"Industry '{industry}' is not in the high/medium-risk lists."
    )
    return [{
        "name": "Total Source of Income",
        "total": float(base_amount or 0.0),
        "risk": risk,
        "reason": reason,
        "value_str": f"₱{base_amount:,.2f}" if has_values else "—",
        "has_values": has_values,
    }]


def _build_expense_item(name: str, total: float, risk: str, reason: str) -> dict:
    canonical_name = _normalize_expense_name(name)
    if not canonical_name:
        canonical_name = str(name or "").strip() or "Unspecified Expense"
    has_values = total is not None and total > 0
    return {
        "name": canonical_name,
        "total": float(total or 0.0),
        "risk": risk,
        "reason": reason,
        "value_str": f"₱{float(total):,.2f}" if has_values else "—",
        "has_values": has_values,
    }


def _parse_expense_detail_items(detail_text: str, risk: str, reason: str) -> list[dict]:
    """
    Parse itemized expenses from free-text detail cells.
    Accepts patterns like:
      "Food 3,000 Gas 1,200"
      "Fuel and Transport: 2500; Electricity: 1800"
    """
    text = (detail_text or "").strip()
    if not text:
        return []

    # Normalize separators found in OCR/exported text.
    normalized = re.sub(r"[\r\n\t]+", "; ", text)
    normalized = re.sub(r"\s{2,}", " ", normalized).strip()
    segments = [s.strip(" ;,|") for s in re.split(r"[;|]+", normalized) if s.strip(" ;,|")]
    if not segments:
        segments = [normalized]

    items = []
    # Patterns:
    #   "Food 3000"
    #   "Fuel: 1,200"
    #   "1,200 Food"
    #   "₱1,200 - Electricity"
    amount_pat = r"[\[\(]?\s*[₱$]?\s*\d[\d,]*(?:\.\d+)?\s*[\]\)]?"
    # Allow practical OCR/export variants in the name part:
    # - parenthetical notes: "Gasoline (Mr. Albert)"
    # - mixed tokens: "LPG (1100/3months)"
    name_pat = r"[A-Za-z][A-Za-z0-9\s/&\-\.\(\)'']{1,120}?"
    name_amount = re.compile(
        rf"(?P<name>{name_pat})\s*[:=\-]?\s*"
        rf"(?P<amt>{amount_pat})"
    )
    amount_name = re.compile(
        rf"(?P<amt>{amount_pat})\s*[:=\-]?\s*"
        rf"(?P<name>{name_pat})"
    )

    def _push_item(raw_name: str, raw_amt: str):
        name = re.sub(r"\s+", " ", (raw_name or "")).strip(" -:/.,[]()")
        if not name:
            return
        amt = _parse_numeric(raw_amt)
        if amt is None or amt <= 0:
            return
        # Keep labels readable and stable.
        clean_name = _normalize_expense_name(name)
        if not clean_name:
            return
        items.append(_build_expense_item(clean_name, amt, risk, reason))

    for seg in segments:
        # Most exported rows are "name ... [amount]" (amount at the end).
        # Parse this first to avoid picking numeric fragments inside the name
        # (e.g., "LPG (1100/3months) [366.67]").
        end_match = re.match(
            rf"^\s*(?P<name>.+?)\s*(?P<amt>{amount_pat})\s*$",
            seg,
        )
        if end_match:
            _push_item(end_match.group("name"), end_match.group("amt"))
            continue

        matched = False
        for m in name_amount.finditer(seg):
            _push_item(m.group("name"), m.group("amt"))
            matched = True
        if matched:
            continue
        for m in amount_name.finditer(seg):
            _push_item(m.group("name"), m.group("amt"))
            matched = True
        if matched:
            continue
        # Last pass over whole segment to catch tightly packed pairs.
        for m in name_amount.finditer(seg + " "):
            _push_item(m.group("name"), m.group("amt"))

    # Deduplicate by normalized name, summing totals
    merged = {}
    for it in items:
        key = it["name"].strip().lower()
        if key not in merged:
            merged[key] = it
        else:
            merged[key]["total"] += it["total"]
            merged[key]["value_str"] = f"₱{merged[key]['total']:,.2f}"
            merged[key]["has_values"] = merged[key]["total"] > 0
    return list(merged.values())


def _parse_expense_name_list(detail_text: str) -> list[str]:
    """
    Parse possible expense names from detail text even when amounts are missing.
    Example: "Food, Water, Electricity" -> ["Food", "Water", "Electricity"].
    """
    text = (detail_text or "").strip()
    if not text:
        return []

    normalized = re.sub(r"[\r\n\t]+", "; ", text)
    normalized = re.sub(r"\s{2,}", " ", normalized).strip()
    tokens = [t.strip(" ;,|:-") for t in re.split(r"[;,\n|]+", normalized) if t.strip(" ;,|:-")]

    names = []
    seen = set()
    for tok in tokens:
        clean = re.sub(r"[\[\(]?\s*[₱$]?\s*\d[\d,]*(?:\.\d+)?\s*[\]\)]?", "", tok)
        clean = re.sub(r"\b(total|expenses?|expense|amount)\b", "", clean, flags=re.I)
        clean = re.sub(r"\s{2,}", " ", clean).strip(" -:/.,[]()")
        if not clean:
            continue
        if not re.search(r"[A-Za-z]", clean):
            continue
        norm = clean.lower()
        if norm in seen:
            continue
        seen.add(norm)
        canonical = _normalize_expense_name(clean)
        if not canonical:
            continue
        names.append(canonical)
    return names


def _build_client_expenses(
    rec: dict,
    biz_detail: str,
    hh_detail: str,
    total_biz_exp: float | None,
    total_hh_exp: float | None,
) -> list[dict]:
    """
    Build the expense list consumed by simulator/charts/report.
    Prefers itemized entries parsed from detail columns; falls back to
    Business/Household totals when itemization is unavailable.
    """
    risk = rec.get("score_label", "LOW")
    industry = rec.get("industry") or "Unspecified Industry"
    reason = (
        f"Industry '{industry}' is marked as {risk} risk."
        if risk in ("HIGH", "MEDIUM", "MODERATE")
        else f"Industry '{industry}' is not in the high/medium-risk lists."
    )

    expenses = []
    expenses.extend(_parse_expense_detail_items(biz_detail, risk, reason))
    expenses.extend(_parse_expense_detail_items(hh_detail, risk, reason))

    # Some rows include name-only lines (e.g. "LPG") mixed with itemized lines.
    # Keep those canonical names as zero-amount rows so override matching still works.
    name_only = _parse_expense_name_list(biz_detail) + _parse_expense_name_list(hh_detail)
    existing = {str((e or {}).get("name") or "").strip().lower() for e in expenses}
    for nm in name_only:
        nk = str(nm or "").strip().lower()
        if not nk or nk in existing:
            continue
        expenses.append(_build_expense_item(nm, 0.0, risk, reason))
        existing.add(nk)

    # If details contain only names (no per-item amounts), synthesize item rows
    # using the provided total so simulator/report can still show itemized lines.
    if not expenses:
        biz_names = _parse_expense_name_list(biz_detail)
        hh_names = _parse_expense_name_list(hh_detail)
        if biz_names and total_biz_exp and total_biz_exp > 0:
            per = float(total_biz_exp) / max(1, len(biz_names))
            for nm in biz_names:
                expenses.append(_build_expense_item(nm, per, risk, reason))
        if hh_names and total_hh_exp and total_hh_exp > 0:
            per = float(total_hh_exp) / max(1, len(hh_names))
            for nm in hh_names:
                expenses.append(_build_expense_item(nm, per, risk, reason))

    # Fallback rows if no itemized values detected.
    if not expenses:
        if total_biz_exp and total_biz_exp > 0:
            expenses.append(_build_expense_item("Business Expenses", total_biz_exp, risk, reason))
        if total_hh_exp and total_hh_exp > 0:
            expenses.append(_build_expense_item("Household / Personal Expenses", total_hh_exp, risk, reason))

    # Final fallback keeps downstream tabs functional.
    if not expenses:
        expenses = _make_compat_expenses(rec)

    return expenses


def _apply_expense_overrides(expenses: list[dict], industry: str) -> str | None:
    """
    Apply per-expense risk overrides and return aggregate client override:
      - "HIGH" if any matched expense is HIGH
      - "MEDIUM" if no HIGH match but any MEDIUM match
      - None otherwise
    """
    if not expenses or not _EXPENSE_RISK_OVERRIDES:
        return None
    best = None
    for item in expenses:
        name = str((item or {}).get("name") or "").strip()
        for key in _expense_override_lookup_keys(name):
            lvl = _EXPENSE_RISK_OVERRIDES.get(key)
            if lvl not in ("HIGH", "MEDIUM", "MODERATE"):
                continue
            if lvl == "MODERATE":
                lvl = "MEDIUM"
            item["risk"] = lvl
            item["reason"] = f"Expense '{name}' is set as {lvl}."
            if lvl == "HIGH":
                return "HIGH"
            if best != "HIGH":
                best = "MEDIUM"
    return best


def _matched_high_expense_name(expenses: list[dict]) -> str:
    """Return first matched HIGH expense override name, else empty string."""
    if not expenses or not _EXPENSE_RISK_OVERRIDES:
        return ""
    for item in expenses:
        name = str((item or {}).get("name") or "").strip()
        if not name:
            continue
        if any(
            _EXPENSE_RISK_OVERRIDES.get(k) == "HIGH"
            for k in _expense_override_lookup_keys(name)
        ):
            return name
    return ""


def _matched_medium_expense_name(expenses: list[dict]) -> str:
    """Return first matched MEDIUM expense override name, else empty string."""
    if not expenses or not _EXPENSE_RISK_OVERRIDES:
        return ""
    for item in expenses:
        name = str((item or {}).get("name") or "").strip()
        if not name:
            continue
        if any(
            str(_EXPENSE_RISK_OVERRIDES.get(k) or "").upper() in ("MEDIUM", "MODERATE")
            for k in _expense_override_lookup_keys(name)
        ):
            return name
    return ""


def _compute_risk_reasoning(
    *,
    industry: str,
    product_name: str,
    product_override: str | None,
    expense_high_name: str,
    expense_medium_name: str = "",
    is_high_industry: bool,
    is_medium_industry: bool = False,
    product_matched_token: str = "",
) -> str:
    """Human-readable explanation for why final risk is HIGH/MEDIUM/LOW."""
    if product_override == "HIGH":
        detail = (product_matched_token or product_name).strip()
        return (
            f"This client is HIGH RISK because the loan product includes '{detail}' "
            "(HIGH product override)."
        )
    if product_override == "LOW":
        return (
            f"This client is LOW RISK because product override for '{product_name}' is set to LOW, "
            "which takes precedence over expense/industry."
        )
    if expense_high_name:
        return f"This client is HIGH RISK because they are using '{expense_high_name}' expenses."
    if expense_medium_name:
        return f"This client is MEDIUM RISK because they are using '{expense_medium_name}' expenses."
    if is_high_industry:
        return f"This client is HIGH RISK because they are under the industry '{industry}'."
    if is_medium_industry:
        return f"This client is MEDIUM RISK because they are under the industry '{industry}'."
    return "This client is LOW RISK because they do not fall in under any HIGH RISK category."
def _normalize_header_cell(val) -> str:
    """Normalize Excel header text (NBSP, unicode spaces, NFKC) for reliable matching."""
    if val is None:
        return ""
    s = str(val).strip()
    if not s:
        return ""
    s = unicodedata.normalize("NFKC", s)
    s = re.sub(r"[\u00a0\u2000-\u200b\u202f\u205f\u3000]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


# ─────────────────────────────────────────────────────────────────────
#  COLUMN PATTERNS (for LATEST_DUMMY.xlsx header)
# ─────────────────────────────────────────────────────────────────────
_COL_PATTERNS = {
    "client_id":         re.compile(r'client\s*id',                     re.I),
    "pn":                re.compile(r'^pn$',                            re.I),
    "applicant":         re.compile(r'applicant',                       re.I),
    "residence":         re.compile(r'residence\s*address',             re.I),
    "office":            re.compile(r'office\s*address',                re.I),
    "industry":          re.compile(r'industry\s*name',                 re.I),
    "spouse_info":       re.compile(r'spouse\s*info',                   re.I),
    # Asset columns: word boundaries + flexible separators (handles odd Excel exports).
    "personal_assets":    re.compile(r'\bpersonal[\s._-]*assets?\b', re.I),
    "business_assets":    re.compile(r'\bbusiness[\s._-]*assets?\b',    re.I),
    "business_inventory": re.compile(r'\bbusiness[\s._-]*inventory\b',  re.I),
    "source_income":     re.compile(r'source\s+of\s+income',            re.I),
    "total_source":      re.compile(r'total\s+source\s+of\s+income',    re.I),
    "biz_exp_detail":    re.compile(r'business\s+expenses?(?:\s+details?)?', re.I),
    "total_biz_exp":     re.compile(r'total\s+business\s+expenses?',    re.I),
    "hhld_exp_detail":   re.compile(r'(household\s*/?\s*personal|household)\s+expenses?(?:\s+details?)?', re.I),
    "total_hhld_exp":    re.compile(r'total\s+household',               re.I),
    "total_net_income":  re.compile(r'total\s+net\s+income',            re.I),
    "amort_history":     re.compile(r'total\s+amortization\s+history',  re.I),
    "current_amort":     re.compile(r'total\s+current\s+amortization',  re.I),
    "total_amortized_cost": re.compile(r'total\s+amortized\s+cost',     re.I),
    "principal_loan":    re.compile(r'principal\s+loan',                re.I),
    "maturity":          re.compile(r'^maturity$',                      re.I),
    "interest_rate":     re.compile(r'interest\s+rate',                 re.I),
    "branch":            re.compile(r'^branch$',                        re.I),
    "loan_class":        re.compile(r'loan\s+class',                    re.I),
    "product_name":      re.compile(r'product\s+name',                  re.I),
    "loan_date":         re.compile(r'loan\s+date',                     re.I),
    "term_unit":         re.compile(r'term\s+unit',                     re.I),
    "term":              re.compile(r'^term$',                          re.I),
    "security":          re.compile(r'^security$',                      re.I),
    "release_tag":       re.compile(r'release\s+tag',                   re.I),
    "loan_amount":       re.compile(r'^loan\s+amount$',                 re.I),
    "loan_status":       re.compile(r'loan\s+status',                   re.I),
    "ao_name":           re.compile(r'ao\s*name',                       re.I),
    "loan_balance":      re.compile(r'^loan\s+balance$',                re.I),
}

def _find_columns(header_row: tuple) -> dict[str, int]:
    cols = {}
    for i, cell_val in enumerate(header_row):
        if cell_val is None:
            continue
        s = _normalize_header_cell(cell_val)
        if not s:
            continue
        for field, pat in _COL_PATTERNS.items():
            if field not in cols and pat.search(s):
                cols[field] = i
    return cols

# ─────────────────────────────────────────────────────────────────────
#  ROW → CLIENT RECORD
# ─────────────────────────────────────────────────────────────────────
def _row_to_client(row: tuple, cols: dict[str, int]) -> dict | None:
    def get(field):
        idx = cols.get(field)
        return row[idx] if idx is not None and idx < len(row) else None

    applicant = str(get("applicant") or "").strip()
    # Skip summary/footer rows commonly present in portfolio exports.
    # These are not real clients but often appear under the Applicant column.
    au = applicant.upper()
    if (not applicant) or re.match(r"^(TOTAL|SUBTOTAL|GRAND\s+TOTAL|AVERAGE|AVG|MEAN)\b", au):
        return None

    client_id = str(get("client_id") or "").strip()
    pn        = str(get("pn") or "").strip()
    industry  = str(get("industry") or "").strip()
    src_text  = str(get("source_income") or "").strip()
    total_source = _parse_numeric(get("total_source"))
    total_net    = _parse_numeric(get("total_net_income"))
    total_biz_exp = _parse_numeric(get("total_biz_exp"))
    total_hh_exp  = _parse_numeric(get("total_hhld_exp"))
    loan_balance = _parse_numeric(get("loan_balance"))
    principal_loan = _parse_numeric(get("principal_loan"))
    current_amort  = _parse_numeric(get("current_amort"))
    amort_history  = _parse_numeric(get("amort_history"))
    personal_assets = _asset_text_field(get("personal_assets"))
    business_assets = _asset_text_field(get("business_assets"))
    business_inventory = _asset_text_field(get("business_inventory"))
    total_amortized_cost = _parse_numeric(get("total_amortized_cost"))
    interest_rate = _parse_numeric(get("interest_rate"))
    loan_amount = _parse_numeric(get("loan_amount"))

    residence = str(get("residence") or "").strip()
    office    = str(get("office") or "").strip()
    spouse_info = str(get("spouse_info") or "").strip()
    maturity = str(get("maturity") or "").strip()
    branch = str(get("branch") or "").strip()
    loan_class = str(get("loan_class") or "").strip()
    product_name = str(get("product_name") or "").strip()
    loan_date = str(get("loan_date") or "").strip()
    term_unit = str(get("term_unit") or "").strip()
    term = str(get("term") or "").strip()
    security = str(get("security") or "").strip()
    release_tag = str(get("release_tag") or "").strip()
    loan_status = str(get("loan_status") or "").strip()
    ao_name = str(get("ao_name") or "").strip()

    # Industry: HIGH/MEDIUM if any split tag is in configured risk sets.
    industry_tags = _extract_industry_tags(industry)
    high_risk_norm = {str(i).strip().lower() for i in _HIGH_RISK_INDUSTRIES}
    medium_risk_norm = {str(i).strip().lower() for i in _MEDIUM_RISK_INDUSTRIES}
    is_high_ind = any(tag.lower() in high_risk_norm for tag in industry_tags)
    is_medium_ind = (not is_high_ind) and any(tag.lower() in medium_risk_norm for tag in industry_tags)
    # Product Name override: any atomic product in the cell may match HIGH.
    pr, pr_matched = lookup_product_risk_override(product_name)
    risk_label = "HIGH" if is_high_ind else ("MEDIUM" if is_medium_ind else "LOW")
    score = 1.8 if risk_label == "HIGH" else (1.2 if risk_label == "MEDIUM" else 0.0)
    score_fg = "#E53E3E" if risk_label == "HIGH" else ("#D4A017" if risk_label == "MEDIUM" else "#2E7D32")
    score_bg = "#FFF5F5" if risk_label == "HIGH" else ("#FFFBF0" if risk_label == "MEDIUM" else "#F0FBE8")

    rec = {
        "client_id":      client_id,
        "pn":             pn,
        "client":         applicant,
        "residence":      residence,
        "office":         office,
        "industry":       industry,
        "spouse_info":    spouse_info,
        "personal_assets": personal_assets,
        "business_assets": business_assets,
        "business_inventory": business_inventory,
        "source_income":  src_text,
        "biz_exp_detail": str(get("biz_exp_detail") or "").strip(),
        "hhld_exp_detail": str(get("hhld_exp_detail") or "").strip(),
        "total_biz_exp":  total_biz_exp,
        "total_hhld_exp": total_hh_exp,
        "total_source":   total_source,
        "net_income":     total_net,
        "loan_balance":   loan_balance,
        "principal_loan": principal_loan,
        "current_amort":  current_amort,
        "amort_history":  amort_history,
        "total_amortized_cost": total_amortized_cost,
        "score":          score,
        "score_label":    risk_label,
        "score_fg":       score_fg,
        "score_bg":       score_bg,
        "maturity":       maturity,
        "interest_rate":  interest_rate,
        "branch":         branch,
        "loan_class":     loan_class,
        "product_name":   product_name,
        "loan_date":      loan_date,
        "term_unit":      term_unit,
        "term":           term,
        "security":       security,
        "release_tag":    release_tag,
        "loan_amount":    loan_amount,
        "loan_status":    loan_status,
        "ao_name":        ao_name,
        "sheet":          "",
        # Compatibility fields used by charts/simulator/report/loanbal tabs.
        "sector":         _industry_to_sector(industry),
        "industry_tags":  industry_tags,
    }
    rec["expenses"] = _build_client_expenses(
        rec,
        rec.get("biz_exp_detail", ""),
        rec.get("hhld_exp_detail", ""),
        total_biz_exp,
        total_hh_exp,
    )
    expense_override = _apply_expense_overrides(rec["expenses"], industry)
    expense_high_name = _matched_high_expense_name(rec["expenses"])
    expense_medium_name = _matched_medium_expense_name(rec["expenses"])

    # Precedence: Product > Expense > Industry.
    if pr in ("HIGH", "MEDIUM"):
        risk_label = pr
    elif expense_override == "HIGH":
        risk_label = "HIGH"
    elif expense_override == "MEDIUM":
        risk_label = "MEDIUM"

    score = 1.8 if risk_label == "HIGH" else (1.2 if risk_label == "MEDIUM" else 0.0)
    score_fg = "#E53E3E" if risk_label == "HIGH" else ("#D4A017" if risk_label == "MEDIUM" else "#2E7D32")
    score_bg = "#FFF5F5" if risk_label == "HIGH" else ("#FFFBF0" if risk_label == "MEDIUM" else "#F0FBE8")
    rec["score"] = score
    rec["score_label"] = risk_label
    rec["score_fg"] = score_fg
    rec["score_bg"] = score_bg
    rec["risk_reasoning"] = _compute_risk_reasoning(
        industry=industry,
        product_name=product_name,
        product_override=pr,
        expense_high_name=expense_high_name,
        expense_medium_name=expense_medium_name,
        is_high_industry=is_high_ind,
        is_medium_industry=is_medium_ind,
        product_matched_token=pr_matched,
    )
    return rec

# ─────────────────────────────────────────────────────────────────────
#  BACKWARD COMPATIBILITY STUBS (for lu_analysis_tab shim)
# ─────────────────────────────────────────────────────────────────────
def _is_summary_sheet(ws) -> bool:
    return True

def _analyse_sheet(ws, sheet_name: str) -> list[dict]:
    return []

def _read_income_from_summary(wb) -> dict[str, dict]:
    return {}

# ─────────────────────────────────────────────────────────────────────
#  PUBLIC ENTRY POINT
# ─────────────────────────────────────────────────────────────────────
def run_lu_analysis(filepath: str) -> dict:
    """
    Load an Excel workbook and return:
    {
        "general":    [client_dict, ...],
        "clients":    {applicant_name: client_dict},
        "sector_map": {sector_name: [client_dict, ...]},
        "income_map": {applicant_name: {"gross": float|None, "net": float|None}},
        "totals":     {"loan_balance": float, "total_source": float,
                       "total_net": float, "current_amort": float},
        "unique_industries": [str, ...],   # distinct industry tags
        "unique_product_names": [str, ...] # distinct atomic product labels (split on comma / etc.)
        "unique_expense_names": [str, ...] # distinct Business/Household expense names
    }
    """
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is not installed.\nRun:  pip install openpyxl")

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    try:
        general: list[dict] = []
        clients: dict[str, dict] = {}
        sector_map: dict[str, list] = {}
        income_map: dict[str, dict] = {}
        unique_industries: set[str] = set()
        unique_product_tokens: dict[str, str] = {}
        unique_expense_names: set[str] = set()

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            cols = {}
            data_start = 1
            for scan_idx, row in enumerate(rows[:_MAX_HEADER_SCAN_ROWS]):
                cols = _find_columns(row)
                if len(cols) >= 5:
                    data_start = scan_idx + 1
                    break

            if not cols:
                continue

            for row in rows[data_start:data_start + _MAX_DATA_ROWS]:
                rec = _row_to_client(row, cols)
                if rec is None:
                    continue
                rec["sheet"] = sheet_name
                name = rec["client"]
                clients[name] = rec
                general.append(rec)
                sector = rec.get("sector", SECTOR_OTHER)
                sector_map.setdefault(sector, []).append(rec)
                income_map[name] = {
                    "gross": rec["total_source"],
                    "net":   rec["net_income"],
                }
                for tag in rec.get("industry_tags", []):
                    unique_industries.add(tag)
                pn = (rec.get("product_name") or "").strip()
                if pn:
                    for tok in split_product_name_tokens(pn):
                        lk = tok.strip().lower()
                        if lk and lk not in unique_product_tokens:
                            unique_product_tokens[lk] = tok.strip()
                for exp in rec.get("expenses", []):
                    nm = str((exp or {}).get("name") or "").strip()
                    if nm:
                        unique_expense_names.add(nm)

    finally:
        wb.close()

    totals = {
        "loan_balance":  sum(r["loan_balance"] or 0 for r in general),
        "total_source":  sum(r["total_source"] or 0 for r in general),
        "total_net":     sum(r["net_income"] or 0 for r in general),
        "current_amort": sum(r["current_amort"] or 0 for r in general),
    }

    return {
        "general":    general,
        "clients":    clients,
        "sector_map": sector_map,
        "income_map": income_map,
        "totals":     totals,
        "unique_industries": sorted(unique_industries, key=lambda s: s.lower()),
        "unique_product_names": sorted(unique_product_tokens.values(), key=lambda s: s.lower()),
        "unique_expense_names": sorted(unique_expense_names, key=lambda s: s.lower()),
    }