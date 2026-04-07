"""
lu_analysis_row_format_patch.py  — v2 (corrected column-map bug)
"""

import re

try:
    import lu_analysis_tab as _orig
except ImportError:
    _orig = None

_EXTRA_SECTOR_KEYWORDS = {
    "tricycle driving":  "Transport",
    "tricycle boundary": "Transport",
    "tricycle service":  "Transport",
    "tricycle":          "Transport",
    "toda":              "Transport",
    "jeepney":           "Transport",
    "gasoline station":  "Wholesale Trade",
    "petron":            "Wholesale Trade",
    "shell":             "Wholesale Trade",
    "caltex":            "Wholesale Trade",
    "bigasan":           "Wholesale Trade",
    "rice retailing":    "Wholesale Trade",
    "sari sari":         "Wholesale Trade",
    "softdrinks":        "Wholesale Trade",
    "lending":           "Consumer Loan",
    "5-6":               "Consumer Loan",
    "paluwagan":         "Consumer Loan",
    "remittance":        "Remittance",
    "ofw":               "Remittance",
    "allotment":         "Remittance",
    "fishing":           "Fishing",
    "bangka":            "Fishing",
    "aquaculture":       "Fishing",
}

_ROW_EXPENSE_PATTERNS = [
    (re.compile(r'\b(fuel|diesel|gasoline|gas\b|petrol|bunker|aldo|ulo)\b',     re.I), "Fuel / Diesel"),
    (re.compile(r'\b(oil\b|lubricant|lube|change\s*oil)\b',                     re.I), "Oil & Lubricants"),
    (re.compile(r'\b(lpg|liquid\s*petroleum|cooking\s*gas|gas\s*stove)\b',      re.I), "LPG / Cooking Gas"),
    (re.compile(r'\b(electricity|electric(?:ity)?|light\b|power\s*bill|elec)\b',re.I), "Utilities (Power)"),
    (re.compile(r'\b(water\s*bill|drinking\s*water|water\b)\b',                 re.I), "Utilities (Water)"),
    (re.compile(r'\b(salary|salaries|wages|payroll|labor|labour|sahod|staff)\b',re.I), "Salaries / Wages"),
    (re.compile(r'\b(rent(?:al)?|lease|apartment\s*rental|branch\s*rent)\b',    re.I), "Rent / Lease"),
    (re.compile(r'\b(repair|maintenance|parts|spare\s*parts|change\s*tire'
                r'|vehicle\s*maint|car\s*maint|tricycle\s*maint)\b',            re.I), "Repairs & Maintenance"),
    (re.compile(r'\b(insurance|premium|comprehensive)\b',                        re.I), "Insurance Premium"),
    (re.compile(r'\b(feed|feeds|bait|fishmeal)\b',                               re.I), "Feed / Bait"),
    (re.compile(r'\b(ice\b|cold\s*storage|refriger)\b',                          re.I), "Ice / Cold Storage"),
    (re.compile(r'\b(toll|tollway)\b',                                           re.I), "Toll Fees"),
    (re.compile(r'\b(tax|taxes|bir|vat|business\s*permit|lto)\b',                re.I), "Taxes"),
    (re.compile(r'\b(communication|telco|phone|internet|load\b|cp\s*load|cable)\b', re.I), "Communication"),
    (re.compile(r'\b(supplies|office\s*supplies|packaging)\b',                   re.I), "Supplies"),
    (re.compile(r'\b(depreciation|amortization|amortisation|amortiz)\b',         re.I), "Depreciation"),
    (re.compile(r'\b(freight|shipping|cargo|delivery|logistics)\b',              re.I), "Freight / Logistics"),
    (re.compile(r'\b(interest|bank\s*charge|financing\s*charge|loan\s*pay)\b',   re.I), "Interest / Bank Charges"),
    (re.compile(r'\b(cost\s*of\s*sales|cos\b|cogs|rice\s*retail)\b',             re.I), "Cost of Sales"),
    (re.compile(r'\b(tricycle\s*expense|vehicle\s*expense|opex|operating\s*exp)\b', re.I), "Vehicle / Operating Expense"),
    (re.compile(r'\b(food|groceries|grocery|snacks)\b',                          re.I), "Food & Groceries"),
    (re.compile(r'\b(school|tuition|education)\b',                               re.I), "School / Education"),
    (re.compile(r'\b(miscellaneous|misc|other\s*expense|other\s*personal)\b',    re.I), "Miscellaneous"),
]

_ROW_EXTRA_RISK = {
    "LPG / Cooking Gas":           ("LOW",      "Household cooking gas; stable unless energy crisis."),
    "Food & Groceries":            ("LOW",      "Household food spending; monitor for inflation impact."),
    "School / Education":          ("LOW",      "Fixed education commitments; stable unless enrolment changes."),
    "Vehicle / Operating Expense": ("MODERATE", "Combined operating cost; may include fuel."),
    "Cost of Sales":               ("HIGH",     "Directly tied to revenue; margin compression risk if COS rises."),
    "Ice / Cold Storage":          ("MODERATE", "Power-sensitive; relevant for perishable-goods businesses."),
}

_BRACKET_RE     = re.compile(r'(.+?)\s*\[([0-9,\.]+)\]\s*(?=\n|$)', re.MULTILINE)
_APPLICANT_RE   = re.compile(r'applicant|client|borrower', re.I)


# ── Format detector ──────────────────────────────────────────────────────────

def _detect_file_format(ws) -> str:
    max_col = min(14, ws.max_column or 13)
    row1    = [str(ws.cell(1, c).value or "").strip() for c in range(1, max_col + 1)]
    joined  = " | ".join(row1).lower()
    if (_APPLICANT_RE.search(row1[0])
            and "source of income" in joined
            and ("business expense" in joined or "household" in joined)):
        r2a = ws.cell(2, 1).value
        if isinstance(r2a, str) and len(r2a.strip()) > 2:
            return "row-based"
    return "column-based"


# ── Column map builder (two-pass, most-specific first) ───────────────────────

def _build_col_map(ws, header_row: int) -> dict[str, int]:
    defaults = {
        "applicant": 1, "income_text": 4, "income_tot": 5,
        "biz_text": 6,  "biz_tot": 7,     "hh_text": 8,
        "hh_tot": 9,    "net_income": 10, "amort_hist": 11, "amort_curr": 12,
    }
    # Ordered most-specific → least-specific so "Total Source Of Income"
    # matches income_tot BEFORE the less-specific income_text pattern fires.
    rules = [
        ("income_tot",  re.compile(r'total\s+source',                    re.I)),
        ("biz_tot",     re.compile(r'total\s+business',                  re.I)),
        ("hh_tot",      re.compile(r'total\s+(household|personal)',       re.I)),
        ("net_income",  re.compile(r'net\s+income',                       re.I)),
        ("amort_hist",  re.compile(r'amortization\s+history|amort.*hist', re.I)),
        ("amort_curr",  re.compile(r'current\s+amortization|amort.*curr', re.I)),
        ("income_text", re.compile(r'source\s+of\s+income',               re.I)),
        ("biz_text",    re.compile(r'business\s+expense',                 re.I)),
        ("hh_text",     re.compile(r'household|personal\s+expense',       re.I)),
        ("applicant",   re.compile(r'applicant|client|borrower',          re.I)),
    ]
    result: dict[str, int] = dict(defaults)
    matched: set[str]      = set()
    for c in range(1, (ws.max_column or 12) + 1):
        hdr = str(ws.cell(header_row, c).value or "").strip()
        if not hdr:
            continue
        for key, pat in rules:
            if key not in matched and pat.search(hdr):
                result[key] = c
                matched.add(key)
                break
    return result


# ── Sector detector ──────────────────────────────────────────────────────────

def _detect_sectors_row(text: str) -> list[str]:
    seen  = {}
    lower = text.lower()
    kw    = dict(_EXTRA_SECTOR_KEYWORDS)
    if _orig:
        kw.update(_orig.SECTOR_KEYWORDS)
    for k in sorted(kw, key=len, reverse=True):
        if k in lower and kw[k] not in seen:
            seen[kw[k]] = True
    return list(seen.keys())


# ── Bracket extractor ────────────────────────────────────────────────────────

def _extract_bracketed_items(text: str) -> list[tuple[str, float]]:
    out = []
    for m in _BRACKET_RE.finditer(text or ""):
        name = re.sub(r'\s+', ' ', m.group(1)).strip()
        if name:
            try:
                out.append((name, float(m.group(2).replace(",", ""))))
            except ValueError:
                pass
    return out


# ── Expense normaliser ───────────────────────────────────────────────────────

def _normalise_expense(raw: str) -> str:
    for pat, cat in _ROW_EXPENSE_PATTERNS:
        if pat.search(raw):
            return cat
    return raw


# ── Risk lookup ──────────────────────────────────────────────────────────────

def _lookup_risk(sector: str, expense: str) -> tuple[str, str]:
    default = ("LOW", "No specific sector-expense sensitivity rule defined; treat as low risk.")
    if _orig:
        t = _orig.SECTOR_EXPENSE_RISK.get(sector, {})
        if expense in t:
            return t[expense]
    return _ROW_EXTRA_RISK.get(expense, default)


# ── Risk score calculator ────────────────────────────────────────────────────

def _compute_risk_score(expenses):
    if _orig:
        return _orig._compute_risk_score(expenses)
    w = {"HIGH": 3, "MODERATE": 2, "LOW": 1}
    if not expenses:
        return (0.0, "N/A", "#9AAACE", "#F5F7FA")
    score = sum(w.get(e["risk"], 1) for e in expenses) / len(expenses)
    for thr, lbl, fg, bg in [
        (2.5,"CRITICAL","#B71C1C","#FFEBEE"),(1.8,"HIGH","#E53E3E","#FFF5F5"),
        (1.2,"MODERATE","#D4A017","#FFFBF0"),(0.0,"LOW","#2E7D32","#F0FBE8"),
    ]:
        if score >= thr:
            return (score, lbl, fg, bg)
    return (score, "LOW", "#2E7D32", "#F0FBE8")


# ── Core analyser ────────────────────────────────────────────────────────────

def _analyse_row_based_sheet(ws, sheet_name: str) -> list[dict]:
    _RISK_ORDER = {"HIGH": 0, "MODERATE": 1, "LOW": 2}

    # Locate header row
    header_row = 1
    for r in range(1, min(11, (ws.max_row or 1) + 1)):
        if _APPLICANT_RE.search(str(ws.cell(r, 1).value or "")):
            header_row = r
            break

    col = _build_col_map(ws, header_row)

    def _num(row_idx, key):
        raw = ws.cell(row_idx, col[key]).value
        if raw is None:
            return None
        if isinstance(raw, (int, float)):
            return float(raw)
        try:
            return float(str(raw).replace(",", "").replace("₱", "").strip())
        except ValueError:
            return None

    results = []
    for row_idx in range(header_row + 1, (ws.max_row or 0) + 1):
        applicant_raw = ws.cell(row_idx, col["applicant"]).value
        if not applicant_raw:
            continue
        applicant_str = str(applicant_raw).strip()
        if re.match(r'^\s*total', applicant_str, re.I):
            continue

        income_text = str(ws.cell(row_idx, col["income_text"]).value or "")
        biz_text    = str(ws.cell(row_idx, col["biz_text"]).value    or "")
        hh_text     = str(ws.cell(row_idx, col["hh_text"]).value     or "")

        total_income = _num(row_idx, "income_tot")
        net_income   = _num(row_idx, "net_income")
        amort_curr   = _num(row_idx, "amort_curr")

        sectors = _detect_sectors_row(f"{income_text}\n{biz_text}\n{hh_text}")
        if not sectors:
            sectors = ["General"]

        raw_items = _extract_bracketed_items(biz_text) + _extract_bracketed_items(hh_text)

        cat_totals: dict[str, float]     = {}
        cat_raw:    dict[str, list[str]] = {}
        for name, val in raw_items:
            cat = _normalise_expense(name)
            cat_totals[cat] = cat_totals.get(cat, 0.0) + val
            cat_raw.setdefault(cat, []).append(name)

        for sector in sectors:
            exp_out = []
            for cat, total in cat_totals.items():
                risk, reason = _lookup_risk(sector, cat)
                rns  = cat_raw.get(cat, [])
                if len(rns) == 1:
                    vstr = f"₱{total:,.2f}  ({rns[0]})"
                else:
                    shown = ", ".join(rns[:3])
                    more  = f"…+{len(rns)-3}" if len(rns) > 3 else ""
                    vstr  = f"₱{total:,.2f}  ({len(rns)} items: {shown}{more})"
                exp_out.append({"name": cat, "risk": risk, "reason": reason,
                                "value_str": vstr, "has_values": True, "total": total})

            if _orig and sector in _orig.SECTOR_EXPENSE_RISK:
                for en, (risk, reason) in _orig.SECTOR_EXPENSE_RISK[sector].items():
                    if risk in ("HIGH","MODERATE") and not any(e["name"]==en for e in exp_out):
                        exp_out.append({"name": en, "risk": risk, "reason": reason,
                                        "value_str": "(not found in file — advisory)",
                                        "has_values": False, "total": 0.0})

            exp_out.sort(key=lambda x: _RISK_ORDER.get(x["risk"], 9))
            score, slabel, sfg, sbg = _compute_risk_score(exp_out)

            parts = []
            if total_income: parts.append(f"Income: ₱{total_income:,.2f}")
            if net_income:   parts.append(f"Net: ₱{net_income:,.2f}")
            if amort_curr:   parts.append(f"Amort: ₱{amort_curr:,.2f}")

            results.append({
                "sector": sector, "sheet": sheet_name, "client": applicant_str,
                "expenses": exp_out, "score": score, "score_label": slabel,
                "score_fg": sfg, "score_bg": sbg,
                "_row_summary": "  |  ".join(parts),
                "_total_income": total_income, "_net_income": net_income, "_amort_curr": amort_curr,
            })
    return results


# ── Patched entry point ──────────────────────────────────────────────────────

def run_lu_analysis_patched(filepath: str) -> dict:
    try:
        import openpyxl as _xl
    except ImportError:
        raise RuntimeError("openpyxl is not installed.\nRun:  pip install openpyxl")
    wb      = _xl.load_workbook(filepath, data_only=True)
    clients: dict[str, list[dict]] = {}
    general: list[dict]            = []
    seen:    set[str]              = set()
    for sname in wb.sheetnames:
        ws  = wb[sname]
        fmt = _detect_file_format(ws)
        if fmt == "row-based":
            rows = _analyse_row_based_sheet(ws, sname)
        else:
            rows = _orig._analyse_sheet(ws, sname) if _orig else []
        for r in rows:
            ck = r["client"]
            clients.setdefault(ck, []).append(r)
            dk = f"{r['sector']}|{ck}"
            if dk not in seen:
                seen.add(dk)
                general.append(r)
    return {"general": general, "clients": clients}


# ── attach() ─────────────────────────────────────────────────────────────────

def attach(cls):
    """Call AFTER lu_analysis_tab.attach(cls)."""
    if _orig:
        _orig.run_lu_analysis = run_lu_analysis_patched
    cls._detect_file_format      = staticmethod(_detect_file_format)
    cls._analyse_row_based_sheet = staticmethod(_analyse_row_based_sheet)
    cls._run_lu_analysis_patched = staticmethod(run_lu_analysis_patched)


# ── Smoke test ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else None
    if not path:
        print("Usage: python lu_analysis_row_format_patch.py <file.xlsx>")
        sys.exit(1)

    print(f"\n{'='*72}")
    print(f"  LU Row-Format Patch — Smoke Test")
    print(f"  File: {path}")
    print(f"{'='*72}\n")

    data = run_lu_analysis_patched(path)
    print(f"  Clients : {len(data['clients'])}")
    print(f"  Entries : {len(data['general'])}\n")

    for client, rows in data["clients"].items():
        for r in rows:
            h = sum(1 for e in r["expenses"] if e["risk"]=="HIGH")
            m = sum(1 for e in r["expenses"] if e["risk"]=="MODERATE")
            l = sum(1 for e in r["expenses"] if e["risk"]=="LOW")
            print(f"  [{r['score_label']:<8} {r['score']:.2f}]  {client:<35}  "
                  f"Sector: {r['sector']:<18}  H={h} M={m} L={l}")
            if r.get("_row_summary"):
                print(f"    └─ {r['_row_summary']}")
            for e in r["expenses"][:3]:
                ico = "🔴" if e["risk"]=="HIGH" else "🟡" if e["risk"]=="MODERATE" else "🟢"
                print(f"       {ico} {e['name']:<30}  {e['value_str'][:55]}")
        print()