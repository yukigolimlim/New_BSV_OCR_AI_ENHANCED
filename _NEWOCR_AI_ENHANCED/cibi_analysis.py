"""
cibi_analysis.py
────────────────────────────────────────────────────────────────────────────────
CIBI Analysis Engine — Banco San Vicente (BSV)

Reads a populated CIBI Excel workbook and produces a full credit analysis
based on:
  • Cash-Flow worksheet  (income / expenses / net disposable)
  • CIBI worksheet       (borrower profile, loan purpose, co-maker, etc.)

Returns a formatted text report that the app can render in the Analysis tab.
────────────────────────────────────────────────────────────────────────────────
"""

from __future__ import annotations

import re
import json
import os
from pathlib import Path
from typing  import Any

# ── Optional imports (fail gracefully) ───────────────────────────────────────
try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# ══════════════════════════════════════════════════════════════════════════════
#  SHEET-NAME CANDIDATES
# ══════════════════════════════════════════════════════════════════════════════

_CASHFLOW_NAMES = {
    "cash flow", "cashflow", "cash_flow", "cf",
    "income", "income & expense", "income and expense",
    "sources of income", "financial", "budget",
}

_CIBI_NAMES = {
    "cibi", "ci", "credit investigation", "borrower", "applicant",
    "loan application", "borrower info", "client info",
}


def _match_sheet(wb, candidates: set[str]):
    """Return the first worksheet whose name (lower-stripped) is in *candidates*."""
    for name in wb.sheetnames:
        if name.strip().lower() in candidates:
            return wb[name]
    # fuzzy: sheet name *contains* any keyword
    for name in wb.sheetnames:
        nl = name.strip().lower()
        for kw in candidates:
            if kw in nl:
                return wb[name]
    return None


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL READING HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _cell_val(ws, row: int, col: int) -> Any:
    """Return a cell value or '' if out-of-range / None."""
    try:
        v = ws.cell(row=row, column=col).value
        return v if v is not None else ""
    except Exception:
        return ""


def _to_float(v) -> float | None:
    """Try to coerce a cell value to float."""
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", "").replace("₱", "").replace(" ", "").strip())
    except Exception:
        return None


def _sheet_to_dict(ws) -> dict[str, Any]:
    """
    Walk every row of *ws* and collect { label : value } pairs.
    A 'label' is any non-empty string cell; the 'value' is the next
    non-empty cell to its right (up to 4 columns).
    """
    data: dict[str, Any] = {}
    if ws is None:
        return data
    for row in ws.iter_rows():
        label = None
        for cell in row:
            v = cell.value
            if v is None or str(v).strip() == "":
                continue
            sv = str(v).strip()
            if label is None:
                label = sv
            else:
                # store under label
                existing = data.get(label)
                if existing is None or existing == "":
                    data[label] = v
                label = sv          # next non-empty becomes next label
    return data


def _sheet_to_rows(ws) -> list[list]:
    """Return all rows as a list-of-lists (strings), trimming trailing empties."""
    rows = []
    if ws is None:
        return rows
    for row in ws.iter_rows(values_only=True):
        cleaned = [str(c).strip() if c is not None else "" for c in row]
        # drop fully-empty rows
        if any(cleaned):
            rows.append(cleaned)
    return rows


# ══════════════════════════════════════════════════════════════════════════════
#  CASH-FLOW PARSER
# ══════════════════════════════════════════════════════════════════════════════

_CF_INCOME_KW  = ("income", "salary", "gross", "revenue", "receipts",
                  "remittance", "business income", "net income", "take home",
                  "allowance", "pension", "rental", "interest income")

_CF_EXPENSE_KW = ("expense", "expenditure", "household", "utilities",
                  "food", "education", "transportation", "medical",
                  "insurance", "rent", "amortization", "obligation",
                  "loan payment", "installment", "credit card")

_CF_TOTAL_KW   = ("total income", "total receipts", "total earnings",
                  "gross income", "net income", "total expenses",
                  "total disbursement", "net disposable", "net cash",
                  "free cash", "surplus", "deficiency")


def parse_cashflow(ws) -> dict:
    """
    Extract structured cash-flow data from a worksheet.

    Returns:
    {
        "income_items":   {label: amount},
        "expense_items":  {label: amount},
        "total_income":   float | None,
        "total_expenses": float | None,
        "net_disposable": float | None,
        "dsr":            float | None,   # debt-service ratio
        "loan_payment":   float | None,
        "raw_rows":       [[str, ...]],   # for AI fallback
    }
    """
    result = {
        "income_items":   {},
        "expense_items":  {},
        "total_income":   None,
        "total_expenses": None,
        "net_disposable": None,
        "dsr":            None,
        "loan_payment":   None,
        "raw_rows":       [],
    }

    if ws is None:
        return result

    rows = _sheet_to_rows(ws)
    result["raw_rows"] = rows

    def _first_numeric(row: list) -> float | None:
        for cell in row[1:]:
            v = _to_float(cell)
            if v is not None and v != 0:
                return v
        return None

    for row in rows:
        if not row:
            continue
        label = row[0].lower()
        amount = _first_numeric(row)

        # ── totals (check first, more specific) ──────────────────────────
        if any(kw in label for kw in ("total income", "total receipts", "gross income",
                                       "total earnings")):
            if amount is not None:
                result["total_income"] = amount
            continue

        if any(kw in label for kw in ("total expense", "total disbursement",
                                       "total outflow", "total expenditure")):
            if amount is not None:
                result["total_expenses"] = amount
            continue

        if any(kw in label for kw in ("net disposable", "net cash", "free cash",
                                       "surplus", "net income after", "disposable")):
            if amount is not None:
                result["net_disposable"] = amount
            continue

        if any(kw in label for kw in ("dsr", "debt service ratio", "debt-service")):
            if amount is not None:
                result["dsr"] = amount
            continue

        if any(kw in label for kw in ("monthly amortization", "loan payment",
                                       "monthly payment", "proposed amortization")):
            if amount is not None:
                result["loan_payment"] = amount
            continue

        # ── income line items ─────────────────────────────────────────────
        if any(kw in label for kw in _CF_INCOME_KW):
            if amount is not None:
                result["income_items"][row[0].strip()] = amount
            continue

        # ── expense line items ────────────────────────────────────────────
        if any(kw in label for kw in _CF_EXPENSE_KW):
            if amount is not None:
                result["expense_items"][row[0].strip()] = amount
            continue

    # ── derive totals if not explicitly found ─────────────────────────────
    if result["total_income"] is None and result["income_items"]:
        result["total_income"] = sum(result["income_items"].values())

    if result["total_expenses"] is None and result["expense_items"]:
        result["total_expenses"] = sum(result["expense_items"].values())

    if result["net_disposable"] is None:
        ti = result["total_income"]
        te = result["total_expenses"]
        lp = result["loan_payment"] or 0
        if ti is not None and te is not None:
            result["net_disposable"] = ti - te - lp

    # ── DSR ───────────────────────────────────────────────────────────────
    if result["dsr"] is None:
        lp = result["loan_payment"]
        ti = result["total_income"]
        if lp and ti and ti > 0:
            result["dsr"] = lp / ti

    return result


# ══════════════════════════════════════════════════════════════════════════════
#  CIBI SHEET PARSER
# ══════════════════════════════════════════════════════════════════════════════

def parse_cibi_sheet(ws) -> dict:
    """
    Extract borrower / loan profile from the CIBI worksheet.

    Returns a flat dict of key fields.
    """
    data = _sheet_to_dict(ws)
    rows = _sheet_to_rows(ws)

    def _find(keys: list[str]) -> str:
        for k in keys:
            kl = k.lower()
            for dk, dv in data.items():
                if kl in dk.lower():
                    v = str(dv).strip()
                    if v and v.lower() not in ("none", "n/a", ""):
                        return v
        return ""

    profile = {
        # Borrower
        "applicant_name":  _find(["full name", "borrower name", "client name", "name of borrower"]),
        "date_of_birth":   _find(["date of birth", "dob", "birthdate"]),
        "civil_status":    _find(["civil status", "marital"]),
        "address":         _find(["address", "residence"]),
        "employer":        _find(["employer", "company", "place of work", "business name"]),
        "position":        _find(["position", "occupation", "job title"]),
        "years_employed":  _find(["years in service", "length of service", "tenure", "years employed"]),
        # Loan
        "loan_purpose":    _find(["purpose", "loan purpose", "use of proceeds"]),
        "loan_amount":     _find(["loan amount", "amount applied", "amount requested", "principal"]),
        "loan_term":       _find(["term", "loan term", "repayment period", "tenor"]),
        "payment_mode":    _find(["payment mode", "mode of payment", "frequency"]),
        "collateral":      _find(["collateral", "security", "mortgage"]),
        # Co-maker / guarantor
        "co_maker":        _find(["co-maker", "co maker", "guarantor", "co-borrower"]),
        "co_maker_income": _find(["co-maker income", "guarantor income"]),
        # Credit history
        "credit_history":  _find(["credit history", "past loan", "existing loan", "outstanding"]),
        "cic_rating":      _find(["cic", "credit score", "credit rating", "risk tier"]),
        "bank_ci":         _find(["bank ci", "bank certification", "bank account"]),
    }

    # pull raw rows for AI fallback
    profile["_raw"] = "\n".join(
        "  |  ".join(c for c in row if c)
        for row in rows[:80]
    )

    return profile


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL FILE READER — MAIN ENTRY
# ══════════════════════════════════════════════════════════════════════════════

def read_cibi_excel(file_path: str) -> dict:
    """
    Open a populated CIBI Excel file and extract structured data.

    Returns:
    {
        "cashflow": {...},      # from parse_cashflow()
        "cibi":     {...},      # from parse_cibi_sheet()
        "sheets":   [str],      # all sheet names
        "raw_all":  str,        # concatenated readable dump (for AI context)
    }
    """
    if not _HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is not installed.\n\nRun:  pip install openpyxl"
        )

    wb = openpyxl.load_workbook(file_path, data_only=True)

    cf_ws   = _match_sheet(wb, _CASHFLOW_NAMES)
    cibi_ws = _match_sheet(wb, _CIBI_NAMES)

    # fallback: use all sheets if no specific match
    if cf_ws is None and cibi_ws is None and wb.sheetnames:
        cf_ws   = wb[wb.sheetnames[0]]
        cibi_ws = wb[wb.sheetnames[-1]] if len(wb.sheetnames) > 1 else cf_ws

    cashflow_data = parse_cashflow(cf_ws)
    cibi_data     = parse_cibi_sheet(cibi_ws)

    # ── Build a combined raw text for AI context ──────────────────────────
    raw_parts = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        raw_parts.append(f"\n{'═'*60}\n  SHEET: {sname}\n{'═'*60}")
        for row in ws.iter_rows(values_only=True):
            row_str = "  |  ".join(str(c).strip() if c is not None else "" for c in row)
            if any(str(c).strip() for c in row if c is not None):
                raw_parts.append(row_str)

    return {
        "cashflow": cashflow_data,
        "cibi":     cibi_data,
        "sheets":   wb.sheetnames,
        "raw_all":  "\n".join(raw_parts),
    }


# ══════════════════════════════════════════════════════════════════════════════
#  GEMINI AI ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════

_BSV_SYSTEM_PROMPT = """
You are a senior credit officer at Banco San Vicente (BSV), a Philippine rural bank.
You are performing a full CIBI (Credit Investigation and Background Investigation) analysis
based on data extracted from a populated CIBI Excel workbook.

Your analysis must follow BSV's standard credit evaluation framework:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

ANALYSIS STRUCTURE (follow exactly):

1. BORROWER PROFILE SUMMARY
2. LOAN REQUEST DETAILS
3. CASH-FLOW ANALYSIS
   A) Income Assessment
   B) Expense Assessment
   C) Net Disposable Income & DSR
   D) Repayment Capacity Verdict
4. CIBI / CREDIT BACKGROUND
   A) Employment & Stability
   B) Credit History & CIC Rating
   C) Bank Record (Bank CI)
   D) Co-maker / Guarantor Assessment
5. COLLATERAL ASSESSMENT
6. RISK FLAGS
7. CREDIT SCORING SUMMARY
8. FINAL RECOMMENDATION
   • Verdict: APPROVE / CONDITIONALLY APPROVE / DECLINE
   • Conditions (if any)
   • Recommended BSV Loan Product

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

BSV DSR POLICY:
  • DSR ≤ 35%  → Low risk, favorable
  • DSR 36–50% → Moderate, needs justification
  • DSR > 50%  → High risk, usually decline unless strong mitigants

BSV NET DISPOSABLE INCOME (NDI) RULE:
  • NDI must be ≥ 20% of gross monthly income after all obligations
  • NDI < 0 → automatic decline unless co-maker income covers gap

Use ₱ symbol for all peso amounts. Be specific, cite actual figures.
Flag any missing data with [DATA NOT FOUND].
"""


def run_cibi_analysis(
    excel_path: str,
    api_key:    str,
    model:      str = "gemini-2.5-flash",
) -> str:
    """
    Full pipeline:
      1. Read the CIBI Excel file
      2. Parse cash-flow and CIBI sheets
      3. Send structured data to Gemini for analysis
      4. Return formatted report text

    Parameters
    ----------
    excel_path : path to the populated CIBI Excel file
    api_key    : Gemini API key
    model      : Gemini model string

    Returns
    -------
    str — formatted analysis report
    """

    # ── Step 1: Read Excel ────────────────────────────────────────────────
    try:
        excel_data = read_cibi_excel(excel_path)
    except ImportError as e:
        return f"⚠ Missing dependency:\n\n{e}"
    except Exception as e:
        return f"⚠ Failed to read Excel file:\n\n{type(e).__name__}: {e}"

    cf   = excel_data["cashflow"]
    cibi = excel_data["cibi"]

    # ── Step 2: Format structured context for AI ──────────────────────────
    def _fmt_peso(v) -> str:
        if v is None: return "[N/A]"
        try:   return f"₱{float(v):,.2f}"
        except: return str(v)

    def _fmt_pct(v) -> str:
        if v is None: return "[N/A]"
        try:
            f = float(v)
            return f"{f*100:.1f}%" if f <= 1 else f"{f:.1f}%"
        except: return str(v)

    income_lines = "\n".join(
        f"    {k}: {_fmt_peso(v)}"
        for k, v in cf["income_items"].items()
    ) or "    [No line items found]"

    expense_lines = "\n".join(
        f"    {k}: {_fmt_peso(v)}"
        for k, v in cf["expense_items"].items()
    ) or "    [No line items found]"

    structured_context = f"""
╔══════════════════════════════════════════════════════════════╗
║           EXTRACTED CIBI EXCEL DATA — BSV ANALYSIS          ║
╚══════════════════════════════════════════════════════════════╝

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  A. BORROWER / LOAN PROFILE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  Applicant Name     : {cibi.get('applicant_name') or '[Not found]'}
  Date of Birth      : {cibi.get('date_of_birth')  or '[Not found]'}
  Civil Status       : {cibi.get('civil_status')   or '[Not found]'}
  Address            : {cibi.get('address')         or '[Not found]'}
  Employer / Business: {cibi.get('employer')        or '[Not found]'}
  Position           : {cibi.get('position')        or '[Not found]'}
  Years in Service   : {cibi.get('years_employed')  or '[Not found]'}

  Loan Amount        : {cibi.get('loan_amount')     or '[Not found]'}
  Loan Purpose       : {cibi.get('loan_purpose')    or '[Not found]'}
  Loan Term          : {cibi.get('loan_term')        or '[Not found]'}
  Payment Mode       : {cibi.get('payment_mode')    or '[Not found]'}
  Collateral         : {cibi.get('collateral')      or '[Not found]'}

  Co-maker / Guarantor : {cibi.get('co_maker')       or '[Not found]'}
  Co-maker Income      : {cibi.get('co_maker_income') or '[Not found]'}

  Credit History     : {cibi.get('credit_history')  or '[Not found]'}
  CIC Rating         : {cibi.get('cic_rating')       or '[Not found]'}
  Bank CI Result     : {cibi.get('bank_ci')          or '[Not found]'}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  B. CASH-FLOW STATEMENT
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  INCOME LINE ITEMS:
{income_lines}

  TOTAL MONTHLY INCOME   : {_fmt_peso(cf['total_income'])}

  EXPENSE LINE ITEMS:
{expense_lines}

  TOTAL MONTHLY EXPENSES : {_fmt_peso(cf['total_expenses'])}

  PROPOSED LOAN PAYMENT  : {_fmt_peso(cf['loan_payment'])}
  NET DISPOSABLE INCOME  : {_fmt_peso(cf['net_disposable'])}
  DEBT-SERVICE RATIO     : {_fmt_pct(cf['dsr'])}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  C. RAW EXCEL CONTENT (CIBI SHEET)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{cibi.get('_raw', '[No raw data]')[:3000]}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  D. RAW EXCEL CONTENT (ALL SHEETS)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{excel_data['raw_all'][:6000]}
"""

    # ── Step 3: Call Gemini ───────────────────────────────────────────────
    if not api_key or api_key == "YOUR_GEMINI_API_KEY_HERE":
        # Return a formatted local-only analysis without AI
        return _local_analysis(cf, cibi)

    try:
        from google import genai as _genai
        from google.genai import types as _gtypes

        client = _genai.Client(api_key=api_key)

        prompt = (
            "Using the CIBI Excel data below, produce a complete BSV credit analysis "
            "following the structure in your system instructions.\n\n"
            + structured_context
        )

        resp = client.models.generate_content(
            model    = model,
            contents = [_gtypes.Content(role="user", parts=[_gtypes.Part(text=prompt)])],
            config   = _gtypes.GenerateContentConfig(
                max_output_tokens  = 4096,
                temperature        = 0.15,
                system_instruction = _BSV_SYSTEM_PROMPT,
            ),
        )

        # extract text
        try:
            text = resp.text or ""
        except Exception:
            text = "".join(
                p.text for p in resp.candidates[0].content.parts
                if hasattr(p, "text") and p.text
            )

        if not text.strip():
            return _local_analysis(cf, cibi)

        return text

    except Exception as e:
        err = str(e).lower()

        # quota fallback
        if any(kw in err for kw in ("quota", "rate", "429", "resource_exhausted")):
            try:
                from google import genai as _genai
                from google.genai import types as _gtypes
                client2 = _genai.Client(api_key=api_key)
                resp2 = client2.models.generate_content(
                    model    = "gemini-2.0-flash",
                    contents = [_gtypes.Content(role="user", parts=[_gtypes.Part(text=prompt)])],
                    config   = _gtypes.GenerateContentConfig(
                        max_output_tokens  = 4096,
                        system_instruction = _BSV_SYSTEM_PROMPT,
                    ),
                )
                try:    return resp2.text or _local_analysis(cf, cibi)
                except: return _local_analysis(cf, cibi)
            except Exception:
                pass

        return f"⚠ Gemini error during CIBI analysis:\n{type(e).__name__}: {e}\n\n" + _local_analysis(cf, cibi)


# ══════════════════════════════════════════════════════════════════════════════
#  LOCAL FALLBACK ANALYSIS (no AI)
# ══════════════════════════════════════════════════════════════════════════════

def _local_analysis(cf: dict, cibi: dict) -> str:
    """
    Generate a rule-based analysis report without AI,
    using only the parsed Excel data.
    """

    def _p(v) -> str:
        if v is None: return "N/A"
        try:   return f"₱{float(v):,.2f}"
        except: return str(v)

    def _pct(v) -> str:
        if v is None: return "N/A"
        try:
            f = float(v)
            return f"{f*100:.1f}%" if f <= 1 else f"{f:.1f}%"
        except: return str(v)

    ti  = cf.get("total_income")
    te  = cf.get("total_expenses")
    nd  = cf.get("net_disposable")
    dsr = cf.get("dsr")
    lp  = cf.get("loan_payment")

    # ── DSR verdict ───────────────────────────────────────────────────────
    dsr_verdict = "N/A"
    dsr_flag    = ""
    if dsr is not None:
        dsr_f = float(dsr) if dsr <= 1 else float(dsr) / 100
        if dsr_f <= 0.35:
            dsr_verdict = "✅ LOW RISK (≤ 35%)"
        elif dsr_f <= 0.50:
            dsr_verdict = "⚠ MODERATE (36–50%) — Needs justification"
            dsr_flag    = "• DSR in moderate range"
        else:
            dsr_verdict = "❌ HIGH RISK (> 50%) — Likely decline"
            dsr_flag    = "• DSR exceeds 50% — repayment stress likely"

    # ── NDI verdict ───────────────────────────────────────────────────────
    ndi_verdict = "N/A"
    ndi_flag    = ""
    if nd is not None and ti is not None and ti > 0:
        ndi_ratio = float(nd) / float(ti)
        if ndi_ratio >= 0.20:
            ndi_verdict = f"✅ Sufficient ({ndi_ratio*100:.1f}% of income)"
        else:
            ndi_verdict = f"❌ Insufficient ({ndi_ratio*100:.1f}% of income — BSV requires ≥ 20%)"
            ndi_flag    = "• Net Disposable Income below BSV minimum threshold (20% of income)"

    # ── Overall verdict ───────────────────────────────────────────────────
    flags = [f for f in (dsr_flag, ndi_flag) if f]
    if not flags:
        overall = "APPROVE"
        overall_note = "All financial indicators are within BSV policy thresholds."
    elif len(flags) == 1 and "moderate" in dsr_verdict.lower():
        overall = "CONDITIONALLY APPROVE"
        overall_note = "Subject to additional justification and/or co-maker requirement."
    else:
        overall = "DECLINE"
        overall_note = "Financial ratios do not meet BSV minimum requirements."

    lines = [
        "═" * 66,
        "  BSV CIBI ANALYSIS REPORT",
        "  Source: Populated CIBI Excel Workbook",
        "═" * 66,
        "",
        "1. BORROWER PROFILE SUMMARY",
        "─" * 44,
        f"  Applicant      : {cibi.get('applicant_name') or 'N/A'}",
        f"  Date of Birth  : {cibi.get('date_of_birth')  or 'N/A'}",
        f"  Civil Status   : {cibi.get('civil_status')   or 'N/A'}",
        f"  Employer       : {cibi.get('employer')        or 'N/A'}",
        f"  Position       : {cibi.get('position')        or 'N/A'}",
        f"  Years in Service: {cibi.get('years_employed') or 'N/A'}",
        "",
        "2. LOAN REQUEST DETAILS",
        "─" * 44,
        f"  Loan Amount    : {cibi.get('loan_amount')  or 'N/A'}",
        f"  Purpose        : {cibi.get('loan_purpose') or 'N/A'}",
        f"  Term           : {cibi.get('loan_term')     or 'N/A'}",
        f"  Payment Mode   : {cibi.get('payment_mode') or 'N/A'}",
        f"  Collateral     : {cibi.get('collateral')   or 'N/A'}",
        "",
        "3. CASH-FLOW ANALYSIS",
        "─" * 44,
        "  A) Income",
    ]

    for k, v in cf.get("income_items", {}).items():
        lines.append(f"     • {k}: {_p(v)}")
    lines += [
        f"     Total Monthly Income   : {_p(ti)}",
        "",
        "  B) Expenses",
    ]
    for k, v in cf.get("expense_items", {}).items():
        lines.append(f"     • {k}: {_p(v)}")
    lines += [
        f"     Total Monthly Expenses : {_p(te)}",
        f"     Proposed Loan Payment  : {_p(lp)}",
        "",
        "  C) Net Disposable Income & DSR",
        f"     Net Disposable Income  : {_p(nd)}",
        f"     NDI Assessment         : {ndi_verdict}",
        f"     Debt-Service Ratio     : {_pct(dsr)}",
        f"     DSR Assessment         : {dsr_verdict}",
        "",
        "4. CIBI / CREDIT BACKGROUND",
        "─" * 44,
        f"  Credit History : {cibi.get('credit_history') or 'N/A'}",
        f"  CIC Rating     : {cibi.get('cic_rating')     or 'N/A'}",
        f"  Bank CI        : {cibi.get('bank_ci')         or 'N/A'}",
        f"  Co-maker       : {cibi.get('co_maker')        or 'N/A'}",
        f"  Co-maker Income: {cibi.get('co_maker_income') or 'N/A'}",
        "",
        "5. COLLATERAL ASSESSMENT",
        "─" * 44,
        f"  {cibi.get('collateral') or 'No collateral data found in CIBI file.'}",
        "",
        "6. RISK FLAGS",
        "─" * 44,
    ]

    if flags:
        for f in flags:
            lines.append(f"  {f}")
    else:
        lines.append("  ✅ No major risk flags identified.")

    lines += [
        "",
        "7. CREDIT SCORING SUMMARY",
        "─" * 44,
        f"  Financial indicators are {'within' if not flags else 'outside'} BSV policy thresholds.",
        f"  DSR: {_pct(dsr)}   |   NDI: {_p(nd)}   |   Income: {_p(ti)}",
        "",
        "═" * 66,
        "8. FINAL RECOMMENDATION",
        "═" * 66,
        f"  VERDICT: {overall}",
        f"  {overall_note}",
        "",
        "  Note: This is a rule-based assessment. For full AI analysis,",
        "  ensure a valid Gemini API key is configured.",
        "═" * 66,
    ]

    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════════════════
#  CONVENIENCE WRAPPER (called from app.py)
# ══════════════════════════════════════════════════════════════════════════════

def run_cibi_analysis_from_text(extracted_text: str, api_key: str) -> str:
    """
    Legacy wrapper: if the caller only has raw extracted text (not an Excel path),
    pass it directly to Gemini for CIBI-style analysis.
    """
    if not api_key or api_key == "YOUR_GEMINI_API_KEY_HERE":
        return (
            "⚠ No Gemini API key configured.\n\n"
            "Set GEMINI_API_KEY in your .env file or in app.py."
        )

    prompt = (
        "Using the document text below, produce a complete BSV CIBI credit analysis. "
        "Focus especially on the cash-flow section and the CIBI (credit investigation) data.\n\n"
        "DOCUMENT TEXT:\n"
        "─" * 56 + "\n"
        + extracted_text[:12_000]
    )

    try:
        from google import genai as _genai
        from google.genai import types as _gtypes

        client = _genai.Client(api_key=api_key)
        resp = client.models.generate_content(
            model    = "gemini-2.5-flash",
            contents = [_gtypes.Content(role="user", parts=[_gtypes.Part(text=prompt)])],
            config   = _gtypes.GenerateContentConfig(
                max_output_tokens  = 4096,
                temperature        = 0.15,
                system_instruction = _BSV_SYSTEM_PROMPT,
            ),
        )
        try:    return resp.text or "⚠ Empty response from Gemini."
        except: return "⚠ Could not read Gemini response."

    except Exception as e:
        return f"⚠ CIBI analysis error:\n{type(e).__name__}: {e}"