"""
analysis.py — DocExtract Pro
==============================
All LLM-based analysis logic.  No UI code lives here.

Public API
----------
  check_loan_applicability(extracted_text) -> (bool, str)
      Fast pre-check: is the document a loan-related file?

  run_credit_analysis(extracted_text) -> str
      Full 10-section credit assessment report.

LLM Backend
-----------
  Primary  : Gemini 2.5 Flash  (gemini-2.5-flash)
  Fallback : Gemini 2.0 Flash  (gemini-2.0-flash)
  API key  : GEMINI_API_KEY from .env or environment
"""
import os
import re
import json
import logging
from pathlib import Path

logger = logging.getLogger(__name__)

# ── Load .env (same folder as this file) ─────────────────────────────────────
_env_path = Path(__file__).parent / ".env"
if _env_path.exists():
    with open(_env_path) as _f:
        for _line in _f:
            _line = _line.strip()
            if _line and not _line.startswith("#") and "=" in _line:
                _k, _v = _line.split("=", 1)
                os.environ.setdefault(_k.strip(), _v.strip())

from utils import SCRIPT_DIR

# ── Model constants ───────────────────────────────────────────────────────────
_PRIMARY_MODEL  = "gemini-2.5-flash"
_FALLBACK_MODEL = "gemini-2.0-flash"


# ══════════════════════════════════════════════════════════════════════════════
#  GEMINI API HELPER
# ══════════════════════════════════════════════════════════════════════════════

def _gemini_call_with_fallback(
    prompt:      str,
    max_tokens:  int   = 3000,
    temperature: float = 0.2,
    label:       str   = "analysis",
) -> tuple[str, str]:
    """
    Call Gemini 2.5 Flash, falling back to Gemini 2.0 Flash on quota/rate errors.

    Returns (response_text, model_used_str)
    where model_used_str is one of: "primary", "fallback", "trimmed"

    Raises RuntimeError if both models fail.
    """
    api_key = os.environ.get("GEMINI_API_KEY", "")
    if not api_key or api_key == "YOUR_GEMINI_API_KEY_HERE":
        raise RuntimeError(
            "GEMINI_API_KEY not found.\n\n"
            "Create a .env file in the same folder as analysis.py:\n"
            "    GEMINI_API_KEY=your_key_here\n\n"
            "Get a free key at: https://aistudio.google.com/app/apikey"
        )

    try:
        from google import genai as _genai
        from google.genai import types as _gtypes
    except ImportError:
        raise RuntimeError(
            "google-genai package not installed.\n"
            "Run:  pip install google-genai"
        )

    client = _genai.Client(api_key=api_key)

    def _safe_text(resp) -> str:
        try:
            return resp.text or ""
        except Exception:
            try:
                return "".join(
                    p.text for p in resp.candidates[0].content.parts
                    if hasattr(p, "text") and p.text
                )
            except Exception:
                return ""

    def _call(model: str, text: str) -> str:
        resp = client.models.generate_content(
            model    = model,
            contents = [_gtypes.Content(
                role  = "user",
                parts = [_gtypes.Part(text=text)]
            )],
            config = _gtypes.GenerateContentConfig(
                max_output_tokens = max_tokens,
                temperature       = temperature,
            ),
        )
        return _safe_text(resp)

    def _trim_prompt(text: str) -> str:
        """Trim to first + last quarter if too large."""
        if len(text) <= 40_000:
            return text
        quarter = len(text) // 4
        return (
            text[:quarter * 2]
            + "\n\n[… context trimmed to fit token limit …]\n\n"
            + text[-quarter:]
        )

    # ── 1. Try primary model ──────────────────────────────────────────────
    try:
        result = _call(_PRIMARY_MODEL, prompt)
        return result, "primary"

    except Exception as e:
        err = str(e).lower()

        # Quota / rate limit → try fallback
        if any(kw in err for kw in ("quota", "rate", "429", "resource_exhausted")):
            logger.warning(
                "[%s] Gemini 2.5 Flash quota hit — switching to %s",
                label, _FALLBACK_MODEL
            )
            try:
                result = _call(_FALLBACK_MODEL, prompt)
                return result, "fallback"
            except Exception as e2:
                raise RuntimeError(
                    f"Both Gemini models quota-limited during {label}.\n\n"
                    f"• {_PRIMARY_MODEL} — quota exceeded\n"
                    f"• {_FALLBACK_MODEL} — also rate-limited\n\n"
                    f"Details: {e2}"
                ) from e2

        # Context too large → trim and retry
        if any(kw in err for kw in ("token", "too long", "context", "size", "413")):
            logger.warning("[%s] Context too large — trimming and retrying", label)
            trimmed = _trim_prompt(prompt)
            try:
                result = _call(_PRIMARY_MODEL, trimmed)
                return result, "trimmed"
            except Exception:
                try:
                    result = _call(_FALLBACK_MODEL, trimmed)
                    return result, "trimmed"
                except Exception as e3:
                    raise RuntimeError(
                        f"Gemini call failed after trimming during {label}: {e3}"
                    ) from e3

        raise RuntimeError(
            f"Gemini API call failed during {label}: {type(e).__name__}: {e}"
        ) from e


# ══════════════════════════════════════════════════════════════════════════════
#  LOAN CATALOG LOADER
# ══════════════════════════════════════════════════════════════════════════════

def _load_loan_catalog() -> str:
    """
    Reads banco_san_vicente_loans.json and returns a compact
    human-readable summary suitable for injection into an LLM prompt.
    """
    json_path = SCRIPT_DIR / "banco_san_vicente_loans.json"
    if not json_path.exists():
        return "[Loan product catalog file not found — skip eligibility section]"

    try:
        with open(json_path, "r", encoding="utf-8") as f:
            catalog = json.load(f)

        lines = ["=== BANCO SAN VICENTE LOAN PRODUCT CATALOG ===\n"]

        for cat in catalog.get("categories", []):
            lines.append(f"\n[{cat['category_name'].upper()}]")

            for p in cat.get("products", []):
                lines.append(f"\n  Product : {p['product_name']}")
                if p.get("purpose"):
                    lines.append(f"  Purpose : {p['purpose']}")
                if p.get("qualification"):
                    lines.append(f"  Qualify : {p['qualification']}")

                la = p.get("loan_amount")
                if la and isinstance(la, dict):
                    amt_parts = []
                    if la.get("minimum"):
                        amt_parts.append(f"min ₱{la['minimum']:,}")
                    if la.get("maximum") and isinstance(la["maximum"], (int, float)):
                        amt_parts.append(f"max ₱{la['maximum']:,}")
                    elif la.get("maximum"):
                        amt_parts.append(f"max: {la['maximum']}")
                    if la.get("first_loan_maximum"):
                        amt_parts.append(f"first-loan max ₱{la['first_loan_maximum']:,}")
                    if la.get("first_cycle"):
                        fc = la["first_cycle"]
                        amt_parts.append(
                            f"first cycle ₱{fc.get('min', 0):,}–₱{fc.get('max', 0):,}"
                        )
                    if la.get("calculation"):
                        amt_parts.append(f"calc: {la['calculation']}")
                    for fk in ("daet", "labo"):
                        fv = la.get("franchise_value", {})
                        if isinstance(fv, dict) and fv.get(fk):
                            amt_parts.append(f"franchise ({fk}) ₱{fv[fk]:,}")
                    if amt_parts:
                        lines.append(f"  Amount  : {', '.join(amt_parts)}")

                for r in (p.get("interest_rates") or []):
                    if isinstance(r, dict) and r.get("term") and r.get("interest_rate"):
                        lines.append(
                            f"  Rate    : {r['term']} @ {r['interest_rate']}"
                            f"  Service Charge {r.get('service_charge', 'N/A')}"
                        )

                col = p.get("collateral")
                if col:
                    if isinstance(col, list):
                        lines.append(f"  Collat  : {', '.join(col)}")
                    elif isinstance(col, str):
                        lines.append(f"  Collat  : {col}")
                    elif isinstance(col, dict):
                        for tier_k, tier_v in col.items():
                            if isinstance(tier_v, dict) and tier_v.get("range"):
                                rng   = tier_v["range"]
                                items = tier_v.get("items", [])
                                lines.append(
                                    f"  Collat ({tier_k} "
                                    f"₱{rng.get('min', 0):,}–₱{rng.get('max', 0):,}): "
                                    f"{', '.join(items)}"
                                )
                            elif isinstance(tier_v, list):
                                lines.append(
                                    f"  Collat ({tier_k}): {', '.join(tier_v)}"
                                )

                if p.get("other_notes"):
                    lines.append(f"  Notes   : {p['other_notes']}")

        return "\n".join(lines)

    except Exception as e:
        logger.exception("Failed to load loan catalog")
        return f"[Could not load loan catalog: {e}]"


_loan_catalog_text: str | None = None


def build_loan_catalog_text() -> str:
    """Return the loan catalog text, loading and caching it on first call."""
    global _loan_catalog_text
    if _loan_catalog_text is None:
        _loan_catalog_text = _load_loan_catalog()
    return _loan_catalog_text


# ══════════════════════════════════════════════════════════════════════════════
#  DYNAMIC XLSX DISCOVERY HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _find_scoring_rubric_file(directory: Path) -> Path | None:
    """
    Locate the credit scoring xlsx in `directory` by CONTENT, not by filename.
    A valid rubric must contain at least one sheet with "template" and one
    with "criteria" or "parameter" in its name.
    """
    try:
        import openpyxl
    except ImportError:
        return None

    for xlsx in sorted(directory.glob("*.xlsx")):
        try:
            wb = openpyxl.load_workbook(str(xlsx), read_only=True, data_only=True)
            names_lower = [s.lower() for s in wb.sheetnames]
            wb.close()
            has_template = any("template" in n for n in names_lower)
            has_criteria = any(
                "criteria" in n or "parameter" in n for n in names_lower
            )
            if has_template and has_criteria:
                return xlsx
        except Exception:
            continue
    return None


def _find_sheet(wb, *keywords):
    """Return first sheet whose name contains ALL keywords (case-insensitive)."""
    for name in wb.sheetnames:
        n = name.lower()
        if all(k.lower() in n for k in keywords):
            return wb[name]
    return None


def _find_header_row(ws, *header_names):
    """
    Scan worksheet for the first row containing ALL given header names.
    Returns {header_name: col_index} or empty dict.
    """
    targets = {h.strip().lower(): h for h in header_names}
    for row in ws.iter_rows(values_only=True):
        found = {}
        for col_i, cell in enumerate(row):
            if isinstance(cell, str):
                cl = cell.strip().lower()
                if cl in targets:
                    found[targets[cl]] = col_i
        if len(found) == len(targets):
            return found
    return {}


def _is_component_header(row: tuple) -> str | None:
    """
    Detect component header rows (CAPACITY / CHARACTER / CAPITAL / CONDITION).
    Returns component name or None.
    """
    _NON_COMPONENTS = {"SCORE", "RISK SCORE MATRIX", "TOTAL"}
    col_b = row[1] if len(row) > 1 else None
    col_c = row[2] if len(row) > 2 else None

    if not isinstance(col_b, str):
        return None

    v = col_b.strip()
    if (v
            and v == v.upper()
            and 2 <= len(v) <= 20
            and col_c is None
            and v not in _NON_COMPONENTS):
        return v
    return None


# ══════════════════════════════════════════════════════════════════════════════
#  CREDIT SCORING FRAMEWORK PARSERS
# ══════════════════════════════════════════════════════════════════════════════

def _parse_criteria_sheet(ws) -> dict:
    """Parse a criteria/parameters sheet (Business or Salary) dynamically."""
    result         = {}
    current_comp   = None
    considerations = []
    col_total      = None
    col_consids    = None

    for row in ws.iter_rows(values_only=True):
        if not any(c is not None for c in row):
            continue

        comp_name = _is_component_header(row)
        if comp_name:
            current_comp   = comp_name
            considerations = []
            col_total      = None
            col_consids    = None
            result[current_comp] = {"considerations": [], "levels": {}}
            continue

        if current_comp is None:
            continue

        col_b = row[1] if len(row) > 1 else None
        col_c = row[2] if len(row) > 2 else None

        if isinstance(col_c, str) and col_c.strip() == "Total Score":
            for ci, cell in enumerate(row):
                if isinstance(cell, str) and cell.strip() == "Total Score":
                    col_total = ci
                    break
            col_consids = col_total + 2 if col_total is not None else 4
            considerations = []
            for cell in row[col_consids:]:
                if cell is None or str(cell).strip() == "":
                    break
                considerations.append(str(cell).strip())
            result[current_comp]["considerations"] = considerations
            continue

        if (isinstance(col_b, (int, float))
                and 1 <= int(col_b) <= 5
                and col_consids is not None):
            level       = int(col_b)
            total_score = (int(col_c) if isinstance(col_c, (int, float)) else None)
            criteria_vals = [
                str(c).strip() if c is not None else ""
                for c in row[col_consids: col_consids + len(considerations)]
            ]
            result[current_comp]["levels"][level] = {
                "score": total_score,
                **dict(zip(considerations, criteria_vals)),
            }

    return result


def _parse_risk_matrix(ws) -> list[tuple]:
    """Extract risk score matrix rows from a criteria sheet."""
    matrix    = []
    in_matrix = False

    for row in ws.iter_rows(values_only=True):
        if not any(c is not None for c in row):
            continue
        col_b = row[1] if len(row) > 1 else None
        if isinstance(col_b, str) and "RISK SCORE MATRIX" in col_b.upper():
            in_matrix = True
            continue
        if in_matrix:
            if (isinstance(col_b, (int, float))
                    and len(row) > 4
                    and isinstance(row[2], (int, float))):
                score_min = int(col_b)
                score_max = int(row[2])
                ecl       = row[3]
                stage     = row[4]
                ecl_str   = (f"{float(ecl) * 100:.4g}%"
                             if isinstance(ecl, (int, float)) else str(ecl))
                matrix.append((score_min, score_max, ecl_str, stage))

    return matrix


def _parse_weights(ws) -> dict:
    """Extract component risk weights dynamically."""
    header_cols = _find_header_row(ws, "Risk Weight", "Component")

    if not header_cols:
        weights = {}
        for row in ws.iter_rows(values_only=True):
            if not any(c is not None for c in row):
                continue
            for ci in range(len(row) - 1):
                weight_val = row[ci]
                comp_val   = row[ci + 1]
                if (isinstance(weight_val, float)
                        and 0 < weight_val < 1
                        and isinstance(comp_val, str)
                        and comp_val.strip().upper() == comp_val.strip()):
                    weights[comp_val.strip().upper()] = (
                        f"{int(round(weight_val * 100))}%"
                    )
        return weights

    col_w = header_cols["Risk Weight"]
    col_c = header_cols["Component"]
    weights = {}
    for row in ws.iter_rows(values_only=True):
        if not any(c is not None for c in row):
            continue
        weight_val = row[col_w] if len(row) > col_w else None
        comp_val   = row[col_c] if len(row) > col_c else None
        if (isinstance(weight_val, float)
                and 0 < weight_val < 1
                and isinstance(comp_val, str)
                and comp_val.strip()):
            comp = comp_val.strip().upper()
            weights[comp] = f"{int(round(weight_val * 100))}%"

    return weights


def _parse_salary_condition(ws) -> dict:
    """Parse a 'Condition if Salary'-style sheet into a single component dict."""
    considerations = []
    level_data: dict[int, dict] = {i: {} for i in range(1, 6)}
    current_consid = None
    value_col      = None

    for row in ws.iter_rows(values_only=True):
        if not any(c is not None for c in row):
            continue

        col_a = row[0] if len(row) > 0 else None

        if isinstance(col_a, str):
            stripped = col_a.strip()
            if re.match(r'^\d+[.)]\s*\S', stripped):
                current_consid = re.sub(r'^\d+[.)]\s*', '', stripped).strip()
                considerations.append(current_consid)
                value_col = None
                continue

        if isinstance(col_a, str) and col_a.strip().lower() == "score":
            value_col = None
            for ci in range(1, len(row)):
                if row[ci] is not None:
                    value_col = ci
                    break
            continue

        if (current_consid
                and isinstance(col_a, (int, float))
                and 1 <= int(col_a) <= 5):
            lvl = int(col_a)
            vc  = value_col if value_col is not None else 1
            val = row[vc] if len(row) > vc else None
            if val is not None:
                level_data[lvl][current_consid] = str(val).strip()

    score_map = {1: 850, 2: 800, 3: 750, 4: 650, 5: 300}
    levels = {
        lvl: {"score": score_map[lvl], **level_data[lvl]}
        for lvl in range(1, 6)
    }
    return {"considerations": considerations, "levels": levels}


# ══════════════════════════════════════════════════════════════════════════════
#  FORMATTING HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _format_component(comp_name: str, weight: str, data: dict) -> str:
    """Render one component block as a compact string for prompt injection."""
    considerations = data.get("considerations", [])
    levels         = data.get("levels", {})
    lines          = [f"{comp_name}  {weight}  | {' | '.join(considerations)}"]
    MAX_VAL = 60
    for consid in considerations:
        row_parts = []
        for lvl in range(1, 6):
            val = levels.get(lvl, {}).get(consid, "")
            if val:
                val = val[:MAX_VAL].rstrip() + ("…" if len(val) > MAX_VAL else "")
                row_parts.append(f"{lvl}={val}")
        if row_parts:
            lines.append(f"  {consid}: {' | '.join(row_parts)}")
    return "\n".join(lines)


def _format_risk_matrix(matrix: list[tuple]) -> str:
    """Render risk matrix as a compact single-line-per-band string."""
    parts = []
    for score_min, score_max, ecl, stage in matrix:
        parts.append(f"{score_min}-{score_max}={ecl} Stg{stage}")
    return " | ".join(parts)


# ══════════════════════════════════════════════════════════════════════════════
#  CREDIT SCORING FRAMEWORK LOADER
# ══════════════════════════════════════════════════════════════════════════════

def _load_credit_scoring_framework() -> str:
    """
    Locate, parse, and format the BSV credit scoring xlsx as a compact
    reference string for LLM prompt injection.
    """
    xlsx_path = _find_scoring_rubric_file(SCRIPT_DIR)

    if xlsx_path is None:
        logger.warning(
            "No credit scoring rubric xlsx found in %s — "
            "scoring framework will use fallback placeholder.", SCRIPT_DIR
        )
        return (
            "[Credit scoring framework file not found.\n"
            f" Searched in: {SCRIPT_DIR}\n"
            " Place an xlsx file containing 'Template' and 'Criteria' sheets\n"
            " in the same folder as analysis.py and restart the app.]"
        )

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

        ws_tmpl_biz  = _find_sheet(wb, "template", "business")
        ws_crit_biz  = (_find_sheet(wb, "parameter", "business")
                        or _find_sheet(wb, "criteria", "business"))
        ws_tmpl_sal  = _find_sheet(wb, "template", "salary")
        ws_crit_sal  = _find_sheet(wb, "criteria", "salary")
        ws_cond_sal  = _find_sheet(wb, "condition", "salary")

        biz_weights = _parse_weights(ws_tmpl_biz) if ws_tmpl_biz else {}
        biz_data    = _parse_criteria_sheet(ws_crit_biz) if ws_crit_biz else {}
        biz_matrix  = _parse_risk_matrix(ws_crit_biz) if ws_crit_biz else []

        sal_weights = _parse_weights(ws_tmpl_sal) if ws_tmpl_sal else {}
        sal_data    = _parse_criteria_sheet(ws_crit_sal) if ws_crit_sal else {}
        sal_matrix  = _parse_risk_matrix(ws_crit_sal) if ws_crit_sal else []

        if ws_cond_sal is not None:
            sal_condition = _parse_salary_condition(ws_cond_sal)
            for key in list(sal_data.keys()):
                if key.strip().upper() == "CONDITION":
                    sal_data[key] = sal_condition
                    break
            else:
                sal_data["CONDITION"] = sal_condition

        lines = [
            f"=== BSV CREDIT SCORING FRAMEWORK ===\n"
            f"(Source: {xlsx_path.name})\n"
        ]
        lines.append(
            "SCORING SCALE: Level→Points: 1=850 2=800 3=750 4=650 5=300\n"
            "Score/consideration = Component points ÷ no. of considerations\n"
            "Average score = Component score × risk weight\n"
            "TOTAL = sum of averages (max 850)\n"
        )

        lines.append("── BUSINESS TEMPLATE (applicant has business income) ──")
        for comp, data in biz_data.items():
            weight = biz_weights.get(comp, biz_weights.get(comp.strip(), "?%"))
            if data:
                lines.append(_format_component(comp.strip(), weight, data))

        lines.append("\n── SALARY TEMPLATE (employed only, no business income) ──")
        for comp, data in sal_data.items():
            weight = sal_weights.get(comp, sal_weights.get(comp.strip(), "?%"))
            if data:
                lines.append(_format_component(comp.strip(), weight, data))

        matrix = biz_matrix or sal_matrix
        if matrix:
            lines.append("\n── RISK SCORE MATRIX ──")
            lines.append(_format_risk_matrix(matrix))

        lines.append(
            "\n── KEY RATIO FORMULAS ──\n"
            "DSR        = Net Income before amort ÷ Total Monthly Amortization\n"
            "Liquidity  = Current Liabilities (≤12 mo) ÷ (Cash + Bank Deposits + Inventory)\n"
            "D/E        = Total Liabilities ÷ Net Worth\n"
            "ROE        = Net Income ÷ Net Worth\n"
            "Asset Turn = Total Sales ÷ Total Assets"
        )

        return "\n".join(lines)

    except Exception as e:
        logger.exception("Failed to parse credit scoring framework from xlsx")
        return f"[Credit scoring framework could not be loaded: {e}]"


_credit_scoring_framework_text: str | None = None


def build_credit_scoring_framework_text() -> str:
    """Return the credit scoring framework text, loading and caching it on first call."""
    global _credit_scoring_framework_text
    if _credit_scoring_framework_text is None:
        _credit_scoring_framework_text = _load_credit_scoring_framework()
    return _credit_scoring_framework_text


# ══════════════════════════════════════════════════════════════════════════════
#  PROMPT TEMPLATES
# ══════════════════════════════════════════════════════════════════════════════

_APPLICABILITY_PROMPT = """\
You are a document classifier for Banco San Vicente, a rural bank in the Philippines.

Your only job is to determine whether the document below is a loan application or \
loan-related document that can be meaningfully analyzed for credit assessment.

LOAN-RELATED documents include:
- Loan application forms (any type)
- Credit investigation / background investigation (CI/BI) forms
- Cashflow statements or income statements submitted with a loan
- Approval forms, credit approval sheets
- Financial statements (balance sheet, income statement) of a borrower
- Collateral appraisal reports
- Any form that is clearly part of a loan application package

NOT LOAN-RELATED documents include:
- General business documents unrelated to lending (e.g. sales reports, inventory lists)
- Medical records, school records, utility bills alone
- Random images, receipts, or personal notes with no financial/loan context
- Blank or near-blank documents
- Documents from a completely unrelated industry with no borrower/loan data

Respond ONLY in this exact JSON format (no explanation, no markdown):
{{"applicable": true, "reason": ""}}
or
{{"applicable": false, "reason": "brief specific reason why this document cannot be used for loan analysis"}}

--- DOCUMENT TEXT (first 2000 chars) ---
{sample}
--- END ---
"""

_ANALYSIS_PROMPT = """\
You are a senior credit analyst at Banco San Vicente, a rural bank in Camarines Norte, Philippines.
Analyze the following loan application data and produce a structured credit assessment report.

{scoring_note}

=== INSTRUCTIONS ===
Produce ALL of the following numbered sections IN ORDER.
Use the BSV Credit Scoring Framework and Loan Product Catalog provided below as your authoritative references.

1. LOAN PRODUCT RECOMMENDATIONS
   Using the BANCO SAN VICENTE LOAN PRODUCT CATALOG provided below,
   match the applicant's profile (occupation, income, collateral, credit standing)
   to the most suitable BSV loan products.
   Do NOT include any product ID or internal code.

   A) RECOMMENDED PRODUCTS (top 1–3 best matches)
      • Product name
      • Why it fits this applicant's profile
      • Estimated loanable amount in Philippine Pesos
      • Applicable interest rate and term (write "Service Charge" in full, never abbreviate as "SC")
      • Documents still needed before submission

   B) UPGRADE PATH  (if the applicant does not fully qualify for their requested loan)
      • Specific steps they must take (build credit history, add collateral, etc.)
      • Which product to apply for now as a stepping stone
      • Realistic timeline to reach eligibility for the desired loan

   C) IF DECLINED ENTIRELY — guidance before re-applying
      • Minimum requirements currently missing
      • Concrete actions with approximate timeframes
      • Which BSV product to target first upon re-application

2. APPLICANT SUMMARY
   - Full name, age, occupation, civil status, address

3. FINANCIAL POSITION
   - Assets, liabilities, net worth
   - Debt-to-Asset Ratio  = Total Liabilities / Total Assets
   - Loan-to-Asset Ratio  = Loan Amount / Total Assets
   - Loan-to-Networth Ratio = Loan Amount / Net Worth

4. INCOME ANALYSIS
   - List all income sources with amounts
   - Monthly net income after expenses
   - If cashflow sheet is blank, note it clearly

5. EXISTING OBLIGATIONS
   - List all existing loans, monthly amortization burden, remaining balance

6. COLLATERAL ASSESSMENT
   - Real properties or assets offered, estimated values

7. CREDIT SCORING (BSV Framework)
   Using the BSV Credit Scoring Framework below, determine whether to apply the
   BUSINESS TEMPLATE or SALARY TEMPLATE based on the applicant's profile,
   then score each component and consideration:

   a) Identify which template applies and why.

   b) For each component (CAPACITY, CHARACTER, CAPITAL if business, CONDITION):
      - State the risk weight for this template
      - For each consideration, assign a score level (1–5) with justification
        based on the criteria in the framework
      - Calculate: Score per Consideration = Component Total Score / No. of Considerations
      - Calculate: Average Score = Component Score × Risk Weight

   c) Calculate the TOTAL SCORE (sum of all average scores, max 850)

   d) Map the total score to the Risk Score Matrix:
      - Risk Level (Low / Low-Moderate / Moderate / Moderate-High / High)
      - ECL% (Expected Credit Loss percentage)
      - Stage (1, 2, or 3)

   e) Show all key financial ratios computed:
      - DSR, Liquidity ratio, D/E ratio, ROE, Asset Turnover
      - For each ratio, state the computed value AND the score level it maps to

   NOTE: If the cashflow sheet is blank and income figures are not available,
   note which ratios cannot be computed and assess qualitatively.
   If the credit scoring sheet in the loan document is all zeros or blank,
   disregard it entirely and use only this BSV framework for scoring.

8. RISK FLAGS
   - Missing documents, data inconsistencies, concerns, or red flags

9. RECOMMENDATION
   - Choose one: APPROVE / CONDITIONALLY APPROVE / DECLINE
   - If CONDITIONALLY APPROVE: list the exact conditions required
   - Brief justification (3–5 sentences)

10. OVERALL RISK RATING
    - Low / Moderate / High with brief reason
    - Reference the score and stage from section 7

Be specific with peso amounts (use ₱ symbol). Use Philippine banking terminology.
Reference exact product names from the catalog.

--- LOAN APPLICATION DATA ---
{extracted_text}
--- END OF DATA ---

--- BANCO SAN VICENTE LOAN PRODUCT CATALOG ---
{loan_catalog_text}
--- END OF CATALOG ---

--- BSV CREDIT SCORING FRAMEWORK ---
{credit_scoring_framework}
--- END OF FRAMEWORK ---
"""


def _strip_fences(text: str) -> str:
    """Remove markdown code fences that some models add despite instructions."""
    text = re.sub(r'^```[a-z]*\n?', '', text.strip())
    text = re.sub(r'\n?```$', '', text)
    return text.strip()


# ══════════════════════════════════════════════════════════════════════════════
#  APPLICABILITY CHECK
# ══════════════════════════════════════════════════════════════════════════════

def check_loan_applicability(extracted_text: str) -> tuple[bool, str]:
    """
    Fast pre-flight call: is this a loan-related document?

    Returns
    -------
    (True,  "")         — proceed with full analysis
    (False, reason_str) — not applicable; reason explains why

    Raises
    ------
    RuntimeError — if the Gemini API call itself fails
    """
    from cic_parser import is_cic_report
    if is_cic_report(extracted_text):
        return True, "Document identified as a CIC Credit Report — fully applicable for credit analysis."

    clean = extracted_text.strip()
    if len(clean) < 80:
        return False, "the extracted text is too short or empty to contain a loan application."

    sample = _sample_text(clean, max_chars=2000)
    prompt = _APPLICABILITY_PROMPT.format(sample=sample)

    try:
        raw, model_used = _gemini_call_with_fallback(
            prompt      = prompt,
            max_tokens  = 120,
            temperature = 0.0,
            label       = "applicability check",
        )
        if model_used == "fallback":
            logger.info("Applicability check used fallback model: %s", _FALLBACK_MODEL)
    except RuntimeError:
        raise

    raw = _strip_fences(raw)

    try:
        data       = json.loads(raw)
        applicable = bool(data.get("applicable", True))
        reason     = str(data.get("reason", "")).strip()
        return applicable, reason
    except json.JSONDecodeError:
        logger.warning("Applicability check returned non-JSON: %r", raw)
        return True, ""


# ══════════════════════════════════════════════════════════════════════════════
#  CREDIT ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════

def run_credit_analysis(extracted_text: str) -> str:
    """
    Full credit assessment report (10 sections).

    Uses Gemini 2.5 Flash (1M token context window) — no truncation needed
    for typical loan documents. Falls back to Gemini 2.0 Flash on quota errors.

    Returns the formatted analysis string.

    Raises
    ------
    RuntimeError — if the Gemini API call fails
    """
    scoring_note             = _detect_scoring_note(extracted_text)
    loan_catalog_text        = build_loan_catalog_text()
    credit_scoring_framework = build_credit_scoring_framework_text()

    # Gemini 2.5 Flash supports 1M input tokens — no truncation needed
    # for typical loan documents. Only trim if truly enormous (>500k chars).
    MAX_EXTRACTED_CHARS = 500_000
    if len(extracted_text) > MAX_EXTRACTED_CHARS:
        half = MAX_EXTRACTED_CHARS // 2
        extracted_text = (
            extracted_text[:half]
            + "\n\n[... middle section truncated to fit token limit ...]\n\n"
            + extracted_text[-half:]
        )

    prompt = _ANALYSIS_PROMPT.format(
        scoring_note             = scoring_note,
        extracted_text           = extracted_text,
        loan_catalog_text        = loan_catalog_text,
        credit_scoring_framework = credit_scoring_framework,
    )

    # Inject CIC-specific instructions if applicable
    from cic_parser import is_cic_report, CIC_ANALYSIS_PROMPT_BLOCK
    if is_cic_report(extracted_text):
        prompt = CIC_ANALYSIS_PROMPT_BLOCK + "\n\n" + prompt

    try:
        result, model_used = _gemini_call_with_fallback(
            prompt      = prompt,
            max_tokens  = 65000,   # Gemini supports much larger outputs
            temperature = 0.2,
            label       = "credit analysis",
        )
    except RuntimeError:
        raise

    result = _strip_fences(result)

    # Notify which model was used
    if model_used == "fallback":
        result = (
            f"⚡ Note: Gemini 2.5 Flash quota reached — "
            f"this analysis was generated by {_FALLBACK_MODEL}.\n"
            f"Quality remains high. Quota resets after a short period.\n"
            + "─" * 56 + "\n\n"
            + result
        )
    elif model_used == "trimmed":
        result = (
            "⚠ Note: Document was very large — context was trimmed to fit token limits.\n"
            "Some details from the middle of the document may not be reflected.\n"
            + "─" * 56 + "\n\n"
            + result
        )

    return result


# ══════════════════════════════════════════════════════════════════════════════
#  PRIVATE HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _detect_scoring_note(extracted_text: str) -> str:
    """
    Inspect the extracted text for a Credit Scoring sheet.
    Returns a plain-English note to inject into the analysis prompt.
    """
    scoring_block = ""
    in_scoring    = False

    for line in extracted_text.splitlines():
        if "CREDIT SCORING" in line.upper():
            in_scoring = True
        elif in_scoring and line.startswith("=== SHEET:"):
            break
        if in_scoring:
            scoring_block += line + "\n"

    if not scoring_block:
        return (
            "NOTE: No Credit Scoring sheet was detected in the document. "
            "Use the BSV Credit Scoring Framework below to perform your own scoring.\n\n"
        )

    numbers  = re.findall(r'\b(\d+(?:\.\d+)?)\b', scoring_block)
    non_zero = [n for n in numbers if float(n) != 0.0]

    if not non_zero:
        return (
            "NOTE: The Credit Scoring sheet in the document exists but contains all zeros — "
            "it has NOT been filled out by the account officer. "
            "Disregard it entirely. Instead, use the BSV Credit Scoring Framework "
            "provided below to compute scores from the financial data in the CI/BI "
            "and Approval Form sheets.\n\n"
        )

    return (
        "NOTE: The Credit Scoring sheet in the document has been filled. "
        "Cross-reference it with the BSV Credit Scoring Framework below "
        "and flag any discrepancies.\n\n"
    )


def _sample_text(text: str, max_chars: int = 2000) -> str:
    """
    Return a representative sample of `text` up to `max_chars`.
    Takes beginning, middle, and end so long documents aren't
    classified purely on their header content.
    """
    if len(text) <= max_chars:
        return text

    third = max_chars // 3
    mid_start = (len(text) - third) // 2
    return (
        text[:third]
        + "\n…\n"
        + text[mid_start: mid_start + third]
        + "\n…\n"
        + text[-third:]
    )