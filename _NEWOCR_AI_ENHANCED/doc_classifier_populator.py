"""
doc_classifier_populator.py — DocExtract Pro
==============================================
Populates a CIBI Excel template with fields extracted by the
Document Classifier (VLM or regex).

Supported doc types and their target sheets/cells
---------------------------------------------------
  SPOUSE_PAYSLIP → CASHFLOW sheet
      A13  : "Spouse Salary — <Employee Name> (<Employer>)"
      G13  : Net Pay (monthly amount)
      (rows 12-16 are the 5 income rows; row 13 = spouse salary slot)

  All other doc types → generic field dump to first blank rows
  in CASHFLOW (extensible — add mappings below as needed).

Public API
----------
  populate_from_classifier(
      template_path, doc_type, fields, output_stem, progress_cb
  ) -> Path
      Opens the template, writes the extracted fields, saves a
      copy to Desktop/DocExtract_Files and returns the Path.
"""
from __future__ import annotations

import os
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Callable, Optional

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


# ── Output folder ─────────────────────────────────────────────────────────────

def _output_dir() -> Path:
    desktop = Path.home() / "Desktop"
    out = desktop / "DocExtract_Files"
    out.mkdir(parents=True, exist_ok=True)
    return out


# ── Helpers ───────────────────────────────────────────────────────────────────

def _field_map(fields: list[tuple[str, str, str]]) -> dict[str, str]:
    """Convert [(icon, label, value), ...] → {label_lower_snake: value}."""
    result = {}
    for _, label, value in fields:
        if value in ("[not found]", "[see raw text]", "", None):
            continue
        key = label.lower().replace(" ", "_").replace("/", "_")
        result[key] = value
    return result


def _parse_amount(raw: str) -> Optional[float]:
    """Parse a peso amount string to float. Returns None on failure."""
    if not raw:
        return None
    try:
        s = re.sub(r"(?i)^PHP", "", str(raw).strip())
        s = re.sub(r"^[P₱,]", "", s)
        s = s.replace(",", "").strip()
        return float(s) if s else None
    except (ValueError, AttributeError):
        return None


def _safe_write(ws, cell_addr: str, value, number_format: str = None):
    """Write a value to a cell, preserving existing style where possible."""
    cell = ws[cell_addr]
    cell.value = value
    if number_format:
        cell.number_format = number_format


# ── SPOUSE_PAYSLIP → CASHFLOW ─────────────────────────────────────────────────

def _populate_spouse_payslip(
    wb,
    fm: dict[str, str],
    cb: Callable,
) -> list[str]:
    """
    Write spouse salary data into the CASHFLOW sheet.

    CASHFLOW layout (income section):
        Row 11  : headers  — A=TYPE OF INCOME, G=MONTHLY, I=MONTHLY TOTALS
        Row 12  : income 1 (primary applicant salary — leave untouched)
        Row 13  : income 2 → SPOUSE SALARY  ← we write here
        Row 14  : income 3 (spare)
        Row 15  : income 4 (spare)
        Row 16  : income 5 (spare)
        Row 17  : TOTAL INCOME (formula — leave untouched)

    Fields written:
        A13 : label  e.g. "Spouse Salary — JUAN DELA CRUZ (BJMP)"
        G13 : net pay amount (float)
        I13 : same net pay amount (monthly total column)
    """
    notes = []

    if "CASHFLOW" not in wb.sheetnames:
        notes.append("⚠  CASHFLOW sheet not found in template.")
        return notes

    ws = wb["CASHFLOW"]
    cb(40, "Writing to CASHFLOW sheet…")

    # Build the label
    name      = fm.get("employee_name", "")
    employer  = fm.get("employer", "")
    period    = fm.get("period_covered", "")

    label_parts = ["Spouse Salary"]
    if name:
        label_parts.append(f"— {name}")
    if employer:
        label_parts.append(f"({employer})")
    if period:
        label_parts.append(f"[{period}]")
    label = " ".join(label_parts)

    # Get net pay
    net_pay_raw = fm.get("net_pay", "")
    net_pay     = _parse_amount(net_pay_raw)

    # Write label to A13
    _safe_write(ws, "A13", label)
    notes.append(f"✅  A13 ← {label}")

    # Write net pay to G13 (MONTHLY column) and I13 (MONTHLY TOTALS column)
    if net_pay is not None:
        _safe_write(ws, "G13", net_pay, number_format='#,##0.00')
        _safe_write(ws, "I13", net_pay, number_format='#,##0.00')
        notes.append(f"✅  G13 / I13 ← ₱{net_pay:,.2f} (Net Pay)")
    else:
        notes.append(f"⚠  Net Pay not found — G13/I13 left blank.")

    # Also write gross pay to a note cell (F13) so the officer can see it
    gross_raw = fm.get("gross_pay", "")
    gross     = _parse_amount(gross_raw)
    if gross is not None:
        _safe_write(ws, "F13", gross, number_format='#,##0.00')
        notes.append(f"ℹ  F13 ← ₱{gross:,.2f} (Gross Pay, semi-monthly col for reference)")

    cb(80, "CASHFLOW sheet updated.")
    return notes


# ── Generic fallback ──────────────────────────────────────────────────────────

def _populate_generic(
    wb,
    doc_type: str,
    fm: dict[str, str],
    cb: Callable,
) -> list[str]:
    """
    Fallback: dumps all extracted fields as key-value pairs into the
    first available sheet, starting at a safe row.
    """
    notes = []
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    cb(40, f"Writing fields to {sheet_name}…")

    start_row = 60  # safe row below most content
    ws[f"A{start_row}"] = f"Extracted fields — {doc_type}"
    ws[f"A{start_row}"].font = Font(bold=True)

    for i, (label, value) in enumerate(fm.items(), start=1):
        ws[f"A{start_row + i}"] = label.replace("_", " ").title()
        ws[f"B{start_row + i}"] = value
        notes.append(f"✅  {sheet_name}!A{start_row + i} ← {label}: {value}")

    cb(80, "Fields written.")
    return notes


# ── Main entry point ──────────────────────────────────────────────────────────

def populate_from_classifier(
    template_path: str,
    doc_type:      str,
    fields:        list[tuple[str, str, str]],
    output_stem:   str = "output",
    progress_cb:   Callable | None = None,
) -> Path:
    """
    Open the template, populate it with extracted fields, save a copy.

    Parameters
    ----------
    template_path : path to the user's Excel template (.xlsx)
    doc_type      : e.g. "SPOUSE_PAYSLIP", "PAYSLIP", etc.
    fields        : [(icon, label, value), ...] from VLM/regex extraction
    output_stem   : base name for the output file
    progress_cb   : optional (pct: int, msg: str) -> None

    Returns
    -------
    Path to the saved output file.
    """
    if not _HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required: pip install openpyxl --break-system-packages"
        )

    def _cb(pct: int, msg: str = ""):
        if progress_cb:
            try:
                progress_cb(pct, msg)
            except Exception:
                pass

    _cb(5, "Opening template…")

    wb = openpyxl.load_workbook(template_path)
    fm = _field_map(fields)

    _cb(20, f"Populating {doc_type} fields…")

    # ── Route to correct populator ────────────────────────────────────────
    if doc_type == "SPOUSE_PAYSLIP":
        notes = _populate_spouse_payslip(wb, fm, _cb)
    else:
        notes = _populate_generic(wb, doc_type, fm, _cb)

    # ── Save output ───────────────────────────────────────────────────────
    _cb(90, "Saving output file…")

    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_name   = f"{output_stem}_{doc_type}_{timestamp}.xlsx"
    out_path   = _output_dir() / out_name

    wb.save(str(out_path))

    _cb(100, f"Saved: {out_path.name}")
    print(f"[populate_from_classifier] Saved to: {out_path}")
    for note in notes:
        print(f"  {note}")

    return out_path