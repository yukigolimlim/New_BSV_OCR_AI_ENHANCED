"""
lu_loanbal_export_patch.py — Sector vs Loan Balance Tab
=========================================================
Displays the sector / client loan-balance exposure table and
provides PDF + Excel export.

This file is the CANONICAL implementation for the Loan Balance tab.
It supersedes _loanbal_* / _generate_loanbal_* functions in lu_ui.py.

Standalone: imports only lu_core and lu_shared.
Attached to app class via attach(cls).

Export behaviour
----------------
  • On-screen table respects the active sector filter.
  • PDF and Excel exports always use the FULL unfiltered dataset
    (by design — a notice is shown when a filter is active).

Public surface
--------------
  attach(cls)
  _build_loanbal_panel(self, parent)
  _loanbal_render(self)
  _loanbal_show_export_menu(self)
  _loanbal_export_pdf(self)
  _loanbal_export_excel(self)
  _generate_loanbal_pdf(all_data, out_path, filepath)
  _generate_loanbal_excel(all_data, out_path, filepath)
"""

import tkinter as tk
import tkinter.ttk as ttk
import customtkinter as ctk
import re
from pathlib import Path
from datetime import datetime
from tkinter import filedialog, messagebox

from lu_core import (
    _parse_numeric,
    SECTOR_OTHER,
)
from lu_shared import (
    F, FF,
    LU_CLIENT_TREE_SPEC, lu_client_row_tuple,
    LU_SECTOR_UNSPECIFIED_LABEL,
    _NAVY_DEEP, _NAVY_MID, _NAVY_LIGHT, _NAVY_MIST, _NAVY_GHOST, _NAVY_PALE,
    _WHITE, _CARD_WHITE, _OFF_WHITE, _BORDER_LIGHT, _BORDER_MID,
    _TXT_NAVY, _TXT_SOFT, _TXT_MUTED, _TXT_ON_LIME,
    _LIME_MID, _LIME_DARK,
    _ACCENT_RED, _ACCENT_GOLD, _ACCENT_SUCCESS,
    _RISK_COLOR, _RISK_BADGE_BG,
    _SECTOR_COLORS, _SECTOR_ICON, _CHART_SECTORS,
    _SECTOR_COLS,
    _make_scrollable, _make_table_frame, _table_header, _table_divider,
    _lu_filter_data_by_query,
    _lu_get_active_sectors, _lu_get_filtered_all_data,
)

try:
    from reportlab.lib.pagesizes import A4, landscape as rl_landscape
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, PageBreak,
    )
    _HAS_RL = True
except ImportError:
    _HAS_RL = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPX = True
except ImportError:
    _HAS_OPX = False


def _build_industry_map(records: list[dict], ordered_names: list[str] | None = None) -> dict[str, list[dict]]:
    """Build industry -> records map using normalized tags with safe fallback."""
    industry_map: dict[str, list[dict]] = {}
    splitter = re.compile(r"\s*(?:,|/|;|&|\band\b)\s*", re.I)
    for rec in records:
        tags = rec.get("industry_tags") or []
        if not tags:
            raw = str(rec.get("industry") or "").strip()
            tags = [t.strip() for t in splitter.split(raw) if t.strip()] if raw else []
        if not tags:
            tags = [LU_SECTOR_UNSPECIFIED_LABEL]
        for tag in tags:
            industry_map.setdefault(tag, []).append(rec)

    if ordered_names:
        ordered_map = {name: industry_map[name] for name in ordered_names if name in industry_map}
        for name, recs in industry_map.items():
            if name not in ordered_map:
                ordered_map[name] = recs
        return ordered_map
    return industry_map


def _risk_label_from_recs(recs: list[dict]) -> str:
    """
    Risk for an industry/sector group from client score labels only.
    Avoids flattening synthetic expense rows (was very slow on large books).
    """
    labels = {str(r.get("score_label") or "LOW").strip().upper() for r in recs}
    if "HIGH" in labels:
        return "HIGH"
    if "MODERATE" in labels:
        return "MODERATE"
    return "LOW"


_LOANBAL_SEARCH_DEBOUNCE_MS = 300


def _loanbal_schedule_render(self):
    """Debounce search typing — full rebuild + matplotlib is expensive."""
    jid = getattr(self, "_loanbal_render_job", None)
    if jid is not None:
        try:
            self.after_cancel(jid)
        except Exception:
            pass
    self._loanbal_render_job = self.after(
        _LOANBAL_SEARCH_DEBOUNCE_MS,
        lambda: _loanbal_run_render_job(self),
    )


def _loanbal_run_render_job(self):
    self._loanbal_render_job = None
    _loanbal_render(self)


# ══════════════════════════════════════════════════════════════════════
#  PANEL BUILDER
# ══════════════════════════════════════════════════════════════════════

def _build_loanbal_panel(self, parent):
    hdr = tk.Frame(parent, bg=_NAVY_MID, height=46)
    hdr.pack(fill="x")
    hdr.pack_propagate(False)

    self._loanbal_hdr_lbl = tk.Label(
        hdr, text="📊  Sector vs Total Loan Balance  —  Exposure Analysis",
        font=F(10, "bold"), fg=_WHITE, bg=_NAVY_MID)
    self._loanbal_hdr_lbl.pack(side="left", padx=20, pady=12)
    tk.Label(hdr, text="🔎", font=F(9), fg=_WHITE, bg=_NAVY_MID).pack(side="left", padx=(8, 4))
    self._loanbal_search_var = tk.StringVar()
    self._loanbal_search_var.trace_add(
        "write",
        lambda *_: (_loanbal_schedule_render(self)
                    if getattr(self, "_lu_all_data", None) else None))
    tk.Entry(
        hdr, textvariable=self._loanbal_search_var,
        font=F(8), relief="flat", bg=_WHITE, fg=_TXT_NAVY,
        insertbackground=_TXT_NAVY, highlightbackground=_NAVY_LIGHT, highlightthickness=1
    ).pack(side="left", padx=(0, 8), ipady=3)
    self._loanbal_match_lbl = tk.Label(
        hdr, text="", font=F(8, "bold"), fg=_WHITE, bg=_NAVY_MID, padx=8, pady=3)
    self._loanbal_match_lbl.pack(side="left", padx=(0, 8), pady=8)

    self._loanbal_export_btn = ctk.CTkButton(
        hdr, text="💾  Export",
        command=lambda: _loanbal_show_export_menu(self),
        width=110, height=30, corner_radius=6,
        fg_color=_LIME_DARK, hover_color=_LIME_MID,
        text_color=_TXT_ON_LIME, font=FF(9, "bold"),
        state="disabled")
    self._loanbal_export_btn.pack(side="right", padx=16, pady=8)

    self._loanbal_body = tk.Frame(parent, bg=_CARD_WHITE)
    self._loanbal_body.pack(fill="both", expand=True)
    tk.Label(self._loanbal_body,
             text="Run an analysis first to view loan balance exposure.",
             font=F(10), fg=_TXT_MUTED, bg=_CARD_WHITE).pack(pady=60)


# ══════════════════════════════════════════════════════════════════════
#  RENDERER
# ══════════════════════════════════════════════════════════════════════

def _loanbal_render(self):
    for w in self._loanbal_body.winfo_children():
        w.destroy()
    # Note: do not call plt.close("all") — it tears down charts in other LU tabs.

    all_data       = _lu_get_filtered_all_data(self)
    q              = getattr(self, "_loanbal_search_var", tk.StringVar(value="")).get().strip()
    all_data       = _lu_filter_data_by_query(all_data, q)
    general        = all_data.get("general", [])
    match_lbl      = getattr(self, "_loanbal_match_lbl", None)
    if match_lbl is not None:
        if q:
            client_names = sorted({(r.get("client") or "").strip() for r in general if r.get("client")})
            if len(client_names) == 1:
                match_lbl.config(text=client_names[0][:28], bg="#4A6FA5")
            else:
                match_lbl.config(text=f"{len(general)} CLIENTS MATCHED", bg="#4A6FA5")
        else:
            match_lbl.config(text="", bg=_NAVY_MID)
    unique_industries = all_data.get("unique_industries", [])
    industry_map   = _build_industry_map(general, unique_industries)
    totals         = all_data.get("totals", {})
    grand_lb       = totals.get("loan_balance", 0) or 0
    active_sectors = _lu_get_active_sectors(self)

    if q:
        self._loanbal_hdr_lbl.config(
            text=f"📊  Loan Balance — Search: {q[:30]}",
            fg=_LIME_MID)
    elif active_sectors:
        self._loanbal_hdr_lbl.config(
            text=f"📊  Loan Balance — Filtered: {' · '.join(active_sectors)}",
            fg=_LIME_MID)
    else:
        self._loanbal_hdr_lbl.config(
            text="📊  Sector vs Total Loan Balance  —  Exposure Analysis",
            fg=_WHITE)

    self._loanbal_export_btn.configure(state="normal")

    if not general:
        tk.Label(self._loanbal_body, text="No data available for this filter.",
                 font=F(10), fg=_TXT_MUTED, bg=_CARD_WHITE).pack(pady=60)
        return

    # Use scrollable body so section charts/tables are always reachable.
    outer, inner, _ = _make_scrollable(self._loanbal_body, _CARD_WHITE)
    outer.pack(fill="both", expand=True)
    pad = tk.Frame(inner, bg=_CARD_WHITE)
    pad.pack(fill="both", expand=True, padx=20, pady=14)

    # ── Grand-total card ──────────────────────────────────────────────
    grand_card = tk.Frame(pad, bg=_NAVY_DEEP,
                          highlightbackground=_NAVY_MID, highlightthickness=1)
    grand_card.pack(fill="x", pady=(0, 14))
    gc_inner = tk.Frame(grand_card, bg=_NAVY_DEEP)
    gc_inner.pack(fill="x", padx=22, pady=14)

    left_gc = tk.Frame(gc_inner, bg=_NAVY_DEEP)
    left_gc.pack(side="left", fill="y")
    if active_sectors:
        tk.Label(left_gc,
                 text=f"💰  FILTERED LOAN BALANCE  ·  {' · '.join(active_sectors)}",
                 font=F(9, "bold"), fg=_LIME_MID, bg=_NAVY_DEEP).pack(anchor="w")
    else:
        tk.Label(left_gc, text="💰  GRAND TOTAL LOAN BALANCE",
                 font=F(9, "bold"), fg=_TXT_MUTED, bg=_NAVY_DEEP).pack(anchor="w")
    tk.Label(left_gc, text=f"₱{grand_lb:,.2f}",
             font=F(20, "bold"), fg=_LIME_MID, bg=_NAVY_DEEP).pack(anchor="w")
    tk.Label(left_gc,
             text=f"{len(general)} clients  ·  {len(industry_map)} industry(ies)",
             font=F(9), fg=_TXT_SOFT, bg=_NAVY_DEEP).pack(anchor="w")
    if active_sectors:
        tk.Label(left_gc,
                 text="⚠  Export button exports the full unfiltered dataset",
                 font=F(7), fg=_ACCENT_GOLD, bg=_NAVY_DEEP).pack(anchor="w", pady=(4, 0))

    right_gc = tk.Frame(gc_inner, bg=_NAVY_DEEP)
    right_gc.pack(side="right", fill="y")
    total_net = sum(r.get("net_income") or 0 for r in general)
    for lbl, val in [
        ("Total Net Income",    f"₱{total_net:,.2f}"),
        ("Avg Loan per Client", f"₱{grand_lb/len(general):,.2f}" if general else "—"),
    ]:
        c = tk.Frame(right_gc, bg=_NAVY_DEEP)
        c.pack(anchor="e", pady=2)
        tk.Label(c, text=lbl, font=F(8), fg=_TXT_MUTED, bg=_NAVY_DEEP).pack(side="left", padx=(0, 8))
        tk.Label(c, text=val, font=F(10, "bold"), fg=_WHITE, bg=_NAVY_DEEP).pack(side="left")

    # ── Sector table ──────────────────────────────────────────────────
    tk.Label(pad, text="Industry Loan Balance Breakdown",
             font=F(11, "bold"), fg=_TXT_NAVY, bg=_CARD_WHITE
             ).pack(anchor="w", pady=(0, 6))

    sector_tf = _make_table_frame(pad, _SECTOR_COLS)
    _table_header(sector_tf, _SECTOR_COLS)
    _table_divider(sector_tf, 1, len(_SECTOR_COLS), _NAVY_MID)

    industry_rows = []
    for industry, recs in industry_map.items():
        n       = len(recs)
        s_lb    = sum(r.get("loan_balance") or 0 for r in recs)
        s_net   = sum(r.get("net_income")   or 0 for r in recs)
        pct     = (s_lb / grand_lb * 100) if grand_lb > 0 else 0.0
        avg_lb  = s_lb / n  if n > 0 else 0.0
        avg_net = s_net / n if n > 0 else 0.0
        risk_label = _risk_label_from_recs(recs)
        industry_rows.append((industry, n, s_lb, pct, avg_lb, avg_net, risk_label))

    # Hard fallback: if industry parsing fails for a file format, derive rows by sector.
    if not industry_rows and general:
        sector_map = {}
        for rec in general:
            sec = (
                (rec.get("sector") or rec.get("industry") or LU_SECTOR_UNSPECIFIED_LABEL).strip()
                or LU_SECTOR_UNSPECIFIED_LABEL
            )
            sector_map.setdefault(sec, []).append(rec)
        for sec, recs in sector_map.items():
            n       = len(recs)
            s_lb    = sum(r.get("loan_balance") or 0 for r in recs)
            s_net   = sum(r.get("net_income")   or 0 for r in recs)
            pct     = (s_lb / grand_lb * 100) if grand_lb > 0 else 0.0
            avg_lb  = s_lb / n  if n > 0 else 0.0
            avg_net = s_net / n if n > 0 else 0.0
            risk_label = _risk_label_from_recs(recs)
            industry_rows.append((sec, n, s_lb, pct, avg_lb, avg_net, risk_label))
    industry_rows.sort(key=lambda x: -x[2])

    tk.Label(
        pad,
        text=f"Showing {len(industry_rows)} sector row(s) from {len(general)} client(s).",
        font=F(8),
        fg=_TXT_MUTED,
        bg=_CARD_WHITE,
    ).pack(anchor="w", pady=(0, 6))

    if not industry_rows:
        tk.Label(
            pad,
            text="No industry-level rows found. Re-scan the file to refresh parsed industries.",
            font=F(9),
            fg=_TXT_MUTED,
            bg=_CARD_WHITE
        ).pack(anchor="w", pady=(4, 12))

    grid_row = 2
    for idx, (industry, n, s_lb, pct, avg_lb, avg_net, risk_label) in enumerate(industry_rows):
        row_bg    = _CARD_WHITE if idx % 2 == 0 else _OFF_WHITE
        col_color = _NAVY_MID
        icon      = "🏭"
        risk_fg   = _RISK_COLOR.get(risk_label, _TXT_SOFT)
        risk_bg   = _RISK_BADGE_BG.get(risk_label, _OFF_WHITE)
        if active_sectors and industry in active_sectors:
            row_bg = "#0E2040"

        stripe = tk.Frame(sector_tf, bg=row_bg)
        stripe.grid(row=grid_row, column=0, columnspan=len(_SECTOR_COLS), sticky="nsew")
        stripe.lower()

        name_fg = _LIME_MID if (active_sectors and industry in active_sectors) else col_color
        val_fg  = _WHITE    if (active_sectors and industry in active_sectors) else _TXT_NAVY

        tk.Label(sector_tf, text=f"  {icon}  {industry}", font=F(9, "bold"),
                 fg=name_fg, bg=row_bg, anchor="w", padx=6, pady=10
                 ).grid(row=grid_row, column=0, sticky="nsew")
        tk.Label(sector_tf, text=str(n), font=F(9),
                 fg=val_fg, bg=row_bg, anchor="center", padx=6, pady=10
                 ).grid(row=grid_row, column=1, sticky="nsew")
        tk.Label(sector_tf, text=f"₱{s_lb:,.2f}", font=F(9, "bold"),
                 fg=val_fg, bg=row_bg, anchor="center", padx=10, pady=10
                 ).grid(row=grid_row, column=2, sticky="nsew")

        pct_cell = tk.Frame(sector_tf, bg=row_bg)
        pct_cell.grid(row=grid_row, column=3, sticky="nsew", padx=6, pady=6)
        tk.Label(pct_cell, text=f"{pct:.1f}%", font=F(9, "bold"),
                 fg=name_fg, bg=row_bg, anchor="center").pack(anchor="w", pady=(2, 1))
        bar_outer = tk.Frame(pct_cell, bg=_BORDER_LIGHT, height=6)
        bar_outer.pack(fill="x", pady=(0, 2))
        bar_outer.pack_propagate(False)
        fill_w = max(3, int(90 * pct / 100))
        tk.Frame(bar_outer, bg=col_color, height=6, width=fill_w).place(x=0, y=0, relheight=1)

        tk.Label(sector_tf, text=f"₱{avg_lb:,.2f}", font=F(9),
                 fg=val_fg, bg=row_bg, anchor="center", padx=10, pady=10
                 ).grid(row=grid_row, column=4, sticky="nsew")
        tk.Label(sector_tf, text=f"₱{avg_net:,.2f}" if avg_net else "—",
                 font=F(9), fg=val_fg, bg=row_bg, anchor="center", padx=10, pady=10
                 ).grid(row=grid_row, column=5, sticky="nsew")

        risk_badge = tk.Frame(sector_tf, bg=row_bg)
        risk_badge.grid(row=grid_row, column=6, sticky="nsew", padx=8, pady=8)
        tk.Label(risk_badge, text=risk_label, font=F(8, "bold"),
                 fg=risk_fg, bg=risk_bg, padx=8, pady=3).pack()
        grid_row += 1

    # Total row
    total_bg = "#EEF3FB"
    for ci in range(len(_SECTOR_COLS)):
        stripe = tk.Frame(sector_tf, bg=total_bg)
        stripe.grid(row=grid_row, column=0, columnspan=len(_SECTOR_COLS), sticky="nsew")
        stripe.lower()
    tk.Label(sector_tf, text="  GRAND TOTAL", font=F(9, "bold"),
             fg=_NAVY_MID, bg=total_bg, anchor="w", padx=6, pady=10
             ).grid(row=grid_row, column=0, sticky="nsew")
    tk.Label(sector_tf, text=str(len(general)), font=F(9, "bold"),
             fg=_NAVY_MID, bg=total_bg, anchor="center", padx=6, pady=10
             ).grid(row=grid_row, column=1, sticky="nsew")
    tk.Label(sector_tf, text=f"₱{grand_lb:,.2f}", font=F(9, "bold"),
             fg=_NAVY_MID, bg=total_bg, anchor="center", padx=10, pady=10
             ).grid(row=grid_row, column=2, sticky="nsew")
    for c in range(3, len(_SECTOR_COLS)):
        tk.Label(sector_tf, text="—", font=F(9), fg=_TXT_MUTED, bg=total_bg,
                 anchor="center").grid(row=grid_row, column=c, sticky="nsew")

    # ── Chart section: share + sector bars ───────────────────────────
    tk.Label(pad, text="Loan Balance Share by Sector",
             font=F(11, "bold"), fg=_TXT_NAVY, bg=_CARD_WHITE
             ).pack(anchor="w", pady=(18, 6))

    chart_card = tk.Frame(
        pad, bg=_WHITE, highlightbackground=_BORDER_MID, highlightthickness=1)
    chart_card.pack(fill="x", pady=(0, 14))

    if industry_rows:
        try:
            import matplotlib
            matplotlib.use("TkAgg")
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

            # Keep chart readable: show top sectors and collapse the tail to "Others".
            sorted_rows = sorted(industry_rows, key=lambda r: r[2], reverse=True)
            pie_max_items = 6
            bar_max_items = 10

            pie_rows = sorted_rows[:pie_max_items]
            pie_other = sorted_rows[pie_max_items:]
            if pie_other:
                other_lb = sum(r[2] for r in pie_other)
                other_pct = sum(r[3] for r in pie_other)
                pie_rows.append(("Others", len(pie_other), other_lb, other_pct, 0.0, 0.0, "LOW"))

            names = [r[0] for r in pie_rows]
            lbs = [r[2] for r in pie_rows]
            pcts = [r[3] for r in pie_rows]
            colors = [plt.cm.tab20(i % 20) for i in range(len(names))]

            bar_rows = sorted_rows[:bar_max_items]
            bar_names = [r[0] for r in bar_rows]
            bar_lbs = [r[2] for r in bar_rows]
            bar_pcts = [r[3] for r in bar_rows]
            bar_colors = [plt.cm.tab20(i % 20) for i in range(len(bar_names))]

            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(10.0, 3.6))
            fig.patch.set_facecolor(_WHITE)
            ax1.set_facecolor(_WHITE)
            ax2.set_facecolor(_WHITE)

            # Donut chart
            wedges, _, _ = ax1.pie(
                lbs,
                labels=None,
                colors=colors,
                startangle=90,
                wedgeprops={"width": 0.42, "edgecolor": _WHITE},
                autopct=lambda p: f"{p:.1f}%" if p >= 6 else "",
                pctdistance=0.82,
                textprops={"fontsize": 7, "color": "#1A2B4A"},
            )
            ax1.text(0, 0.05, f"₱{grand_lb/1e6:.2f}M", ha="center", va="center",
                     fontsize=10, fontweight="bold", color="#1A2B4A")
            ax1.text(0, -0.13, "total", ha="center", va="center",
                     fontsize=8, color=_TXT_MUTED)
            ax1.set_title("Loan Balance Share", fontsize=9, color="#1A2B4A", pad=8)
            legend_labels = [f"{n[:20]}  {p:.1f}%" for n, p in zip(names, pcts)]
            ax1.legend(
                wedges, legend_labels,
                loc="lower center", bbox_to_anchor=(0.5, -0.18),
                ncol=3, fontsize=6, frameon=False
            )

            # Horizontal bar chart
            y = list(range(len(bar_names)))
            ax2.barh(y, bar_lbs, color=bar_colors, height=0.62)
            ax2.set_yticks(y, [n[:28] for n in bar_names], fontsize=7)
            ax2.invert_yaxis()
            ax2.tick_params(axis="x", labelsize=7)
            ax2.spines[["top", "right"]].set_visible(False)
            ax2.set_title("Loan Balance by Sector", fontsize=9, color="#1A2B4A", pad=8)
            for i, (v, p) in enumerate(zip(bar_lbs, bar_pcts)):
                ax2.text(v, i, f" {p:.1f}%", va="center", fontsize=7, color="#1A2B4A")
            ax2.xaxis.set_major_formatter(
                matplotlib.ticker.FuncFormatter(
                    lambda x, _: f"₱{x/1e6:.1f}M" if x >= 1e6 else f"₱{x:,.0f}"
                )
            )

            fig.subplots_adjust(wspace=0.20, bottom=0.32, left=0.08, right=0.98, top=0.92)
            self._loanbal_chart_canvas = FigureCanvasTkAgg(fig, master=chart_card)
            self._loanbal_chart_canvas.get_tk_widget().pack(fill="x", padx=8, pady=8)
            plt.close(fig)
        except Exception:
            tk.Label(chart_card,
                     text="Chart preview unavailable (matplotlib issue).",
                     font=F(9), fg=_TXT_MUTED, bg=_WHITE).pack(pady=18)
    else:
        tk.Label(chart_card,
                 text="No sector rows available for charting.",
                 font=F(9), fg=_TXT_MUTED, bg=_WHITE).pack(pady=18)

    # ── Client table: Excel column order; Treeview replaces one Label per cell (major perf win)
    tk.Label(
        pad,
        text="Individual clients — columns match your uploaded Excel (scroll horizontally)",
        font=F(11, "bold"),
        fg=_TXT_NAVY,
        bg=_CARD_WHITE,
    ).pack(anchor="w", pady=(20, 6))

    cli_box = tk.Frame(pad, bg=_CARD_WHITE, height=360)
    cli_box.pack(fill="both", expand=True)
    cli_box.pack_propagate(False)

    vsb_cli = ttk.Scrollbar(cli_box, orient="vertical")
    vsb_cli.pack(side="right", fill="y")
    hsb_cli = ttk.Scrollbar(cli_box, orient="horizontal")
    hsb_cli.pack(side="bottom", fill="x")

    lb_cols = tuple(c[0] for c in LU_CLIENT_TREE_SPEC)
    tstyle = ttk.Style()
    tstyle.theme_use("default")
    tstyle.configure(
        "LB.Treeview",
        background=_WHITE,
        foreground=_TXT_NAVY,
        rowheight=22,
        fieldbackground=_WHITE,
        bordercolor=_BORDER_MID,
        font=("Segoe UI", 8),
    )
    tstyle.configure(
        "LB.Treeview.Heading",
        background=_NAVY_MID,
        foreground=_WHITE,
        font=("Segoe UI", 8, "bold"),
        relief="flat",
    )
    tstyle.map("LB.Treeview", background=[("selected", _NAVY_GHOST)])

    cli_tree = ttk.Treeview(
        cli_box,
        columns=lb_cols,
        show="headings",
        style="LB.Treeview",
        yscrollcommand=vsb_cli.set,
        xscrollcommand=hsb_cli.set,
        selectmode="browse",
    )
    cli_tree.pack(side="left", fill="both", expand=True)
    vsb_cli.config(command=cli_tree.yview)
    hsb_cli.config(command=cli_tree.xview)

    _stretch_lb = {
        "client", "industry", "source_income", "biz_exp_detail", "hhld_exp_detail", "product_name",
        "personal_assets", "business_assets", "business_inventory",
    }
    for cid, heading, _field, width, anchor, _k in LU_CLIENT_TREE_SPEC:
        cli_tree.heading(cid, text=heading)
        cli_tree.column(
            cid,
            width=width,
            minwidth=36,
            anchor=anchor,
            stretch=(cid in _stretch_lb),
        )

    clients_sorted = sorted(general, key=lambda r: -(r.get("loan_balance") or 0))
    for idx, rec in enumerate(clients_sorted):
        rl = rec.get("score_label", "N/A")
        tag = rl if rl in ("HIGH", "LOW") else "NA"
        if tag == "NA" and idx % 2 == 1:
            tag = "alt"
        cli_tree.insert("", "end", values=lu_client_row_tuple(rec), tags=(tag,))

    cli_tree.tag_configure("HIGH", background="#FFF5F5", foreground=_ACCENT_RED)
    cli_tree.tag_configure("LOW", background="#F0FBE8", foreground=_ACCENT_SUCCESS)
    cli_tree.tag_configure("NA", background=_WHITE, foreground=_TXT_MUTED)
    cli_tree.tag_configure("alt", background=_OFF_WHITE)

    def _cli_tree_wheel(ev):
        cli_tree.yview_scroll(int(-1 * (ev.delta / 120)), "units")
        return "break"

    cli_tree.bind("<MouseWheel>", _cli_tree_wheel)


# ══════════════════════════════════════════════════════════════════════
#  EXPORT MENU
# ══════════════════════════════════════════════════════════════════════

def _loanbal_show_export_menu(self):
    active_sectors = _lu_get_active_sectors(self)
    q = getattr(self, "_loanbal_search_var", tk.StringVar(value="")).get().strip()
    if active_sectors:
        sector_text = " · ".join(active_sectors)
        pdf_label = f"📄  Export PDF — Full Report (filter: {sector_text[:30]})"
        xl_label  = f"📊  Export Excel — Current Filter ({sector_text[:30]})"
    elif q:
        pdf_label = "📄  Export PDF — Full Report (current view filtered)"
        xl_label  = "📊  Export Excel — Current Filter"
    else:
        pdf_label = "📄  Export PDF — Sector & Client Loan Balance"
        xl_label  = "📊  Export Excel — Current Filter"

    menu = tk.Menu(
        self._lu_loanbal_view, tearoff=0,
        font=F(9), bg=_WHITE, fg=_TXT_NAVY,
        activebackground="#D6E4F7", activeforeground=_NAVY_DEEP,
        relief="flat", bd=1)
    menu.add_command(label=pdf_label, command=lambda: _loanbal_export_pdf(self))
    menu.add_command(label=xl_label,  command=lambda: _loanbal_export_excel(self))
    try:
        btn = self._loanbal_export_btn
        btn.update_idletasks()
        menu.tk_popup(btn.winfo_rootx(), btn.winfo_rooty() + btn.winfo_height())
    finally:
        menu.grab_release()


# ══════════════════════════════════════════════════════════════════════
#  PDF EXPORT
# ══════════════════════════════════════════════════════════════════════

def _loanbal_export_pdf(self):
    if not self._lu_all_data:
        messagebox.showwarning("No Data", "Run an analysis first.")
        return
    if not _HAS_RL:
        messagebox.showerror("Missing Library",
                             "reportlab is not installed.\nRun:  pip install reportlab")
        return
    default_name = f"LoanBalance_Exposure_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    path = filedialog.asksaveasfilename(
        title="Save Loan Balance PDF",
        defaultextension=".pdf",
        filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
        initialfile=default_name)
    if not path:
        return
    try:
        _generate_loanbal_pdf(self._lu_all_data, path, filepath=self._lu_filepath or "")
        messagebox.showinfo("Export Complete", f"PDF saved to:\n{path}")
    except Exception as ex:
        messagebox.showerror("PDF Export Error", str(ex))


def _generate_loanbal_pdf(all_data, out_path, filepath=""):
    PAGE  = rl_landscape(A4)
    doc   = SimpleDocTemplate(out_path, pagesize=PAGE,
                               leftMargin=1.5*cm, rightMargin=1.5*cm,
                               topMargin=1.8*cm, bottomMargin=1.8*cm)
    styles = getSampleStyleSheet()
    navy   = rl_colors.HexColor("#1A3A6B")
    lime   = rl_colors.HexColor("#5A9E28")
    white  = rl_colors.white
    off    = rl_colors.HexColor("#F5F7FA")
    mist   = rl_colors.HexColor("#EEF3FB")
    border = rl_colors.HexColor("#C5D0E8")
    red    = rl_colors.HexColor("#E53E3E")
    gold   = rl_colors.HexColor("#D4A017")
    green  = rl_colors.HexColor("#2E7D32")

    title_s = ParagraphStyle("LBTitle", parent=styles["Title"],
                              fontSize=16, textColor=navy, spaceAfter=2)
    sub_s   = ParagraphStyle("LBSub",   parent=styles["Normal"],
                              fontSize=8,  textColor=rl_colors.HexColor("#9AAACE"), leading=10)
    h2_s    = ParagraphStyle("LBH2",    parent=styles["Normal"],
                              fontSize=10, textColor=navy, leading=14, spaceBefore=10)
    body_s  = ParagraphStyle("LBBody",  parent=styles["Normal"],
                              fontSize=8,  textColor=rl_colors.HexColor("#1A2B4A"), leading=11)
    muted_s = ParagraphStyle("LBMuted", parent=styles["Normal"],
                              fontSize=7,  textColor=rl_colors.HexColor("#9AAACE"), leading=10)

    RISK_RL = {"HIGH": red, "MODERATE": gold, "LOW": green, "N/A": rl_colors.grey}

    general    = all_data.get("general", [])
    unique_industries = all_data.get("unique_industries", [])
    industry_map = _build_industry_map(general, unique_industries)
    totals     = all_data.get("totals", {})
    grand_lb   = totals.get("loan_balance", 0) or 0
    now        = datetime.now().strftime("%B %d, %Y  %H:%M")
    fname      = Path(filepath).name if filepath else "—"

    def _rgb_hex(color):
        try:
            return f"{int(color.red*255):02X}{int(color.green*255):02X}{int(color.blue*255):02X}"
        except Exception:
            return "000000"

    story = []
    story.append(Paragraph("Sector vs Loan Balance — Exposure Analysis", title_s))
    story.append(Paragraph(
        f"File: {fname}    Generated: {now}    "
        f"Total Clients: {len(general)}    "
        f"Grand Total Loan Balance: ₱{grand_lb:,.2f}", sub_s))
    story.append(Spacer(1, 0.3*cm))
    story.append(HRFlowable(width="100%", thickness=1.5, color=navy))
    story.append(Spacer(1, 0.4*cm))
    story.append(Paragraph("Industry Loan Balance Breakdown", h2_s))
    story.append(Spacer(1, 0.2*cm))

    sector_rows_data = []
    for industry, recs in industry_map.items():
        n       = len(recs)
        s_lb    = sum(r.get("loan_balance") or 0 for r in recs)
        s_net   = sum(r.get("net_income")   or 0 for r in recs)
        pct     = (s_lb / grand_lb * 100) if grand_lb > 0 else 0.0
        avg_lb  = s_lb / n  if n > 0 else 0.0
        avg_net = s_net / n if n > 0 else 0.0
        risk_label = _risk_label_from_recs(recs)
        sector_rows_data.append((industry, n, s_lb, pct, avg_lb, avg_net, risk_label))
    sector_rows_data.sort(key=lambda x: -x[2])

    pg_w   = PAGE[0] - 3*cm
    s_cols = [pg_w*0.22, pg_w*0.09, pg_w*0.17, pg_w*0.12,
              pg_w*0.17, pg_w*0.14, pg_w*0.09]

    sec_hdr_s = ParagraphStyle("SecHdr", parent=styles["Normal"],
                               fontSize=8, textColor=white, leading=11)
    sec_tbl_data = [[
        Paragraph("<b>Sector</b>",             sec_hdr_s),
        Paragraph("<b># Clients</b>",          sec_hdr_s),
        Paragraph("<b>Total Loan Balance</b>", sec_hdr_s),
        Paragraph("<b>% of Total</b>",         sec_hdr_s),
        Paragraph("<b>Avg Loan / Client</b>",  sec_hdr_s),
        Paragraph("<b>Avg Net Income</b>",     sec_hdr_s),
        Paragraph("<b>Risk Profile</b>",       sec_hdr_s),
    ]]
    for industry, n, s_lb, pct, avg_lb, avg_net, risk_label in sector_rows_data:
        icon    = "🏭"
        col_hex = "1A3A6B"
        risk_col = RISK_RL.get(risk_label, rl_colors.grey)
        sec_tbl_data.append([
            Paragraph(f"<font color='#{col_hex}'><b>{icon}  {industry}</b></font>", body_s),
            Paragraph(str(n), body_s),
            Paragraph(f"<b>₱{s_lb:,.2f}</b>", body_s),
            Paragraph(f"{pct:.1f}%", body_s),
            Paragraph(f"₱{avg_lb:,.2f}", body_s),
            Paragraph(f"₱{avg_net:,.2f}" if avg_net else "—", body_s),
            Paragraph(
                f"<font color='#{_rgb_hex(risk_col)}'><b>{risk_label}</b></font>",
                body_s),
        ])
    sec_tbl_data.append([
        Paragraph("<b>GRAND TOTAL</b>", body_s),
        Paragraph(f"<b>{len(general)}</b>", body_s),
        Paragraph(f"<b>₱{grand_lb:,.2f}</b>", body_s),
        Paragraph("<b>100.0%</b>", body_s),
        Paragraph(f"<b>₱{grand_lb/len(general):,.2f}</b>" if general else "—", body_s),
        Paragraph("—", body_s),
        Paragraph("—", body_s),
    ])

    sec_style = TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  navy),
        ("FONTSIZE",      (0, 0), (-1, -1), 8),
        ("LEADING",       (0, 0), (-1, -1), 11),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 6),
        ("BOX",           (0, 0), (-1, -1), 0.5, border),
        ("INNERGRID",     (0, 0), (-1, -1), 0.3, border),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        *[("BACKGROUND",  (0, i), (-1, i), off if i % 2 == 0 else white)
          for i in range(1, len(sec_tbl_data) - 1)],
        ("BACKGROUND",    (0, -1), (-1, -1), mist),
        ("FONTNAME",      (0, -1), (-1, -1), "Helvetica-Bold"),
        ("LINEABOVE",     (0, -1), (-1, -1), 1.2, navy),
    ])
    sec_tbl = Table(sec_tbl_data, colWidths=s_cols, repeatRows=1)
    sec_tbl.setStyle(sec_style)
    story.append(sec_tbl)

    # Client breakdown page
    story.append(PageBreak())
    story.append(Paragraph("Individual Client Loan Balance", title_s))
    story.append(Paragraph(
        f"File: {fname}    Generated: {now}    "
        f"Sorted by: Loan Balance (Descending)    "
        f"Grand Total: ₱{grand_lb:,.2f}", sub_s))
    story.append(Spacer(1, 0.3*cm))
    story.append(HRFlowable(width="100%", thickness=1.5, color=navy))
    story.append(Spacer(1, 0.4*cm))

    c_cols = [pg_w*0.07, pg_w*0.07, pg_w*0.16, pg_w*0.12,
              pg_w*0.09, pg_w*0.09, pg_w*0.07, pg_w*0.09,
              pg_w*0.08, pg_w*0.08, pg_w*0.09, pg_w*0.09]
    hdr_s  = ParagraphStyle("CLIHdr", parent=styles["Normal"],
                             fontSize=7, textColor=white, leading=9)
    cli_tbl_data = [[
        Paragraph("<b>Client ID</b>",      hdr_s),
        Paragraph("<b>PN</b>",             hdr_s),
        Paragraph("<b>Client</b>",         hdr_s),
        Paragraph("<b>Sector</b>",         hdr_s),
        Paragraph("<b>Principal Loan</b>", hdr_s),
        Paragraph("<b>Loan Balance</b>",   hdr_s),
        Paragraph("<b>% of Total</b>",     hdr_s),
        Paragraph("<b>Net Income</b>",     hdr_s),
        Paragraph("<b>Current Amort</b>",  hdr_s),
        Paragraph("<b>Amort History</b>",  hdr_s),
        Paragraph("<b>Total Source</b>",   hdr_s),
        Paragraph("<b>Risk</b>",           hdr_s),
    ]]
    clients_sorted = sorted(general, key=lambda r: -(r.get("loan_balance") or 0))
    for rec in clients_sorted:
        lb     = rec.get("loan_balance") or 0
        pl     = rec.get("principal_loan") or 0
        net    = rec.get("net_income") or 0
        amrt_c = rec.get("current_amort") or 0
        amrt_h = rec.get("amort_history") or 0
        total_s = rec.get("total_source") or 0
        pct     = (lb / grand_lb * 100) if grand_lb > 0 else 0.0
        rl      = rec.get("score_label", "N/A")
        industries = rec.get("industry_tags") or ([rec.get("industry")] if rec.get("industry") else [])
        ind_text = " · ".join(industries) if industries else "—"
        col_hex = "1A3A6B"
        risk_col = RISK_RL.get(rl, rl_colors.grey)
        icon    = "🏭"
        cli_tbl_data.append([
            Paragraph(rec.get("client_id", "—"), muted_s),
            Paragraph(rec.get("pn", "—"), muted_s),
            Paragraph(f"<b>{rec['client'][:30]}</b>", body_s),
            Paragraph(f"<font color='#{col_hex}'> {ind_text[:24]}</font>", body_s),
            Paragraph(f"{pl:,.2f}" if pl else "—", body_s),
            Paragraph(f"<b>{lb:,.2f}</b>", body_s),
            Paragraph(f"{pct:.2f}%", body_s),
            Paragraph(f"{net:,.2f}" if net else "—", body_s),
            Paragraph(f"{amrt_c:,.2f}" if amrt_c else "—", body_s),
            Paragraph(f"{amrt_h:,.2f}" if amrt_h else "—", body_s),
            Paragraph(f"{total_s:,.2f}" if total_s else "—", body_s),
            Paragraph(
                f"<font color='#{_rgb_hex(risk_col)}'><b>{rl}</b></font>",
                body_s),
        ])

    cli_style = TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  navy),
        ("FONTSIZE",      (0, 0), (-1, -1), 7),
        ("LEADING",       (0, 0), (-1, -1), 9),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING",   (0, 0), (-1, -1), 5),
        ("BOX",           (0, 0), (-1, -1), 0.5, border),
        ("INNERGRID",     (0, 0), (-1, -1), 0.25, border),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        *[("BACKGROUND",  (0, i), (-1, i), off if i % 2 == 0 else white)
          for i in range(1, len(cli_tbl_data))],
    ])
    cli_tbl = Table(cli_tbl_data, colWidths=c_cols, repeatRows=1)
    cli_tbl.setStyle(cli_style)
    story.append(cli_tbl)
    doc.build(story)


# ══════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════

def _loanbal_export_excel(self):
    if not self._lu_all_data:
        messagebox.showwarning("No Data", "Run an analysis first.")
        return
    if not _HAS_OPX:
        messagebox.showerror("Missing Library",
                             "openpyxl is not installed.\nRun:  pip install openpyxl")
        return
    filtered_data = _lu_get_filtered_all_data(self)
    q = getattr(self, "_loanbal_search_var", tk.StringVar(value="")).get().strip()
    filtered_data = _lu_filter_data_by_query(filtered_data, q)
    active_sectors = _lu_get_active_sectors(self) or []

    if q:
        default_name = f"LoanBalance_Filtered_Search_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    elif active_sectors:
        slug = "_".join(s.replace("/", "_").replace(" ", "_") for s in active_sectors[:2])
        default_name = f"LoanBalance_Filtered_{slug}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    else:
        default_name = f"LoanBalance_Exposure_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    path = filedialog.asksaveasfilename(
        title="Save Loan Balance Excel",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialfile=default_name)
    if not path:
        return
    try:
        _generate_loanbal_excel(filtered_data, path, filepath=self._lu_filepath or "")
        messagebox.showinfo("Export Complete", f"Excel saved to:\n{path}")
    except Exception as ex:
        messagebox.showerror("Excel Export Error", str(ex))


def _build_loanbal_export_payload_from_records(records: list[dict]) -> dict:
    """Build a minimal all_data dict for sector/client Excel export from a client subset."""
    tot_lb = sum(r.get("loan_balance") or 0 for r in records)
    tot_src = sum(r.get("total_source") or 0 for r in records)
    tot_net = sum(r.get("net_income") or 0 for r in records)
    tot_am = sum(r.get("current_amort") or 0 for r in records)
    unique_industries = sorted(
        {
            (tag or "").strip()
            for rec in records
            for tag in (
                rec.get("industry_tags")
                or ([rec.get("industry")] if rec.get("industry") else [])
            )
            if (tag or "").strip()
        },
        key=str.lower,
    )
    return {
        "general": list(records),
        "clients": {r["client"]: r for r in records if r.get("client")},
        "sector_map": {},
        "income_map": {},
        "totals": {
            "loan_balance": tot_lb,
            "total_source": tot_src,
            "total_net": tot_net,
            "current_amort": tot_am,
        },
        "unique_industries": unique_industries,
        "unique_product_names": [],
        "unique_expense_names": [],
    }


def _generate_loanbal_excel(
    all_data,
    out_path,
    filepath="",
    *,
    document_title: str | None = None,
    client_sheet_title: str | None = None,
    client_sheet_subtitle_suffix: str = "",
):
    wb  = openpyxl.Workbook()
    now = datetime.now().strftime("%B %d, %Y  %H:%M")
    fname = Path(filepath).name if filepath else "—"
    main_title = document_title or "Sector vs Loan Balance — Exposure Analysis"
    cli_title = client_sheet_title or "Client Loan Balance — Individual Breakdown"

    general    = all_data.get("general", [])
    unique_industries = all_data.get("unique_industries", [])
    industry_map = _build_industry_map(general, unique_industries)
    totals     = all_data.get("totals", {})
    grand_lb   = totals.get("loan_balance", 0) or 0

    NUM_FMT = '#,##0.00'
    RISK_FC = {"HIGH": "E53E3E", "MODERATE": "D4A017", "LOW": "2E7D32", "N/A": "9AAACE"}

    def fill(hex_str):
        return PatternFill("solid", fgColor=hex_str.lstrip("#"))

    FILLS = {
        "hdr":   fill("1A3A6B"),
        "total": fill("EEF3FB"),
        "alt":   fill("F5F7FA"),
        "white": fill("FFFFFF"),
    }

    def thin_border():
        s = Side(style="thin", color="C5D0E8")
        return Border(left=s, right=s, top=s, bottom=s)

    def thick_bottom():
        s   = Side(style="thin",   color="C5D0E8")
        bot = Side(style="medium", color="1A3A6B")
        return Border(left=s, right=s, top=s, bottom=bot)

    # ── Sheet 1: Sector Summary ───────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Sector Summary"

    ws1.merge_cells("A1:G1")
    ws1["A1"] = main_title
    ws1["A1"].font      = Font(bold=True, size=14, color="0A1628")
    ws1["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[1].height = 24

    ws1.merge_cells("A2:G2")
    ws1["A2"] = (f"File: {fname}    Generated: {now}    "
                 f"Total Clients: {len(general)}    "
                 f"Grand Total Loan Balance: ₱{grand_lb:,.2f}")
    ws1["A2"].font      = Font(size=8, color="9AAACE")
    ws1["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws1.row_dimensions[2].height = 16
    ws1.row_dimensions[3].height = 6

    SEC_HDRS   = ["Sector", "# Clients", "Total Loan Balance", "% of Total",
                  "Avg Loan / Client", "Avg Net Income", "Risk Profile"]
    SEC_WIDTHS = [28, 12, 22, 14, 22, 20, 14]
    for ci, (h, w) in enumerate(zip(SEC_HDRS, SEC_WIDTHS), 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w
        c = ws1.cell(4, ci, h)
        c.fill      = FILLS["hdr"]
        c.font      = Font(bold=True, color="FFFFFF", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()
    ws1.row_dimensions[4].height = 22

    sector_rows_data = []
    for industry, recs in industry_map.items():
        n       = len(recs)
        s_lb    = sum(r.get("loan_balance") or 0 for r in recs)
        s_net   = sum(r.get("net_income")   or 0 for r in recs)
        pct     = (s_lb / grand_lb * 100) if grand_lb > 0 else 0.0
        avg_lb  = s_lb / n  if n > 0 else 0.0
        avg_net = s_net / n if n > 0 else 0.0
        risk_label = _risk_label_from_recs(recs)
        sector_rows_data.append((industry, n, s_lb, pct, avg_lb, avg_net, risk_label))
    sector_rows_data.sort(key=lambda x: -x[2])

    data_start = 5
    for idx, (industry, n, s_lb, pct, avg_lb, avg_net, risk_label) in enumerate(sector_rows_data):
        row_num   = data_start + idx
        row_fill  = FILLS["alt"] if idx % 2 == 0 else FILLS["white"]
        icon      = "🏭"
        sec_color = "1A3A6B"
        risk_fc   = RISK_FC.get(risk_label, "9AAACE")
        values = [f"{icon}  {industry}", n, s_lb, pct / 100, avg_lb,
                  avg_net if avg_net else None, risk_label]
        fmts   = [None, "0", NUM_FMT, '0.00%', NUM_FMT, NUM_FMT, None]
        bolds  = [True, False, True, False, False, False, True]
        colors = [sec_color, "1A2B4A", "1A2B4A", "1A2B4A", "1A2B4A", "1A2B4A", risk_fc]
        for ci, (val, fmt, bold, color) in enumerate(zip(values, fmts, bolds, colors), 1):
            c = ws1.cell(row_num, ci, val)
            c.fill      = row_fill
            c.font      = Font(bold=bold, size=9, color=color)
            c.border    = thin_border()
            c.alignment = Alignment(
                horizontal="left" if ci == 1 else "center",
                vertical="center", indent=1 if ci == 1 else 0)
            if fmt:
                c.number_format = fmt
        ws1.row_dimensions[row_num].height = 18

    total_row  = data_start + len(sector_rows_data)
    avg_grand  = grand_lb / len(general) if general else 0
    grand_vals = ["GRAND TOTAL", len(general), grand_lb, 1.0, avg_grand, None, "—"]
    grand_fmts = [None, "0", NUM_FMT, '0.00%', NUM_FMT, None, None]
    for ci, (val, fmt) in enumerate(zip(grand_vals, grand_fmts), 1):
        c = ws1.cell(total_row, ci, val)
        c.fill      = FILLS["total"]
        c.font      = Font(bold=True, size=9, color="0A1628")
        c.border    = thick_bottom()
        c.alignment = Alignment(
            horizontal="left" if ci == 1 else "center",
            vertical="center", indent=1 if ci == 1 else 0)
        if fmt:
            c.number_format = fmt
    ws1.row_dimensions[total_row].height = 20
    ws1.freeze_panes = "A5"

    # ── Sheet 2: Client Breakdown ─────────────────────────────────────
    cli_cols = [
        ("Client ID", "client_id", 12, "text"),
        ("PN", "pn", 12, "text"),
        ("Applicant", "client", 28, "text"),
        ("Residence Address", "residence", 26, "text"),
        ("Office Address", "office", 26, "text"),
        ("Industry Name", "industry", 24, "text"),
        ("Spouse Info", "spouse_info", 20, "text"),
        ("Personal Assets", "personal_assets", 36, "text"),
        ("Business Assets", "business_assets", 36, "text"),
        ("Business Inventory", "business_inventory", 36, "text"),
        ("Source of Income", "source_income", 28, "text"),
        ("Total Source Of Income", "total_source", 18, "num"),
        ("Business Expenses", "biz_exp_detail", 28, "text"),
        ("Total Business Expenses", "total_biz_exp", 18, "num"),
        ("Household / Personal Expenses", "hhld_exp_detail", 30, "text"),
        ("Total Household / Personal Expenses", "total_hhld_exp", 24, "num"),
        ("Total Net Income", "net_income", 16, "num"),
        ("Total Amortization History", "amort_history", 20, "num"),
        ("Total Current Amortization", "current_amort", 20, "num"),
        ("Loan Balance", "loan_balance", 16, "num"),
        ("Total Amortized Cost", "total_amortized_cost", 18, "num"),
        ("Principal Loan", "principal_loan", 16, "num"),
        ("Maturity", "maturity", 14, "text"),
        ("Interest Rate", "interest_rate", 12, "text"),
        ("Branch", "branch", 14, "text"),
        ("Loan Class", "loan_class", 14, "text"),
        ("Product Name", "product_name", 20, "text"),
        ("Loan Date", "loan_date", 14, "text"),
        ("Term Unit", "term_unit", 12, "text"),
        ("Term", "term", 10, "text"),
        ("Security", "security", 14, "text"),
        ("Release Tag", "release_tag", 14, "text"),
        ("Loan Amount", "loan_amount", 16, "num"),
        ("Loan Status", "loan_status", 14, "text"),
        ("AO Name", "ao_name", 18, "text"),
        ("Risk Label", "score_label", 12, "text"),
        ("Risk Reasoning", "risk_reasoning", 46, "text"),
    ]

    _cli_end = get_column_letter(len(cli_cols))
    ws2 = wb.create_sheet("Client Breakdown")
    ws2.merge_cells(f"A1:{_cli_end}1")
    ws2["A1"] = cli_title
    ws2["A1"].font      = Font(bold=True, size=14, color="0A1628")
    ws2["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws2.row_dimensions[1].height = 24

    ws2.merge_cells(f"A2:{_cli_end}2")
    extra = f"    {client_sheet_subtitle_suffix}" if client_sheet_subtitle_suffix else ""
    ws2["A2"] = (f"File: {fname}    Generated: {now}    "
                 f"Total Clients: {len(general)}    "
                 f"Grand Total Loan Balance: ₱{grand_lb:,.2f}    "
                 f"Sorted by: Loan Balance (Descending){extra}")
    ws2["A2"].font      = Font(size=8, color="9AAACE")
    ws2["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws2.row_dimensions[2].height = 16
    ws2.row_dimensions[3].height = 6

    for ci, (h, _key, w, _kind) in enumerate(cli_cols, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
        c = ws2.cell(4, ci, h)
        c.fill      = FILLS["hdr"]
        c.font      = Font(bold=True, color="FFFFFF", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin_border()
    ws2.row_dimensions[4].height = 22

    clients_sorted = sorted(general, key=lambda r: -(r.get("loan_balance") or 0))
    sum_fields = {
        "total_source": 0.0,
        "total_biz_exp": 0.0,
        "total_hhld_exp": 0.0,
        "net_income": 0.0,
        "amort_history": 0.0,
        "current_amort": 0.0,
        "loan_balance": 0.0,
        "total_amortized_cost": 0.0,
        "principal_loan": 0.0,
        "loan_amount": 0.0,
    }

    for idx, rec in enumerate(clients_sorted):
        row_num = 5 + idx
        rl      = rec.get("score_label", "N/A")
        risk_fc = RISK_FC.get(rl, "9AAACE")
        row_fill = FILLS["alt"] if idx % 2 == 0 else FILLS["white"]

        row_vals = []
        for _hdr, key, _w, kind in cli_cols:
            val = rec.get(key, "")
            if kind in ("num", "pct"):
                num = _parse_numeric(val)
                val = num if num is not None else None
                if kind == "num" and key in sum_fields and num is not None:
                    sum_fields[key] += num
            row_vals.append(val)

        text_line_est = 1
        for ci, val in enumerate(row_vals, 1):
            _hdr, key, _w, kind = cli_cols[ci - 1]
            c = ws2.cell(row_num, ci, val)
            c.fill   = row_fill
            c.border = thin_border()
            if kind == "num":
                c.font = Font(bold=True, size=9, color="1A2B4A")
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = NUM_FMT
            elif kind == "pct":
                c.font = Font(size=9, color="1A2B4A")
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.number_format = '0.00%'
            elif key == "score_label":
                c.font      = Font(bold=True, size=9, color=risk_fc)
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif key in (
                "client", "residence", "office", "source_income", "biz_exp_detail", "hhld_exp_detail",
                "personal_assets", "business_assets", "business_inventory", "risk_reasoning",
            ):
                c.font      = Font(bold=True, size=9, color="1A2B4A")
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
            else:
                c.font = Font(size=9, color="6B7FA3")
                if kind == "text":
                    # Prevent Excel from visually spilling long text into adjacent columns.
                    c.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                else:
                    c.alignment = Alignment(horizontal="center", vertical="center")

            if kind == "text":
                txt = str(val or "")
                if txt:
                    width_units = max(10, int(cli_cols[ci - 1][2]))
                    approx_chars_per_line = max(12, int(width_units * 1.1))
                    est_lines = max(1, txt.count("\n") + 1, (len(txt) // approx_chars_per_line) + 1)
                    text_line_est = max(text_line_est, est_lines)

        ws2.row_dimensions[row_num].height = min(160, max(18, 8 + text_line_est * 13))

    total_row  = 5 + len(clients_sorted)
    total_vals = []
    total_fmts = []
    for _hdr, key, _w, kind in cli_cols:
        if key == "client_id":
            total_vals.append("GRAND TOTAL")
            total_fmts.append(None)
        elif kind == "num" and key in sum_fields:
            total_vals.append(sum_fields[key])
            total_fmts.append(NUM_FMT)
        elif kind == "pct" and key == "interest_rate":
            total_vals.append(None)
            total_fmts.append(None)
        else:
            total_vals.append(None)
            total_fmts.append(None)

    for ci, (val, fmt) in enumerate(zip(total_vals, total_fmts), 1):
        _hdr, key, _w, kind = cli_cols[ci - 1]
        c = ws2.cell(total_row, ci, val)
        c.fill   = FILLS["total"]
        c.font   = Font(bold=True, size=9, color="0A1628")
        c.border = thick_bottom()
        if kind in ("num", "pct"):
            c.alignment = Alignment(horizontal="right", vertical="center")
        elif key in ("client_id", "client"):
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        else:
            c.alignment = Alignment(horizontal="center", vertical="center")
        if fmt:
            c.number_format = fmt
    ws2.row_dimensions[total_row].height = 20
    ws2.freeze_panes = "A5"

    wb.save(out_path)


# ══════════════════════════════════════════════════════════════════════
#  ATTACH
# ══════════════════════════════════════════════════════════════════════

def attach(cls):
    """
    Attach Loan Balance tab methods to the app class.
    Call AFTER lu_tab_analysis.attach(cls).
    """
    cls._build_loanbal_panel         = _build_loanbal_panel
    cls._loanbal_schedule_render     = _loanbal_schedule_render
    cls._loanbal_render              = _loanbal_render
    cls._loanbal_show_export_menu    = _loanbal_show_export_menu
    cls._loanbal_export_pdf       = _loanbal_export_pdf
    cls._loanbal_export_excel     = _loanbal_export_excel
    cls._generate_loanbal_pdf     = staticmethod(_generate_loanbal_pdf)
    cls._generate_loanbal_excel   = staticmethod(_generate_loanbal_excel)
    cls._loanbal_export_btn       = None   # placeholder; set by _build_loanbal_panel