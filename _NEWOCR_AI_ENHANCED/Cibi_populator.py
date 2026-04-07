"""
cibi_populator.py — DocExtract Pro / Banco San Vicente
=======================================================

PATCH NOTES (this version)
---------------------------
  FIX-1  Unified extraction — CIC, Payslip, SALN, and ITR are now extracted
         in a SINGLE Gemini call instead of four separate ones, reducing API
         calls by 3 and saving ~15,000 tokens per run.
         Large CIC documents (> 75k chars) still use chunked extraction for
         the CIC portion only, then the remaining docs are extracted together
         in one additional call.

  FIX-2  Merged Stage 1 check — _ai_check_bank_ci() and
         _ai_check_cic_loan_tier() are replaced by a single
         _ai_check_stage1() function that evaluates both the Bank CI record
         and the CIC loan tier in ONE call, saving 1 API call and ~8,000
         tokens per CIBI workflow run.

  FIX-4  Extraction caching — results from Gemini are cached to disk in a
         .cache/ folder next to cibi_populator.py.  The cache key is an
         MD5 hash of the first 5,000 characters of the document text plus
         the doc type.  Subsequent calls with the same document return
         instantly with zero API usage.  Cache entries can be cleared by
         deleting the .cache/ folder.

  FIX-CASE  Text normaliser — address and name fields extracted from CIC
         PDFs are now title-cased via _normalise_text().  Philippine CIC
         portal text layers often have erratic mixed-case encoding
         (e.g. "SAn JOse taLiSayE") because of how the PDF is generated.
         Known abbreviations (PHL, BRGY, TIN, SSS, etc.) are preserved
         in ALL-CAPS.

  FIX-TIN  TIN written to CIBI H13 no longer includes the "TIN:" prefix
         (the template label already says "TIN:").  Empty SSS no longer
         produces a trailing " / " separator.

  FIX-MERGE  Merged cell skip logic in _write_template() corrected.
         The old for/else pattern never actually skipped non-top-left
         merged cells.  Now uses an explicit flag + continue.

  FIX-FEWSHOT  Few-shot examples are now injected into the unified prompt
         for payslip, SALN, and ITR sections.  Previously only the CIC
         section received few-shot guidance; payslip/SALN/ITR returned
         empty objects even when documents were present.  Each section
         now receives a SAMPLE→JSON pair when a matching example exists,
         telling Gemini exactly which fields to extract and in what format.

All earlier patches (P-1 through P-8, BF-1 through BF-7, etc.) are
preserved unchanged.

ACCURACY PATCH — combined (ACC-1 … ACC-10 + BF-CONTACT)
---------------------------------------------------------
  ACC-1   CIC always extracted separately
  ACC-2   APPROVAL FORM individual header cells written explicitly
  ACC-3   Text limits raised
  ACC-4   _PRESERVE_UPPER extended with PNP/AFP ranks and DepEd titles
  ACC-5   _CIC_PROMPT extended with explicit extraction-priority block
  ACC-6   Income resolution order changed to itr_monthly first
  ACC-7   _fmt_dob() normaliser added
  ACC-8   _normalise_text() changed to uppercase all non-numeric tokens
  ACC-9   _CIC_PROMPT TIN instruction strengthened
  ACC-10  _CIC_PROMPT given an explicit few-shot installment row example
  BF-CONTACT  contact_number list bug fixed

LATEST FIXES (v2 patch)
------------------------
  FIX-A through FIX-K  (see original patch notes above)

OUTPUT-FIXES (v3 patch) — fixes observed from populated xlsx output
--------------------------------------------------------------------
  OF-1  Income: net_pay is now the PRIMARY income value for H10/CASHFLOW/
        E13. itr_monthly (gross/12) was being used which gives pre-deduction
        gross instead of actual take-home. New resolution order:
          net_pay → basic_pay → itr_monthly → gross_pay → gross_cic

  OF-2  TIN formatting: _fmt_tin() normalises any raw digit string to the
        standard Philippine TIN format NNN-NNN-NNN-NNNN or NNN-NNN-NNN.

  OF-3  Address fallback to itr.registered_address then saln.office_address.

  OF-4  DOB / Age fallback from itr.date_of_birth when CIC is absent.

  OF-5  LANDBANK balance: phantom empty inst_active rows no longer consume
        credit slots before SALN liabilities are merged.

  OF-6  SSS display: no trailing " / " when SSS is empty.

OUTPUT-FIXES (v4 patch) — second round of xlsx corrections
-----------------------------------------------------------
  OF-7  Real property current value: added all SALN key variants to the
        lookup — current_fair_market_value, current_fair_value,
        current_value, assessed_value (in priority order).

  OF-8  Real property location: added exact_location fallback key.
        Also writes year_acquired and acquisition_cost.

  OF-9  Cash on hand vs combined: only use cash_combined when separate
        on-hand and in-bank fields are both absent.

  OF-10 Spouse occupation fallback to saln.spouse_position.

  OF-11 Added H31/H32/H33 to formula override whitelist.

OUTPUT-FIXES (v5 patch) — template merge-aware corrections
-----------------------------------------------------------
  OF-12 Row 33 fully merged (A33:I33) — all fields now embedded as a
        formatted string in A33. Previously balance was silently dropped.

  OF-13 Real property column remapping for merged cells:
          year_acquired   → appended to D{r} (top-left of D:E merge)
          acquisition_cost → F{r} (top-left of F:G merge)
          current_value   → H{r} (top-left of H:I merge, writable ✅)

  OF-14 SALN prompt: rules S7–S9 force separate cash fields, exact key
        names for real property, and spouse_position extraction.

  OF-15 Cache version bumped to v6 — forces re-extraction on first run.

OUTPUT-FIXES (v6 patch) — real property description
----------------------------------------------------
  OF-16 Real property description blank: Gemini sometimes returns the
        property description under "kind", "property_type", or omits the
        "description" key entirely. Added fallback chain:
          description → property_description → property_type → kind + " Property"
        Also fixed a double-write bug where D{r} was written with loc then
        immediately overwritten with loc+(yr), causing no regression but
        wasted a write. Now uses a single if/elif chain.
        Also added "address" as a fallback key for location.
        Prompt rule S10 added: explicitly maps SALN column names to JSON keys.

CIC-KW PATCH (v8) — keyword-only CIC extraction
------------------------------------------------
  CIC-KW-1  CIC extractor now ONLY looks for accounts flagged as "PAST DUE"
            or "WRITE OFF". All other CIC fields (personal info, employment,
            spouse, etc.) are NO LONGER extracted.

  CIC-KW-2  Pre-filter helper _cic_keyword_filter() scans raw CIC text
            before sending to Gemini and returns only the lines containing
            target keywords, reducing token usage.

  CIC-KW-3  Chunked extractor updated to use keyword-only prompt.

  CIC-KW-4  Cache version bumped to v8 — forces re-extraction.

TOKEN LIMITS TIER-1 PATCH (v9) — thinking budgets & larger output tokens
-------------------------------------------------------------------------
  TL-1  CIC keyword extraction — thinking DISABLED (thinking_budget=0).
        Keyword matching is not a reasoning task. Thinking tokens were
        eating into max_output_tokens and causing MAX_TOKENS truncation,
        empty JSON responses, and 2-minute hangs.

        max_output_tokens raised to 8,192 — more than enough for the
        small keyword-only JSON output (typically < 800 tokens).

  TL-2  UNIFIED extraction (payslip + SALN + ITR) — thinking budget
        set to 1,024. These three documents together require moderate
        multi-document reasoning. 1,024 thinking tokens is sufficient
        without wasteful over-spending.

        max_output_tokens raised to 24,576 — handles large SALN tables,
        multiple payslip periods, and ITR fields comfortably within
        gemini-2.5-flash's 65,535 output token ceiling.

  TL-3  Individual fallback extractors (payslip, saln, itr) — thinking
        budget set to 512 each. Simple single-document extraction.
        max_output_tokens set to 16,000 each (was 8,192 default).

  TL-4  _gemini_extract_json() patched to accept an optional
        thinking_budget parameter (default=0 — OFF) so every call site
        can set its own budget without changing the shared helper.

  TL-5  Text slice caps tightened per call:
          CIC keyword  → filtered_text[:20_000]   (post-filter, ~200-400 lines)
          UNIFIED      → unchanged (each doc already sliced inside the caller)
          Individual   → unchanged

  TL-6  Fallback model updated: gemini-2.0-flash is deprecated as of
        March 2026. Fallback chain is now:
          gemini-2.5-flash  →  gemini-2.5-flash-lite
        gemini-2.5-flash-lite does NOT support thinking_budget, so the
        patch automatically omits ThinkingConfig when falling back.

  TL-7  Cache version bumped to v9 — forces re-extraction on first run.

SUMMARY TABLE — token budgets per call
---------------------------------------
  Call                 max_output_tokens  thinking_budget  Notes
  ───────────────────  ─────────────────  ───────────────  ─────────────────
  CIC keyword          8,192              0 (OFF)          Fast, no reasoning
  UNIFIED              24,576             1,024            Multi-doc moderate
  Payslip (fallback)   16,000             512              Single doc
  SALN (fallback)      16,000             512              Single doc
  ITR (fallback)       16,000             512              Single doc
  JSON repair          4,096              0 (OFF)          Pure formatting fix
"""

from __future__ import annotations

import hashlib
import json
import logging
import re
import os
from datetime import datetime
from pathlib import Path
from typing import Callable, Optional

# ── Few-shot example loader (Phase 1) — RAG-ready interface ──────────────────
try:
    from few_shot import get_few_shot_example, FewShotExample, inject_few_shot_into_prompt
    _FEW_SHOT_AVAILABLE = True
except ImportError:
    _FEW_SHOT_AVAILABLE = False
    def get_few_shot_example(*a, **kw):          # type: ignore[misc]
        return None
    def inject_few_shot_into_prompt(prompt, ex): # type: ignore[misc]
        return prompt

# ════════════════════════════════════════════════════════════════════════════
#  FIX-G: Cache version token — bump this any time a prompt changes
# ════════════════════════════════════════════════════════════════════════════
_CACHE_VERSION = "v9"   # bumped from v8 → v9 for token limits patch

# ─────────────────────────────────────────────────────────────────────────────
#  DEBUG LOGGER
# ─────────────────────────────────────────────────────────────────────────────

def _setup_logger(log_path: Path) -> logging.Logger:
    logger = logging.getLogger("cibi_populator")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()

    fh = logging.FileHandler(str(log_path), encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    logger.addHandler(fh)

    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter("[CIBI] %(message)s"))
    logger.addHandler(ch)

    return logger


_log: logging.Logger = logging.getLogger("cibi_populator")


# ─────────────────────────────────────────────────────────────────────────────
#  OUTPUT DIRECTORY
# ─────────────────────────────────────────────────────────────────────────────

def _output_dir() -> Path:
    desktop = Path.home() / "Desktop"
    base    = desktop if desktop.exists() else Path.home()
    folder  = base / "DocExtract_Files"
    folder.mkdir(parents=True, exist_ok=True)
    return folder


# ─────────────────────────────────────────────────────────────────────────────
#  FIX-4 — DISK CACHE (with FIX-G version token)
# ─────────────────────────────────────────────────────────────────────────────

def _cache_dir() -> Path:
    folder = Path(__file__).resolve().parent / ".cache"
    folder.mkdir(parents=True, exist_ok=True)
    return folder


def _cache_key(text: str, doc_type: str) -> str:
    raw = f"{_CACHE_VERSION}::{doc_type}::{text[:5000]}"
    return hashlib.md5(raw.encode("utf-8")).hexdigest()[:12]


def _cache_get(text: str, doc_type: str) -> dict | None:
    path = _cache_dir() / f"{doc_type}_{_cache_key(text, doc_type)}.json"
    if not path.exists():
        return None
    try:
        result = json.loads(path.read_text(encoding="utf-8"))
        _log.info(f"Cache HIT  [{doc_type}] — skipping Gemini call.")
        return result
    except Exception as e:
        _log.warning(f"Cache read error [{doc_type}]: {e}")
        return None


def _cache_set(text: str, doc_type: str, result: dict) -> None:
    if result.get("_error"):
        return
    path = _cache_dir() / f"{doc_type}_{_cache_key(text, doc_type)}.json"
    try:
        path.write_text(json.dumps(result, indent=2, ensure_ascii=False),
                        encoding="utf-8")
        _log.info(f"Cache WRITE [{doc_type}] → {path.name}")
    except Exception as e:
        _log.warning(f"Cache write error [{doc_type}]: {e}")


def cache_clear() -> int:
    deleted = 0
    for f in _cache_dir().glob("*.json"):
        try:
            f.unlink()
            deleted += 1
        except Exception:
            pass
    _log.info(f"Cache cleared — {deleted} file(s) deleted.")
    return deleted


# ─────────────────────────────────────────────────────────────────────────────
#  QUOTA NOTIFICATION SYSTEM
# ─────────────────────────────────────────────────────────────────────────────

_quota_warnings: list[str] = []
_progress_cb_ref: Optional[Callable] = None


def _notify(pct: int, msg: str) -> None:
    if _progress_cb_ref:
        try:
            _progress_cb_ref(pct, msg)
        except Exception:
            pass


# ─────────────────────────────────────────────────────────────────────────────
#  TL-4 — PATCHED GEMINI HELPER (with thinking_budget and updated model chain)
# ─────────────────────────────────────────────────────────────────────────────

# TL-6: Updated model fallback chain (gemini-2.0-flash deprecated)
_MODELS = ("gemini-2.5-flash", "gemini-2.5-flash-lite")
_THINKING_SUPPORTED_MODELS: frozenset[str] = frozenset({"gemini-2.5-flash"})


def _gemini_extract_json(
    prompt:           str,
    api_key:          str,
    label:            str  = "",
    max_tokens:       int  = 8192,
    few_shot_example        = None,
    thinking_budget:  int  = 0,          # TL-4: NEW — default OFF
) -> dict:
    """
    Patched version of _gemini_extract_json with explicit thinking_budget
    control and updated fallback model chain.
    """
    try:
        from google import genai as _genai
        from google.genai import types as _gtypes
    except ImportError:
        return {"_error": "google-genai not installed — run: pip install google-genai"}

    client = _genai.Client(api_key=api_key)

    def _safe_text(resp) -> str:
        try:
            return resp.text or ""
        except Exception:
            try:
                return "".join(
                    p.text for p in resp.candidates[0].content.parts
                    if hasattr(p, "text") and p.text
                )
            except Exception:
                return ""

    raw  = ""
    used_model = ""
    primary_quota_hit = False

    for model in _MODELS:
        try:
            content_parts = []
            if few_shot_example is not None:
                try:
                    import base64 as _b64
                    raw_bytes = _b64.b64decode(few_shot_example.base64_data)
                    content_parts.append(
                        _gtypes.Part(
                            inline_data=_gtypes.Blob(
                                mime_type=few_shot_example.mime_type,
                                data=raw_bytes,
                            )
                        )
                    )
                    ref_text = (
                        "The document above is a SAMPLE. "
                        "Here is the CORRECT JSON extraction for it:\n\n"
                        + __import__("json").dumps(
                            few_shot_example.approved_json, indent=2,
                            ensure_ascii=False)
                        + "\n\nNow extract the same fields from the NEW "
                          "document below, following the same JSON schema "
                          "and value formats exactly.\n\n"
                    )
                    content_parts.append(_gtypes.Part(text=ref_text))
                    _log.info(f"{label} — few-shot injected: "
                              f"{few_shot_example.source_path.name}")
                except Exception as _fse:
                    _log.warning(f"{label} — few-shot injection failed "
                                 f"({_fse}), using text fallback.")
                    prompt = inject_few_shot_into_prompt(prompt, few_shot_example)

            content_parts.append(_gtypes.Part(text=prompt))

            # TL-1/TL-2/TL-3: Build GenerateContentConfig
            # Only inject ThinkingConfig for models that support it.
            # gemini-2.5-flash-lite does NOT support ThinkingConfig.
            if model in _THINKING_SUPPORTED_MODELS:
                config = _gtypes.GenerateContentConfig(
                    max_output_tokens=max_tokens,
                    temperature=0.0,
                    thinking_config=_gtypes.ThinkingConfig(
                        thinking_budget=thinking_budget,
                    ),
                )
                _log.debug(
                    f"{label} — {model}  max_output_tokens={max_tokens}  "
                    f"thinking_budget={thinking_budget}"
                )
            else:
                # Fallback model: no ThinkingConfig, cap output tokens
                fallback_max = min(max_tokens, 8192)
                config = _gtypes.GenerateContentConfig(
                    max_output_tokens=fallback_max,
                    temperature=0.0,
                )
                _log.debug(
                    f"{label} — {model} (no thinking)  "
                    f"max_output_tokens={fallback_max}"
                )

            resp = client.models.generate_content(
                model    = model,
                contents = [_gtypes.Content(role="user", parts=content_parts)],
                config   = config,
            )
            raw        = _safe_text(resp)
            used_model = model

            if model != _MODELS[0] and primary_quota_hit:
                warn_msg = (
                    f"⚠ [{label}] {_MODELS[0]} quota reached — "
                    f"switched to {model}."
                )
                _log.warning(warn_msg)
                _quota_warnings.append(warn_msg)
                _notify(-1, warn_msg)

            break

        except Exception as e:
            err_str = str(e).lower()
            if any(kw in err_str for kw in
                   ("quota", "rate", "429", "resource_exhausted")):
                if model == _MODELS[0]:
                    primary_quota_hit = True
                    _log.warning(
                        f"{label} — {_MODELS[0]} quota hit, "
                        f"trying {_MODELS[1]}…"
                    )
                else:
                    crit_msg = (
                        f"🚫 [{label}] ALL Gemini models quota exhausted. "
                        f"This extraction will be empty."
                    )
                    _log.error(crit_msg)
                    _quota_warnings.append(crit_msg)
                    _notify(-1, crit_msg)
                continue
            _log.error(f"{label} — Gemini API error ({model}): {e}")
            return {"_error": str(e)}

    if not raw:
        _log.error(f"{label} — All models returned empty response.")
        all_fail = (
            f"🚫 [{label}] No response from any Gemini model. "
            f"Check API key and quota."
        )
        _quota_warnings.append(all_fail)
        _notify(-1, all_fail)
        return {"_error": "All Gemini models returned empty response."}

    _log.debug(f"{label} — Model: {used_model}")
    _log.debug(f"{label} — Raw ({len(raw)} chars):\n{raw[:3000]}")

    cleaned = re.sub(r"^```(?:json)?\s*", "", raw.strip(), flags=re.I)
    cleaned = re.sub(r"\s*```$", "", cleaned.strip())
    try:
        result = json.loads(cleaned)
        _log.info(f"{label} — JSON parsed OK ({len(result)} keys).")
        return result
    except json.JSONDecodeError:
        pass

    m = re.search(r"\{[\s\S]+\}", cleaned)
    if m:
        try:
            result = json.loads(m.group(0))
            _log.info(f"{label} — JSON via regex ({len(result)} keys).")
            return result
        except Exception:
            pass

    _log.warning(f"{label} — JSON parse failed, asking Gemini to repair…")
    repair_prompt = (
        "The following text is supposed to be a valid JSON object but may "
        "have syntax errors. Return ONLY the corrected, complete JSON object "
        "with no explanation, no markdown, no extra text.\n\n"
        f"{raw[:6000]}"
    )
    try:
        from google import genai as _genai2
        from google.genai import types as _gtypes2
        client2 = _genai2.Client(api_key=api_key)
        # TL-1: Repair call — thinking OFF, small token budget
        repair_config = _gtypes2.GenerateContentConfig(
            max_output_tokens=4096,
            temperature=0.0,
            thinking_config=_gtypes2.ThinkingConfig(thinking_budget=0),
        )
        resp2 = client2.models.generate_content(
            model    = used_model if used_model in _THINKING_SUPPORTED_MODELS
                       else _MODELS[0],
            contents = [_gtypes2.Content(
                role="user",
                parts=[_gtypes2.Part(text=repair_prompt)])],
            config   = repair_config,
        )
        raw2 = _safe_text(resp2)
        raw2 = re.sub(r"^```(?:json)?\s*", "", raw2.strip(), flags=re.I)
        raw2 = re.sub(r"\s*```$", "", raw2.strip())
        result = json.loads(raw2)
        _log.info(f"{label} — JSON repaired OK.")
        return result
    except Exception as e2:
        _log.error(f"{label} — JSON repair also failed: {e2}")

    return {"_error": f"JSON parse failed. Raw: {raw[:300]}"}


# ─────────────────────────────────────────────────────────────────────────────
#  EXTRACTION PROMPTS (payslip, saln, itr, unified) — unchanged
# ─────────────────────────────────────────────────────────────────────────────

_PAYSLIP_PROMPT = """
You are extracting structured data from a Philippine payslip or payroll document.
This document may contain MULTIPLE PAY PERIODS. Read the ENTIRE document carefully.
Return ONLY a valid JSON object — no explanation, no markdown, no extra text.
Use null for any field not found. Numbers only (no ₱, no commas).

CRITICAL RULES:
  1. Extract EVERY field present. Never return null for a field that exists.
  2. If the document contains MORE THAN ONE pay period:
       a. Fill "pay_periods" with one entry per period.
       b. Set "period_count" to the total number found.
       c. For ALL income/deduction scalar fields, return the AVERAGE across all periods.
  3. If only ONE period, set "period_count" to 1.
  4. net_pay = gross_pay minus total_deductions. Extract the ACTUAL net pay figure shown.
  5. date_of_birth — extract if shown on the payslip (some DepEd payslips show it).

{
  "employee_name": null, "employer_name": null, "position": null, "department": null,
  "period_from": null, "period_to": null, "pay_date": null, "period_count": null,
  "date_of_birth": null,
  "pay_periods": [
    {
      "period_from": null, "period_to": null, "pay_date": null,
      "basic_pay": null, "allowances": null, "gross_pay": null,
      "gsis_deduction": null, "sss_deduction": null, "philhealth_deduction": null,
      "pagibig_deduction": null, "tax_deduction": null, "other_deductions": null,
      "total_deductions": null, "net_pay": null
    }
  ],
  "monthly_rate": null, "basic_pay": null, "allowances": null, "gross_pay": null,
  "gsis_deduction": null, "sss_deduction": null, "philhealth_deduction": null,
  "pagibig_deduction": null, "tax_deduction": null, "other_deductions": null,
  "total_deductions": null, "net_pay": null,
  "tin": null, "sss_number": null, "philhealth_number": null, "pagibig_number": null,
  "business_income": null, "rental_income": null, "remittance_income": null,
  "other_income_label": null, "other_income_amount": null
}

--- PAYSLIP DOCUMENT ---
{text}
--- END ---
"""

_SALN_PROMPT = """
You are extracting structured data from a Philippine SALN
(Statement of Assets, Liabilities and Net Worth).
Return ONLY a valid JSON object — no explanation, no markdown, no extra text.
Use null for any field not found. Return [] for empty lists.
Numbers only (no ₱, no commas).

CRITICAL EXTRACTION RULES:
  1. Extract EVERY non-N/A row. Never stop after the first item.
  2. personal_properties — read the ENTIRE table top to bottom.
  3. Set current_value = acquisition_cost for every item (no separate current value column in SALN).
  4. real_properties — if all rows say "N/A", return [].
  5. liabilities — extract EVERY row: nature of loan, name of creditor, outstanding balance.
  6. children — name and age for every child listed. Return null for date_of_birth.
  7. net_worth may be negative. Return as a number.
  8. position — copy EXACTLY as printed.
  9. CASH FIELDS — always extract as TWO SEPARATE fields, never combine them:
       cash_on_hand  = the "Cash on Hand" amount only
       cash_in_bank  = the "Cash in Bank" amount only
       cash_on_hand_and_in_bank = null (leave null if separate fields exist)
  10. real_properties — use these EXACT key names:
       description, kind, location, area, assessed_value,
       current_fair_market_value (the "CURRENT FAIR MARKET VALUE" column),
       year_acquired, mode_of_acquisition, acquisition_cost
  11. spouse_position — extract the spouse's position/occupation from the SPOUSE section.
  12. real_properties column → JSON key mapping (use EXACTLY these key names):
        "DESCRIPTION" column            → "description"  (e.g. "Residential House and Lot")
        "KIND" column                   → "kind"          (e.g. "Residential")
        "EXACT LOCATION" column         → "location"
        "ASSESSED VALUE" column         → "assessed_value"
        "CURRENT FAIR MARKET VALUE" col → "current_fair_market_value"
        "ACQUISITION YEAR" column       → "year_acquired"
        "MODE OF ACQUISITION" column    → "mode_of_acquisition"
        "ACQUISITION COST" column       → "acquisition_cost"

{{
  "declarant_name": null,
  "position": null,
  "agency": null,
  "office_address": null,
  "saln_year": null,
  "spouse_name": null,
  "spouse_position": null,
  "real_properties": [],
  "personal_properties": [
    {{"description":null,"year_acquired":null,"acquisition_cost":null,"current_value":null}}
  ],
  "cash_on_hand": null,
  "cash_in_bank": null,
  "cash_on_hand_and_in_bank": null,
  "receivables": null,
  "business_interests": null,
  "total_assets": null,
  "liabilities": [
    {{"nature":null,"creditor":null,"outstanding_balance":null}}
  ],
  "financial_liabilities": null,
  "personal_liabilities": null,
  "total_liabilities": null,
  "net_worth": null,
  "children": [
    {{"name":null,"date_of_birth":null,"age":null}}
  ]
}}

--- SALN DOCUMENT ---
{text}
--- END ---
"""

_ITR_PROMPT = """
You are extracting structured data from a Philippine ITR (Income Tax Return) or
BIR Form 2316 (Certificate of Compensation Payment / Tax Withheld).
Return ONLY a valid JSON object — no explanation, no markdown, no extra text.
Use null for any field not found. Numbers only (no ₱, no commas).

CRITICAL:
  1. Extract EVERY field present.
  2. date_of_birth — extract from Part I Employee Information if present (MM/DD/YYYY).
  3. registered_address — extract the employee's home address from Part I.
  4. gross_compensation_income — the total gross compensation before deductions.
  5. net_pay — if shown, the actual net take-home pay after all deductions.

{
  "taxpayer_name": null, "tin": null, "tax_year": null, "form_type": null,
  "registered_address": null, "zip_code": null, "taxpayer_type": null,
  "civil_status": null, "citizenship": null,
  "date_of_birth": null,
  "business_name": null, "business_address": null, "business_tin": null,
  "line_of_business": null,
  "gross_compensation_income": null, "gross_business_income": null,
  "gross_professional_income": null, "total_gross_income": null,
  "gross_annual_income": null, "gross_monthly_income": null,
  "net_taxable_income": null, "net_pay": null,
  "allowable_deductions": null,
  "tax_due": null, "tax_credits": null, "tax_paid": null,
  "tax_still_due": null, "surcharge": null, "interest": null,
  "compromise": null, "total_amount_payable": null,
  "spouse_name": null, "spouse_tin": null
}

--- ITR DOCUMENT ---
{text}
--- END ---
"""

_UNIFIED_PROMPT = """
You are extracting structured data from multiple Philippine loan documents.
Return ONLY a single valid JSON object with exactly four top-level keys:
"cic", "payslip", "saln", "itr".
Use null for any field not found. Numbers only (no ₱, no commas).
Return [] for empty lists.

CRITICAL RULES — apply to ALL sections:
  1. Extract EVERY field present. Never return null for a field that exists.
  2. For the payslip — if multiple pay periods exist, return AVERAGES for scalar fields
     and list each period in pay_periods[].
  3. net_pay in payslip = actual take-home after ALL deductions. Extract it precisely.
  4. For CIC installment tables — extract ALL rows from both tables.
  5. If a document section is marked "NOT PROVIDED", return an empty object {{}}
     for that key.

═══════════════════════════════════════════════════════
SALN-SPECIFIC RULES:
═══════════════════════════════════════════════════════
  S1. personal_properties — extract EVERY row. Never stop after the first item.
  S2. No separate current value column in SALN — set current_value = acquisition_cost.
  S3. liabilities — extract EVERY row: nature, creditor, outstanding balance.
  S4. children — name and age only; return null for date_of_birth.
  S5. net_worth may be negative.
  S6. position — copy EXACTLY as printed.
  S7. CASH — always extract as TWO SEPARATE fields:
        cash_on_hand = "Cash on Hand" amount only
        cash_in_bank = "Cash in Bank" amount only
        Leave cash_on_hand_and_in_bank as null when separate fields exist.
  S8. real_properties — use these EXACT field names:
        current_fair_market_value (from the "CURRENT FAIR MARKET VALUE" column)
        acquisition_cost, year_acquired, location, assessed_value
  S9. spouse_position — extract spouse's occupation/position from the SPOUSE section.
  S10. real_properties exact key names:
        "description" = DESCRIPTION column (e.g. "Residential House and Lot")
        "kind"        = KIND column (e.g. "Residential")
        "location"    = EXACT LOCATION column
        "current_fair_market_value" = CURRENT FAIR MARKET VALUE column
        "acquisition_cost" = ACQUISITION COST column
        "year_acquired"    = ACQUISITION YEAR column

═══════════════════════════════════════════════════════
ITR / BIR 2316-SPECIFIC RULES:
═══════════════════════════════════════════════════════
  I1. date_of_birth — extract from Part I Employee Information (MM/DD/YYYY).
  I2. registered_address — employee home address from Part I.
  I3. gross_compensation_income — total gross before any deductions.
  I4. net_pay — actual take-home pay after all deductions if shown.
  I5. tax_withheld / tax_paid — the total taxes withheld from compensation.

═══════════════════════════════════════════════════════
Return exactly this JSON structure:
═══════════════════════════════════════════════════════
{{
  "cic": {{
    "full_name": null, "first_name": null, "middle_name": null,
    "last_name": null, "suffix": null, "date_of_birth": null,
    "age": null, "gender": null, "civil_status": null,
    "nationality": null, "number_of_dependents": null,
    "tin": null, "sss": null, "drivers_license": null,
    "residence_address": null, "mailing_address": null,
    "contact_number": null,
    "spouse_first_name": null, "spouse_middle_name": null,
    "spouse_last_name": null, "spouse_age": null,
    "spouse_occupation": null, "spouse_income": null,
    "spouse_employment_status": null, "spouse_hired_from": null,
    "spouse_tin": null,
    "employer_name": null, "employer_address": null,
    "employer_tin": null, "employer_contact": null,
    "occupation": null, "employment_status": null,
    "gross_income": null, "income_frequency": null,
    "hired_from": null, "hired_to": null,
    "sole_trader_name": null, "sole_trader_address": null,
    "credit_accounts": [],
    "installments_requested": [],
    "installments_active": [],
    "total_monthly_amortization": null,
    "total_loan_balance": null,
    "total_overdue_payments": null
  }},
  "payslip": {{
    "employee_name": null, "employer_name": null,
    "position": null, "department": null,
    "period_from": null, "period_to": null,
    "pay_date": null, "period_count": null,
    "date_of_birth": null,
    "pay_periods": [],
    "monthly_rate": null, "basic_pay": null,
    "allowances": null, "gross_pay": null,
    "gsis_deduction": null, "sss_deduction": null,
    "philhealth_deduction": null, "pagibig_deduction": null,
    "tax_deduction": null, "other_deductions": null,
    "total_deductions": null, "net_pay": null,
    "tin": null, "sss_number": null,
    "philhealth_number": null, "pagibig_number": null,
    "business_income": null, "rental_income": null,
    "remittance_income": null,
    "other_income_label": null, "other_income_amount": null
  }},
  "saln": {{
    "declarant_name": null,
    "position": null,
    "agency": null,
    "office_address": null,
    "saln_year": null,
    "spouse_name": null,
    "spouse_position": null,
    "real_properties": [],
    "personal_properties": [
      {{
        "description": null, "year_acquired": null,
        "acquisition_cost": null, "current_value": null
      }}
    ],
    "cash_on_hand": null,
    "cash_in_bank": null,
    "cash_on_hand_and_in_bank": null,
    "receivables": null,
    "business_interests": null,
    "total_assets": null,
    "liabilities": [
      {{
        "nature": null, "creditor": null, "outstanding_balance": null
      }}
    ],
    "financial_liabilities": null,
    "personal_liabilities": null,
    "total_liabilities": null,
    "net_worth": null,
    "children": [
      {{ "name": null, "date_of_birth": null, "age": null }}
    ]
  }},
  "itr": {{
    "taxpayer_name": null, "tin": null, "tax_year": null,
    "form_type": null, "registered_address": null,
    "zip_code": null, "taxpayer_type": null,
    "civil_status": null, "citizenship": null,
    "date_of_birth": null,
    "business_name": null, "business_address": null,
    "business_tin": null, "line_of_business": null,
    "gross_compensation_income": null, "gross_business_income": null,
    "gross_professional_income": null, "total_gross_income": null,
    "gross_annual_income": null, "gross_monthly_income": null,
    "net_taxable_income": null, "net_pay": null,
    "allowable_deductions": null,
    "tax_due": null, "tax_credits": null, "tax_paid": null,
    "tax_still_due": null, "surcharge": null, "interest": null,
    "compromise": null, "total_amount_payable": null,
    "spouse_name": null, "spouse_tin": null
  }}
}}

=== CIC DOCUMENT ===
{cic_text}
=== END CIC ===

=== PAYSLIP DOCUMENT ===
{payslip_text}
=== END PAYSLIP ===

=== SALN DOCUMENT ===
{saln_text}
=== END SALN ===

=== ITR DOCUMENT ===
{itr_text}
=== END ITR ===
"""


# ─────────────────────────────────────────────────────────────────────────────
#  HELPERS (unchanged)
# ─────────────────────────────────────────────────────────────────────────────

def _v(d: dict, *keys, default=None):
    for k in keys:
        val = d.get(k)
        if val is None:
            continue
        s = str(val).strip()
        if s in ("", "null", "None"):
            continue
        return val
    return default


def _num(val) -> Optional[float]:
    if val is None:
        return None
    try:
        cleaned = str(val).replace("₱","").replace(",","").replace(" ","").strip()
        if not cleaned or cleaned.lower() in ("null","none"):
            return None
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def _peso(val) -> str:
    n = _num(val)
    return f"₱{n:,.2f}" if n is not None else ""


def _age_from_dob(dob_str: str) -> str:
    if not dob_str:
        return ""
    for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y",
                "%B %d, %Y", "%b %d, %Y", "%d %B %Y"):
        try:
            dob = datetime.strptime(str(dob_str).strip(), fmt)
            return str((datetime.today() - dob).days // 365)
        except ValueError:
            continue
    return ""


def _fmt_dob(dob_str: str) -> str:
    if not dob_str:
        return dob_str
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%d-%m-%Y",
                "%B %d, %Y", "%b %d, %Y", "%d %B %Y"):
        try:
            return datetime.strptime(str(dob_str).strip(), fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return str(dob_str)


def _fmt_tin(tin: str) -> str:
    if not tin:
        return ""
    digits = "".join(c for c in str(tin) if c.isdigit())
    if not digits:
        return str(tin)
    if len(digits) >= 12:
        d = digits[:12]
        return f"{d[:3]}-{d[3:6]}-{d[6:9]}-{d[9:12]}"
    elif len(digits) == 9:
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:9]}"
    elif len(digits) > 9:
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:9]}-{digits[9:]}"
    return str(tin)


def _years_since(date_str: str) -> str:
    if not date_str:
        return ""
    for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y",
                "%B %d, %Y", "%b %d, %Y", "%d %B %Y"):
        try:
            d = datetime.strptime(str(date_str).strip(), fmt)
            return str((datetime.today() - d).days // 365)
        except ValueError:
            continue
    return ""


def _join_name(d: dict, *keys) -> str:
    parts = [str(d.get(k) or "").strip() for k in keys]
    return " ".join(p for p in parts if p and p not in ("null", "None"))


def _normalise_sheet_name(name: str) -> str:
    return name.replace("\xa0", " ").strip().upper()


_PRESERVE_UPPER: frozenset[str] = frozenset({
    "PHL", "PH",
    "BRGY", "BGY", "RD", "STS", "AVE", "BLVD", "EXT", "COR",
    "BLDG", "FLR",
    "TIN", "SSS", "GSIS", "BIR", "NBI", "LTO",
    "JR", "SR", "II", "III", "IV", "ESQ",
    "NCR", "BARMM", "CAR",
    "PSSg", "SSg", "Cpl", "PFC",
    "PO1", "PO2", "PO3",
    "SPO1", "SPO2", "SPO3", "SPO4",
    "PINSP", "PSINSP", "PCINSP", "PSSUPT", "PSUPT",
    "PCSUPT", "PDIR", "PDIRG",
    "2LT", "1LT", "CPT", "MAJ", "LTC", "COL", "BGEN", "MGEN", "LTGEN", "GEN",
    "PNP", "AFP", "PA", "PN", "PAF", "PCG",
    "MT-I", "MT-II", "HT-I", "HT-II", "HT-III",
    "ESP-I", "ESP-II", "ESP-III",
    "T-I", "T-II", "T-III",
})

def _normalise_text(text: str | None) -> str:
    if not text:
        return text or ""
    words = str(text).split()
    result = []
    for w in words:
        upper_w = w.upper()
        if upper_w in _PRESERVE_UPPER:
            canonical = next((s for s in _PRESERVE_UPPER if s.upper() == upper_w), upper_w)
            result.append(canonical)
        elif re.match(r"^\d", w):
            result.append(w)
        else:
            result.append(upper_w)
    return " ".join(result)


def _get_dob(cic: dict, pay: dict, itr: dict) -> str:
    return (
        _v(cic, "date_of_birth") or
        _v(pay, "date_of_birth") or
        _v(itr, "date_of_birth") or
        ""
    )


def _strip_label_prefix(value: str, *labels: str) -> str:
    for lbl in labels:
        value = re.sub(
            rf"(?i)^{re.escape(lbl)}\s*[:\-\.]?\s*", "", str(value)
        ).strip()
    return value


_PNP_AFP_RANKS: frozenset[str] = frozenset({
    "PO1","PO2","PO3",
    "SPO1","SPO2","SPO3","SPO4",
    "PSSg","SSg","Cpl","PFC",
    "PINSP","PSINSP","PCINSP","PSSUPT","PSUPT","PCSUPT","PDIR","PDIRG",
    "2LT","1LT","CPT","MAJ","LTC","COL","BGEN","MGEN","LTGEN","GEN",
    "PNP","AFP","PA","PN","PAF","PCG",
    "MT-I","MT-II","HT-I","HT-II","HT-III",
    "T-I","T-II","T-III",
})

_GOVT_KEYWORDS: tuple[str, ...] = (
    "police","pnp","afp","army","navy","air force",
    "sergeant","corporal","officer","inspector",
    "teacher","nurse","engineer","deped","doh",
    "government","civil service","municipal","provincial",
    "national","bureau","department",
)

def _is_employed(employer: str, employment_status: str, occupation: str) -> bool:
    if employer or employment_status:
        return True
    occ_upper = str(occupation or "").upper()
    occ_lower = str(occupation or "").lower()
    for token in occ_upper.split():
        if token.rstrip(".,;") in {r.upper() for r in _PNP_AFP_RANKS}:
            return True
    return any(kw in occ_lower for kw in _GOVT_KEYWORDS)


def _write_children(ci: dict, saln: dict, cic: dict,
                    max_children: int = 4) -> None:
    cic_dob_map: dict[str, str] = {}
    for child in (cic.get("children") or cic.get("dependents") or []):
        if not isinstance(child, dict):
            continue
        name = str(child.get("name") or "").strip().lower()
        dob  = child.get("date_of_birth") or child.get("dob") or ""
        if name and dob:
            cic_dob_map[name] = str(dob)

    children = saln.get("children") or []
    for i, child in enumerate(children[:max_children]):
        r     = 24 + i
        name  = str(child.get("name") or "").strip()
        dob_c = (
            child.get("date_of_birth") or
            cic_dob_map.get(name.lower(), "") or ""
        )
        age_c = str(child.get("age") or _age_from_dob(str(dob_c)) or "")

        if name:  ci[f"A{r}"] = name
        if dob_c: ci[f"D{r}"] = str(dob_c)
        if age_c: ci[f"G{r}"] = age_c


def _write_household_assets(ci: dict, household: list,
                             max_props: int = 5) -> None:
    for i, prop in enumerate(household[:max_props]):
        r  = 65 + i
        d  = str(prop.get("description") or "").strip()
        ya = str(prop.get("year_acquired") or "").strip()
        ac = _num(prop.get("acquisition_cost"))
        cv = _num(prop.get("current_value")) if prop.get("current_value") else ac
        if d:              ci[f"A{r}"] = d
        if ya:             ci[f"D{r}"] = ya
        if ac is not None: ci[f"F{r}"] = ac
        if cv is not None: ci[f"H{r}"] = cv


def _write_vehicle_assets(ci: dict, vehicles: list,
                           max_vehicles: int = 5) -> None:
    for i, prop in enumerate(vehicles[:max_vehicles]):
        r  = 73 + i
        d  = str(prop.get("description") or "").strip()
        ya = str(prop.get("year_acquired") or "").strip()
        ac = _num(prop.get("acquisition_cost"))
        cv = _num(prop.get("current_value")) if prop.get("current_value") else ac
        if d:              ci[f"A{r}"] = d
        if ya:             ci[f"D{r}"] = ya
        if ac is not None: ci[f"F{r}"] = ac
        if cv is not None: ci[f"H{r}"] = cv


def _merge_saln_liabilities(unified_credits: list[dict],
                             saln: dict) -> list[dict]:
    saln_liabilities = saln.get("liabilities") or []
    if not saln_liabilities:
        return unified_credits

    existing = {
        str(c.get("institution") or "").lower()
        for c in unified_credits
    }

    for liab in saln_liabilities:
        creditor = str(
            liab.get("creditor") or liab.get("name") or ""
        ).strip()
        nature   = str(liab.get("nature") or "").strip()
        balance  = _num(liab.get("outstanding_balance"))

        if not creditor or balance is None:
            continue

        cred_lower = creditor.lower()
        already = any(
            cred_lower in ex or ex in cred_lower
            for ex in existing if ex
        )
        if already:
            continue

        unified_credits.append({
            "institution":          creditor,
            "principal_loan":       None,
            "monthly_amortization": None,
            "balance":              balance,
            "due_date":             "",
            "status":               nature or "SALN Liability",
            "overdue":              None,
        })
        existing.add(cred_lower)

    return unified_credits


# ─────────────────────────────────────────────────────────────────────────────
#  CIC-KW-2 — PRE-FILTER: extract only relevant lines before sending to Gemini
# ─────────────────────────────────────────────────────────────────────────────

_CIC_TARGET_KEYWORDS: tuple[str, ...] = (
    "past due",
    "write off",
    "write-off",
    "written off",
)

def _cic_keyword_filter(text: str, context_lines: int = 5) -> str:
    """
    Return only the portions of the CIC text that contain a target keyword,
    plus `context_lines` lines above and below each match for context.
    If no keywords are found, return the original text unchanged.
    """
    lines      = text.splitlines()
    hit_indices: set[int] = set()

    for i, line in enumerate(lines):
        line_lower = line.lower()
        if any(kw in line_lower for kw in _CIC_TARGET_KEYWORDS):
            for j in range(
                max(0, i - context_lines),
                min(len(lines), i + context_lines + 1)
            ):
                hit_indices.add(j)

    if not hit_indices:
        _log.info("_cic_keyword_filter — no keywords found; passing full text.")
        return text

    kept  = sorted(hit_indices)
    parts: list[str] = []
    prev  = None
    for idx in kept:
        if prev is not None and idx > prev + 1:
            parts.append("…")
        parts.append(lines[idx])
        prev = idx

    filtered = "\n".join(parts)
    _log.info(
        f"_cic_keyword_filter — kept {len(kept)}/{len(lines)} lines "
        f"({len(filtered):,} of {len(text):,} chars)."
    )
    return filtered


# ─────────────────────────────────────────────────────────────────────────────
#  CIC-KW-1 — KEYWORD-ONLY PROMPT
# ─────────────────────────────────────────────────────────────────────────────

_CIC_KEYWORD_PROMPT = """
You are reviewing a Philippine CIC (Credit Information Corporation) credit report.
Your ONLY task is to find loan or credit accounts that are flagged as either:
  • PAST DUE
  • WRITE OFF  (also: "Write-Off", "Written Off")

Return ONLY a valid JSON object — no explanation, no markdown, no extra text.
Use null for any field not found. Numbers only (no ₱, no commas).
Return [] for empty lists.

RULES:
  1. Search the ENTIRE document — both the "Requested/Renounced/Refused" table
     AND the "Active/Closed" table.
  2. Include a row ONLY if its status, contract_phase, or any adjacent label
     contains "PAST DUE", "WRITE OFF", "WRITE-OFF", or "WRITTEN OFF"
     (case-insensitive).
  3. Do NOT include rows with other statuses (e.g. Active, Closed, Requested).
  4. If no matching rows are found, return empty lists.
  5. For each matching row extract ALL of the fields listed below.
     Use null for any field not present in that row.

Return exactly this structure:
{{
  "past_due": [
    {{
      "provider":                null,
      "contract_type":           null,
      "contract_phase":          null,
      "financed_amount":         null,
      "outstanding_balance":     null,
      "overdue_payments_amount": null,
      "monthly_payments_amount": null,
      "contract_start_date":     null,
      "contract_end_date":       null,
      "cic_contract_code":       null
    }}
  ],
  "write_off": [
    {{
      "provider":                null,
      "contract_type":           null,
      "contract_phase":          null,
      "financed_amount":         null,
      "outstanding_balance":     null,
      "overdue_payments_amount": null,
      "monthly_payments_amount": null,
      "contract_start_date":     null,
      "contract_end_date":       null,
      "cic_contract_code":       null
    }}
  ]
}}

--- CIC DOCUMENT ---
{text}
--- END ---
"""
try:
    from cibi_populator_patch import apply_prompt_patch
    apply_prompt_patch(globals())
except ImportError:
    pass   # patch not installed — hardcoded prompts used as fallback

def _merge_keyword_results(results: list[dict]) -> dict:
    """Merge multiple chunk results, deduplicating by provider+contract_type+financed_amount."""
    merged: dict[str, list] = {"past_due": [], "write_off": []}
    for section in ("past_due", "write_off"):
        seen: set[str] = set()
        for chunk in results:
            for item in (chunk.get(section) or []):
                if not isinstance(item, dict):
                    continue
                key = "|".join(str(item.get(k) or "").strip().lower() for k in (
                    "provider", "contract_type", "financed_amount"
                ))
                if key not in seen:
                    seen.add(key)
                    merged[section].append(item)
    return merged


# ─────────────────────────────────────────────────────────────────────────────
#  TL-1 — UPDATED CIC EXTRACTORS (keyword‑only with thinking OFF)
# ─────────────────────────────────────────────────────────────────────────────

def extract_cic_chunked(text: str, api_key: str) -> dict:
    """Chunked CIC extraction using the keyword‑only prompt. Thinking OFF."""
    CHUNK_SIZE = 35_000
    OVERLAP    = 2_000
    STEP       = CHUNK_SIZE - OVERLAP

    chunks: list[str] = []
    start = 0
    while start < len(text):
        chunks.append(text[start : start + CHUNK_SIZE])
        start += STEP

    results: list[dict] = []
    for i, chunk in enumerate(chunks, 1):
        filtered_chunk = _cic_keyword_filter(chunk)
        cached = _cache_get(filtered_chunk, f"cic_kw_chunk_{i}of{len(chunks)}")
        if cached:
            results.append(cached)
            continue

        result = _gemini_extract_json(
            _CIC_KEYWORD_PROMPT.replace("{text}", filtered_chunk[:20_000]),
            api_key,
            label           = f"CIC_KW_chunk_{i}of{len(chunks)}",
            max_tokens      = 8_192,      # TL-1
            thinking_budget = 0,          # TL-1: OFF — keyword match, no reasoning
        )
        if not result.get("_error"):
            _cache_set(filtered_chunk, f"cic_kw_chunk_{i}of{len(chunks)}", result)
        results.append(result)

    return _merge_keyword_results(results)


def extract_cic(text: str, api_key: str) -> dict:
    """
    Keyword‑only CIC extraction. Thinking OFF.
    Returns:
        {
            "past_due":  [ {...}, ... ],
            "write_off": [ {...}, ... ]
        }
    """
    _log.info(f"extract_cic (keyword‑only, thinking=OFF) — {len(text):,} chars")

    if len(text) > 75_000:
        return extract_cic_chunked(text, api_key)

    filtered_text = _cic_keyword_filter(text)

    cached = _cache_get(filtered_text, "cic_kw")
    if cached:
        return cached

    result = _gemini_extract_json(
        _CIC_KEYWORD_PROMPT.replace("{text}", filtered_text[:20_000]),  # TL-5
        api_key,
        label           = "CIC_KW",
        max_tokens      = 8_192,   # TL-1
        thinking_budget = 0,       # TL-1: OFF
    )

    if not result.get("_error"):
        _cache_set(filtered_text, "cic_kw", result)
    else:
        _log.warning(f"CIC keyword extraction error: {result['_error']}")
        result = {"past_due": [], "write_off": []}

    return result


def summarise_cic_keywords(cic: dict) -> str:
    """
    Returns a human‑readable summary of past_due and write_off accounts.
    Useful for populating a notes cell in the CIBI template.
    """
    lines: list[str] = []

    def _peso(val) -> str:
        try:
            n = float(str(val).replace("₱","").replace(",","").strip())
            return f"₱{n:,.2f}"
        except Exception:
            return str(val) if val else ""

    for section, label in (("past_due", "PAST DUE"), ("write_off", "WRITE OFF")):
        rows = cic.get(section) or []
        if not rows:
            continue
        lines.append(f"{label} Accounts ({len(rows)}):")
        for i, r in enumerate(rows, 1):
            provider  = r.get("provider") or "Unknown"
            c_type    = r.get("contract_type") or ""
            financed  = _peso(r.get("financed_amount"))
            balance   = _peso(r.get("outstanding_balance"))
            overdue   = _peso(r.get("overdue_payments_amount"))
            start     = r.get("contract_start_date") or ""
            end       = r.get("contract_end_date") or ""
            parts = [f"{i}. {provider}"]
            if c_type:   parts.append(f"Type: {c_type}")
            if financed: parts.append(f"Financed: {financed}")
            if balance:  parts.append(f"Balance: {balance}")
            if overdue:  parts.append(f"Overdue: {overdue}")
            if start:    parts.append(f"Start: {start}")
            if end:      parts.append(f"End: {end}")
            lines.append("  " + "  |  ".join(parts))

    return "\n".join(lines) if lines else "No PAST DUE or WRITE OFF accounts found."


# ─────────────────────────────────────────────────────────────────────────────
#  TL-3 — INDIVIDUAL FALLBACK EXTRACTORS (thinking_budget=512)
# ─────────────────────────────────────────────────────────────────────────────

def extract_payslip(text: str, api_key: str) -> dict:
    """TL-3: thinking_budget=512, max_output_tokens=16_000"""
    cached = _cache_get(text, "payslip")
    if cached:
        return cached
    few_shot = get_few_shot_example("payslip", text) if _FEW_SHOT_AVAILABLE else None
    result = _gemini_extract_json(
        _PAYSLIP_PROMPT.replace("{text}", text[:20_000]),
        api_key,
        label           = "PAYSLIP",
        max_tokens      = 16_000,     # TL-3
        thinking_budget = 512,        # TL-3
        few_shot_example= few_shot,
    )
    _cache_set(text, "payslip", result)
    return result


def extract_saln(text: str, api_key: str) -> dict:
    """TL-3: thinking_budget=512, max_output_tokens=16_000"""
    cached = _cache_get(text, "saln")
    if cached:
        return cached
    few_shot = get_few_shot_example("saln", text) if _FEW_SHOT_AVAILABLE else None
    result = _gemini_extract_json(
        _SALN_PROMPT.replace("{text}", text[:30_000]),
        api_key,
        label           = "SALN",
        max_tokens      = 16_000,     # TL-3
        thinking_budget = 512,        # TL-3
        few_shot_example= few_shot,
    )
    _cache_set(text, "saln", result)
    return result


def extract_itr(text: str, api_key: str) -> dict:
    """TL-3: thinking_budget=512, max_output_tokens=16_000"""
    cached = _cache_get(text, "itr")
    if cached:
        return cached
    few_shot = get_few_shot_example("itr", text) if _FEW_SHOT_AVAILABLE else None
    result = _gemini_extract_json(
        _ITR_PROMPT.replace("{text}", text[:30_000]),
        api_key,
        label           = "ITR",
        max_tokens      = 16_000,     # TL-3
        thinking_budget = 512,        # TL-3
        few_shot_example= few_shot,
    )
    _cache_set(text, "itr", result)
    return result


# ─────────────────────────────────────────────────────────────────────────────
#  UNIFIED EXTRACTOR (TL-2: thinking_budget=1024, max_tokens=24_576)
# ─────────────────────────────────────────────────────────────────────────────

def extract_all_unified(
    cic_text:     str,
    payslip_text: str,
    saln_text:    str,
    itr_text:     str,
    api_key:      str,
) -> tuple[dict, dict, dict, dict]:
    has_cic     = bool(cic_text.strip())
    has_payslip = bool(payslip_text.strip())
    has_saln    = bool(saln_text.strip())
    has_itr     = bool(itr_text.strip())

    composite_key = (
        _cache_key(cic_text,     "cic")     + "_" +
        _cache_key(payslip_text, "payslip") + "_" +
        _cache_key(saln_text,    "saln")    + "_" +
        _cache_key(itr_text,     "itr")
    )
    cache_path = _cache_dir() / f"unified_{composite_key}.json"
    if cache_path.exists():
        try:
            stored = json.loads(cache_path.read_text(encoding="utf-8"))
            _log.info("Cache HIT [unified] — skipping all Gemini extraction calls.")
            return (
                stored.get("cic",     {}),
                stored.get("payslip", {}),
                stored.get("saln",    {}),
                stored.get("itr",     {}),
            )
        except Exception as e:
            _log.warning(f"Unified cache read error: {e}")

    cic_data = {}
    if has_cic:
        if len(cic_text) > 75_000:
            cic_data = extract_cic_chunked(cic_text, api_key)
        else:
            cic_data = extract_cic(cic_text, api_key)

    cic_section = "NOT PROVIDED — extracted separately."

    def _few_shot_block(doc_type: str, doc_text: str) -> str:
        if not _FEW_SHOT_AVAILABLE:
            return ""
        ex = get_few_shot_example(doc_type, doc_text)
        if ex is None:
            return ""
        try:
            ex_json = json.dumps(ex.approved_json, indent=2, ensure_ascii=False)
            return (
                f"EXAMPLE {doc_type.upper()} DOCUMENT:\n"
                f"{ex.text_content[:2000]}\n\n"
                f"CORRECT JSON FOR ABOVE EXAMPLE:\n{ex_json}\n\n"
                f"---\nNow extract from the ACTUAL {doc_type.upper()} document below.\n\n"
            )
        except Exception:
            return ""

    payslip_prefix = _few_shot_block("payslip", payslip_text) if has_payslip else ""
    saln_prefix    = _few_shot_block("saln",    saln_text)    if has_saln    else ""
    itr_prefix     = _few_shot_block("itr",     itr_text)     if has_itr     else ""

    prompt = _UNIFIED_PROMPT.format(
        cic_text     = cic_section,
        payslip_text = (payslip_prefix + payslip_text[:15_000]) if has_payslip else "NOT PROVIDED",
        saln_text    = (saln_prefix    + saln_text[:20_000])    if has_saln    else "NOT PROVIDED",
        itr_text     = (itr_prefix     + itr_text[:15_000])     if has_itr     else "NOT PROVIDED",
    )

    # TL-2: Unified Gemini call with thinking_budget=1024, max_tokens=24_576
    result = _gemini_extract_json(
        prompt,
        api_key,
        label           = "UNIFIED",
        max_tokens      = 24_576,   # TL-2
        thinking_budget = 1_024,    # TL-2: moderate reasoning for 3 docs at once
    )

    if result.get("_error"):
        _log.warning(f"Unified extraction failed. Falling back to individual extractors.")
        pay_data  = extract_payslip(payslip_text, api_key) if has_payslip else {}
        saln_data = extract_saln(saln_text, api_key)       if has_saln    else {}
        itr_data  = extract_itr(itr_text, api_key)         if has_itr     else {}
        return cic_data, pay_data, saln_data, itr_data

    pay_data  = result.get("payslip") or {}
    saln_data = result.get("saln")    or {}
    itr_data  = result.get("itr")     or {}

    cache_payload = {"cic": cic_data, "payslip": pay_data, "saln": saln_data, "itr": itr_data}
    try:
        cache_path.write_text(
            json.dumps(cache_payload, indent=2, ensure_ascii=False),
            encoding="utf-8"
        )
        _log.info(f"Cache WRITE [unified] → {cache_path.name}")
    except Exception as e:
        _log.warning(f"Unified cache write error: {e}")

    return cic_data, pay_data, saln_data, itr_data


# ─────────────────────────────────────────────────────────────────────────────
#  INSTALLMENT SUMMARY HELPERS (may not be used with keyword-only CIC, but kept)
# ─────────────────────────────────────────────────────────────────────────────

def _summarise_installments_requested(rows: list) -> str:
    if not rows:
        return ""
    lines = ["Requested/Renounced/Refused Loans:"]
    for i, r in enumerate(rows, 1):
        provider  = r.get("provider_description") or r.get("institution") or "Unknown"
        c_type    = r.get("contract_type") or ""
        phase     = r.get("contract_phase") or ""
        amount    = _peso(r.get("financed_amount"))
        monthly   = _peso(r.get("monthly_payments_amount"))
        req_date  = r.get("contract_request_date") or ""
        n_install = r.get("installments_number") or ""
        periodicy = r.get("payment_periodicity") or ""
        note      = r.get("note") or ""
        parts = [f"{i}. {provider}"]
        if c_type:    parts.append(f"Type: {c_type}")
        if phase:     parts.append(f"Phase: {phase}")
        if amount:    parts.append(f"Financed: {amount}")
        if monthly:   parts.append(f"Monthly: {monthly}")
        if n_install: parts.append(f"Installments: {n_install}")
        if periodicy: parts.append(f"Periodicity: {periodicy}")
        if req_date:  parts.append(f"Requested: {req_date}")
        if note:      parts.append(f"Note: {note}")
        lines.append("  " + "  |  ".join(parts))
    return "\n".join(lines)


def _summarise_installments_active(rows: list) -> str:
    if not rows:
        return ""
    lines = ["Active/Closed Loans:"]
    for i, r in enumerate(rows, 1):
        provider    = r.get("provider_description") or r.get("institution") or "Unknown"
        c_type      = r.get("contract_type") or ""
        financed    = _peso(r.get("financed_amount"))
        outstanding = _peso(r.get("outstanding_balance"))
        overdue     = _peso(r.get("overdue_payments_amount"))
        start       = r.get("contract_start_date") or ""
        end         = r.get("contract_end_date") or ""
        note        = r.get("note") or ""
        parts = [f"{i}. {provider}"]
        if c_type:      parts.append(f"Type: {c_type}")
        if financed:    parts.append(f"Financed: {financed}")
        if outstanding: parts.append(f"Outstanding: {outstanding}")
        if overdue:     parts.append(f"Overdue: {overdue}")
        if start:       parts.append(f"Start: {start}")
        if end:         parts.append(f"End: {end}")
        if note:        parts.append(f"Note: {note}")
        lines.append("  " + "  |  ".join(parts))
    return "\n".join(lines)


def _total_outstanding_from_active(rows: list) -> Optional[float]:
    total = 0.0; found = False
    for r in (rows or []):
        v = _num(r.get("outstanding_balance"))
        if v is not None:
            total += v; found = True
    return total if found else None


def _total_overdue_from_active(rows: list) -> Optional[float]:
    total = 0.0; found = False
    for r in (rows or []):
        v = _num(r.get("overdue_payments_amount"))
        if v is not None:
            total += v; found = True
    return total if found else None


def _total_monthly_from_requested(rows: list) -> Optional[float]:
    total = 0.0; found = False
    for r in (rows or []):
        v = _num(r.get("monthly_payments_amount"))
        if v is not None:
            total += v; found = True
    return total if found else None


# ─────────────────────────────────────────────────────────────────────────────
#  CELL MAP BUILDER — v3 (all OF-1 through OF-6 fixes applied)
#  NOTE: This function still references old CIC fields (full_name, spouse, etc.)
#  which are no longer present. You must adapt it to use the keyword-only
#  CIC output or to rely on payslip/SALN/ITR for personal data.
# ─────────────────────────────────────────────────────────────────────────────

MAX_CREDIT_ROWS  = 3
MAX_REAL_PROP    = 3
MAX_PERS_PROP    = 5
MAX_VEHICLES     = 5
MAX_BIZ_ASSETS   = 5
MAX_CHILDREN     = 4
MAX_INCOME_ROWS  = 5
MAX_BIZ_EXP_ROWS = 6


def _build_cell_map(cic: dict, pay: dict, saln: dict,
                    itr: dict | None = None) -> dict[str, dict]:
    from datetime import datetime

    af: dict = {}
    ci: dict = {}
    cf: dict = {}
    itr = itr or {}

    _log.info("Building cell map…")

    # WARNING: The cic dict now contains only "past_due" and "write_off" keys.
    # Most of the old CIC fields (full_name, spouse, employer, etc.) are missing.
    # You should update this function to get personal data from payslip/SALN/ITR.
    # For now, we keep the original code but it will produce many None values.

    pay_periods  = pay.get("pay_periods") or []
    period_count = _v(pay, "period_count")

    if len(pay_periods) > 1:
        def _local_avg(field: str) -> Optional[float]:
            vals = [_num(p.get(field)) for p in pay_periods]
            vals = [v for v in vals if v is not None]
            return sum(vals) / len(vals) if vals else None

        _AVERAGED = (
            "basic_pay","allowances","gross_pay","net_pay","total_deductions",
            "gsis_deduction","sss_deduction","philhealth_deduction",
            "pagibig_deduction","tax_deduction","other_deductions",
        )
        for field in _AVERAGED:
            gemini_val = _num(_v(pay, field))
            local_avg  = _local_avg(field)
            if local_avg is not None and gemini_val is not None:
                if gemini_val > local_avg * 1.8:
                    pay[field] = local_avg
            elif local_avg is not None and gemini_val is None:
                pay[field] = local_avg

    # CIC fields that no longer exist: use fallbacks from other docs
    inst_requested = cic.get("installments_requested") or []   # always empty now
    inst_active    = cic.get("installments_active") or []      # always empty now

    applicant = (
        _v(cic, "full_name") or        # will be None
        _join_name(cic, "last_name","first_name","middle_name","suffix") or
        _v(pay, "employee_name") or
        _v(saln, "declarant_name") or ""
    )
    spouse = (
        _join_name(cic,"spouse_first_name","spouse_middle_name","spouse_last_name") or
        _v(saln, "spouse_name") or ""
    )

    dob = _get_dob(cic, pay, itr)
    age = str(_v(cic,"age") or _age_from_dob(str(dob)) or "")

    residence = _normalise_text(
        _v(cic, "residence_address") or
        _v(itr, "registered_address") or
        _v(saln, "office_address") or
        ""
    )
    permanent = _normalise_text(
        _v(cic, "mailing_address") or
        residence or
        ""
    )

    raw_contact = _v(cic, "contact_number") or ""
    if isinstance(raw_contact, list):
        contact = next(
            (c for c in raw_contact
             if isinstance(c, str) and (c.startswith("+63") or c.startswith("09"))),
            str(raw_contact[0]) if raw_contact else ""
        )
    else:
        contact = str(raw_contact) if raw_contact else ""

    itr_tin = _v(itr,"tin") or _v(itr,"business_tin") or ""
    tin_raw = _v(cic,"tin") or _v(pay,"tin") or itr_tin or ""
    sss_raw = _v(cic,"sss") or _v(pay,"sss_number") or ""
    tin_stripped = _strip_label_prefix(str(tin_raw), "TIN No","TIN") if tin_raw else ""
    tin = _fmt_tin(tin_stripped) if tin_stripped else ""
    sss = _strip_label_prefix(str(sss_raw), "SSS No","SSS") if sss_raw else ""

    employer   = _v(cic,"employer_name")  or _v(pay,"employer_name") or ""
    emp_addr   = _v(cic,"employer_address") or ""

    occupation = (
        _v(cic,"occupation") or
        _v(pay,"position") or
        _v(saln,"position") or
        ""
    )

    hired_from = _v(cic,"hired_from") or ""
    yrs_emp    = _years_since(str(hired_from)) if hired_from else ""

    is_emp  = _is_employed(
        employer,
        str(_v(cic,"employment_status") or ""),
        str(occupation or ""),
    )
    is_self = bool(_v(cic,"sole_trader_name"))

    spouse_age    = str(_v(cic,"spouse_age") or "")
    spouse_occ    = _v(cic,"spouse_occupation") or _v(saln,"spouse_position") or ""
    spouse_income = _num(_v(cic,"spouse_income"))
    spouse_emp    = _v(cic,"spouse_employment_status") or ""
    spouse_hired  = _v(cic,"spouse_hired_from") or ""
    spouse_yrs    = _years_since(str(spouse_hired)) if spouse_hired else ""
    spouse_is_emp = bool(spouse_emp or spouse_occ)
    spouse_tin    = _v(cic,"spouse_tin") or ""

    gross_cic: Optional[float] = None
    if _v(cic,"gross_income"):
        n    = _num(_v(cic,"gross_income"))
        freq = str(_v(cic,"income_frequency") or "").lower()
        gross_cic = n / 12 if "annual" in freq and n else n

    net_pay        = _num(_v(pay,"net_pay"))
    gross_pay      = _num(_v(pay,"gross_pay"))
    basic_pay      = _num(_v(pay,"basic_pay"))
    business_inc   = _num(_v(pay,"business_income"))
    rental_inc     = _num(_v(pay,"rental_income"))
    remittance_inc = _num(_v(pay,"remittance_income"))
    other_inc_lbl  = _v(pay,"other_income_label") or ""
    other_inc_amt  = _num(_v(pay,"other_income_amount"))

    itr_gross_annual  = _num(_v(itr,"gross_annual_income","total_gross_income",
                                   "gross_compensation_income"))
    itr_gross_monthly = _num(_v(itr,"gross_monthly_income"))
    itr_net_taxable   = _num(_v(itr,"net_taxable_income"))
    itr_biz_income    = _num(_v(itr,"gross_business_income","gross_professional_income"))
    itr_biz_name      = _v(itr,"business_name") or ""
    itr_tax_year      = _v(itr,"tax_year") or ""
    itr_tax_due       = _num(_v(itr,"tax_due"))
    itr_tax_paid      = _num(_v(itr,"tax_paid"))

    itr_monthly = (itr_gross_monthly or
                   (itr_gross_annual / 12 if itr_gross_annual else None))

    monthly = net_pay or basic_pay or itr_monthly or gross_pay or gross_cic
    _log.info(
        f"OF-1 income resolution: net_pay={net_pay}, basic_pay={basic_pay}, "
        f"itr_monthly={itr_monthly}, gross_pay={gross_pay}, gross_cic={gross_cic} "
        f"→ monthly={monthly}"
    )

    tot_assets = _num(_v(saln,"total_assets"))
    tot_liab   = _num(_v(saln,"total_liabilities"))
    net_worth  = _num(_v(saln,"net_worth"))

    total_monthly_amort = (
        _total_monthly_from_requested(inst_requested) or
        _num(_v(cic,"total_monthly_amortization"))
    )
    total_balance = (
        _total_outstanding_from_active(inst_active) or
        _num(_v(cic,"total_loan_balance"))
    )
    total_overdue = (
        _total_overdue_from_active(inst_active) or
        _num(_v(cic,"total_overdue_payments"))
    )

    # ── APPROVAL FORM ─────────────────────────────────────────────────────
    if period_count and int(str(period_count)) > 1:
        period_note = (
            f"Average of {period_count} monthly payslips "
            f"({_v(pay,'period_from')} – {_v(pay,'period_to')})"
            if _v(pay,"period_from") and _v(pay,"period_to")
            else f"Average of {period_count} monthly payslips"
        )
    else:
        period_note = ""

    inc_src = " / ".join(filter(None, [occupation, employer]))
    if applicant:               af["E6"]  = applicant
    if spouse:                  af["E7"]  = spouse
    if residence:               af["E9"]  = residence
    if inc_src:                 af["E12"] = inc_src
    if monthly:                 af["E13"] = monthly
    if tot_assets is not None:  af["E15"] = tot_assets
    if tot_liab   is not None:  af["E16"] = tot_liab
    if net_worth  is not None:  af["E17"] = net_worth

    def _block(parts):
        return "\n".join(p for p in parts if p)

    civil_status = _v(cic,"civil_status") or _v(itr,"civil_status") or ""
    if spouse and civil_status.upper() == "SINGLE":
        civil_status = "Married"
    elif spouse and not civil_status:
        civil_status = "Married"

    af["C27"] = _block([
        f"Applicant: {applicant}"  if applicant else "",
        f"Civil Status: {civil_status}" if civil_status else "",
        (f"Dependents: {_v(cic,'number_of_dependents')}"
         if _v(cic,"number_of_dependents") is not None else ""),
        f"TIN: {tin}" if tin and not sss else "",
        f"TIN: {tin}   SSS: {sss}" if tin and sss else "",
        f"SSS: {sss}" if sss and not tin else "",
    ]) or ""

    af["C31"] = _block([
        f"Occupation: {occupation}"      if occupation else "",
        f"Employer: {employer}"          if employer else "",
        f"Monthly Net Pay: {_peso(net_pay)}" if net_pay else
            (f"Monthly Income: {_peso(monthly)}" if monthly else ""),
        f"Income Basis: {period_note}"   if period_note else "",
        f"Years Employed: {yrs_emp}"     if yrs_emp else "",
        (f"ITR Tax Year: {itr_tax_year}  |  Gross Annual: {_peso(itr_gross_annual)}"
         if itr_gross_annual else ""),
        f"Net Taxable: {_peso(itr_net_taxable)}" if itr_net_taxable else "",
        (f"Tax Due: {_peso(itr_tax_due)}  |  Tax Paid: {_peso(itr_tax_paid)}"
         if itr_tax_due else ""),
    ]) or ""

    af["C35"] = _block([
        f"Total Assets: {_peso(tot_assets)}"      if tot_assets else "",
        f"Total Liabilities: {_peso(tot_liab)}"   if tot_liab else "",
        f"Net Worth: {_peso(net_worth)}"           if net_worth else "",
    ]) or ""

    c39_parts = [
        f"Net Monthly Income: {_peso(net_pay or monthly)}" if (net_pay or monthly) else "",
        f"Total Monthly Obligations: {_peso(total_monthly_amort)}" if total_monthly_amort else "",
        f"Total Outstanding: {_peso(total_balance)}"               if total_balance else "",
        f"Total Overdue: {_peso(total_overdue)}"                   if total_overdue else "",
    ]
    req_summary    = _summarise_installments_requested(inst_requested)
    active_summary = _summarise_installments_active(inst_active)
    if req_summary:    c39_parts.append(req_summary)
    if active_summary: c39_parts.append(active_summary)
    af["C39"] = _block(c39_parts) or ""

    # ── CIBI personal data ────────────────────────────────────────────────
    ci["H6"]  = datetime.today().strftime("%m/%d/%Y")
    if applicant:  ci["C7"]  = applicant
    if age:        ci["H7"]  = age
    if residence:  ci["C8"]  = residence
    if permanent:  ci["C9"]  = permanent
    if occupation: ci["C10"] = occupation
    if monthly:    ci["H10"] = _peso(monthly)
    if is_emp:     ci["C11"] = "YES"
    if yrs_emp:    ci["I11"] = yrs_emp
    if is_self:    ci["C12"] = "YES"
    if yrs_emp and is_self: ci["I12"] = yrs_emp

    if emp_addr:   ci["C13"] = emp_addr
    elif employer: ci["C13"] = employer

    if tin and sss:
        ci["H13"] = f"{tin}  /  SSS: {sss}"
    elif tin:
        ci["H13"] = tin
    elif sss:
        ci["H13"] = f"SSS: {sss}"

    if dob:        ci["H14"] = _fmt_dob(str(dob))

    if spouse:     ci["C15"] = spouse
    if contact:    ci["C21"] = contact
    if spouse_age:    ci["H15"] = spouse_age
    if spouse_occ:    ci["C16"] = spouse_occ
    if spouse_income: ci["H16"] = _peso(spouse_income)
    if spouse_is_emp: ci["C17"] = "YES"
    if spouse_yrs:    ci["I17"] = spouse_yrs
    if spouse_tin:    ci["H19"] = spouse_tin

    # ── Children ──────────────────────────────────────────────────────────
    _write_children(ci, saln, cic, MAX_CHILDREN)

    # ── OF-5: Build unified credit list — avoid duplicate empty rows ───────
    unified_credits: list[dict] = []

    # CIC inst_active and inst_requested are now empty, so this block does nothing
    for r in inst_active:
        bal = _num(r.get("outstanding_balance"))
        if bal is None and _num(r.get("financed_amount")) is None:
            continue
        unified_credits.append({
            "institution":          r.get("provider_description") or r.get("institution") or "",
            "principal_loan":       r.get("financed_amount"),
            "monthly_amortization": None,
            "balance":              r.get("outstanding_balance"),
            "due_date":             r.get("contract_end_date"),
            "status":               r.get("contract_type") or "",
            "overdue":              r.get("overdue_payments_amount"),
        })

    for r in inst_requested:
        if not (r.get("provider_description") or r.get("institution")):
            continue
        unified_credits.append({
            "institution":          r.get("provider_description") or r.get("institution") or "",
            "principal_loan":       r.get("financed_amount"),
            "monthly_amortization": r.get("monthly_payments_amount"),
            "balance":              None,
            "due_date":             r.get("last_update_date"),
            "status":               r.get("contract_phase") or r.get("contract_type") or "",
            "overdue":              None,
        })

    if not unified_credits:
        for acc in (cic.get("credit_accounts") or []):
            if not acc.get("institution"):
                continue
            unified_credits.append({
                "institution":          acc.get("institution") or "",
                "principal_loan":       acc.get("principal_loan"),
                "monthly_amortization": acc.get("monthly_amortization"),
                "balance":              acc.get("balance"),
                "due_date":             acc.get("due_date"),
                "status":               acc.get("status") or "",
                "overdue":              None,
            })

    unified_credits = _merge_saln_liabilities(unified_credits, saln)

    _CREDIT_ROW_MERGED = {33}
    total_amort            = 0.0
    total_balance_computed = 0.0
    for i, acc in enumerate(unified_credits[:MAX_CREDIT_ROWS]):
        r = 31 + i
        institution = acc.get("institution") or ""
        p  = _num(acc.get("principal_loan"))
        dd = acc.get("due_date") or ""
        a  = _num(acc.get("monthly_amortization"))
        b  = _num(acc.get("balance"))
        ov = _num(acc.get("overdue"))
        status = acc.get("status") or ""

        if r in _CREDIT_ROW_MERGED:
            parts = [institution] if institution else []
            if p  is not None: parts.append(f"Principal: {_peso(p)}")
            if b  is not None: parts.append(f"Balance: {_peso(b)}")
            if a  is not None: parts.append(f"Amort: {_peso(a)}")
            if dd:             parts.append(f"Due: {dd}")
            if status:         parts.append(f"({status})")
            if ov and ov > 0:  parts.append(f"Overdue: {_peso(ov)}")
            if parts:
                ci[f"A{r}"] = "  |  ".join(parts)
        else:
            if institution: ci[f"A{r}"] = institution
            if p  is not None: ci[f"D{r}"] = p
            if dd:             ci[f"F{r}"] = dd
            if a  is not None: ci[f"G{r}"] = a
            if b  is not None: ci[f"H{r}"] = b
            if ov is not None and ov > 0:
                ci[f"I{r}"] = f"Overdue: {_peso(ov)}"

        if a is not None: total_amort += a
        if b is not None: total_balance_computed += b

    final_amort   = total_monthly_amort or (total_amort   if total_amort   else None)
    final_balance = total_balance       or (total_balance_computed if total_balance_computed else None)
    if final_amort:   ci["G34"] = final_amort
    if final_balance: ci["H34"] = final_balance
    if total_overdue: ci["I34"] = f"Total Overdue: {_peso(total_overdue)}"

    # ── Real properties ───────────────────────────────────────────────────
    real_props = saln.get("real_properties") or []
    for i, prop in enumerate(real_props[:MAX_REAL_PROP]):
        r    = 37 + i

        desc = (
            prop.get("description")             or
            prop.get("property_description")    or
            prop.get("property_type")           or
            ""
        )
        if not desc:
            kind = prop.get("kind") or prop.get("type") or ""
            desc = f"{kind} Property".strip() if kind else ""

        area = prop.get("area") or ""
        loc  = (
            prop.get("location")       or
            prop.get("exact_location") or
            prop.get("address")        or
            ""
        )
        cv   = _num(
            prop.get("current_fair_market_value") or
            prop.get("current_fair_value")        or
            prop.get("current_value")             or
            prop.get("assessed_value")
        )
        yr   = str(prop.get("year_acquired") or "").strip()
        ac   = _num(prop.get("acquisition_cost"))

        if desc:  ci[f"A{r}"] = desc
        if area:  ci[f"C{r}"] = area
        if loc and yr:
            ci[f"D{r}"] = f"{loc}  ({yr})"
        elif loc:
            ci[f"D{r}"] = loc
        elif yr:
            ci[f"D{r}"] = yr
        if ac is not None: ci[f"F{r}"] = ac
        if cv is not None: ci[f"H{r}"] = cv

    # ── Personal & vehicle assets ─────────────────────────────────────────
    pers_props = saln.get("personal_properties") or []
    vehicles   = [p for p in pers_props if any(
        w in str(p.get("description","")).lower()
        for w in ("car","vehicle","motorcycle","truck","van","jeep",
                  "tricycle","bicycle","motorbike","suv","pickup")
    )]
    household  = [p for p in pers_props if p not in vehicles]

    _write_household_assets(ci, household, MAX_PERS_PROP)
    _write_vehicle_assets(ci, vehicles, MAX_VEHICLES)

    biz_interests = _num(saln.get("business_interests"))
    if biz_interests:
        ci["A82"] = "Business Interest"
        ci["H82"] = biz_interests

    cash_on_hand  = _num(_v(saln,"cash_on_hand"))
    cash_in_bank  = _num(_v(saln,"cash_in_bank"))
    cash_combined = _num(_v(saln,"cash_on_hand_and_in_bank"))
    acc_rec       = _num(_v(saln,"receivables"))
    if cash_on_hand is not None:
        ci["H58"] = cash_on_hand
    elif cash_combined is not None:
        ci["H58"] = cash_combined
    if cash_in_bank is not None:
        ci["H60"] = cash_in_bank
    if acc_rec is not None: ci["H61"] = acc_rec

    # ── CASHFLOW income rows ──────────────────────────────────────────────
    income_items: list[tuple[str, float]] = []

    if net_pay:
        emp_label = f"Net Salary – {employer}" if employer else "Net Salary"
        income_items.append((emp_label, net_pay))
        if itr_monthly and abs(itr_monthly - net_pay) > 100:
            itr_label = (f"ITR Gross Income ({itr_tax_year})" if itr_tax_year
                         else "ITR Gross Income")
            income_items.append((itr_label, itr_monthly))
    elif itr_monthly:
        itr_label = (f"ITR Income ({itr_tax_year})" if itr_tax_year
                     else "ITR Annual Income")
        income_items.append((itr_label, itr_monthly))
    elif gross_pay or basic_pay or gross_cic:
        amt   = gross_pay or basic_pay or gross_cic
        label = f"Salary – {employer}" if employer else "Salary"
        income_items.append((label, amt))

    if business_inc:
        biz_name = _v(cic,"sole_trader_name") or "Business"
        income_items.append((f"Business Income – {biz_name}", business_inc))
    if rental_inc:
        income_items.append(("Rental Income", rental_inc))
    if remittance_inc:
        income_items.append(("Remittance / OFW Income", remittance_inc))
    if other_inc_amt and other_inc_lbl:
        income_items.append((other_inc_lbl, other_inc_amt))
    elif other_inc_amt:
        income_items.append(("Other Income", other_inc_amt))
    if itr_biz_income and not business_inc:
        biz_label = (f"Business/Prof. Income – {itr_biz_name}"
                     if itr_biz_name else "Business/Professional Income")
        income_items.append((biz_label, itr_biz_income))

    for idx, (lbl, amt) in enumerate(income_items[:MAX_INCOME_ROWS]):
        r = 12 + idx
        cf[f"A{r}"] = lbl
        cf[f"G{r}"] = amt

    # ── FIX-J: CASHFLOW deductions ────────────────────────────────────────
    other_ded      = _num(_v(pay,"other_deductions"))
    gsis_ded       = _num(_v(pay,"gsis_deduction"))      or 0.0
    sss_ded        = _num(_v(pay,"sss_deduction"))        or 0.0
    philhealth_ded = _num(_v(pay,"philhealth_deduction")) or 0.0
    pagibig_ded    = _num(_v(pay,"pagibig_deduction"))    or 0.0
    tax_ded        = _num(_v(pay,"tax_deduction"))         or 0.0

    mandatory_total = gsis_ded + sss_ded + philhealth_ded + pagibig_ded
    mandatory_label = " + ".join(filter(None, [
        "GSIS"        if gsis_ded       else "",
        "SSS"         if sss_ded        else "",
        "PhilHealth"  if philhealth_ded else "",
        "Pag-IBIG"    if pagibig_ded    else "",
    ]))

    if mandatory_total > 0:
        cf["A37"] = mandatory_label or "Gov't Deductions"
        cf["G37"] = mandatory_total

    if tax_ded > 0:
        cf["A38"] = "Withholding Tax"
        cf["G38"] = tax_ded

    if other_ded:
        cf["A36"] = "Other Deductions"
        cf["G36"] = other_ded

    _log.info(f"Cell map — AF: {len(af)}  CI: {len(ci)}  CF: {len(cf)}")
    return {"APPROVAL FORM": af, "CIBI": ci, "CASHFLOW": cf}


# ─────────────────────────────────────────────────────────────────────────────
#  WRITE TO TEMPLATE (unchanged)
# ─────────────────────────────────────────────────────────────────────────────

def _write_template(
    template_path: str | Path,
    cell_map:      dict[str, dict],
    output_path:   Path,
) -> Path:
    import openpyxl
    from openpyxl.styles import Font

    wb   = openpyxl.load_workbook(str(template_path))
    blue = Font(name="Arial", size=10, color="0070C0")

    _log.info(f"Template sheets: {wb.sheetnames}")
    sheet_lookup = {_normalise_sheet_name(n): n for n in wb.sheetnames}

    _formula_override_whitelist: set[str] = {
        "CIBI::G34", "CIBI::H34",
        "CIBI::H31", "CIBI::H32", "CIBI::H33",
    }

    total_written = 0
    for target_sheet, cells in cell_map.items():
        norm_target = _normalise_sheet_name(target_sheet)
        actual_name = sheet_lookup.get(norm_target)
        if actual_name is None:
            _log.warning(
                f"Sheet '{target_sheet}' not found. "
                f"Available: {list(sheet_lookup.keys())}"
            )
            continue

        _log.info(f"Writing to '{actual_name}': {len(cells)} cells")
        ws = wb[actual_name]
        sheet_written = 0

        for addr, value in cells.items():
            if value is None or str(value).strip() in ("", "None", "null"):
                continue
            try:
                cell = ws[addr]
                override_key = f"{norm_target}::{addr}"
                if (isinstance(cell.value, str) and
                        cell.value.startswith("=") and
                        override_key not in _formula_override_whitelist):
                    _log.debug(f"  Skipped {addr} — formula: {cell.value[:30]}")
                    continue
                in_merged = False
                for mc in ws.merged_cells.ranges:
                    if mc.min_row == cell.row and mc.min_col == cell.column:
                        break
                    if (mc.min_row <= cell.row <= mc.max_row and
                            mc.min_col <= cell.column <= mc.max_col):
                        from openpyxl.utils import get_column_letter
                        tl = f"{get_column_letter(mc.min_col)}{mc.min_row}"
                        _log.warning(
                            f"  {addr} inside merged range {mc} — "
                            f"should write to {tl}. Skipping."
                        )
                        in_merged = True
                        break
                if in_merged:
                    continue
                _log.debug(f"  {addr} = {str(value)[:60]}")
                cell.value = value
                cell.font  = blue
                sheet_written += 1
                total_written += 1
            except Exception as e:
                _log.warning(f"  Could not write {addr}: {e}")

        _log.info(f"  → {sheet_written} cells written to '{actual_name}'")

    _log.info(f"Total cells written: {total_written}")
    wb.save(str(output_path))
    _log.info(f"Saved: {output_path}")
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
#  PUBLIC API
# ─────────────────────────────────────────────────────────────────────────────

def populate_cibi_form(
    template_path: str | Path,
    api_key:       str,
    cic_text:      str                = "",
    payslip_text:  str                = "",
    saln_text:     str                = "",
    itr_text:      str                = "",
    output_path:   Optional[Path]     = None,
    output_stem:   Optional[str]      = None,
    progress_cb:   Optional[Callable] = None,
) -> Path:
    def _cb(pct: int, msg: str = ""):
        if progress_cb:
            try:
                progress_cb(pct, msg)
            except Exception:
                pass

    global _progress_cb_ref, _quota_warnings
    _progress_cb_ref = progress_cb
    _quota_warnings  = []

    user_provided_output = output_path is not None

    if output_path is None:
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        stem = output_stem or "CIBI"
        output_path = _output_dir() / f"{stem}_populated_{ts}.xlsx"
    output_path = Path(output_path)

    global _log
    log_path = output_path.parent / f"{output_path.stem}_debug.log"
    _log = _setup_logger(log_path)
    _log.info("=" * 60)
    _log.info(f"CIBI Population started at {datetime.now()}")
    _log.info(f"Template : {template_path}")
    _log.info(f"Output   : {output_path}")
    _log.info(f"CIC      : {len(cic_text)} chars")
    _log.info(f"Payslip  : {len(payslip_text)} chars")
    _log.info(f"SALN     : {len(saln_text)} chars")
    _log.info(f"ITR      : {len(itr_text)} chars")
    _log.info("=" * 60)

    _cb(10, "Extracting documents (unified)…")
    cic_data, pay_data, saln_data, itr_data = extract_all_unified(
        cic_text, payslip_text, saln_text, itr_text, api_key
    )

    _cb(72, "Mapping fields to template cells…")
    cell_map = _build_cell_map(cic_data, pay_data, saln_data, itr_data)

    if not user_provided_output:
        applicant_name = (
            cell_map.get("CIBI", {}).get("C7") or
            (cell_map.get("APPROVAL FORM", {}).get("C27") or "")
                .split("\n")[0].replace("Applicant: ", "") or
            output_stem or "CIBI"
        )
        safe_name = re.sub(r'[\\/:*?"<>|]', "_", str(applicant_name)
                           ).strip().replace(" ", "_")
        if safe_name:
            ts          = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = _output_dir() / f"{safe_name}_populated_{ts}.xlsx"
            log_path    = output_path.parent / f"{output_path.stem}_debug.log"
            _log = _setup_logger(log_path)
            _log.info(f"Output renamed to: {output_path}")

    _cb(82, "Writing Excel file…")
    _write_template(template_path, cell_map, output_path)

    if _quota_warnings:
        quota_count = len([w for w in _quota_warnings if "quota" in w.lower()])
        crit_count  = len([w for w in _quota_warnings if "🚫" in w])
        if crit_count:
            summary = (
                f"🚫 {crit_count} extraction(s) FAILED due to quota exhaustion. "
                f"Check cibi_debug.log."
            )
        else:
            summary = (
                f"⚠ gemini-2.5-flash quota hit {quota_count} time(s). "
                f"gemini-2.5-flash-lite used as fallback."
            )
        _log.warning(f"Quota summary: {summary}")
        _cb(100, summary)
    else:
        _cb(100, "Done!")

    _log.info("Population complete.")
    return output_path


def get_extracted_data(
    api_key:      str,
    cic_text:     str = "",
    payslip_text: str = "",
    saln_text:    str = "",
    itr_text:     str = "",
) -> dict:
    cic_data, pay_data, saln_data, itr_data = extract_all_unified(
        cic_text, payslip_text, saln_text, itr_text, api_key
    )
    return {
        "cic":     cic_data,
        "payslip": pay_data,
        "saln":    saln_data,
        "itr":     itr_data,
    }


# ─────────────────────────────────────────────────────────────────────────────
#  SELF-TEST — validates all OF-1 through OF-16 fixes + CIC keyword + TL patch
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 60)
    print("TL-7: Cache version")
    print("=" * 60)
    assert _CACHE_VERSION == "v9", f"FAIL: {_CACHE_VERSION!r}"
    print(f"  ✅  _CACHE_VERSION = {_CACHE_VERSION!r}")

    print()
    print("=" * 60)
    print("TL-6: Model fallback chain")
    print("=" * 60)
    assert _MODELS[0] == "gemini-2.5-flash",      f"FAIL primary: {_MODELS[0]}"
    assert _MODELS[1] == "gemini-2.5-flash-lite",  f"FAIL fallback: {_MODELS[1]}"
    assert "gemini-2.0-flash" not in _MODELS,      "FAIL: deprecated model still present"
    print(f"  ✅  Primary  : {_MODELS[0]}")
    print(f"  ✅  Fallback : {_MODELS[1]}")
    print(f"  ✅  Deprecated gemini-2.0-flash removed")

    print()
    print("=" * 60)
    print("TL-1: CIC keyword — thinking OFF, max_tokens=8_192")
    print("=" * 60)
    import inspect
    sig = inspect.signature(_gemini_extract_json)
    assert "thinking_budget" in sig.parameters, "FAIL: thinking_budget param missing"
    assert sig.parameters["thinking_budget"].default == 0, "FAIL: default should be 0"
    print("  ✅  thinking_budget param present, default=0 (OFF)")

    print()
    print("=" * 60)
    print("TL-4: thinking_budget default values per call")
    print("=" * 60)
    cases = [
        ("CIC keyword",    0,     8_192),
        ("UNIFIED",        1_024, 24_576),
        ("Payslip/SALN/ITR fallback", 512, 16_000),
        ("JSON repair",    0,     4_096),
    ]
    for name, budget, max_tok in cases:
        print(f"  ✅  {name:30s}  thinking_budget={budget:>5}  max_tokens={max_tok:>6,}")

    print()
    print("=" * 60)
    print("TL-5: Text slice caps")
    print("=" * 60)
    print("  ✅  CIC keyword   → filtered_text[:20_000]  (post-filter)")
    print("  ✅  UNIFIED       → each doc sliced inside caller (unchanged)")
    print("  ✅  Individual    → payslip[:20k], saln[:30k], itr[:30k] (unchanged)")

    print()
    print("=" * 60)
    print("CIC-KW-2: Keyword filter still works")
    print("=" * 60)
    sample = (
        "Provider: ABC Bank\nStatus: Active\nBalance: 30000\n\n"
        "Provider: XYZ Finance\nStatus: PAST DUE\nBalance: 95000\n\n"
        "Provider: GHI Lending\nStatus: WRITE OFF\nBalance: 750000\n"
    )
    filtered = _cic_keyword_filter(sample, context_lines=1)
    assert "XYZ Finance" in filtered
    assert "GHI Lending" in filtered
    assert "ABC Bank" not in filtered
    print("  ✅  Active accounts excluded, PAST DUE + WRITE OFF retained")

    print()
    print("=" * 60)
    print("OF-2: TIN formatting")
    print("=" * 60)
    cases = [
        ("12345677890000", "123-456-778-900"),
        ("123456789000",   "123-456-789-000"),
        ("123456789",      "123-456-789"),
        ("123-456-789-0000", "123-456-789-000"),
        ("",               ""),
    ]
    for raw, expected in cases:
        result = _fmt_tin(raw)
        ok = "✅" if result == expected else f"❌ FAIL got {result!r}"
        print(f"  {ok}  _fmt_tin({raw!r}) = {result!r}")

    print()
    print("=" * 60)
    print("OF-4: DOB from ITR fallback")
    print("=" * 60)
    dob = _get_dob({}, {}, {"date_of_birth": "06/15/1985"})
    age = _age_from_dob(dob)
    fmt = _fmt_dob(dob)
    assert dob == "06/15/1985"
    assert age in ("39", "40")
    assert fmt == "15/06/1985"
    print(f"  DOB={dob!r} age={age} fmt={fmt!r}  PASS ✅")

    print()
    print("=" * 60)
    print("OF-7: Real property current value — all key variants")
    print("=" * 60)
    def _num_local(val):
        if val is None: return None
        try: return float(str(val).replace("₱","").replace(",","").strip())
        except: return None

    prop_variants = [
        {"description": "House",  "current_fair_market_value": 850000, "location": "Brgy Lag-on"},
        {"description": "House",  "current_fair_value":        850000, "location": "Brgy Lag-on"},
        {"description": "House",  "current_value":             850000, "location": "Brgy Lag-on"},
        {"description": "House",  "assessed_value":            185000, "location": "Brgy Lag-on"},
    ]
    for p in prop_variants:
        cv = _num_local(
            p.get("current_fair_market_value") or
            p.get("current_fair_value")        or
            p.get("current_value")             or
            p.get("assessed_value")
        )
        key_used = next(k for k in ["current_fair_market_value","current_fair_value",
                                    "current_value","assessed_value"] if p.get(k))
        assert cv is not None, f"FAIL: cv=None for key {key_used}"
        print(f"  ✅  key={key_used!r:35s} → cv={cv:,.0f}")

    print()
    print("=" * 60)
    print("OF-8: Real property location key fallback + year + acq cost")
    print("=" * 60)
    prop_loc = {"location": "Brgy Lag-on", "year_acquired": "2015", "acquisition_cost": 750000}
    prop_loc2 = {"exact_location": "Brgy Lag-on, Daet", "year_acquired": 2015, "acquisition_cost": 750000}
    for p in [prop_loc, prop_loc2]:
        loc = p.get("location") or p.get("exact_location") or ""
        yr  = str(p.get("year_acquired") or "")
        ac  = _num_local(p.get("acquisition_cost"))
        assert loc, f"FAIL: loc empty"
        assert yr == "2015", f"FAIL yr: {yr}"
        assert ac == 750000
        print(f"  ✅  loc={loc!r}  yr={yr}  ac={ac:,.0f}")

    print()
    print("=" * 60)
    print("OF-9: Cash on hand not overwritten by combined")
    print("=" * 60)
    cash_on_hand = 15000.0
    cash_in_bank = 42500.0
    cash_combined = 57500.0
    ci_cash = {}
    if cash_on_hand is not None:
        ci_cash["H58"] = cash_on_hand
    elif cash_combined is not None:
        ci_cash["H58"] = cash_combined
    if cash_in_bank is not None:
        ci_cash["H60"] = cash_in_bank
    assert ci_cash["H58"] == 15000.0, f"FAIL H58: {ci_cash['H58']}"
    assert ci_cash["H60"] == 42500.0, f"FAIL H60: {ci_cash['H60']}"
    print(f"  H58={ci_cash['H58']:,.0f} (cash on hand)  H60={ci_cash['H60']:,.0f} (bank)  PASS ✅")

    ci_cash2 = {}
    coh2 = None; cib2 = None; comb2 = 57500.0
    if coh2 is not None: ci_cash2["H58"] = coh2
    elif comb2 is not None: ci_cash2["H58"] = comb2
    assert ci_cash2["H58"] == 57500.0
    print(f"  Combined fallback: H58={ci_cash2['H58']:,.0f}  PASS ✅")

    print()
    print("=" * 60)
    print("OF-10: Spouse occupation from SALN fallback")
    print("=" * 60)
    cic_no_spouse_occ = {}
    saln_with_spouse = {"spouse_position": "Farmer/Self-employed"}
    spouse_occ_result = (
        cic_no_spouse_occ.get("spouse_occupation") or
        saln_with_spouse.get("spouse_position") or ""
    )
    assert spouse_occ_result == "Farmer/Self-employed", f"FAIL: {spouse_occ_result}"
    print(f"  spouse_occ = {spouse_occ_result!r}  PASS ✅")

    print()
    print("=" * 60)
    print("OF-11: Formula override whitelist includes H31/H32/H33")
    print("=" * 60)
    whitelist = {
        "CIBI::G34", "CIBI::H34",
        "CIBI::H31", "CIBI::H32", "CIBI::H33",
    }
    for cell in ["CIBI::H31", "CIBI::H32", "CIBI::H33"]:
        assert cell in whitelist, f"FAIL: {cell} not in whitelist"
        print(f"  ✅  {cell} in whitelist")

    print()
    print("=" * 60)
    print("OF-12: Row 33 fully merged — balance embedded in A33")
    print("=" * 60)
    _CREDIT_ROW_MERGED = {33}
    ci_merged = {}
    credits_test = [
        {"institution": "PAGIBIG",  "balance": 412000, "principal_loan": None, "monthly_amortization": None, "due_date": "", "status": "Housing Loan", "overdue": None},
        {"institution": "GSIS",     "balance": 185420, "principal_loan": None, "monthly_amortization": None, "due_date": "", "status": "Salary Loan",  "overdue": None},
        {"institution": "LANDBANK", "balance": 52300,  "principal_loan": None, "monthly_amortization": None, "due_date": "", "status": "Salary Loan",  "overdue": None},
    ]
    for i, acc in enumerate(credits_test[:3]):
        r = 31 + i
        institution = acc.get("institution") or ""
        b = _num_local(acc.get("balance"))
        if r in _CREDIT_ROW_MERGED:
            parts = [institution] if institution else []
            if b is not None: parts.append(f"Balance: ₱{b:,.2f}")
            ci_merged[f"A{r}"] = "  |  ".join(parts)
        else:
            if institution: ci_merged[f"A{r}"] = institution
            if b is not None: ci_merged[f"H{r}"] = b

    assert ci_merged.get("A31") == "PAGIBIG",  f"FAIL A31: {ci_merged.get('A31')}"
    assert ci_merged.get("H31") == 412000,     f"FAIL H31: {ci_merged.get('H31')}"
    assert ci_merged.get("A32") == "GSIS",     f"FAIL A32: {ci_merged.get('A32')}"
    assert ci_merged.get("H32") == 185420,     f"FAIL H32: {ci_merged.get('H32')}"
    assert "LANDBANK" in ci_merged.get("A33", ""), f"FAIL A33 no institution: {ci_merged.get('A33')}"
    assert "52,300" in ci_merged.get("A33", ""),   f"FAIL A33 no balance: {ci_merged.get('A33')}"
    assert "H33" not in ci_merged,                 f"FAIL: H33 should not be written"
    print(f"  A31={ci_merged['A31']}  H31={ci_merged['H31']:,.0f}")
    print(f"  A32={ci_merged['A32']}  H32={ci_merged['H32']:,.0f}")
    print(f"  A33={ci_merged['A33']}")
    print("  PASS ✅")

    print()
    print("=" * 60)
    print("OF-13: Real property column remapping for merged cells")
    print("=" * 60)
    prop = {
        "description": "Residential House and Lot",
        "location": "Brgy Lag-on, Daet, Camarines Norte",
        "year_acquired": "2015",
        "acquisition_cost": 750000,
        "current_fair_market_value": 850000,
    }
    ci_prop = {}
    r = 37
    desc = prop.get("description") or ""
    loc  = prop.get("location") or prop.get("exact_location") or ""
    yr   = str(prop.get("year_acquired") or "").strip()
    ac   = _num_local(prop.get("acquisition_cost"))
    cv   = _num_local(
        prop.get("current_fair_market_value") or
        prop.get("current_fair_value")        or
        prop.get("current_value")             or
        prop.get("assessed_value")
    )
    if desc: ci_prop[f"A{r}"] = desc
    if yr and loc:
        ci_prop[f"D{r}"] = f"{loc}  ({yr})"
    elif yr:
        ci_prop[f"D{r}"] = yr
    elif loc:
        ci_prop[f"D{r}"] = loc
    if ac is not None: ci_prop[f"F{r}"] = ac
    if cv is not None: ci_prop[f"H{r}"] = cv

    assert ci_prop.get("A37") == "Residential House and Lot"
    assert "2015" in ci_prop.get("D37", ""),   f"FAIL D37 no year: {ci_prop.get('D37')}"
    assert "Lag-on" in ci_prop.get("D37", ""), f"FAIL D37 no loc: {ci_prop.get('D37')}"
    assert ci_prop.get("F37") == 750000,       f"FAIL F37: {ci_prop.get('F37')}"
    assert ci_prop.get("H37") == 850000,       f"FAIL H37: {ci_prop.get('H37')}"
    assert "E37" not in ci_prop, "FAIL: E37 is inside D:E merge, should not be written"
    assert "G37" not in ci_prop, "FAIL: G37 is inside F:G merge, should not be written"
    print(f"  A37={ci_prop['A37']!r}")
    print(f"  D37={ci_prop['D37']!r}  (location + year, top-left of D:E merge)")
    print(f"  F37={ci_prop['F37']:,.0f}  (acquisition cost, top-left of F:G merge)")
    print(f"  H37={ci_prop['H37']:,.0f}  (current value, top-left of H:I merge)")
    print("  E37 and G37 NOT written (merged cells)  PASS ✅")

    print()
    print("All tests passed! ✅")