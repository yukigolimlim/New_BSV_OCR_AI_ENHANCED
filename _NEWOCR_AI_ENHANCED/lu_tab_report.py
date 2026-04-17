"""
lu_tab_report.py — Report Tab
================================
Plain-text risk report with PDF and Excel export.

Standalone: imports only lu_core and lu_shared.
Attached to app class via attach(cls).

Public surface
--------------
  attach(cls)
  _build_report_panel(self, parent)
  _report_show_placeholder(self)
  _report_render(self)
  _report_print(self)
  _lu_show_export_menu(self)
  _export_pdf(self)
  _export_excel(self)
  _generate_pdf(results, out_path, filepath, client_name, sector_filter)
  _generate_excel(results, out_path, filepath, client_name, sector_filter)
"""

import tkinter as tk
import customtkinter as ctk
from pathlib import Path
from datetime import datetime
from tkinter import filedialog, messagebox

from lu_core import GENERAL_CLIENT, _compute_risk_score
from lu_shared import (
    F, FF,
    _NAVY_DEEP, _NAVY_MID, _NAVY_LIGHT, _NAVY_MIST, _NAVY_GHOST, _NAVY_PALE,
    _WHITE, _CARD_WHITE, _OFF_WHITE, _BORDER_LIGHT, _BORDER_MID,
    _TXT_NAVY, _TXT_SOFT, _TXT_MUTED, _TXT_ON_LIME,
    _LIME_MID, _LIME_DARK, _LIME_PALE,
    _ACCENT_RED, _ACCENT_GOLD, _ACCENT_SUCCESS,
    _SECTOR_COLORS, _SECTOR_ICON,
    _lu_filter_data_by_query,
    _lu_get_active_sectors, _lu_get_filtered_all_data,
)

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, PageBreak,
    )
    _HAS_RL = True
except ImportError:
    _HAS_RL = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPX = True
except ImportError:
    _HAS_OPX = False


# ══════════════════════════════════════════════════════════════════════
#  PANEL BUILDER
# ══════════════════════════════════════════════════════════════════════

def _build_report_panel(self, parent):
    hdr = tk.Frame(parent, bg=_NAVY_MID, height=46)
    hdr.pack(fill="x")
    hdr.pack_propagate(False)
    tk.Label(hdr, text="📄  Risk Analysis Report",
             font=F(10, "bold"), fg=_WHITE, bg=_NAVY_MID).pack(side="left", padx=20, pady=12)
    tk.Label(hdr, text="🔎", font=F(9), fg=_WHITE, bg=_NAVY_MID).pack(side="left", padx=(8, 4))
    self._report_search_var = tk.StringVar()
    self._report_search_var.trace_add(
        "write", lambda *_: _report_render(self) if getattr(self, "_lu_all_data", None) else None)
    tk.Entry(
        hdr, textvariable=self._report_search_var,
        font=F(8), relief="flat", bg=_WHITE, fg=_TXT_NAVY,
        insertbackground=_TXT_NAVY, highlightbackground=_NAVY_LIGHT, highlightthickness=1
    ).pack(side="left", padx=(0, 8), ipady=3)
    self._report_match_lbl = tk.Label(
        hdr, text="", font=F(8, "bold"), fg=_WHITE, bg=_NAVY_MID, padx=8, pady=3)
    self._report_match_lbl.pack(side="left", padx=(0, 8), pady=8)

    ctk.CTkButton(
        hdr, text="🖨  Print",
        command=lambda: _report_print(self),
        width=80, height=30, corner_radius=6,
        fg_color=_NAVY_LIGHT, hover_color=_NAVY_MID,
        text_color=_WHITE, font=FF(8, "bold")
    ).pack(side="right", padx=(0, 8), pady=8)

    ctk.CTkButton(
        hdr, text="📊  Export Excel",
        command=lambda: _export_excel(self),
        width=130, height=30, corner_radius=6,
        fg_color=_LIME_DARK, hover_color=_LIME_MID,
        text_color=_TXT_ON_LIME, font=FF(8, "bold")
    ).pack(side="right", padx=(0, 4), pady=8)

    ctk.CTkButton(
        hdr, text="📄  Export PDF",
        command=lambda: _export_pdf(self),
        width=120, height=30, corner_radius=6,
        fg_color=_LIME_DARK, hover_color=_LIME_MID,
        text_color=_TXT_ON_LIME, font=FF(8, "bold")
    ).pack(side="right", padx=(0, 4), pady=8)

    body = tk.Frame(parent, bg=_CARD_WHITE)
    body.pack(fill="both", expand=True)
    rsb = tk.Scrollbar(body, relief="flat", troughcolor=_OFF_WHITE,
                       bg=_BORDER_LIGHT, width=8, bd=0)
    rsb.pack(side="right", fill="y")
    self._report_text = tk.Text(
        body, font=("Consolas", 9), fg=_TXT_NAVY, bg=_WHITE,
        relief="flat", bd=0, padx=28, pady=20, wrap="word",
        yscrollcommand=rsb.set, state="disabled")
    self._report_text.pack(side="left", fill="both", expand=True)
    rsb.config(command=self._report_text.yview)

    t = self._report_text
    t.tag_configure("title",         font=("Consolas", 13, "bold"), foreground=_NAVY_DEEP)
    t.tag_configure("h1",            font=("Consolas", 11, "bold"), foreground=_NAVY_MID)
    t.tag_configure("h2",            font=("Consolas", 10, "bold"), foreground=_NAVY_LIGHT)
    t.tag_configure("high",          font=("Consolas",  9, "bold"), foreground=_ACCENT_RED)
    t.tag_configure("mod",           font=("Consolas",  9, "bold"), foreground=_ACCENT_GOLD)
    t.tag_configure("low",           font=("Consolas",  9, "bold"), foreground=_ACCENT_SUCCESS)
    t.tag_configure("muted",         font=("Consolas",  8),         foreground=_TXT_MUTED)
    t.tag_configure("normal",        font=("Consolas",  9),         foreground=_TXT_NAVY)
    t.tag_configure("rule",          font=("Consolas",  8),         foreground=_BORDER_MID)
    t.tag_configure("client_name",   font=("Consolas", 11, "bold"), foreground=_NAVY_MID)
    t.tag_configure("general_tag",   font=("Consolas",  9, "bold"), foreground=_LIME_DARK)
    t.tag_configure("perclient_tag", font=("Consolas",  9, "bold"), foreground="#8B5CF6")
    t.tag_configure("sector_tag",    font=("Consolas",  9, "bold"), foreground=_LIME_MID)
    for key, fg in [("score_critical", "#B71C1C"), ("score_high", _ACCENT_RED),
                    ("score_moderate", _ACCENT_GOLD), ("score_low", _ACCENT_SUCCESS)]:
        t.tag_configure(key, font=("Consolas", 9, "bold"), foreground=fg)
    _report_show_placeholder(self)


# ══════════════════════════════════════════════════════════════════════
#  PLACEHOLDER + RENDERER
# ══════════════════════════════════════════════════════════════════════

def _report_show_placeholder(self):
    self._report_text.config(state="normal")
    self._report_text.delete("1.0", "end")
    self._report_text.insert("end", "Run an analysis first to generate the report.", "muted")
    self._report_text.config(state="disabled")


def _report_render(self):
    t              = self._report_text
    client         = self._lu_active_client
    is_general     = (client == GENERAL_CLIENT)
    all_data       = self._lu_all_data
    filtered_data  = _lu_get_filtered_all_data(self)
    q              = getattr(self, "_report_search_var", tk.StringVar(value="")).get().strip()
    filtered_data  = _lu_filter_data_by_query(filtered_data, q)
    match_lbl      = getattr(self, "_report_match_lbl", None)
    if match_lbl is not None:
        if q:
            rows = filtered_data.get("general", [])
            client_names = sorted({(r.get("client") or "").strip() for r in rows if r.get("client")})
            if len(client_names) == 1:
                match_lbl.config(text=client_names[0][:28], bg="#4A6FA5")
            else:
                match_lbl.config(text=f"{len(rows)} CLIENTS MATCHED", bg="#4A6FA5")
        else:
            match_lbl.config(text="", bg=_NAVY_MID)
    active_sectors = _lu_get_active_sectors(self)

    if is_general:
        results = filtered_data.get("general", [])
    else:
        rec = all_data.get("clients", {}).get(client)
        results = [rec] if (rec and (not q or _lu_filter_data_by_query({"general": [rec]}, q).get("general"))) else []

    now   = datetime.now().strftime("%B %d, %Y  %H:%M")
    fname = Path(self._lu_filepath).name if self._lu_filepath else "—"
    rule  = "─" * 78 + "\n"
    dbl   = "═" * 78 + "\n"
    RISK_TAG = {"HIGH": "high", "MODERATE": "mod", "LOW": "low"}

    t.config(state="normal")
    t.delete("1.0", "end")
    t.insert("end", dbl, "rule")
    t.insert("end", "  LU RISK ANALYSIS REPORT\n", "title")
    t.insert("end", dbl, "rule")
    t.insert("end", f"  File     : {fname}\n", "normal")
    t.insert("end", f"  Generated: {now}\n", "normal")

    if is_general:
        totals = filtered_data.get("totals", {})
        if q:
            t.insert("end", "  Mode     : ", "normal")
            t.insert("end", f"SEARCH — {q}\n", "sector_tag")
        elif active_sectors:
            t.insert("end", "  Mode     : ", "normal")
            t.insert("end", f"SECTOR FILTER — {' · '.join(active_sectors)}\n", "sector_tag")
        else:
            t.insert("end", "  Mode     : ", "normal")
            t.insert("end", "GENERAL VIEW — All Clients\n", "general_tag")
        t.insert("end", f"  Clients  : {len(results)}\n", "normal")
        t.insert("end", f"  Total Loan Balance : ₱{totals.get('loan_balance',0):,.2f}\n", "normal")
        t.insert("end", f"  Total Net Income   : ₱{totals.get('total_net',0):,.2f}\n", "normal")
    else:
        rec = results[0] if results else {}
        t.insert("end", "  Mode     : ", "normal")
        t.insert("end", "PER-CLIENT VIEW\n", "perclient_tag")
        t.insert("end", "  Client   : ", "normal")
        t.insert("end", f"{rec.get('client','—')}\n", "client_name")
        t.insert("end", f"  Client ID: {rec.get('client_id','—')}\n", "normal")
        t.insert("end", f"  PN       : {rec.get('pn','—')}\n", "normal")
        t.insert("end", f"  Industry : {rec.get('industry','—')}\n", "normal")
        t.insert("end", f"  Sector   : {rec.get('sector','—')}\n", "normal")
        t.insert("end", f"  Loan Bal : ₱{rec.get('loan_balance') or 0:,.2f}\n", "normal")
        t.insert("end", f"  Net Inc  : ₱{rec.get('net_income') or 0:,.2f}\n", "normal")
        t.insert("end", "  Risk Score: ", "normal")
        lbl = rec.get("score_label", "N/A")
        t.insert("end", f"{lbl}  ({rec.get('score',0):.2f})\n", f"score_{lbl.lower()}")

    t.insert("end", dbl, "rule")
    t.insert("end", "\n")

    for rec in results:
        icon = _SECTOR_ICON.get(rec.get("sector", ""), "📋")
        t.insert("end", rule, "rule")
        t.insert("end", f"  {icon}  CLIENT: ", "h1")
        t.insert("end", f"{rec['client']}\n", "client_name")
        t.insert("end", f"  Industry: {rec.get('industry','—')}  |  Sector: {rec.get('sector','—')}\n", "normal")
        t.insert("end", f"  ID: {rec.get('client_id','—')}  PN: {rec.get('pn','—')}\n", "normal")
        t.insert("end",
                 f"  Total Source Income: ₱{rec.get('total_source') or 0:,.2f}  "
                 f"Net Income: ₱{rec.get('net_income') or 0:,.2f}  "
                 f"Loan Balance: ₱{rec.get('loan_balance') or 0:,.2f}\n", "normal")
        t.insert("end", rule, "rule")

        exps = rec.get("expenses", [])
        client_label = str(rec.get("score_label") or "").upper()
        # Use effective risk for display so UI stays consistent with final
        # precedence (Product > Expense > Industry).
        effective = lambda e: "HIGH" if client_label == "HIGH" else str(e.get("risk") or "LOW").upper()
        h = sum(1 for e in exps if effective(e) == "HIGH")
        m = sum(1 for e in exps if effective(e) == "MODERATE")
        l = sum(1 for e in exps if effective(e) == "LOW")
        t.insert("end", "  Risk summary: ", "normal")
        t.insert("end", f"HIGH×{h} ", "high")
        t.insert("end", f"MODERATE×{m} ", "mod")
        t.insert("end", f"LOW×{l}\n\n", "low")

        col_w = [26, 10, 44]
        t.insert("end", f"  {'EXPENSE ITEM':<{col_w[0]}} {'RISK':<{col_w[1]}} {'IMPACT REASON'}\n", "h2")
        t.insert("end", "  " + "─" * 76 + "\n", "rule")
        for exp in exps:
            exp_risk = effective(exp)
            tag  = RISK_TAG.get(exp_risk, "normal")
            name = exp["name"] if len(exp["name"]) <= col_w[0] else exp["name"][:col_w[0]-1] + "…"
            rsn  = exp["reason"][:col_w[2]] if len(exp["reason"]) > col_w[2] else exp["reason"]
            t.insert("end", f"  {name:<{col_w[0]}} {exp_risk:<{col_w[1]}} {rsn}\n", tag)
            t.insert("end", "  " + " " * (col_w[0]+col_w[1]+1) + f"↳ {exp['value_str']}\n", "muted")
        t.insert("end", "\n")

    t.insert("end", dbl, "rule")
    t.insert("end", "  END OF REPORT\n", "muted")
    t.insert("end", dbl, "rule")
    t.config(state="disabled")
    t.yview_moveto(0)


def _report_print(self):
    try:
        import subprocess, platform, tempfile, os
        content = self._report_text.get("1.0", "end")
        with tempfile.NamedTemporaryFile(mode="w", suffix=".txt",
                                         delete=False, encoding="utf-8") as f:
            f.write(content)
            tmp = f.name
        if platform.system() == "Windows":
            os.startfile(tmp, "print")
        else:
            subprocess.run(["lpr", tmp])
    except Exception as ex:
        messagebox.showerror("Print Error", str(ex))


# ══════════════════════════════════════════════════════════════════════
#  EXPORT MENU  (top-bar Export button)
# ══════════════════════════════════════════════════════════════════════

def _lu_show_export_menu(self):
    client         = self._lu_active_client
    is_general     = (client == GENERAL_CLIENT)
    active_sectors = _lu_get_active_sectors(self)

    if active_sectors:
        label_pdf = f"📄  Export PDF — Sector: {' · '.join(active_sectors)[:35]}"
        label_xl  = f"📊  Export Excel — Sector: {' · '.join(active_sectors)[:35]}"
    elif is_general:
        label_pdf = "📄  Export General PDF report"
        label_xl  = "📊  Export General Excel workbook"
    else:
        label_pdf = f"📄  Export PDF — {client}"
        label_xl  = f"📊  Export Excel — {client}"

    menu = tk.Menu(self._lu_analysis_frame, tearoff=0,
                   font=F(9), bg=_WHITE, fg=_TXT_NAVY,
                   activebackground=_NAVY_GHOST, activeforeground=_NAVY_DEEP,
                   relief="flat", bd=1)
    menu.add_command(label=label_pdf, command=lambda: _export_pdf(self))
    menu.add_command(label=label_xl,  command=lambda: _export_excel(self))
    menu.add_separator()
    menu.add_command(label="🖨  Print report", command=lambda: _report_print(self))
    try:
        menu.tk_popup(
            self._lu_export_btn.winfo_rootx(),
            self._lu_export_btn.winfo_rooty() + self._lu_export_btn.winfo_height())
    finally:
        menu.grab_release()


# ══════════════════════════════════════════════════════════════════════
#  PDF EXPORT
# ══════════════════════════════════════════════════════════════════════

def _rgb_hex(color) -> str:
    try:
        return f"{int(color.red*255):02X}{int(color.green*255):02X}{int(color.blue*255):02X}"
    except Exception:
        return "000000"


def _export_pdf(self):
    if not self._lu_all_data:
        messagebox.showwarning("No Data", "Run an analysis first.")
        return
    if not _HAS_RL:
        messagebox.showerror("Missing Library",
                             "reportlab is not installed.\nRun:  pip install reportlab")
        return
    client         = self._lu_active_client
    is_general     = (client == GENERAL_CLIENT)
    active_sectors = _lu_get_active_sectors(self)
    q              = getattr(self, "_report_search_var", tk.StringVar(value="")).get().strip()

    if q:
        default_name = f"LU_Risk_Search_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    elif active_sectors:
        slug = "_".join(s.replace("/", "_").replace(" ", "_") for s in active_sectors)
        default_name = f"LU_Risk_Sector_{slug}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    elif is_general:
        default_name = f"LU_Risk_General_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    else:
        default_name = f"LU_Risk_{client.replace(' ','_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"

    path = filedialog.asksaveasfilename(
        title="Save PDF Report", defaultextension=".pdf",
        filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
        initialfile=default_name)
    if not path:
        return
    try:
        filtered_data = _lu_get_filtered_all_data(self)
        filtered_data = _lu_filter_data_by_query(filtered_data, q)
        if is_general or active_sectors:
            results = filtered_data.get("general", [])
            cn = None
        else:
            all_data = self._lu_all_data
            rec = all_data.get("clients", {}).get(client)
            results = [rec] if (rec and (not q or _lu_filter_data_by_query({"general": [rec]}, q).get("general"))) else []
            cn = client
        _generate_pdf(results, path,
                      filepath=self._lu_filepath or "",
                      client_name=cn,
                      sector_filter=active_sectors)
        messagebox.showinfo("Export Complete", f"PDF saved to:\n{path}")
    except Exception as ex:
        messagebox.showerror("PDF Export Error", str(ex))


def _generate_pdf(results, out_path, filepath="", client_name=None, sector_filter=None):
    styles  = getSampleStyleSheet()
    doc     = SimpleDocTemplate(out_path, pagesize=A4,
                                leftMargin=2*cm, rightMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
    navy    = rl_colors.HexColor("#1A3A6B")
    red     = rl_colors.HexColor("#E53E3E")
    gold    = rl_colors.HexColor("#D4A017")
    green   = rl_colors.HexColor("#2E7D32")
    white   = rl_colors.white
    off     = rl_colors.HexColor("#F5F7FA")
    border  = rl_colors.HexColor("#C5D0E8")
    crimson = rl_colors.HexColor("#B71C1C")

    title_style  = ParagraphStyle("LUTitle",  parent=styles["Title"],    fontSize=18, textColor=navy, spaceAfter=4)
    h1_style     = ParagraphStyle("LUH1",     parent=styles["Heading1"], fontSize=13, textColor=white, spaceAfter=4)
    body_style   = ParagraphStyle("LUBody",   parent=styles["Normal"],   fontSize=8,  leading=11,
                                  textColor=rl_colors.HexColor("#1A2B4A"))
    muted_style  = ParagraphStyle("LUMuted",  parent=styles["Normal"],   fontSize=7,  leading=10,
                                  textColor=rl_colors.HexColor("#9AAACE"))
    client_style = ParagraphStyle("LUClient", parent=styles["Normal"],   fontSize=13, leading=16, textColor=navy, spaceAfter=2)
    mode_style   = ParagraphStyle("LUMode",   parent=styles["Normal"],   fontSize=9,  leading=12,
                                  textColor=rl_colors.HexColor("#5A9E28"))

    RISK_COLOR_RL  = {"HIGH": red,    "MODERATE": gold,    "LOW": green}
    SCORE_COLOR_RL = {"CRITICAL": crimson, "HIGH": red, "MODERATE": gold, "LOW": green}

    story = [Paragraph("LU Risk Analysis Report", title_style)]
    now   = datetime.now().strftime("%B %d, %Y  %H:%M")
    fname = Path(filepath).name if filepath else "—"

    if sector_filter:
        story.append(Paragraph(f"SECTOR FILTER — {' · '.join(sector_filter)}", mode_style))
        story.append(Paragraph(f"Clients in filter: {len(results)}", body_style))
    elif client_name and results:
        rec = results[0]
        story.append(Paragraph("PER-CLIENT VIEW", mode_style))
        story.append(Paragraph(f"<b>Client: {client_name}</b>", client_style))
        story.append(Paragraph(
            f"Industry: {rec.get('industry','—')}  |  Sector: {rec.get('sector','—')}",
            body_style))
        story.append(Paragraph(
            f"Loan Balance: ₱{rec.get('loan_balance') or 0:,.2f}   "
            f"Net Income: ₱{rec.get('net_income') or 0:,.2f}",
            body_style))
        score, label, _, _ = _compute_risk_score(
            rec.get("expenses", []),
            str(rec.get("industry") or ""),
            str(rec.get("product_name") or ""),
        )
        sc = SCORE_COLOR_RL.get(label, green)
        story.append(Paragraph(
            f"Risk Score: <font color='#{_rgb_hex(sc)}'><b>{label} ({score:.2f})</b></font>",
            body_style))
    else:
        story.append(Paragraph("GENERAL VIEW — All Clients", mode_style))

    story += [
        Paragraph(f"File: {fname}    Generated: {now}", muted_style),
        Spacer(1, 0.3*cm),
        HRFlowable(width="100%", thickness=1, color=border),
        Spacer(1, 0.4*cm),
    ]

    for rec in results:
        story.append(PageBreak())
        icon       = _SECTOR_ICON.get(rec.get("sector", ""), "")
        client_lbl = rec.get("client", "—")
        hdr_tbl    = Table([[Paragraph(f"{icon}  {client_lbl}", h1_style)]],
                           colWidths=[17*cm])
        hdr_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), navy),
            ("TOPPADDING",    (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("LEFTPADDING",   (0, 0), (-1, -1), 10),
        ]))
        story += [hdr_tbl, Spacer(1, 0.1*cm)]
        story.append(Paragraph(
            f"Industry: {rec.get('industry','—')}  |  Sector: {rec.get('sector','—')}  |  "
            f"ID: {rec.get('client_id','—')}  PN: {rec.get('pn','—')}",
            muted_style))
        story.append(Paragraph(
            f"Total Source: ₱{rec.get('total_source') or 0:,.2f}   "
            f"Net Income: ₱{rec.get('net_income') or 0:,.2f}   "
            f"Loan Balance: ₱{rec.get('loan_balance') or 0:,.2f}   "
            f"Current Amort: ₱{rec.get('current_amort') or 0:,.2f}",
            body_style))
        story.append(Spacer(1, 0.2*cm))

        col_w    = [4.5*cm, 2.2*cm, 6.3*cm, 4*cm]
        tbl_data = [[Paragraph(f"<b>{h}</b>", body_style)
                     for h in ["Expense Item", "Risk", "Impact Reason", "Value / Amount"]]]
        client_label = str(rec.get("score_label") or "").upper()
        for exp in rec.get("expenses", []):
            exp_risk = "HIGH" if client_label == "HIGH" else str(exp.get("risk") or "LOW").upper()
            rc = RISK_COLOR_RL.get(exp_risk, green)
            tbl_data.append([
                Paragraph(f"<b>{exp['name']}</b>", body_style),
                Paragraph(
                    f"<font color='#{_rgb_hex(rc)}'><b>{exp_risk}</b></font>",
                    body_style),
                Paragraph(exp["reason"], body_style),
                Paragraph(
                    f"<font color='#9AAACE'><i>{exp['value_str']}</i></font>"
                    if not exp["has_values"] else exp["value_str"],
                    body_style),
            ])
        tbl = Table(tbl_data, colWidths=col_w)
        tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0),  navy),
            ("FONTSIZE",      (0, 0), (-1, -1), 8),
            ("LEADING",       (0, 0), (-1, -1), 11),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING",   (0, 0), (-1, -1), 6),
            ("VALIGN",        (0, 0), (-1, -1), "TOP"),
            ("BOX",           (0, 0), (-1, -1), 0.5, border),
            ("INNERGRID",     (0, 0), (-1, -1), 0.3, border),
            *[("BACKGROUND",  (0, i), (-1, i), off if i % 2 == 0 else white)
              for i in range(1, len(tbl_data))],
        ]))
        story += [tbl, Spacer(1, 0.3*cm)]

    doc.build(story)


# ══════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════

def _export_excel(self):
    if not self._lu_all_data:
        messagebox.showwarning("No Data", "Run an analysis first.")
        return
    if not _HAS_OPX:
        messagebox.showerror("Missing Library",
                             "openpyxl is not installed.\nRun:  pip install openpyxl")
        return
    client         = self._lu_active_client
    is_general     = (client == GENERAL_CLIENT)
    active_sectors = _lu_get_active_sectors(self)
    q              = getattr(self, "_report_search_var", tk.StringVar(value="")).get().strip()

    if q:
        default_name = f"LU_Risk_Search_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    elif active_sectors:
        slug = "_".join(s.replace("/", "_").replace(" ", "_") for s in active_sectors)
        default_name = f"LU_Risk_Sector_{slug}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    elif is_general:
        default_name = f"LU_Risk_General_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    else:
        default_name = f"LU_Risk_{client.replace(' ','_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    path = filedialog.asksaveasfilename(
        title="Save Excel Report", defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialfile=default_name)
    if not path:
        return
    try:
        filtered_data = _lu_get_filtered_all_data(self)
        filtered_data = _lu_filter_data_by_query(filtered_data, q)
        if is_general or active_sectors:
            results = filtered_data.get("general", [])
            cn = None
        else:
            all_data = self._lu_all_data
            rec = all_data.get("clients", {}).get(client)
            results = [rec] if (rec and (not q or _lu_filter_data_by_query({"general": [rec]}, q).get("general"))) else []
            cn = client
        _generate_excel(results, path,
                        filepath=self._lu_filepath or "",
                        client_name=cn,
                        sector_filter=active_sectors)
        messagebox.showinfo("Export Complete", f"Excel saved to:\n{path}")
    except Exception as ex:
        messagebox.showerror("Excel Export Error", str(ex))


def _generate_excel(results, out_path, filepath="", client_name=None, sector_filter=None):
    wb  = openpyxl.Workbook()
    now = datetime.now().strftime("%B %d, %Y  %H:%M")
    fname = Path(filepath).name if filepath else "—"

    RISK_FONT_COLOR = {"HIGH": "E53E3E", "MODERATE": "D4A017", "LOW": "2E7D32"}

    def fill(hex_str):
        return PatternFill("solid", fgColor=hex_str.lstrip("#"))

    FILLS = {
        "nav_hdr":  fill("1A3A6B"),
        "col_hdr":  fill("EEF3FB"),
        "HIGH":     fill("FFF5F5"),
        "MODERATE": fill("FFFBF0"),
        "LOW":      fill("F0FBE8"),
    }

    def border_thin():
        s = Side(style="thin", color="C5D0E8")
        return Border(left=s, right=s, top=s, bottom=s)

    for rec in results:
        safe_name = rec.get("client", "Client")[:28].replace("/", "-").replace("\\", "-")
        ws = wb.create_sheet(title=safe_name)
        if wb.active and not wb.active.title:
            wb.remove(wb.active)

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 48
        ws.column_dimensions["D"].width = 24

        ws.merge_cells("A1:D1")
        icon  = _SECTOR_ICON.get(rec.get("sector", ""), "")
        ws["A1"] = f"{icon}  {rec.get('client','')}"
        ws["A1"].fill      = FILLS["nav_hdr"]
        ws["A1"].font      = Font(bold=True, color="FFFFFF", size=12)
        ws["A1"].alignment = Alignment(vertical="center", horizontal="left", indent=1)

        ws.merge_cells("A2:D2")
        ws["A2"] = (f"Loan Balance: ₱{rec.get('loan_balance') or 0:,.2f}   "
                    f"Net Income: ₱{rec.get('net_income') or 0:,.2f}   "
                    f"Risk: {rec.get('score_label','N/A')}")
        ws["A2"].fill      = fill("EEF3FB")
        ws["A2"].font      = Font(size=9, color="1A3A6B")
        ws["A2"].alignment = Alignment(vertical="center", horizontal="left", indent=1)

        for ci, hdr_text in enumerate(
                ["Expense Item", "Risk Level", "Impact Reason", "Value / Amount"], 1):
            c = ws.cell(3, ci, hdr_text)
            c.fill      = FILLS["col_hdr"]
            c.font      = Font(bold=True, size=9, color="4A6FA5")
            c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
            c.border    = border_thin()

        client_label = str(rec.get("score_label") or "").upper()
        for idx, exp in enumerate(rec.get("expenses", []), start=4):
            exp_risk = "HIGH" if client_label == "HIGH" else str(exp.get("risk") or "LOW").upper()
            row_fill = FILLS.get(exp_risk, FILLS["LOW"])
            risk_col = RISK_FONT_COLOR.get(exp_risk, "2E7D32")
            for ci, (val, bold, color, halign, wrap) in enumerate([
                (exp["name"],       True,  "1A2B4A", "left",   False),
                (exp_risk,          True,  risk_col, "center", False),
                (exp["reason"],     False, "6B7FA3", "left",   True),
                (exp["value_str"],  False,
                 "9AAACE" if not exp["has_values"] else "1A2B4A", "left", True),
            ], 1):
                c = ws.cell(idx, ci, val)
                c.fill      = row_fill
                c.font      = Font(bold=bold, size=9 if ci < 3 else 8, color=color,
                                   italic=(ci == 4 and not exp["has_values"]))
                c.alignment = Alignment(vertical="top", horizontal=halign,
                                        wrap_text=wrap, indent=1)
                c.border    = border_thin()
            ws.row_dimensions[idx].height = 30

    # Remove default empty sheet if results exist
    if results and "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    elif not results:
        ws = wb.active
        ws.title = "No Data"
        ws["A1"] = "No results to export."

    wb.save(out_path)


# ══════════════════════════════════════════════════════════════════════
#  ATTACH
# ══════════════════════════════════════════════════════════════════════

def attach(cls):
    """Attach Report-tab and export methods to the app class."""
    cls._build_report_panel    = _build_report_panel
    cls._report_show_placeholder = _report_show_placeholder
    cls._report_render         = _report_render
    cls._report_print          = _report_print
    cls._lu_show_export_menu   = _lu_show_export_menu
    cls._export_pdf            = _export_pdf
    cls._export_excel          = _export_excel
    cls._generate_pdf          = staticmethod(_generate_pdf)
    cls._generate_excel        = staticmethod(_generate_excel)

    # Kept for backward compatibility with lu_analysis_tab shim
    cls._generate_pdf          = staticmethod(_generate_pdf)
