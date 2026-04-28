"""
admin_logs.py — DocExtract Pro (Banco San Vicente)
===================================================
Admin Logs tab — displays all rows from the `logs` table in PostgreSQL.
Only visible to users with role = 'super admin'.

Columns: id, user_id, email, action, description, time

Attach via:
    import admin_logs as _admin_logs_mod
    _admin_logs_mod.attach(DocExtractorApp)   # inside ui_panels.attach()
"""

import tkinter as tk
import customtkinter as ctk
from datetime import datetime

# ── Design tokens (mirrors your existing app palette) ─────────────────────────
_SB_BG         = "#0B1622"
_SB_ACCENT     = "#5BBF3E"
_SB_ACCENT2    = "#7DD65C"
_CARD_BG       = "#FFFFFF"

NAVY_DEEP      = "#0B1F3A"
NAVY_LIGHT     = "#1E3A5F"
NAVY_MID       = "#2E5C8A"
NAVY_MIST      = "#EAF0F8"
NAVY_GHOST     = "#D6E4F0"
TXT_NAVY       = "#1A2E42"
TXT_SOFT       = "#7A94B0"
TXT_MUTED      = "#A8BCCE"
WHITE          = "#FFFFFF"
OFF_WHITE      = "#F8FAFC"
ACCENT_RED     = "#D94040"
ACCENT_GOLD    = "#C89A2E"
ACCENT_SUCCESS = "#2E7D4F"
BORDER_LIGHT   = "#E2EAF4"
BORDER_MID     = "#C8D8EC"
LIME_DARK      = "#3D8F26"

# Action badge colours  {keyword_in_action: (bg, fg)}
_ACTION_COLORS = {
    "login":    ("#E3F5E8", "#1A6B35"),
    "logout":   ("#FFF4E0", "#8A5A00"),
    "create":   ("#E8F0FF", "#1A3A8A"),
    "update":   ("#EEF0FF", "#3A2E8A"),
    "delete":   ("#FFE8E8", "#8A1A1A"),
    "export":   ("#E0F7FA", "#005F6B"),
    "upload":   ("#E8F5E9", "#1B5E20"),
    "error":    ("#FFE8E8", "#8A1A1A"),
    "view":     ("#F3E8FF", "#5A1A8A"),
    "search":   ("#E8F4FF", "#0A4A7A"),
}
_ACTION_DEFAULT = ("#EDF1F7", "#3A4A5E")

_PAGE_SIZE = 20

# Column config: (key, header, pixel_width, anchor)
# pixel_width = 0 means "expand to fill remaining space"
_COLUMNS = [
    ("id",          "ID",          55,  "center"),
    ("user_id",     "User ID",     72,  "center"),
    ("email",       "Email",       210, "w"),
    ("action",      "Action",      120, "center"),
    ("description", "Description", 0,   "w"),
    ("time",        "Timestamp",   162, "center"),
]


# ─────────────────────────────────────────────────────────────────────────────
#  PANEL BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def _build_logs_panel(self, parent):
    BG = _CARD_BG
    frame = tk.Frame(parent, bg=BG)
    self._logs_frame = frame

    # ── Internal state ────────────────────────────────────────────────
    self._logs_all_rows  = []
    self._logs_filtered  = []
    self._logs_page      = 0
    self._logs_sort_col  = "id"
    self._logs_sort_asc  = False   # newest first by default

    # ── TOP HEADER ───────────────────────────────────────────────────
    top = tk.Frame(frame, bg=BG)
    top.pack(fill="x", padx=28, pady=(20, 0))

    left_hdr = tk.Frame(top, bg=BG)
    left_hdr.pack(side="left", fill="y")

    tk.Frame(left_hdr, bg=_SB_ACCENT, width=4).pack(side="left", fill="y", padx=(0, 14))

    title_col = tk.Frame(left_hdr, bg=BG)
    title_col.pack(side="left")
    tk.Label(title_col, text="System Logs",
             font=_F(self, 18, "bold"), fg=NAVY_DEEP, bg=BG).pack(anchor="w")
    tk.Label(title_col, text="All recorded activity  •  Super Admin only",
             font=_F(self, 8), fg=TXT_SOFT, bg=BG).pack(anchor="w", pady=(1, 0))

    right_hdr = tk.Frame(top, bg=BG)
    right_hdr.pack(side="right", fill="y", pady=4)

    self._logs_count_lbl = tk.Label(
        right_hdr, text="— records",
        font=_F(self, 8, "bold"),
        fg=_SB_ACCENT2, bg="#0A1E0A",
        padx=14, pady=5
    )
    self._logs_count_lbl.pack(side="right", padx=(8, 0))

    ctk.CTkButton(
        right_hdr, text="📊  Generate Report",
        command=lambda: _export_logs_excel(self),
        width=140, height=32, corner_radius=6,
        fg_color=LIME_DARK, hover_color=_SB_ACCENT,
        text_color=WHITE, font=_F(self, 9, "bold"),
        border_width=0
    ).pack(side="right", padx=(0, 8))

    ctk.CTkButton(
        right_hdr, text="↻  Refresh",
        command=lambda: _load_logs(self),
        width=100, height=32, corner_radius=6,
        fg_color=NAVY_LIGHT, hover_color=NAVY_MID,
        text_color=WHITE, font=_F(self, 9, "bold"),
        border_width=0
    ).pack(side="right")

    # ── STATS CARDS ──────────────────────────────────────────────────
    stats_row = tk.Frame(frame, bg=BG)
    stats_row.pack(fill="x", padx=28, pady=(16, 0))

    self._logs_stat_lbls = {}
    _stats_cfg = [
        ("total",   "Total Logs",   "📋", "#2E5C8A"),
        ("today",   "Today",        "📅", "#3D8F26"),
        ("users",   "Unique Users", "👤", "#8A5A00"),
        ("actions", "Action Types", "⚡", "#5A1A8A"),
    ]
    for key, label, icon, accent_col in _stats_cfg:
        card = tk.Frame(stats_row, bg=NAVY_DEEP,
                        highlightbackground="#1E3A5F", highlightthickness=1)
        card.pack(side="left", fill="x", expand=True, padx=(0, 10))

        tk.Frame(card, bg=accent_col, height=3).pack(fill="x")

        inner = tk.Frame(card, bg=NAVY_DEEP)
        inner.pack(fill="x", padx=14, pady=10)

        tk.Label(inner, text=f"{icon}  {label}",
                 font=_F(self, 7), fg="#4A6E9A", bg=NAVY_DEEP).pack(anchor="w")

        val_lbl = tk.Label(inner, text="—",
                           font=_F(self, 16, "bold"), fg=WHITE, bg=NAVY_DEEP)
        val_lbl.pack(anchor="w")
        self._logs_stat_lbls[key] = val_lbl

    # ── FILTER BAR ───────────────────────────────────────────────────
    fbar = tk.Frame(frame, bg=BG)
    fbar.pack(fill="x", padx=28, pady=(16, 0))

    search_wrap = tk.Frame(fbar, bg=WHITE,
                           highlightbackground=BORDER_MID, highlightthickness=1)
    search_wrap.pack(side="left", fill="x", expand=True, padx=(0, 10))

    tk.Label(search_wrap, text="🔍", font=("Segoe UI Emoji", 10),
             bg=WHITE, fg=TXT_SOFT).pack(side="left", padx=(10, 4))

    self._logs_search_var = tk.StringVar()
    _PLACEHOLDER = "Search by email, action, description…"
    se = tk.Entry(search_wrap, textvariable=self._logs_search_var,
                  font=_F(self, 10), fg=TXT_MUTED, bg=WHITE,
                  relief="flat", bd=0, width=32, insertbackground=NAVY_MID)
    se.pack(side="left", fill="x", expand=True, pady=7)
    se.insert(0, _PLACEHOLDER)

    def _fi(e):
        if se.get() == _PLACEHOLDER:
            se.delete(0, "end")
            se.config(fg=TXT_NAVY)
    def _fo(e):
        if not se.get().strip():
            se.delete(0, "end")
            se.insert(0, _PLACEHOLDER)
            se.config(fg=TXT_MUTED)
    se.bind("<FocusIn>",  _fi)
    se.bind("<FocusOut>", _fo)
    self._logs_search_var.trace_add("write", lambda *a: _apply_filter(self))

    clr = tk.Label(search_wrap, text="✕", font=_F(self, 9, "bold"),
                   fg=ACCENT_RED, bg=WHITE, cursor="hand2", padx=10)
    clr.pack(side="right")

    def _clr_search():
        self._logs_search_var.set("")
        se.delete(0, "end")
        se.insert(0, _PLACEHOLDER)
        se.config(fg=TXT_MUTED)
    clr.bind("<Button-1>", lambda e: _clr_search())

    self._logs_action_var = tk.StringVar(value="All Actions")
    self._logs_action_menu = ctk.CTkOptionMenu(
        fbar,
        variable=self._logs_action_var,
        values=["All Actions"],
        command=lambda v: _apply_filter(self),
        width=150, height=34, corner_radius=6,
        fg_color=NAVY_DEEP, button_color=NAVY_LIGHT,
        button_hover_color=NAVY_MID,
        dropdown_fg_color=NAVY_DEEP,
        dropdown_hover_color="#162438",
        text_color=WHITE, dropdown_text_color=WHITE,
        font=_F(self, 9),
    )
    self._logs_action_menu.pack(side="left", padx=(0, 10))

    self._logs_page_info = tk.Label(fbar, text="",
                                    font=_F(self, 8), fg=TXT_MUTED, bg=BG)
    self._logs_page_info.pack(side="right")

    # ── TABLE ────────────────────────────────────────────────────────
    tbl_wrap = tk.Frame(frame, bg=BORDER_LIGHT, padx=1, pady=1)
    tbl_wrap.pack(fill="both", expand=True, padx=28, pady=(12, 0))

    tbl_card = tk.Frame(tbl_wrap, bg=WHITE)
    tbl_card.pack(fill="both", expand=True)

    # Header
    self._logs_hdr_row = tk.Frame(tbl_card, bg=NAVY_DEEP, height=40)
    self._logs_hdr_row.pack(fill="x")
    self._logs_hdr_row.pack_propagate(False)
    _build_header_row(self)

    # Scrollable body
    body_outer = tk.Frame(tbl_card, bg=WHITE)
    body_outer.pack(fill="both", expand=True)

    vsb = tk.Scrollbar(body_outer, orient="vertical", relief="flat",
                       troughcolor=OFF_WHITE, bg=BORDER_LIGHT, width=8, bd=0)
    vsb.pack(side="right", fill="y")

    self._logs_canvas = tk.Canvas(body_outer, bg=WHITE,
                                  highlightthickness=0, yscrollcommand=vsb.set)
    self._logs_canvas.pack(side="left", fill="both", expand=True)
    vsb.config(command=self._logs_canvas.yview)

    self._logs_body = tk.Frame(self._logs_canvas, bg=WHITE)
    _cw = self._logs_canvas.create_window((0, 0), window=self._logs_body, anchor="nw")

    def _body_cfg(e):
        self._logs_canvas.configure(scrollregion=self._logs_canvas.bbox("all"))
    def _canvas_cfg(e):
        self._logs_canvas.itemconfig(_cw, width=e.width)
    self._logs_body.bind("<Configure>", _body_cfg)
    self._logs_canvas.bind("<Configure>", _canvas_cfg)

    def _mw(e):
        self._logs_canvas.yview_scroll(int(-1*(e.delta/120)), "units")
    self._logs_canvas.bind("<MouseWheel>", _mw)
    self._logs_body.bind("<MouseWheel>", _mw)

    # ── PAGINATION FOOTER ────────────────────────────────────────────
    footer = tk.Frame(frame, bg=BG)
    footer.pack(fill="x", padx=28, pady=(10, 16))

    self._logs_prev_btn = ctk.CTkButton(
        footer, text="◀  Prev",
        command=lambda: _page_logs(self, -1),
        width=88, height=32, corner_radius=6,
        fg_color=NAVY_MIST, hover_color=NAVY_GHOST,
        text_color=NAVY_MID, font=_F(self, 9, "bold"),
        border_width=1, border_color=BORDER_MID
    )
    self._logs_prev_btn.pack(side="left", padx=(0, 6))

    self._logs_next_btn = ctk.CTkButton(
        footer, text="Next  ▶",
        command=lambda: _page_logs(self, 1),
        width=88, height=32, corner_radius=6,
        fg_color=NAVY_MIST, hover_color=NAVY_GHOST,
        text_color=NAVY_MID, font=_F(self, 9, "bold"),
        border_width=1, border_color=BORDER_MID
    )
    self._logs_next_btn.pack(side="left")

    self._logs_status_lbl = tk.Label(
        footer, text="", font=_F(self, 8), fg=TXT_SOFT, bg=BG)
    self._logs_status_lbl.pack(side="left", padx=(14, 0))

    # Auto-load on first render
    self.after(400, lambda: _load_logs(self))


# ─────────────────────────────────────────────────────────────────────────────
#  HEADER ROW  (rebuilt on every sort so the indicator updates)
# ─────────────────────────────────────────────────────────────────────────────

def _build_header_row(self):
    for w in self._logs_hdr_row.winfo_children():
        w.destroy()

    for col_key, col_label, col_w, _ in _COLUMNS:
        is_expand = (col_w == 0)

        cell = tk.Frame(self._logs_hdr_row, bg=NAVY_DEEP,
                        width=col_w if not is_expand else 1)
        cell.pack(side="left", fill="both", expand=is_expand)
        cell.pack_propagate(False)

        indicator = ""
        if self._logs_sort_col == col_key:
            indicator = "  ▲" if self._logs_sort_asc else "  ▼"

        is_active = (self._logs_sort_col == col_key)
        lbl = tk.Label(
            cell,
            text=col_label + indicator,
            font=_F(self, 8, "bold"),
            fg=WHITE if is_active else "#A8C8E8",
            bg=NAVY_DEEP,
            anchor="center", cursor="hand2", padx=8
        )
        lbl.pack(fill="both", expand=True)

        def _sort(e=None, k=col_key):
            _sort_logs(self, k)
        lbl.bind("<Button-1>", _sort)
        cell.bind("<Button-1>", _sort)

        def _he(e, l=lbl): l.config(fg=WHITE)
        def _hl(e, l=lbl, k=col_key):
            l.config(fg=WHITE if self._logs_sort_col == k else "#A8C8E8")
        lbl.bind("<Enter>", _he)
        lbl.bind("<Leave>", _hl)

        tk.Frame(self._logs_hdr_row, bg="#1E3A5F", width=1).pack(
            side="left", fill="y", pady=8)


# ─────────────────────────────────────────────────────────────────────────────
#  DATA LOADING
# ─────────────────────────────────────────────────────────────────────────────

def _load_logs(self):
    _set_status(self, "Loading…", ACCENT_GOLD)
    try:
        conn = self.get_conn()
        if conn is None:
            _set_status(self, "✘ No database connection", ACCENT_RED)
            return
        cur = conn.cursor()
        cur.execute(
            "SELECT id, user_id, email, action, description, time "
            "FROM logs ORDER BY id DESC"
        )
        rows = cur.fetchall()
        cur.close()
    except Exception as e:
        _set_status(self, f"✘ {e}", ACCENT_RED)
        return

    today_str   = datetime.now().strftime("%Y-%m-%d")
    today_count = 0

    self._logs_all_rows = []
    for r in rows:
        ts = _fmt_time(r[5])
        if ts.startswith(today_str):
            today_count += 1
        self._logs_all_rows.append({
            "id":          str(r[0]),
            "user_id":     str(r[1]) if r[1] is not None else "—",
            "email":       r[2] or "—",
            "action":      r[3] or "—",
            "description": r[4] or "—",
            "time":        ts,
        })

    total   = len(self._logs_all_rows)
    u_users = len({r["user_id"] for r in self._logs_all_rows})
    u_acts  = len({r["action"]  for r in self._logs_all_rows})

    self._logs_stat_lbls["total"].config(text=f"{total:,}")
    self._logs_stat_lbls["today"].config(text=f"{today_count:,}")
    self._logs_stat_lbls["users"].config(text=f"{u_users:,}")
    self._logs_stat_lbls["actions"].config(text=f"{u_acts:,}")
    self._logs_count_lbl.config(text=f"{total:,} records")

    actions = sorted({r["action"] for r in self._logs_all_rows if r["action"] != "—"})
    self._logs_action_menu.configure(values=["All Actions"] + actions)
    self._logs_action_var.set("All Actions")

    self._logs_page = 0
    _apply_filter(self)
    _set_status(self, f"✔  Loaded {total:,} rows", ACCENT_SUCCESS)


def _fmt_time(val):
    if val is None:
        return "—"
    try:
        if isinstance(val, str):
            val = datetime.fromisoformat(val)
        return val.strftime("%Y-%m-%d  %H:%M:%S")
    except Exception:
        return str(val)


# ─────────────────────────────────────────────────────────────────────────────
#  FILTER & SORT
# ─────────────────────────────────────────────────────────────────────────────

def _apply_filter(self):
    raw   = self._logs_search_var.get().strip().lower()
    query = "" if "search by" in raw else raw
    act   = self._logs_action_var.get()

    rows = self._logs_all_rows
    if act and act != "All Actions":
        rows = [r for r in rows if r["action"].lower() == act.lower()]
    if query:
        rows = [r for r in rows if any(query in str(v).lower() for v in r.values())]

    self._logs_filtered = list(rows)
    self._logs_page = 0
    _render_page(self)


def _sort_logs(self, col_key):
    if self._logs_sort_col == col_key:
        self._logs_sort_asc = not self._logs_sort_asc
    else:
        self._logs_sort_col = col_key
        self._logs_sort_asc = True

    def _key(r):
        v = r.get(col_key, "")
        if col_key in ("id", "user_id"):
            try: return int(v)
            except: return 0
        return str(v).lower()

    self._logs_filtered.sort(key=_key, reverse=not self._logs_sort_asc)
    self._logs_page = 0
    _build_header_row(self)
    _render_page(self)


# ─────────────────────────────────────────────────────────────────────────────
#  PAGINATION
# ─────────────────────────────────────────────────────────────────────────────

def _page_logs(self, direction):
    total_pages = max(1, (len(self._logs_filtered) + _PAGE_SIZE - 1) // _PAGE_SIZE)
    new_page = self._logs_page + direction
    if 0 <= new_page < total_pages:
        self._logs_page = new_page
        _render_page(self)


# ─────────────────────────────────────────────────────────────────────────────
#  TABLE RENDER
# ─────────────────────────────────────────────────────────────────────────────

def _render_page(self):
    for w in self._logs_body.winfo_children():
        w.destroy()

    rows        = self._logs_filtered
    total       = len(rows)
    total_pages = max(1, (total + _PAGE_SIZE - 1) // _PAGE_SIZE)
    start       = self._logs_page * _PAGE_SIZE
    end         = min(start + _PAGE_SIZE, total)
    page_rows   = rows[start:end]

    s = start + 1 if total else 0
    self._logs_page_info.config(
        text=f"Page {self._logs_page + 1} / {total_pages}  ·  "
             f"{s}–{end} of {total:,} rows"
    )
    self._logs_prev_btn.configure(
        state="normal" if self._logs_page > 0 else "disabled")
    self._logs_next_btn.configure(
        state="normal" if self._logs_page < total_pages - 1 else "disabled")

    if not page_rows:
        holder = tk.Frame(self._logs_body, bg=WHITE, height=160)
        holder.pack(fill="x")
        tk.Label(holder, text="No records match your filter",
                 font=_F(self, 11), fg=TXT_MUTED, bg=WHITE
                 ).place(relx=0.5, rely=0.5, anchor="center")
        self._logs_canvas.after(
            30, lambda: self._logs_canvas.configure(
                scrollregion=self._logs_canvas.bbox("all")))
        return

    for i, row in enumerate(page_rows):
        row_bg = WHITE if i % 2 == 0 else "#F6F9FF"

        rf = tk.Frame(self._logs_body, bg=row_bg, height=38)
        rf.pack(fill="x")
        rf.pack_propagate(False)

        hover_widgets = []   # (widget, normal_bg)

        for col_key, _, col_w, col_anchor in _COLUMNS:
            is_expand = (col_w == 0)

            cell = tk.Frame(rf, bg=row_bg, width=col_w if not is_expand else 1)
            cell.pack(side="left", fill="both", expand=is_expand)
            cell.pack_propagate(False)
            hover_widgets.append((cell, row_bg))

            val = row.get(col_key, "")

            if col_key == "action":
                # Coloured pill badge
                bbg, bfg = _action_color(val)
                badge_frame = tk.Frame(cell, bg=row_bg)
                badge_frame.place(relx=0.5, rely=0.5, anchor="center")
                tk.Label(badge_frame, text=val,
                         font=_F(self, 8, "bold"),
                         fg=bfg, bg=bbg,
                         padx=9, pady=2).pack()
                hover_widgets.append((badge_frame, row_bg))
            else:
                short = _trunc(val, col_w if not is_expand else 500)
                lbl = tk.Label(
                    cell, text=short,
                    font=_F(self, 9),
                    fg=NAVY_MID if col_key == "id" else TXT_NAVY,
                    bg=row_bg, anchor=col_anchor, padx=8
                )
                lbl.pack(fill="both", expand=True)
                hover_widgets.append((lbl, row_bg))

            # Column separator
            tk.Frame(rf, bg=BORDER_LIGHT, width=1).pack(
                side="left", fill="y", pady=6)

        # Bottom hairline
        tk.Frame(self._logs_body, bg=BORDER_LIGHT, height=1).pack(fill="x")

        # Hover highlight
        def _enter(e, rf=rf, hw=hover_widgets):
            rf.config(bg=NAVY_MIST)
            for w, _ in hw:
                try:
                    if isinstance(w, tk.Label):
                        w.config(bg=NAVY_MIST)
                    elif isinstance(w, tk.Frame):
                        w.config(bg=NAVY_MIST)
                except Exception:
                    pass

        def _leave(e, rf=rf, hw=hover_widgets):
            rf.config(bg=hw[0][1])
            for w, orig in hw:
                try:
                    if isinstance(w, tk.Label):
                        w.config(bg=orig)
                    elif isinstance(w, tk.Frame):
                        w.config(bg=orig)
                except Exception:
                    pass

        rf.bind("<Enter>", _enter)
        rf.bind("<Leave>", _leave)
        for w, _ in hover_widgets:
            w.bind("<Enter>", _enter)
            w.bind("<Leave>", _leave)

    self._logs_canvas.yview_moveto(0)
    self._logs_canvas.after(
        30, lambda: self._logs_canvas.configure(
            scrollregion=self._logs_canvas.bbox("all")))


# ─────────────────────────────────────────────────────────────────────────────
#  SMALL HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _action_color(action: str):
    key = action.lower().strip()
    for k, v in _ACTION_COLORS.items():
        if k in key:
            return v
    return _ACTION_DEFAULT


def _trunc(text: str, pixel_width: int, ratio: float = 6.5) -> str:
    if pixel_width <= 0:
        return text
    max_c = max(4, int(pixel_width / ratio) - 2)
    return (text[:max_c - 1] + "…") if len(text) > max_c else text


def _set_status(self, msg: str, color: str = TXT_SOFT):
    try:
        if self._logs_status_lbl.winfo_exists():
            self._logs_status_lbl.config(text=msg, fg=color)
    except Exception:
        pass


def _F(self, size: int, weight: str = "normal") -> tuple:
    """Return a scaled font tuple, delegates to self.F() when available."""
    try:
        return self.F(size, weight)
    except Exception:
        zoom = getattr(self, "_ui_zoom", 1.0)
        return ("Segoe UI", max(6, int(round(size * zoom))), weight)


# ─────────────────────────────────────────────────────────────────────────────
#  PUBLIC HELPER — write a log entry from anywhere in the app
# ─────────────────────────────────────────────────────────────────────────────

def insert_log(app, action: str, description: str = ""):
    """
    Insert a row into the logs table.

    Usage (from any module that has a reference to the app instance):
        from admin_logs import insert_log
        insert_log(self, "login", "User authenticated successfully")
    """
    try:
        conn = app.get_conn()
        if conn is None:
            return
        user_id = getattr(app, "_current_user_id", None)
        email   = getattr(app, "_current_username", "") or ""
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO logs (user_id, email, action, description) "
            "VALUES (%s, %s, %s, %s)",
            (user_id, email, action, description)
        )
        conn.commit()
        cur.close()
    except Exception as e:
        print(f"[admin_logs] insert_log error: {e}")


# ─────────────────────────────────────────────────────────────────────────────
#  ATTACH
# ─────────────────────────────────────────────────────────────────────────────
def _export_logs_excel(self):
    import os
    import tkinter.filedialog as fd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    rows = self._logs_filtered if self._logs_filtered else self._logs_all_rows
    if not rows:
        _set_status(self, "✘ No data to export", ACCENT_RED)
        return

    save_path = fd.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel Workbook", "*.xlsx")],
        initialfile=f"system_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        title="Save Report As"
    )
    if not save_path:
        return

    _set_status(self, "Generating report…", ACCENT_GOLD)

    wb = Workbook()
    ws = wb.active
    ws.title = "System Logs"

    # ── Styles ──────────────────────────────────────────────────────
    hdr_font    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill    = PatternFill("solid", start_color="0B1F3A")
    hdr_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font   = Font(name="Arial", size=9)
    cell_align  = Alignment(vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="C8D8EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill("solid", start_color="F6F9FF")

    # ── Title block ─────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    ws["A1"] = "Banco San Vicente — System Logs Report"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="0B1F3A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%B %d, %Y  %H:%M:%S')}   |   Total Records: {len(rows):,}"
    ws["A2"].font = Font(name="Arial", size=9, color="7A94B0")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 6   # spacer

    # ── Column headers ───────────────────────────────────────────────
    headers = ["ID", "User ID", "Email", "Action", "Description", "Timestamp"]
    col_widths = [8, 10, 34, 18, 60, 22]

    for col_i, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=4, column=col_i, value=hdr)
        cell.font    = hdr_font
        cell.fill    = hdr_fill
        cell.alignment = hdr_align
        cell.border  = border
        ws.column_dimensions[get_column_letter(col_i)].width = width

    ws.row_dimensions[4].height = 22

    # ── Data rows ────────────────────────────────────────────────────
    for row_i, row in enumerate(rows, start=5):
        is_alt = (row_i % 2 == 0)
        row_data = [
            row["id"], row["user_id"], row["email"],
            row["action"], row["description"], row["time"]
        ]
        for col_i, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=value)
            cell.font   = cell_font
            cell.border = border
            if col_i in (1, 2, 4, 6):   # ID, User ID, Action, Timestamp → centred
                cell.alignment = center_align
            else:
                cell.alignment = cell_align
            if is_alt:
                cell.fill = alt_fill
        ws.row_dimensions[row_i].height = 16

    # ── Summary sheet ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 18

    ws2["A1"] = "Report Summary"
    ws2["A1"].font = Font(name="Arial", bold=True, size=13, color="0B1F3A")
    ws2.row_dimensions[1].height = 26

    summary_data = [
        ("Total Records",   len(rows)),
        ("Unique Users",    len({r["user_id"] for r in rows})),
        ("Action Types",    len({r["action"]  for r in rows})),
        ("Generated At",    datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]

    # Action breakdown
    from collections import Counter
    action_counts = Counter(r["action"] for r in rows)

    summary_data.append(("", ""))
    summary_data.append(("Action", "Count"))
    for action, count in sorted(action_counts.items(), key=lambda x: -x[1]):
        summary_data.append((action, count))

    for s_i, (label, value) in enumerate(summary_data, start=3):
        ca = ws2.cell(row=s_i, column=1, value=label)
        cb = ws2.cell(row=s_i, column=2, value=value)
        if label in ("Action", ""):
            ca.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
            cb.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
            ca.fill = PatternFill("solid", start_color="1E3A5F")
            cb.fill = PatternFill("solid", start_color="1E3A5F")
            ca.alignment = center_align
            cb.alignment = center_align
        else:
            ca.font = Font(name="Arial", size=9, color="1A2E42")
            cb.font = Font(name="Arial", size=9, color="2E5C8A")
            cb.alignment = center_align
        ca.border = border
        cb.border = border

    # ── Freeze panes & auto-filter ───────────────────────────────────
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:F{4 + len(rows)}"

    try:
        wb.save(save_path)
        _set_status(self, f"✔  Report saved: {os.path.basename(save_path)}", ACCENT_SUCCESS)
    except Exception as e:
        _set_status(self, f"✘ Save failed: {e}", ACCENT_RED)
def attach(cls):
    """Called by ui_panels.attach() — injects _build_logs_panel into the app class."""
    cls._build_logs_panel = _build_logs_panel